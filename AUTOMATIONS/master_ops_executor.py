#!/usr/bin/env python3
"""Execute or plan actions from enhanced Master Ops workbook."""

from __future__ import annotations

import argparse
import json
import subprocess
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List
from zipfile import BadZipFile

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent.parent
OUT_DIR = BASE_DIR / "output" / "master_ops_exec"


COMMANDS_BY_OP: Dict[str, List[str]] = {
    "D01": [
        "python3 AUTOMATIONS/gumroad_autolist_packager.py --write",
    ],
    "S01": [
        "python3 AUTOMATIONS/freelance_demand_scanner.py --hourly",
        "python3 AUTOMATIONS/auto_freelance_responder.py --dry-run",
    ],
    "S05": [
        "python3 AUTOMATIONS/email_sender.py --preview --outreach AUTOMATIONS/outreach/HOT_BATCH_FEB13_COMPLIANT.csv --max-sends 25",
    ],
    "C12": [
        "python3 AUTOMATIONS/email_sender.py --preview --outreach AUTOMATIONS/outreach/HOT_BATCH_FEB13_COMPLIANT.csv --max-sends 25",
    ],
    "N61": [
        "python3 AUTOMATIONS/local_biz_pipeline.py --urls-file AUTOMATIONS/sample_local_biz_urls.csv --dry-run",
    ],
    "N68": [
        "python3 AUTOMATIONS/app_packager.py --write",
        "python3 AUTOMATIONS/deploy_guard.py --tick",
    ],
    "S02": [
        "python3 AUTOMATIONS/local_biz_pipeline.py --urls-file AUTOMATIONS/sample_local_biz_urls.csv --dry-run",
    ],
}

COMMANDS_BY_LANE: Dict[str, List[str]] = {
    "freelance_arbitrage": [
        "python3 AUTOMATIONS/freelance_demand_scanner.py --hourly",
        "python3 AUTOMATIONS/auto_freelance_responder.py --dry-run",
        "python3 AUTOMATIONS/freelance_packager.py --write",
    ],
    "gumroad_listings": [
        "python3 AUTOMATIONS/gumroad_autolist_packager.py --write",
    ],
    "cold_outreach_warmup": [
        "python3 AUTOMATIONS/email_sender.py --preview --outreach AUTOMATIONS/outreach/HOT_BATCH_FEB13_COMPLIANT.csv --max-sends 25",
    ],
    "rbi_intent_sniping": [
        "python3 AUTOMATIONS/clawdbot_rbi_engine.py --tick --max-intents 180 --max-syndication 420 --max-directories 900 --max-jobs 200 --max-keywords 260 --max-community 180",
    ],
    "app_factory": [
        "python3 AUTOMATIONS/app_packager.py --write",
        "python3 AUTOMATIONS/deploy_guard.py --tick",
    ],
    "ecom_arb": [
        "python3 AUTOMATIONS/ecom_arb_engine.py --hourly --top 2",
        "python3 AUTOMATIONS/ecom_autopilot.py --tick --top 12 --min-margin 20 --min-profit 3",
        "python3 AUTOMATIONS/ecom_distributor.py --distribute-all",
    ],
    "other": [
        "python3 AUTOMATIONS/alpha_monitor.py --cron",
    ],
}


def now_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def clean(v: Any) -> str:
    return str(v or "").strip()


def safe_float(v: Any, default: float = 999.0) -> float:
    try:
        return float(str(v).strip())
    except Exception:
        return default


def find_latest_workbook(explicit: str) -> Path:
    if explicit:
        p = Path(explicit).expanduser().resolve()
        if p.exists():
            return p
        raise FileNotFoundError(f"Workbook not found: {p}")
    candidates = sorted(BASE_DIR.glob("PRINTMAXX_MASTER_OPS_ENHANCED_*.xlsx"))
    if not candidates:
        raise FileNotFoundError("No PRINTMAXX_MASTER_OPS_ENHANCED_*.xlsx found")
    last_error = ""
    for candidate in reversed(candidates):
        try:
            wb = load_workbook(candidate, read_only=True, data_only=True)
            wb.close()
            return candidate
        except Exception as exc:
            last_error = str(exc)
            continue
    raise FileNotFoundError(f"No readable enhanced workbook found. last_error={last_error}")


def read_priority_rows(path: Path) -> List[Dict[str, Any]]:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except BadZipFile as exc:
        raise FileNotFoundError(f"Workbook is invalid/corrupt: {path} ({exc})") from exc
    if "PRIORITY_AUTOMATION_EXEC" not in wb.sheetnames:
        return []
    ws_priority = wb["PRIORITY_AUTOMATION_EXEC"]
    headers = [ws_priority.cell(1, c).value for c in range(1, ws_priority.max_column + 1)]
    idx = {clean(h): i + 1 for i, h in enumerate(headers) if h is not None}

    status_map: Dict[str, Dict[str, Any]] = {}
    if "AUTO_STATUS_LIVE" in wb.sheetnames:
        ws_status = wb["AUTO_STATUS_LIVE"]
        s_headers = [ws_status.cell(1, c).value for c in range(1, ws_status.max_column + 1)]
        s_idx = {clean(h): i + 1 for i, h in enumerate(s_headers) if h is not None}
        if "OP_ID" in s_idx:
            for r in range(2, ws_status.max_row + 1):
                op_id = clean(ws_status.cell(r, s_idx["OP_ID"]).value)
                if not op_id:
                    continue
                status_map[op_id] = {
                    "LANE": clean(ws_status.cell(r, s_idx.get("LANE", 0)).value) if s_idx.get("LANE") else "other",
                    "READINESS": clean(ws_status.cell(r, s_idx.get("READINESS", 0)).value) if s_idx.get("READINESS") else "",
                    "BLOCKER_KEY": clean(ws_status.cell(r, s_idx.get("BLOCKER_KEY", 0)).value) if s_idx.get("BLOCKER_KEY") else "NONE",
                    "APPROVAL_OK": clean(ws_status.cell(r, s_idx.get("APPROVAL_OK", 0)).value) if s_idx.get("APPROVAL_OK") else "",
                    "SOURCE_ARTIFACT": clean(ws_status.cell(r, s_idx.get("SOURCE_ARTIFACT", 0)).value) if s_idx.get("SOURCE_ARTIFACT") else "",
                }

    required = {"RANK", "OP_ID", "OP_NAME"}
    if not required.issubset(set(idx.keys())):
        return []

    rows: List[Dict[str, Any]] = []
    for r in range(2, ws_priority.max_row + 1):
        op_id = clean(ws_priority.cell(r, idx["OP_ID"]).value)
        if not op_id:
            continue
        status = status_map.get(op_id, {})
        readiness = clean(ws_priority.cell(r, idx.get("READINESS", 0)).value) if idx.get("READINESS") else ""
        lane = clean(ws_priority.cell(r, idx.get("LANE", 0)).value) if idx.get("LANE") else ""
        blocker = clean(ws_priority.cell(r, idx.get("BLOCKER_KEY", 0)).value) if idx.get("BLOCKER_KEY") else ""

        rows.append(
            {
                "RANK": ws_priority.cell(r, idx["RANK"]).value,
                "OP_ID": op_id,
                "OP_NAME": ws_priority.cell(r, idx["OP_NAME"]).value,
                "READINESS": readiness or status.get("READINESS", "BUILD"),
                "LANE": lane or status.get("LANE", "other"),
                "BLOCKER_KEY": blocker or status.get("BLOCKER_KEY", "NONE"),
                "APPROVAL_OK": clean(ws_priority.cell(r, idx.get("APPROVAL_OK", 0)).value) if idx.get("APPROVAL_OK") else status.get("APPROVAL_OK", ""),
                "SOURCE_ARTIFACT": clean(ws_priority.cell(r, idx.get("SOURCE_ARTIFACT", 0)).value) if idx.get("SOURCE_ARTIFACT") else status.get("SOURCE_ARTIFACT", ""),
                "AUTOMATION_SCORE_100": ws_priority.cell(r, idx.get("AUTOMATION_SCORE_100", 0)).value if idx.get("AUTOMATION_SCORE_100") else "",
                "SIGNAL_COUNT": ws_priority.cell(r, idx.get("SIGNAL_COUNT", 0)).value if idx.get("SIGNAL_COUNT") else "",
            }
        )
    rows.sort(key=lambda x: safe_float(x.get("RANK"), 999.0))
    return rows


def commands_for_row(row: Dict[str, Any]) -> List[str]:
    op_id = clean(row.get("OP_ID"))
    lane = clean(row.get("LANE")).lower() or "other"
    if op_id in COMMANDS_BY_OP:
        return list(COMMANDS_BY_OP[op_id])
    if lane in COMMANDS_BY_LANE:
        return list(COMMANDS_BY_LANE[lane])
    return list(COMMANDS_BY_LANE["other"])


def plan_actions(
    rows: List[Dict[str, Any]],
    limit: int,
    include_build: bool,
    allow_blocked_prework: bool,
    max_per_lane: int,
) -> List[Dict[str, Any]]:
    planned: List[Dict[str, Any]] = []
    lane_counts: Dict[str, int] = defaultdict(int)
    for row in rows:
        if len(planned) >= max(1, limit):
            break
        readiness = clean(row.get("READINESS")).upper() or "BUILD"
        blocker = clean(row.get("BLOCKER_KEY")).upper() or "NONE"
        op_id = clean(row.get("OP_ID"))
        lane = clean(row.get("LANE")).lower() or "other"
        if readiness not in {"READY", "BLOCKED", "BUILD"}:
            continue
        if readiness == "BUILD" and not include_build:
            continue
        is_blocked = blocker not in {"NONE", ""}
        if is_blocked and not allow_blocked_prework:
            continue
        if lane_counts[lane] >= max(1, max_per_lane):
            continue
        cmds = commands_for_row(row)
        if not cmds:
            continue
        lane_counts[lane] += 1
        planned.append(
            {
                "op_id": op_id,
                "op_name": row.get("OP_NAME") or "",
                "rank": row.get("RANK"),
                "lane": lane,
                "readiness": readiness,
                "blocked_prework": bool(is_blocked),
                "blocker_key": blocker,
                "approval_ok": clean(row.get("APPROVAL_OK")),
                "source_artifact": clean(row.get("SOURCE_ARTIFACT")),
                "commands": cmds,
                "automation_score_100": row.get("AUTOMATION_SCORE_100", ""),
                "signal_count": row.get("SIGNAL_COUNT", ""),
            }
        )
    return planned


def run_commands(commands: List[str], timeout_sec: int) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for cmd in commands:
        try:
            proc = subprocess.run(
                ["bash", "-lc", cmd],
                cwd=BASE_DIR,
                capture_output=True,
                text=True,
                timeout=timeout_sec,
                check=False,
            )
            merged = ((proc.stdout or "") + ("\n" + proc.stderr if proc.stderr else "")).strip()
            out.append(
                {
                    "command": cmd,
                    "status": "OK" if proc.returncode == 0 else "FAILED",
                    "return_code": int(proc.returncode),
                    "output_tail": merged[-400:],
                }
            )
        except subprocess.TimeoutExpired:
            out.append({"command": cmd, "status": "TIMEOUT", "return_code": 124, "output_tail": f"Timed out after {timeout_sec}s"})
    return out


def write_outputs(payload: Dict[str, Any]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    manifest = OUT_DIR / "manifest.json"
    latest = OUT_DIR / "latest.md"
    manifest.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")

    lines: List[str] = []
    lines.append("# Master Ops Executor")
    lines.append("")
    lines.append(f"Generated: {payload.get('generated_at', '')}")
    lines.append(f"Workbook: {payload.get('workbook', '')}")
    lines.append(f"Mode: {payload.get('mode', '')}")
    lines.append("")
    lines.append("## Planned Ops")
    lines.append("")
    planned = payload.get("planned", [])
    if not planned:
        lines.append("- None")
    else:
        for item in planned:
            lines.append(f"- {item.get('op_id', '')}: {item.get('op_name', '')}")
            lines.append(
                f"  lane={item.get('lane', '')} readiness={item.get('readiness', '')} blocked_prework={item.get('blocked_prework', False)} blocker={item.get('blocker_key', '')}"
            )
            lines.append(f"  rank={item.get('rank', '')} score={item.get('automation_score_100', '')} signal={item.get('signal_count', '')}")
            for cmd in item.get("commands", []):
                lines.append(f"  - `{cmd}`")
    lines.append("")
    lines.append("## Execution Results")
    lines.append("")
    results = payload.get("results", [])
    if not results:
        lines.append("- Dry run only (no commands executed).")
    else:
        for item in results:
            lines.append(f"- `{item.get('command', '')}` -> {item.get('status', '')} (rc={item.get('return_code', '')})")
    latest.write_text("\n".join(lines) + "\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Plan/execute from enhanced Master Ops workbook")
    ap.add_argument("--workbook", default="", help="Enhanced workbook path (default: latest)")
    ap.add_argument("--top", type=int, default=5, help="Max READY ops to plan")
    ap.add_argument("--include-build", action="store_true", help="Include BUILD ops with discovery commands")
    ap.add_argument(
        "--allow-blocked-prework",
        action="store_true",
        default=True,
        help="Allow safe prework commands even when account blockers remain (default: enabled)",
    )
    ap.add_argument("--no-allow-blocked-prework", dest="allow_blocked_prework", action="store_false", help="Skip blocked ops")
    ap.add_argument("--max-per-lane", type=int, default=2, help="Cap planned ops per lane")
    ap.add_argument("--apply", action="store_true", help="Execute mapped commands (default dry-run)")
    ap.add_argument("--timeout-sec", type=int, default=900, help="Per-command timeout")
    return ap.parse_args()


def main() -> int:
    args = parse_args()
    workbook = find_latest_workbook(args.workbook)
    rows = read_priority_rows(workbook)
    planned = plan_actions(
        rows,
        limit=max(1, args.top),
        include_build=bool(args.include_build),
        allow_blocked_prework=bool(args.allow_blocked_prework),
        max_per_lane=max(1, args.max_per_lane),
    )
    flattened_cmds: List[str] = []
    for p in planned:
        flattened_cmds.extend([str(c) for c in p.get("commands", [])])
    results = run_commands(flattened_cmds, timeout_sec=max(60, args.timeout_sec)) if args.apply else []

    payload = {
        "generated_at": now_iso(),
        "workbook": str(workbook),
        "mode": "apply" if args.apply else "dry_run",
        "planned": planned,
        "results": results,
    }
    write_outputs(payload)
    print("master_ops_executor: wrote")
    print(f"- {OUT_DIR / 'manifest.json'}")
    print(f"- {OUT_DIR / 'latest.md'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
