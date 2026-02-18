#!/usr/bin/env python3
"""Clone and enhance PRINTMAXX_MASTER_OPS.xlsx with live automation intelligence.

Adds sheets that map static ops planning to current runtime signals, blockers,
and expansion opportunities.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


BASE_DIR = Path(__file__).resolve().parent.parent
OPS_DIR = BASE_DIR / "OPS"
OUTPUT_DIR = BASE_DIR / "output"
LEDGER_DIR = BASE_DIR / "LEDGER"

DEFAULT_INPUT = BASE_DIR / "PRINTMAXX_MASTER_OPS.xlsx"


HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
PLAYBOOK_SECTION_TITLES = {
    "LLM ALPHA",
    "INFRA STACK",
    "SETUP INSTRUCTIONS",
    "ALGORITHM GUIDE",
    "SHADOWBAN AVOIDANCE",
    "LLM-IN-THE-LOOP",
    "MANUAL FIRST",
    "AUTOMATE AFTER",
}


def now_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def safe_int(v: object, default: int = 0) -> int:
    try:
        return int(float(str(v).strip()))
    except Exception:
        return default


def safe_float(v: object, default: float = 0.0) -> float:
    try:
        return float(str(v).strip())
    except Exception:
        return default


def load_json(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
        return payload if isinstance(payload, dict) else {}
    except Exception:
        return {}


def count_csv_rows(path: Path) -> int:
    if not path.exists():
        return 0
    try:
        with open(path, "r", encoding="utf-8", errors="replace", newline="") as f:
            reader = csv.reader(f)
            count = -1
            for _ in reader:
                count += 1
        return max(0, count)
    except Exception:
        return 0


def parse_pending_approvals(path: Path) -> List[str]:
    if not path.exists():
        return []
    out: List[str] = []
    for raw in path.read_text(encoding="utf-8", errors="replace").splitlines():
        s = raw.strip()
        if s.startswith("- [ ]"):
            out.append(s)
    return out


def parse_approved_keys(path: Path) -> List[str]:
    if not path.exists():
        return []
    approved: List[str] = []
    with open(path, "r", encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            key = (row.get("key") or "").strip().upper()
            status = (row.get("status") or "").strip().upper()
            if key and status in {"APPROVED", "ACTIVE", "TRUE", "YES"}:
                approved.append(key)
    return sorted(set(approved))


def load_all_ops(ws) -> List[Dict[str, Any]]:
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    out: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        row = {}
        has_data = False
        for c, h in enumerate(headers, start=1):
            key = str(h).strip() if h is not None else f"COL_{c}"
            value = ws.cell(r, c).value
            if value not in (None, ""):
                has_data = True
            row[key] = value
        if has_data:
            out.append(row)
    return out


def load_priority_launch(ws) -> List[Dict[str, Any]]:
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    out: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        row = {}
        has_data = False
        for c, h in enumerate(headers, start=1):
            key = str(h).strip() if h is not None else f"COL_{c}"
            value = ws.cell(r, c).value
            if value not in (None, ""):
                has_data = True
            row[key] = value
        if has_data:
            out.append(row)
    return out


def delete_if_exists(wb, sheet_name: str) -> None:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)


def stylize_header(ws, row_idx: int, total_cols: int) -> None:
    for c in range(1, total_cols + 1):
        cell = ws.cell(row_idx, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def set_widths(ws, widths: Dict[int, float]) -> None:
    for idx, width in widths.items():
        col_letter = ws.cell(1, idx).column_letter
        ws.column_dimensions[col_letter].width = width


def clean_text(v: Any) -> str:
    s = str(v or "").strip()
    return re.sub(r"\s+", " ", s)


def lane_from_op(op_name: str, category: str, op_id: str) -> str:
    text = f"{op_name} {category} {op_id}".lower()
    if any(k in text for k in ["gumroad", "digital product", "template", "notion"]):
        return "gumroad_listings"
    if any(k in text for k in ["freelance", "upwork", "fiverr", "client"]):
        return "freelance_arbitrage"
    if any(k in text for k in ["cold email", "outreach", "email", "lead gen", "lead"]):
        return "cold_outreach_warmup"
    if any(k in text for k in ["rbi", "reddit", "x/twitter", "content", "syndication", "community"]):
        return "rbi_intent_sniping"
    if any(k in text for k in ["app", "mobile", "pwa", "saas"]):
        return "app_factory"
    if any(k in text for k in ["ecom", "etsy", "ebay", "amazon", "shop"]):
        return "ecom_arb"
    return "other"


def blocker_for_lane(lane: str) -> str:
    mapping = {
        "gumroad_listings": "GUMROAD_ACCOUNT",
        "freelance_arbitrage": "FIVERR_UPWORK_ACCOUNT",
        "cold_outreach_warmup": "EMAIL_INFRA",
        "rbi_intent_sniping": "X_MULTI_ACCOUNT_STACK",
        "ecom_arb": "ACCOUNT_EBAY/ETSY/AMAZON",
        "app_factory": "STORE_ACCOUNT_AND_PAYMENT",
    }
    return mapping.get(lane, "NONE")


def command_template_for_lane(lane: str) -> str:
    templates = {
        "freelance_arbitrage": "python3 AUTOMATIONS/freelance_demand_scanner.py --hourly && python3 AUTOMATIONS/auto_freelance_responder.py --dry-run",
        "gumroad_listings": "python3 AUTOMATIONS/gumroad_autolist_packager.py --write",
        "cold_outreach_warmup": "python3 AUTOMATIONS/email_sender.py --preview --outreach AUTOMATIONS/outreach/HOT_BATCH_FEB13_COMPLIANT.csv --max-sends 25",
        "rbi_intent_sniping": "python3 AUTOMATIONS/clawdbot_rbi_engine.py --tick --max-intents 180 --max-syndication 420 --max-directories 900",
        "app_factory": "python3 AUTOMATIONS/app_packager.py --write && python3 AUTOMATIONS/deploy_guard.py --tick",
        "ecom_arb": "python3 AUTOMATIONS/ecom_arb_engine.py --hourly --top 2 && python3 AUTOMATIONS/ecom_autopilot.py --tick --top 12 --min-margin 20 --min-profit 3",
        "other": "python3 AUTOMATIONS/alpha_monitor.py --cron",
    }
    return templates.get(lane, templates["other"])


def extract_deep_playbook_index(ws) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    current_playbook = ""
    current_title = ""
    current_section = "OVERVIEW"
    step_index = 0

    for r in range(1, ws.max_row + 1):
        col_b = clean_text(ws.cell(r, 2).value)
        col_c = clean_text(ws.cell(r, 3).value)
        if not col_b and not col_c:
            continue

        header_match = re.match(r"^(OP\d+):\s*(.+)$", col_b, flags=re.IGNORECASE)
        if header_match:
            current_playbook = header_match.group(1).upper()
            current_title = clean_text(header_match.group(2))
            current_section = "OVERVIEW"
            step_index = 0
            rows.append(
                {
                    "PLAYBOOK_OP": current_playbook,
                    "PLAYBOOK_TITLE": current_title,
                    "SECTION": current_section,
                    "ROW_TYPE": "PLAYBOOK_HEADER",
                    "STEP_INDEX": step_index,
                    "TEXT": col_b,
                    "LANE": lane_from_op(current_title, "deep_playbook", current_playbook),
                    "SOURCE_ROW": r,
                }
            )
            continue

        if col_b and not col_c and col_b.upper() in PLAYBOOK_SECTION_TITLES:
            current_section = col_b.upper()
            rows.append(
                {
                    "PLAYBOOK_OP": current_playbook,
                    "PLAYBOOK_TITLE": current_title,
                    "SECTION": current_section,
                    "ROW_TYPE": "SECTION",
                    "STEP_INDEX": step_index,
                    "TEXT": col_b,
                    "LANE": lane_from_op(current_title, "deep_playbook", current_playbook),
                    "SOURCE_ROW": r,
                }
            )
            continue

        step_index += 1
        content = col_c or col_b
        rows.append(
            {
                "PLAYBOOK_OP": current_playbook,
                "PLAYBOOK_TITLE": current_title,
                "SECTION": current_section,
                "ROW_TYPE": "STEP",
                "STEP_INDEX": step_index,
                "TEXT": content,
                "LANE": lane_from_op(current_title, "deep_playbook", current_playbook),
                "SOURCE_ROW": r,
            }
        )
    return rows


def extract_alpha_thesis_index(ws) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    in_table = False
    for r in range(1, ws.max_row + 1):
        c2 = clean_text(ws.cell(r, 2).value)
        c3 = clean_text(ws.cell(r, 3).value)
        c4 = clean_text(ws.cell(r, 4).value)
        if not in_table:
            if c2.upper() == "OPPORTUNITY" and "EDGE" in c3.upper():
                in_table = True
            continue
        if c2.upper().startswith("THE PRINTMAXX THESIS"):
            break
        if not c2 or not c3:
            continue
        lane = lane_from_op(c2, "alpha_thesis", "")
        rows.append(
            {
                "ALPHA_ID": f"A{len(rows) + 1:03d}",
                "OPPORTUNITY": c2,
                "WHY_LLM_EDGE": c3,
                "EDGE_DURATION": c4,
                "LANE": lane,
                "BLOCKER_KEY": blocker_for_lane(lane),
                "COMMAND_TEMPLATE": command_template_for_lane(lane),
                "SOURCE_ROW": r,
            }
        )
    return rows


def lane_metrics() -> Dict[str, Dict[str, Any]]:
    freelance = load_json(OUTPUT_DIR / "freelance" / "manifest.json")
    ecom = load_json(OUTPUT_DIR / "ecom" / "manifest.json")
    ecom_autolist = load_json(OUTPUT_DIR / "ecom_autolist" / "manifest.json")
    clawdbot = load_json(OUTPUT_DIR / "clawdbot" / "manifest.json")
    sidecar = load_json(OUTPUT_DIR / "clawwork_sidecar" / "manifest.json")
    claw_counts = clawdbot.get("counts", {}) if isinstance(clawdbot.get("counts"), dict) else {}
    sidecar_totals = sidecar.get("totals", {}) if isinstance(sidecar.get("totals"), dict) else {}

    cold_rows = count_csv_rows(OUTPUT_DIR / "cold_emails" / "cold_emails_ready.csv")
    return {
        "freelance_arbitrage": {
            "signal_count": safe_int(freelance.get("draft_count"), 0),
            "artifact": "output/freelance/manifest.json",
            "expected_profit_hint": safe_float(sidecar_totals.get("expected_profit_usd"), 0.0),
        },
        "gumroad_listings": {
            "signal_count": safe_int(ecom_autolist.get("pending_count"), 0),
            "artifact": "output/ecom_autolist/manifest.json",
            "expected_profit_hint": safe_float(sidecar_totals.get("expected_profit_usd"), 0.0),
        },
        "cold_outreach_warmup": {
            "signal_count": cold_rows,
            "artifact": "output/cold_emails/cold_emails_ready.csv",
            "expected_profit_hint": safe_float(sidecar_totals.get("expected_profit_usd"), 0.0),
        },
        "rbi_intent_sniping": {
            "signal_count": safe_int(claw_counts.get("intent_rows"), 0) + safe_int(claw_counts.get("syndication_rows"), 0),
            "artifact": "output/clawdbot/manifest.json",
            "expected_profit_hint": safe_float(sidecar_totals.get("expected_profit_usd"), 0.0),
        },
        "app_factory": {
            "signal_count": len(ecom.get("top_arb", [])) if isinstance(ecom.get("top_arb"), list) else 0,
            "artifact": "output/apps/manifest.json",
            "expected_profit_hint": 0.0,
        },
        "ecom_arb": {
            "signal_count": len(ecom.get("top_arb", [])) if isinstance(ecom.get("top_arb"), list) else 0,
            "artifact": "output/ecom/manifest.json",
            "expected_profit_hint": 0.0,
        },
        "other": {
            "signal_count": 0,
            "artifact": "n/a",
            "expected_profit_hint": 0.0,
        },
    }


def status_score(priority: Any, signal_count: int, blocked: bool, automation_level: Any) -> int:
    score = 30
    p = str(priority or "").strip().upper()
    if p == "HIGHEST":
        score += 30
    elif p == "HIGH":
        score += 20
    elif p == "MEDIUM":
        score += 10

    a = str(automation_level or "").strip().upper()
    if a == "HIGH":
        score += 15
    elif a == "MEDIUM":
        score += 8
    elif a == "LOW":
        score += 3

    if signal_count > 0:
        score += min(25, int(signal_count ** 0.5))
    if blocked:
        score -= 20
    return max(0, min(100, score))


def build_enhanced_rows(
    ops_rows: List[Dict[str, Any]],
    pending: List[str],
    approved_keys: List[str],
    lane_info: Dict[str, Dict[str, Any]],
) -> List[Dict[str, Any]]:
    pending_blob = "\n".join(pending).upper()
    out: List[Dict[str, Any]] = []
    for row in ops_rows:
        op_id = str(row.get("OP_ID") or "").strip()
        op_name = str(row.get("OP_NAME") or "").strip()
        category = str(row.get("CATEGORY") or "").strip()
        priority = row.get("PRIORITY")
        auto_level = row.get("AUTOMATION_LEVEL")

        lane = lane_from_op(op_name, category, op_id)
        metrics = lane_info.get(lane, lane_info["other"])
        signal_count = safe_int(metrics.get("signal_count"), 0)
        blocker = blocker_for_lane(lane)
        blocked = blocker != "NONE" and blocker in pending_blob
        approval_ok = blocker == "NONE" or blocker in approved_keys
        readiness = "BLOCKED" if blocked else ("READY" if signal_count > 0 else "BUILD")
        score = status_score(priority, signal_count, blocked, auto_level)

        next_action = "Run discovery"
        if lane == "freelance_arbitrage":
            next_action = "Run freelance scan + draft responder"
        elif lane == "gumroad_listings":
            next_action = "Generate Gumroad listing pack + approve AUTO_LIST_ECOM"
        elif lane == "cold_outreach_warmup":
            next_action = "Generate preview sends (no-live) and clear EMAIL_INFRA"
        elif lane == "rbi_intent_sniping":
            next_action = "Run clawdbot_rbi tick + review top draft queues"
        elif lane == "app_factory":
            next_action = "Package app + deploy guard + store submission queue"
        elif lane == "ecom_arb":
            next_action = "Run ecom arb scan + distributor matrix update"

        out.append(
            {
                "OP_ID": op_id,
                "CATEGORY": category,
                "OP_NAME": op_name,
                "LANE": lane,
                "READINESS": readiness,
                "AUTOMATION_SCORE_100": score,
                "SIGNAL_COUNT": signal_count,
                "BLOCKER_KEY": blocker,
                "APPROVAL_OK": "YES" if approval_ok else "NO",
                "NEXT_AUTOMATION_ACTION": next_action,
                "SOURCE_ARTIFACT": str(metrics.get("artifact", "n/a")),
                "EXPECTED_PROFIT_HINT_USD": float(metrics.get("expected_profit_hint", 0.0)),
                "UPDATED_AT": now_iso(),
            }
        )
    out.sort(key=lambda r: (safe_int(r.get("AUTOMATION_SCORE_100"), 0), safe_int(r.get("SIGNAL_COUNT"), 0)), reverse=True)
    return out


def write_sheet(ws, headers: List[str], rows: List[Dict[str, Any]]) -> None:
    ws.append(headers)
    stylize_header(ws, 1, len(headers))
    for row in rows:
        ws.append([row.get(h, "") for h in headers])


def write_system_evidence_sheet(wb, lane_info: Dict[str, Dict[str, Any]], pending: List[str], approved: List[str]) -> None:
    name = "SYSTEM_EVIDENCE"
    delete_if_exists(wb, name)
    ws = wb.create_sheet(name)
    ws.append(["KEY", "VALUE", "SOURCE", "NOTES"])
    stylize_header(ws, 1, 4)
    rows = [
        ("generated_at", now_iso(), "AUTOMATIONS/master_ops_enhancer.py", "Workbook enhancement run timestamp"),
        ("pending_approvals_count", len(pending), "OPS/HUMAN_LOOP_QUEUE.md", "Keys still blocking live execution"),
        ("approved_keys_count", len(approved), "OPS/HUMAN_APPROVALS.csv", "Persistent approvals"),
    ]
    for lane, info in lane_info.items():
        rows.append((f"{lane}.signal_count", safe_int(info.get("signal_count"), 0), str(info.get("artifact", "")), "Live queue/signal volume"))
    for row in rows:
        ws.append(list(row))
    set_widths(ws, {1: 34, 2: 18, 3: 58, 4: 64})


def write_etc_expansion_sheet(wb, enhanced_rows: List[Dict[str, Any]]) -> None:
    name = "ETC_EXPANSION_QUEUE"
    delete_if_exists(wb, name)
    ws = wb.create_sheet(name)
    headers = [
        "RANK",
        "OP_ID",
        "OP_NAME",
        "CATEGORY",
        "LANE",
        "AUTOMATION_SCORE_100",
        "SIGNAL_COUNT",
        "EXPANSION_LOGIC",
        "NEXT_ACTION",
    ]
    ws.append(headers)
    stylize_header(ws, 1, len(headers))

    by_lane: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in enhanced_rows:
        by_lane[str(row.get("LANE", "other"))].append(row)

    expanded: List[Dict[str, Any]] = []
    for lane, rows in by_lane.items():
        rows_sorted = sorted(rows, key=lambda r: safe_int(r.get("AUTOMATION_SCORE_100"), 0), reverse=True)
        seed = rows_sorted[:5]
        for row in seed:
            expanded.append(
                {
                    "OP_ID": row.get("OP_ID", ""),
                    "OP_NAME": row.get("OP_NAME", ""),
                    "CATEGORY": row.get("CATEGORY", ""),
                    "LANE": lane,
                    "AUTOMATION_SCORE_100": row.get("AUTOMATION_SCORE_100", 0),
                    "SIGNAL_COUNT": row.get("SIGNAL_COUNT", 0),
                    "EXPANSION_LOGIC": "ETC rule: expand adjacent ops in same lane/category.",
                    "NEXT_ACTION": row.get("NEXT_AUTOMATION_ACTION", ""),
                }
            )
    expanded.sort(key=lambda r: (safe_int(r.get("AUTOMATION_SCORE_100"), 0), safe_int(r.get("SIGNAL_COUNT"), 0)), reverse=True)
    for i, row in enumerate(expanded, start=1):
        ws.append(
            [
                i,
                row["OP_ID"],
                row["OP_NAME"],
                row["CATEGORY"],
                row["LANE"],
                row["AUTOMATION_SCORE_100"],
                row["SIGNAL_COUNT"],
                row["EXPANSION_LOGIC"],
                row["NEXT_ACTION"],
            ]
        )
    set_widths(ws, {1: 8, 2: 12, 3: 42, 4: 16, 5: 24, 6: 18, 7: 14, 8: 54, 9: 54})


def write_priority_exec_sheet(wb, priority_rows: List[Dict[str, Any]], enhanced_index: Dict[str, Dict[str, Any]]) -> None:
    name = "PRIORITY_AUTOMATION_EXEC"
    delete_if_exists(wb, name)
    ws = wb.create_sheet(name)
    headers = [
        "RANK",
        "OP_ID",
        "OP_NAME",
        "TIME_TO_FIRST_$",
        "READINESS",
        "LANE",
        "AUTOMATION_SCORE_100",
        "SIGNAL_COUNT",
        "BLOCKER_KEY",
        "APPROVAL_OK",
        "SOURCE_ARTIFACT",
        "NEXT_AUTOMATION_ACTION",
    ]
    ws.append(headers)
    stylize_header(ws, 1, len(headers))
    rows_sorted = sorted(priority_rows, key=lambda r: safe_float(r.get("RANK"), 999.0))
    for row in rows_sorted:
        op_id = str(row.get("OP_ID") or "").strip()
        e = enhanced_index.get(op_id, {})
        ws.append(
            [
                row.get("RANK", ""),
                op_id,
                row.get("OP_NAME", ""),
                row.get("TIME_TO_FIRST_$", ""),
                e.get("READINESS", "BUILD"),
                e.get("LANE", "other"),
                e.get("AUTOMATION_SCORE_100", 0),
                e.get("SIGNAL_COUNT", 0),
                e.get("BLOCKER_KEY", "NONE"),
                e.get("APPROVAL_OK", "NO"),
                e.get("SOURCE_ARTIFACT", "n/a"),
                e.get("NEXT_AUTOMATION_ACTION", "Run discovery"),
            ]
        )
    set_widths(ws, {1: 8, 2: 12, 3: 40, 4: 16, 5: 14, 6: 22, 7: 20, 8: 14, 9: 28, 10: 12, 11: 38, 12: 52})


def write_playbook_index_sheet(wb, playbook_rows: List[Dict[str, Any]]) -> None:
    name = "DEEP_PLAYBOOK_INDEX"
    delete_if_exists(wb, name)
    ws = wb.create_sheet(name)
    headers = [
        "PLAYBOOK_OP",
        "PLAYBOOK_TITLE",
        "SECTION",
        "ROW_TYPE",
        "STEP_INDEX",
        "LANE",
        "TEXT",
        "SOURCE_ROW",
    ]
    write_sheet(ws, headers, playbook_rows)
    set_widths(ws, {1: 14, 2: 40, 3: 24, 4: 16, 5: 12, 6: 22, 7: 120, 8: 10})


def write_alpha_index_sheet(wb, alpha_rows: List[Dict[str, Any]]) -> None:
    name = "ALPHA_THESIS_INDEX"
    delete_if_exists(wb, name)
    ws = wb.create_sheet(name)
    headers = [
        "ALPHA_ID",
        "OPPORTUNITY",
        "LANE",
        "BLOCKER_KEY",
        "EDGE_DURATION",
        "COMMAND_TEMPLATE",
        "WHY_LLM_EDGE",
        "SOURCE_ROW",
    ]
    write_sheet(ws, headers, alpha_rows)
    set_widths(ws, {1: 10, 2: 44, 3: 24, 4: 28, 5: 20, 6: 68, 7: 92, 8: 10})


def build_venture_automation_rows(
    priority_rows: List[Dict[str, Any]],
    enhanced_index: Dict[str, Dict[str, Any]],
    alpha_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for p in sorted(priority_rows, key=lambda r: safe_float(r.get("RANK"), 999.0)):
        op_id = clean_text(p.get("OP_ID"))
        if not op_id:
            continue
        e = enhanced_index.get(op_id, {})
        lane = clean_text(e.get("LANE") or "other")
        out.append(
            {
                "SOURCE": "PRIORITY_LAUNCH",
                "VENTURE_ID": op_id,
                "VENTURE_NAME": clean_text(p.get("OP_NAME")),
                "LANE": lane,
                "READINESS": clean_text(e.get("READINESS") or "BUILD"),
                "BLOCKER_KEY": clean_text(e.get("BLOCKER_KEY") or "NONE"),
                "AUTOMATION_SCORE_100": safe_int(e.get("AUTOMATION_SCORE_100"), 0),
                "SIGNAL_COUNT": safe_int(e.get("SIGNAL_COUNT"), 0),
                "COMMAND_TEMPLATE": command_template_for_lane(lane),
                "NEXT_ACTION": clean_text(e.get("NEXT_AUTOMATION_ACTION") or "Run discovery"),
            }
        )

    for a in alpha_rows:
        lane = clean_text(a.get("LANE") or "other")
        out.append(
            {
                "SOURCE": "LLM_ALPHA_THESIS",
                "VENTURE_ID": clean_text(a.get("ALPHA_ID")),
                "VENTURE_NAME": clean_text(a.get("OPPORTUNITY")),
                "LANE": lane,
                "READINESS": "EXPLORE",
                "BLOCKER_KEY": clean_text(a.get("BLOCKER_KEY") or "NONE"),
                "AUTOMATION_SCORE_100": 0,
                "SIGNAL_COUNT": 0,
                "COMMAND_TEMPLATE": clean_text(a.get("COMMAND_TEMPLATE") or command_template_for_lane(lane)),
                "NEXT_ACTION": "Open low-cost probe and validate unit economics",
            }
        )

    out.sort(key=lambda r: (safe_int(r.get("AUTOMATION_SCORE_100"), 0), safe_int(r.get("SIGNAL_COUNT"), 0)), reverse=True)
    return out


def write_venture_map_sheet(wb, venture_rows: List[Dict[str, Any]]) -> None:
    name = "VENTURE_AUTOMATION_MAP"
    delete_if_exists(wb, name)
    ws = wb.create_sheet(name)
    headers = [
        "SOURCE",
        "VENTURE_ID",
        "VENTURE_NAME",
        "LANE",
        "READINESS",
        "BLOCKER_KEY",
        "AUTOMATION_SCORE_100",
        "SIGNAL_COUNT",
        "COMMAND_TEMPLATE",
        "NEXT_ACTION",
    ]
    write_sheet(ws, headers, venture_rows)
    set_widths(ws, {1: 18, 2: 12, 3: 44, 4: 24, 5: 14, 6: 28, 7: 20, 8: 14, 9: 72, 10: 54})


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Enhance PRINTMAXX_MASTER_OPS.xlsx with live automation intelligence")
    ap.add_argument("--input", default=str(DEFAULT_INPUT), help="Path to source workbook")
    ap.add_argument("--output", default="", help="Path to enhanced workbook output")
    return ap.parse_args()


def main() -> int:
    args = parse_args()
    src = Path(args.input).expanduser().resolve()
    if not src.exists():
        print(f"master_ops_enhancer: missing source workbook: {src}")
        return 1

    date_tag = datetime.now().strftime("%Y-%m-%d")
    out = Path(args.output).expanduser().resolve() if args.output else src.parent / f"PRINTMAXX_MASTER_OPS_ENHANCED_{date_tag}.xlsx"
    shutil.copy2(src, out)

    wb = load_workbook(out)
    if "ALL OPS MASTER" not in wb.sheetnames:
        print("master_ops_enhancer: source workbook missing 'ALL OPS MASTER'")
        return 1
    if "PRIORITY LAUNCH" not in wb.sheetnames:
        print("master_ops_enhancer: source workbook missing 'PRIORITY LAUNCH'")
        return 1

    ops_rows = load_all_ops(wb["ALL OPS MASTER"])
    priority_rows = load_priority_launch(wb["PRIORITY LAUNCH"])
    pending = parse_pending_approvals(OPS_DIR / "HUMAN_LOOP_QUEUE.md")
    approved = parse_approved_keys(OPS_DIR / "HUMAN_APPROVALS.csv")
    lane_info = lane_metrics()

    enhanced_rows = build_enhanced_rows(ops_rows, pending, approved, lane_info)
    enhanced_index = {str(r.get("OP_ID", "")): r for r in enhanced_rows if str(r.get("OP_ID", "")).strip()}
    deep_playbook_rows = extract_deep_playbook_index(wb["DEEP PLAYBOOK"]) if "DEEP PLAYBOOK" in wb.sheetnames else []
    alpha_rows = extract_alpha_thesis_index(wb["LLM ALPHA THESIS"]) if "LLM ALPHA THESIS" in wb.sheetnames else []
    venture_rows = build_venture_automation_rows(priority_rows, enhanced_index, alpha_rows)

    delete_if_exists(wb, "AUTO_STATUS_LIVE")
    ws_status = wb.create_sheet("AUTO_STATUS_LIVE")
    status_headers = [
        "OP_ID",
        "CATEGORY",
        "OP_NAME",
        "LANE",
        "READINESS",
        "AUTOMATION_SCORE_100",
        "SIGNAL_COUNT",
        "BLOCKER_KEY",
        "APPROVAL_OK",
        "NEXT_AUTOMATION_ACTION",
        "SOURCE_ARTIFACT",
        "EXPECTED_PROFIT_HINT_USD",
        "UPDATED_AT",
    ]
    write_sheet(ws_status, status_headers, enhanced_rows)
    set_widths(ws_status, {1: 12, 2: 16, 3: 38, 4: 22, 5: 12, 6: 20, 7: 12, 8: 28, 9: 12, 10: 56, 11: 42, 12: 22, 13: 20})

    write_priority_exec_sheet(wb, priority_rows, enhanced_index)
    write_etc_expansion_sheet(wb, enhanced_rows)
    write_playbook_index_sheet(wb, deep_playbook_rows)
    write_alpha_index_sheet(wb, alpha_rows)
    write_venture_map_sheet(wb, venture_rows)
    write_system_evidence_sheet(wb, lane_info, pending, approved)

    wb.save(out)
    print("master_ops_enhancer: wrote")
    print(f"- {out}")
    print(f"- rows_scored={len(enhanced_rows)}")
    print(f"- playbook_rows_indexed={len(deep_playbook_rows)}")
    print(f"- alpha_rows_indexed={len(alpha_rows)}")
    print(f"- venture_rows={len(venture_rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
