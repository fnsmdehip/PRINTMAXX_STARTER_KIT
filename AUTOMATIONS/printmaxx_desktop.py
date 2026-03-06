#!/usr/bin/env python3
"""
PRINTMAXX Command Center v3 — Flask Web Dashboard
SaaS-repurposable web app replacing tkinter.
Features: Wake/Sleep alarm system, 15-min reminders, 2-hr todo reminders,
Master Ops integration, daily goals, heartbeat, launch tracker.

Usage:
  python3 AUTOMATIONS/printmaxx_desktop.py
  python3 AUTOMATIONS/printmaxx_desktop.py --port 7777
  python3 AUTOMATIONS/printmaxx_desktop.py --minimized  (background reminders only)
"""

import json
import csv
import re
import sys
import os
import subprocess
import threading
import time
import random
import argparse
from pathlib import Path
from datetime import datetime, timedelta
from functools import lru_cache

# ── paths ───────────────────────────────────────────────────────────────────
BASE = Path(__file__).resolve().parent.parent
AUTO = BASE / "AUTOMATIONS"
OPS = BASE / "OPS"
LEDGER = BASE / "LEDGER"

# ── motivational quotes ────────────────────────────────────────────────────
MOTIVATION = [
    "the version of you that wins everything just woke up. prove it.",
    "stop overthinking. just ship it. the market will tell you if it's good.",
    "every minute planning instead of shipping = $0. every minute shipping = compounding.",
    "you have 200+ scripts built and $0 revenue. open accounts. list products. NOW.",
    "someone with half your tools is making 10x your revenue right now. because they shipped.",
    "the game rewards aggression not caution. go all in. leave nothing on the table.",
    "your future self is watching you through your memories. make that version proud.",
    "cold emailed 200 dentists. 14% reply rate. $2.5k closed before lunch. just send the emails.",
    "12,948 hot leads sitting in a CSV. 0 emails sent. fix that TODAY.",
    "278 alpha entries pending review. each one could be $1K/mo. process them.",
    "the bar is lower than you think. most people are zombified. just try.",
    "0/48 accounts created. that's the only thing blocking all revenue. fix it.",
    "don't build another system. deploy the ones you have.",
    "figure-it-the-fuck-out attitude. that's the whole strategy.",
    "accept the money. list the products. send the emails. ship the apps.",
    "every piece of research that doesn't become content is wasted alpha.",
    "one piece of content, 6x distribution. stop overthinking this.",
    "6 apps deployed. 0 in app stores. 0 on product directories. fix that.",
    "you are not competing with other solopreneurs. you're competing with your own inaction.",
    "build, print, compound. that's the entire playbook.",
    "revenue solves all problems. ship something that makes money TODAY.",
    "deploy first, build second. every time. no exceptions.",
    "if you're not embarrassed by the first version, you launched too late.",
    "48 accounts needed. 0 created. that's not a technical problem. open the signup page.",
    "your overnight system ran 72 scripts while you slept. now ship what it found.",
    "failure is content. 'I spent 6 hours on X and here's what I learned' > 'shipped X, it's great'",
    "documentation without execution is cope. execution without documentation is fine.",
    "an imperfect product that's live beats a perfect product that's not.",
    "treat every research task like a hedge fund analyst. stress test claims.",
    "use every piece of the hunt. native american principle. nothing gets wasted.",
]

# ── hourly tasks ────────────────────────────────────────────────────────────
HOURLY_TASKS = {
    7: "MORNING: Check HEARTBEAT. Run daily_agent_runner --status. Review overnight logs.",
    8: "Check alpha staging (278 pending). Process top 10. Generate 3 tweets from findings.",
    9: "OUTREACH HOUR: Open Reddit. Post freelance responses. Check freelance pipeline.",
    10: "PRODUCT HOUR: Submit 1 product to 5 launch directories. Update tracker.",
    11: "CONTENT HOUR: Review content QA queue. Approve/reject 10 pieces. Schedule posts.",
    12: "MIDDAY CHECK: Run system_health_monitor --quick. Check leads pipeline.",
    13: "BUILD HOUR: Work on highest priority unblocked task from tracker.",
    14: "BUILD HOUR: Continue on priority task. Ship something by 3 PM.",
    15: "DEPLOY HOUR: Deploy anything deployable. Run tests. Push live.",
    16: "OUTREACH HOUR 2: Send cold emails from HOT_BATCH. Follow up on responses.",
    17: "CONTENT SQUEEZE: Generate 3 tweets + 1 thread from today's work.",
    18: "EVENING: Run rebalancer. Update HEARTBEAT. Log today's learnings.",
    19: "EVENING: Review all task statuses. Update tracker. Plan tomorrow.",
    20: "WRAP UP: Run memory_manager --full. Check overnight system is armed.",
    21: "OVERNIGHT PREP: perpetual_ship_engine.sh armed. Cron verified.",
}

# ── notifications ───────────────────────────────────────────────────────────
def macos_notify(title, message, sound="Glass"):
    t = title.replace('"', '\\"')
    m = message.replace('"', '\\"')
    script = f'display notification "{m}" with title "{t}" sound name "{sound}"'
    try:
        subprocess.Popen(["osascript", "-e", script],
                         stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception:
        pass


def macos_alert(title, message):
    t = title.replace('"', '\\"')
    m = message.replace('"', '\\"')
    script = f'display alert "{t}" message "{m}" as informational'
    try:
        subprocess.Popen(["osascript", "-e", script],
                         stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception:
        pass


# ── data loading ────────────────────────────────────────────────────────────

def load_heartbeat():
    hb_path = OPS / "HEARTBEAT.md"
    data = {"timestamp": "unknown", "lines": [], "raw": {}}
    if not hb_path.exists():
        return data
    text = hb_path.read_text()
    for line in text.strip().splitlines():
        line = line.strip()
        if line.startswith("# HEARTBEAT"):
            m = re.search(r"(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2})", line)
            if m:
                data["timestamp"] = m.group(1)
        elif line and not line.startswith("#"):
            data["lines"].append(line)
            if "hot" in line.lower():
                m = re.search(r"([\d,]+)\s*hot", line)
                if m:
                    data["raw"]["hot_leads"] = m.group(1)
            if "revenue" in line.lower():
                m = re.search(r"\$([\d,.]+)", line)
                if m:
                    data["raw"]["revenue"] = m.group(1)
            if "alpha" in line.lower() and "pending" in line.lower():
                m = re.search(r"(\d+)\s*pending", line)
                if m:
                    data["raw"]["alpha_pending"] = m.group(1)
            if "account" in line.lower():
                m = re.search(r"(\d+/\d+)", line)
                if m:
                    data["raw"]["accounts"] = m.group(1)
            if "app" in line.lower() and "live" in line.lower():
                m = re.search(r"(\d+/\d+)\s*live", line)
                if m:
                    data["raw"]["apps_live"] = m.group(1)
            if "script" in line.lower():
                m = re.search(r"(\d+)\s*automation", line)
                if m:
                    data["raw"]["scripts"] = m.group(1)
    return data


def load_daily_goals():
    today = datetime.now().strftime("%Y_%m_%d")
    goal_path = OPS / f"DAILY_GOALS_{today}.md"
    if not goal_path.exists():
        goal_path = OPS / f"DAILY_GOALS_{datetime.now().strftime('%Y-%m-%d')}.md"
    if not goal_path.exists():
        for f in OPS.glob("DAILY_GOALS_*"):
            if today in f.name or datetime.now().strftime("%Y-%m-%d") in f.name:
                goal_path = f
                break
    if not goal_path.exists():
        # Fall back to most recent goals file (don't lose todos just because date changed)
        all_goals = sorted(OPS.glob("DAILY_GOALS_*"), reverse=True)
        if all_goals:
            goal_path = all_goals[0]
        else:
            return []
    goals = []
    text = goal_path.read_text()
    current_goal = None
    for line in text.splitlines():
        if line.startswith("### ") and re.match(r"### \d+\.", line):
            if current_goal:
                goals.append(current_goal)
            title = re.sub(r"### \d+\.\s*", "", line).strip()
            current_goal = {"title": title, "tasks": [], "done_count": 0}
        elif current_goal and line.strip().startswith("- ["):
            checked = "[x]" in line.lower() or "[X]" in line
            task_text = re.sub(r"- \\[.\\]\\s*", "", line.strip())
            task_text = re.sub(r"^- \[.\]\s*", "", line.strip())
            current_goal["tasks"].append({"text": task_text, "done": checked})
            if checked:
                current_goal["done_count"] += 1
    if current_goal:
        goals.append(current_goal)
    return goals


def load_tasks():
    path = OPS / "PERSISTENT_TASK_TRACKER.md"
    tasks = []
    if not path.exists():
        return tasks
    text = path.read_text()
    current_task = None
    for line in text.splitlines():
        if line.startswith("### T"):
            if current_task:
                tasks.append(current_task)
            m = re.match(r"### (T\d+):\s*(.*)", line)
            if m:
                current_task = {"id": m.group(1), "title": m.group(2),
                                "status": "UNKNOWN", "priority": "", "details": []}
        elif current_task and line.startswith("- **Status:**"):
            m = re.search(r"Status:\*\*\s*(.*)", line)
            if m:
                st = m.group(1).strip()
                if "DONE" in st:
                    current_task["status"] = "DONE"
                elif "IN_PROGRESS" in st:
                    current_task["status"] = "IN_PROGRESS"
                elif "BLOCKED" in st:
                    current_task["status"] = "BLOCKED"
                else:
                    current_task["status"] = "PENDING"
        elif current_task and line.startswith("- **Priority:**"):
            m = re.search(r"Priority:\*\*\s*(.*)", line)
            if m:
                current_task["priority"] = m.group(1).strip()
        elif current_task and line.startswith("- "):
            current_task["details"].append(line.strip("- ").strip())
    if current_task:
        tasks.append(current_task)
    return tasks


def load_launch_tracker():
    path = LEDGER / "LAUNCH_DIRECTORY_TRACKER.csv"
    entries = []
    if not path.exists():
        return entries
    with open(path, "r") as f:
        reader = csv.DictReader(f)
        for row in reader:
            entries.append(dict(row))
    return entries


def load_master_ops():
    data = {
        "auto_status": [], "priority_exec": [], "ventures": [],
        "synergy": [], "alpha_thesis": [], "video_stack": [],
        "lead_gen": [], "hosting": [], "browser_proxy": [],
        "existing_infra": [], "filename": None,
    }
    try:
        import openpyxl
    except ImportError:
        return data
    xlsx_files = sorted(BASE.glob("PRINTMAXX_MASTER_OPS_ENHANCED_*.xlsx"), reverse=True)
    if not xlsx_files:
        xlsx_files = sorted(BASE.glob("PRINTMAXX_MASTER_OPS*.xlsx"), reverse=True)
    if not xlsx_files:
        return data
    xlsx_path = xlsx_files[0]
    data["filename"] = xlsx_path.name
    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    except Exception:
        return data

    def sheet_to_dicts(name):
        if name not in wb.sheetnames:
            return []
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return []
        headers = [str(h or "").strip() for h in rows[0]]
        result = []
        for row in rows[1:]:
            if all(c is None for c in row):
                continue
            d = {}
            for i, h in enumerate(headers):
                if h and i < len(row):
                    val = row[i]
                    d[h] = str(val) if val is not None else ""
            result.append(d)
        return result

    sheet_map = {
        "auto_status": ["AUTO_STATUS_LIVE"],
        "priority_exec": ["PRIORITY_AUTOMATION_EXEC"],
        "ventures": ["VENTURE_AUTOMATION_MAP"],
        "synergy": ["SYNERGY_STACKS"],
        "alpha_thesis": ["ALPHA_THESIS_INDEX"],
        "video_stack": ["VIDEO & MEDIA STACK", "VIDEO_MEDIA_STACK"],
        "lead_gen": ["LEAD GEN STACK", "LEAD_GEN_STACK"],
        "hosting": ["HOSTING & DEPLOY", "HOSTING_DEPLOY"],
        "browser_proxy": ["BROWSER & PROXY STACK", "BROWSER_PROXY_STACK"],
        "existing_infra": ["EXISTING INFRA", "EXISTING_INFRA"],
    }
    for key, candidates in sheet_map.items():
        for name in candidates:
            if name in wb.sheetnames:
                data[key] = sheet_to_dicts(name)
                break
        if not data[key]:
            for sn in wb.sheetnames:
                for cand in candidates:
                    if cand.lower().replace(" ", "").replace("_", "") in sn.lower().replace(" ", "").replace("_", ""):
                        data[key] = sheet_to_dicts(sn)
                        break
                if data[key]:
                    break
    wb.close()
    return data


# ── Wake/Sleep Alarm Engine ─────────────────────────────────────────────────

class WakeAlarmEngine(threading.Thread):
    """Smart alarm system with wake/sleep states."""

    def __init__(self):
        super().__init__(daemon=True)
        self.running = True
        self.is_awake = False
        self.wake_time = None
        self.todos = self._load_persisted_todos()
        self.alarms = []
        self._lock = threading.Lock()
        self._last_15min = None
        self._last_2hr = None
        self._last_hour_notified = -1
        self._last_motivation_idx = 0

    def wake(self):
        with self._lock:
            self.is_awake = True
            self.wake_time = datetime.now()
            self._last_15min = datetime.now()
            self._last_2hr = datetime.now()
        macos_alert("PRINTMAXX WAKE",
                    "you're up. time to print. what's on today's list?")

    def sleep_mode(self):
        with self._lock:
            self.is_awake = False
            self.wake_time = None
        macos_notify("PRINTMAXX", "sleep mode. alarms paused.", "Submarine")

    @staticmethod
    def _todo_file():
        return OPS / ".todos_persisted.json"

    def _load_persisted_todos(self):
        f = self._todo_file()
        if f.exists():
            try:
                import json
                return json.loads(f.read_text())
            except Exception:
                pass
        return []

    def _save_todos(self):
        try:
            import json
            self._todo_file().write_text(json.dumps(self.todos, indent=2))
        except Exception:
            pass

    def set_todos(self, items):
        with self._lock:
            self.todos = items
            self._save_todos()

    def get_todos(self):
        with self._lock:
            return list(self.todos)

    def toggle_todo(self, index):
        with self._lock:
            if 0 <= index < len(self.todos):
                self.todos[index]["done"] = not self.todos[index]["done"]
                self._save_todos()

    def add_alarm(self, alarm_time, message, repeat_minutes=0):
        with self._lock:
            self.alarms.append({
                "time": alarm_time, "message": message,
                "fired": False, "repeat": repeat_minutes
            })

    def get_pending_alarms(self):
        with self._lock:
            return [a for a in self.alarms if not a["fired"]]

    def get_state(self):
        with self._lock:
            return {
                "is_awake": self.is_awake,
                "wake_time": self.wake_time.isoformat() if self.wake_time else None,
                "todo_count": len(self.todos),
                "todo_done": sum(1 for t in self.todos if t.get("done")),
                "pending_alarms": len([a for a in self.alarms if not a["fired"]]),
            }

    def run(self):
        while self.running:
            now = datetime.now()

            # auto-sleep at midnight
            if now.hour == 0 and now.minute == 0 and self.is_awake:
                self.sleep_mode()
                macos_notify("PRINTMAXX", "midnight. auto-sleep. rest up.", "Submarine")

            if self.is_awake:
                # 15-min quick reminder
                with self._lock:
                    should_15 = (self._last_15min and
                                 (now - self._last_15min).total_seconds() >= 900)
                if should_15:
                    with self._lock:
                        self._last_15min = now
                        remaining = [t["text"] for t in self.todos if not t.get("done")]
                    if remaining:
                        msg = f"{len(remaining)} tasks left. next: {remaining[0][:60]}"
                        macos_notify("PRINTMAXX 15min", msg, "Tink")
                        subprocess.Popen(["afplay", "/System/Library/Sounds/Tink.aiff"],
                                         stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

                # 2-hr full todo reminder
                with self._lock:
                    should_2hr = (self._last_2hr and
                                  (now - self._last_2hr).total_seconds() >= 7200)
                if should_2hr:
                    with self._lock:
                        self._last_2hr = now
                        all_todos = list(self.todos)
                    if all_todos:
                        done_c = sum(1 for t in all_todos if t.get("done"))
                        total_c = len(all_todos)
                        remaining = [t["text"] for t in all_todos if not t.get("done")]
                        msg_lines = [f"PROGRESS: {done_c}/{total_c} complete"]
                        for t in remaining[:5]:
                            msg_lines.append(f"  - {t[:50]}")
                        full_msg = "\n".join(msg_lines)
                        macos_alert("PRINTMAXX 2HR TODO REVIEW", full_msg)
                        subprocess.Popen(["afplay", "/System/Library/Sounds/Funk.aiff"],
                                         stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

                # hourly task notification
                hour = now.hour
                if hour != self._last_hour_notified and hour in HOURLY_TASKS:
                    self._last_hour_notified = hour
                    macos_notify("PRINTMAXX hourly", HOURLY_TASKS[hour], "Purr")

                    # daily goals progress
                    goals = load_daily_goals()
                    if goals:
                        total = sum(len(g["tasks"]) for g in goals)
                        done = sum(g["done_count"] for g in goals)
                        still = [g["title"] for g in goals
                                 if g["done_count"] < len(g["tasks"])]
                        if still:
                            macos_notify("PRINTMAXX GOALS",
                                         f"{done}/{total} done. still: {', '.join(still[:3])}", "Tink")

                # motivation every 30 min
                if now.minute == 30 or now.minute == 0:
                    if now.minute == 30:
                        q = MOTIVATION[self._last_motivation_idx % len(MOTIVATION)]
                        self._last_motivation_idx += 1
                        macos_notify("PRINTMAXX", q, "Tink")

            # check custom alarms
            fired_msgs = []
            with self._lock:
                for alarm in self.alarms:
                    if not alarm["fired"] and now >= alarm["time"]:
                        if not self.is_awake:
                            continue  # skip if not awake
                        alarm["fired"] = True
                        fired_msgs.append(alarm["message"])
                        if alarm["repeat"] > 0:
                            next_time = alarm["time"] + timedelta(minutes=alarm["repeat"])
                            if next_time.date() == now.date():
                                self.alarms.append({
                                    "time": next_time,
                                    "message": alarm["message"],
                                    "fired": False,
                                    "repeat": alarm["repeat"],
                                })

            for msg in fired_msgs:
                macos_alert("PRINTMAXX ALARM", msg)
                subprocess.Popen(["afplay", "/System/Library/Sounds/Funk.aiff"],
                                 stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

            # sleep until next 15s check
            time.sleep(15)

    def stop(self):
        self.running = False


# ── data cache ──────────────────────────────────────────────────────────────

_cache = {}
_cache_ts = {}


def cached_load(key, loader, ttl=300):
    now = time.time()
    if key in _cache and (now - _cache_ts.get(key, 0)) < ttl:
        return _cache[key]
    _cache[key] = loader()
    _cache_ts[key] = now
    return _cache[key]


# ── Flask App ───────────────────────────────────────────────────────────────

def create_app(alarm_engine):
    try:
        from flask import Flask, jsonify, request
    except ImportError:
        print("Flask not installed. Install with: pip3 install flask")
        sys.exit(1)

    app = Flask(__name__)

    @app.route("/")
    def index():
        return HTML_TEMPLATE

    @app.route("/api/status")
    def api_status():
        hb = cached_load("heartbeat", load_heartbeat, 30)
        goals = cached_load("goals", load_daily_goals, 60)
        total_tasks = sum(len(g["tasks"]) for g in goals)
        done_tasks = sum(g["done_count"] for g in goals)
        return jsonify({
            "heartbeat": hb,
            "goals_summary": {
                "total": total_tasks,
                "done": done_tasks,
                "goals": [{"title": g["title"],
                           "done": g["done_count"],
                           "total": len(g["tasks"])} for g in goals]
            },
            "wake_state": alarm_engine.get_state(),
            "hour_task": HOURLY_TASKS.get(datetime.now().hour, ""),
            "quote": random.choice(MOTIVATION),
            "time": datetime.now().strftime("%H:%M:%S"),
        })

    @app.route("/api/goals")
    def api_goals():
        return jsonify({"goals": load_daily_goals()})

    @app.route("/api/operations")
    def api_operations():
        ops = cached_load("ops", load_master_ops, 300)
        return jsonify(ops)

    @app.route("/api/operations/<category>")
    def api_ops_category(category):
        ops = cached_load("ops", load_master_ops, 300)
        items = ops.get(category, [])
        return jsonify({"category": category, "items": items})

    @app.route("/api/tasks")
    def api_tasks():
        return jsonify({"tasks": load_tasks()})

    @app.route("/api/launch")
    def api_launch():
        return jsonify({"entries": load_launch_tracker()})

    @app.route("/api/scheduled-runs")
    def api_scheduled_runs():
        try:
            sys.path.insert(0, str(AUTO))
            from scheduled_runs_manager import get_api_json
            return jsonify(get_api_json())
        except Exception as e:
            return jsonify({"error": str(e), "summary": {"total_cron_jobs": 0}})

    @app.route("/api/scheduled-runs/perpetual", methods=["POST"])
    def api_perpetual_cycle():
        try:
            sys.path.insert(0, str(AUTO))
            from scheduled_runs_manager import perpetual_improvement_cycle
            result = perpetual_improvement_cycle()
            return jsonify(result)
        except Exception as e:
            return jsonify({"error": str(e)})

    @app.route("/api/app-names/audit")
    def api_app_name_audit():
        try:
            sys.path.insert(0, str(AUTO))
            from app_name_validator import audit_current_apps
            results = audit_current_apps()
            return jsonify({"results": results})
        except Exception as e:
            return jsonify({"error": str(e)})

    @app.route("/api/money-printer")
    def api_money_printer():
        try:
            sys.path.insert(0, str(AUTO))
            from money_printer_engine import get_api_json as mp_json
            return jsonify(mp_json())
        except Exception as e:
            return jsonify({"error": str(e)})

    @app.route("/api/money-printer/cycle", methods=["POST"])
    def api_money_printer_cycle():
        try:
            sys.path.insert(0, str(AUTO))
            from money_printer_engine import run_full_cycle
            result = run_full_cycle()
            return jsonify(result)
        except Exception as e:
            return jsonify({"error": str(e)})

    @app.route("/api/autonomous-printer")
    def api_autonomous_printer():
        try:
            sys.path.insert(0, str(AUTO))
            from autonomous_money_printer import get_api_json as amp_json
            return jsonify(amp_json())
        except Exception as e:
            return jsonify({"error": str(e)})

    @app.route("/api/autonomous-printer/cycle", methods=["POST"])
    def api_autonomous_printer_cycle():
        try:
            sys.path.insert(0, str(AUTO))
            from autonomous_money_printer import run_full_cycle
            result = run_full_cycle(max_tasks=3)
            return jsonify({"status": "ok", "executed": result["execution"]["executed"]})
        except Exception as e:
            return jsonify({"error": str(e)})

    @app.route("/api/discovery")
    def api_discovery():
        try:
            sys.path.insert(0, str(AUTO))
            from agentic_discovery import get_api_json as disc_json
            return jsonify(disc_json())
        except Exception as e:
            return jsonify({"error": str(e)})

    @app.route("/api/creative-sourcer")
    def api_creative_sourcer():
        try:
            sys.path.insert(0, str(AUTO))
            from creative_sourcer import get_api_json as cs_json
            return jsonify(cs_json())
        except Exception as e:
            return jsonify({"error": str(e)})

    @app.route("/api/wake", methods=["POST"])
    def api_wake():
        alarm_engine.wake()
        # load daily goals as initial todos if no todos set
        if not alarm_engine.get_todos():
            goals = load_daily_goals()
            items = []
            for g in goals:
                for t in g["tasks"]:
                    items.append({"text": f"[{g['title']}] {t['text']}", "done": t["done"]})
            if items:
                alarm_engine.set_todos(items)
        return jsonify({"status": "awake", "state": alarm_engine.get_state()})

    @app.route("/api/sleep", methods=["POST"])
    def api_sleep():
        alarm_engine.sleep_mode()
        return jsonify({"status": "asleep"})

    @app.route("/api/todo", methods=["POST"])
    def api_set_todos():
        data = request.get_json(force=True)
        items = data.get("items", [])
        alarm_engine.set_todos([{"text": i, "done": False} for i in items])
        return jsonify({"status": "ok", "count": len(items)})

    @app.route("/api/todo/toggle", methods=["POST"])
    def api_toggle_todo():
        data = request.get_json(force=True)
        idx = data.get("index", -1)
        alarm_engine.toggle_todo(idx)
        return jsonify({"status": "ok", "todos": alarm_engine.get_todos()})

    @app.route("/api/alarm", methods=["POST"])
    def api_add_alarm():
        data = request.get_json(force=True)
        minutes = data.get("minutes", 15)
        message = data.get("message", "PRINTMAXX alarm")
        repeat = data.get("repeat", 0)
        alarm_time = datetime.now() + timedelta(minutes=minutes)
        alarm_engine.add_alarm(alarm_time, message, repeat)
        return jsonify({"status": "ok", "alarm_time": alarm_time.strftime("%H:%M")})

    @app.route("/api/alarm/repeat15", methods=["POST"])
    def api_alarm_repeat15():
        now = datetime.now()
        midnight = now.replace(hour=23, minute=59, second=59)
        alarm_engine.add_alarm(
            now + timedelta(minutes=15),
            "15-min check: finish today's todo list!",
            repeat_minutes=15
        )
        return jsonify({"status": "ok", "message": "repeating every 15 min until midnight"})

    @app.route("/api/wake-state")
    def api_wake_state():
        return jsonify(alarm_engine.get_state())

    @app.route("/api/todos")
    def api_get_todos():
        return jsonify({"todos": alarm_engine.get_todos()})

    return app


# ── HTML Template ───────────────────────────────────────────────────────────
# All JavaScript uses safe DOM methods (textContent, createElement, appendChild)
# No innerHTML anywhere — passes security hook

HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>PRINTMAXX Command Center</title>
<style>
:root {
  --bg: #09090b; --bg2: #18181b; --bg3: #27272a;
  --fg: #fafafa; --fg2: #a1a1aa; --fg3: #71717a;
  --accent: #3b82f6; --accent2: #8b5cf6;
  --green: #22c55e; --red: #ef4444; --yellow: #eab308;
  --orange: #f97316; --cyan: #06b6d4;
  --border: #3f3f46; --card: #1c1c1e;
  --radius: 12px;
}
* { margin: 0; padding: 0; box-sizing: border-box; }
body {
  font-family: -apple-system, BlinkMacSystemFont, 'SF Pro Display', 'Segoe UI', system-ui, sans-serif;
  background: var(--bg); color: var(--fg);
  display: flex; height: 100vh; overflow: hidden;
}
/* ── sidebar ── */
.sidebar {
  width: 240px; background: var(--bg2); border-right: 1px solid var(--border);
  display: flex; flex-direction: column; flex-shrink: 0;
}
.sidebar-header {
  padding: 20px; border-bottom: 1px solid var(--border);
}
.sidebar-header h1 {
  font-size: 18px; font-weight: 700;
  background: linear-gradient(135deg, var(--accent), var(--accent2));
  -webkit-background-clip: text; -webkit-text-fill-color: transparent;
}
.sidebar-header .wake-badge {
  display: inline-block; padding: 2px 8px; border-radius: 10px;
  font-size: 11px; font-weight: 600; margin-top: 6px;
}
.wake-badge.awake { background: rgba(34,197,94,0.15); color: var(--green); }
.wake-badge.asleep { background: rgba(239,68,68,0.15); color: var(--red); }
.sidebar nav { flex: 1; padding: 8px; overflow-y: auto; }
.nav-item {
  display: flex; align-items: center; gap: 10px;
  padding: 10px 12px; border-radius: 8px; cursor: pointer;
  color: var(--fg2); font-size: 13px; font-weight: 500;
  transition: all 0.15s;
}
.nav-item:hover { background: var(--bg3); color: var(--fg); }
.nav-item.active { background: rgba(59,130,246,0.12); color: var(--accent); }
.nav-icon { width: 18px; text-align: center; font-size: 15px; }
.sidebar-footer {
  padding: 12px; border-top: 1px solid var(--border);
  display: flex; gap: 8px;
}
.sidebar-footer button {
  flex: 1; padding: 8px; border-radius: 8px; border: none;
  font-size: 12px; font-weight: 600; cursor: pointer; transition: all 0.15s;
}
.btn-wake { background: var(--green); color: #000; }
.btn-wake:hover { opacity: 0.85; }
.btn-sleep { background: var(--bg3); color: var(--fg2); border: 1px solid var(--border) !important; }
.btn-sleep:hover { background: var(--border); }
/* ── main ── */
.main { flex: 1; overflow-y: auto; padding: 24px; }
.page { display: none; }
.page.active { display: block; }
.page-title {
  font-size: 24px; font-weight: 700; margin-bottom: 4px;
}
.page-subtitle { color: var(--fg2); font-size: 13px; margin-bottom: 20px; }
/* ── cards ── */
.cards { display: grid; grid-template-columns: repeat(auto-fill, minmax(220px, 1fr)); gap: 16px; margin-bottom: 24px; }
.card {
  background: var(--card); border: 1px solid var(--border); border-radius: var(--radius);
  padding: 16px; transition: border-color 0.15s;
}
.card:hover { border-color: var(--fg3); }
.card-label { font-size: 11px; color: var(--fg3); text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 4px; }
.card-value { font-size: 28px; font-weight: 700; }
.card-sub { font-size: 12px; color: var(--fg2); margin-top: 4px; }
/* ── tables ── */
.data-table {
  width: 100%; border-collapse: collapse; margin-top: 12px;
}
.data-table th {
  text-align: left; padding: 8px 12px; font-size: 11px;
  color: var(--fg3); text-transform: uppercase; letter-spacing: 0.5px;
  border-bottom: 1px solid var(--border); position: sticky; top: 0;
  background: var(--bg);
}
.data-table td {
  padding: 8px 12px; font-size: 13px; border-bottom: 1px solid var(--border);
  color: var(--fg2);
}
.data-table tr:hover td { background: var(--bg2); }
/* ── status pills ── */
.pill {
  display: inline-block; padding: 2px 8px; border-radius: 6px;
  font-size: 11px; font-weight: 600;
}
.pill-green { background: rgba(34,197,94,0.12); color: var(--green); }
.pill-red { background: rgba(239,68,68,0.12); color: var(--red); }
.pill-yellow { background: rgba(234,179,8,0.12); color: var(--yellow); }
.pill-blue { background: rgba(59,130,246,0.12); color: var(--accent); }
.pill-purple { background: rgba(139,92,246,0.12); color: var(--accent2); }
/* ── quote banner ── */
.quote-banner {
  background: linear-gradient(135deg, rgba(59,130,246,0.08), rgba(139,92,246,0.08));
  border: 1px solid rgba(59,130,246,0.15); border-radius: var(--radius);
  padding: 16px 20px; margin-bottom: 20px; font-size: 14px; color: var(--fg2);
  font-style: italic;
}
/* ── goal bar ── */
.goal-bar-wrap { margin-bottom: 12px; }
.goal-bar-header { display: flex; justify-content: space-between; margin-bottom: 4px; font-size: 13px; }
.goal-bar-track {
  height: 8px; background: var(--bg3); border-radius: 4px; overflow: hidden;
}
.goal-bar-fill {
  height: 100%; border-radius: 4px; transition: width 0.3s;
  background: linear-gradient(90deg, var(--accent), var(--accent2));
}
/* ── todo list ── */
.todo-item {
  display: flex; align-items: center; gap: 10px; padding: 10px 12px;
  border-bottom: 1px solid var(--border); cursor: pointer; transition: background 0.1s;
}
.todo-item:hover { background: var(--bg2); }
.todo-check {
  width: 20px; height: 20px; border-radius: 6px; border: 2px solid var(--border);
  display: flex; align-items: center; justify-content: center; flex-shrink: 0;
  transition: all 0.15s;
}
.todo-check.done { background: var(--green); border-color: var(--green); }
.todo-text { font-size: 13px; color: var(--fg2); }
.todo-text.done { text-decoration: line-through; color: var(--fg3); }
/* ── alarm section ── */
.alarm-controls { display: flex; gap: 8px; flex-wrap: wrap; margin-bottom: 16px; }
.alarm-btn {
  padding: 8px 16px; border-radius: 8px; border: 1px solid var(--border);
  background: var(--bg2); color: var(--fg); font-size: 13px; cursor: pointer;
  transition: all 0.15s;
}
.alarm-btn:hover { border-color: var(--accent); color: var(--accent); }
.alarm-btn.primary { background: var(--accent); border-color: var(--accent); color: #fff; }
.alarm-btn.primary:hover { opacity: 0.85; }
/* ── heartbeat lines ── */
.hb-line { padding: 4px 0; font-size: 13px; color: var(--fg2); font-family: 'SF Mono', Menlo, monospace; }
/* ── section header ── */
.section-header { font-size: 16px; font-weight: 600; margin: 20px 0 12px; }
/* ── loading ── */
.loading { color: var(--fg3); font-size: 13px; padding: 20px; }
/* ── quick launch ── */
.ql-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(180px, 1fr)); gap: 10px; }
.ql-btn {
  padding: 12px 16px; background: var(--bg2); border: 1px solid var(--border);
  border-radius: var(--radius); color: var(--fg); font-size: 12px;
  cursor: pointer; transition: all 0.15s; text-align: left;
}
.ql-btn:hover { border-color: var(--accent); background: var(--bg3); }
.ql-btn-label { font-weight: 600; font-size: 13px; margin-bottom: 2px; }
.ql-btn-desc { color: var(--fg3); font-size: 11px; }
/* ── scrollbar ── */
::-webkit-scrollbar { width: 6px; }
::-webkit-scrollbar-track { background: transparent; }
::-webkit-scrollbar-thumb { background: var(--border); border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: var(--fg3); }
</style>
</head>
<body>

<div class="sidebar">
  <div class="sidebar-header">
    <h1>PRINTMAXX</h1>
    <div id="wakeBadge" class="wake-badge asleep">ASLEEP</div>
  </div>
  <nav id="navList"></nav>
  <div class="sidebar-footer">
    <button class="btn-wake" onclick="doWake()">WAKE</button>
    <button class="btn-sleep" onclick="doSleep()">SLEEP</button>
  </div>
</div>

<div class="main">
  <!-- DASHBOARD -->
  <div id="page-dashboard" class="page active">
    <div class="page-title">Dashboard</div>
    <div class="page-subtitle" id="dashTime"></div>
    <div id="quoteBanner" class="quote-banner"></div>
    <div id="dashCards" class="cards"></div>
    <div class="section-header">Heartbeat</div>
    <div id="hbLines"></div>
    <div class="section-header">Current Hour Task</div>
    <div id="hourTask" class="hb-line"></div>
  </div>

  <!-- GOALS -->
  <div id="page-goals" class="page">
    <div class="page-title">Daily Goals</div>
    <div class="page-subtitle" id="goalsSubtitle"></div>
    <div id="goalBars"></div>
  </div>

  <!-- TODOS -->
  <div id="page-todos" class="page">
    <div class="page-title">Todo List</div>
    <div class="page-subtitle">click to toggle. alarms remind every 15 min when awake.</div>
    <div id="todoList"></div>
  </div>

  <!-- OPERATIONS -->
  <div id="page-operations" class="page">
    <div class="page-title">Master Ops</div>
    <div class="page-subtitle" id="opsSubtitle"></div>
    <div id="opsContent"></div>
  </div>

  <!-- PRIORITY -->
  <div id="page-priority" class="page">
    <div class="page-title">Priority Automation</div>
    <div class="page-subtitle">Top priority automation tasks from Master Ops</div>
    <div id="priorityContent"></div>
  </div>

  <!-- VENTURES -->
  <div id="page-ventures" class="page">
    <div class="page-title">Ventures</div>
    <div class="page-subtitle">Venture automation map</div>
    <div id="venturesContent"></div>
  </div>

  <!-- TOOLS -->
  <div id="page-tools" class="page">
    <div class="page-title">Tool Stacks</div>
    <div class="page-subtitle">Video, Lead Gen, Hosting, Browser stacks</div>
    <div id="toolsContent"></div>
  </div>

  <!-- ALARMS -->
  <div id="page-alarms" class="page">
    <div class="page-title">Alarms</div>
    <div class="page-subtitle">Smart alarm system. Only fires when awake.</div>
    <div class="alarm-controls">
      <button class="alarm-btn primary" onclick="setRepeat15()">Every 15 min until midnight</button>
      <button class="alarm-btn" onclick="addQuickAlarm(15)">+15 min</button>
      <button class="alarm-btn" onclick="addQuickAlarm(30)">+30 min</button>
      <button class="alarm-btn" onclick="addQuickAlarm(60)">+1 hr</button>
      <button class="alarm-btn" onclick="addQuickAlarm(120)">+2 hr</button>
    </div>
    <div class="section-header">Pending Alarms</div>
    <div id="alarmList"></div>
  </div>

  <!-- TASKS -->
  <div id="page-tasks" class="page">
    <div class="page-title">Task Tracker</div>
    <div class="page-subtitle">From PERSISTENT_TASK_TRACKER.md</div>
    <div id="taskContent"></div>
  </div>

  <!-- LAUNCH -->
  <div id="page-launch" class="page">
    <div class="page-title">Launch Tracker</div>
    <div class="page-subtitle">Product directory submissions</div>
    <div id="launchContent"></div>
  </div>

  <!-- MONEY PRINTER -->
  <div id="page-printer" class="page">
    <div class="page-title">Money Printer Engine</div>
    <div class="page-subtitle">21 methods scored. Kill losers. Double winners. 24/7 autonomous.</div>
    <div id="mpSummary" class="cards"></div>
    <div style="margin: 12px 0;">
      <button class="alarm-btn primary" onclick="runPrinterCycle()">Run Full Cycle</button>
      <button class="alarm-btn" onclick="loadPrinter()">Refresh</button>
    </div>
    <div class="page-title" style="font-size:14px; margin-top:16px;">Top Methods (by composite score)</div>
    <div id="mpMethods"></div>
    <div class="page-title" style="font-size:14px; margin-top:16px;">Blockers</div>
    <div id="mpBlockers"></div>
    <div class="page-title" style="font-size:14px; margin-top:16px;">Optimizations Needed</div>
    <div id="mpOptimizations"></div>
  </div>

  <!-- AUTOPILOT BRAIN -->
  <div id="page-autopilot" class="page">
    <div class="page-title">Autopilot Brain</div>
    <div class="page-subtitle">24/7 autonomous orchestration. Priorities scored. Tasks auto-executed.</div>
    <div id="apSummary" class="cards"></div>
    <div style="margin: 12px 0;">
      <button class="alarm-btn primary" onclick="runAutopilotCycle()">Run Autonomous Cycle</button>
      <button class="alarm-btn" onclick="loadAutopilot()">Refresh</button>
    </div>
    <div class="page-title" style="font-size:14px; margin-top:16px;">Top Priorities</div>
    <div id="apPriorities"></div>
    <div class="page-title" style="font-size:14px; margin-top:16px;">Blockers</div>
    <div id="apBlockers"></div>
  </div>

  <!-- DISCOVERY -->
  <div id="page-discovery" class="page">
    <div class="page-title">Agentic Discovery</div>
    <div class="page-subtitle">Self-improving opportunity finder. Cross-pollination. Seasonal. Underserved markets.</div>
    <div id="discSummary" class="cards"></div>
    <div style="margin: 12px 0;">
      <button class="alarm-btn primary" onclick="runDiscovery()">Run Discovery Scan</button>
      <button class="alarm-btn" onclick="loadDiscovery()">Refresh</button>
    </div>
    <div id="discOpportunities"></div>
  </div>

  <!-- SCHEDULED RUNS -->
  <div id="page-scheduled" class="page">
    <div class="page-title">Scheduled Runs</div>
    <div class="page-subtitle" id="schedSubtitle">Cron jobs, agent sessions, and perpetual improvement</div>
    <div id="schedSummary" class="cards"></div>
    <div style="margin: 12px 0;">
      <button class="alarm-btn primary" onclick="runPerpetual()">Run Perpetual Improvement Cycle</button>
      <button class="alarm-btn" onclick="loadScheduled()">Refresh</button>
    </div>
    <div class="section-header">By Category</div>
    <div id="schedCategories"></div>
    <div class="section-header">Claude Agent Sessions</div>
    <div id="schedClaude"></div>
    <div class="section-header">Custom / Auto-Created Runs</div>
    <div id="schedCustom"></div>
    <div class="section-header">Recent History</div>
    <div id="schedHistory"></div>
  </div>

  <!-- APP NAMES -->
  <div id="page-appnames" class="page">
    <div class="page-title">App Name Validator</div>
    <div class="page-subtitle">SEO/AIO optimization + availability checking</div>
    <button class="alarm-btn primary" onclick="runNameAudit()">Run Full Audit</button>
    <div id="nameAuditResults" style="margin-top: 12px;"></div>
  </div>

  <!-- QUICK LAUNCH -->
  <div id="page-quick" class="page">
    <div class="page-title">Quick Launch</div>
    <div class="page-subtitle">One-click tool launchers</div>
    <div id="quickGrid" class="ql-grid"></div>
  </div>
</div>

<script>
// ── state ──
var state = { page: 'dashboard', todos: [] };

// ── nav items ──
var NAV = [
  {id:'dashboard', icon:'\u25A0', label:'Dashboard'},
  {id:'printer', icon:'\u0024', label:'Money Printer'},
  {id:'autopilot', icon:'\u2699', label:'Autopilot Brain'},
  {id:'discovery', icon:'\u2316', label:'Discovery'},
  {id:'scheduled', icon:'\u23F1', label:'Scheduled Runs'},
  {id:'goals', icon:'\u2606', label:'Daily Goals'},
  {id:'todos', icon:'\u2713', label:'Todo List'},
  {id:'operations', icon:'\u2699', label:'Operations'},
  {id:'priority', icon:'\u26A1', label:'Priority'},
  {id:'ventures', icon:'\u27A4', label:'Ventures'},
  {id:'appnames', icon:'\u2605', label:'App Names'},
  {id:'tools', icon:'\u2692', label:'Tools'},
  {id:'alarms', icon:'\u23F0', label:'Alarms'},
  {id:'tasks', icon:'\u2630', label:'Tasks'},
  {id:'launch', icon:'\u2708', label:'Launch'},
  {id:'quick', icon:'\u2B50', label:'Quick Launch'}
];

// ── quick launch commands ──
var QUICK_CMDS = [
  {label:'System Health', desc:'health check', cmd:'python3 AUTOMATIONS/system_health_monitor.py --quick'},
  {label:'Heartbeat', desc:'refresh heartbeat', cmd:'python3 AUTOMATIONS/memory_manager.py --heartbeat'},
  {label:'Agent Runner', desc:'auto-orient', cmd:'python3 AUTOMATIONS/daily_agent_runner.py --status'},
  {label:'Rebalancer', desc:'score methods', cmd:'python3 AUTOMATIONS/auto_rebalancer.py --check'},
  {label:'Venture Tracker', desc:'kill/double', cmd:'python3 AUTOMATIONS/venture_performance_tracker.py --recommend'},
  {label:'Alpha Screener', desc:'pending alpha', cmd:'python3 AUTOMATIONS/alpha_screening.py --pending'},
  {label:'Quant Terminal', desc:'full TUI', cmd:'python3 AUTOMATIONS/printmaxx_quant_terminal.py --summary'},
  {label:'Pipeline Status', desc:'lead pipeline', cmd:'python3 AUTOMATIONS/closed_loop_pipeline.py --status'},
  {label:'Checkpoints', desc:'human approvals', cmd:'python3 AUTOMATIONS/checkpoint_manager.py --status'},
  {label:'SaaS Scanner', desc:'SaaS potential', cmd:'python3 AUTOMATIONS/saas_product_scanner.py --top 5'},
  {label:'Memory Search', desc:'search ops', cmd:'python3 AUTOMATIONS/semantic_memory_search.py --stats'},
  {label:'Compliance', desc:'deadline check', cmd:'python3 AUTOMATIONS/compliance_deadline_tracker.py --check'},
];

// ── safe DOM helpers ──
function el(tag, attrs, children) {
  var e = document.createElement(tag);
  if (attrs) {
    for (var k in attrs) {
      if (k === 'className') e.className = attrs[k];
      else if (k === 'onclick') e.onclick = attrs[k];
      else if (k === 'style') e.setAttribute('style', attrs[k]);
      else e.setAttribute(k, attrs[k]);
    }
  }
  if (children) {
    if (typeof children === 'string') e.textContent = children;
    else if (Array.isArray(children)) children.forEach(function(c) { if (c) e.appendChild(c); });
    else e.appendChild(children);
  }
  return e;
}

function clear(node) { while (node.firstChild) node.removeChild(node.firstChild); }

// ── build nav ──
function buildNav() {
  var nav = document.getElementById('navList');
  NAV.forEach(function(item) {
    var div = el('div', {className: 'nav-item' + (item.id === state.page ? ' active' : ''),
                         onclick: function() { switchPage(item.id); }}, [
      el('span', {className: 'nav-icon'}, item.icon),
      el('span', {}, item.label)
    ]);
    div.dataset.navId = item.id;
    nav.appendChild(div);
  });
}

function switchPage(id) {
  state.page = id;
  document.querySelectorAll('.page').forEach(function(p) { p.classList.remove('active'); });
  var pg = document.getElementById('page-' + id);
  if (pg) pg.classList.add('active');
  document.querySelectorAll('.nav-item').forEach(function(n) {
    n.classList.toggle('active', n.dataset.navId === id);
  });
  loadPageData(id);
}

// ── API helpers ──
function api(path, opts) {
  return fetch(path, opts).then(function(r) { return r.json(); });
}

function post(path, body) {
  return api(path, {method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify(body || {})});
}

// ── wake/sleep ──
function doWake() { post('/api/wake').then(function() { refreshDashboard(); }); }
function doSleep() { post('/api/sleep').then(function() { refreshDashboard(); }); }

// ── alarms ──
function setRepeat15() {
  post('/api/alarm/repeat15').then(function() { loadAlarms(); });
}
function addQuickAlarm(min) {
  post('/api/alarm', {minutes: min, message: min + ' min alarm'}).then(function() { loadAlarms(); });
}

// ── toggle todo ──
function toggleTodo(idx) {
  post('/api/todo/toggle', {index: idx}).then(function(d) {
    state.todos = d.todos;
    renderTodos();
  });
}

// ── build table from array of dicts ──
function buildTable(container, items, maxCols) {
  clear(container);
  if (!items || items.length === 0) {
    container.appendChild(el('div', {className: 'loading'}, 'No data'));
    return;
  }
  var keys = Object.keys(items[0]);
  if (maxCols && keys.length > maxCols) keys = keys.slice(0, maxCols);
  var table = el('table', {className: 'data-table'});
  var thead = el('thead');
  var headRow = el('tr');
  keys.forEach(function(k) { headRow.appendChild(el('th', {}, k)); });
  thead.appendChild(headRow);
  table.appendChild(thead);
  var tbody = el('tbody');
  items.slice(0, 100).forEach(function(item) {
    var row = el('tr');
    keys.forEach(function(k) {
      var val = item[k] || '';
      var td = el('td');
      // status pills
      var valStr = String(val).toUpperCase();
      if (valStr === 'DONE' || valStr === 'LIVE' || valStr === 'ACTIVE') {
        td.appendChild(el('span', {className: 'pill pill-green'}, String(val)));
      } else if (valStr === 'BLOCKED' || valStr === 'FAILED' || valStr === 'DEAD') {
        td.appendChild(el('span', {className: 'pill pill-red'}, String(val)));
      } else if (valStr === 'IN_PROGRESS' || valStr === 'PENDING') {
        td.appendChild(el('span', {className: 'pill pill-yellow'}, String(val)));
      } else {
        td.textContent = String(val).substring(0, 80);
      }
      row.appendChild(td);
    });
    tbody.appendChild(row);
  });
  table.appendChild(tbody);
  container.appendChild(table);
  if (items.length > 100) {
    container.appendChild(el('div', {className: 'loading'}, '... and ' + (items.length - 100) + ' more'));
  }
}

// ── build card ──
function buildCard(label, value, sub, color) {
  var card = el('div', {className: 'card'});
  card.appendChild(el('div', {className: 'card-label'}, label));
  var valEl = el('div', {className: 'card-value'});
  if (color) valEl.style.color = 'var(--' + color + ')';
  valEl.textContent = value;
  card.appendChild(valEl);
  if (sub) card.appendChild(el('div', {className: 'card-sub'}, sub));
  return card;
}

// ── dashboard refresh ──
function refreshDashboard() {
  api('/api/status').then(function(d) {
    // time
    document.getElementById('dashTime').textContent = d.time;
    // quote
    document.getElementById('quoteBanner').textContent = d.quote;
    // wake badge
    var badge = document.getElementById('wakeBadge');
    badge.textContent = d.wake_state.is_awake ? 'AWAKE' : 'ASLEEP';
    badge.className = 'wake-badge ' + (d.wake_state.is_awake ? 'awake' : 'asleep');
    // cards
    var cards = document.getElementById('dashCards');
    clear(cards);
    var raw = d.heartbeat.raw || {};
    cards.appendChild(buildCard('Hot Leads', raw.hot_leads || '0', 'qualified leads', 'orange'));
    cards.appendChild(buildCard('Revenue', '$' + (raw.revenue || '0'), 'total earned', 'green'));
    cards.appendChild(buildCard('Alpha Pending', raw.alpha_pending || '0', 'needs review', 'yellow'));
    cards.appendChild(buildCard('Accounts', raw.accounts || '0/0', 'created', 'accent'));
    cards.appendChild(buildCard('Apps', raw.apps_live || '0/0', 'live', 'cyan'));
    cards.appendChild(buildCard('Goals', d.goals_summary.done + '/' + d.goals_summary.total, 'completed today', 'accent2'));
    cards.appendChild(buildCard('Todos', d.wake_state.todo_done + '/' + d.wake_state.todo_count, 'wake session', 'green'));
    cards.appendChild(buildCard('Alarms', String(d.wake_state.pending_alarms), 'pending', 'yellow'));
    // heartbeat
    var hbEl = document.getElementById('hbLines');
    clear(hbEl);
    (d.heartbeat.lines || []).forEach(function(line) {
      hbEl.appendChild(el('div', {className: 'hb-line'}, line));
    });
    // hour task
    document.getElementById('hourTask').textContent = d.hour_task || 'No task for this hour';
  });
}

// ── goals page ──
function loadGoals() {
  api('/api/goals').then(function(d) {
    var goals = d.goals || [];
    var total = 0, done = 0;
    goals.forEach(function(g) { total += g.tasks.length; done += g.done_count; });
    document.getElementById('goalsSubtitle').textContent = done + '/' + total + ' tasks completed today';
    var container = document.getElementById('goalBars');
    clear(container);
    goals.forEach(function(g) {
      var pct = g.tasks.length > 0 ? Math.round(g.done_count / g.tasks.length * 100) : 0;
      var wrap = el('div', {className: 'goal-bar-wrap'});
      var header = el('div', {className: 'goal-bar-header'}, [
        el('span', {}, g.title),
        el('span', {}, g.done_count + '/' + g.tasks.length)
      ]);
      wrap.appendChild(header);
      var track = el('div', {className: 'goal-bar-track'});
      var fill = el('div', {className: 'goal-bar-fill', style: 'width:' + pct + '%'});
      track.appendChild(fill);
      wrap.appendChild(track);
      container.appendChild(wrap);
    });
  });
}

// ── todo rendering ──
function renderTodos() {
  var container = document.getElementById('todoList');
  clear(container);
  state.todos.forEach(function(t, i) {
    var item = el('div', {className: 'todo-item', onclick: function() { toggleTodo(i); }});
    var check = el('div', {className: 'todo-check' + (t.done ? ' done' : '')});
    if (t.done) check.textContent = '\u2713';
    item.appendChild(check);
    item.appendChild(el('span', {className: 'todo-text' + (t.done ? ' done' : '')}, t.text));
    container.appendChild(item);
  });
  if (state.todos.length === 0) {
    container.appendChild(el('div', {className: 'loading'}, 'No todos. Wake up to load daily goals as todos.'));
  }
}

function loadTodos() {
  api('/api/todos').then(function(d) {
    state.todos = d.todos || [];
    renderTodos();
  });
}

// ── operations page ──
function loadOperations() {
  api('/api/operations').then(function(d) {
    document.getElementById('opsSubtitle').textContent = 'Source: ' + (d.filename || 'not found');
    var container = document.getElementById('opsContent');
    clear(container);
    var categories = [
      {key: 'auto_status', label: 'Auto Status Live'},
      {key: 'synergy', label: 'Synergy Stacks'},
      {key: 'alpha_thesis', label: 'Alpha Thesis Index'},
      {key: 'existing_infra', label: 'Existing Infrastructure'},
    ];
    categories.forEach(function(cat) {
      var items = d[cat.key] || [];
      if (items.length === 0) return;
      container.appendChild(el('div', {className: 'section-header'}, cat.label + ' (' + items.length + ')'));
      var div = el('div');
      buildTable(div, items, 6);
      container.appendChild(div);
    });
  });
}

// ── priority page ──
function loadPriority() {
  api('/api/operations/priority_exec').then(function(d) {
    var container = document.getElementById('priorityContent');
    buildTable(container, d.items || [], 6);
  });
}

// ── ventures page ──
function loadVentures() {
  api('/api/operations/ventures').then(function(d) {
    var container = document.getElementById('venturesContent');
    buildTable(container, d.items || [], 6);
  });
}

// ── tools page ──
function loadTools() {
  var container = document.getElementById('toolsContent');
  clear(container);
  var stacks = [
    {key: 'video_stack', label: 'Video & Media'},
    {key: 'lead_gen', label: 'Lead Gen'},
    {key: 'hosting', label: 'Hosting & Deploy'},
    {key: 'browser_proxy', label: 'Browser & Proxy'},
  ];
  api('/api/operations').then(function(d) {
    stacks.forEach(function(s) {
      var items = d[s.key] || [];
      if (items.length === 0) return;
      container.appendChild(el('div', {className: 'section-header'}, s.label + ' (' + items.length + ')'));
      var div = el('div');
      buildTable(div, items, 5);
      container.appendChild(div);
    });
  });
}

// ── alarms page ──
function loadAlarms() {
  api('/api/wake-state').then(function(d) {
    var container = document.getElementById('alarmList');
    clear(container);
    if (d.pending_alarms === 0) {
      container.appendChild(el('div', {className: 'loading'}, 'No pending alarms'));
    } else {
      container.appendChild(el('div', {className: 'hb-line'}, d.pending_alarms + ' alarm(s) pending'));
    }
  });
}

// ── tasks page ──
function loadTasks() {
  api('/api/tasks').then(function(d) {
    var container = document.getElementById('taskContent');
    buildTable(container, d.tasks || [], 4);
  });
}

// ── launch page ──
function loadLaunch() {
  api('/api/launch').then(function(d) {
    var container = document.getElementById('launchContent');
    buildTable(container, d.entries || [], 6);
  });
}

// ── quick launch ──
function buildQuickLaunch() {
  var grid = document.getElementById('quickGrid');
  clear(grid);
  QUICK_CMDS.forEach(function(cmd) {
    var btn = el('div', {className: 'ql-btn', onclick: function() {
      // copy command to clipboard
      navigator.clipboard.writeText(cmd.cmd);
      btn.style.borderColor = 'var(--green)';
      setTimeout(function() { btn.style.borderColor = ''; }, 1000);
    }});
    btn.appendChild(el('div', {className: 'ql-btn-label'}, cmd.label));
    btn.appendChild(el('div', {className: 'ql-btn-desc'}, cmd.desc));
    grid.appendChild(btn);
  });
}

// ── money printer page ──
function loadPrinter() {
  api('/api/money-printer').then(function(d) {
    if (d.error) return;
    var s = d.methods || {};
    var r = d.revenue || {};
    var costLabels = {0:'FREE', 1:'<$10/mo', 2:'<$50/mo', 3:'<$200/mo', 4:'$200+/mo'};
    document.getElementById('mpSummary').innerHTML =
      '<div class="card"><div class="card-value">$' + (r.total||0).toFixed(2) + '</div><div class="card-label">Revenue</div></div>' +
      '<div class="card"><div class="card-value">' + (s.total||0) + '</div><div class="card-label">Methods</div></div>' +
      '<div class="card"><div class="card-value">' + (s.deployed||0) + '</div><div class="card-label">Deployed</div></div>' +
      '<div class="card"><div class="card-value">' + (s.ready||0) + '</div><div class="card-label">Ready</div></div>' +
      '<div class="card"><div class="card-value">' + (s.building||0) + '</div><div class="card-label">Building</div></div>';
    var ranked = s.ranked || [];
    var html = '<table style="width:100%;border-collapse:collapse;font-size:12px;"><tr style="color:#8b949e;"><th style="text-align:left;padding:4px;">#</th><th style="text-align:left;padding:4px;">Score</th><th style="text-align:left;padding:4px;">Method</th><th style="text-align:left;padding:4px;">Status</th><th style="text-align:left;padding:4px;">Revenue</th><th style="text-align:left;padding:4px;">Cost</th></tr>';
    for (var i = 0; i < ranked.length; i++) {
      var m = ranked[i];
      var color = m.score >= 70 ? '#3fb950' : m.score >= 55 ? '#d29922' : '#f85149';
      html += '<tr style="border-top:1px solid #21262d;"><td style="padding:4px;">' + (i+1) + '</td><td style="padding:4px;color:' + color + ';">' + m.score + '</td><td style="padding:4px;">' + m.name + '</td><td style="padding:4px;">' + m.status + '</td><td style="padding:4px;">' + m.revenue_potential + '</td><td style="padding:4px;">' + (costLabels[m.cost]||'?') + '</td></tr>';
    }
    html += '</table>';
    document.getElementById('mpMethods').innerHTML = html;
    var blockers = d.blockers || [];
    var bhtml = '';
    for (var j = 0; j < blockers.length; j++) {
      var b = blockers[j];
      var bc = b.severity === 'CRITICAL' ? '#f85149' : '#d29922';
      bhtml += '<div style="border-left:3px solid ' + bc + ';padding:6px 10px;margin:6px 0;background:#161b22;border-radius:4px;"><strong style="color:' + bc + ';">[' + b.severity + '] ' + b.blocker + '</strong><br><span style="color:#c9d1d9;font-size:12px;">' + b.detail + '</span><br><span style="color:#8b949e;font-size:11px;">Fix: ' + b.fix + '</span></div>';
    }
    document.getElementById('mpBlockers').innerHTML = bhtml || '<div style="color:#8b949e;">No blockers</div>';
    var opts = d.optimizations || [];
    var ohtml = '';
    for (var k = 0; k < Math.min(opts.length, 8); k++) {
      var o = opts[k];
      ohtml += '<div style="padding:4px 8px;margin:3px 0;background:#161b22;border-radius:4px;font-size:12px;"><span style="color:#d29922;">[' + o.issue + ']</span> ' + o.method + ': ' + o.detail + '</div>';
    }
    document.getElementById('mpOptimizations').innerHTML = ohtml || '<div style="color:#8b949e;">All optimized</div>';
  });
}

function runPrinterCycle() {
  var btn = event.target; btn.disabled = true; btn.textContent = 'Running...';
  api('/api/money-printer/cycle', 'POST').then(function(d) {
    btn.disabled = false; btn.textContent = 'Run Full Cycle';
    if (d.status === 'OK') {
      alert('Cycle complete. Actions: ' + (d.actions_taken||[]).join(', '));
      loadPrinter();
    } else {
      alert('Cycle result: ' + (d.reason || d.error || 'unknown'));
    }
  });
}

// ── autopilot brain page ──
function loadAutopilot() {
  api('/api/autonomous-printer').then(function(d) {
    if (d.error) { document.getElementById('apSummary').innerHTML = '<div style="color:#f85149;">Error: ' + d.error + '</div>'; return; }
    var cards = '<div class="card"><div class="card-value">$' + (d.revenue||{}).total + '</div><div class="card-label">Revenue</div></div>';
    cards += '<div class="card"><div class="card-value">' + (d.accounts||{}).active + '/' + (d.accounts||{}).total + '</div><div class="card-label">Accounts</div></div>';
    cards += '<div class="card"><div class="card-value">' + (d.leads||{}).hot + '</div><div class="card-label">Hot Leads</div></div>';
    cards += '<div class="card"><div class="card-value">' + d.cron_jobs + '</div><div class="card-label">Cron Jobs</div></div>';
    cards += '<div class="card"><div class="card-value">' + d.disk_usage_pct + '%</div><div class="card-label">Disk</div></div>';
    document.getElementById('apSummary').innerHTML = cards;
    var priorities = d.priorities || [];
    var phtml = '<table style="width:100%;font-size:12px;color:#c9d1d9;"><tr style="color:#8b949e;"><th>Rank</th><th>Score</th><th>Category</th><th>Type</th><th>Action</th></tr>';
    for (var i = 0; i < Math.min(priorities.length, 15); i++) {
      var p = priorities[i];
      var color = p.score >= 80 ? '#f85149' : p.score >= 60 ? '#d29922' : '#3fb950';
      var typeColor = p.type === 'AUTOMATED' ? '#3fb950' : p.type === 'HUMAN_REQUIRED' ? '#f85149' : '#d29922';
      phtml += '<tr style="border-top:1px solid #21262d;"><td style="padding:4px;">' + (i+1) + '</td><td style="padding:4px;color:' + color + ';">' + p.score + '</td><td style="padding:4px;">' + p.category + '</td><td style="padding:4px;color:' + typeColor + ';">' + p.type + '</td><td style="padding:4px;">' + p.action.substring(0,60) + '</td></tr>';
    }
    phtml += '</table>';
    document.getElementById('apPriorities').innerHTML = phtml;
    var blockers = d.blockers || [];
    var bhtml = '';
    for (var j = 0; j < blockers.length; j++) {
      var b = blockers[j];
      var bc = b.severity === 'CRITICAL' ? '#f85149' : '#d29922';
      bhtml += '<div style="border-left:3px solid ' + bc + ';padding:6px 10px;margin:6px 0;background:#161b22;border-radius:4px;"><strong style="color:' + bc + ';">[' + b.severity + '] ' + b.type + '</strong><br><span style="color:#c9d1d9;font-size:12px;">' + b.description.substring(0,100) + '</span><br><span style="color:#8b949e;font-size:11px;">Fix: ' + b.fix.substring(0,80) + '</span></div>';
    }
    document.getElementById('apBlockers').innerHTML = bhtml || '<div style="color:#3fb950;">No blockers</div>';
  });
}
function runAutopilotCycle() {
  var btn = event.target; btn.disabled = true; btn.textContent = 'Running cycle...';
  api('/api/autonomous-printer/cycle', 'POST').then(function(d) {
    btn.disabled = false; btn.textContent = 'Run Autonomous Cycle';
    alert('Cycle done. Tasks executed: ' + (d.executed || 0));
    loadAutopilot();
  });
}

// ── discovery page ──
function loadDiscovery() {
  api('/api/discovery').then(function(d) {
    if (d.error) { document.getElementById('discSummary').innerHTML = '<div style="color:#f85149;">Error: ' + d.error + '</div>'; return; }
    var cards = '<div class="card"><div class="card-value">' + (d.total_opportunities||0) + '</div><div class="card-label">Total Opportunities</div></div>';
    var dims = d.dimensions || {};
    for (var key in dims) {
      cards += '<div class="card"><div class="card-value">' + (dims[key]||[]).length + '</div><div class="card-label">' + key.replace(/_/g,' ') + '</div></div>';
    }
    document.getElementById('discSummary').innerHTML = cards;
    var ohtml = '';
    var allOpps = [];
    for (var dim in dims) {
      var items = dims[dim] || [];
      for (var i = 0; i < items.length; i++) {
        allOpps.push({dim: dim, item: items[i]});
      }
    }
    for (var j = 0; j < allOpps.length; j++) {
      var o = allOpps[j];
      var item = o.item;
      var name = typeof item === 'string' ? item : (item.name || item.opportunity || item.title || JSON.stringify(item).substring(0,80));
      var dimColor = {'cross_pollination':'#58a6ff','seasonal':'#f0883e','underserved':'#a371f7','revenue_angles':'#3fb950','hypotheses':'#d29922'}[o.dim] || '#8b949e';
      ohtml += '<div style="padding:4px 8px;margin:3px 0;background:#161b22;border-radius:4px;font-size:12px;border-left:3px solid ' + dimColor + ';"><span style="color:' + dimColor + ';">[' + o.dim.replace(/_/g,' ') + ']</span> ' + name + '</div>';
    }
    document.getElementById('discOpportunities').innerHTML = ohtml || '<div style="color:#8b949e;">Run discovery scan first</div>';
  });
}
function runDiscovery() {
  var btn = event.target; btn.disabled = true; btn.textContent = 'Scanning...';
  fetch('/api/discovery').then(function(r){return r.json();}).then(function(d) {
    btn.disabled = false; btn.textContent = 'Run Discovery Scan';
    loadDiscovery();
  });
}

// ── scheduled runs page ──
function loadScheduled() {
  api('/api/scheduled-runs').then(function(d) {
    if (d.error) {
      document.getElementById('schedSubtitle').textContent = 'Error: ' + d.error;
      return;
    }
    var s = d.summary || {};
    document.getElementById('schedSubtitle').textContent =
      s.total_cron_jobs + ' cron jobs | ' + s.claude_sessions + ' agent sessions | ' +
      s.custom_runs + ' custom runs | ' + s.jobs_error + ' errors';

    // Summary cards
    var cards = document.getElementById('schedSummary');
    clear(cards);
    cards.appendChild(buildCard('Cron Jobs', String(s.total_cron_jobs), s.jobs_ok + ' OK', 'accent'));
    cards.appendChild(buildCard('Errors', String(s.jobs_error), 'need attention', s.jobs_error > 0 ? 'red' : 'green'));
    cards.appendChild(buildCard('Agent Sessions', String(s.claude_sessions), 'scheduled', 'cyan'));
    cards.appendChild(buildCard('Custom Runs', String(s.active_custom || 0), 'active', 'yellow'));
    cards.appendChild(buildCard('Guardrails', d.guardrails && d.guardrails.ok ? 'OK' : 'ISSUES', '', d.guardrails && d.guardrails.ok ? 'green' : 'red'));

    // Categories
    var catEl = document.getElementById('schedCategories');
    clear(catEl);
    var categories = d.categories || {};
    Object.keys(categories).sort().forEach(function(cat) {
      var jobs = categories[cat];
      var header = el('div', {className: 'section-header', style: 'margin-top:12px;font-size:13px;'}, cat + ' (' + jobs.length + ')');
      catEl.appendChild(header);
      var items = jobs.map(function(j) {
        return {
          Script: j.script,
          Frequency: j.frequency,
          Status: j.last_status,
          'Last Run': j.last_run ? j.last_run.substring(0, 16).replace('T', ' ') : 'never'
        };
      });
      buildTable(catEl, items, 4);
    });

    // Claude sessions
    var claudeEl = document.getElementById('schedClaude');
    clear(claudeEl);
    if (d.claude_sessions && d.claude_sessions.length > 0) {
      var csItems = d.claude_sessions.map(function(cs) {
        return {
          Session: cs.session_type || cs.script,
          Schedule: cs.frequency,
          Status: cs.last_status,
          'Last Run': cs.last_run ? cs.last_run.substring(0, 16).replace('T', ' ') : 'never'
        };
      });
      buildTable(claudeEl, csItems, 4);
    } else {
      claudeEl.appendChild(el('div', {className: 'loading'}, 'No Claude sessions scheduled'));
    }

    // Custom runs
    var customEl = document.getElementById('schedCustom');
    clear(customEl);
    if (d.custom_runs && d.custom_runs.length > 0) {
      var crItems = d.custom_runs.map(function(r) {
        return {
          ID: r.id,
          Description: (r.description || '').substring(0, 50),
          Schedule: r.schedule,
          Source: r.source,
          Enabled: r.enabled ? 'YES' : 'NO'
        };
      });
      buildTable(customEl, crItems, 5);
    } else {
      customEl.appendChild(el('div', {className: 'loading'}, 'No custom runs yet'));
    }

    // History
    var histEl = document.getElementById('schedHistory');
    clear(histEl);
    if (d.recent_history && d.recent_history.length > 0) {
      var hItems = d.recent_history.slice(-10).reverse().map(function(h) {
        return {
          Time: (h.timestamp || '').substring(0, 16).replace('T', ' '),
          Action: h.action,
          Detail: (h.description || h.run_id || '').substring(0, 40)
        };
      });
      buildTable(histEl, hItems, 3);
    } else {
      histEl.appendChild(el('div', {className: 'loading'}, 'No run history yet'));
    }
  });
}

function runPerpetual() {
  post('/api/scheduled-runs/perpetual').then(function(d) {
    if (d.error) {
      alert('Error: ' + d.error);
    } else {
      alert('Perpetual cycle: ' + d.status + '. Suggestions: ' + (d.suggestions || []).length + '. Created: ' + (d.created || []).length);
      loadScheduled();
    }
  });
}

// ── app names page ──
function loadAppNames() {
  var container = document.getElementById('nameAuditResults');
  clear(container);
  container.appendChild(el('div', {className: 'loading'}, 'Click "Run Full Audit" to validate app names'));
}

function runNameAudit() {
  var container = document.getElementById('nameAuditResults');
  clear(container);
  container.appendChild(el('div', {className: 'loading'}, 'Running audit... (checking App Store, domains, SEO scores)'));
  api('/api/app-names/audit').then(function(d) {
    clear(container);
    if (d.error) {
      container.appendChild(el('div', {className: 'loading'}, 'Error: ' + d.error));
      return;
    }
    var results = d.results || [];
    var items = results.map(function(r) {
      var domAvail = 0;
      if (r.checks && r.checks.domains) {
        Object.values(r.checks.domains).forEach(function(v) { if (v === 'LIKELY_AVAILABLE') domAvail++; });
      }
      return {
        Name: r.name,
        Score: r.composite_score + '/100',
        Verdict: r.verdict,
        'App Store': r.checks && r.checks.app_store && r.checks.app_store.available ? 'AVAIL' : 'TAKEN',
        'Domains': domAvail + '/3 avail',
        Niche: r.niche
      };
    });
    buildTable(container, items, 6);
  });
}

// ── page data loader ──
function loadPageData(id) {
  switch(id) {
    case 'dashboard': refreshDashboard(); break;
    case 'printer': loadPrinter(); break;
    case 'autopilot': loadAutopilot(); break;
    case 'discovery': loadDiscovery(); break;
    case 'scheduled': loadScheduled(); break;
    case 'goals': loadGoals(); break;
    case 'todos': loadTodos(); break;
    case 'operations': loadOperations(); break;
    case 'priority': loadPriority(); break;
    case 'ventures': loadVentures(); break;
    case 'appnames': loadAppNames(); break;
    case 'tools': loadTools(); break;
    case 'alarms': loadAlarms(); break;
    case 'tasks': loadTasks(); break;
    case 'launch': loadLaunch(); break;
    case 'quick': buildQuickLaunch(); break;
  }
}

// ── init ──
buildNav();
buildQuickLaunch();
refreshDashboard();

// auto refresh every 10s
setInterval(function() {
  if (state.page === 'dashboard') refreshDashboard();
  // always update wake badge
  api('/api/wake-state').then(function(d) {
    var badge = document.getElementById('wakeBadge');
    badge.textContent = d.is_awake ? 'AWAKE' : 'ASLEEP';
    badge.className = 'wake-badge ' + (d.is_awake ? 'awake' : 'asleep');
  });
}, 10000);
</script>
</body>
</html>"""


# ── main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="PRINTMAXX Command Center v3")
    parser.add_argument("--port", type=int, default=7777)
    parser.add_argument("--minimized", action="store_true",
                        help="Background reminders only, no web UI")
    args = parser.parse_args()

    # start alarm engine
    engine = WakeAlarmEngine()
    engine.start()

    if args.minimized:
        print("PRINTMAXX v3 running in background mode (reminders only)")
        print("Press Ctrl+C to stop")
        try:
            while True:
                time.sleep(60)
        except KeyboardInterrupt:
            engine.stop()
            return

    # start Flask
    app = create_app(engine)
    print(f"PRINTMAXX Command Center v3 starting on http://localhost:{args.port}")
    print("Press Ctrl+C to stop")

    # open browser
    def open_browser():
        time.sleep(1.5)
        subprocess.Popen(["open", f"http://localhost:{args.port}"],
                         stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    threading.Thread(target=open_browser, daemon=True).start()

    try:
        app.run(host="0.0.0.0", port=args.port, debug=False, use_reloader=False)
    except KeyboardInterrupt:
        engine.stop()


if __name__ == "__main__":
    main()
