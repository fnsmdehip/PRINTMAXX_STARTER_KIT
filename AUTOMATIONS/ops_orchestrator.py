#!/usr/bin/env python3

from __future__ import annotations
"""
PRINTMAXX Master Ops Orchestrator — The Brain of the Entire Operation.

Hedge-fund-grade operations management for 182+ solopreneur money methods.
Implements: State Machine Lifecycle, Adaptive Scheduling, Kelly Criterion
Allocation, Multi-Source Signal Fusion, Anti-Fragile Learning, Self-Healing.

Usage:
    python3 AUTOMATIONS/ops_orchestrator.py --run           # Execute priority queue
    python3 AUTOMATIONS/ops_orchestrator.py --status        # Show all ops states
    python3 AUTOMATIONS/ops_orchestrator.py --adapt         # Recalculate adaptive frequencies
    python3 AUTOMATIONS/ops_orchestrator.py --signals       # Show fused signal dashboard
    python3 AUTOMATIONS/ops_orchestrator.py --kelly         # Show Kelly allocation
    python3 AUTOMATIONS/ops_orchestrator.py --heal          # Diagnose and fix broken ops
    python3 AUTOMATIONS/ops_orchestrator.py --learn "text"  # Log a learning
    python3 AUTOMATIONS/ops_orchestrator.py --summary       # Quick system summary (10 lines)
    python3 AUTOMATIONS/ops_orchestrator.py --dashboard     # Full TUI dashboard
"""

import os
import sys
import csv
import json
import math
import time
import glob as globmod
import subprocess
from datetime import datetime, timedelta
from pathlib import Path
from collections import defaultdict

# ---------------------------------------------------------------------------
# PATHS
# ---------------------------------------------------------------------------
BASE = Path(__file__).resolve().parent.parent
LEDGER = BASE / "LEDGER"
FINANCIALS = BASE / "FINANCIALS"
AUTOMATIONS = BASE / "AUTOMATIONS"
OPS_DIR = BASE / "OPS"
PRODUCTS = BASE / "PRODUCTS"

STATE_FILE = LEDGER / "OPS_ORCHESTRATOR_STATE.json"
LOG_FILE = LEDGER / "OPS_ORCHESTRATOR_LOG.jsonl"
LEARNINGS_FILE = LEDGER / "RBI_STRATEGIC" / "LEARNINGS.jsonl"

# ---------------------------------------------------------------------------
# OP LIFECYCLE STATES
# ---------------------------------------------------------------------------
STATES = [
    "DORMANT",      # Not yet started, no resources allocated
    "SCOUTING",     # Research phase, gathering data
    "ACTIVE",       # Live, receiving resources
    "OPTIMIZING",   # Running A/B tests, improving conversion
    "SCALING",      # Proven winner, receiving max resources
    "MATURE",       # Steady-state, minimal maintenance
    "EXIT",         # Winding down, reallocating resources
]

# State transition rules: (from_state, to_state, condition_name)
TRANSITIONS = {
    ("DORMANT", "SCOUTING"):    "priority_score > 60",
    ("SCOUTING", "ACTIVE"):     "has_infrastructure AND has_accounts",
    ("ACTIVE", "OPTIMIZING"):   "first_revenue > 0",
    ("OPTIMIZING", "SCALING"):  "roi > 200% for 30+ days",
    ("SCALING", "MATURE"):      "growth_rate < 5%/month",
    ("*", "EXIT"):              "negative_roi for 60+ days",
    ("EXIT", "DORMANT"):        "new_alpha_signal detected",
}

# ---------------------------------------------------------------------------
# DYNAMIC OPS CATALOG — Reads ALL ops from PRINTMAXX_MASTER_OPS.xlsx
# Source of truth for which ops exist. Runtime state (revenue, wins, etc.)
# is stored in OPS_ORCHESTRATOR_STATE.json and merged on load.
# ---------------------------------------------------------------------------

# Category defaults for base_freq_hours, scalability, market_signal
# Used when XLSX doesn't provide these (it provides REVENUE_RANGE and AUTOMATION_LEVEL)
CATEGORY_DEFAULTS = {
    "CONTENT":   {"base_freq_hours": 6,   "scalability": 8, "market_signal": 7},
    "ECOM":      {"base_freq_hours": 24,  "scalability": 7, "market_signal": 7},
    "DIGITAL":   {"base_freq_hours": 48,  "scalability": 9, "market_signal": 7},
    "SERVICE":   {"base_freq_hours": 12,  "scalability": 7, "market_signal": 8},
    "APP":       {"base_freq_hours": 24,  "scalability": 10, "market_signal": 7},
    "PERSONA":   {"base_freq_hours": 12,  "scalability": 8, "market_signal": 8},
    "INVEST":    {"base_freq_hours": 4,   "scalability": 6, "market_signal": 7},
    "COMMUNITY": {"base_freq_hours": 48,  "scalability": 8, "market_signal": 7},
    "AFFILIATE": {"base_freq_hours": 48,  "scalability": 9, "market_signal": 8},
    "GROWTH":    {"base_freq_hours": 24,  "scalability": 5, "market_signal": 6},
}

# Map automation level text to numeric scalability bonus
AUTOMATION_SCALABILITY = {
    "High": 2,
    "Medium": 0,
    "Low": -1,
    "Manual": -2,
}

# Map revenue range text to market signal bonus
def _parse_revenue_signal(rev_range):
    """Estimate market signal from revenue range string like '$1K-20K/mo'."""
    if not rev_range or rev_range == '$0' or rev_range == '$0 (support)':
        return 4
    rev = rev_range.upper().replace(',', '')
    if '50K' in rev or '68K' in rev or '100K' in rev:
        return 10
    if '20K' in rev or '10K' in rev:
        return 9
    if '5K' in rev:
        return 8
    if '2K' in rev or '3K' in rev:
        return 7
    if '1K' in rev or '500' in rev:
        return 6
    return 5


def _extract_platforms(platforms_str):
    """Extract platform names from the PLATFORMS column for account needs."""
    if not platforms_str or platforms_str == 'None' or platforms_str == 'N/A':
        return []
    platforms = [p.strip().lower().replace('/', '_').replace(' ', '_')
                 for p in platforms_str.split(',')]
    # Normalize common platform names
    normalized = []
    for p in platforms:
        if p in ('x_twitter', 'x', 'twitter'):
            normalized.append('twitter')
        elif p in ('ios', 'android', 'web', 'app_store', 'google_play'):
            normalized.append('apple_dev')
        elif p in ('email', 'cold_email'):
            normalized.append('email_provider')
        elif p in ('various', 'all', 'n_a', ''):
            continue
        else:
            normalized.append(p)
    return list(set(normalized))


def _extract_infra_files(source_file, exists_in_system):
    """Extract infrastructure file paths from SOURCE_FILE column."""
    if not source_file or source_file == 'None' or source_file == 'N/A':
        return []
    files = [f.strip() for f in source_file.split(',') if f.strip()]
    return [f for f in files if f and f != 'None' and f != 'N/A']


def load_ops_from_xlsx():
    """Load ALL ops from PRINTMAXX_MASTER_OPS.xlsx dynamically.

    Returns a list of op dicts with the same structure as the old DEFAULT_OPS.
    """
    try:
        import openpyxl
    except ImportError:
        print("WARNING: openpyxl not installed. Using empty ops list.")
        return []

    xlsx_path = BASE / "PRINTMAXX_MASTER_OPS.xlsx"
    if not xlsx_path.exists():
        print(f"WARNING: {xlsx_path} not found. Using empty ops list.")
        return []

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        print(f"WARNING: Failed to open XLSX: {e}. Using empty ops list.")
        return []

    if 'ALL OPS MASTER' not in wb.sheetnames:
        print("WARNING: 'ALL OPS MASTER' sheet not found in XLSX.")
        return []

    sheet = wb['ALL OPS MASTER']

    # Read headers to build column index (handles column reordering)
    headers = {}
    for i, cell in enumerate(sheet[1]):
        if cell.value:
            headers[str(cell.value).strip().upper()] = i

    ops = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        op_id = str(row[headers.get('OP_ID', 0)]) if row[headers.get('OP_ID', 0)] else None
        if not op_id or op_id in ('TOTAL OPS:', 'None', ''):
            continue

        cat = str(row[headers.get('CATEGORY', 1)]) if row[headers.get('CATEGORY', 1)] else 'UNKNOWN'
        cat = cat.upper()
        name = str(row[headers.get('OP_NAME', 2)]) if row[headers.get('OP_NAME', 2)] else 'Unnamed'
        desc = str(row[headers.get('DESCRIPTION', 3)]) if row[headers.get('DESCRIPTION', 3)] else ''
        rev_range = str(row[headers.get('REVENUE_RANGE', 4)]) if row[headers.get('REVENUE_RANGE', 4)] else '$0'
        automation = str(row[headers.get('AUTOMATION_LEVEL', 5)]) if row[headers.get('AUTOMATION_LEVEL', 5)] else 'Manual'
        source_file = str(row[headers.get('SOURCE_FILE', 10)]) if len(row) > 10 and row[headers.get('SOURCE_FILE', 10)] else ''
        exists_in_sys = str(row[headers.get('EXISTS_IN_SYSTEM', 9)]) if len(row) > 9 and row[headers.get('EXISTS_IN_SYSTEM', 9)] else 'NO'
        platforms_str = str(row[headers.get('PLATFORMS', 15)]) if len(row) > 15 and row[headers.get('PLATFORMS', 15)] else ''

        # Get category defaults
        cat_defaults = CATEGORY_DEFAULTS.get(cat, {"base_freq_hours": 24, "scalability": 5, "market_signal": 5})

        # Calculate scalability from category default + automation bonus
        scalability = cat_defaults["scalability"] + AUTOMATION_SCALABILITY.get(automation, 0)
        scalability = max(1, min(scalability, 10))

        # Calculate market signal from revenue range
        market_signal = _parse_revenue_signal(rev_range)

        # Extract infra files and account needs
        infra_files = _extract_infra_files(source_file, exists_in_sys)
        needs_accounts = _extract_platforms(platforms_str)

        ops.append({
            "op_id": op_id,
            "name": name,
            "category": cat.lower(),
            "description": desc,
            "revenue_range": rev_range,
            "automation_level": automation,
            "infra_files": infra_files,
            "needs_accounts": needs_accounts,
            "base_freq_hours": cat_defaults["base_freq_hours"],
            "scalability": scalability,
            "market_signal": market_signal,
        })

    return ops


# Load ops from XLSX at module level
DEFAULT_OPS = load_ops_from_xlsx()


# ---------------------------------------------------------------------------
# UTILITY FUNCTIONS
# ---------------------------------------------------------------------------

def read_csv(path, max_rows=2000):
    """Read CSV, return list of dicts. Silent on failure."""
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            return [row for i, row in enumerate(csv.DictReader(f)) if i < max_rows]
    except Exception:
        return []


def read_jsonl(path, max_lines=1000):
    """Read JSONL, return list of dicts."""
    results = []
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            for i, line in enumerate(f):
                if i >= max_lines:
                    break
                line = line.strip()
                if line:
                    try:
                        results.append(json.loads(line))
                    except json.JSONDecodeError:
                        pass
    except FileNotFoundError:
        pass
    return results


def append_jsonl(path, entry):
    """Append a single JSON object to a JSONL file."""
    os.makedirs(path.parent, exist_ok=True)
    with open(path, "a", encoding="utf-8") as f:
        f.write(json.dumps(entry) + "\n")


def file_exists(rel_path):
    """Check if a file or directory exists relative to BASE."""
    p = BASE / rel_path
    return p.exists()


def file_age_hours(path):
    """Return hours since file was last modified."""
    try:
        return (time.time() - os.path.getmtime(path)) / 3600
    except Exception:
        return 99999


def now_iso():
    return datetime.now().isoformat()


def log_event(event_type, op_id="SYSTEM", details=None):
    """Append to activity log."""
    entry = {
        "timestamp": now_iso(),
        "event": event_type,
        "op_id": op_id,
        "details": details or {},
    }
    append_jsonl(LOG_FILE, entry)


# ---------------------------------------------------------------------------
# STATE MANAGEMENT
# ---------------------------------------------------------------------------

class OrchestratorState:
    """Persistent state for all ops, frequencies, Kelly fractions, history."""

    def __init__(self):
        self.ops = {}           # op_id -> op dict
        self.meta = {
            "created_at": now_iso(),
            "last_run": None,
            "total_runs": 0,
            "version": "1.0.0",
        }

    def save(self):
        os.makedirs(STATE_FILE.parent, exist_ok=True)
        data = {
            "meta": self.meta,
            "ops": self.ops,
        }
        with open(STATE_FILE, "w") as f:
            json.dump(data, f, indent=2, default=str)

    @classmethod
    def _make_op_entry(cls, op):
        """Create a fresh op state entry from a DEFAULT_OPS dict."""
        oid = op["op_id"]
        return {
            "op_id": oid,
            "name": op["name"],
            "category": op["category"],
            "description": op.get("description", ""),
            "revenue_range": op.get("revenue_range", "$0"),
            "automation_level": op.get("automation_level", "Manual"),
            "state": "DORMANT",
            "infra_files": op.get("infra_files", []),
            "needs_accounts": op.get("needs_accounts", []),
            "base_freq_hours": op.get("base_freq_hours", 24),
            "current_freq_hours": op.get("base_freq_hours", 24),
            "scalability": op.get("scalability", 5),
            "market_signal": op.get("market_signal", 5),
            "priority_score": 0,
            "kelly_fraction": 0.0,
            "total_revenue": 0.0,
            "total_hours": 0.0,
            "total_runs": 0,
            "consecutive_failures": 0,
            "win_count": 0,
            "loss_count": 0,
            "last_run": None,
            "last_revenue_date": None,
            "first_revenue_date": None,
            "momentum": 0.0,
            "signal_confidence": 0.0,
            "created_at": now_iso(),
            "state_changed_at": now_iso(),
            "notes": "",
        }

    @classmethod
    def load(cls):
        """Load from disk, then merge with XLSX source of truth.

        Strategy:
        - XLSX is the source of truth for WHICH ops exist and their metadata
          (name, category, description, revenue_range, automation_level, etc.)
        - State file (JSON) preserves RUNTIME data (state, revenue, wins, losses,
          momentum, kelly_fraction, last_run, etc.)
        - On load: start from XLSX ops, then overlay runtime data from state file.
        - Ops in state file but NOT in XLSX are preserved (legacy ops).
        - Ops in XLSX but NOT in state file get fresh default entries.
        """
        state = cls()
        saved_ops = {}

        # Step 1: Load saved runtime state if it exists
        if STATE_FILE.exists():
            try:
                with open(STATE_FILE) as f:
                    data = json.load(f)
                state.meta = data.get("meta", state.meta)
                saved_ops = data.get("ops", {})
            except (json.JSONDecodeError, KeyError):
                pass

        # Step 2: Build ops from XLSX (source of truth for op definitions)
        xlsx_op_ids = set()
        for op in DEFAULT_OPS:
            oid = op["op_id"]
            xlsx_op_ids.add(oid)

            if oid in saved_ops:
                # Op exists in saved state - start with saved runtime data
                entry = saved_ops[oid]
                # Update metadata fields from XLSX (name, category, etc. may have changed)
                entry["name"] = op["name"]
                entry["category"] = op["category"]
                entry["description"] = op.get("description", entry.get("description", ""))
                entry["revenue_range"] = op.get("revenue_range", entry.get("revenue_range", "$0"))
                entry["automation_level"] = op.get("automation_level", entry.get("automation_level", "Manual"))
                entry["infra_files"] = op.get("infra_files", entry.get("infra_files", []))
                entry["needs_accounts"] = op.get("needs_accounts", entry.get("needs_accounts", []))
                entry["base_freq_hours"] = op.get("base_freq_hours", entry.get("base_freq_hours", 24))
                entry["scalability"] = op.get("scalability", entry.get("scalability", 5))
                entry["market_signal"] = op.get("market_signal", entry.get("market_signal", 5))
                state.ops[oid] = entry
            else:
                # New op from XLSX - create fresh entry
                state.ops[oid] = cls._make_op_entry(op)

        # Step 3: Preserve any legacy ops from state file not in XLSX
        for oid, op_data in saved_ops.items():
            if oid not in xlsx_op_ids:
                state.ops[oid] = op_data

        # Step 4: Log and save
        new_count = len(xlsx_op_ids - set(saved_ops.keys()))
        legacy_count = len(set(saved_ops.keys()) - xlsx_op_ids)
        if new_count > 0 or legacy_count > 0 or not saved_ops:
            log_event("STATE_MERGED", details={
                "xlsx_ops": len(xlsx_op_ids),
                "saved_ops": len(saved_ops),
                "new_from_xlsx": new_count,
                "legacy_preserved": legacy_count,
                "total_ops": len(state.ops),
            })
            state.save()

        return state


# ---------------------------------------------------------------------------
# DATA INGESTION — Read all external signal sources
# ---------------------------------------------------------------------------

def get_accounts():
    """Read LEDGER/ACCOUNTS.csv, return dict of platform -> is_active."""
    rows = read_csv(LEDGER / "ACCOUNTS.csv")
    active = {}
    for r in rows:
        platform = r.get("Platform", r.get("platform", "")).strip().lower()
        status = r.get("Status", r.get("status", "")).strip().upper()
        if platform:
            active[platform] = status in ("CREATED", "ACTIVE", "WARMED")
    return active


def get_revenue_by_method():
    """Read FINANCIALS/REVENUE_TRACKER.csv for revenue per op."""
    rows = read_csv(FINANCIALS / "REVENUE_TRACKER.csv")
    revenue = defaultdict(float)
    for r in rows:
        method = r.get("method_id", r.get("method", ""))
        try:
            amt = float(str(r.get("amount", r.get("revenue", "0"))).replace("$", "").replace(",", ""))
            revenue[method] += amt
        except (ValueError, TypeError):
            pass
    return dict(revenue)


def get_overnight_status():
    """Read most recent overnight status JSON. Return list of script results."""
    pattern = str(AUTOMATIONS / "logs" / "overnight_status_*.json")
    files = sorted(globmod.glob(pattern))
    if not files:
        return []
    results = []
    try:
        with open(files[-1], "r") as f:
            for line in f:
                line = line.strip()
                if line and line.startswith("{"):
                    try:
                        results.append(json.loads(line))
                    except json.JSONDecodeError:
                        pass
    except Exception:
        pass
    return results


def get_freelance_demand_signals():
    """Read LEDGER/FREELANCE_DEMAND_SCAN.csv for demand signals."""
    rows = read_csv(LEDGER / "FREELANCE_DEMAND_SCAN.csv")
    signals = []
    for r in rows:
        try:
            score = int(r.get("score", 0))
        except (ValueError, TypeError):
            score = 0
        matched = r.get("matched_services", "")
        signals.append({
            "score": score,
            "source": r.get("source", ""),
            "title": r.get("title", "")[:80],
            "matched_services": matched,
            "budget": r.get("budget", ""),
        })
    return signals


def get_ecom_arb_signals():
    """Read LEDGER/ECOM_ARB_OPPORTUNITIES.csv for ecom arb signals."""
    rows = read_csv(LEDGER / "ECOM_ARB_OPPORTUNITIES.csv")
    signals = []
    for r in rows:
        try:
            margin = float(r.get("margin_pct", 0))
            composite = float(r.get("composite_score", 0))
        except (ValueError, TypeError):
            margin = 0
            composite = 0
        action = r.get("action", "SKIP")
        if action in ("LIST", "WATCH"):
            signals.append({
                "product": r.get("product", ""),
                "margin_pct": margin,
                "composite_score": composite,
                "action": action,
                "category": r.get("category", ""),
            })
    return signals


def get_trend_signals():
    """Read LEDGER/TREND_SIGNALS.csv for trend signals."""
    rows = read_csv(LEDGER / "TREND_SIGNALS.csv")
    signals = []
    for r in rows:
        try:
            score = int(r.get("score", 0))
            strength = int(r.get("strength", 0))
        except (ValueError, TypeError):
            score = 0
            strength = 0
        if score >= 45:
            signals.append({
                "signal": r.get("signal", "")[:80],
                "score": score,
                "strength": strength,
                "signal_type": r.get("signal_type", ""),
                "product_matches": r.get("product_matches", ""),
            })
    return signals


def get_alpha_signals():
    """Read LEDGER/ALPHA_STAGING.csv for alpha intelligence signals."""
    rows = read_csv(LEDGER / "ALPHA_STAGING.csv")
    signals = []
    for r in rows:
        status = r.get("status", "").strip().upper()
        if status == "APPROVED":
            signals.append({
                "alpha_id": r.get("alpha_id", ""),
                "category": r.get("category", ""),
                "roi_potential": r.get("roi_potential", ""),
                "tactic": r.get("tactic", "")[:100],
            })
    return signals


# ---------------------------------------------------------------------------
# SIGNAL FUSION ENGINE
# ---------------------------------------------------------------------------

# Maps signal categories to op categories (lowercased to match XLSX categories)
SIGNAL_TO_OP_MAP = {
    "freelance": ["service"],
    "website": ["service"],
    "automation": ["service", "digital"],
    "cold_email": ["service", "growth"],
    "social_media": ["content", "persona"],
    "video_editing": ["content", "service"],
    "data_entry": ["service"],
    "copywriting": ["content", "service"],
    "logo": ["service"],
    "presentation": ["service"],
    "app_development": ["app"],
    "trading": ["invest"],
    "community_management": ["community"],
    "affiliate_marketing": ["affiliate"],
    "ecommerce": ["ecom"],
}

ALPHA_CAT_TO_OP_MAP = {
    "APP_FACTORY": ["app"],
    "CONTENT_FARM": ["content"],
    "CONTENT_FORMAT": ["content"],
    "OUTBOUND": ["service", "growth"],
    "AI_INFLUENCER": ["persona"],
    "SEO_GEO_ASO": ["growth", "content"],
    "TOOL_ALPHA": ["growth"],
    "MONETIZATION": ["digital", "app", "community", "ecom"],
    "GROWTH_HACK": ["content", "affiliate", "growth"],
    "ECOM": ["ecom"],
    "SERVICE": ["service"],
    "INVEST": ["invest"],
}


def fuse_signals(state):
    """
    Combine signals from ALL scanners. When multiple signals point to the same
    op category, confidence MULTIPLIES.

    Returns: dict of op_id -> fused_confidence (0-100)
    """
    # Gather per-category signal counts and strength
    category_signals = defaultdict(lambda: {"count": 0, "total_strength": 0, "sources": set()})

    # 1. Freelance demand signals
    for sig in get_freelance_demand_signals():
        if sig["score"] >= 50:
            for svc in sig["matched_services"].split(","):
                svc = svc.strip().lower()
                if svc in SIGNAL_TO_OP_MAP:
                    for cat in SIGNAL_TO_OP_MAP[svc]:
                        category_signals[cat]["count"] += 1
                        category_signals[cat]["total_strength"] += sig["score"]
                        category_signals[cat]["sources"].add("freelance_demand")

    # 2. Ecom arb signals
    for sig in get_ecom_arb_signals():
        if sig["composite_score"] > 30:
            category_signals["ecom"]["count"] += 1
            category_signals["ecom"]["total_strength"] += sig["composite_score"]
            category_signals["ecom"]["sources"].add("ecom_arb")

    # 3. Trend signals
    # Map trend product_matches to XLSX categories (lowercased)
    TREND_TO_CATEGORY = {
        "digital_product": "digital",
        "health_wellness": "app",
        "fitness": "app",
        "service_opportunity": "service",
        "dropship_product": "ecom",
        "content_opportunity": "content",
        "community": "community",
        "affiliate": "affiliate",
        "investment": "invest",
        "persona": "persona",
        "growth": "growth",
        "app": "app",
    }
    for sig in get_trend_signals():
        matches = sig.get("product_matches", "").lower()
        for keyword, cat in TREND_TO_CATEGORY.items():
            if keyword in matches:
                category_signals[cat]["count"] += 1
                category_signals[cat]["total_strength"] += sig["score"]
                category_signals[cat]["sources"].add("trends")

    # 4. Alpha intelligence signals
    for sig in get_alpha_signals():
        alpha_cat = sig.get("category", "").upper()
        if alpha_cat in ALPHA_CAT_TO_OP_MAP:
            for cat in ALPHA_CAT_TO_OP_MAP[alpha_cat]:
                category_signals[cat]["count"] += 1
                roi = sig.get("roi_potential", "MEDIUM")
                strength = {"HIGHEST": 90, "HIGH": 70, "MEDIUM": 50, "LOW": 30}.get(roi, 40)
                category_signals[cat]["total_strength"] += strength
                category_signals[cat]["sources"].add("alpha")

    # 5. Overnight system health signals
    overnight = get_overnight_status()
    failure_cats = set()
    success_cats = set()
    for entry in overnight:
        script = entry.get("script", "")
        status = entry.get("status", "")
        # Map scripts to XLSX categories (lowercased)
        if "leads" in script:
            if status == "SUCCESS":
                success_cats.add("service")
                success_cats.add("growth")
            else:
                failure_cats.add("service")
        elif "ecom" in script or "arb" in script:
            if status == "SUCCESS":
                success_cats.add("ecom")
            else:
                failure_cats.add("ecom")
        elif "alpha" in script:
            if status == "SUCCESS":
                success_cats.add("growth")
        elif "platform" in script or "niche" in script:
            if status == "SUCCESS":
                success_cats.add("content")
        elif "app" in script:
            if status == "SUCCESS":
                success_cats.add("app")
        elif "affiliate" in script:
            if status == "SUCCESS":
                success_cats.add("affiliate")

    # Calculate fused confidence per category
    category_confidence = {}
    for cat, data in category_signals.items():
        n_sources = len(data["sources"])
        avg_strength = data["total_strength"] / max(data["count"], 1)
        # Multi-source multiplier: 1 source = 1x, 2 = 1.5x, 3 = 2x, 4+ = 2.5x
        source_multiplier = 1.0 + 0.5 * min(n_sources - 1, 3) if n_sources > 0 else 0
        raw_confidence = avg_strength * source_multiplier
        # Penalize categories with overnight failures
        if cat in failure_cats and cat not in success_cats:
            raw_confidence *= 0.7
        category_confidence[cat] = min(raw_confidence, 100)

    # Map category confidence to individual ops
    op_confidence = {}
    for oid, op in state.ops.items():
        cat = op.get("category", "")
        base = category_confidence.get(cat, 0)
        # Boost ops that have infrastructure ready
        infra_boost = 0
        for f in op.get("infra_files", []):
            if file_exists(f):
                infra_boost += 5
        op_confidence[oid] = min(base + infra_boost, 100)

    return op_confidence, category_confidence


# ---------------------------------------------------------------------------
# PRIORITY SCORING ENGINE
# ---------------------------------------------------------------------------

def compute_priority_scores(state):
    """
    Score each op 0-100 based on:
    - Infrastructure readiness (0-20)
    - Account readiness (0-15)
    - Signal confidence (0-25)
    - Market signal (0-15)
    - Scalability (0-10)
    - Momentum (0-15)
    """
    accounts = get_accounts()
    revenue_map = get_revenue_by_method()
    op_confidence, _ = fuse_signals(state)

    for oid, op in state.ops.items():
        score = 0

        # Infrastructure readiness (0-20)
        infra_files = op.get("infra_files", [])
        if infra_files:
            ready = sum(1 for f in infra_files if file_exists(f))
            score += int((ready / len(infra_files)) * 20)
        else:
            score += 10  # No infra needed = partial credit

        # Account readiness (0-15)
        needed = op.get("needs_accounts", [])
        if needed:
            ready = sum(1 for a in needed if accounts.get(a.lower(), False))
            score += int((ready / len(needed)) * 15)
        else:
            score += 15  # No accounts needed = full credit

        # Signal confidence from fusion engine (0-25)
        confidence = op_confidence.get(oid, 0)
        score += int(confidence * 0.25)

        # Market signal (0-15, scaled from 1-10 input)
        ms = op.get("market_signal", 5)
        score += int(ms * 1.5)

        # Scalability (0-10, scaled from 1-10 input)
        sc = op.get("scalability", 5)
        score += sc

        # Momentum bonus (0-15)
        momentum = op.get("momentum", 0)
        score += int(min(abs(momentum), 1.0) * 15) if momentum > 0 else 0

        op["priority_score"] = min(score, 100)
        op["signal_confidence"] = confidence


# ---------------------------------------------------------------------------
# KELLY CRITERION ALLOCATION
# ---------------------------------------------------------------------------

def compute_kelly_fractions(state):
    """
    Allocate TIME like a hedge fund allocates capital using Kelly Criterion.
    f* = (p*b - q) / b
    where p = win_rate, b = avg_win/avg_loss, q = 1-p
    """
    total_kelly = 0.0
    kelly_map = {}

    for oid, op in state.ops.items():
        wins = op.get("win_count", 0)
        losses = op.get("loss_count", 0)
        total = wins + losses

        if total < 2:
            # Insufficient data — use priority score as proxy
            # Higher priority = higher fraction, but capped at 1/n
            n_ops = max(len(state.ops), 1)
            kelly = op.get("priority_score", 50) / (100 * n_ops) * 2
        else:
            p = wins / total  # win rate
            q = 1 - p

            # Estimate payoff ratio from revenue and hours
            rev = op.get("total_revenue", 0)
            hrs = op.get("total_hours", 0)

            if hrs > 0 and rev > 0:
                b = rev / max(hrs * 15, 1)  # $15/hr is our opportunity cost
            else:
                b = 1.0  # neutral assumption

            # Kelly formula
            if b > 0:
                kelly = (p * b - q) / b
            else:
                kelly = 0

        # Clamp: no negative allocation, max 25% to any single op
        kelly = max(kelly, 0)
        kelly = min(kelly, 0.25)

        # Zero allocation for EXIT state ops
        if op.get("state") == "EXIT":
            kelly = 0

        kelly_map[oid] = kelly
        total_kelly += kelly

    # Normalize so fractions sum to 1.0
    if total_kelly > 0:
        for oid in kelly_map:
            kelly_map[oid] /= total_kelly
    else:
        # Equal allocation as fallback
        n = max(len(kelly_map), 1)
        for oid in kelly_map:
            kelly_map[oid] = 1.0 / n

    # Store in state
    for oid, fraction in kelly_map.items():
        state.ops[oid]["kelly_fraction"] = round(fraction, 6)


# ---------------------------------------------------------------------------
# ADAPTIVE SCHEDULING
# ---------------------------------------------------------------------------

def adapt_frequencies(state):
    """
    Adjust run frequency based on performance.
    High performers: increase frequency (shorter intervals).
    Low performers: decrease frequency (longer intervals).
    Formula: new_freq = base_freq / (1 + momentum * 2)
    """
    for oid, op in state.ops.items():
        base = op.get("base_freq_hours", 24)
        momentum = op.get("momentum", 0)
        consecutive_failures = op.get("consecutive_failures", 0)

        # Momentum factor: positive = run more often
        if momentum > 0:
            freq = base / (1 + momentum * 2)
        elif momentum < 0:
            # Slow down, but with exponential backoff for consecutive failures
            backoff = min(2 ** consecutive_failures, 16) if consecutive_failures > 0 else 1
            freq = base * (1 + abs(momentum)) * backoff
        else:
            freq = base

        # Clamp between 1 hour and 720 hours (30 days)
        freq = max(1, min(freq, 720))

        op["current_freq_hours"] = round(freq, 1)

    log_event("FREQUENCIES_ADAPTED")


# ---------------------------------------------------------------------------
# STATE MACHINE TRANSITIONS
# ---------------------------------------------------------------------------

def evaluate_transitions(state):
    """Evaluate state transitions for all ops based on real data."""
    accounts = get_accounts()
    revenue_map = get_revenue_by_method()
    transitions_made = []

    for oid, op in state.ops.items():
        current_state = op.get("state", "DORMANT")
        new_state = current_state
        reason = ""

        # DORMANT -> SCOUTING: priority score > 60
        if current_state == "DORMANT" and op.get("priority_score", 0) > 60:
            new_state = "SCOUTING"
            reason = f"priority_score={op['priority_score']} > 60"

        # SCOUTING -> ACTIVE: has infrastructure AND has accounts
        elif current_state == "SCOUTING":
            infra_files = op.get("infra_files", [])
            has_infra = all(file_exists(f) for f in infra_files) if infra_files else True
            needed = op.get("needs_accounts", [])
            has_accts = all(accounts.get(a.lower(), False) for a in needed) if needed else True
            if has_infra and has_accts:
                new_state = "ACTIVE"
                reason = "infrastructure ready AND accounts active"

        # ACTIVE -> OPTIMIZING: first revenue > 0
        elif current_state == "ACTIVE":
            rev = revenue_map.get(oid, 0) + op.get("total_revenue", 0)
            if rev > 0:
                new_state = "OPTIMIZING"
                reason = f"first_revenue=${rev:.2f}"
                if not op.get("first_revenue_date"):
                    op["first_revenue_date"] = now_iso()

        # OPTIMIZING -> SCALING: high ROI sustained
        elif current_state == "OPTIMIZING":
            rev = op.get("total_revenue", 0)
            hrs = op.get("total_hours", 0)
            if hrs > 0 and rev > 0:
                roi_pct = (rev / (hrs * 15) - 1) * 100  # $15/hr opportunity cost
                # Check if we have 30+ days of data
                first_rev = op.get("first_revenue_date")
                if first_rev:
                    try:
                        days = (datetime.now() - datetime.fromisoformat(first_rev)).days
                    except (ValueError, TypeError):
                        days = 0
                    if roi_pct > 200 and days >= 30:
                        new_state = "SCALING"
                        reason = f"ROI={roi_pct:.0f}% for {days} days"

        # SCALING -> MATURE: growth rate < 5%/month
        elif current_state == "SCALING":
            # Simplified: if momentum near zero, transition to MATURE
            if abs(op.get("momentum", 0)) < 0.05:
                new_state = "MATURE"
                reason = f"growth_rate stabilized (momentum={op.get('momentum', 0):.3f})"

        # ANY -> EXIT: negative ROI for extended period
        if current_state in ("ACTIVE", "OPTIMIZING", "SCALING", "MATURE"):
            rev = op.get("total_revenue", 0)
            hrs = op.get("total_hours", 0)
            if hrs > 20 and rev == 0:
                new_state = "EXIT"
                reason = f"0 revenue after {hrs:.0f} hours invested"

        # EXIT -> DORMANT: new alpha signal
        if current_state == "EXIT" and op.get("signal_confidence", 0) > 70:
            new_state = "DORMANT"
            reason = f"new_alpha_signal (confidence={op['signal_confidence']:.0f})"

        if new_state != current_state:
            op["state"] = new_state
            op["state_changed_at"] = now_iso()
            transitions_made.append((oid, current_state, new_state, reason))
            log_event("STATE_TRANSITION", oid, {
                "from": current_state,
                "to": new_state,
                "reason": reason,
            })

    return transitions_made


# ---------------------------------------------------------------------------
# SELF-HEALING
# ---------------------------------------------------------------------------

def diagnose_and_heal(state):
    """
    Detect failures from overnight status, diagnose, and attempt fixes.
    Returns list of diagnosis/action tuples.
    """
    overnight = get_overnight_status()
    diagnoses = []

    # Count successes and failures by script category
    failures = defaultdict(list)
    successes = set()

    for entry in overnight:
        script = entry.get("script", "")
        status = entry.get("status", "")
        details = entry.get("details", "")

        if status == "FAILED":
            failures["failed"].append({"script": script, "details": details})
        elif status == "TIMEOUT":
            failures["timeout"].append({"script": script, "details": details})
        elif status == "SUCCESS":
            successes.add(script)

    healing_actions = []

    # Diagnose timeout patterns
    timeout_scripts = [f["script"] for f in failures.get("timeout", [])]
    if timeout_scripts:
        # Check if it's all lead scrapers
        lead_timeouts = [s for s in timeout_scripts if s.startswith("leads_")]
        if len(lead_timeouts) > 5:
            diagnoses.append({
                "issue": f"Lead scraper timeouts ({len(lead_timeouts)} scripts)",
                "diagnosis": "DuckDuckGo/Google rate limiting or network timeout",
                "severity": "MEDIUM",
                "action": "Increase timeout to 300s or switch to cached leads",
                "auto_fixable": False,
            })
            # Update affected ops
            for oid, op in state.ops.items():
                if op.get("category") == "outbound":
                    op["consecutive_failures"] = op.get("consecutive_failures", 0) + 1

        non_lead_timeouts = [s for s in timeout_scripts if not s.startswith("leads_")]
        for script in non_lead_timeouts:
            diagnoses.append({
                "issue": f"Timeout: {script}",
                "diagnosis": "External API or browser automation timeout",
                "severity": "LOW",
                "action": f"Retry with longer timeout or fallback scraper",
                "auto_fixable": False,
            })

    # Diagnose failed scripts
    for failure in failures.get("failed", []):
        script = failure["script"]
        details = failure["details"]
        diagnoses.append({
            "issue": f"Failed: {script} ({details})",
            "diagnosis": "Script error or missing dependency",
            "severity": "HIGH" if "exit code 2" in details else "MEDIUM",
            "action": f"Check AUTOMATIONS/{script}.py for errors, run manually to debug",
            "auto_fixable": False,
        })

    # Check for ops stuck in wrong state
    accounts = get_accounts()
    for oid, op in state.ops.items():
        s = op.get("state", "DORMANT")
        if s == "SCOUTING":
            needed = op.get("needs_accounts", [])
            missing = [a for a in needed if not accounts.get(a.lower(), False)]
            if missing:
                diagnoses.append({
                    "issue": f"{oid} stuck in SCOUTING (missing accounts: {', '.join(missing)})",
                    "diagnosis": "Account creation is the bottleneck",
                    "severity": "HIGH",
                    "action": f"Create accounts: {', '.join(missing)}. See OPS/ACCOUNT_CREATION_NOW.md",
                    "auto_fixable": False,
                })

    # Overall system health score
    total = len(overnight)
    success_count = len(successes)
    health = (success_count / max(total, 1)) * 100 if total > 0 else 0
    diagnoses.insert(0, {
        "issue": f"System health: {health:.0f}% ({success_count}/{total} scripts OK)",
        "diagnosis": "Overnight run summary",
        "severity": "INFO",
        "action": "No action needed" if health > 70 else "Review failed scripts",
        "auto_fixable": False,
    })

    return diagnoses


# ---------------------------------------------------------------------------
# MOMENTUM CALCULATOR
# ---------------------------------------------------------------------------

def update_momentum(state):
    """
    Calculate momentum for each op based on recent performance trajectory.
    Momentum = weighted average of recent revenue changes.
    Positive = accelerating, Negative = decelerating.
    """
    revenue_map = get_revenue_by_method()

    for oid, op in state.ops.items():
        old_rev = op.get("total_revenue", 0)
        new_rev = revenue_map.get(oid, 0) + old_rev  # Cumulative

        if old_rev > 0:
            pct_change = (new_rev - old_rev) / old_rev
            # Exponential moving average: new_momentum = 0.7 * old + 0.3 * change
            old_momentum = op.get("momentum", 0)
            op["momentum"] = round(0.7 * old_momentum + 0.3 * pct_change, 4)
        elif new_rev > 0:
            op["momentum"] = 0.5  # First revenue = positive signal
        else:
            # Decay momentum toward zero if no revenue
            old_momentum = op.get("momentum", 0)
            op["momentum"] = round(old_momentum * 0.9, 4)


# ---------------------------------------------------------------------------
# EXECUTION ENGINE
# ---------------------------------------------------------------------------

def get_execution_queue(state):
    """
    Build priority-ordered queue of ops to run right now.
    Filters by: state, frequency, and Kelly allocation.
    """
    now = datetime.now()
    queue = []

    for oid, op in state.ops.items():
        s = op.get("state", "DORMANT")
        if s in ("EXIT",):
            continue

        freq = op.get("current_freq_hours", 24)
        last_run = op.get("last_run")

        # Check if enough time has passed
        if last_run:
            try:
                lr = datetime.fromisoformat(last_run)
                hours_since = (now - lr).total_seconds() / 3600
                if hours_since < freq:
                    continue
            except (ValueError, TypeError):
                pass

        queue.append({
            "op_id": oid,
            "name": op.get("name", ""),
            "state": s,
            "priority": op.get("priority_score", 0),
            "kelly": op.get("kelly_fraction", 0),
            "freq": freq,
            "category": op.get("category", ""),
        })

    # Sort by priority descending, then Kelly descending
    queue.sort(key=lambda x: (-x["priority"], -x["kelly"]))
    return queue


# ---------------------------------------------------------------------------
# CLI DISPLAY FUNCTIONS
# ---------------------------------------------------------------------------

def display_status(state):
    """Show all ops with states, priorities, and scores."""
    compute_priority_scores(state)
    compute_kelly_fractions(state)

    # Group by state
    by_state = defaultdict(list)
    for oid, op in state.ops.items():
        by_state[op.get("state", "DORMANT")].append(op)

    print(f"\n{'='*76}")
    print(f"  PRINTMAXX OPS ORCHESTRATOR — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"  {len(state.ops)} ops tracked | Version {state.meta.get('version', '?')}")
    print(f"{'='*76}\n")

    state_icons = {
        "DORMANT": "[ ]", "SCOUTING": "[?]", "ACTIVE": "[>]",
        "OPTIMIZING": "[~]", "SCALING": "[^]", "MATURE": "[=]", "EXIT": "[x]",
    }

    for s in STATES:
        ops = by_state.get(s, [])
        icon = state_icons.get(s, "[ ]")
        print(f"  {icon} {s} ({len(ops)} ops)")
        if ops:
            ops.sort(key=lambda x: -x.get("priority_score", 0))
            for op in ops[:8]:  # Show top 8 per state
                kelly_pct = op.get("kelly_fraction", 0) * 100
                rev = op.get("total_revenue", 0)
                freq = op.get("current_freq_hours", 24)
                print(f"      {op['op_id']:<10} {op['name']:<35} "
                      f"P:{op.get('priority_score', 0):3}/100  "
                      f"K:{kelly_pct:4.1f}%  "
                      f"Rev:${rev:>8.2f}  "
                      f"Freq:{freq:5.1f}h")
            if len(ops) > 8:
                print(f"      ... and {len(ops) - 8} more")
        print()

    print(f"{'='*76}\n")


def display_signals(state):
    """Show fused signal dashboard."""
    op_confidence, category_confidence = fuse_signals(state)

    print(f"\n{'='*76}")
    print(f"  SIGNAL FUSION DASHBOARD — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"{'='*76}\n")

    print("  CATEGORY CONFIDENCE (multi-source fusion):")
    print(f"  {'─'*60}")
    sorted_cats = sorted(category_confidence.items(), key=lambda x: -x[1])
    for cat, conf in sorted_cats:
        bar_len = int(conf / 2)
        bar = "#" * bar_len + "." * (50 - bar_len)
        print(f"  {cat:<16} [{bar}] {conf:5.1f}")
    print()

    # Signal source counts
    freelance = get_freelance_demand_signals()
    ecom = get_ecom_arb_signals()
    trends = get_trend_signals()
    alpha = get_alpha_signals()

    print("  SIGNAL SOURCES:")
    print(f"  {'─'*60}")
    print(f"  Freelance demand entries:  {len(freelance):>5} (score >= 50: {sum(1 for s in freelance if s['score'] >= 50)})")
    print(f"  Ecom arb opportunities:   {len(ecom):>5} (listable: {sum(1 for s in ecom if s['action'] == 'LIST')})")
    print(f"  Trend signals:            {len(trends):>5} (strong: {sum(1 for s in trends if s['score'] >= 55)})")
    print(f"  Alpha intelligence:       {len(alpha):>5} (approved)")
    overnight = get_overnight_status()
    successes = sum(1 for e in overnight if e.get("status") == "SUCCESS")
    failures = sum(1 for e in overnight if e.get("status") in ("FAILED", "TIMEOUT"))
    print(f"  Overnight scripts:        {len(overnight):>5} (OK:{successes} FAIL:{failures})")
    print()

    # Top 10 ops by fused confidence
    print("  TOP 10 OPS BY FUSED CONFIDENCE:")
    print(f"  {'─'*60}")
    sorted_ops = sorted(op_confidence.items(), key=lambda x: -x[1])[:10]
    for oid, conf in sorted_ops:
        op = state.ops.get(oid, {})
        name = op.get("name", "?")[:35]
        s = op.get("state", "?")
        print(f"  {oid:<10} {name:<35} {conf:5.1f}  [{s}]")

    print(f"\n{'='*76}\n")


def display_kelly(state):
    """Show Kelly criterion allocation."""
    compute_kelly_fractions(state)

    print(f"\n{'='*76}")
    print(f"  KELLY CRITERION TIME ALLOCATION")
    print(f"{'='*76}\n")

    daily_hours = 16  # Available hours per day
    sorted_ops = sorted(state.ops.items(), key=lambda x: -x[1].get("kelly_fraction", 0))

    print(f"  {'Op ID':<10} {'Name':<35} {'Kelly%':>7} {'Hours/Day':>10} {'State':<12}")
    print(f"  {'─'*76}")

    total_allocated = 0
    for oid, op in sorted_ops:
        kelly = op.get("kelly_fraction", 0)
        if kelly < 0.001:
            continue
        hours = kelly * daily_hours
        total_allocated += hours
        s = op.get("state", "?")
        print(f"  {oid:<10} {op['name']:<35} {kelly*100:6.2f}% {hours:9.2f}h  [{s}]")

    print(f"  {'─'*76}")
    print(f"  {'TOTAL':<46} {'100.00%':>7} {total_allocated:9.2f}h")
    print()

    # Show zero-allocation ops
    zero_ops = [(oid, op) for oid, op in state.ops.items() if op.get("kelly_fraction", 0) < 0.001]
    if zero_ops:
        print(f"  ZERO ALLOCATION ({len(zero_ops)} ops): {', '.join(oid for oid, _ in zero_ops[:10])}")
        if len(zero_ops) > 10:
            print(f"    ... and {len(zero_ops) - 10} more")
    print(f"\n{'='*76}\n")


def display_heal(state):
    """Show diagnosis and healing recommendations."""
    diagnoses = diagnose_and_heal(state)

    print(f"\n{'='*76}")
    print(f"  SELF-HEALING DIAGNOSTIC — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"{'='*76}\n")

    severity_order = {"INFO": 0, "LOW": 1, "MEDIUM": 2, "HIGH": 3}
    diagnoses.sort(key=lambda x: -severity_order.get(x.get("severity", "LOW"), 0))

    for d in diagnoses:
        sev = d.get("severity", "LOW")
        icon = {"INFO": "[i]", "LOW": "[.]", "MEDIUM": "[!]", "HIGH": "[X]"}.get(sev, "[?]")
        print(f"  {icon} [{sev}] {d['issue']}")
        print(f"      Diagnosis: {d['diagnosis']}")
        print(f"      Action: {d['action']}")
        auto = "YES" if d.get("auto_fixable") else "NO"
        print(f"      Auto-fix: {auto}")
        print()

    print(f"{'='*76}\n")


def display_summary(state):
    """Quick 10-line system summary."""
    compute_priority_scores(state)
    compute_kelly_fractions(state)

    # Count by state
    state_counts = defaultdict(int)
    for op in state.ops.values():
        state_counts[op.get("state", "DORMANT")] += 1

    total_rev = sum(op.get("total_revenue", 0) for op in state.ops.values())
    top_ops = sorted(state.ops.values(), key=lambda x: -x.get("priority_score", 0))[:3]
    overnight = get_overnight_status()
    ok = sum(1 for e in overnight if e.get("status") == "SUCCESS")
    fail = sum(1 for e in overnight if e.get("status") in ("FAILED", "TIMEOUT"))

    print(f"\n{'='*60}")
    print(f"  OPS ORCHESTRATOR SUMMARY — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"{'='*60}")
    print(f"  Ops: {len(state.ops)} tracked | Rev: ${total_rev:.2f}")
    state_line = " | ".join(f"{s}:{c}" for s, c in sorted(state_counts.items()))
    print(f"  States: {state_line}")
    print(f"  Overnight: {ok} OK, {fail} failed out of {len(overnight)}")
    print(f"  Top 3 priority ops:")
    for op in top_ops:
        print(f"    {op['op_id']:<10} {op['name']:<30} P:{op.get('priority_score', 0)}/100 [{op.get('state', '?')}]")
    accts = get_accounts()
    active = sum(1 for v in accts.values() if v)
    print(f"  Accounts: {active}/{len(accts)} active")
    print(f"  Run: python3 AUTOMATIONS/ops_orchestrator.py --dashboard")
    print(f"{'='*60}\n")


def display_dashboard(state):
    """Full TUI-style dashboard combining all panels."""
    compute_priority_scores(state)
    compute_kelly_fractions(state)

    print(f"\n{'='*76}")
    print(f"  PRINTMAXX MASTER OPS ORCHESTRATOR DASHBOARD")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M')} | {len(state.ops)} ops | v{state.meta.get('version', '?')}")
    print(f"{'='*76}")

    # Panel 1: State distribution
    state_counts = defaultdict(int)
    for op in state.ops.values():
        state_counts[op.get("state", "DORMANT")] += 1

    print(f"\n  STATE DISTRIBUTION")
    print(f"  {'─'*60}")
    for s in STATES:
        c = state_counts.get(s, 0)
        bar = "#" * min(c, 40) + f" ({c})"
        print(f"  {s:<12} {bar}")

    # Panel 2: Execution queue
    queue = get_execution_queue(state)
    print(f"\n  EXECUTION QUEUE (top 10 ready now)")
    print(f"  {'─'*60}")
    for item in queue[:10]:
        print(f"  {item['op_id']:<10} {item['name']:<30} P:{item['priority']:3} [{item['state']}]")
    if not queue:
        print("  [No ops ready to run]")

    # Panel 3: Signal summary
    _, cat_conf = fuse_signals(state)
    top_cats = sorted(cat_conf.items(), key=lambda x: -x[1])[:5]
    print(f"\n  TOP SIGNAL CATEGORIES")
    print(f"  {'─'*60}")
    for cat, conf in top_cats:
        bar_len = int(conf / 2.5)
        bar = "#" * bar_len
        print(f"  {cat:<16} {bar} {conf:.0f}")

    # Panel 4: Transitions
    transitions = evaluate_transitions(state)
    if transitions:
        print(f"\n  STATE TRANSITIONS THIS RUN")
        print(f"  {'─'*60}")
        for oid, from_s, to_s, reason in transitions:
            print(f"  {oid:<10} {from_s} -> {to_s}: {reason}")
    else:
        print(f"\n  No state transitions triggered")

    # Panel 5: Health
    diagnoses = diagnose_and_heal(state)
    high_severity = [d for d in diagnoses if d.get("severity") in ("HIGH", "MEDIUM")]
    if high_severity:
        print(f"\n  HEALTH ALERTS ({len(high_severity)})")
        print(f"  {'─'*60}")
        for d in high_severity[:5]:
            print(f"  [{d['severity']}] {d['issue'][:60]}")

    # Panel 6: Revenue + Kelly
    total_rev = sum(op.get("total_revenue", 0) for op in state.ops.values())
    top_kelly = sorted(state.ops.items(), key=lambda x: -x[1].get("kelly_fraction", 0))[:5]
    print(f"\n  REVENUE & ALLOCATION")
    print(f"  {'─'*60}")
    print(f"  Total tracked revenue: ${total_rev:.2f}")
    print(f"  Top Kelly allocations:")
    for oid, op in top_kelly:
        k = op.get("kelly_fraction", 0) * 100
        print(f"    {oid:<10} {op['name']:<30} {k:.1f}%")

    print(f"\n{'='*76}\n")


def execute_run(state):
    """Execute the priority queue: update state, run transitions, adapt."""
    print(f"\n  Running orchestrator cycle...")

    # 1. Update momentum
    update_momentum(state)

    # 2. Compute priority scores
    compute_priority_scores(state)

    # 3. Evaluate state transitions
    transitions = evaluate_transitions(state)

    # 4. Compute Kelly fractions
    compute_kelly_fractions(state)

    # 5. Adapt frequencies
    adapt_frequencies(state)

    # 6. Get execution queue
    queue = get_execution_queue(state)

    # 7. Update run metadata
    state.meta["last_run"] = now_iso()
    state.meta["total_runs"] = state.meta.get("total_runs", 0) + 1

    # 8. Save state
    state.save()

    # 9. Log
    log_event("RUN_COMPLETE", details={
        "transitions": len(transitions),
        "queue_size": len(queue),
        "total_ops": len(state.ops),
    })

    # 10. Display results
    print(f"\n  Orchestrator cycle complete.")
    print(f"  Transitions: {len(transitions)}")
    for oid, from_s, to_s, reason in transitions:
        print(f"    {oid}: {from_s} -> {to_s} ({reason})")
    print(f"  Execution queue: {len(queue)} ops ready")
    for item in queue[:5]:
        print(f"    {item['op_id']}: {item['name']} [P:{item['priority']}]")
    print(f"  State saved to {STATE_FILE.relative_to(BASE)}")

    # 11. Log learning from this run
    learning = {
        "timestamp": now_iso(),
        "op_id": "ORCHESTRATOR",
        "event": "cycle_complete",
        "outcome": f"{len(transitions)} transitions, {len(queue)} ops in queue",
        "learning": f"Run #{state.meta['total_runs']}. "
                    f"Top priority: {queue[0]['op_id'] if queue else 'none'}. "
                    f"Transitions: {', '.join(f'{t[0]}:{t[1]}->{t[2]}' for t in transitions[:3])}",
    }
    append_jsonl(LEARNINGS_FILE, learning)

    print(f"\n")


def log_learning(text, state):
    """Log a manual learning entry."""
    entry = {
        "timestamp": now_iso(),
        "op_id": "MANUAL",
        "event": "manual_learning",
        "outcome": "logged",
        "learning": text,
    }
    append_jsonl(LEARNINGS_FILE, entry)
    log_event("LEARNING_LOGGED", details={"text": text[:200]})

    # Count total learnings
    count = 0
    try:
        with open(LEARNINGS_FILE) as f:
            count = sum(1 for _ in f)
    except Exception:
        pass

    print(f"\n  Learning logged to {LEARNINGS_FILE.relative_to(BASE)}")
    print(f"  Total learnings: {count}")
    print(f"  Text: {text[:100]}{'...' if len(text) > 100 else ''}")
    print()


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    args = sys.argv[1:]
    state = OrchestratorState.load()

    if not args:
        display_summary(state)
        return

    if "--run" in args:
        execute_run(state)

    elif "--status" in args:
        compute_priority_scores(state)
        compute_kelly_fractions(state)
        evaluate_transitions(state)
        state.save()
        display_status(state)

    elif "--adapt" in args:
        compute_priority_scores(state)
        adapt_frequencies(state)
        state.save()
        print("\n  Adaptive frequencies recalculated and saved.\n")
        # Show top changes
        sorted_ops = sorted(state.ops.items(), key=lambda x: x[1].get("current_freq_hours", 24))
        print(f"  {'Op ID':<10} {'Name':<35} {'Base':>6} {'Current':>8} {'Momentum':>9}")
        print(f"  {'─'*70}")
        for oid, op in sorted_ops[:15]:
            base = op.get("base_freq_hours", 24)
            curr = op.get("current_freq_hours", 24)
            mom = op.get("momentum", 0)
            changed = " *" if abs(curr - base) > 0.5 else ""
            print(f"  {oid:<10} {op['name']:<35} {base:5.1f}h {curr:7.1f}h {mom:8.4f}{changed}")
        print()

    elif "--signals" in args:
        compute_priority_scores(state)
        display_signals(state)

    elif "--kelly" in args:
        compute_priority_scores(state)
        compute_kelly_fractions(state)
        state.save()
        display_kelly(state)

    elif "--heal" in args:
        display_heal(state)

    elif "--learn" in args:
        idx = args.index("--learn")
        if idx + 1 < len(args):
            log_learning(args[idx + 1], state)
        else:
            print("Usage: --learn 'what you learned'")

    elif "--summary" in args:
        compute_priority_scores(state)
        compute_kelly_fractions(state)
        display_summary(state)

    elif "--dashboard" in args:
        execute_run(state)
        display_dashboard(state)

    else:
        print("""
PRINTMAXX Master Ops Orchestrator — The Brain of the Operation

Usage:
  python3 AUTOMATIONS/ops_orchestrator.py --run           Execute priority queue cycle
  python3 AUTOMATIONS/ops_orchestrator.py --status        Show all ops with states
  python3 AUTOMATIONS/ops_orchestrator.py --adapt         Recalculate adaptive frequencies
  python3 AUTOMATIONS/ops_orchestrator.py --signals       Show fused signal dashboard
  python3 AUTOMATIONS/ops_orchestrator.py --kelly         Show Kelly criterion allocation
  python3 AUTOMATIONS/ops_orchestrator.py --heal          Diagnose and fix broken ops
  python3 AUTOMATIONS/ops_orchestrator.py --learn "text"  Log a learning
  python3 AUTOMATIONS/ops_orchestrator.py --summary       Quick system summary
  python3 AUTOMATIONS/ops_orchestrator.py --dashboard     Full TUI dashboard
""")


if __name__ == "__main__":
    main()
