#!/usr/bin/env python3

from __future__ import annotations
"""
PRINTMAXX Ralph Loop Factory - Create and manage Ralph Wiggum loops for any op.

Generates canonical Ralph loop structure (prompt.md, prd.json, progress.txt, run.sh)
for any of the 182+ ops in the master spreadsheet. Follows the 10 Ralph best practices:
static prompts, filesystem memory, append-only logs, one task per iteration, quality gates,
agent picks task, small stories, git commits, max iterations, backpressure.

Usage:
    python3 AUTOMATIONS/ralph_loop_factory.py --create C01
    python3 AUTOMATIONS/ralph_loop_factory.py --create-batch CONTENT
    python3 AUTOMATIONS/ralph_loop_factory.py --list
    python3 AUTOMATIONS/ralph_loop_factory.py --status
    python3 AUTOMATIONS/ralph_loop_factory.py --run C01
    python3 AUTOMATIONS/ralph_loop_factory.py --run-batch CONTENT
    python3 AUTOMATIONS/ralph_loop_factory.py --results C01
    python3 AUTOMATIONS/ralph_loop_factory.py --quality C01
"""

import os
import sys
import json
import csv
import argparse
import subprocess
import textwrap
from datetime import datetime
from pathlib import Path

BASE = Path(__file__).resolve().parent.parent
RALPH_DIR = BASE / "ralph" / "loops"
LEDGER = BASE / "LEDGER"
OPS = BASE / "OPS"
AUTOMATIONS = BASE / "AUTOMATIONS"

PROJECT_DIR = str(BASE)

# ============================================================================
# OP REGISTRY: Dynamically loaded from PRINTMAXX_MASTER_OPS.xlsx
# ============================================================================

# Category descriptions for prompt generation
CATEGORY_DESCRIPTIONS = {
    "CONTENT": "Content creation, distribution, engagement, and monetization",
    "ECOM": "E-commerce, product listings, marketplace arbitrage, and POD",
    "DIGITAL": "Digital products, templates, courses, SaaS, and info products",
    "SERVICE": "Freelance, agency, cold outreach, and service-based revenue",
    "APP": "App building, deployment, ASO, and app monetization",
    "PERSONA": "AI influencer personas, content creation, and platform monetization",
    "INVEST": "Trading, domain flipping, portfolio allocation, and investments",
    "COMMUNITY": "Paid communities, membership platforms, and community monetization",
    "AFFILIATE": "Affiliate marketing, referral programs, and comparison sites",
    "GROWTH": "Cross-cutting growth tactics, automation, infrastructure, and systems",
}

# Default task templates by category (used when generating Ralph loops)
CATEGORY_TASK_TEMPLATES = {
    "CONTENT": [
        "research audience and trending formats",
        "create content calendar and batch plan",
        "produce initial content batch (10+ pieces)",
        "distribute across target platforms",
        "build engagement farming templates",
        "set up analytics and tracking",
        "repurpose top performers to other formats",
        "optimize and scale winning content",
    ],
    "ECOM": [
        "research products and market demand",
        "create optimized listings with SEO tags",
        "set up pricing and margin calculations",
        "manage inventory and fulfillment pipeline",
        "build marketing and promotion strategy",
        "handle fulfillment and delivery workflow",
        "analyze sales data and performance metrics",
        "scale winners and kill underperformers",
    ],
    "DIGITAL": [
        "research and validate product idea",
        "build the product or template",
        "create marketplace listing with sales copy",
        "set up pricing tiers and bundles",
        "build distribution and delivery workflow",
        "create marketing assets and funnels",
        "collect customer feedback and iterate",
        "scale to additional marketplaces",
    ],
    "SERVICE": [
        "generate leads and build prospect lists",
        "write proposals and pitch templates",
        "build portfolio samples and case studies",
        "create pricing packages (3 tiers)",
        "set up delivery workflow and SOPs",
        "manage client pipeline and communication",
        "develop upsell and cross-sell strategy",
        "collect testimonials and scale referrals",
    ],
    "APP": [
        "research market and competitor apps",
        "design and prototype UI/UX",
        "build MVP with core features",
        "test and QA (Simulator, Lighthouse > 90)",
        "deploy to production (web + app store)",
        "ASO optimization and marketing launch",
        "implement monetization (subscriptions, ads, affiliate)",
        "iterate based on user feedback and metrics",
    ],
    "PERSONA": [
        "develop character backstory and visual identity",
        "create content calendar and posting schedule",
        "set up platform accounts and profiles",
        "produce initial content batch (images, videos, text)",
        "build engagement strategy and community interaction",
        "set up monetization (subscriptions, tips, products)",
        "grow audience with cross-promotion and tactics",
        "manage brand consistency and compliance",
    ],
    "INVEST": [
        "research market and identify opportunities",
        "develop investment thesis and criteria",
        "calculate position sizing and risk limits",
        "execute entry positions",
        "monitor positions and market conditions",
        "plan and execute exit strategy",
        "review portfolio performance",
        "rebalance and reallocate capital",
    ],
    "COMMUNITY": [
        "choose platform and set up community space",
        "define content strategy and value proposition",
        "build member acquisition funnel",
        "create engagement framework (events, discussions)",
        "set up monetization (tiers, products, sponsorships)",
        "implement moderation and community guidelines",
        "plan and host events (AMAs, workshops, challenges)",
        "grow membership with referrals and partnerships",
    ],
    "AFFILIATE": [
        "research and sign up for affiliate programs",
        "create content around promoted products",
        "build comparison pages and review content",
        "place tracking links strategically",
        "drive traffic with SEO, social, and paid",
        "optimize conversion rates and EPC",
        "build relationships with affiliate managers",
        "scale top-performing programs and content",
    ],
    "GROWTH": [
        "design the growth system or infrastructure",
        "implement and configure tools",
        "test with small-scale pilot",
        "optimize based on initial results",
        "automate recurring workflows",
        "set up monitoring and alerting",
        "document SOPs for handoff and scaling",
        "scale across all ops and platforms",
    ],
}


def _load_ops_from_xlsx():
    """Load ALL ops from PRINTMAXX_MASTER_OPS.xlsx dynamically.

    Returns (OP_CATEGORIES dict, ALL_OPS flat dict) in the same structure
    as the previous hardcoded versions.
    """
    try:
        import openpyxl
    except ImportError:
        print("WARNING: openpyxl not installed. Using empty ops registry.")
        return {}, {}

    xlsx_path = BASE / "PRINTMAXX_MASTER_OPS.xlsx"
    if not xlsx_path.exists():
        print(f"WARNING: {xlsx_path} not found. Using empty ops registry.")
        return {}, {}

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        print(f"WARNING: Failed to open XLSX: {e}. Using empty ops registry.")
        return {}, {}

    if 'ALL OPS MASTER' not in wb.sheetnames:
        print("WARNING: 'ALL OPS MASTER' sheet not found in XLSX.")
        return {}, {}

    sheet = wb['ALL OPS MASTER']

    # Build column index from headers
    headers = {}
    for i, cell in enumerate(sheet[1]):
        if cell.value:
            headers[str(cell.value).strip().upper()] = i

    # Collect ops by category
    categories = {}
    all_ops = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        op_id = str(row[headers.get('OP_ID', 0)]) if row[headers.get('OP_ID', 0)] else None
        if not op_id or op_id in ('TOTAL OPS:', 'None', ''):
            continue

        cat = str(row[headers.get('CATEGORY', 1)]).upper() if row[headers.get('CATEGORY', 1)] else 'UNKNOWN'
        name = str(row[headers.get('OP_NAME', 2)]) if row[headers.get('OP_NAME', 2)] else 'Unnamed'
        desc = str(row[headers.get('DESCRIPTION', 3)]) if row[headers.get('DESCRIPTION', 3)] else ''
        platforms_str = str(row[headers.get('PLATFORMS', 15)]) if len(row) > 15 and row[headers.get('PLATFORMS', 15)] else 'All'

        # Parse platforms
        platforms = [p.strip() for p in platforms_str.split(',') if p.strip() and p.strip() not in ('None', 'N/A')]
        if not platforms:
            platforms = ['All']

        # Get task template for this category
        tasks = list(CATEGORY_TASK_TEMPLATES.get(cat, CATEGORY_TASK_TEMPLATES["GROWTH"]))

        # Customize task descriptions with op-specific context from the description
        # First task always includes the specific op context
        if desc and len(desc) > 10:
            tasks[0] = f"Research and plan: {name} - {desc[:100]}"

        # Initialize category if needed
        if cat not in categories:
            categories[cat] = {
                "description": CATEGORY_DESCRIPTIONS.get(cat, f"{cat} operations"),
                "ops": {},
            }

        # Add op to category
        op_entry = {
            "name": name,
            "platforms": platforms,
            "tasks": tasks,
        }
        categories[cat]["ops"][op_id] = op_entry

        # Add to flat lookup with category context
        all_ops[op_id] = {
            **op_entry,
            "category": cat,
            "category_desc": categories[cat]["description"],
        }

    return categories, all_ops


# Load at module level
OP_CATEGORIES, ALL_OPS = _load_ops_from_xlsx()


# ============================================================================
# PROMPT TEMPLATE GENERATORS (category-specific)
# ============================================================================

def generate_prompt_md(op_id: str, op: dict) -> str:
    """Generate a static prompt.md for a Ralph loop, tailored to the op category."""
    category = op["category"]
    tasks_list = "\n".join(f"- {t}" for t in op["tasks"])
    platforms = ", ".join(op.get("platforms", ["All"]))

    category_instructions = {
        "CONTENT": textwrap.dedent("""\
            ## Content-Specific Instructions
            - ALL content MUST follow `.claude/rules/copy-style.md` (non-negotiable)
            - No AI vocabulary: no "leverage," "comprehensive," "innovative," "seamless"
            - Consequence-first hooks, specific numbers, @pipelineabuser energy
            - Each content piece needs: hook, value delivery, CTA
            - Platform-specific formatting (character limits, hashtags, aspect ratios)
            - Track engagement metrics in output for future optimization
            - Save all content to `CONTENT/social/` or `CONTENT/` subdirectories
            - Generate minimum 3 variants per concept for A/B testing"""),
        "SERVICE": textwrap.dedent("""\
            ## Service-Specific Instructions
            - Focus on REAL deliverables: proposals, sample work, portfolio pieces
            - Every service needs: pricing tiers (3 levels), delivery timeline, scope doc
            - Lead generation is the bottleneck - always produce outreach material
            - Use existing templates from `MONEY_METHODS/LOCAL_BIZ/templates/` where applicable
            - Cold emails follow `AUTOMATIONS/generate_cold_emails.py` format
            - Track pipeline in CSV: lead -> contact -> proposal -> close -> deliver
            - Build SOPs so VAs can replicate the delivery process
            - Every iteration should produce something a client could see"""),
        "APP": textwrap.dedent("""\
            ## App-Specific Instructions
            - Quality standards: `MONEY_METHODS/APP_FACTORY/APP_QUALITY_STANDARDS.md`
            - iOS rejection prevention: `MONEY_METHODS/APP_FACTORY/IOS_REJECTION_PREVENTION.md`
            - Every app needs: onboarding (4+ screens), paywall, analytics, native plugins
            - Use Capacitor for iOS wrapping, PWA for web deployment
            - Lighthouse score > 90 before submission
            - No hover states on mobile (use active: instead)
            - RevenueCat for subscription management
            - Test in iOS Simulator before marking any task complete
            - ASO: research keywords, competitor screenshots, compelling description"""),
        "ECOM": textwrap.dedent("""\
            ## Ecom-Specific Instructions
            - Every listing needs: title (SEO optimized), description (benefit-focused), tags, pricing
            - Use existing data from `LEDGER/ECOM_ARB_OPPORTUNITIES.csv` for product selection
            - POD designs: create actual design specs, not just descriptions
            - Track all listings in CSV: platform, product, price, cost, margin, status
            - Gumroad listings follow `PRODUCTS/GUMROAD_READY_LISTINGS.md` format
            - Etsy listings follow `PRODUCTS/ECOM_LISTINGS_READY/ETSY_LISTINGS_COMPLETE.md` format
            - Calculate actual margins including platform fees, shipping, payment processing
            - Scale winners (margin > 40%), kill losers (margin < 15%)"""),
        "DIGITAL": textwrap.dedent("""\
            ## Digital Product-Specific Instructions
            - Every product needs: listing copy, cover image spec, pricing tiers, delivery method
            - Validate demand before building (search volume, competitor analysis, Reddit mentions)
            - Gumroad: use `PRODUCTS/GUMROAD_READY_LISTINGS.md` format
            - Notion templates: test in actual Notion, include preview images
            - Courses: outline first, then build module by module with real exercises
            - Chrome extensions: test in browser, submit to Chrome Web Store
            - Track: downloads, revenue, refund rate, customer feedback
            - Bundle related products for higher AOV ($9 single -> $29 bundle)"""),
        "PERSONA": textwrap.dedent("""\
            ## Persona-Specific Instructions
            - Platform compliance is CRITICAL - check TOS before every action
            - AI-generated content MUST be disclosed where required (FTC + platform rules)
            - Fanvue allows AI personas explicitly - OnlyFans does NOT
            - Content calendar: daily posting schedule with content type rotation
            - Build subscriber funnel: free content -> tiered paid access
            - Subreddit posting: follow each sub's rules, no spam, genuine engagement
            - Track: subscribers, revenue per sub, churn rate, content performance
            - Reference `MONEY_METHODS/AI_INFLUENCER/AI_NSFW_EXECUTION_FULL.md` for compliance"""),
        "INVEST": textwrap.dedent("""\
            ## Investment-Specific Instructions
            - NEVER risk more than 5% of capital on any single position
            - Meme coins: max $5-20/bet, track in `LEDGER/MEMECOIN_PORTFOLIO.csv`
            - Domain flipping: research exact match domains, check trademark conflicts
            - Algo trading: paper trade first, backtest minimum 90 days before live
            - All positions tracked in `FINANCIALS/INVESTMENT_PORTFOLIO.csv`
            - Kelly Criterion for position sizing (calculate win rate first)
            - Set stop-loss and take-profit levels before entry
            - Review portfolio weekly, rebalance monthly"""),
        "COMMUNITY": textwrap.dedent("""\
            ## Community-Specific Instructions
            - Community value must be genuine - no empty paid groups
            - Platform comparison: Skool ($99/mo, built-in gamification), Discord (free, flexible),
              Telegram (free, mobile-first), Whop (marketplace discovery)
            - Free tier to paid conversion: provide real value free, premium = access + exclusivity
            - Engagement is the metric: daily active members, not total members
            - Moderation plan required before launch (rules, enforcement, escalation)
            - Events drive retention: weekly AMAs, monthly challenges, quarterly reviews
            - Cross-promote with newsletter, apps, and content for member acquisition
            - Track: MRR, churn rate, DAU/MAU ratio, NPS score"""),
        "AFFILIATE": textwrap.dedent("""\
            ## Affiliate-Specific Instructions
            - Sign up for programs with real accounts only (human step if needed)
            - Create genuine content around products (not thin affiliate spam)
            - FTC disclosure required on ALL affiliate content
            - Track: clicks, conversions, revenue, EPC (earnings per click)
            - Programmatic SEO pages at `builds/programmatic_seo/`
            - Integrate affiliate links into existing apps and content
            - Reference `OPS/AFFILIATE_LAUNCH_CHECKLIST.md` for program list
            - Build comparison tables and "best X for Y" content (highest converting format)"""),
        "GROWTH": textwrap.dedent("""\
            ## Growth-Specific Instructions
            - Account creation requires human-in-loop for credentials
            - Anti-detect browser setup: `OPS/BROWSER_AUTOMATION_SETUP.md`
            - Follow platform rate limits strictly - account bans = catastrophic
            - Warmup schedules in `ralph/loops/social_setup/output/`
            - Track account health: engagement rate, follower growth, restriction status
            - Cross-pollination matrix: `LEDGER/CROSS_POLLINATION_MATRIX.csv`
            - Every system must be documented for handoff (SOPs, not tribal knowledge)
            - Automation scripts go in `AUTOMATIONS/` with CLI flags"""),
    }

    category_section = category_instructions.get(category, "")

    prompt = textwrap.dedent(f"""\
    # Ralph Loop: {op["name"]} ({op_id})

    **Category:** {category} - {op["category_desc"]}
    **Platforms:** {platforms}
    **Op ID:** {op_id}

    **Working directory:** {PROJECT_DIR}

    ---

    ## Your Mission

    You are a Ralph Wiggum loop agent executing tasks for the **{op["name"]}** operation.
    Each iteration, you:
    1. Read `prd.json` in this directory to find the next task where `passes` is `false`
    2. Read `progress.txt` for learnings from previous iterations
    3. Execute ONE task completely
    4. Verify output quality against quality gates
    5. Update `prd.json` (set `passes: true` for completed task)
    6. Append learnings to `progress.txt`
    7. Exit cleanly (context is destroyed, memory lives in files)

    ## Available Tasks

    {tasks_list}

    {category_section}

    ## Quality Gates (CHECK BEFORE MARKING DONE)

    Every task output must pass ALL of these:
    1. **Real output exists** - an actual file was created/modified (not just planned)
    2. **Content quality** - passes copy-style.md checks (no AI vocabulary, specific numbers)
    3. **Actionable** - someone could use this output immediately without follow-up
    4. **Integrated** - output connects to existing project infrastructure
    5. **Tracked** - any new data logged to appropriate LEDGER CSV

    If quality gate fails, DO NOT mark the task as done. Instead:
    - Append failure reason to `progress.txt`
    - Try to fix the issue in this iteration
    - If unfixable, leave `passes: false` and exit (next iteration gets a fresh start)

    ## File Paths Reference

    | Resource | Path |
    |----------|------|
    | Project root | `{PROJECT_DIR}` |
    | This loop dir | `{PROJECT_DIR}/ralph/loops/{op_id.lower()}` |
    | PRD (task list) | `{PROJECT_DIR}/ralph/loops/{op_id.lower()}/prd.json` |
    | Progress log | `{PROJECT_DIR}/ralph/loops/{op_id.lower()}/progress.txt` |
    | Copy style rules | `{PROJECT_DIR}/.claude/rules/copy-style.md` |
    | LEDGER (data) | `{PROJECT_DIR}/LEDGER/` |
    | Products | `{PROJECT_DIR}/PRODUCTS/` |
    | Content | `{PROJECT_DIR}/CONTENT/` |
    | Money Methods | `{PROJECT_DIR}/MONEY_METHODS/` |
    | Automations | `{PROJECT_DIR}/AUTOMATIONS/` |

    ## Backpressure Rules

    - If you notice output quality dropping, STOP and append a note to progress.txt
    - If a task is too large, break it into subtasks in progress.txt (next iteration picks up)
    - If blocked on a human action, note it in progress.txt and move to next unblocked task
    - Maximum ONE task per iteration. Do it well. Exit cleanly.

    ## Output Format

    Every file you create must include at the top:
    ```
    # Generated by Ralph Loop: {op_id} - {op["name"]}
    # Iteration: [number]
    # Date: [ISO date]
    # Task: [task description]
    ```

    ## Iteration Steps (EXACT ORDER)

    1. `cat prd.json` - find first task where `passes: false`
    2. `cat progress.txt` - read previous learnings
    3. Execute the task (create files, run scripts, etc.)
    4. Verify output against quality gates
    5. Update prd.json: set `passes: true` for completed task
    6. Append to progress.txt: what was done, what was learned, any issues
    7. If ALL tasks have `passes: true`, output `<promise>COMPLETE</promise>`
    8. Exit
    """)
    return prompt


def generate_prd_json(op_id: str, op: dict) -> dict:
    """Generate prd.json with atomic tasks for the op."""
    stories = []
    for i, task in enumerate(op["tasks"]):
        stories.append({
            "id": f"{op_id}-T{i+1:02d}",
            "description": task.capitalize(),
            "passes": False,
            "priority": i + 1,
            "dependencies": [],
            "estimated_iterations": 1,
            "output_files": [],
            "notes": ""
        })
    return {
        "op_id": op_id,
        "op_name": op["name"],
        "category": op["category"],
        "created": datetime.now().isoformat(),
        "stories": stories
    }


def generate_run_sh(op_id: str, op: dict) -> str:
    """Generate the canonical Ralph loop runner script."""
    loop_dir = f"{PROJECT_DIR}/ralph/loops/{op_id.lower()}"
    return textwrap.dedent(f"""\
    #!/bin/bash
    # Ralph Loop Runner: {op["name"]} ({op_id})
    # Generated: {datetime.now().isoformat()}
    #
    # Canonical Ralph Wiggum pattern:
    # - Fresh context every iteration
    # - State lives in files (prd.json, progress.txt)
    # - One task per iteration
    # - Quality gates before marking done
    # - Backpressure on quality drops

    LOOP_DIR="{loop_dir}"
    PROJECT_DIR="{PROJECT_DIR}"
    MAX_ITERATIONS=${{1:-20}}
    SLEEP_BETWEEN=${{2:-5}}

    cd "$PROJECT_DIR"

    # SAFETY: Load guardrails wrapper
    source "$PROJECT_DIR/AUTOMATIONS/guardrails_wrapper.sh"

    echo "=== Ralph Loop: {op["name"]} ({op_id}) ==="
    echo "Max iterations: $MAX_ITERATIONS"
    echo "Sleep between: ${{SLEEP_BETWEEN}}s"
    echo ""

    iteration=1

    while [ $iteration -le $MAX_ITERATIONS ]; do
        echo "--- Iteration $iteration / $MAX_ITERATIONS ---"
        echo "$(date +%Y-%m-%dT%H:%M:%S) Starting iteration $iteration" >> "$LOOP_DIR/progress.txt"

        # Check if all tasks complete
        if python3 -c "
    import json
    with open('$LOOP_DIR/prd.json') as f:
        prd = json.load(f)
    if all(s['passes'] for s in prd['stories']):
        print('COMPLETE')
        exit(0)
    exit(1)
    " 2>/dev/null; then
            echo "All tasks complete!"
            break
        fi

        # Run one iteration with fresh context
        cat "$LOOP_DIR/prompt.md" | claude --print --dangerously-skip-permissions

        echo ""
        echo "Iteration $iteration complete. Sleeping ${{SLEEP_BETWEEN}}s..."
        iteration=$((iteration + 1))
        sleep "$SLEEP_BETWEEN"
    done

    echo ""
    echo "=== Loop finished after $((iteration - 1)) iterations ==="
    echo "$(date +%Y-%m-%dT%H:%M:%S) Loop finished after $((iteration - 1)) iterations" >> "$LOOP_DIR/progress.txt"
    """)


# ============================================================================
# FACTORY OPERATIONS
# ============================================================================

def create_loop(op_id: str) -> bool:
    """Create a complete Ralph loop for an op."""
    op_id = op_id.upper()
    if op_id not in ALL_OPS:
        print(f"ERROR: Unknown op ID '{op_id}'. Use --list-ops to see available ops.")
        return False

    op = ALL_OPS[op_id]
    loop_dir = RALPH_DIR / op_id.lower()

    if loop_dir.exists():
        print(f"WARNING: Loop already exists at {loop_dir}")
        print("  Use --status to check its state, or delete the directory to recreate.")
        return False

    loop_dir.mkdir(parents=True, exist_ok=True)

    # 1. Generate prompt.md (static, never changes)
    prompt_path = loop_dir / "prompt.md"
    prompt_path.write_text(generate_prompt_md(op_id, op))

    # 2. Generate prd.json (task list)
    prd_path = loop_dir / "prd.json"
    prd_path.write_text(json.dumps(generate_prd_json(op_id, op), indent=2))

    # 3. Create empty progress.txt
    progress_path = loop_dir / "progress.txt"
    progress_path.write_text(
        f"# Progress Log: {op['name']} ({op_id})\n"
        f"# Created: {datetime.now().isoformat()}\n"
        f"# Category: {op['category']}\n\n"
    )

    # 4. Generate run.sh
    run_path = loop_dir / "run.sh"
    run_path.write_text(generate_run_sh(op_id, op))
    os.chmod(str(run_path), 0o755)

    print(f"CREATED loop: {op_id} - {op['name']}")
    print(f"  Directory: {loop_dir}")
    print(f"  Files: prompt.md, prd.json, progress.txt, run.sh")
    print(f"  Tasks: {len(op['tasks'])}")
    print(f"  Run: bash {loop_dir}/run.sh")
    return True


def create_batch(category: str) -> int:
    """Create loops for all ops in a category."""
    category = category.upper()
    if category not in OP_CATEGORIES:
        print(f"ERROR: Unknown category '{category}'.")
        print(f"  Available: {', '.join(OP_CATEGORIES.keys())}")
        return 0

    created = 0
    for op_id in OP_CATEGORIES[category]["ops"]:
        if create_loop(op_id):
            created += 1
        print()
    print(f"Created {created} loops for category {category}")
    return created


def list_loops():
    """List all created Ralph loops and their status."""
    print("=" * 80)
    print("RALPH LOOP REGISTRY")
    print("=" * 80)

    # List available ops (not yet created)
    created_ops = set()
    if RALPH_DIR.exists():
        for d in RALPH_DIR.iterdir():
            if d.is_dir() and (d / "prd.json").exists():
                created_ops.add(d.name.upper())

    print(f"\nTotal ops defined: {len(ALL_OPS)}")
    print(f"Loops created: {len(created_ops)}")
    print()

    for cat, data in OP_CATEGORIES.items():
        print(f"--- {cat}: {data['description']} ---")
        for op_id, op_data in data["ops"].items():
            status = "CREATED" if op_id in created_ops else "available"
            marker = "[X]" if op_id in created_ops else "[ ]"

            # Check progress if created
            extra = ""
            if op_id in created_ops:
                loop_dir = RALPH_DIR / op_id.lower()
                prd_path = loop_dir / "prd.json"
                if prd_path.exists():
                    try:
                        prd = json.loads(prd_path.read_text())
                        done = sum(1 for s in prd["stories"] if s["passes"])
                        total = len(prd["stories"])
                        extra = f" ({done}/{total} tasks done)"
                    except Exception:
                        extra = " (prd.json parse error)"

            print(f"  {marker} {op_id}: {op_data['name']} - {status}{extra}")
        print()


def check_status():
    """Detailed status of all created loops."""
    print("=" * 80)
    print("RALPH LOOP STATUS DASHBOARD")
    print(f"Timestamp: {datetime.now().isoformat()}")
    print("=" * 80)

    if not RALPH_DIR.exists():
        print("\nNo loops directory found.")
        return

    loops = []
    for d in sorted(RALPH_DIR.iterdir()):
        if not d.is_dir():
            continue
        prd_path = d / "prd.json"
        if not prd_path.exists():
            continue

        op_id = d.name.upper()
        try:
            prd = json.loads(prd_path.read_text())
            stories = prd.get("stories", [])
            done = sum(1 for s in stories if s["passes"])
            total = len(stories)
            pct = (done / total * 100) if total > 0 else 0

            # Check progress.txt for last activity
            progress_path = d / "progress.txt"
            last_line = ""
            if progress_path.exists():
                lines = progress_path.read_text().strip().split("\n")
                for line in reversed(lines):
                    if line.strip() and not line.startswith("#"):
                        last_line = line.strip()[:60]
                        break

            status = "COMPLETE" if done == total else "IN_PROGRESS" if done > 0 else "PENDING"
            loops.append({
                "op_id": op_id,
                "name": prd.get("op_name", d.name),
                "category": prd.get("category", "?"),
                "done": done,
                "total": total,
                "pct": pct,
                "status": status,
                "last_activity": last_line,
            })
        except Exception as e:
            loops.append({
                "op_id": op_id,
                "name": d.name,
                "category": "?",
                "done": 0,
                "total": 0,
                "pct": 0,
                "status": "ERROR",
                "last_activity": str(e)[:60],
            })

    if not loops:
        print("\nNo loops created yet. Use --create <OP_ID> to create one.")
        return

    # Summary
    complete = sum(1 for l in loops if l["status"] == "COMPLETE")
    in_progress = sum(1 for l in loops if l["status"] == "IN_PROGRESS")
    pending = sum(1 for l in loops if l["status"] == "PENDING")

    print(f"\n  COMPLETE: {complete}  |  IN_PROGRESS: {in_progress}  |  PENDING: {pending}  |  TOTAL: {len(loops)}")
    print()

    # Table
    print(f"{'OP':>6}  {'STATUS':>12}  {'PROG':>8}  {'NAME':<35}  {'LAST ACTIVITY':<50}")
    print("-" * 115)
    for l in loops:
        bar_len = 8
        filled = int(l["pct"] / 100 * bar_len)
        bar = "#" * filled + "." * (bar_len - filled)
        prog = f"{l['done']}/{l['total']}"
        print(f"{l['op_id']:>6}  {l['status']:>12}  {prog:>8}  {l['name']:<35}  {l['last_activity']:<50}")


def run_loop(op_id: str):
    """Launch a Ralph loop for an op."""
    op_id = op_id.upper()
    loop_dir = RALPH_DIR / op_id.lower()
    run_path = loop_dir / "run.sh"

    if not run_path.exists():
        print(f"ERROR: Loop not found for {op_id}. Create it first with --create {op_id}")
        return

    print(f"Launching Ralph loop for {op_id}...")
    print(f"  Script: {run_path}")
    print(f"  To stop: Ctrl+C")
    print()

    subprocess.run(["bash", str(run_path)], cwd=str(BASE))


def run_batch(category: str):
    """Launch all loops in a category sequentially."""
    category = category.upper()
    if category not in OP_CATEGORIES:
        print(f"ERROR: Unknown category '{category}'.")
        return

    for op_id in OP_CATEGORIES[category]["ops"]:
        loop_dir = RALPH_DIR / op_id.lower()
        if (loop_dir / "run.sh").exists():
            print(f"\n=== Running {op_id} ===")
            subprocess.run(["bash", str(loop_dir / "run.sh")], cwd=str(BASE))
        else:
            print(f"SKIP: {op_id} (not created)")


def show_results(op_id: str):
    """Show results and progress for a loop."""
    op_id = op_id.upper()
    loop_dir = RALPH_DIR / op_id.lower()

    if not loop_dir.exists():
        print(f"ERROR: Loop not found for {op_id}")
        return

    print(f"=== Results: {op_id} ===\n")

    # Show PRD status
    prd_path = loop_dir / "prd.json"
    if prd_path.exists():
        prd = json.loads(prd_path.read_text())
        print(f"Op: {prd.get('op_name', '?')}")
        print(f"Category: {prd.get('category', '?')}")
        print(f"Created: {prd.get('created', '?')}")
        print()

        stories = prd.get("stories", [])
        done = sum(1 for s in stories if s["passes"])
        print(f"Tasks: {done}/{len(stories)} complete")
        print()

        for s in stories:
            marker = "[X]" if s["passes"] else "[ ]"
            print(f"  {marker} {s['id']}: {s['description']}")
            if s.get("output_files"):
                for f in s["output_files"]:
                    print(f"        -> {f}")
            if s.get("notes"):
                print(f"        Note: {s['notes']}")
        print()

    # Show progress log
    progress_path = loop_dir / "progress.txt"
    if progress_path.exists():
        content = progress_path.read_text().strip()
        if content:
            print("--- Progress Log ---")
            lines = content.split("\n")
            # Show last 30 lines
            for line in lines[-30:]:
                print(f"  {line}")
            if len(lines) > 30:
                print(f"  ... ({len(lines) - 30} earlier lines omitted)")


def quality_check(op_id: str):
    """Run quality gate checks on a loop's output."""
    op_id = op_id.upper()
    loop_dir = RALPH_DIR / op_id.lower()

    if not loop_dir.exists():
        print(f"ERROR: Loop not found for {op_id}")
        return

    print(f"=== Quality Check: {op_id} ===\n")

    prd_path = loop_dir / "prd.json"
    if not prd_path.exists():
        print("ERROR: prd.json not found")
        return

    prd = json.loads(prd_path.read_text())
    stories = prd.get("stories", [])
    issues = []

    for s in stories:
        if not s["passes"]:
            continue  # skip incomplete tasks

        task_issues = []

        # Check if output files exist
        if s.get("output_files"):
            for f in s["output_files"]:
                fpath = BASE / f
                if not fpath.exists():
                    task_issues.append(f"Output file missing: {f}")
                else:
                    content = fpath.read_text()
                    # Check for AI vocabulary
                    ai_words = ["leverage", "comprehensive", "innovative", "seamless",
                                "game-changer", "cutting-edge", "delve", "robust",
                                "empower", "unlock"]
                    found_ai = [w for w in ai_words if w.lower() in content.lower()]
                    if found_ai:
                        task_issues.append(f"AI vocabulary in {f}: {', '.join(found_ai)}")

                    # Check for em dashes
                    if "\u2014" in content:
                        task_issues.append(f"Em dashes found in {f}")

        if task_issues:
            issues.append({"task": s["id"], "issues": task_issues})
            print(f"  FAIL {s['id']}: {s['description']}")
            for issue in task_issues:
                print(f"    - {issue}")
        else:
            print(f"  PASS {s['id']}: {s['description']}")

    print()
    total = sum(1 for s in stories if s["passes"])
    failed = len(issues)
    print(f"Completed tasks: {total}")
    print(f"Quality issues: {failed}")
    if failed == 0 and total > 0:
        print("Quality: ALL CLEAR")
    elif failed > 0:
        print(f"Quality: {failed} task(s) need attention")


def list_ops():
    """List all available op IDs."""
    print("Available Op IDs:")
    print()
    for cat, data in OP_CATEGORIES.items():
        print(f"  {cat}:")
        for op_id, op_data in data["ops"].items():
            print(f"    {op_id}: {op_data['name']}")
    print()
    print(f"Categories: {', '.join(OP_CATEGORIES.keys())}")


# ============================================================================
# CLI
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="PRINTMAXX Ralph Loop Factory - Create and manage Ralph loops for any op"
    )
    parser.add_argument("--create", metavar="OP_ID", help="Create loop for an op (e.g. C01, S01)")
    parser.add_argument("--create-batch", metavar="CATEGORY", help="Create loops for all ops in category")
    parser.add_argument("--list", action="store_true", help="List all loops and their status")
    parser.add_argument("--list-ops", action="store_true", help="List all available op IDs")
    parser.add_argument("--status", action="store_true", help="Detailed status dashboard")
    parser.add_argument("--run", metavar="OP_ID", help="Launch loop for an op")
    parser.add_argument("--run-batch", metavar="CATEGORY", help="Launch all loops in a category")
    parser.add_argument("--results", metavar="OP_ID", help="Show results/progress for a loop")
    parser.add_argument("--quality", metavar="OP_ID", help="Run quality check on a loop's output")

    args = parser.parse_args()

    if args.create:
        create_loop(args.create)
    elif args.create_batch:
        create_batch(args.create_batch)
    elif args.list:
        list_loops()
    elif args.list_ops:
        list_ops()
    elif args.status:
        check_status()
    elif args.run:
        run_loop(args.run)
    elif args.run_batch:
        run_batch(args.run_batch)
    elif args.results:
        show_results(args.results)
    elif args.quality:
        quality_check(args.quality)
    else:
        parser.print_help()
        print("\nQuick Start:")
        print("  python3 AUTOMATIONS/ralph_loop_factory.py --list-ops    # See all available ops")
        print("  python3 AUTOMATIONS/ralph_loop_factory.py --create S01  # Create a loop")
        print("  python3 AUTOMATIONS/ralph_loop_factory.py --list        # See created loops")
        print("  python3 AUTOMATIONS/ralph_loop_factory.py --status      # Detailed dashboard")


if __name__ == "__main__":
    main()
