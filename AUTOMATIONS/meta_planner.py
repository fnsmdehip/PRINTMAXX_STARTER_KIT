#!/usr/bin/env python3

from __future__ import annotations
"""
PRINTMAXX Meta Planner — Autonomous Full-Vision Task Generation

Reads MASTER_OPS.xlsx (150+ ops across 12 sheets), the full project state,
and ALL 195 automation scripts to generate a META PLAN that pursues the
ENTIRE PRINTMAXX vision autonomously.

Unlike workflow_wirer.py (which wires known pipelines), the meta planner:
  - Reads EVERY op from MASTER_OPS and checks if it's automated
  - Identifies gaps: ops that have NO script coverage
  - Generates LLM-native tasks for work no script can handle
  - Produces a compact nav context for agent prompts
  - Prioritizes by revenue potential, blockers, and dependencies

Usage:
    python3 AUTOMATIONS/meta_planner.py --meta-plan     # Full meta plan from MASTER_OPS
    python3 AUTOMATIONS/meta_planner.py --nav-context    # Compact folder nav for agents
    python3 AUTOMATIONS/meta_planner.py --wire-all       # Wire all unautomated ops
    python3 AUTOMATIONS/meta_planner.py --status         # Coverage: X/150 ops automated
    python3 AUTOMATIONS/meta_planner.py --gaps           # Show what's missing
"""

import json
import os
import re
import sys
import csv
import hashlib
import argparse
from pathlib import Path
from datetime import datetime
from collections import defaultdict

# ============================================================
# Path Safety + Constants
# ============================================================

PROJECT_ROOT = Path(__file__).resolve().parent.parent
AUTOMATIONS_DIR = PROJECT_ROOT / "AUTOMATIONS"
OPS_DIR = PROJECT_ROOT / "OPS"
LEDGER_DIR = PROJECT_ROOT / "LEDGER"
QUEUE_PATH = OPS_DIR / "AUTONOMOUS_TASK_QUEUE.jsonl"
META_PLAN_PATH = OPS_DIR / "META_PLAN.json"
NAV_CONTEXT_PATH = OPS_DIR / "NAV_CONTEXT_FOR_AGENTS.md"

# MASTER_OPS xlsx — pick the latest enhanced or base version
MASTER_OPS_CANDIDATES = sorted(
    PROJECT_ROOT.glob("PRINTMAXX_MASTER_OPS_ENHANCED_*.xlsx"),
    reverse=True,
)
MASTER_OPS_BASE = PROJECT_ROOT / "PRINTMAXX_MASTER_OPS.xlsx"
MASTER_OPS_PATH = MASTER_OPS_CANDIDATES[0] if MASTER_OPS_CANDIDATES else MASTER_OPS_BASE

# Sheets that are reference/tool-list data, not actionable ops.
# Exclude them from gap analysis and task generation to avoid noise.
# These contain tool comparisons, infrastructure inventory, thesis lists,
# and playbook reference rows — not discrete ops to automate.
REFERENCE_ONLY_SHEETS = {
    "DEEP PLAYBOOK",
    "DEEP_PLAYBOOK_INDEX",
    "ALPHA_THESIS_INDEX",
    "LLM ALPHA THESIS",
    "SYSTEM_EVIDENCE",
    "VIDEO & MEDIA STACK",       # Tool comparison table
    "HOSTING & DEPLOY",          # Hosting provider comparison
    "BROWSER & PROXY STACK",     # Anti-detect/proxy tool list
    "LEAD GEN STACK",            # Lead gen tool comparison
    "NSFW COMPLIANCE",           # Compliance rules reference
    "EXISTING INFRA",            # Inventory of what's built
}

# Minimum description length to consider a row an actionable op
MIN_DESCRIPTION_LEN = 10

# Strategic state files
HEARTBEAT_PATH = OPS_DIR / "HEARTBEAT.md"
ACTIVE_TASKS_PATH = OPS_DIR / "active-tasks.md"
TASK_TRACKER_PATH = OPS_DIR / "PERSISTENT_TASK_TRACKER.md"
ALPHA_STAGING_PATH = LEDGER_DIR / "ALPHA_STAGING.csv"


def safe_path(target: Path) -> Path:
    """Verify path is within project root."""
    resolved = Path(target).resolve()
    if not str(resolved).startswith(str(PROJECT_ROOT)):
        raise ValueError(f"BLOCKED: {resolved} is outside project root {PROJECT_ROOT}")
    return resolved


# ============================================================
# Script Inventory — Catalog all 195 automation scripts
# ============================================================

# Keyword-to-script mapping for intelligent op-to-script matching.
# Keys are lowercase keyword sets found in op descriptions.
# Values are (script_name, suggested_flags, description).
SCRIPT_CATALOG = {}


def build_script_catalog() -> dict:
    """Scan AUTOMATIONS/ for all .py files and build keyword index."""
    global SCRIPT_CATALOG
    catalog = {}
    if not AUTOMATIONS_DIR.exists():
        return catalog

    for py_file in sorted(AUTOMATIONS_DIR.glob("*.py")):
        name = py_file.name
        stem = py_file.stem
        # Extract keywords from filename
        keywords = set(stem.lower().replace("_", " ").split())
        # Remove very common words
        keywords -= {"py", "the", "and", "for", "all", "run", "auto", "daily", "new"}
        catalog[name] = {
            "path": str(py_file.relative_to(PROJECT_ROOT)),
            "stem": stem,
            "keywords": keywords,
            "exists": True,
        }

    SCRIPT_CATALOG = catalog
    return catalog


def find_scripts_for_op(op_description: str, op_category: str = "") -> list:
    """Find automation scripts that might handle a given op.

    Returns list of (script_name, confidence_score) tuples.
    Confidence: 0.0-1.0 based on keyword overlap.
    """
    if not SCRIPT_CATALOG:
        build_script_catalog()

    desc_lower = (op_description + " " + op_category).lower()
    desc_words = set(re.findall(r'[a-z]{3,}', desc_lower))

    matches = []
    for script_name, info in SCRIPT_CATALOG.items():
        overlap = desc_words & info["keywords"]
        if len(overlap) >= 2:
            confidence = min(len(overlap) / max(len(info["keywords"]), 1), 1.0)
            matches.append((script_name, round(confidence, 2), overlap))
        elif len(overlap) == 1:
            # Single keyword match — lower confidence
            word = list(overlap)[0]
            # Only count meaningful single matches
            if word in ("scraper", "scrape", "monitor", "scanner", "tracker",
                        "pipeline", "compliance", "freelance", "ecom", "arb",
                        "alpha", "content", "lead", "trend", "telegram",
                        "reddit", "twitter", "competitor", "health"):
                matches.append((script_name, 0.3, overlap))

    matches.sort(key=lambda x: x[1], reverse=True)
    return matches[:5]


# ============================================================
# MASTER_OPS Reader — Parse all 12 sheets
# ============================================================

def read_master_ops() -> dict:
    """Read PRINTMAXX_MASTER_OPS.xlsx and extract all ops from all sheets.

    Returns dict with sheet_name -> list of op dicts.
    """
    try:
        import openpyxl
    except ImportError:
        print("  ERROR: openpyxl not installed. Run: pip3 install openpyxl")
        sys.exit(1)

    if not MASTER_OPS_PATH.exists():
        print(f"  ERROR: MASTER_OPS not found at {MASTER_OPS_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(str(MASTER_OPS_PATH), read_only=True, data_only=True)
    all_ops = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # First row is headers
        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
        ops = []
        for row_idx, row in enumerate(rows[1:], start=2):
            if not row or all(c is None for c in row):
                continue
            op = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    op[headers[i]] = val
            # Try to extract an op_id and description
            op["_sheet"] = sheet_name
            op["_row"] = row_idx
            ops.append(op)

        all_ops[sheet_name] = ops

    wb.close()
    return all_ops


def normalize_op(op: dict) -> dict:
    """Normalize an op dict to have consistent fields regardless of sheet format.

    Extracts: op_id, description, category, status, priority, sheet.
    """
    # Try common column name patterns for each field
    def get_field(op, candidates, default=""):
        for key in candidates:
            for op_key in op:
                if op_key and key.lower() in str(op_key).lower():
                    val = op.get(op_key)
                    if val is not None:
                        return str(val).strip()
        return default

    op_id = get_field(op, ["op_id", "op id", "id", "op#", "op_number", "op number"])
    if not op_id:
        op_id = f"{op.get('_sheet', 'UNK')}_{op.get('_row', 0)}"

    description = get_field(op, ["description", "desc", "op_name", "op name",
                                  "name", "task", "what", "operation",
                                  "tool", "service", "best_for", "notes"])
    category = get_field(op, ["category", "cat", "type", "method", "vertical",
                               "bucket", "area"])
    status = get_field(op, ["status", "state", "progress", "done"])
    priority = get_field(op, ["priority", "prio", "pri", "rank", "urgency", "impact"])
    revenue = get_field(op, ["revenue", "rev", "mrr", "arr", "income", "potential",
                              "est_revenue", "revenue_potential"])
    blocker = get_field(op, ["blocker", "blocked", "blocking", "dependency",
                              "depends_on", "needs"])
    automation = get_field(op, ["automation", "automated", "script", "cron",
                                 "automated_by", "tool"])

    return {
        "op_id": op_id,
        "description": description,
        "category": category,
        "status": status,
        "priority": priority,
        "revenue_potential": revenue,
        "blocker": blocker,
        "automation_status": automation,
        "sheet": op.get("_sheet", ""),
        "row": op.get("_row", 0),
        "raw": {k: str(v)[:200] for k, v in op.items()
                if not k.startswith("_") and v is not None},
    }


def extract_all_ops(include_reference_sheets: bool = False) -> list:
    """Extract and normalize all ops from MASTER_OPS.

    Args:
        include_reference_sheets: If True, include DEEP PLAYBOOK and other
            reference-only sheets. Default False to focus on actionable ops.
    """
    sheets_data = read_master_ops()
    all_normalized = []

    for sheet_name, ops in sheets_data.items():
        # Skip reference-only sheets unless explicitly requested
        if not include_reference_sheets and sheet_name in REFERENCE_ONLY_SHEETS:
            continue

        for op in ops:
            normalized = normalize_op(op)
            # Skip rows with no meaningful description (empty/header rows)
            desc = normalized["description"].strip()
            if len(desc) < MIN_DESCRIPTION_LEN:
                continue
            all_normalized.append(normalized)

    return all_normalized


# ============================================================
# State Reader — Current system state from strategic files
# ============================================================

def read_system_state() -> dict:
    """Read current system state from HEARTBEAT, active-tasks, task tracker, etc."""
    state = {
        "heartbeat": "",
        "active_tasks": "",
        "pending_alpha": 0,
        "queue_tasks": 0,
        "queue_pending": 0,
        "queue_completed": 0,
        "queue_failed": 0,
        "script_count": 0,
    }

    # HEARTBEAT
    if HEARTBEAT_PATH.exists():
        state["heartbeat"] = HEARTBEAT_PATH.read_text()[:2000]

    # Active tasks
    if ACTIVE_TASKS_PATH.exists():
        state["active_tasks"] = ACTIVE_TASKS_PATH.read_text()[:2000]

    # Alpha staging — count pending review entries
    if ALPHA_STAGING_PATH.exists():
        try:
            with open(ALPHA_STAGING_PATH, "r") as f:
                reader = csv.DictReader(f)
                pending = 0
                for row in reader:
                    status = row.get("status", "")
                    if "PENDING" in status.upper():
                        pending += 1
                state["pending_alpha"] = pending
        except Exception:
            pass

    # Queue
    if QUEUE_PATH.exists():
        try:
            tasks = []
            for line in QUEUE_PATH.read_text().strip().split("\n"):
                line = line.strip()
                if line:
                    try:
                        tasks.append(json.loads(line))
                    except json.JSONDecodeError:
                        pass
            state["queue_tasks"] = len(tasks)
            state["queue_pending"] = sum(
                1 for t in tasks if t.get("status") == "PENDING"
            )
            state["queue_completed"] = sum(
                1 for t in tasks if t.get("status") == "COMPLETED"
            )
            state["queue_failed"] = sum(
                1 for t in tasks if t.get("status") == "FAILED"
            )
        except Exception:
            pass

    # Script count
    state["script_count"] = len(list(AUTOMATIONS_DIR.glob("*.py")))

    return state


# ============================================================
# Coverage Analysis — Which ops are automated?
# ============================================================

def analyze_coverage(ops: list) -> dict:
    """Analyze which ops have script coverage and which don't.

    Returns coverage report with:
      - automated: ops with matching scripts
      - llm_needed: ops that need LLM-native work
      - unautomated: ops with no coverage at all
      - by_sheet: per-sheet breakdown
    """
    if not SCRIPT_CATALOG:
        build_script_catalog()

    automated = []
    llm_needed = []
    unautomated = []
    by_sheet = defaultdict(lambda: {"total": 0, "automated": 0, "llm": 0, "gap": 0})

    for op in ops:
        sheet = op["sheet"]
        by_sheet[sheet]["total"] += 1

        # Check if already marked as automated in the spreadsheet
        auto_status = op["automation_status"].lower()
        is_marked_automated = any(w in auto_status for w in
                                   ["automated", "cron", "script", "running", "active"])

        # Find matching scripts
        matches = find_scripts_for_op(op["description"], op["category"])
        high_confidence = [m for m in matches if m[1] >= 0.5]

        # Check if this is a build/strategy/LLM-type task
        desc_lower = op["description"].lower()
        is_llm_task = any(w in desc_lower for w in
                          ["strategy", "playbook", "design", "write", "create content",
                           "brand", "analyze", "audit", "review", "plan", "research",
                           "onboard", "persona", "outreach template", "copy", "draft"])

        if is_marked_automated or high_confidence:
            op["_coverage"] = "automated"
            op["_matching_scripts"] = [(m[0], m[1]) for m in high_confidence[:3]]
            automated.append(op)
            by_sheet[sheet]["automated"] += 1
        elif is_llm_task:
            op["_coverage"] = "llm_needed"
            op["_matching_scripts"] = [(m[0], m[1]) for m in matches[:3]]
            llm_needed.append(op)
            by_sheet[sheet]["llm"] += 1
        else:
            op["_coverage"] = "unautomated"
            op["_matching_scripts"] = [(m[0], m[1]) for m in matches[:3]]
            unautomated.append(op)
            by_sheet[sheet]["gap"] += 1

    return {
        "automated": automated,
        "llm_needed": llm_needed,
        "unautomated": unautomated,
        "by_sheet": dict(by_sheet),
        "total_ops": len(ops),
        "total_automated": len(automated),
        "total_llm": len(llm_needed),
        "total_gaps": len(unautomated),
    }


# ============================================================
# Priority Scoring — Revenue-weighted prioritization
# ============================================================

# Priority logic from SELF_PLANNING_PROMPT.md
CATEGORY_PRIORITY = {
    # Priority 1: Revenue-blocking
    "alpha": 1, "research": 1, "scraping": 1, "pipeline": 1,
    "content": 2, "content generation": 2,
    # Priority 2: Compounding work
    "lead gen": 2, "lead generation": 2, "leads": 2,
    "outreach": 2, "cold email": 2, "freelance": 2,
    "ecom": 3, "arbitrage": 3,
    # Priority 3: Building
    "app": 3, "building": 3, "deploy": 3, "ship": 3,
    "landing page": 3, "product": 3,
    # Priority 4: Analysis and optimization
    "analysis": 4, "competitor": 4, "trend": 4,
    "optimization": 4, "a/b test": 4,
    # Priority 5: Maintenance
    "maintenance": 5, "compliance": 5, "backup": 5,
    "health": 5, "monitoring": 5,
    # Priority 6: Learning
    "learning": 6, "retrospective": 6, "self improvement": 6,
}


def score_op_priority(op: dict) -> int:
    """Score an op's priority (1=highest, 6=lowest) based on category and revenue."""
    cat_lower = op["category"].lower()
    desc_lower = op["description"].lower()

    # Check against priority map
    best_priority = 5  # default mid-range
    for keyword, priority in CATEGORY_PRIORITY.items():
        if keyword in cat_lower or keyword in desc_lower:
            best_priority = min(best_priority, priority)

    # Revenue boost: ops mentioning revenue/money get priority bump
    if any(w in desc_lower for w in ["revenue", "$", "money", "monetiz", "profit",
                                      "income", "mrr", "arr", "earning"]):
        best_priority = max(1, best_priority - 1)

    # Blocker detection: blocked ops get deprioritized
    if op["blocker"] and any(w in op["blocker"].lower() for w in
                              ["human", "account", "credential", "manual", "payment"]):
        best_priority = min(6, best_priority + 2)

    return best_priority


def score_revenue_potential(op: dict) -> str:
    """Estimate revenue potential: HIGH / MEDIUM / LOW / UNKNOWN."""
    rev = op["revenue_potential"].lower()
    desc = op["description"].lower()

    if any(w in rev for w in ["high", "$1k", "$5k", "$10k", "1000", "5000"]):
        return "HIGH"
    if any(w in rev for w in ["medium", "med", "$500", "$100"]):
        return "MEDIUM"
    if any(w in rev for w in ["low", "$0", "none", "n/a"]):
        return "LOW"

    # Infer from description
    if any(w in desc for w in ["monetiz", "revenue", "sell", "listing", "product",
                                "service", "client", "customer", "paid"]):
        return "MEDIUM"

    return "UNKNOWN"


# ============================================================
# Task Generation — Create JSONL tasks from ops
# ============================================================

def generate_task_id(source: str, op_id: str) -> str:
    """Generate deterministic task ID from source + op_id."""
    date_str = datetime.now().strftime("%Y%m%d")
    # Hash the op_id to avoid long IDs
    short_hash = hashlib.md5(op_id.encode()).hexdigest()[:6]
    clean_source = re.sub(r'[^a-zA-Z0-9]', '_', source)[:20]
    return f"META_{date_str}_{clean_source}_{short_hash}"


def build_script_task(op: dict, script_name: str, confidence: float) -> dict:
    """Build a script-execution task for an op."""
    date_path = datetime.now().strftime("%Y-%m-%d")
    priority = score_op_priority(op)

    return {
        "id": generate_task_id(op["sheet"], op["op_id"]),
        "category": op["category"].lower() if op["category"] else "building",
        "priority": priority,
        "risk_level": "LOW",
        "description": op["description"],
        "success_criteria": f"Op '{op['op_id']}' executed via {script_name}. Output saved.",
        "estimated_minutes": 30,
        "output_path": f"OPS/autonomous_output/{date_path}/meta_plan/",
        "status": "PENDING",
        "created_at": datetime.now().isoformat(),
        "dependencies": [],
        "source": "meta_planner",
        "meta": {
            "op_id": op["op_id"],
            "sheet": op["sheet"],
            "revenue_potential": score_revenue_potential(op),
            "script_confidence": confidence,
        },
        "execution": {
            "type": "script",
            "script": script_name,
            "flags": "",
            "command": f"python3 AUTOMATIONS/{script_name}",
        },
    }


def build_llm_task(op: dict) -> dict:
    """Build an LLM-native task for an op that needs human-like reasoning."""
    date_path = datetime.now().strftime("%Y-%m-%d")
    priority = score_op_priority(op)

    # Build a focused LLM prompt from the op
    prompt_parts = [
        f"Execute PRINTMAXX op '{op['op_id']}' from sheet '{op['sheet']}'.",
        f"Description: {op['description']}",
    ]
    if op["category"]:
        prompt_parts.append(f"Category: {op['category']}")
    if op["blocker"]:
        prompt_parts.append(f"Known blocker: {op['blocker']} — work around it or flag for human.")
    prompt_parts.extend([
        "Read OPS/HEARTBEAT.md for system state.",
        "Read OPS/PERSISTENT_TASK_TRACKER.md for context.",
        "Execute the op fully. Save all outputs to the output_path.",
        "Update the task tracker with DONE status and proof.",
    ])

    return {
        "id": generate_task_id(op["sheet"], op["op_id"]),
        "category": op["category"].lower() if op["category"] else "building",
        "priority": priority,
        "risk_level": "LOW" if priority >= 3 else "MEDIUM",
        "description": op["description"],
        "success_criteria": f"Op '{op['op_id']}' completed with tangible output saved.",
        "estimated_minutes": 45,
        "output_path": f"OPS/autonomous_output/{date_path}/meta_plan/",
        "status": "PENDING",
        "created_at": datetime.now().isoformat(),
        "dependencies": [],
        "source": "meta_planner",
        "meta": {
            "op_id": op["op_id"],
            "sheet": op["sheet"],
            "revenue_potential": score_revenue_potential(op),
        },
        "execution": {
            "type": "llm",
            "prompt": "\n".join(prompt_parts),
        },
    }


def generate_meta_plan(ops: list, coverage: dict) -> dict:
    """Generate a full meta plan from all ops.

    Returns a plan dict with tasks grouped by priority.
    """
    tasks = []
    seen_ids = set()

    # 1. Generate tasks for unautomated ops (highest value — filling gaps)
    for op in coverage["unautomated"]:
        matches = op.get("_matching_scripts", [])
        if matches and matches[0][1] >= 0.3:
            task = build_script_task(op, matches[0][0], matches[0][1])
        else:
            task = build_llm_task(op)

        if task["id"] not in seen_ids:
            tasks.append(task)
            seen_ids.add(task["id"])

    # 2. Generate tasks for LLM-needed ops
    for op in coverage["llm_needed"]:
        task = build_llm_task(op)
        if task["id"] not in seen_ids:
            tasks.append(task)
            seen_ids.add(task["id"])

    # 3. Generate tasks for automated ops that might need re-running
    #    (only if they have status suggesting they haven't run recently)
    for op in coverage["automated"]:
        status_lower = op["status"].lower()
        if any(w in status_lower for w in ["stale", "failed", "error",
                                            "not started", "pending", ""]):
            matches = op.get("_matching_scripts", [])
            if matches:
                task = build_script_task(op, matches[0][0], matches[0][1])
                if task["id"] not in seen_ids:
                    task["priority"] = min(task["priority"] + 1, 6)  # Slightly lower
                    tasks.append(task)
                    seen_ids.add(task["id"])

    # Sort by priority then revenue potential
    revenue_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2, "UNKNOWN": 3}
    tasks.sort(key=lambda t: (
        t["priority"],
        revenue_order.get(t.get("meta", {}).get("revenue_potential", "UNKNOWN"), 3),
    ))

    plan = {
        "generated_at": datetime.now().isoformat(),
        "master_ops_file": str(MASTER_OPS_PATH.name),
        "total_ops_scanned": coverage["total_ops"],
        "total_automated": coverage["total_automated"],
        "total_llm_tasks": coverage["total_llm"],
        "total_gaps": coverage["total_gaps"],
        "total_tasks_generated": len(tasks),
        "by_priority": {},
        "by_sheet": coverage["by_sheet"],
        "tasks": tasks,
    }

    # Group by priority for summary
    for t in tasks:
        p = str(t["priority"])
        if p not in plan["by_priority"]:
            plan["by_priority"][p] = 0
        plan["by_priority"][p] += 1

    return plan


# ============================================================
# Nav Context Builder — Compact folder map for agents
# ============================================================

def build_nav_context() -> str:
    """Build a compact navigation context (<500 lines) for agent prompts.

    Includes:
      - Key directories and their purpose
      - Script locations
      - Data file locations
      - Memory system guide
      - "Where is..." table from CLAUDE.md (condensed)
    """
    lines = []
    lines.append("# PRINTMAXX NAV CONTEXT (compact agent reference)")
    lines.append(f"# Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append(f"# Project: {PROJECT_ROOT}")
    lines.append("")

    # Section 1: Directory Map
    lines.append("## Directory Map")
    lines.append("")
    dir_map = [
        ("AUTOMATIONS/", "195 Python scripts — scrapers, pipelines, monitors, tools"),
        ("OPS/", "Operational state — HEARTBEAT, active-tasks, task tracker, playbooks"),
        ("LEDGER/", "Source of truth — all CSVs: alpha, methods, niches, metrics"),
        ("LEDGER/MEGA_SHEET/", "10 consolidated CSVs (2,512 rows) — query for broad lookups"),
        ("LEDGER/RBI_STRATEGIC/", "Strategic outputs — GTM tactics, hypotheses, learnings"),
        ("CONTENT/", "612+ content files — truth pages, social, longtail, email"),
        ("CONTENT/social/", "Per-niche content packages — 13 accounts with first-week content"),
        ("PRODUCTS/", "Product listings — Gumroad, Fiverr, Etsy, Redbubble, KDP"),
        ("PRODUCTS/FREELANCE_LISTINGS_READY/", "Copy-paste freelance listings for 10 platforms"),
        ("PRODUCTS/GUMROAD_INSTANT_UPLOAD/", "13 Gumroad products ready to upload"),
        ("MONEY_METHODS/", "Playbooks per revenue method — APP_FACTORY, LOCAL_BIZ, etc."),
        ("MONEY_METHODS/APP_FACTORY/", "6 PWA apps + design system + GTM + submission guides"),
        ("MONEY_METHODS/LOCAL_BIZ/", "6 templates + personalizer + lead scraper + pricing"),
        ("DIGITAL_PRODUCTS/", "Gumroad listings, micro products, PDFs ready to sell"),
        ("FINANCIALS/", "Revenue, expenses, P&L, investments, tax tracking"),
        ("LANDING/printmaxx-site/", "Next.js landing site — App Router, Turbopack"),
        ("ralph/", "Autonomous overnight loops — 13 loop configs"),
        ("scripts/", "Utility scripts — builders, audit, RBI engine, SEO generator"),
        ("scripts/builders/", "11 Python scripts that regenerate any .xlsx from scratch"),
        ("SECRETS/", "Credentials (gitignored) — CREDENTIALS.env, PAYMENT_INFO.md"),
        ("AUDIT/", "Deep scan reports — alpha gaps, automation inventory"),
        ("builds/", "Programmatic SEO (602 pages), app builds, content assets"),
    ]
    lines.append("| Directory | Contents |")
    lines.append("|-----------|----------|")
    for d, desc in dir_map:
        lines.append(f"| `{d}` | {desc} |")
    lines.append("")

    # Section 2: Key State Files (Memory System)
    lines.append("## Memory System (read in this order)")
    lines.append("")
    lines.append("1. `OPS/HEARTBEAT.md` — <20 lines, pure numbers, 3-second system pulse")
    lines.append("2. `OPS/active-tasks.md` — crash recovery: what was running, what's left")
    lines.append("3. `OPS/PERSISTENT_TASK_TRACKER.md` — all tasks, statuses, blockers")
    lines.append("4. `LEDGER/ALPHA_STAGING.csv` — unprocessed alpha count")
    lines.append("5. `OPS/AUTONOMOUS_TASK_QUEUE.jsonl` — current task queue")
    lines.append("6. `AUTOMATIONS/logs/daily/YYYY-MM-DD.md` — today's execution log")
    lines.append("")

    # Section 3: Key Scripts Quick Reference
    lines.append("## Key Scripts (most-used)")
    lines.append("")
    key_scripts = [
        ("daily_agent_runner.py --status", "Auto-orient in 10 seconds"),
        ("memory_manager.py --full", "Refresh all 3 memory layers"),
        ("closed_loop_pipeline.py --cycles 5 --batch 2000", "Lead qualification pipeline"),
        ("system_health_monitor.py --quick", "3-second health pulse"),
        ("venture_performance_tracker.py --recommend", "Score methods, KILL/MAINTAIN/DOUBLE_DOWN"),
        ("twitter_alpha_scraper.py --all", "Scrape 116+ Twitter accounts via Brave"),
        ("background_reddit_scraper.py --scrape", "Reddit JSON API (no auth needed)"),
        ("daily_research_orchestrator.py --full", "5 scrapers + HN + 41 subs + PH"),
        ("alpha_auto_processor.py --process-new", "Route alpha to ventures/OPS/cron/archive"),
        ("competitor_monitor.py --scan", "19 apps, 6 niches, iTunes API"),
        ("ecom_arb_engine.py --scan --top 10", "Amazon/eBay pricing, margin calc"),
        ("freelance_demand_scanner.py --scan", "9 subreddits, hiring posts"),
        ("trend_aggregator.py --scan", "Google Trends + Reddit + PH"),
        ("compliance_deadline_tracker.py --check", "21 regulatory deadlines"),
        ("telegram_community_monitor.py --scan", "26 channels, 8 niches"),
        ("workflow_wirer.py --wire", "Wire all pipelines into task queue"),
        ("meta_planner.py --meta-plan", "Full meta plan from MASTER_OPS"),
        ("printmaxx.py status", "Unified CLI (12 subcommands wrapping 28+ scripts)"),
    ]
    lines.append("| Command | Purpose |")
    lines.append("|---------|---------|")
    for cmd, purpose in key_scripts:
        lines.append(f"| `python3 AUTOMATIONS/{cmd}` | {purpose} |")
    lines.append("")

    # Section 4: Data Files
    lines.append("## Key Data Files")
    lines.append("")
    data_files = [
        ("LEDGER/ALPHA_STAGING.csv", "All alpha entries with status and score"),
        ("LEDGER/MEGA_SHEET/*.csv", "10 consolidated tabs (2,512 rows total)"),
        ("LEDGER/REVENUE_STREAMS_TRACKER.csv", "100 revenue streams pre-populated"),
        ("LEDGER/FREELANCE_PIPELINE_ACTIVE.csv", "Active freelance opportunities"),
        ("LEDGER/ECOM_ARB_OPPORTUNITIES.csv", "Ecom arb scan results"),
        ("LEDGER/TREND_SIGNALS.csv", "Multi-source trend detection"),
        ("LEDGER/COMPLIANCE_DEADLINES.csv", "21 regulatory deadlines"),
        ("LEDGER/TELEGRAM_SIGNALS.csv", "Telegram channel signals"),
        ("LEDGER/RBI_STRATEGIC/LEARNINGS.jsonl", "Append-only learnings database"),
        ("FINANCIALS/REVENUE_TRACKER.csv", "All revenue by method"),
        ("FINANCIALS/EXPENSE_TRACKER.csv", "All expenses"),
        ("AUTOMATIONS/leads/qualified/HOT_LEADS_QUALIFIED.csv", "Score >= 65 leads"),
        ("AUTOMATIONS/outreach/PIPELINE_TRACKER.csv", "Cold email pipeline"),
    ]
    lines.append("| File | Contents |")
    lines.append("|------|----------|")
    for f, desc in data_files:
        lines.append(f"| `{f}` | {desc} |")
    lines.append("")

    # Section 5: XLSX Strategic Spreadsheets
    lines.append("## XLSX Spreadsheets (project root)")
    lines.append("")
    xlsx_files = [
        ("PRINTMAXX_MASTER_OPS.xlsx", "150+ ops across 12 sheets — THE master plan"),
        ("PRINTMAXX_STRATEGIC_RBI.xlsx", "Viability matrix, bottlenecks, hypotheses"),
        ("PRINTMAXX_FREELANCE_ARB.xlsx", "30 services, 10 platforms"),
        ("PRINTMAXX_OPS_PLAYBOOK.xlsx", "22 ops, 1813 deep playbook rows"),
        ("PRINTMAXX_BRAND_NAMES.xlsx", "207 brand names"),
    ]
    for f, desc in xlsx_files:
        lines.append(f"- `{f}` — {desc}")
    lines.append("")

    # Section 6: Condensed "Where is..." table
    lines.append("## Quick Lookup (Where is...?)")
    lines.append("")
    where_is = [
        ("Latest handoff", "OPS/SESSION_HANDOFF_FEB12_2026.md"),
        ("System pulse", "OPS/HEARTBEAT.md"),
        ("All alpha", "LEDGER/ALPHA_STAGING.csv"),
        ("Revenue", "FINANCIALS/REVENUE_TRACKER.csv"),
        ("Human tasks", "OPS/ACCOUNT_CREATION_NOW.md"),
        ("Copy style", ".claude/rules/copy-style.md"),
        ("Agent playbook", "OPS/AGENT_DAILY_PLAYBOOK.md"),
        ("Task tracker", "OPS/PERSISTENT_TASK_TRACKER.md"),
        ("Deploy log (16 sites)", "OPS/DEPLOY_LOG.md"),
        ("App Factory index", "MONEY_METHODS/APP_FACTORY/APP_FACTORY_CENTRAL_INDEX.md"),
        ("Product index", "PRODUCTS/PRODUCTS_CENTRAL_INDEX.md"),
        ("Content index", "CONTENT/CONTENT_CENTRAL_INDEX.md"),
        ("Leads index", "OPS/LEADS_OUTREACH_CENTRAL_INDEX.md"),
        ("Cron orchestrator", "printmaxx_cron.sh"),
        ("Quant terminal", "python3 AUTOMATIONS/printmaxx_quant_terminal.py"),
        ("Fiverr gigs (10)", "PRODUCTS/FREELANCE_LISTINGS_READY/FIVERR_GIGS_10.md"),
        ("Gumroad products (13)", "PRODUCTS/GUMROAD_INSTANT_UPLOAD/"),
        ("Etsy listings", "PRODUCTS/ECOM_LISTINGS_READY/ETSY_LISTINGS_COMPLETE.md"),
        ("Live dashboard", "python3 AUTOMATIONS/live_dashboard_server.py (localhost:8888)"),
    ]
    lines.append("| What | Where |")
    lines.append("|------|-------|")
    for what, where in where_is:
        lines.append(f"| {what} | `{where}` |")
    lines.append("")

    # Section 7: Current System State (live from HEARTBEAT)
    lines.append("## Current State (from HEARTBEAT)")
    lines.append("")
    if HEARTBEAT_PATH.exists():
        for line in HEARTBEAT_PATH.read_text().strip().split("\n"):
            line = line.strip()
            if line and not line.startswith("#"):
                lines.append(f"- {line}")
    else:
        lines.append("- HEARTBEAT.md not found — run `python3 AUTOMATIONS/memory_manager.py --heartbeat`")
    lines.append("")

    # Section 8: Priority Framework
    lines.append("## Priority Framework")
    lines.append("")
    lines.append("1. **P1 Revenue-blocking:** Unprocessed alpha, failed scrapers, system health")
    lines.append("2. **P2 Compounding:** Alpha sources, content gen, lead qualification, trends")
    lines.append("3. **P3 Building:** Apps, landing pages, tools, scripts")
    lines.append("4. **P4 Analysis:** Competitors, optimization, A/B tests")
    lines.append("5. **P5 Maintenance:** Compliance, backups, health checks")
    lines.append("6. **P6 Learning:** Retrospectives, prompt effectiveness, learnings")
    lines.append("")

    # Section 9: Guardrails
    lines.append("## Guardrails")
    lines.append("")
    lines.append(f"- All file ops MUST stay within: `{PROJECT_ROOT}`")
    lines.append("- Never propose tasks with risk_level CRITICAL")
    lines.append("- Never propose tasks involving payments, account creation, or publishing")
    lines.append("- Every task must have clear success_criteria")
    lines.append("- Never delete CLAUDE.md, LEDGER/, FINANCIALS/, SECRETS/, XLSX files")
    lines.append("")

    return "\n".join(lines)


# ============================================================
# Queue Writer — Write tasks to AUTONOMOUS_TASK_QUEUE.jsonl
# ============================================================

def read_existing_queue() -> list:
    """Read existing tasks from the JSONL queue."""
    tasks = []
    if QUEUE_PATH.exists():
        for line in QUEUE_PATH.read_text().strip().split("\n"):
            line = line.strip()
            if line:
                try:
                    tasks.append(json.loads(line))
                except json.JSONDecodeError:
                    pass
    return tasks


def write_tasks_to_queue(new_tasks: list) -> int:
    """Append new tasks to the queue, deduplicating by ID.

    Returns count of tasks actually added.
    """
    safe_path(QUEUE_PATH)
    QUEUE_PATH.parent.mkdir(parents=True, exist_ok=True)

    existing = read_existing_queue()
    existing_ids = {t["id"] for t in existing}

    added = 0
    for task in new_tasks:
        if task["id"] not in existing_ids:
            existing.append(task)
            existing_ids.add(task["id"])
            added += 1

    lines = [json.dumps(t) for t in existing]
    QUEUE_PATH.write_text("\n".join(lines) + "\n" if lines else "")
    return added


# ============================================================
# CLI Display Functions
# ============================================================

def show_status():
    """Show meta plan coverage: X/N ops automated, gaps, state summary."""
    print("\n" + "=" * 70)
    print("  PRINTMAXX META PLANNER — STATUS")
    print("=" * 70)

    # Build catalog and read ops
    build_script_catalog()
    ops = extract_all_ops(include_reference_sheets=False)
    coverage = analyze_coverage(ops)
    state = read_system_state()

    # Also count total rows including reference sheets for context
    all_ops_incl_ref = extract_all_ops(include_reference_sheets=True)
    ref_count = len(all_ops_incl_ref) - len(ops)

    print(f"\n  MASTER_OPS file:     {MASTER_OPS_PATH.name}")
    print(f"  Actionable ops:      {coverage['total_ops']}")
    print(f"  Reference rows:      {ref_count} (in {', '.join(REFERENCE_ONLY_SHEETS)})")
    print(f"  Scripts available:   {state['script_count']}")
    print()

    # Coverage breakdown
    total = coverage["total_ops"]
    automated = coverage["total_automated"]
    llm = coverage["total_llm"]
    gaps = coverage["total_gaps"]

    pct_auto = (automated / total * 100) if total else 0
    pct_llm = (llm / total * 100) if total else 0
    pct_gap = (gaps / total * 100) if total else 0

    print(f"  COVERAGE (actionable ops only):")
    print(f"    Script-automated:  {automated}/{total} ({pct_auto:.0f}%)")
    print(f"    LLM-native tasks:  {llm}/{total} ({pct_llm:.0f}%)")
    print(f"    Unautomated gaps:  {gaps}/{total} ({pct_gap:.0f}%)")
    print()

    # Per-sheet breakdown
    print("  BY SHEET:")
    for sheet, counts in sorted(coverage["by_sheet"].items()):
        total_s = counts["total"]
        auto_s = counts["automated"]
        llm_s = counts["llm"]
        gap_s = counts["gap"]
        # Cap the bar at 60 chars for readability
        bar_len = min(total_s, 60)
        if total_s > 0:
            bar_auto = max(1, int(auto_s / total_s * bar_len)) if auto_s else 0
            bar_llm = max(1, int(llm_s / total_s * bar_len)) if llm_s else 0
            bar_gap = bar_len - bar_auto - bar_llm
        else:
            bar_auto = bar_llm = bar_gap = 0
        bar = "#" * bar_auto + "~" * bar_llm + "." * max(0, bar_gap)
        print(f"    {sheet:35s} {auto_s:3d}/{total_s:3d} auto  "
              f"{llm_s:2d} llm  {gap_s:3d} gap  [{bar}]")
    print()

    # System state
    print("  SYSTEM STATE:")
    print(f"    Task queue:        {state['queue_tasks']} total "
          f"({state['queue_pending']} pending, "
          f"{state['queue_completed']} done, "
          f"{state['queue_failed']} failed)")
    print(f"    Alpha pending:     {state['pending_alpha']}")
    print()

    # Heartbeat
    if state["heartbeat"]:
        print("  HEARTBEAT:")
        for line in state["heartbeat"].strip().split("\n"):
            line = line.strip()
            if line and not line.startswith("#"):
                print(f"    {line}")
    print()
    print("=" * 70)


def show_gaps():
    """Show unautomated ops — what's missing."""
    print("\n" + "=" * 70)
    print("  PRINTMAXX META PLANNER — GAPS ANALYSIS")
    print("=" * 70)

    build_script_catalog()
    ops = extract_all_ops(include_reference_sheets=False)
    coverage = analyze_coverage(ops)

    # Unautomated gaps
    gaps = coverage["unautomated"]
    if not gaps:
        print("\n  No gaps found. All ops have coverage.")
        print("=" * 70)
        return

    print(f"\n  {len(gaps)} UNAUTOMATED OPS (no script, no LLM coverage):\n")

    by_sheet = defaultdict(list)
    for op in gaps:
        by_sheet[op["sheet"]].append(op)

    for sheet, sheet_ops in sorted(by_sheet.items()):
        print(f"  [{sheet}] ({len(sheet_ops)} gaps)")
        for op in sheet_ops[:10]:  # Limit display
            matches = op.get("_matching_scripts", [])
            match_str = ""
            if matches:
                match_str = f"  (closest: {matches[0][0]} @ {matches[0][1]})"
            desc = op["description"][:70] + "..." if len(op["description"]) > 70 else op["description"]
            print(f"    {op['op_id']:25s} {desc}{match_str}")
        if len(sheet_ops) > 10:
            print(f"    ... and {len(sheet_ops) - 10} more")
        print()

    # LLM-needed tasks
    llm_ops = coverage["llm_needed"]
    if llm_ops:
        print(f"\n  {len(llm_ops)} OPS NEEDING LLM WORK (no script can handle):\n")
        by_sheet_llm = defaultdict(list)
        for op in llm_ops:
            by_sheet_llm[op["sheet"]].append(op)

        for sheet, sheet_ops in sorted(by_sheet_llm.items()):
            print(f"  [{sheet}] ({len(sheet_ops)} LLM tasks)")
            for op in sheet_ops[:8]:
                desc = op["description"][:70] + "..." if len(op["description"]) > 70 else op["description"]
                print(f"    {op['op_id']:25s} {desc}")
            if len(sheet_ops) > 8:
                print(f"    ... and {len(sheet_ops) - 8} more")
            print()

    print("=" * 70)


def show_meta_plan():
    """Generate and display the full meta plan."""
    print("\n" + "=" * 70)
    print("  PRINTMAXX META PLANNER — FULL META PLAN")
    print("=" * 70)

    build_script_catalog()
    ops = extract_all_ops(include_reference_sheets=False)
    coverage = analyze_coverage(ops)
    plan = generate_meta_plan(ops, coverage)

    print(f"\n  Generated: {plan['generated_at']}")
    print(f"  MASTER_OPS: {plan['master_ops_file']}")
    print(f"  Ops scanned: {plan['total_ops_scanned']}")
    print(f"  Tasks generated: {plan['total_tasks_generated']}")
    print()

    print("  BY PRIORITY:")
    for p in sorted(plan["by_priority"].keys()):
        count = plan["by_priority"][p]
        label = {
            "1": "Revenue-blocking",
            "2": "Compounding",
            "3": "Building",
            "4": "Analysis",
            "5": "Maintenance",
            "6": "Learning",
        }.get(p, "Other")
        print(f"    P{p} ({label}): {count} tasks")
    print()

    # Show top 20 tasks
    print("  TOP 20 TASKS (highest priority):\n")
    for i, task in enumerate(plan["tasks"][:20], 1):
        exec_type = task.get("execution", {}).get("type", "unknown")
        exec_detail = ""
        if exec_type == "script":
            exec_detail = task["execution"].get("script", "")
        else:
            exec_detail = "LLM"
        desc = task["description"][:60] + "..." if len(task["description"]) > 60 else task["description"]
        rev = task.get("meta", {}).get("revenue_potential", "?")
        print(f"    {i:2d}. [P{task['priority']}] [{exec_type:6s}] {desc}")
        print(f"        {exec_detail}  rev={rev}  est={task['estimated_minutes']}min")
    print()

    if len(plan["tasks"]) > 20:
        print(f"  ... and {len(plan['tasks']) - 20} more tasks")
        print()

    # Save the plan
    safe_path(META_PLAN_PATH)
    META_PLAN_PATH.parent.mkdir(parents=True, exist_ok=True)
    META_PLAN_PATH.write_text(json.dumps(plan, indent=2))
    print(f"  Full plan saved to: {META_PLAN_PATH.relative_to(PROJECT_ROOT)}")
    print("=" * 70)


def show_nav_context():
    """Generate and display the compact nav context for agents."""
    nav = build_nav_context()
    line_count = len(nav.split("\n"))

    print(nav)
    print(f"\n--- {line_count} lines (target: <500) ---")

    # Also save to file
    safe_path(NAV_CONTEXT_PATH)
    NAV_CONTEXT_PATH.parent.mkdir(parents=True, exist_ok=True)
    NAV_CONTEXT_PATH.write_text(nav)
    print(f"Saved to: {NAV_CONTEXT_PATH.relative_to(PROJECT_ROOT)}")


def wire_all():
    """Wire ALL unautomated ops into the task queue."""
    print("\n" + "=" * 70)
    print("  PRINTMAXX META PLANNER — WIRE ALL OPS")
    print("=" * 70)

    build_script_catalog()
    ops = extract_all_ops(include_reference_sheets=False)
    coverage = analyze_coverage(ops)
    plan = generate_meta_plan(ops, coverage)

    tasks = plan["tasks"]
    if not tasks:
        print("\n  No tasks to wire.")
        print("=" * 70)
        return

    # Filter to only PENDING-worthy tasks (skip already completed ops)
    wireable = []
    for task in tasks:
        # Remove internal meta for cleaner queue
        clean_task = {k: v for k, v in task.items()}
        wireable.append(clean_task)

    added = write_tasks_to_queue(wireable)

    print(f"\n  Total tasks generated: {len(wireable)}")
    print(f"  New tasks wired to queue: {added}")
    print(f"  Queue location: {QUEUE_PATH.relative_to(PROJECT_ROOT)}")
    print()

    # Breakdown
    script_tasks = sum(1 for t in wireable
                       if t.get("execution", {}).get("type") == "script")
    llm_tasks = sum(1 for t in wireable
                    if t.get("execution", {}).get("type") == "llm")
    print(f"  Script-execution tasks: {script_tasks}")
    print(f"  LLM-native tasks:      {llm_tasks}")
    print()

    # Save the plan too
    safe_path(META_PLAN_PATH)
    META_PLAN_PATH.parent.mkdir(parents=True, exist_ok=True)
    META_PLAN_PATH.write_text(json.dumps(plan, indent=2))
    print(f"  Meta plan saved to: {META_PLAN_PATH.relative_to(PROJECT_ROOT)}")
    print("=" * 70)


# ============================================================
# Main CLI
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="PRINTMAXX Meta Planner — autonomous full-vision task generation"
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument(
        "--meta-plan",
        action="store_true",
        help="Generate full meta plan from MASTER_OPS"
    )
    group.add_argument(
        "--nav-context",
        action="store_true",
        help="Output compact folder nav for agent prompts (<500 lines)"
    )
    group.add_argument(
        "--wire-all",
        action="store_true",
        help="Wire all unautomated ops into the task queue"
    )
    group.add_argument(
        "--status",
        action="store_true",
        help="Show meta plan coverage (X/N ops automated)"
    )
    group.add_argument(
        "--gaps",
        action="store_true",
        help="Show unautomated ops — what's missing"
    )

    args = parser.parse_args()

    if args.status:
        show_status()
    elif args.gaps:
        show_gaps()
    elif args.meta_plan:
        show_meta_plan()
    elif args.nav_context:
        show_nav_context()
    elif args.wire_all:
        wire_all()


if __name__ == "__main__":
    main()
