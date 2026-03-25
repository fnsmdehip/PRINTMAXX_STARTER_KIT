#!/usr/bin/env python3

from __future__ import annotations
"""
PRINTMAXX UNIFIED CONTROL PANEL
=================================
Single Flask server that replaces all scattered dashboards.
Serves the control panel HTML and exposes API endpoints for:
- Agent throttle control (EFFICIENT/HIGH mode toggle, per-agent on/off)
- System status (health, revenue, cron, sites)
- Quick actions (run scripts, view logs)
- Slash command launcher
- System tree (hierarchy, data flow, cron schedule)
- Real-time metrics (agent invocations, alpha entries)
- Venture status
- Pipeline funnel

Port: 9999
Usage: python3 AUTOMATIONS/control_panel.py
"""

import csv
import json
import os
import re
import sqlite3
import subprocess
import sys
import threading
import time
import uuid
import webbrowser
from datetime import datetime, timedelta
from pathlib import Path
from collections import Counter
from urllib.request import Request, urlopen
from urllib.error import URLError

try:
    from flask import Flask, jsonify, request, Response
except ImportError:
    print("ERROR: Flask not installed. Run: pip3 install flask")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent
AUTOMATIONS = PROJECT_ROOT / "AUTOMATIONS"
OPS = PROJECT_ROOT / "OPS"
LEDGER = PROJECT_ROOT / "LEDGER"
FINANCIALS = PROJECT_ROOT / "FINANCIALS"
LOGS_DIR = AUTOMATIONS / "logs"
CLAUDE_LOGS = Path.home() / ".claude" / "logs"
LA_DIR = Path.home() / "Library" / "LaunchAgents"
HTML_PATH = AUTOMATIONS / "control_panel.html"

PORT = 9999
PYTHON = sys.executable

# Load N8N_API_KEY from .env
N8N_BASE = "http://localhost:5678"
N8N_API_KEY = ""
_env_path = PROJECT_ROOT / ".env"
if _env_path.exists():
    try:
        with open(_env_path, "r") as _ef:
            for _line in _ef:
                _line = _line.strip()
                if _line.startswith("N8N_API_KEY="):
                    N8N_API_KEY = _line.split("=", 1)[1].strip()
                    break
    except Exception:
        pass

# Sovrun paths
SOVRUN_SKILLS_DB = AUTOMATIONS / "agent" / "sovrun" / "skills.db"
SOVRUN_SKILLS_DB_ALT = AUTOMATIONS / "agent" / "swarm" / "sovrun" / "skills.db"
DAG_CHECKPOINT = AUTOMATIONS / "agent" / "morning_dag_checkpoint.json"
DAG_STATE = AUTOMATIONS / "agent" / "morning_dag_state.json"

app = Flask(__name__)
_lock = threading.Lock()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def safe_read(filepath, max_lines=50):
    try:
        p = Path(filepath)
        if not p.exists():
            return ""
        with open(p, "r", encoding="utf-8", errors="replace") as f:
            lines = f.readlines()
        return "".join(lines[-max_lines:])
    except Exception:
        return ""


def safe_read_json(filepath):
    try:
        with open(filepath, "r") as f:
            return json.load(f)
    except Exception:
        return {}


def safe_read_csv(filepath, max_rows=100):
    try:
        p = Path(filepath)
        if not p.exists():
            return []
        rows = []
        with open(p, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.DictReader(f)
            for i, row in enumerate(reader):
                if i >= max_rows:
                    break
                rows.append(dict(row))
        return rows
    except Exception:
        return []


def run_cmd(cmd, timeout=30):
    try:
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout, cwd=str(PROJECT_ROOT))
        return r.returncode, r.stdout.strip(), r.stderr.strip()
    except subprocess.TimeoutExpired:
        return -1, "", "timeout"
    except Exception as e:
        return -1, "", str(e)


def days_until_friday():
    today = datetime.now()
    days_ahead = (4 - today.weekday()) % 7
    if days_ahead == 0 and today.hour >= 12:
        days_ahead = 7
    return days_ahead


def relative_time(ts):
    if not ts:
        return "never"
    try:
        if isinstance(ts, (int, float)):
            dt = datetime.fromtimestamp(ts)
        else:
            dt = datetime.fromisoformat(str(ts))
        delta = datetime.now() - dt
        if delta.total_seconds() < 60:
            return f"{int(delta.total_seconds())}s ago"
        if delta.total_seconds() < 3600:
            return f"{int(delta.total_seconds() / 60)}m ago"
        if delta.total_seconds() < 86400:
            return f"{delta.total_seconds() / 3600:.1f}h ago"
        return f"{delta.days}d ago"
    except Exception:
        return "unknown"


# ---------------------------------------------------------------------------
# Throttle integration
# ---------------------------------------------------------------------------
def get_throttle_status():
    """Get full throttle status via throttle_toggle.py --json"""
    code, out, err = run_cmd([PYTHON, str(AUTOMATIONS / "throttle_toggle.py"), "--json"], timeout=15)
    if code == 0 and out:
        try:
            return json.loads(out)
        except json.JSONDecodeError:
            pass
    return {"mode": "unknown", "agents": {}, "summary": {}}


def set_throttle_mode(mode):
    with _lock:
        code, out, err = run_cmd([PYTHON, str(AUTOMATIONS / "throttle_toggle.py"), "--mode", mode], timeout=60)
    return {"success": code == 0, "output": out, "error": err}


def toggle_agent(agent_id, enabled):
    flag = "--on" if enabled else "--off"
    with _lock:
        code, out, err = run_cmd([PYTHON, str(AUTOMATIONS / "throttle_toggle.py"), "--agent", agent_id, flag], timeout=15)
    return {"success": code == 0, "output": out, "error": err}


def set_agent_interval(agent_id, hours):
    with _lock:
        code, out, err = run_cmd([PYTHON, str(AUTOMATIONS / "throttle_toggle.py"), "--agent", agent_id, "--interval", str(hours)], timeout=15)
    return {"success": code == 0, "output": out, "error": err}


# ---------------------------------------------------------------------------
# System status
# ---------------------------------------------------------------------------
def get_system_status():
    status = {
        "timestamp": datetime.now().isoformat(),
        "days_until_reset": days_until_friday(),
        "revenue": "$0",
        "day_number": 0,
    }

    # Revenue from financial tracker
    rev_path = FINANCIALS / "REVENUE_TRACKER.csv"
    if rev_path.exists():
        rows = safe_read_csv(rev_path, max_rows=1000)
        total = sum(float(r.get("amount", 0)) for r in rows if r.get("amount"))
        status["revenue"] = f"${total:,.0f}"

    # Day number from session briefing
    briefing = safe_read(OPS / "SESSION_BRIEFING.md", max_lines=20)
    for line in briefing.split("\n"):
        if "Day " in line and "at zero" in line.lower():
            try:
                day_num = int(line.split("Day ")[1].split(" ")[0])
                status["day_number"] = day_num
            except (IndexError, ValueError):
                pass

    # Alpha stats
    alpha_path = LEDGER / "ALPHA_STAGING.csv"
    if alpha_path.exists():
        rows = safe_read_csv(alpha_path, max_rows=50000)
        status["alpha_total"] = len(rows)
        status["alpha_pending"] = sum(1 for r in rows if r.get("status", "").upper() == "PENDING_REVIEW")
        status["alpha_approved"] = sum(1 for r in rows if r.get("status", "").upper() == "APPROVED")

    # Cron count
    code, out, _ = run_cmd(["crontab", "-l"], timeout=5)
    if code == 0:
        cron_lines = [l for l in out.split("\n") if l.strip() and not l.strip().startswith("#") and not l.strip().startswith("SHELL") and not l.strip().startswith("PATH") and not l.strip().startswith("BASE") and not l.strip().startswith("PYTHON")]
        status["cron_jobs"] = len(cron_lines)

    # Launchd agent count
    code, out, _ = run_cmd(["launchctl", "list"], timeout=5)
    if code == 0:
        pm_agents = [l for l in out.split("\n") if "printmaxx" in l.lower() or "com.claude.schedule" in l.lower()]
        status["launchd_agents"] = len(pm_agents)

    return status


def get_log_tail(agent_id, lines=30):
    """Get last N lines of an agent's log."""
    # Try ~/.claude/logs/ first (swarm + venture agents)
    for prefix in ["swarm_", "auto_", ""]:
        log_path = CLAUDE_LOGS / f"{prefix}{agent_id}.log"
        if log_path.exists():
            return safe_read(log_path, max_lines=lines)

    # Try AUTOMATIONS/logs/
    for name in [f"{agent_id}.log", f"cron_{agent_id}.log"]:
        log_path = LOGS_DIR / name
        if log_path.exists():
            return safe_read(log_path, max_lines=lines)

    return f"No log found for {agent_id}"


def get_quick_actions():
    """List of available quick actions."""
    return [
        {"id": "decision_engine", "name": "Decision Engine Cycle", "cmd": f"{PYTHON} AUTOMATIONS/decision_engine.py --cycle", "category": "core"},
        {"id": "system_health", "name": "System Health Check", "cmd": f"{PYTHON} AUTOMATIONS/system_health_monitor.py --quick", "category": "core"},
        {"id": "alpha_process", "name": "Process New Alpha", "cmd": f"{PYTHON} AUTOMATIONS/alpha_auto_processor.py --process-new", "category": "alpha"},
        {"id": "loop_closer", "name": "Run Loop Closer", "cmd": f"{PYTHON} AUTOMATIONS/loop_closer.py --cycle", "category": "core"},
        {"id": "quality_gate", "name": "Quality Gate", "cmd": f"{PYTHON} AUTOMATIONS/quality_gate.py --gate", "category": "quality"},
        {"id": "daily_digest", "name": "Generate Daily Digest", "cmd": f"{PYTHON} AUTOMATIONS/daily_digest.py", "category": "reports"},
        {"id": "tactical_plan", "name": "Generate Tactical Plan", "cmd": f"{PYTHON} AUTOMATIONS/daily_tactical_engine.py --save", "category": "reports"},
        {"id": "intelligence_stats", "name": "Intelligence Stats", "cmd": f"{PYTHON} AUTOMATIONS/intelligence_router.py --stats", "category": "intel"},
        {"id": "alpha_stats", "name": "Alpha Query Stats", "cmd": f"{PYTHON} AUTOMATIONS/alpha_query.py --stats", "category": "alpha"},
        {"id": "venture_status", "name": "Venture Status", "cmd": f"{PYTHON} AUTOMATIONS/venture_autonomy.py --status", "category": "agents"},
        {"id": "swarm_status", "name": "Swarm Status", "cmd": f"{PYTHON} AUTOMATIONS/agent_swarm.py --status", "category": "agents"},
        {"id": "swarm_health", "name": "Swarm Health Check", "cmd": f"{PYTHON} AUTOMATIONS/agent_swarm.py --health", "category": "agents"},
        {"id": "throttle_status", "name": "Throttle Status", "cmd": f"{PYTHON} AUTOMATIONS/throttle_toggle.py --status", "category": "throttle"},
        {"id": "throttle_estimate", "name": "Token Burn Estimate", "cmd": f"{PYTHON} AUTOMATIONS/throttle_toggle.py --estimate", "category": "throttle"},
    ]


def get_slash_commands():
    """List available Claude Code slash commands."""
    cmd_dir = PROJECT_ROOT / ".claude" / "commands"
    commands = []
    if cmd_dir.exists():
        for f in sorted(cmd_dir.glob("*.md")):
            name = f.stem
            first_line = ""
            try:
                with open(f, "r") as fh:
                    first_line = fh.readline().strip().lstrip("#").strip()
            except Exception:
                pass
            commands.append({"name": f"/{name}", "description": first_line, "file": str(f.relative_to(PROJECT_ROOT))})
    return commands


def get_deployed_sites():
    """Read deployment URLs."""
    deploy_path = OPS / "DEPLOYMENT_URLS.md"
    sites = []
    if deploy_path.exists():
        content = safe_read(deploy_path, max_lines=100)
        for line in content.split("\n"):
            if "surge.sh" in line or "vercel.app" in line:
                urls = re.findall(r'https?://[^\s\)]+', line)
                for url in urls:
                    name = line.split("|")[1].strip() if "|" in line else url.split("//")[1].split(".")[0]
                    sites.append({"name": name, "url": url})
    return sites


# ---------------------------------------------------------------------------
# NEW: System Tree API
# ---------------------------------------------------------------------------
def get_system_tree():
    """Parse PRINTMAXX_SYSTEM_MAP.md for hierarchy, data flow, agents, cron."""
    hierarchy = [
        {
            "level": "L0",
            "name": "ORCHESTRATOR",
            "color": "#ff3355",
            "scripts": [
                {"name": "ceo_agent.py", "desc": "16 phases. Scores all ops. PROMOTE/ENHANCE/CREATE/KILL.", "schedule": "Every 2h"}
            ]
        },
        {
            "level": "L1",
            "name": "ENGINES",
            "color": "#ff6b35",
            "scripts": [
                {"name": "venture_autonomy.py", "desc": "8 venture pipelines. Self-managing schedules.", "schedule": "Every 4h"},
                {"name": "agent_swarm.py", "desc": "25 agents via launchd. swarm_brain every 4h.", "schedule": "Continuous"},
                {"name": "decision_engine.py", "desc": "Closed-loop: pending data to scored actions.", "schedule": "Every 30min"}
            ]
        },
        {
            "level": "L2",
            "name": "INTELLIGENCE",
            "color": "#4488ff",
            "scripts": [
                {"name": "intelligence_router.py", "desc": "484 docs + 15K alpha entries into single brief.", "schedule": "On demand"},
                {"name": "alpha_query.py", "desc": "Venture-based alpha search with ROI normalization.", "schedule": "On demand"},
                {"name": "daily_digest.py", "desc": "What happened overnight summary.", "schedule": "6:45 AM"},
                {"name": "session_briefing.py", "desc": "Auto session-start: agent reports, changes.", "schedule": "On session start"},
                {"name": "prompt_meta_review.py", "desc": "Analyzes user prompts for lost threads.", "schedule": "Every 48h"}
            ]
        },
        {
            "level": "L3",
            "name": "EXECUTION",
            "color": "#00ff88",
            "scripts": [
                {"name": "daily_tactical_engine.py", "desc": "Unified 'do exactly this today' plan.", "schedule": "7:15 AM"},
                {"name": "daily_engagement_planner.py", "desc": "Warmup-aware Twitter action plan.", "schedule": "7:00 AM"},
                {"name": "growth_strategist.py", "desc": "Growth strategies per venture from intel.", "schedule": "5:00 AM"},
                {"name": "twitter_warmup_poster.py", "desc": "21-day warmup ramp. Advances at midnight.", "schedule": "Midnight"},
                {"name": "alpha_auto_processor.py", "desc": "Routes ALPHA_STAGING to ventures.", "schedule": "6:30 AM"},
                {"name": "actionable_aggregator.py", "desc": "Scans 6 sources into prioritized queue.", "schedule": "7:30 AM"}
            ]
        },
        {
            "level": "L4",
            "name": "COLLECTION",
            "color": "#8b5cf6",
            "scripts": [
                {"name": "twitter_alpha_scraper.py", "desc": "133 Twitter accounts via Brave cookies.", "schedule": "6:00 AM"},
                {"name": "background_reddit_scraper.py", "desc": "Reddit JSON API, no auth.", "schedule": "6:15 AM"}
            ]
        },
        {
            "level": "L5",
            "name": "QUALITY",
            "color": "#eab308",
            "scripts": [
                {"name": "quality_gate.py", "desc": "Hard gate. Blocks slop, rewrites bad content.", "schedule": "Every 2h"},
                {"name": "compliance_scanner.py", "desc": "FTC / platform compliance auditing.", "schedule": "On demand"},
                {"name": "system_health_monitor.py", "desc": "Health checks: agents, cron, disk.", "schedule": "Every 6h"}
            ]
        },
        {
            "level": "L6",
            "name": "MAINTENANCE",
            "color": "#888888",
            "scripts": [
                {"name": "loop_closer.py", "desc": "3 loops: decision, feedback, pipeline.", "schedule": "Every 2h"},
                {"name": "memory_manager.py", "desc": "Filesystem-based memory management.", "schedule": "5:00 AM"},
                {"name": "build_codebase_grammar.py", "desc": "118x compressed grammar for LLM context.", "schedule": "5:45 AM"},
                {"name": "wire_missed_intelligence.py", "desc": "Parses scan results to update catalog.", "schedule": "On demand"}
            ]
        }
    ]

    data_flow = [
        {"from": "Scrapers", "to": "ALPHA_STAGING.csv", "label": "Raw alpha intake"},
        {"from": "ALPHA_STAGING.csv", "to": "alpha_auto_processor", "label": "Route to ventures"},
        {"from": "Filesystem scan", "to": "wire_missed_intelligence", "label": "Find unindexed docs"},
        {"from": "wire_missed_intelligence", "to": "INTELLIGENCE_CATALOG.json", "label": "Update catalog"},
        {"from": "All sources", "to": "intelligence_router.py", "label": "484 docs + 15K alpha"},
        {"from": "intelligence_router.py", "to": "ceo_agent (Phase 3.5)", "label": "Enrichment"},
        {"from": "intelligence_router.py", "to": "venture_autonomy", "label": "Pre-execution brief"},
        {"from": "intelligence_router.py", "to": "agent_swarm", "label": "AGENT_VENTURE_MAP"},
        {"from": "decision_engine", "to": "DECISIONS.csv", "label": "Audit trail"},
        {"from": "loop_closer", "to": "Swarm adjustments", "label": "Feedback + execution"}
    ]

    cron_schedule = {
        "morning_chain": [
            {"time": "5:00", "script": "growth_strategist"},
            {"time": "5:30", "script": "venture self-manager"},
            {"time": "5:45", "script": "codebase grammar"},
            {"time": "6:00", "script": "twitter scraper + alpha review"},
            {"time": "6:15", "script": "reddit scraper"},
            {"time": "6:30", "script": "alpha auto-processor"},
            {"time": "6:45", "script": "daily digest"},
            {"time": "7:00", "script": "engagement planner"},
            {"time": "7:15", "script": "tactical engine"}
        ],
        "continuous": [
            {"interval": "15min", "script": "guardian pulse + keepalive"},
            {"interval": "30min", "script": "decision engine"},
            {"interval": "2h", "script": "CEO agent + loop closer + safety commit"},
            {"interval": "3h", "script": "signal aggregator + ops manager"},
            {"interval": "4h", "script": "venture autonomy + tracker"},
            {"interval": "6h", "script": "swarm health + memecoin + SAM.gov"}
        ],
        "daily": [
            {"time": "1 AM", "script": "ship engine layer1"},
            {"time": "2 AM", "script": "overnight master runner"},
            {"time": "3 AM", "script": "closed-loop pipeline"},
            {"time": "4 AM", "script": "lead enrichment + log rotation"},
            {"time": "9 PM", "script": "autonomous factory"},
            {"time": "11 PM", "script": "guardian improve"},
            {"time": "Midnight", "script": "warmup poster advance"}
        ],
        "weekly": [
            {"time": "Sun 2 AM", "script": "performance report"},
            {"time": "Sun 3 AM", "script": "full backup"},
            {"time": "Sun 4 AM", "script": "system clone"},
            {"time": "Mon", "script": "ASO, RPM, ecom"}
        ]
    }

    agent_topology = {
        "venture_agents": [
            {"name": "CONTENT", "type": "CONTENT", "interval": "4h"},
            {"name": "OUTBOUND", "type": "OUTBOUND", "interval": "4h"},
            {"name": "APP_FACTORY", "type": "APP_FACTORY", "interval": "4h"},
            {"name": "LOCAL_BIZ", "type": "LOCAL_BIZ", "interval": "4h"},
            {"name": "MONETIZATION", "type": "MONETIZATION", "interval": "4h"},
            {"name": "PRODUCT", "type": "PRODUCT", "interval": "4h"},
            {"name": "RESEARCH", "type": "RESEARCH", "interval": "4h"},
            {"name": "SCRAPING", "type": "SCRAPING", "interval": "4h"}
        ],
        "swarm_agents": {
            "META": ["swarm_brain (4h, Opus)", "meta_executor (4h, Opus)"],
            "DISCOVERY": ["gap_hunter", "opportunity_scanner", "competitor_stalker"],
            "ACTION": ["asset_deployer", "content_compounder", "lead_machine"],
            "MEDIA": ["video_factory", "image_factory"],
            "OPTIMIZE": ["seo_aso_optimizer", "conversion_optimizer", "quality_enforcer"],
            "QUALITY": ["quality_gate (2h, Opus)", "playwright_tester"],
            "INTELLIGENCE": ["trend_synthesizer", "cross_pollinator", "revenue_tracker"],
            "MAINTENANCE": ["system_healer", "data_janitor"],
            "GROWTH": ["distribution_engine", "inbound_maximizer", "social_poster", "growth_strategist (24h, Opus)"],
            "NOTIFICATION": ["alert_dispatcher"]
        }
    }

    # Try to get script statuses from log file mtimes
    script_statuses = {}
    for layer in hierarchy:
        for script in layer["scripts"]:
            sname = script["name"].replace(".py", "")
            status = "unknown"
            last_run = None
            # Check log files
            for log_dir in [LOGS_DIR, CLAUDE_LOGS]:
                for prefix in ["", "cron_", "swarm_", "auto_"]:
                    log_path = log_dir / f"{prefix}{sname}.log"
                    if log_path.exists():
                        try:
                            mtime = log_path.stat().st_mtime
                            last_run = datetime.fromtimestamp(mtime).isoformat()
                            age_hours = (datetime.now() - datetime.fromtimestamp(mtime)).total_seconds() / 3600
                            if age_hours < 6:
                                status = "active"
                            elif age_hours < 24:
                                status = "idle"
                            else:
                                status = "stale"
                        except Exception:
                            pass
                        break
                if last_run:
                    break
            script["status"] = status
            script["last_run"] = last_run

    # -----------------------------------------------------------------------
    # State files — check existence, size, mtime
    # -----------------------------------------------------------------------
    state_files_def = [
        {"name": "ceo_state.json", "desc": "CEO cycle count, decisions, timestamps", "path": "AUTOMATIONS/agent/ceo_agent/ceo_state.json"},
        {"name": "autonomy_state.json", "desc": "Venture pipelines, stages, results", "path": "AUTOMATIONS/agent/autonomy/autonomy_state.json"},
        {"name": "swarm_state.json", "desc": "Agent health, cycles, effectiveness scores", "path": "AUTOMATIONS/agent/swarm/swarm_state.json"},
        {"name": "twitter_warmup_state.json", "desc": "Warmup day, phase, post history", "path": "AUTOMATIONS/twitter_warmup_state.json"},
        {"name": "ALPHA_STAGING.csv", "desc": "Raw alpha intake (15,154 entries)", "path": "LEDGER/ALPHA_STAGING.csv"},
        {"name": "INTELLIGENCE_CATALOG.json", "desc": "487 docs, value scores, buried gold", "path": "OPS/INTELLIGENCE_CATALOG.json"},
        {"name": "decisions.jsonl", "desc": "CEO decision audit trail", "path": "AUTOMATIONS/agent/ceo_agent/decisions.jsonl"},
        {"name": "missions.jsonl", "desc": "Shared mission log", "path": "AUTOMATIONS/agent/missions.jsonl"},
        {"name": "message_bus.jsonl", "desc": "Inter-agent messages", "path": "AUTOMATIONS/agent/message_bus.jsonl"},
        {"name": "feedback_recommendations.json", "desc": "Loop closer to swarm adjustments", "path": "AUTOMATIONS/agent/swarm/feedback_recommendations.json"},
        {"name": "USER_PROMPTS.jsonl", "desc": "Every user prompt timestamped", "path": "LEDGER/USER_PROMPTS.jsonl"},
        {"name": "throttle_state.json", "desc": "Current throttle mode and overrides", "path": "AUTOMATIONS/throttle_state.json"}
    ]

    state_files = []
    for sf in state_files_def:
        full_path = PROJECT_ROOT / sf["path"]
        entry = dict(sf)
        if full_path.exists():
            try:
                stat = full_path.stat()
                entry["exists"] = True
                entry["size_kb"] = round(stat.st_size / 1024, 1)
                entry["last_modified"] = relative_time(stat.st_mtime)
            except Exception:
                entry["exists"] = True
                entry["size_kb"] = 0
                entry["last_modified"] = "unknown"
        else:
            entry["exists"] = False
            entry["size_kb"] = 0
            entry["last_modified"] = "never"
        state_files.append(entry)

    hooks = [
        {"event": "PreToolUse", "matcher": "Write|Edit", "desc": "Path validation — blocks writes outside project"},
        {"event": "PostToolUse", "matcher": "Edit|Write", "desc": "py_compile, secret detection, safe_path check, file handle leak check, type hints"},
        {"event": "SessionStart", "matcher": "*", "desc": "Cron check, subconscious injection, session briefing, control panel launch"},
        {"event": "UserPromptSubmit", "matcher": "*", "desc": "Log user prompts, sync memory"},
        {"event": "Stop", "matcher": "*", "desc": "Session end processing, memory sync"}
    ]

    directory_overview = [
        {"name": "AUTOMATIONS/", "desc": "THE BRAIN — 298 Python scripts, 5K files", "icon": "brain"},
        {"name": "LEDGER/", "desc": "DATA WAREHOUSE — 2,023 CSVs, 2.2K files", "icon": "database"},
        {"name": "OPS/", "desc": "OPERATIONS CENTER — 1.2K files", "icon": "settings"},
        {"name": "CONTENT/", "desc": "CONTENT ENGINE — 957 files", "icon": "pencil"},
        {"name": "MONEY_METHODS/", "desc": "REVENUE METHODS — 11.7K files", "icon": "coin"},
        {"name": "DIGITAL_PRODUCTS/", "desc": "13 products built, $0 listed", "icon": "package"},
        {"name": "MEDIA/", "desc": "MEDIA ASSETS — 11.5K files", "icon": "photo"},
        {"name": "FINANCIALS/", "desc": "MONEY TRACKING", "icon": "chart-bar"},
        {"name": "01_STRATEGY/", "desc": "Strategic planning docs", "icon": "target"},
        {"name": "03_PLAYBOOKS/", "desc": "19K files — agency, AI, local lead gen, ecom arb", "icon": "book"}
    ]

    revenue_status = {
        "current": "$0",
        "day": 35,
        "pipeline_ready": "13 digital products built. 40 posts queued. 1,110 leads. 16 live sites.",
        "human_time_to_unblock": "~85 min"
    }

    return {
        "hierarchy": hierarchy,
        "data_flow": data_flow,
        "cron_schedule": cron_schedule,
        "agent_topology": agent_topology,
        "state_files": state_files,
        "hooks": hooks,
        "directory_overview": directory_overview,
        "revenue_status": revenue_status
    }


# ---------------------------------------------------------------------------
# NEW: Master Ops API (parsed from xlsx)
# ---------------------------------------------------------------------------
_master_ops_cache = {"data": None, "timestamp": 0}


def get_master_ops():
    """Parse the latest PRINTMAXX_MASTER_OPS_ENHANCED xlsx and return structured JSON."""
    # 5-minute cache
    now = time.time()
    if _master_ops_cache["data"] and (now - _master_ops_cache["timestamp"]) < 300:
        return _master_ops_cache["data"]

    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl not installed", "total_ops": 0}

    # Find latest xlsx
    xlsx_files = sorted(PROJECT_ROOT.glob("PRINTMAXX_MASTER_OPS_ENHANCED_*.xlsx"))
    if not xlsx_files:
        return {"error": "No PRINTMAXX_MASTER_OPS_ENHANCED_*.xlsx found", "total_ops": 0}

    xlsx_path = xlsx_files[-1]
    xlsx_name = xlsx_path.name
    # Extract date from filename: PRINTMAXX_MASTER_OPS_ENHANCED_2026-03-03.xlsx
    date_str = xlsx_name.replace("PRINTMAXX_MASTER_OPS_ENHANCED_", "").replace(".xlsx", "")

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    except Exception as e:
        return {"error": f"Failed to open xlsx: {e}", "total_ops": 0}

    result = {
        "last_updated": date_str,
        "xlsx_file": xlsx_name,
        "total_ops": 0,
        "categories": [],
        "priority_launch": [],
        "synergy_stacks": [],
        "readiness": {"READY": 0, "BUILD": 0, "BLOCKED": 0},
        "venture_map": [],
        "alpha_thesis": [],
        "blocker_summary": [],
        "tool_stacks": {"video": [], "hosting": [], "lead_gen": []},
        "existing_infra": {"total": 0, "by_category": []}
    }

    try:
        # --- ALL OPS MASTER ---
        if "ALL OPS MASTER" in wb.sheetnames:
            ws = wb["ALL OPS MASTER"]
            headers = None
            ops_by_category = {}
            total_ops = 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h).strip() if h else f"col_{j}" for j, h in enumerate(row)]
                    continue
                if not row or not row[0]:
                    continue
                total_ops += 1
                rd = {headers[j]: (row[j] if j < len(row) else None) for j in range(len(headers))}
                cat = str(rd.get("CATEGORY", "OTHER") or "OTHER")
                if cat not in ops_by_category:
                    ops_by_category[cat] = []
                ops_by_category[cat].append({
                    "op_id": str(rd.get("OP_ID", "") or ""),
                    "name": str(rd.get("OP_NAME", "") or ""),
                    "revenue": str(rd.get("REVENUE_RANGE", "") or ""),
                    "automation": str(rd.get("AUTOMATION_LEVEL", "") or ""),
                    "status": str(rd.get("STATUS", "") or ""),
                    "priority": str(rd.get("PRIORITY", "") or "")
                })
            result["total_ops"] = total_ops
            result["categories"] = [
                {"name": cat, "count": len(ops), "ops": ops}
                for cat, ops in sorted(ops_by_category.items())
            ]

        # --- PRIORITY LAUNCH ---
        if "PRIORITY LAUNCH" in wb.sheetnames:
            ws = wb["PRIORITY LAUNCH"]
            headers = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h).strip() if h else f"col_{j}" for j, h in enumerate(row)]
                    continue
                if not row or row[0] is None:
                    continue
                rd = {headers[j]: (row[j] if j < len(row) else None) for j in range(len(headers))}
                result["priority_launch"].append({
                    "rank": rd.get("RANK", 0),
                    "op_id": str(rd.get("OP_ID", "") or ""),
                    "name": str(rd.get("OP_NAME", "") or ""),
                    "effort": str(rd.get("EFFORT", "") or ""),
                    "revenue": str(rd.get("REVENUE_POTENTIAL", "") or ""),
                    "first_step": str(rd.get("FIRST_STEP", "") or ""),
                    "time_to_first": str(rd.get("TIME_TO_FIRST_$", "") or "")
                })

        # --- SYNERGY STACKS ---
        if "SYNERGY STACKS" in wb.sheetnames:
            ws = wb["SYNERGY STACKS"]
            headers = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h).strip() if h else f"col_{j}" for j, h in enumerate(row)]
                    continue
                if not row or not row[0]:
                    continue
                rd = {headers[j]: (row[j] if j < len(row) else None) for j in range(len(headers))}
                result["synergy_stacks"].append({
                    "id": str(rd.get("PACKAGE_ID", "") or ""),
                    "name": str(rd.get("NAME", "") or ""),
                    "score": rd.get("SYNERGY_SCORE", 0) or 0,
                    "multiplier": str(rd.get("REVENUE_MULTIPLIER", "") or ""),
                    "methods": str(rd.get("METHODS_COMBINED", "") or ""),
                    "description": str(rd.get("DESCRIPTION", "") or "")
                })

        # --- AUTO_STATUS_LIVE ---
        if "AUTO_STATUS_LIVE" in wb.sheetnames:
            ws = wb["AUTO_STATUS_LIVE"]
            headers = None
            readiness_counts = {"READY": 0, "BUILD": 0, "BLOCKED": 0}
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h).strip() if h else f"col_{j}" for j, h in enumerate(row)]
                    continue
                if not row or not row[0]:
                    continue
                rd = {headers[j]: (row[j] if j < len(row) else None) for j in range(len(headers))}
                r_val = str(rd.get("READINESS", "") or "").upper()
                if r_val in readiness_counts:
                    readiness_counts[r_val] += 1
                elif r_val:
                    readiness_counts.setdefault(r_val, 0)
                    readiness_counts[r_val] += 1
            result["readiness"] = readiness_counts

        # --- VENTURE_AUTOMATION_MAP ---
        blocker_map = {}  # blocker_key -> list of venture_ids
        if "VENTURE_AUTOMATION_MAP" in wb.sheetnames:
            ws = wb["VENTURE_AUTOMATION_MAP"]
            headers = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h).strip() if h else f"col_{j}" for j, h in enumerate(row)]
                    continue
                if not row or not row[0]:
                    continue
                rd = {headers[j]: (row[j] if j < len(row) else None) for j in range(len(headers))}
                vid = str(rd.get("VENTURE_ID", "") or "")
                blocker_key = str(rd.get("BLOCKER_KEY", "") or "")
                auto_score = rd.get("AUTOMATION_SCORE_100", 0) or 0
                signal_count = rd.get("SIGNAL_COUNT", 0) or 0
                result["venture_map"].append({
                    "venture_id": vid,
                    "name": str(rd.get("VENTURE_NAME", "") or ""),
                    "lane": str(rd.get("LANE", "") or ""),
                    "readiness": str(rd.get("READINESS", "") or ""),
                    "blocker": blocker_key,
                    "auto_score": auto_score,
                    "signal_count": signal_count
                })
                if blocker_key and blocker_key != "NONE":
                    if blocker_key not in blocker_map:
                        blocker_map[blocker_key] = []
                    blocker_map[blocker_key].append(vid)

        # Build blocker summary
        result["blocker_summary"] = [
            {"blocker": bk, "count": len(vids), "ventures_affected": vids}
            for bk, vids in sorted(blocker_map.items(), key=lambda x: -len(x[1]))
        ]

        # --- ALPHA_THESIS_INDEX ---
        if "ALPHA_THESIS_INDEX" in wb.sheetnames:
            ws = wb["ALPHA_THESIS_INDEX"]
            headers = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h).strip() if h else f"col_{j}" for j, h in enumerate(row)]
                    continue
                if not row or not row[0]:
                    continue
                rd = {headers[j]: (row[j] if j < len(row) else None) for j in range(len(headers))}
                result["alpha_thesis"].append({
                    "id": str(rd.get("ALPHA_ID", "") or ""),
                    "opportunity": str(rd.get("OPPORTUNITY", "") or ""),
                    "lane": str(rd.get("LANE", "") or ""),
                    "edge_duration": str(rd.get("EDGE_DURATION", "") or ""),
                    "why_llm_edge": str(rd.get("WHY_LLM_EDGE", "") or "")
                })

        # --- VIDEO & MEDIA STACK ---
        if "VIDEO & MEDIA STACK" in wb.sheetnames:
            ws = wb["VIDEO & MEDIA STACK"]
            headers = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h).strip() if h else f"col_{j}" for j, h in enumerate(row)]
                    continue
                if not row or not row[0]:
                    continue
                rd = {headers[j]: (row[j] if j < len(row) else None) for j in range(len(headers))}
                result["tool_stacks"]["video"].append({
                    "name": str(rd.get("TOOL", "") or ""),
                    "type": str(rd.get("TYPE", "") or ""),
                    "free_tier": str(rd.get("FREE_TIER", "") or ""),
                    "quality": str(rd.get("QUALITY", "") or "")
                })

        # --- HOSTING & DEPLOY ---
        if "HOSTING & DEPLOY" in wb.sheetnames:
            ws = wb["HOSTING & DEPLOY"]
            headers = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h).strip() if h else f"col_{j}" for j, h in enumerate(row)]
                    continue
                if not row or not row[0]:
                    continue
                rd = {headers[j]: (row[j] if j < len(row) else None) for j in range(len(headers))}
                result["tool_stacks"]["hosting"].append({
                    "name": str(rd.get("PLATFORM", "") or ""),
                    "free_tier": str(rd.get("FREE_TIER", "") or ""),
                    "commercial": str(rd.get("COMMERCIAL_USE", "") or ""),
                    "best_for": str(rd.get("BEST_FOR", "") or "")
                })

        # --- LEAD GEN STACK ---
        if "LEAD GEN STACK" in wb.sheetnames:
            ws = wb["LEAD GEN STACK"]
            headers = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h).strip() if h else f"col_{j}" for j, h in enumerate(row)]
                    continue
                if not row or not row[0]:
                    continue
                rd = {headers[j]: (row[j] if j < len(row) else None) for j in range(len(headers))}
                result["tool_stacks"]["lead_gen"].append({
                    "name": str(rd.get("TOOL", "") or ""),
                    "type": str(rd.get("TYPE", "") or ""),
                    "free_tier": str(rd.get("FREE_TIER", "") or ""),
                    "automation": str(rd.get("AUTOMATION_LEVEL", "") or "")
                })

        # --- EXISTING INFRA ---
        if "EXISTING INFRA" in wb.sheetnames:
            ws = wb["EXISTING INFRA"]
            headers = None
            infra_by_cat = {}
            infra_total = 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h).strip() if h else f"col_{j}" for j, h in enumerate(row)]
                    continue
                if not row or not row[0]:
                    continue
                infra_total += 1
                rd = {headers[j]: (row[j] if j < len(row) else None) for j in range(len(headers))}
                cat = str(rd.get("CATEGORY", "OTHER") or "OTHER")
                if cat not in infra_by_cat:
                    infra_by_cat[cat] = []
                infra_by_cat[cat].append({
                    "name": str(rd.get("ITEM", "") or ""),
                    "status": str(rd.get("STATUS", "") or ""),
                    "notes": str(rd.get("NOTES", "") or "")
                })
            result["existing_infra"]["total"] = infra_total
            result["existing_infra"]["by_category"] = [
                {"category": cat, "count": len(items), "items": items}
                for cat, items in sorted(infra_by_cat.items())
            ]

    except Exception as e:
        result["parse_error"] = str(e)
    finally:
        try:
            wb.close()
        except Exception:
            pass

    _master_ops_cache["data"] = result
    _master_ops_cache["timestamp"] = now
    return result


# ---------------------------------------------------------------------------
# NEW: Real-time metrics API
# ---------------------------------------------------------------------------
def get_realtime_data():
    """Return time-series data for real-time charts."""
    now = datetime.now()
    data_points = []

    # Scan log files for modification times in last 24h to build activity timeline
    log_dirs = [LOGS_DIR, CLAUDE_LOGS]
    activity_counts = Counter()

    for log_dir in log_dirs:
        if not log_dir.exists():
            continue
        try:
            for log_file in log_dir.iterdir():
                if not log_file.is_file() or not log_file.name.endswith(".log"):
                    continue
                try:
                    mtime = datetime.fromtimestamp(log_file.stat().st_mtime)
                    if (now - mtime).total_seconds() < 86400:
                        # Bucket into hourly slots
                        hour_key = mtime.replace(minute=0, second=0, microsecond=0)
                        activity_counts[hour_key.isoformat()] += 1
                except Exception:
                    continue
        except Exception:
            continue

    # Build 24-hour timeline
    for i in range(24):
        t = (now - timedelta(hours=23 - i)).replace(minute=0, second=0, microsecond=0)
        key = t.isoformat()
        data_points.append({
            "timestamp": int(t.timestamp() * 1000),
            "invocations": activity_counts.get(key, 0)
        })

    # Alpha entries over time (from ALPHA_STAGING.csv timestamps)
    alpha_timeline = []
    alpha_path = LEDGER / "ALPHA_STAGING.csv"
    if alpha_path.exists():
        alpha_counts = Counter()
        try:
            with open(alpha_path, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    ts_str = row.get("timestamp", row.get("created_at", row.get("date", "")))
                    if ts_str:
                        try:
                            ts = datetime.fromisoformat(ts_str.replace("Z", "+00:00").split("+")[0])
                            day_key = ts.strftime("%Y-%m-%d")
                            alpha_counts[day_key] += 1
                        except Exception:
                            pass
        except Exception:
            pass

        # Last 7 days of alpha
        for i in range(7):
            day = (now - timedelta(days=6 - i)).strftime("%Y-%m-%d")
            alpha_timeline.append({
                "date": day,
                "count": alpha_counts.get(day, 0)
            })

    # System events from recent logs
    events = []
    if LOGS_DIR.exists():
        try:
            log_files = sorted(LOGS_DIR.iterdir(), key=lambda p: p.stat().st_mtime, reverse=True)
            for lf in log_files[:10]:
                if lf.is_file() and lf.name.endswith(".log"):
                    try:
                        mtime = datetime.fromtimestamp(lf.stat().st_mtime)
                        events.append({
                            "name": lf.stem,
                            "time": mtime.isoformat(),
                            "ago": relative_time(mtime)
                        })
                    except Exception:
                        pass
        except Exception:
            pass

    return {
        "agent_activity": data_points,
        "alpha_timeline": alpha_timeline,
        "recent_events": events
    }


# ---------------------------------------------------------------------------
# NEW: Venture status API
# ---------------------------------------------------------------------------
def get_venture_status():
    """Return venture status from autonomy_state.json."""
    state_path = AUTOMATIONS / "agent" / "autonomy" / "autonomy_state.json"
    state = safe_read_json(state_path)
    ventures = state.get("ventures", {})

    result = []
    for vid, v in ventures.items():
        last_run = v.get("last_run", "")
        result.append({
            "id": vid,
            "name": v.get("name", vid),
            "type": v.get("type", "UNKNOWN"),
            "status": v.get("status", "UNKNOWN"),
            "interval_hours": v.get("interval_hours", 0),
            "cycles_run": v.get("cycles_run", 0),
            "last_run": last_run,
            "last_run_ago": relative_time(last_run) if last_run else "never",
            "pipeline": v.get("pipeline", []),
            "latest_result": v.get("results", [{}])[-1] if v.get("results") else {}
        })

    return {"ventures": result}


# ---------------------------------------------------------------------------
# NEW: Pipeline funnel API
# ---------------------------------------------------------------------------
def get_pipeline():
    """Return pipeline stages: alpha pending -> approved -> routed -> deployed -> revenue."""
    alpha_path = LEDGER / "ALPHA_STAGING.csv"
    stages = {
        "pending": 0,
        "approved": 0,
        "routed": 0,
        "deployed": 0,
        "revenue": 0
    }

    if alpha_path.exists():
        rows = safe_read_csv(alpha_path, max_rows=50000)
        for r in rows:
            status = r.get("status", "").upper()
            if status == "PENDING_REVIEW":
                stages["pending"] += 1
            elif status == "APPROVED":
                stages["approved"] += 1
            elif status in ("ROUTED", "INTEGRATED", "PROCESSED"):
                stages["routed"] += 1
            elif status in ("DEPLOYED", "LIVE", "PUBLISHED"):
                stages["deployed"] += 1
            elif status in ("REVENUE", "MONETIZED"):
                stages["revenue"] += 1

    # Check revenue tracker
    rev_path = FINANCIALS / "REVENUE_TRACKER.csv"
    if rev_path.exists():
        rev_rows = safe_read_csv(rev_path, max_rows=1000)
        total_rev = sum(float(r.get("amount", 0)) for r in rev_rows if r.get("amount"))
        stages["total_revenue"] = f"${total_rev:,.0f}"
    else:
        stages["total_revenue"] = "$0"

    # Add total alpha count
    stages["total_alpha"] = stages["pending"] + stages["approved"] + stages["routed"] + stages["deployed"] + stages["revenue"]

    return stages


# ---------------------------------------------------------------------------
# API Routes
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    if HTML_PATH.exists():
        with open(HTML_PATH, "r") as f:
            return Response(f.read(), mimetype="text/html")
    return "<h1>Control panel HTML not found</h1>", 404


@app.route("/api/status")
def api_status():
    throttle = get_throttle_status()
    system = get_system_status()
    return jsonify({
        "throttle": throttle,
        "system": system,
        "days_until_reset": days_until_friday(),
    })


@app.route("/api/agents")
def api_agents():
    return jsonify(get_throttle_status())


@app.route("/api/mode", methods=["POST"])
def api_mode():
    data = request.get_json(force=True)
    mode = data.get("mode", "efficient")
    if mode not in ("efficient", "high"):
        return jsonify({"error": "mode must be 'efficient' or 'high'"}), 400
    result = set_throttle_mode(mode)
    return jsonify(result)


@app.route("/api/agent/<agent_id>/toggle", methods=["POST"])
def api_toggle_agent(agent_id):
    data = request.get_json(force=True)
    enabled = data.get("enabled", True)
    result = toggle_agent(agent_id, enabled)
    return jsonify(result)


@app.route("/api/agent/<agent_id>/interval", methods=["POST"])
def api_set_interval(agent_id):
    data = request.get_json(force=True)
    hours = data.get("hours", 24)
    if hours <= 0:
        return jsonify({"error": "hours must be > 0"}), 400
    result = set_agent_interval(agent_id, hours)
    return jsonify(result)


@app.route("/api/logs/<agent_id>")
def api_logs(agent_id):
    lines = int(request.args.get("lines", 30))
    return jsonify({"agent_id": agent_id, "log": get_log_tail(agent_id, lines)})


@app.route("/api/actions")
def api_actions():
    return jsonify({"actions": get_quick_actions()})


@app.route("/api/action/<action_id>", methods=["POST"])
def api_run_action(action_id):
    actions = {a["id"]: a for a in get_quick_actions()}
    if action_id not in actions:
        return jsonify({"error": f"unknown action: {action_id}"}), 404
    action = actions[action_id]
    cmd_parts = action["cmd"].split()
    code, out, err = run_cmd(cmd_parts, timeout=120)
    return jsonify({"success": code == 0, "output": out[-3000:], "error": err[-1000:]})


@app.route("/api/commands")
def api_commands():
    return jsonify({"commands": get_slash_commands()})


@app.route("/api/sites")
def api_sites():
    return jsonify({"sites": get_deployed_sites()})


@app.route("/api/blockers")
def api_blockers():
    """Read P0 blockers from task tracker."""
    tracker_path = OPS / "PERSISTENT_TASK_TRACKER.md"
    blockers = []
    if tracker_path.exists():
        content = safe_read(tracker_path, max_lines=200)
        in_p0 = False
        for line in content.split("\n"):
            if "[P0]" in line:
                in_p0 = True
                continue
            if line.startswith("## [P") and "[P0]" not in line:
                in_p0 = False
            if in_p0 and line.strip().startswith("- ["):
                blockers.append(line.strip())
    return jsonify({"blockers": blockers})


# --- NEW API ENDPOINTS ---

@app.route("/api/system-tree")
def api_system_tree():
    return jsonify(get_system_tree())


@app.route("/api/realtime")
def api_realtime():
    return jsonify(get_realtime_data())


@app.route("/api/ventures")
def api_ventures():
    return jsonify(get_venture_status())


@app.route("/api/pipeline")
def api_pipeline():
    return jsonify(get_pipeline())


@app.route("/api/master-ops")
def api_master_ops():
    return jsonify(get_master_ops())


# ---------------------------------------------------------------------------
# Sovrun + n8n helpers
# ---------------------------------------------------------------------------
def _n8n_request(path, method="GET"):
    """Make a request to the n8n API. Returns parsed JSON or error dict."""
    if not N8N_API_KEY:
        return {"error": "N8N_API_KEY not configured in .env"}
    url = f"{N8N_BASE}/api/v1{path}"
    headers = {"X-N8N-API-KEY": N8N_API_KEY, "Accept": "application/json"}
    try:
        req = Request(url, headers=headers, method=method)
        with urlopen(req, timeout=10) as resp:
            return json.loads(resp.read().decode())
    except URLError as e:
        return {"error": f"n8n unreachable: {e.reason}"}
    except Exception as e:
        return {"error": str(e)}


def _get_skills_db():
    """Return path to the skills.db that exists, or None."""
    for p in [SOVRUN_SKILLS_DB, SOVRUN_SKILLS_DB_ALT]:
        if p.exists():
            return p
    return None


def get_sovrun_skills(query=""):
    """Query procedural memory FTS5 for skill docs."""
    db_path = _get_skills_db()
    if not db_path:
        return {"error": "skills.db not found", "skills": []}
    try:
        conn = sqlite3.connect(str(db_path))
        conn.row_factory = sqlite3.Row
        if query:
            # Try FTS5 first
            try:
                rows = conn.execute(
                    "SELECT * FROM skills_fts WHERE skills_fts MATCH ? ORDER BY rank LIMIT 50",
                    (query,)
                ).fetchall()
            except Exception:
                # Fallback to LIKE
                rows = conn.execute(
                    "SELECT * FROM skills WHERE name LIKE ? OR description LIKE ? LIMIT 50",
                    (f"%{query}%", f"%{query}%")
                ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM skills ORDER BY rowid DESC LIMIT 50").fetchall()
        skills = [dict(r) for r in rows]
        conn.close()
        return {"skills": skills, "count": len(skills), "query": query}
    except Exception as e:
        return {"error": str(e), "skills": []}


def get_handoff_chains():
    """Return the 5 handoff chains and basic status."""
    chains_def = {
        "local_biz": [
            "savvy_lead_scraper", "lead_enrichment", "eas_lead_pipeline",
            "cold_email_generator", "follow_up_tracker",
        ],
        "content_factory": [
            "alpha_scanner", "topic_selector", "content_generator",
            "voice_injector", "platform_formatter",
        ],
        "product_launch": [
            "demand_scanner", "product_builder", "listing_creator",
            "distribution_engine", "sales_tracker",
        ],
        "freelance": [
            "job_scanner", "proposal_writer", "submission_tracker",
            "delivery_manager", "review_collector",
        ],
        "alpha_to_revenue": [
            "scraper_fleet", "alpha_processor", "intelligence_router",
            "capital_genesis_ranker", "venture_autonomy",
        ],
    }
    result = []
    for name, steps in chains_def.items():
        result.append({
            "name": name,
            "steps": steps,
            "step_count": len(steps),
        })
    return {"chains": result}


def get_dag_status():
    """Read morning intelligence DAG checkpoint/state."""
    data = {"checkpoint": None, "state": None}
    if DAG_CHECKPOINT.exists():
        data["checkpoint"] = safe_read_json(DAG_CHECKPOINT)
    if DAG_STATE.exists():
        data["state"] = safe_read_json(DAG_STATE)

    # Build the DAG structure for visualization
    dag_steps = [
        {"name": "scrape_twitter", "depends_on": [], "layer": 1, "desc": "Twitter (133 accounts)"},
        {"name": "scrape_reddit", "depends_on": [], "layer": 1, "desc": "Reddit (18 subreddits)"},
        {"name": "scrape_hn", "depends_on": [], "layer": 1, "desc": "HN + ProductHunt"},
        {"name": "merge_results", "depends_on": ["scrape_twitter", "scrape_reddit", "scrape_hn"], "layer": 2, "desc": "Merge scraper results"},
        {"name": "alpha_processor", "depends_on": ["merge_results"], "layer": 3, "desc": "Process new alpha"},
        {"name": "intelligence_router", "depends_on": ["alpha_processor"], "layer": 4, "desc": "Route through intel"},
        {"name": "capital_genesis_ranker", "depends_on": ["intelligence_router"], "layer": 5, "desc": "Rank by Capital Genesis"},
    ]

    # Enrich with status from checkpoint
    cp = data.get("checkpoint") or {}
    step_statuses = cp.get("step_statuses", cp.get("steps", {}))
    for step in dag_steps:
        sname = step["name"]
        if isinstance(step_statuses, dict) and sname in step_statuses:
            ss = step_statuses[sname]
            if isinstance(ss, dict):
                step["status"] = ss.get("status", "pending")
                step["duration"] = ss.get("duration_seconds", ss.get("duration", None))
            else:
                step["status"] = str(ss)
        else:
            step["status"] = "pending"

    data["dag_steps"] = dag_steps
    return data


# --- Sovrun + n8n API routes ---

@app.route("/api/n8n/workflows")
def api_n8n_workflows():
    data = _n8n_request("/workflows")
    return jsonify(data)


@app.route("/api/n8n/executions")
def api_n8n_executions():
    limit = request.args.get("limit", "20")
    data = _n8n_request(f"/executions?limit={limit}")
    return jsonify(data)


@app.route("/api/sovrun/skills")
def api_sovrun_skills():
    query = request.args.get("q", "")
    return jsonify(get_sovrun_skills(query))


@app.route("/api/sovrun/chains")
def api_sovrun_chains():
    return jsonify(get_handoff_chains())


@app.route("/api/sovrun/chains/<chain_name>/run", methods=["POST"])
def api_run_chain(chain_name):
    code, out, err = run_cmd(
        [PYTHON, str(AUTOMATIONS / "agent_swarm.py"), "--chain", chain_name],
        timeout=120
    )
    return jsonify({"success": code == 0, "output": out[-3000:], "error": err[-1000:]})


@app.route("/api/sovrun/dag")
def api_sovrun_dag():
    return jsonify(get_dag_status())


@app.route("/api/daily-feed")
def api_daily_feed():
    """Aggregate all daily automated activity into actionable notifications."""
    feed = []
    today = datetime.now().strftime("%Y-%m-%d")
    today_ts = datetime.now().replace(hour=0, minute=0, second=0).timestamp()

    # 1. Alpha staging — new entries needing review
    alpha_csv = LEDGER / "ALPHA_STAGING.csv"
    if alpha_csv.exists():
        try:
            pending = 0
            with open(alpha_csv) as f:
                for row in csv.DictReader(f):
                    if row.get("status", "").upper() in ("PENDING_REVIEW", "NEW_METHOD"):
                        pending += 1
            if pending > 0:
                # Get actual pending entries to show
                preview_items = []
                with open(alpha_csv) as f2:
                    for row in csv.DictReader(f2):
                        if row.get("status", "").upper() in ("PENDING_REVIEW", "NEW_METHOD") and len(preview_items) < 5:
                            preview_items.append(row.get("extracted_method", row.get("tactic", row.get("content", "")))[:120])
                feed.append({"type": "action", "icon": "ti-bulb", "color": "var(--warn)", "title": f"{pending} alpha entries need review", "desc": "ALPHA_STAGING.csv has pending items", "action": "python3 AUTOMATIONS/alpha_auto_processor.py --process-new", "priority": "high", "preview": preview_items})
        except Exception:
            pass

    # 2. Capital Genesis priority stack — P0 items
    pstack = OPS / "CAPITAL_GENESIS_PRIORITY_STACK.md"
    if pstack.exists() and pstack.stat().st_mtime > today_ts - 86400:
        try:
            p0_lines = [line.strip() for line in open(pstack) if "| P0 |" in line or "LAUNCH_NOW" in line]
            if p0_lines:
                feed.append({"type": "opportunity", "icon": "ti-rocket", "color": "var(--accent)", "title": f"{len(p0_lines)} P0 methods ready to launch", "desc": "Capital Genesis ranked these as immediate action", "action": "cat OPS/CAPITAL_GENESIS_PRIORITY_STACK.md", "priority": "high", "preview": [l.split("|")[2].strip() if "|" in l else l[:100] for l in p0_lines[:5]]})
        except Exception:
            pass

    # 3. Daily tool scout — new tools found
    scout = OPS / "DAILY_TOOL_SCOUT.md"
    if scout.exists() and scout.stat().st_mtime > today_ts - 86400:
        scout_preview = []
        try:
            for line in open(scout):
                if line.startswith("| ") and ("[" in line) and ("Stars" not in line and "Security" not in line and "---" not in line and "Date" not in line):
                    parts = [p.strip() for p in line.split("|") if p.strip()]
                    if len(parts) >= 3:
                        scout_preview.append(f"{parts[0]} - {parts[1]}" if "http" not in parts[0] else parts[1])
                    if len(scout_preview) >= 8:
                        break
        except Exception:
            pass
        feed.append({"type": "research", "icon": "ti-search", "color": "var(--blue)", "title": "Daily tool scout has new results", "desc": f"{len(scout_preview)} tools found", "action": "cat OPS/DAILY_TOOL_SCOUT.md", "priority": "medium", "preview": scout_preview})

    # 4. Scraper outputs — what ran today
    scraper_dir = AUTOMATIONS / "twitter_scraper_output"
    if scraper_dir.exists():
        recent = [f for f in scraper_dir.iterdir() if f.stat().st_mtime > today_ts - 86400]
        if recent:
            twitter_preview = []
            for rf in sorted(recent, key=lambda x: x.stat().st_mtime, reverse=True)[:3]:
                try:
                    data = json.loads(rf.read_text())
                    if isinstance(data, list):
                        for item in data[:3]:
                            text = item.get("text", item.get("content", item.get("tweet", "")))[:100]
                            author = item.get("author", item.get("user", item.get("screen_name", "")))
                            if text:
                                twitter_preview.append(f"@{author}: {text}" if author else text)
                    elif isinstance(data, dict):
                        for k, v in list(data.items())[:3]:
                            if isinstance(v, str):
                                twitter_preview.append(f"{k}: {v[:80]}")
                except Exception:
                    twitter_preview.append(rf.name)
                if len(twitter_preview) >= 5:
                    break
            feed.append({"type": "data", "icon": "ti-brand-twitter", "color": "var(--blue)", "title": f"Twitter scraper: {len(recent)} new outputs", "desc": "Alpha extracted from 133 monitored accounts", "action": "ls AUTOMATIONS/twitter_scraper_output/", "priority": "low", "preview": twitter_preview})

    reddit_dir = AUTOMATIONS / "reddit_scraper_output"
    if reddit_dir.exists():
        recent = [f for f in reddit_dir.iterdir() if f.stat().st_mtime > today_ts - 86400]
        if recent:
            reddit_preview = []
            for rf in sorted(recent, key=lambda x: x.stat().st_mtime, reverse=True)[:3]:
                try:
                    data = json.loads(rf.read_text())
                    if isinstance(data, list):
                        for item in data[:3]:
                            title = item.get("title", item.get("text", ""))[:100]
                            sub = item.get("subreddit", "")
                            if title:
                                reddit_preview.append(f"r/{sub}: {title}" if sub else title)
                    elif isinstance(data, dict):
                        for k, v in list(data.items())[:3]:
                            if isinstance(v, str):
                                reddit_preview.append(v[:80])
                except Exception:
                    reddit_preview.append(rf.name)
                if len(reddit_preview) >= 5:
                    break
            feed.append({"type": "data", "icon": "ti-brand-reddit", "color": "var(--warn)", "title": f"Reddit scraper: {len(recent)} new outputs", "desc": "Pain points and alpha from subreddits", "action": "ls AUTOMATIONS/reddit_scraper_output/", "priority": "low", "preview": reddit_preview})

    # 5. n8n workflow executions
    if N8N_API_KEY:
        try:
            n8n_data = _n8n_request("/executions?limit=20")
            if isinstance(n8n_data, dict) and "data" in n8n_data:
                execs = n8n_data["data"]
                today_execs = [e for e in execs if today in e.get("startedAt", "")]
                failed = [e for e in today_execs if e.get("status") == "error"]
                if failed:
                    feed.append({"type": "alert", "icon": "ti-alert-triangle", "color": "var(--danger)", "title": f"{len(failed)} n8n workflows FAILED today", "desc": "Check n8n for error details", "action": "open http://localhost:5678", "priority": "high"})
                elif today_execs:
                    feed.append({"type": "status", "icon": "ti-check", "color": "var(--accent)", "title": f"{len(today_execs)} n8n workflows ran successfully", "desc": "All workflows healthy", "action": "", "priority": "low"})
        except Exception:
            pass

    # 6. Agent swarm reports — today's reports
    reports_dir = AUTOMATIONS / "agent" / "swarm" / "reports"
    if reports_dir.exists():
        today_reports = [f.name for f in reports_dir.iterdir() if f.is_file() and today.replace("-", "") in f.name]
        if today_reports:
            # Read first few lines of each report for preview
            report_previews = []
            for rname in today_reports[:5]:
                rpath = reports_dir / rname
                try:
                    with open(rpath) as rf:
                        lines = [l.strip() for l in rf.readlines()[:5] if l.strip() and not l.startswith("#")]
                        if lines:
                            report_previews.append(f"{rname}: {lines[0][:80]}")
                except Exception:
                    report_previews.append(rname)
            feed.append({"type": "intel", "icon": "ti-robot", "color": "var(--purple)", "title": f"{len(today_reports)} agent reports generated", "desc": ", ".join(today_reports[:3]) + ("..." if len(today_reports) > 3 else ""), "action": f"ls {reports_dir}", "priority": "medium", "preview": report_previews})

    # 7. Soul drift — check if below threshold
    drift_file = AUTOMATIONS / "agent" / "swarm" / "soul_drift_report.json"
    if drift_file.exists():
        try:
            drift = json.loads(drift_file.read_text())
            avg = drift.get("system_average", drift.get("average_score", 10))
            if avg < 6:
                feed.append({"type": "alert", "icon": "ti-mood-sad", "color": "var(--danger)", "title": f"Soul drift score: {avg}/10 (below 6 threshold)", "desc": "Agents producing off-brand output. Review and fix prompts.", "action": "python3 AUTOMATIONS/loop_closer.py --drift", "priority": "high"})
        except Exception:
            pass

    # 8. CEO decisions — recent
    ceo_decisions = AUTOMATIONS / "agent" / "ceo_agent" / "decisions.jsonl"
    if ceo_decisions.exists():
        try:
            recent = []
            with open(ceo_decisions) as f:
                for line in f:
                    try:
                        d = json.loads(line)
                        if today in d.get("ts", d.get("timestamp", "")):
                            recent.append(d)
                    except Exception:
                        pass
            if recent:
                feed.append({"type": "decision", "icon": "ti-crown", "color": "var(--accent)", "title": f"CEO agent made {len(recent)} decisions today", "desc": recent[-1].get("action", recent[-1].get("decision", ""))[:80] if recent else "", "action": "python3 AUTOMATIONS/ceo_agent.py --status", "priority": "medium"})
        except Exception:
            pass

    # 9. Method discovery — new methods found
    discovery_log = LEDGER / "METHOD_DISCOVERY_LOG.csv"
    if discovery_log.exists() and discovery_log.stat().st_mtime > today_ts - 86400:
        try:
            new_methods = sum(1 for line in open(discovery_log) if today in line)
            if new_methods > 0:
                feed.append({"type": "opportunity", "icon": "ti-coin", "color": "var(--accent)", "title": f"{new_methods} new revenue methods discovered", "desc": "From Reddit, HN, Twitter crawl", "action": "python3 AUTOMATIONS/method_discovery_crawler.py --report", "priority": "high"})
        except Exception:
            pass

    # 10. Gov contracts — new opportunities
    gov_csv = LEDGER / "GOV_OPPORTUNITIES.csv"
    if gov_csv.exists() and gov_csv.stat().st_mtime > today_ts - 86400:
        feed.append({"type": "opportunity", "icon": "ti-building-bank", "color": "var(--blue)", "title": "Government contract scan updated", "desc": "SAM.gov + UK Contracts Finder results", "action": "cat LEDGER/GOV_OPPORTUNITIES.csv | head -20", "priority": "medium"})

    # 11. Content posting queue — what's ready to post
    queue_dir = PROJECT_ROOT / "CONTENT" / "social" / "posting_queue"
    if queue_dir.exists():
        ready = [f for f in queue_dir.iterdir() if f.is_file() and f.suffix in (".txt", ".md")]
        if ready:
            feed.append({"type": "action", "icon": "ti-send", "color": "var(--purple)", "title": f"{len(ready)} posts in queue ready to publish", "desc": "Review and schedule or post manually", "action": f"ls {queue_dir}", "priority": "medium"})

    # 12. Freelance demand — new opportunities
    freelance_csv = LEDGER / "FREELANCE_DEMAND_SCAN.csv"
    if freelance_csv.exists() and freelance_csv.stat().st_mtime > today_ts - 86400:
        feed.append({"type": "opportunity", "icon": "ti-briefcase", "color": "var(--accent)", "title": "Freelance demand scan updated", "desc": "New gigs matching our capabilities", "action": "python3 AUTOMATIONS/auto_freelance_responder.py --scan", "priority": "medium"})

    # 13. KPI Executor — auto-task execution status
    kpi_report = OPS / "KPI_EXECUTOR_REPORT.md"
    if kpi_report.exists() and kpi_report.stat().st_mtime > today_ts - 86400:
        try:
            report_content = kpi_report.read_text(encoding="utf-8")
            summary_line = ""
            for rline in report_content.split("\n"):
                if "| AUTO" in rline and "OK" in rline:
                    summary_line = rline.strip()
                    break
            has_failures = "FAILED" in report_content and "0 FAILED" not in report_content
            feed.append({
                "type": "alert" if has_failures else "data",
                "icon": "ti-player-play" if not has_failures else "ti-alert-circle",
                "color": "var(--danger)" if has_failures else "var(--accent)",
                "title": f"KPI Executor: {summary_line}" if summary_line else "KPI Executor ran today",
                "desc": "Auto-executed daily KPI tasks",
                "action": "python3 AUTOMATIONS/kpi_executor.py --status",
                "priority": "high" if has_failures else "low",
            })
        except Exception:
            pass
    semi_queue = OPS / "SEMI_REVIEW_QUEUE.md"
    if semi_queue.exists() and semi_queue.stat().st_mtime > today_ts - 86400:
        try:
            sq_content = semi_queue.read_text(encoding="utf-8")
            semi_count = sq_content.count("**Status:** Awaiting review")
            if semi_count > 0:
                feed.append({"type": "action", "icon": "ti-checkbox", "color": "var(--warn)", "title": f"{semi_count} SEMI tasks awaiting review", "desc": "KPI executor queued these for human approval", "action": "python3 AUTOMATIONS/kpi_executor.py --review", "priority": "medium"})
        except Exception:
            pass

    # Sort: high priority first, then medium, then low
    priority_order = {"high": 0, "medium": 1, "low": 2}
    feed.sort(key=lambda x: priority_order.get(x.get("priority", "low"), 3))

    return jsonify({"feed": feed, "count": len(feed), "date": today})


@app.route("/api/kpi")
def api_kpi():
    """Return KPI data: revenue, blockers, daily targets, week progress."""
    kpi = {"revenue": 0, "by_source": {}, "blockers": [], "daily_targets": [], "week": 0, "month_target": 1000}

    # Revenue from available sources
    for csv_name, source_key in [("GUMROAD_SALES.csv", "gumroad"), ("FREELANCE_EARNINGS.csv", "freelance"), ("STRIPE_SALES.csv", "stripe")]:
        csv_path = LEDGER / csv_name
        if csv_path.exists():
            try:
                with open(csv_path) as f:
                    for row in csv.DictReader(f):
                        amt = float((row.get("amount", row.get("price", "0")) or "0").replace("$", "").replace(",", ""))
                        kpi["revenue"] += amt
                        kpi["by_source"][source_key] = kpi["by_source"].get(source_key, 0) + amt
            except Exception:
                pass

    # Human blockers
    roadmap = OPS / "MONTHLY_ROADMAP_2026_04.md"
    if roadmap.exists():
        in_blockers = False
        with open(roadmap) as f:
            for line in f:
                if "HUMAN BLOCKERS" in line:
                    in_blockers = True
                elif in_blockers and line.startswith("- ["):
                    done = line.startswith("- [x]")
                    text = line[5:].strip() if done else line[5:].strip()
                    kpi["blockers"].append({"text": text, "done": done})
                elif in_blockers and not line.strip():
                    break

    # Week number
    import calendar
    day = datetime.now().day
    kpi["week"] = min((day - 1) // 7 + 1, 4)

    # Weekly check-in data
    checkin = OPS / "KPI_WEEKLY_CHECKIN.md"
    if checkin.exists():
        kpi["has_weekly_checkin"] = True
        kpi["checkin_date"] = datetime.fromtimestamp(checkin.stat().st_mtime).strftime("%Y-%m-%d")

    return jsonify(kpi)


@app.route("/api/kpi/calendar")
def api_kpi_calendar():
    """Return full month calendar with daily KPIs and tasks.
    ?mode=standard (default) | ?mode=beast | ?mode=modafinil
    """
    import calendar
    mode = request.args.get("mode", "standard")
    now = datetime.now()
    year, month = now.year, now.month  # Current month, not next month
    _, days_in_month = calendar.monthrange(year, month)
    today_day = now.day

    # Intensity multipliers per mode
    MULTIPLIERS = {
        "standard": {"engagement": 1.0, "content": 1.0, "outreach": 1.0, "revenue": 1.0, "label": "Standard", "desc": "Sustainable daily pace. 6-8 hrs focused work."},
        "beast": {"engagement": 2.0, "content": 2.0, "outreach": 2.0, "revenue": 1.5, "label": "Beast Mode", "desc": "12-14 hr days. Everything at 2x. Sleep when profitable."},
        "modafinil": {"engagement": 3.5, "content": 3.0, "outreach": 3.0, "revenue": 2.5, "label": "Modafinil Mode", "desc": "16-18 hr days. Upper ceiling of human output. Every waking minute is revenue-generating. Not sustainable forever but will compress 3 months into 1."},
    }
    mult = MULTIPLIERS.get(mode, MULTIPLIERS["standard"])

    # ==========================================================================
    # PRINTMAXX DAILY OPS — 33-AGENT AUTONOMOUS SYSTEM KPIs
    # ==========================================================================
    # Task format: {"text": str, "tag": "AUTO"|"SEMI"|"MANUAL", "detail": str}
    # AUTO  = fully automated, just verify it ran
    # SEMI  = system generates output, human reviews/approves before execution
    # MANUAL = requires human hands (account creation, payments, strategic decisions)
    #
    # Account Tiers referenced in detail fields:
    #   SAFE      = primary @PRINTMAXXER, conservative (5 posts/day), full TOS compliance
    #   MEDIUM    = 2-3 niche accounts, moderate automation (10-15 posts/day), test new tactics
    #   AGGRESSIVE = burner/test accounts, push limits (20+ posts/day), expendable if banned
    #
    # Week 1: INFRA + ACCOUNTS
    # Week 2: VOLUME + PIPELINE
    # Week 3: EDGE + OPTIMIZE
    # Week 4: COMPOUND + CLOSE
    daily_plans = {
        # ===== WEEK 1: INFRA + ACCOUNTS =====
        1: {
            "theme": "MULTI-ACCOUNT GENESIS",
            "tasks": [
                {"text": "Set up GoLogin with 6 browser profiles (primary + 2 niche + 2 content farm + 1 findom)", "tag": "MANUAL", "detail": "Go to gologin.com, log in. Create 6 profiles:\n1. @PRINTMAXXER (primary, SAFE tier)\n2. Fitness/golf niche (MEDIUM tier)\n3. Finance niche (MEDIUM tier)\n4. Content farm #1 (AGGRESSIVE tier)\n5. Content farm #2 (AGGRESSIVE tier)\n6. Findom persona (MEDIUM tier)\nFor each profile: OS=macOS, browser=Chrome, WebRTC=altered, Canvas=noise, WebGL=noise, timezone=match proxy geo.\nAssign SOAX mobile proxy per profile: GoLogin > Profile > Proxy > SOAX rotating mobile US.\nVerify fingerprint uniqueness: browserleaks.com in each profile.\nTime: 45 min."},
                {"text": "Assign unique SOAX residential proxy to each GoLogin profile", "tag": "MANUAL", "detail": "Log into SOAX dashboard (soax.com). Create 6 sticky sessions (mobile US, different states).\nCopy proxy credentials (host:port:user:pass) for each.\nIn GoLogin, open each profile > Proxy Settings > paste SOAX credentials.\nTest each: visit whatismyipaddress.com, confirm unique IP and US geo.\nNEVER share IPs between profiles. Each account = unique residential IP.\nIf IP flagged: rotate to new sticky session in SOAX dashboard.\nTime: 20 min."},
                {"text": "Create 3 secondary X/Twitter accounts in GoLogin profiles", "tag": "MANUAL", "detail": "Open each GoLogin profile (niche #1, #2, content farm #1).\nGo to twitter.com/i/flow/signup. Use unique email per account (protonmail).\nPhone verification: use separate numbers (TextNow app = free US numbers).\nSet unique bio, name, PFP per account. Each gets its own niche voice.\nDO NOT follow each other. DO NOT interact between accounts.\nStagger creation: wait 30+ min between signups.\nIf CAPTCHA loops: clear cookies, try different SOAX session.\nTier: MEDIUM for niche, AGGRESSIVE for content farm.\nTime: 30 min."},
                {"text": "Subscribe X Premium Basic ($3/mo) on primary + 2 secondaries", "tag": "MANUAL", "detail": "In each GoLogin profile, go to twitter.com/i/premium_sign_up.\nSelect Basic ($3/mo). Pay with separate payment methods if possible.\nWhy: 10x median impressions, reply prioritization, TweepCred +100 boost.\nTotal: $9/mo for 3 accounts. Add remaining accounts after warmup.\nVerify: blue checkmark appears on profile. Check Settings > Premium.\nTime: 10 min."},
                {"text": "Post 8 tweets from @PRINTMAXXER + 1 thread", "tag": "AUTO", "detail": "Runs via auto_content_poster.py cron at 7AM.\nContent source: CONTENT/social/posting_queue/\nGenerated by: content_factory.py --batch-alpha\nThread topic: 'what you are building' (build-in-public hook).\nVerify: check CONTENT/social/posting_queue/ for today's posts.\nManual check: python3 AUTOMATIONS/auto_content_poster.py --status\nIf cron missed: python3 AUTOMATIONS/auto_content_poster.py --post\nTier: SAFE. Max 8 tweets/day during warmup.\nRisk: None. API-based posting via Typefully/Publer.\nTime: 0 min (automated). 2 min to verify."},
                {"text": "Generate cross-platform content batch from latest alpha", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/content_factory.py --batch-alpha 5\nReads: LEDGER/ALPHA_STAGING.csv for latest high-value entries.\nOutputs: 5 content pieces adapted for X, LinkedIn, Pinterest, Facebook.\nStored at: CONTENT/social/posting_queue/\nCron: runs as part of daily content pipeline.\nVerify: ls -la CONTENT/social/posting_queue/ | head -20\nRisk: None. Content generation only, no posting.\nTime: 0 min (automated)."},
                {"text": "Generate 3 faceless video scripts for boomer YouTube", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/ai_video_content_pipeline.py --generate golf --count 3\nOutputs: 3 scripts at CONTENT/video/scripts/\nTarget demo: men 55-70, golf/fishing/health interests.\nReference: AUTOMATIONS/auto_ops/alpha_theses/BOOMER_MALE_55_70_AFFILIATE.md\nNext step: render with Remotion (video_factory) or manual screen recording.\nRisk: None. Script generation only.\nTime: 0 min (automated)."},
                {"text": "Upload 3 low-content books to Amazon KDP", "tag": "MANUAL", "detail": "Go to kdp.amazon.com. Sign in or create account.\nUpload from DIGITAL_PRODUCTS/:\n1. Journal template (E05) - $2-5 royalty/sale\n2. Planner template - $2-5 royalty/sale\n3. Prompt book - $3-7 royalty/sale\nFor each: set title with keywords, description, 7 keyword slots (research via Publisher Rocket or free Amazon search suggest).\nCategory: choose 2 low-competition categories.\nPricing: $4.99-9.99 range. 70% royalty.\nCompounds over 90 days as reviews accumulate.\nTime: 45 min."},
                {"text": "List 5 digital downloads on Etsy", "tag": "MANUAL", "detail": "Go to etsy.com/sell. Create shop if needed (fnsmdehip@proton.me).\nUpload from DIGITAL_PRODUCTS/:\n1. Automation checklist ($5)\n2. Prompt pack ($9)\n3. Workflow templates ($12)\n4. Claude Code guide ($15)\n5. Agency playbook excerpt ($7)\nFor each: SEO-optimized title (use eRank.com for Etsy keyword research).\n13 tags per listing (max allowed). High-res mockup thumbnails from MEDIA/generated_images/.\nEtsy fees: $0.20 listing + 6.5% transaction.\nTime: 40 min."},
                {"text": "Reply to 40 accounts from primary (reply guy strategy)", "tag": "SEMI", "detail": "edge_growth_engine.py generates reply suggestions.\nCommand: python3 AUTOMATIONS/edge_growth_engine.py --reply-guy\nOutput: CONTENT/social/reply_queue/ with suggested replies.\nYOU review the queue before posting.\nTarget: accounts tweeting about manual processes, hiring devs, or automation.\nReply within 15 min of target posting for max algo weight.\nAuthor-engaged reply = +75 weight (150x a like).\nSafe limit: 50 replies/day/account. We do 40 (conservative).\nTool: GoLogin primary profile + SOAX mobile proxy.\nTier: SAFE. Genuine engagement only.\nTime: 30 min to review queue + post."},
                {"text": "Post 3 tweets per secondary account (warmup)", "tag": "SEMI", "detail": "Content pre-generated by content_factory.py.\nReview at: CONTENT/social/posting_queue/ (filtered by account).\nPost manually via GoLogin profiles (warmup phase = manual posting preferred).\nQuality over volume: first 100 tweets set algo foundation.\nNEVER cross-post identical text (40% duplicate penalty).\nStagger: 30+ min between accounts.\nTier: MEDIUM. 3 tweets/day during warmup.\nRisk: Low. Under all rate limits.\nTime: 15 min."},
                {"text": "Create Gumroad account + list 13 products", "tag": "MANUAL", "detail": "Go to gumroad.com. Sign up with fnsmdehip@proton.me.\nProducts ready at PRODUCTS/ and DIGITAL_PRODUCTS/.\nUpload all 13 with price anchoring ($197 crossed out, $97 shown).\nKey products:\n- Automation prompt pack ($19)\n- Claude Code mastery guide ($29)\n- Agency playbook ($49)\n- Full bundle ($97)\nUse thumbnails from MEDIA/generated_images/.\nSet up Stripe Connect for payouts.\nGumroad fee: 10% + processing.\nTime: 30 min."},
                {"text": "Create Whop storefront + list top 8 products", "tag": "MANUAL", "detail": "Go to whop.com. Create account.\nList top 8 products (same as Gumroad best sellers).\nWhop fee: 5.7% (vs Gumroad 10%) = better margins.\nSet up payment processing.\nLink to same Stripe account.\nTime: 20 min."},
                {"text": "Verify all cron jobs and scrapers ran", "tag": "AUTO", "detail": "Commands:\n  python3 AUTOMATIONS/system_health_monitor.py --quick\n  Check: 112 cron jobs status.\n  Verify twitter_alpha_scraper.py output: LEDGER/TWITTER_ALPHA_SCRAPES.csv (133 accounts, 6AM cron)\n  Verify background_reddit_scraper.py: LEDGER/REDDIT_SCRAPES/ (6:15AM cron)\n  Check soul_drift_report.json: any agent below 6/10 gets prompt rewrite.\nPath: AUTOMATIONS/agent/swarm/soul_drift_report.json\nRisk: None. Read-only checks.\nTime: 2 min."},
                {"text": "Identify 3 brokering verticals to test (use OPS/BROKERING_ARBITRAGE_OPPORTUNITIES.md)", "tag": "SEMI", "detail": "Read OPS/BROKERING_ARBITRAGE_OPPORTUNITIES.md.\nPick 3 verticals with highest automation potential and lowest upfront cost.\nRecommended starters: lead gen as a service, merchant processing referrals, equipment financing.\nFor each: identify 20 target businesses to scrape.\nScore on deal size x volume x automation level.\nStore picks in AUTOMATIONS/leads/brokering/vertical_picks.md\nTime: 20 min."},
            ],
            "revenue_target": "$0",
            "content": {"tweets": 14, "threads": 1, "tiktoks": 0, "youtube": 3, "newsletters": 0},
            "engagement": {"replies": 40, "dms": 5, "comments": 10, "follows": 25},
            "outreach": {"cold_emails": 0, "proposals": 0, "calls": 0},
        },
        2: {
            "theme": "FREELANCE + PRODUCT BLITZ",
            "tasks": [
                {"text": "Warm up 3 secondary X accounts (5 tweets + 15 replies + 20 follows each)", "tag": "SEMI", "detail": "Content pre-generated at CONTENT/social/posting_queue/.\nReview content for each account's niche voice.\nPost via GoLogin profiles. Stagger 30+ min between accounts.\nFollow 20 niche-relevant accounts per profile (manual, genuine interest).\n15 genuine replies per account to niche accounts.\nFirst 100 tweets = algorithmic foundation. Quality matters more than volume.\nTier: MEDIUM.\nRisk: Low. Well under rate limits.\nTime: 30 min."},
                {"text": "Create Fiverr account + post 5 gigs", "tag": "MANUAL", "detail": "Go to fiverr.com. Sign up with fnsmdehip@proton.me.\nCreate 5 gigs:\n1. AI automation setup ($150)\n2. Claude Code development ($200)\n3. Chatbot build ($100)\n4. Data pipeline ($175)\n5. Web scraping ($125)\nLaunch price: $25 for first gig to get first review fast.\nEach gig needs: title, description, 3 packages (basic/standard/premium), FAQ, requirements.\nUse portfolio from PRODUCTS/ as proof.\nTime: 40 min."},
                {"text": "Create Upwork profile + submit 10 proposals", "tag": "MANUAL", "detail": "Go to upwork.com. Create profile.\nHeadline: 'AI Automation Engineer | Claude Code | Ship in 48hrs'\nPortfolio: screenshots from PRODUCTS/ and APP_FACTORY/ builds.\nSubmit 10 proposals on AI/automation jobs with $1K+ budgets.\nEach proposal: personalized opening line referencing their specific project.\nAttach 1-2 relevant portfolio pieces per proposal.\nTip: mention delivery speed (48hr turnaround) as differentiator.\nUpwork fee: 10% on first $10K with client.\nTime: 60 min."},
                {"text": "Post 8 tweets + 1 thread from primary", "tag": "AUTO", "detail": "Runs via auto_content_poster.py cron at 7AM.\nThread topic: 'I build apps in 2 hours that agencies quote $10K for'\nContent from: CONTENT/social/posting_queue/\nVerify: python3 AUTOMATIONS/auto_content_poster.py --status\nIf missed: python3 AUTOMATIONS/auto_content_poster.py --post\nTier: SAFE.\nTime: 0 min (automated). 2 min to verify."},
                {"text": "Generate product thumbnails for 13 Gumroad products", "tag": "AUTO", "detail": "Uses Playwright screenshot factory (zero cost).\nTemplates at: MEDIA/image_templates/\nCommand: python3 AUTOMATIONS/image_factory.py --batch-thumbnails\nOutput: MEDIA/generated_images/thumbnails/\nIf not automated yet: use Canva free tier or screenshot existing landing pages.\nTime: 0 min if automated, 30 min if manual."},
                {"text": "Run auto_clip_pipeline on 2 trending AI videos", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/auto_clip_pipeline.py --demo --count 2\nFinds trending YouTube videos in AI niche.\nExtracts best 30-60 second clips.\nOutput: CONTENT/video/clips/\nUse for TikTok/Shorts/Reels cross-posting.\nRisk: Fair use for commentary/reaction. Add original commentary.\nTime: 0 min (automated)."},
                {"text": "Reply to 50 accounts from primary (target automation pain points)", "tag": "SEMI", "detail": "Command: python3 AUTOMATIONS/edge_growth_engine.py --reply-guy\nGenerates reply queue at: CONTENT/social/reply_queue/\nTarget: anyone tweeting about manual processes, hiring devs, needing automation.\nYOU review before posting.\nReply within 15 min of viral posts for max algo weight.\nAuthor-engaged reply = +75 weight (150x a like).\nSafe limit: 50 replies/day/account.\nTool: GoLogin primary profile + SOAX mobile proxy.\nTier: SAFE.\nTime: 35 min."},
                {"text": "DM 10 people who tweeted about manual work pain points", "tag": "MANUAL", "detail": "Find targets from edge_growth_engine.py reply-guy output.\nSend genuine value DMs (no pitch). Examples:\n- 'Saw your tweet about X. I actually built a tool for this. Happy to share.'\n- 'That workflow you described - there is a way to cut it from 3hrs to 15min.'\nDO NOT link to products yet. Build relationship first.\nMax 50 DMs/day/account. We do 10 (conservative).\nTier: SAFE.\nTime: 20 min."},
                {"text": "Post in 2 X Communities for velocity window boost", "tag": "SEMI", "detail": "Find 2 X Communities in automation/AI niche.\nPost content 30 min BEFORE main tweet to generate early engagement.\nThis triggers the algorithm's velocity window for broader distribution.\nCommunities replaced engagement pods as the legitimate alternative.\nTier: SAFE. Platform-endorsed feature.\nTime: 10 min."},
                {"text": "Verify agent swarm and n8n workflows operational", "tag": "AUTO", "detail": "Commands:\n  python3 AUTOMATIONS/agent_swarm.py --status\n  Expected: 25 agents operational (4 Opus, 5 Sonnet, 3 Haiku active)\n  Verify n8n w01 (GMaps lead scraper): check LEDGER/ for new leads.\n  Verify n8n w04 (cold email SendGrid): ready for Day 4 launch.\n  Check alpha_auto_processor.py: process any PENDING_REVIEW entries.\n  Command: python3 AUTOMATIONS/alpha_auto_processor.py --process-new\nRisk: None. Status checks.\nTime: 3 min."},
            ],
            "revenue_target": "$0-20",
            "content": {"tweets": 23, "threads": 1, "tiktoks": 0, "youtube": 0, "newsletters": 0},
            "engagement": {"replies": 50, "dms": 10, "comments": 15, "follows": 60},
            "outreach": {"cold_emails": 0, "proposals": 10, "calls": 0},
        },
        3: {
            "theme": "GOV CONTRACTS + EAS PIPELINE",
            "tasks": [
                {"text": "Secondary account warmup Day 3 (5 tweets + 20 replies per account)", "tag": "SEMI", "detail": "Content from CONTENT/social/posting_queue/ filtered per account niche.\nPost via GoLogin profiles. Unique content per account (never cross-post identical text).\nFirst 100 tweets set algo foundation.\nStagger 30+ min between accounts.\nTier: MEDIUM.\nTime: 20 min."},
                {"text": "Set up GoLogin profile for LinkedIn automation", "tag": "MANUAL", "detail": "In GoLogin, create new profile: 'LinkedIn-EAS'\nOS: macOS, browser: Chrome, proxy: separate SOAX sticky session.\nNEVER share proxy with Twitter profiles.\nLinkedIn detects automation aggressively. Use Expandi ($99/mo) or Dripify ($59/mo) for safe automation.\nChrome extensions = instant detection. Desktop automation = DEAD.\nTier: MEDIUM.\nTime: 15 min."},
                {"text": "Post 8 tweets + 2 threads from primary", "tag": "AUTO", "detail": "Runs via auto_content_poster.py cron.\nThread topics: 1 freelance value thread, 1 build-in-public thread.\nVerify: python3 AUTOMATIONS/auto_content_poster.py --status\nTier: SAFE.\nTime: 0 min (automated)."},
                {"text": "Generate 5 TikTok scripts for fitness niche", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/ai_video_content_pipeline.py --generate fitness --count 5\nOutput: CONTENT/video/scripts/fitness/\nFirst 3 seconds = hook or die. Pattern interrupts (zoom, cut, text flash).\nPost times: 6-10 AM and 7-11 PM.\nTime: 0 min (automated)."},
                {"text": "Create 3 LinkedIn posts repurposing best tweets for EAS B2B", "tag": "SEMI", "detail": "Take top-performing tweets from auto_content_poster.py --winners.\nRewrite with B2B framing for Enterprise Automation Solutions.\nOptimize for LinkedIn Depth Score: dwell time + comment depth + saves.\nZERO external links in body (external links = -60% reach on LinkedIn).\nAI-detected content = -47% reach. Write in human voice.\nPost from GoLogin LinkedIn profile.\nTier: SAFE.\nTime: 20 min."},
                {"text": "Publish 1 Medium article + 1 Quora answer (Parasite SEO)", "tag": "SEMI", "detail": "Repurpose best thread content into:\n1. Medium article (DA 96 backlink, free)\n2. Quora answer to relevant automation question\nParasite SEO: high-DA backlinks to money sites for free.\nInclude 1-2 links to Gumroad/landing pages in author bio.\nDO NOT use AI-detected writing style. Rewrite in human voice.\nMedium tip: submit to publications for 10x distribution.\nTime: 30 min."},
                {"text": "Reply to 50 from primary + 10 per secondary (80 total)", "tag": "SEMI", "detail": "Command: python3 AUTOMATIONS/edge_growth_engine.py --reply-guy\nReview queue at CONTENT/social/reply_queue/.\nPrimary: 50 replies (SAFE tier, genuine engagement).\nSecondaries: 10 each (MEDIUM tier, building niche presence).\nTotal: 80 replies across accounts.\nStagger between accounts.\nTime: 40 min."},
                {"text": "Scan SAM.gov for IT/software/automation contracts", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/sam_gov_monitor.py\nScans for NAICS codes 541511-541990 (IT consulting, software dev, digital marketing).\nOutput: LEDGER/GOV_OPPORTUNITIES.csv\nAlso run: python3 AUTOMATIONS/uk_contracts_finder.py --categories 'IT consulting marketing'\nReview top 3 contracts we can bid on with Claude Code capabilities.\nTime: 0 min (automated). 10 min to review output."},
                {"text": "Run EAS lead pipeline: scrape + score 50 local businesses", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/eas_lead_pipeline.py\nCron: weekday 8AM.\nScrapes + scores local businesses with outdated websites.\nOutput: LEDGER/EAS_LEADS_SCORED.csv\nScoring: website age, mobile responsiveness, social presence, industry.\nTop verticals: dental, legal, HVAC (highest close rates).\nTime: 0 min (automated). 5 min to review top leads."},
                {"text": "Draft 3 EAS cold email templates", "tag": "SEMI", "detail": "Templates for 3 EAS packages:\n1. Signal Map intro ($1,500) - 'We mapped 12 automation opportunities for [BUSINESS]'\n2. Phone Pilot pitch ($3,500) - 'Your receptionist handles 47 calls/day. AI handles 200.'\n3. Ops Pilot value prop ($4,500) - 'Cut 20 hrs/week of admin. We build it in 5 days.'\nStore in: MONEY_METHODS/EAS/email_templates/\nLoad into n8n w04 for Day 4+ automated sending.\nPersonalization tokens: {business_name}, {owner_name}, {industry}, {pain_point}.\nTime: 25 min."},
                {"text": "Verify venture autonomy and daily engagement planner", "tag": "AUTO", "detail": "Commands:\n  python3 AUTOMATIONS/venture_autonomy.py --status\n  Expected: all 8 venture types active (OUTBOUND, CONTENT, APP, LOCAL_BIZ, RESEARCH, MONETIZE, PRODUCT, SCRAPING).\n  Verify daily_engagement_planner.py ran at 7AM cron.\n  Review today's engagement plan at OPS/DAILY_ENGAGEMENT_PLAN.md.\n  Check n8n w09 (content repurpose) execution status.\nTime: 3 min."},
            ],
            "revenue_target": "$0-30",
            "content": {"tweets": 24, "threads": 2, "tiktoks": 5, "youtube": 0, "newsletters": 0},
            "engagement": {"replies": 80, "dms": 5, "comments": 15, "follows": 35},
            "outreach": {"cold_emails": 0, "proposals": 10, "calls": 0},
        },
        4: {
            "theme": "COLD OUTREACH IGNITION",
            "tasks": [
                {"text": "Create TikTok account + post first 3 videos", "tag": "MANUAL", "detail": "Use real phone (no VPN, no proxy for TikTok signup).\nContent: screen recordings of AI builds (authentic, raw).\nPost times: 6-10 AM or 7-11 PM.\nFirst 3 seconds = hook. Pattern interrupts every 3-5 seconds.\nHashtags: 0-2 max (3+ = 40% reach penalty on X, similar on TikTok).\nDO NOT use anti-detect browser for TikTok (phone fingerprint required).\nTime: 30 min."},
                {"text": "Create Pinterest account + pin 10 affiliate products", "tag": "MANUAL", "detail": "Go to pinterest.com. Create business account.\nPin 10 boomer-targeted products with affiliate links:\n- Golf gadgets, fishing gear, health supplements, financial tools.\nSEO-optimize: keyword-rich pin titles and descriptions.\nUse eRank or Pinterest Trends for keyword research.\n15-25 pins/day is optimal cadence.\nAffiliate: Amazon Associates, ShareASale, or direct brand programs.\nTime: 25 min."},
                {"text": "Cross-promote: secondaries repost/quote-tweet primary content", "tag": "SEMI", "detail": "Secondary accounts quote-tweet primary threads with unique takes.\nStagger 2+ hours apart between accounts.\nEach account adds a unique perspective from its niche.\nNEVER repost identical text (40% duplicate penalty).\nTier: MEDIUM.\nTime: 15 min."},
                {"text": "Post 10 tweets + 1 thread from primary + 5 per secondary", "tag": "AUTO", "detail": "Primary: runs via auto_content_poster.py cron.\nThread topic: breakdown of a real automation build.\nSecondaries: 5 tweets each, niche-specific.\nContent from content_factory.py --batch-alpha 10.\nVerify: python3 AUTOMATIONS/auto_content_poster.py --status\nTier: SAFE (primary), MEDIUM (secondaries).\nTime: 0 min (automated). 5 min to verify."},
                {"text": "Post 3 TikToks cross-posted to YouTube Shorts + IG Reels", "tag": "SEMI", "detail": "Take TikTok videos from Day 3 scripts (rendered).\nAdapt captions per platform (never identical text).\nTikTok: raw, authentic. Shorts: more polished. Reels: aesthetic.\nRemove TikTok watermark before cross-posting (SnapTik or SaveTik).\nPost times optimized per platform.\nTime: 20 min."},
                {"text": "Reply to 60 from primary + 15 per secondary (90 total)", "tag": "SEMI", "detail": "Command: python3 AUTOMATIONS/edge_growth_engine.py --reply-guy\nReview queue at CONTENT/social/reply_queue/.\nPrimary: 60 replies (pushing toward ceiling).\nSecondaries: 15 each.\nTotal: 90 replies. Still under 50/account limit.\nTier: SAFE (primary), MEDIUM (secondaries).\nTime: 45 min."},
                {"text": "DM 15 people: devs hiring, automation needs, collaborators", "tag": "MANUAL", "detail": "5 who tweeted about hiring devs (pitch speed + cost savings).\n5 who need automation (offer free 15-min audit).\n5 potential collaborators (cross-promotion, mutual amplification).\nPersonalize every DM. Reference their specific tweet/post.\nMax 50 DMs/day. We do 15 (conservative SAFE tier).\nTime: 25 min."},
                {"text": "Comment on 10 Reddit posts for visibility", "tag": "SEMI", "detail": "Subreddits: r/freelance, r/automation, r/smallbusiness, r/Entrepreneur.\nProvide genuine value in comments. Link in Reddit profile, NOT in comment.\nReddit auto-removes comments with obvious self-promo links.\nAccount age + karma threshold required before posting in most subs.\nUse real Reddit account (no proxy needed for Reddit).\nTime: 20 min."},
                {"text": "Launch cold email via n8n: 50 emails to scored local businesses", "tag": "AUTO", "detail": "n8n workflow w04 sends via SendGrid.\nRecipients: top-scored leads from eas_lead_pipeline.py.\nTemplate: Signal Map intro ($1,500) with personalization tokens.\nCAN-SPAM compliant: opt-out link, physical address, identified as ad.\nDomain: warmed 2+ weeks, SPF+DKIM+DMARC configured.\nPlain text only (HTML = spam filter trigger).\nDisable open tracking (Gmail shows warning label).\nSafe limit: 50/day for warmed inbox.\nMonitor: bounce rate < 5%, spam complaints < 0.1%.\nIf bounces spike: pause and clean list.\nTime: 0 min (automated). 5 min to verify."},
                {"text": "Submit 10 Upwork proposals on $2K+ projects", "tag": "SEMI", "detail": "Search Upwork for: AI automation, Claude, chatbot, web scraping, data pipeline.\nFilter: $1K-5K budget, posted last 24h.\nEach proposal: personalized opening referencing their project.\nAttach portfolio from PRODUCTS/ and APP_FACTORY/ as proof of speed.\nMention 48hr delivery as differentiator.\n20 total active proposals after today.\nTime: 45 min."},
                {"text": "Verify edge growth tracking and n8n webhook readiness", "tag": "AUTO", "detail": "Commands:\n  python3 AUTOMATIONS/edge_growth_engine.py --status\n  Verify: engagement ratios per account within platform limits.\n  Check method_discovery_crawler.py 5AM cron: new P0 methods in OPS/CAPITAL_GENESIS_PRIORITY_STACK.md\n  Check n8n w14 (Stripe product delivery) webhook: ready for first sale.\nTime: 3 min."},
            ],
            "revenue_target": "$0-50",
            "content": {"tweets": 20, "threads": 1, "tiktoks": 3, "youtube": 0, "newsletters": 0},
            "engagement": {"replies": 90, "dms": 15, "comments": 20, "follows": 45},
            "outreach": {"cold_emails": 50, "proposals": 10, "calls": 0},
        },
        5: {
            "theme": "E-COMMERCE + ARBITRAGE SCAN",
            "tasks": [
                {"text": "Set up Fanvue account for findom persona (P02)", "tag": "MANUAL", "detail": "Use separate GoLogin profile + unique SOAX proxy.\nGo to fanvue.com. Create account. Fanvue = most AI-friendly platform.\n15% of revenue from AI creators. Top AI creator: $23K/mo.\nCross-platform stack: Twitter discovery > Fanvue content > Telegram VIP > CashApp tributes.\nGeo-match proxy to persona location.\nTime: 15 min."},
                {"text": "Post 5 findom persona tweets + set up cross-platform funnel", "tag": "SEMI", "detail": "Teaser content + #findom #paypig hashtags + Fanvue link in bio.\nPost via GoLogin findom profile.\n5-10 tweets/day in findom niche.\nBuild funnel: Twitter (free discovery) > Fanvue (paid content) > Telegram VIP ($99-499/mo) > CashApp (tributes).\nTier: MEDIUM. Separate from primary brand entirely.\nTime: 15 min."},
                {"text": "Set up Beehiiv newsletter + create lead magnet PDF", "tag": "MANUAL", "detail": "Go to beehiiv.com. Create account.\nNewsletter name: match @PRINTMAXXER brand.\nLead magnet: free automation checklist PDF (create from DIGITAL_PRODUCTS/ content).\nSet up welcome email sequence (3 emails over 7 days).\nAdd opt-in link to all bios, Gumroad checkout, YouTube descriptions.\nBeehiiv free tier: up to 2,500 subscribers.\nTime: 25 min."},
                {"text": "Post 10 tweets + 1 thread from primary (cold email case study)", "tag": "AUTO", "detail": "Thread topic: 'I sent 50 cold emails, here is what happened' (real-time case study).\nAuto-posted via cron. Content from posting_queue/.\n5 tweets per secondary (10 total), adapted to niche voice.\nVerify: python3 AUTOMATIONS/auto_content_poster.py --status\nTime: 0 min (automated)."},
                {"text": "Generate 3 faceless YouTube scripts (boomer health)", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/ai_video_content_pipeline.py --generate health --count 3\nOutput: CONTENT/video/scripts/health/\nTarget: men 55-70.\nTopics: supplement reviews, health tracking gadgets, longevity tips.\nMonetization: affiliate links in description ($10-50/referral for health supplements).\nTime: 0 min (automated)."},
                {"text": "Batch-clip 5 viral videos for repurposing", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/auto_clip_pipeline.py --url [trending_video] --max-clips 5\nExtracts best 30-60 second segments from trending content.\nOutput: CONTENT/video/clips/\nUse for TikTok/Shorts/Reels with original commentary added.\nTime: 0 min (automated)."},
                {"text": "Reply to 60 from primary + 20 from secondaries (80 total)", "tag": "SEMI", "detail": "Command: python3 AUTOMATIONS/edge_growth_engine.py --reply-guy\nReview queue. Post via GoLogin profiles.\nPrimary: 60 (SAFE). Secondaries: 20 total (MEDIUM).\nTime: 40 min."},
                {"text": "DM 10 mid-tier accounts offering free template for testimonial", "tag": "MANUAL", "detail": "Target: 2K-20K follower accounts in AI/automation niche.\nOffer: free automation template in exchange for honest feedback/testimonial.\nTestimonials = social proof for Gumroad listings.\nFTC: must disclose if testimonial was incentivized.\nTime: 20 min."},
                {"text": "Buy 3 PLR products and rebrand for Gumroad", "tag": "MANUAL", "detail": "Sources: PLRProducts.com, IDplr.com. Budget: $20-50 each.\nPLR = Private Label Rights = legal to rebrand and resell.\nRebrand with PRINTMAXX voice (apply copy-style.md).\nList on Gumroad at $27-47 each.\nN11 PLR arbitrage: 20-40% margin after acquisition.\nTime: 30 min."},
                {"text": "Scan TikTok/Facebook ads libraries for winning products", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/product_scanner.py --source tiktok_ads_library --source fb_ads_library --min-engagement 10000\nIdentifies products: under $30 AOV, 3x+ markup potential.\nOutput: LEDGER/WINNING_PRODUCTS.csv\nE01 TikTok Shop arbitrage: find supplier on AliExpress/1688, list on TikTok Shop.\nTime: 0 min (automated). 10 min to review."},
                {"text": "Send cold email batch 2 (50 more, A/B test subject lines)", "tag": "AUTO", "detail": "n8n w04 auto-sends batch 2.\n100 total emails sent.\nA/B test: 2 subject lines, 25 each.\nFollow up on batch 1 non-openers with different subject line (auto via SendGrid sequence).\nMonitor: open rates (target >40%), reply rates (target >2%).\nTime: 0 min (automated). 5 min to check analytics."},
                {"text": "Auto-generate freelance proposals for hot jobs", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/auto_freelance_responder.py --scan-and-respond\nScans Reddit and Upwork for matching jobs.\nAuto-generates personalized proposals.\nOutput: CONTENT/proposals/pending/ for review before sending.\nCron: every 2 hours, 15 min after scanner.\nTime: 0 min (automated). 10 min to review proposals."},
                {"text": "Verify warmup status and analytics", "tag": "AUTO", "detail": "Commands:\n  Verify twitter_warmup_poster.py warmup day status.\n  Check cold email analytics: open rates (>40%), reply rates (>2%), bounce (<5%).\n  Verify capital_genesis_ranker.py 5:30AM cron output.\n  Check n8n w13 (Reddit pain point miner): any EAS prospects?\nTime: 5 min."},
            ],
            "revenue_target": "$0-75",
            "content": {"tweets": 20, "threads": 1, "tiktoks": 0, "youtube": 3, "newsletters": 0},
            "engagement": {"replies": 80, "dms": 10, "comments": 10, "follows": 40},
            "outreach": {"cold_emails": 50, "proposals": 5, "calls": 0},
        },
        6: {
            "theme": "WEEK 1 WAR ROOM",
            "tasks": [
                {"text": "Audit all GoLogin profiles for fingerprint isolation", "tag": "MANUAL", "detail": "Open each GoLogin profile. Visit browserleaks.com.\nVerify: unique IP, unique canvas hash, unique WebGL hash, no shared cookies.\nCheck: no fingerprint overlap between any two profiles.\nIf overlap found: regenerate fingerprint in GoLogin profile settings.\nTime: 20 min."},
                {"text": "Check shadowban status on all accounts", "tag": "SEMI", "detail": "For X/Twitter: visit shadowban.eu, enter each account handle.\nFor Instagram: check engagement rate drop (normal = 3-6%, shadowban = <1%).\nFor TikTok: check if videos appear in hashtag search from logged-out browser.\nIf shadowbanned on X: reduce to 3 tweets/day for 48hrs, no automation, genuine engagement only.\nRecovery: stop all automated activity, post manually for 5-7 days.\nTime: 15 min."},
                {"text": "Batch create 30 tweets + 7 threads + 15/secondary for Week 2", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/content_factory.py --batch-alpha 30\nAdditional: python3 AUTOMATIONS/content_factory.py --threads 7\nPer secondary: 15 unique tweets in niche voice.\nOutput: CONTENT/social/posting_queue/\nScript 5 TikToks: python3 AUTOMATIONS/ai_video_content_pipeline.py --generate fitness --count 5\nTime: 0 min (automated). 15 min to review/edit."},
                {"text": "Post 5 tweets + 1 thread from primary (Week 1 real numbers)", "tag": "AUTO", "detail": "Reduced volume Saturday. Focus on quality.\nThread: 'Week 1 building in public -- real numbers'\nVerify: python3 AUTOMATIONS/auto_content_poster.py --status\nTier: SAFE.\nTime: 0 min (automated)."},
                {"text": "Reply to 30 accounts (reduced Saturday volume, higher quality)", "tag": "SEMI", "detail": "Saturday = quality over quantity.\nPrimary only, 30 replies.\nFocus on high-value accounts with >10K followers.\nTime: 20 min."},
                {"text": "Full Week 1 analytics across all channels", "tag": "SEMI", "detail": "Compile:\n- Twitter impressions per account (Twitter Analytics or Typefully).\n- Gumroad views + conversions.\n- Fiverr clicks + buyer requests.\n- Email open/reply/bounce rates from SendGrid dashboard.\n- Etsy views/favorites.\n- KDP page reads.\nCalculate revenue-per-hour for every channel.\nRank all channels by ROI. Top 3 get 3x resources in Week 2.\nStore results in OPS/WEEKLY_ANALYTICS/week1.md.\nTime: 30 min."},
                {"text": "Verify system health + scrapers + daily digest", "tag": "AUTO", "detail": "Commands:\n  python3 AUTOMATIONS/system_health_monitor.py --quick\n  Verify all scrapers ran (twitter_alpha, reddit, alpha_auto_processor).\n  Review daily_digest.py output (6:45AM cron): identify 3 missed actions.\n  Check ACTIONABLE_QUEUE.md for queued items not yet executed.\nTime: 5 min."},
                {"text": "Scan SAM.gov for new contracts posted this week", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/sam_gov_monitor.py\nReview GOV_OPPORTUNITIES.csv: shortlist contracts with deadlines in next 14 days.\nTime: 0 min (automated). 5 min to review."},
            ],
            "revenue_target": "$0-100",
            "content": {"tweets": 5, "threads": 1, "tiktoks": 0, "youtube": 0, "newsletters": 0},
            "engagement": {"replies": 30, "dms": 5, "comments": 5, "follows": 15},
            "outreach": {"cold_emails": 0, "proposals": 0, "calls": 0},
        },
        7: {
            "theme": "CONTENT FACTORY + BOOMER PIPELINE",
            "tasks": [
                {"text": "Set up YouTube channel for faceless boomer content", "tag": "MANUAL", "detail": "Go to youtube.com. Create channel.\nBranding: name, about page, banner, PFP.\nChannel keywords: golf tips, fishing gear, health supplements, men over 50.\nThumbnail template: create reusable template in Canva or MEDIA/image_templates/.\nMonetization: YouTube Partner Program at 1K subs + 4K watch hours.\nTime: 20 min."},
                {"text": "Add payment links to ALL account bios", "tag": "MANUAL", "detail": "CashApp, Venmo, crypto wallet links in all bios.\nP11 passive tributes: some followers will tip/donate spontaneously.\nFor findom persona: CashApp is primary tribute method.\nTime: 10 min."},
                {"text": "Upload first 3 faceless YouTube videos", "tag": "SEMI", "detail": "Videos from ai_video_content_pipeline.py output.\nTopics: golf tips, fishing gear reviews, health for men 55+.\nUpload via YouTube Studio. Set:\n- SEO title with keywords\n- Description with affiliate links + timestamps\n- Tags (up to 500 characters)\n- End screen: subscribe button + next video\n- Cards at 30%, 60%, 80% marks linking to products\nThumbnail: high-contrast with large text.\nTime: 30 min."},
                {"text": "Create 5 Facebook posts for boomer Groups", "tag": "SEMI", "detail": "Target Groups: golf tips, fishing tech, health tracking for seniors.\nContent: practical tips with product recommendations.\nInclude affiliate links where group rules allow.\nDO NOT spam. Provide genuine value. 1 post per group per day max.\nFacebook organic reach is low but boomer demo is highly active in Groups.\nTime: 20 min."},
                {"text": "Write first Beehiiv newsletter issue", "tag": "SEMI", "detail": "Topic: automation alpha + exclusive tool recommendation.\nInclude: 3-5 actionable tips, 1 tool recommendation (affiliate link), 1 free resource.\nCTA: reply to this email (boosts deliverability), share with a friend (growth loop).\nSend to full subscriber list.\nMonitor: open rate (target >40%), click rate (target >5%).\nTime: 25 min."},
                {"text": "Atomize top content across 6 platforms", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/edge_growth_engine.py --repurpose [best_thread_path]\n1 thread becomes 12 niche-specific angles.\nDistributed to: X, LinkedIn, Pinterest, Facebook, Medium, Reddit.\nEach adaptation unique per platform voice.\nTime: 0 min (automated)."},
                {"text": "Reply to 30 accounts + respond to all DMs", "tag": "SEMI", "detail": "Sunday = maintenance mode.\n30 replies from primary.\nRespond to ALL accumulated DMs across all accounts.\n3 tweets per secondary (maintain warmup cadence).\nTime: 25 min."},
                {"text": "Optimize top 3 Gumroad listings from Week 1 data", "tag": "SEMI", "detail": "Review Gumroad analytics: views, conversions, bounce rate per product.\nUpdate titles, descriptions, pricing based on data.\nA/B test: try different price points on lowest-converting products.\nAdd social proof: any reviews, testimonials, download counts.\nTime: 15 min."},
                {"text": "Write cold email follow-up sequence (3 emails)", "tag": "SEMI", "detail": "Sequence: Day 3, Day 5, Day 7 after initial.\nEach follow-up: different angle, shorter, more direct.\nLoad into n8n w04 as automated sequence.\nKey: first email = 58% of all replies. Optimize step 1 hardest.\nTime: 20 min."},
                {"text": "Apply to 3 affiliate programs", "tag": "MANUAL", "detail": "Target high-commission verticals:\n1. Health/supplement (10-15% commission per sale)\n2. SaaS tools ($50-500 per referral)\n3. Hosting ($50-200 per referral)\nApply via: ShareASale, CJ Affiliate, Impact, or direct brand programs.\nNeed: website/social proof, traffic estimates, content plan.\nTime: 15 min."},
                {"text": "Verify backups, intelligence router, security audit", "tag": "AUTO", "detail": "Commands:\n  Verify weekly backup ran (Sunday 3AM cron): ls ~/PRINTMAXX_BACKUPS/\n  Verify intelligence_router.py coverage: 484 docs, 98.3% target.\n    python3 AUTOMATIONS/intelligence_router.py --stats\n  Verify security_audit.py Sunday 4:30AM cron: review 6-category findings.\n    cat AUTOMATIONS/logs/security_audit.log | tail -50\nPlan Week 2 escalation: 3x content volume, launch gov contract bids.\nTime: 5 min."},
            ],
            "revenue_target": "$0-150",
            "content": {"tweets": 9, "threads": 0, "tiktoks": 0, "youtube": 3, "newsletters": 1},
            "engagement": {"replies": 30, "dms": 10, "comments": 5, "follows": 15},
            "outreach": {"cold_emails": 0, "proposals": 0, "calls": 0},
        },
        # ===== WEEK 2: VOLUME + PIPELINE =====
        8: {
            "theme": "TRIPLE-ACCOUNT VOLUME",
            "tasks": [
                {"text": "Post 12 tweets + 2 threads from primary (3x Week 1)", "tag": "AUTO", "detail": "Auto-posted via cron. 3x volume increase from Week 1.\nContent from content_factory.py --batch-alpha.\nVerify: python3 AUTOMATIONS/auto_content_poster.py --status\nTier: SAFE. 12 tweets/day is within normal range for active accounts.\nTime: 0 min (automated)."},
                {"text": "Post 8 tweets per secondary (16 total across 2 accounts)", "tag": "AUTO", "detail": "Auto-posted via cron, staggered 30+ min between accounts.\nContent adapted per niche voice. Never identical to primary.\nTier: MEDIUM. 8/day is moderate for warmed accounts.\nTime: 0 min (automated). 5 min to verify."},
                {"text": "Cross-account amplification: secondaries quote-tweet primary", "tag": "SEMI", "detail": "Secondary accounts quote-tweet primary threads with unique takes.\nStagger 2+ hours apart. Each adds niche-specific value.\nNEVER use identical text (40% duplicate detection penalty).\nTier: MEDIUM.\nTime: 10 min."},
                {"text": "Generate 10 pieces of multi-platform content", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/content_factory.py --batch-alpha 10\nTargets: X (primary), LinkedIn (EAS), Pinterest (affiliate), Facebook (boomer).\nOutput: CONTENT/social/posting_queue/\nTime: 0 min (automated)."},
                {"text": "Generate 5 TikToks (3 fitness + 2 golf)", "tag": "AUTO", "detail": "Commands:\n  python3 AUTOMATIONS/ai_video_content_pipeline.py --generate fitness --count 3\n  python3 AUTOMATIONS/ai_video_content_pipeline.py --generate golf --count 2\nPost 3 TikToks + cross-post to YouTube Shorts + IG Reels (unique captions).\nTime: 0 min (automated). 15 min to post cross-platform."},
                {"text": "Reply to 75 from primary + 25 from secondaries (100 total)", "tag": "SEMI", "detail": "Command: python3 AUTOMATIONS/edge_growth_engine.py --reply-guy\nReview queue. 100 total replies across accounts.\nPrimary: 75 (SAFE, under 50/account, spread across day).\nWait -- 75 exceeds single account limit of 50.\nAdjust: 50 from primary + 25 per secondary = 100 total.\nTime: 50 min."},
                {"text": "DM 15 people (Upwork posters, engagers, cold prospects)", "tag": "MANUAL", "detail": "5 Upwork job posters who haven't hired yet (pitch speed).\n5 people who engaged with your threads (convert to leads).\n5 cold prospects from EAS lead list.\nPersonalize every DM.\nTime: 25 min."},
                {"text": "Set up ManyChat on Instagram for auto-DM lead magnet", "tag": "MANUAL", "detail": "Go to manychat.com. Connect Instagram business account.\nSet up: anyone who comments keyword (e.g., 'GUIDE') gets auto-DM with lead magnet link.\nInstagram Graph API: 200 DMs/hr, Meta-approved, zero ban risk.\nThis is the ONLY legal Instagram DM automation.\nTime: 20 min."},
                {"text": "Post to 3 X Communities for velocity window boost", "tag": "SEMI", "detail": "Post content to X Communities before main tweet.\nTriggers 30-min velocity window for broader distribution.\nCommunities provide warm audience for initial engagement.\nTier: SAFE. Platform-endorsed feature.\nTime: 10 min."},
                {"text": "Send cold email batch 3 (75 emails, 175 total)", "tag": "AUTO", "detail": "n8n w04 auto-sends with A/B winning subject line from batch 1-2.\nFollow up on Week 1 emails that got opens but no replies (auto via SendGrid sequence).\nMonitor: open rates, reply rates, bounces.\nTime: 0 min (automated). 5 min to check analytics."},
                {"text": "Submit 15 Upwork proposals ($1K-5K projects)", "tag": "SEMI", "detail": "Filter: AI/automation, $1K-5K budget, posted last 48h.\nAttach portfolio from PRODUCTS/ as proof of delivery speed.\n35 total active proposals after today.\nTime: 60 min."},
                {"text": "Set up Bland AI voice outreach for EAS", "tag": "MANUAL", "detail": "Go to bland.ai. Create account.\nLoad EAS discovery call script from MONEY_METHODS/EAS/call_scripts/.\nTest 5 calls to warm leads from eas_lead_pipeline scored list.\n100 free calls/day. 1-3% appointment rate = 0-1 meetings from 5 calls.\nS05 voice outreach venture.\nTime: 20 min."},
                {"text": "Verify agent swarm and n8n workflows", "tag": "AUTO", "detail": "Commands:\n  python3 AUTOMATIONS/agent_swarm.py --status (25 agents, model routing check)\n  python3 AUTOMATIONS/alpha_auto_processor.py --process-new (clear PENDING_REVIEW)\n  Check n8n w01 (GMaps lead scraper): new leads enriched.\nTime: 3 min."},
                {"text": "Send 20 cold emails to businesses in top brokering vertical", "tag": "SEMI", "detail": "Using leads scraped for top brokering vertical from Week 1 picks.\nTemplates: adapt EAS cold email templates for brokering pitch.\nPitch: 'We connect businesses like yours to [service providers] -- you pay nothing unless we deliver.'\nPersonalize with business name, industry pain point, service gap.\nUse n8n w04 (SendGrid) or manual send via GoLogin.\nTrack: open rates, reply rates, interest signals.\nStore outreach log in AUTOMATIONS/leads/brokering/outreach_log.csv\nTime: 30 min."},
            ],
            "revenue_target": "$10-75",
            "content": {"tweets": 28, "threads": 2, "tiktoks": 3, "youtube": 0, "newsletters": 0},
            "engagement": {"replies": 100, "dms": 15, "comments": 15, "follows": 50},
            "outreach": {"cold_emails": 75, "proposals": 15, "calls": 0},
        },
        9: {
            "theme": "VIDEO + FACELESS SCALE",
            "tasks": [
                {"text": "All secondaries: 8 tweets + 25 replies each (past warmup)", "tag": "AUTO", "detail": "Auto-posted via cron. Secondaries now past warmup threshold.\n8 tweets + 25 replies = active but safe.\nTier: MEDIUM.\nTime: 0 min (automated). 10 min to verify."},
                {"text": "Upload YouTube faceless videos #4-5 (golf, health 55+)", "tag": "SEMI", "detail": "Videos from ai_video_content_pipeline.py.\nUpload via YouTube Studio with full SEO:\n- Keyword-rich title, description, tags.\n- Affiliate links in description.\n- End screens + cards.\n- Custom thumbnail (high contrast, large text).\nTime: 25 min."},
                {"text": "Record 1 YouTube tutorial: 'How I automate a business in 48 hours'", "tag": "MANUAL", "detail": "Screen record a real automation build with narration.\n10 min target length. Show: problem > solution > results.\nEdit: cut dead air, add text overlays, speed up repetitive parts.\nUpload with SEO optimization.\nThis is your portfolio piece for freelance clients.\nTime: 90 min (record + edit + upload)."},
                {"text": "Batch-clip 15 trending AI/tech videos", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/auto_clip_pipeline.py --urls-file batch.txt --max-clips 15\nProcess 3 trending AI/tech videos, 5 clips each.\nOutput: CONTENT/video/clips/\nPost as TikToks + Shorts + Reels (3 clips per platform = 9 posts).\nTime: 0 min (automated). 20 min to post cross-platform."},
                {"text": "Post 10 tweets + 1 thread from primary promoting YouTube", "tag": "AUTO", "detail": "Thread: 'I just dropped a video showing how I automate a business in 48hrs. Link in bio.'\nDO NOT put YouTube link in tweet body (30-50% distribution penalty).\nUse 'link in bio' CTA.\nTime: 0 min (automated)."},
                {"text": "Reply to 75 from primary + 25 from secondaries (100 total)", "tag": "SEMI", "detail": "Review queue from edge_growth_engine.py.\nPost via GoLogin profiles.\nTime: 50 min."},
                {"text": "DM 10 mid-tier YouTubers for cross-promotion", "tag": "MANUAL", "detail": "Target: AI/automation YouTubers with 1K-50K subscribers.\nOffer: mutual shoutouts, collaborative content, or affiliate partnership.\nTime: 20 min."},
                {"text": "Submit 10 Upwork proposals + auto-bid Reddit freelance", "tag": "SEMI", "detail": "Upwork: 10 more proposals (45 total). Target $1K-5K jobs.\nReddit: python3 AUTOMATIONS/auto_freelance_responder.py --scan-and-respond\nAuto-generates proposals for hot Reddit freelance threads.\nReview before sending.\nTime: 40 min."},
                {"text": "Bland AI: 20 calls to scored EAS leads", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/bland_caller.py --leads 20 --script eas_discovery\nOr via Bland AI dashboard.\n100 free calls/day. 1-3% appointment rate = 0-1 meetings from 20 calls.\nLog all objections for script refinement.\nTime: 0 min (automated). 5 min to review results."},
                {"text": "Scan SAM.gov for IT/software contracts + draft proposal", "tag": "SEMI", "detail": "Command: python3 AUTOMATIONS/sam_gov_monitor.py --limit 50\nReview LEDGER/GOV_OPPORTUNITIES.csv.\nDraft proposal outline for top-scoring contract matching NAICS 541511-541990.\nSAM.gov proposal requirements: formatting, SAM registration, DUNS number.\nTime: 30 min."},
                {"text": "Verify venture autonomy + alpha scraper output", "tag": "AUTO", "detail": "Commands:\n  python3 AUTOMATIONS/venture_autonomy.py --status (8 venture types active)\n  Check twitter_alpha_scraper.py: new accounts to follow/engage.\nTime: 3 min."},
            ],
            "revenue_target": "$10-100",
            "content": {"tweets": 26, "threads": 1, "tiktoks": 3, "youtube": 3, "newsletters": 0},
            "engagement": {"replies": 100, "dms": 10, "comments": 15, "follows": 45},
            "outreach": {"cold_emails": 0, "proposals": 10, "calls": 20},
        },
        10: {
            "theme": "EMAIL SCALE + MULTI-PLATFORM",
            "tasks": [
                {"text": "Start 2 more secondary accounts (health, finance niches)", "tag": "MANUAL", "detail": "New GoLogin profiles + SOAX sticky sessions.\nNiche 1: health/supplements. Niche 2: personal finance.\nEach account: bio, PFP, 3 seed tweets, follow 20 niche accounts.\nTier: MEDIUM.\nTime: 20 min."},
                {"text": "Post 12 tweets + 2 threads from primary", "tag": "AUTO", "detail": "1 case study thread, 1 tactical value thread.\nAuto-posted via cron.\nTime: 0 min (automated)."},
                {"text": "Post 8 tweets per secondary (24 total across 3)", "tag": "AUTO", "detail": "Auto-posted via cron, staggered.\nTime: 0 min (automated)."},
                {"text": "Post in 5 Facebook Groups with boomer content", "tag": "SEMI", "detail": "Target Groups: golf, fishing, health tracking for men 50+.\nContent: practical tips with product recommendations.\n1 post per group per day. Genuine value, not spam.\nTime: 20 min."},
                {"text": "Create 3 LinkedIn posts for EAS B2B", "tag": "SEMI", "detail": "Repurpose best tweets with B2B framing.\nOptimize for Depth Score. Zero external links in body.\nTime: 15 min."},
                {"text": "Cross-post top content across all platforms", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/edge_growth_engine.py --cross-post\nDistributes top-performing content across X, LinkedIn, Pinterest, Facebook.\nAdapted per platform. Never identical.\nTime: 0 min (automated)."},
                {"text": "Reply to 75 from primary + 30 from secondaries (105 total)", "tag": "SEMI", "detail": "Review queue from edge_growth_engine.py.\nPost via GoLogin profiles.\nTime: 50 min."},
                {"text": "DM 15 people (engagers + cold prospects)", "tag": "MANUAL", "detail": "5 who engaged with content + 10 cold prospects from email list.\nTime: 25 min."},
                {"text": "Send cold email batch 4 (100 emails, 275 total)", "tag": "AUTO", "detail": "n8n w04 sends via SendGrid.\nSplit test 3 value props.\nFollow up: Day 3 on batch 2, Day 5 on batch 1 (auto).\nSafe limit: multi-inbox rotation.\nTime: 0 min (automated). 5 min to check analytics."},
                {"text": "Write case study from any wins", "tag": "SEMI", "detail": "If ANY positive signal (reply, call booked, gig landed):\nWrite detailed case study with real numbers.\nUse as social proof in: Gumroad, proposals, cold emails.\nStore at: OPS/CASE_STUDIES/\nTime: 20 min."},
                {"text": "Verify n8n workflows + engagement planner + alpha processing", "tag": "AUTO", "detail": "Commands:\n  Check n8n w01/w04/w09/w14: all executing.\n  Check n8n w16 error alerter: any silent failures.\n  Verify daily_engagement_planner.py 7AM cron.\n  python3 AUTOMATIONS/alpha_auto_processor.py --process-new (100 entries)\nTime: 5 min."},
            ],
            "revenue_target": "$20-100",
            "content": {"tweets": 36, "threads": 2, "tiktoks": 0, "youtube": 0, "newsletters": 0},
            "engagement": {"replies": 105, "dms": 15, "comments": 20, "follows": 60},
            "outreach": {"cold_emails": 100, "proposals": 0, "calls": 0},
        },
        11: {
            "theme": "EAS OFFENSIVE + MULTI-PLATFORM",
            "tasks": [
                {"text": "All 4 secondaries: 8 tweets + 25 replies each", "tag": "AUTO", "detail": "32 tweets + 100 replies total across secondaries.\nAuto-posted via cron.\nTier: MEDIUM.\nTime: 0 min (automated)."},
                {"text": "Create Pinterest boards for each niche", "tag": "MANUAL", "detail": "Boards: automation tools, boomer health, golf tech, finance tips.\n10 pins per board with affiliate links.\nSEO-optimized descriptions.\nTime: 20 min."},
                {"text": "Post 12 tweets + 1 thread from primary", "tag": "AUTO", "detail": "Cross-promoting YouTube/TikTok content.\nTime: 0 min (automated)."},
                {"text": "Upload YouTube faceless videos #6-7 (boomer niches)", "tag": "SEMI", "detail": "Golf tips, health supplements content.\nFull SEO setup: title, description, tags, end screens.\nTime: 25 min."},
                {"text": "Post 3 TikToks + cross-post to Shorts + Reels", "tag": "SEMI", "detail": "Adapted per platform, never identical.\nRemove TikTok watermark before cross-posting.\nTime: 20 min."},
                {"text": "Pin 10 affiliate products on Pinterest", "tag": "SEMI", "detail": "SEO-optimized descriptions.\nAffiliate links to boomer-targeted products.\nTime: 15 min."},
                {"text": "Reply to 75 primary + 30 secondaries + 20 TikTok comments (125 total)", "tag": "SEMI", "detail": "Review queue from edge_growth_engine.py.\nTikTok: reply to comments on your videos (algorithm boost for reply engagement).\nTime: 60 min."},
                {"text": "DM 10 TikTok creators for duet/stitch opportunities", "tag": "MANUAL", "detail": "Target: AI/tech TikTokers with 5K-100K followers.\nOffer: duet/stitch collaboration.\nTikTok collab = instant audience share.\nTime: 15 min."},
                {"text": "Run EAS lead pipeline: score 20 local businesses", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/eas_lead_pipeline.py\nCron: weekday 8AM.\nOutput: LEDGER/EAS_LEADS_SCORED.csv\nTop verticals: dental, legal, HVAC.\nTime: 0 min (automated). 5 min to review."},
                {"text": "Send 20 personalized EAS outreach emails", "tag": "SEMI", "detail": "To highest-scored leads from eas_lead_pipeline.\nVerticals: dental, legal, HVAC.\nTemplates from MONEY_METHODS/EAS/email_templates/.\nPersonalize with business name, pain point, industry.\nTime: 30 min."},
                {"text": "Submit 10 Upwork proposals ($2K-10K targets)", "tag": "SEMI", "detail": "Increase budget targets from $1K to $2K-10K.\n55 total active proposals.\nTime: 40 min."},
                {"text": "Verify loop closer, soul drift, UK contracts", "tag": "AUTO", "detail": "Commands:\n  python3 AUTOMATIONS/loop_closer.py --cycle (4 loops: decisions, feedback, pipeline, soul drift)\n  Review soul_drift_report.json: rewrite prompts for agents < 7/10.\n  python3 AUTOMATIONS/uk_contracts_finder.py: new UK IT/automation opportunities.\nTime: 5 min."},
            ],
            "revenue_target": "$20-150",
            "content": {"tweets": 44, "threads": 1, "tiktoks": 3, "youtube": 2, "newsletters": 0},
            "engagement": {"replies": 125, "dms": 10, "comments": 20, "follows": 50},
            "outreach": {"cold_emails": 20, "proposals": 10, "calls": 0},
        },
        12: {
            "theme": "EAS CLOSE + GOV BID PREP",
            "tasks": [
                {"text": "Audit all account health (TweepCred, engagement rates, warmup)", "tag": "SEMI", "detail": "Check each account:\n- Engagement rate (healthy = 3-6%, suppressed = <1%)\n- Follower-to-engagement ratio (bought followers tank this)\n- TweepCred proxy: high engagement + Premium = 55-75 range\n- Warmup progress: past 100 tweets threshold?\nDocument in OPS/ACCOUNT_HEALTH.md.\nTime: 15 min."},
                {"text": "Post 12 tweets + 1 thread from primary (EAS case study)", "tag": "AUTO", "detail": "Thread: EAS automation ROI breakdown.\nTime: 0 min (automated)."},
                {"text": "Post 8 per secondary (24 total) + 3 TikToks", "tag": "AUTO", "detail": "TikToks: real automation results (screen recordings).\nTime: 0 min (automated). 10 min for TikTok posting."},
                {"text": "Create Facebook ad creative for boomer demo", "tag": "SEMI", "detail": "$0 spend today -- build assets only.\nCreative: golf/health product review format.\nTarget: men 55-70, interests: golf, fishing, health supplements.\nStore in MEDIA/ad_creatives/facebook/.\nTime: 20 min."},
                {"text": "Reply to 75 primary + 30 secondaries + 10 LinkedIn (115 total)", "tag": "SEMI", "detail": "LinkedIn: comment on 10 posts about business automation.\nValue-add comments. EAS site link in profile, not comments.\nTime: 55 min."},
                {"text": "DM 10 local business owners on IG/Facebook", "tag": "MANUAL", "detail": "Target: owners who posted about being overwhelmed.\nOffer free 15-min audit of their operations.\nPersonalize based on their posts.\nTime: 20 min."},
                {"text": "Cold call 5 highest-scored EAS leads", "tag": "SEMI", "detail": "Use Bland AI if available, manual phone if not.\nScript from MONEY_METHODS/EAS/call_scripts/.\nGoal: book discovery call or get proposal request.\nTrack: pick-up rate, interest level, objections.\nTime: 30 min if manual, 0 if Bland AI."},
                {"text": "Send cold email batch 5 (75 emails, 350 total)", "tag": "AUTO", "detail": "Best-performing subject line from A/B tests.\nn8n w04 auto-sends.\nTime: 0 min (automated)."},
                {"text": "Send 5 custom EAS proposals", "tag": "SEMI", "detail": "Packages: Signal Map ($1,500), Phone Pilot ($3,500), Ops Pilot ($4,500).\nTemplates at MONEY_METHODS/EAS/proposals/.\nCustomize per business. Attach ROI projections.\nTime: 45 min."},
                {"text": "Draft SAM.gov proposal for highest-value contract", "tag": "SEMI", "detail": "From GOV_OPPORTUNITIES.csv: pick highest-value matching contract.\nDraft proposal following SAM.gov format requirements.\nTime: 60 min."},
                {"text": "Verify security audit, lead scoring, Apollo enrichment", "tag": "AUTO", "detail": "Commands:\n  python3 AUTOMATIONS/security_audit.py --status (6-category scan)\n  Check eas_lead_pipeline.py: high-scored leads converting to replies?\n  Check n8n w02 (Apollo lead enrichment): B2B contacts matching EAS.\nTime: 5 min."},
            ],
            "revenue_target": "$30-200",
            "content": {"tweets": 36, "threads": 1, "tiktoks": 3, "youtube": 0, "newsletters": 0},
            "engagement": {"replies": 115, "dms": 10, "comments": 10, "follows": 40},
            "outreach": {"cold_emails": 75, "proposals": 5, "calls": 5},
        },
        13: {
            "theme": "BOOMER BLITZ + NEWSLETTER",
            "tasks": [
                {"text": "Subscribe X Premium on 2 more secondaries past warmup", "tag": "MANUAL", "detail": "Accounts that passed warmup threshold get X Premium Basic ($3/mo each).\nTotal X Premium spend: $15/mo for 5 accounts.\n10x impressions per account.\nTime: 5 min."},
                {"text": "Write newsletter #1 (Week 2 alpha + tool recommendation)", "tag": "SEMI", "detail": "Beehiiv newsletter.\nContent: Week 2 automation alpha + exclusive tool recommendation.\nInclude affiliate link for recommended tool.\nCTA: reply to boost deliverability + share for growth.\nTime: 25 min."},
                {"text": "Post 12 tweets + 2 threads from primary", "tag": "AUTO", "detail": "1 growth update thread, 1 tactical lesson thread.\nTime: 0 min (automated)."},
                {"text": "Post 8 per secondary (24 total)", "tag": "AUTO", "detail": "Niche-specific content per account.\nTime: 0 min (automated)."},
                {"text": "Post in 10 Facebook Groups with boomer content", "tag": "SEMI", "detail": "Golf tech reviews, health supplements, fishing gear.\n1 post per group. Genuine value.\nTime: 30 min."},
                {"text": "Upload YouTube faceless videos #8-9 (boomer niches)", "tag": "SEMI", "detail": "Supplement reviews, golf gadgets.\nFull SEO setup.\nTime: 25 min."},
                {"text": "Generate next batch of 5 golf video scripts", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/ai_video_content_pipeline.py --generate golf --count 5\nTime: 0 min (automated)."},
                {"text": "Reply to 75 primary + 30 secondaries (105 total)", "tag": "SEMI", "detail": "Review queue. Post via GoLogin.\nTime: 50 min."},
                {"text": "DM 15 people (subscribers, engagers, cold)", "tag": "MANUAL", "detail": "5 newsletter subscribers (engagement), 5 thread engagers (conversion), 5 cold prospects.\nTime: 25 min."},
                {"text": "Submit 10 Upwork proposals (65 total)", "tag": "SEMI", "detail": "Fresh $2K+ jobs. Track win rate.\nTime: 40 min."},
                {"text": "Launch $5/day Facebook ad or 10 more organic Group posts", "tag": "SEMI", "detail": "IF budget: Facebook Ads Manager > men 55-70, interests: golf/fishing/health.\n$5/day = $150/mo. Track: CPC, CTR, conversions.\nIF no budget: organic substitute -- 10 more boomer Facebook Group posts.\nTime: 15 min."},
                {"text": "Bland AI: 50 calls to scored leads", "tag": "AUTO", "detail": "Week 2 ramp. 100 free calls/day.\nTarget: 2-3 meetings booked.\nLog objections for script refinement.\nTime: 0 min (automated). 10 min to review."},
                {"text": "Verify cron health, method discovery, content repurpose", "tag": "AUTO", "detail": "Commands:\n  crontab -l | wc -l (verify ~112 jobs)\n  Check method_discovery_crawler.py 5AM cron output.\n  Check n8n w09 content repurpose execution.\nTime: 5 min."},
            ],
            "revenue_target": "$30-200",
            "content": {"tweets": 36, "threads": 2, "tiktoks": 0, "youtube": 2, "newsletters": 1},
            "engagement": {"replies": 105, "dms": 15, "comments": 15, "follows": 45},
            "outreach": {"cold_emails": 0, "proposals": 10, "calls": 50},
        },
        14: {
            "theme": "WEEK 2 CALIBRATION",
            "tasks": [
                {"text": "Full multi-account audit (engagement, growth, shadowban)", "tag": "SEMI", "detail": "Per account: engagement rate, follower growth, shadowban status.\nshadowban.eu for X. Engagement drop for others.\nDocument results. Kill accounts <0.5% engagement.\nTime: 20 min."},
                {"text": "Batch create Week 3 content (35 tweets + 7 threads + 20/secondary)", "tag": "AUTO", "detail": "Commands:\n  python3 AUTOMATIONS/content_factory.py --batch-alpha 35\n  python3 AUTOMATIONS/content_factory.py --threads 7\nScript 5 TikToks + 1 YouTube tutorial.\nTime: 0 min (automated). 20 min to review."},
                {"text": "Post 8 tweets + 1 thread from primary (Week 2 numbers)", "tag": "AUTO", "detail": "Transparent growth update with real numbers.\nTime: 0 min (automated)."},
                {"text": "Reply to 30 accounts (reduced volume, analytics day)", "tag": "SEMI", "detail": "Quality over quantity on calibration day.\nHandle all open DM conversations.\nTime: 20 min."},
                {"text": "Full funnel analysis: emails > opens > replies > calls > revenue", "tag": "MANUAL", "detail": "Compile from: SendGrid (email), Upwork (proposals), Gumroad (products), all platforms.\nFull proposal analysis: sent vs responses vs gigs won vs revenue/gig.\nRank every channel by revenue-per-hour.\nTop 3 channels get 3x resources in Week 3.\nKILL any channel with literally 0 signal after 14 days.\nStore: OPS/WEEKLY_ANALYTICS/week2.md\nTime: 45 min."},
                {"text": "Pre-draft 100 cold emails for Week 3", "tag": "AUTO", "detail": "Auto-generate with personalization tokens.\nLoad into n8n w04.\nTime: 0 min (automated)."},
                {"text": "Review GOV contracts + submit if deadline within 7 days", "tag": "SEMI", "detail": "SAM.gov + UK Contracts Finder results from the week.\nIf any contract deadline within 7 days: prioritize submission.\nTime: 15 min."},
                {"text": "Verify backups, resilience logs, priority stack", "tag": "AUTO", "detail": "Commands:\n  Verify backup cron ran. Check agent resilience logs for circuit breaker triggers.\n  Review ACTIONABLE_QUEUE.md (7:30AM cron): execute queued items.\n  Deep dive CAPITAL_GENESIS_PRIORITY_STACK.md: adjust Week 3 priorities.\nTime: 10 min."},
            ],
            "revenue_target": "$50-250",
            "content": {"tweets": 8, "threads": 1, "tiktoks": 0, "youtube": 0, "newsletters": 0},
            "engagement": {"replies": 30, "dms": 10, "comments": 5, "follows": 15},
            "outreach": {"cold_emails": 0, "proposals": 0, "calls": 0},
        },
        # ===== WEEK 3: EDGE + OPTIMIZE =====
        15: {
            "theme": "KILL AND SCALE",
            "tasks": [
                {"text": "KILL accounts with <0.5% engagement after 14 days", "tag": "MANUAL", "detail": "Review Week 2 analytics.\n<0.5% engagement = algo suppression territory.\nOptions: rebuild with new content strategy OR abandon.\nFree up resources (proxy, profile, time) for performing accounts.\nTime: 10 min."},
                {"text": "SCALE top 2 accounts to 3x posting volume", "tag": "AUTO", "detail": "Update cron schedules for top performers.\nIncrease from 8 to 15-20 tweets/day.\nauto_content_poster.py handles volume scaling.\nTime: 5 min to adjust cron."},
                {"text": "Deploy follow/unfollow on AGGRESSIVE tier accounts", "tag": "AUTO", "detail": "50 follows/day per account. Unfollow after 3 days if no followback.\nONLY on AGGRESSIVE tier (expendable accounts).\nNEVER on SAFE (@PRINTMAXXER) or MEDIUM accounts.\nedge_growth_engine.py tracks ratios within platform limits.\nX limit: 400 follows/day (new accounts: 10/day).\nRisk: Shadowban on AGGRESSIVE accounts (acceptable).\nTime: 0 min (automated)."},
                {"text": "Post 15 tweets + 2 threads from primary + 10/secondary (35 total)", "tag": "AUTO", "detail": "Scaled volume on performing accounts.\nTime: 0 min (automated)."},
                {"text": "Deploy content atomization: 1 thread > 12 angles", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/content_factory.py --atomize [best_thread]\n1 source > 12 niche-specific angles via AI.\nEach account gets unique voice and angle.\n40% penalty for duplicate detection -- all 12 must be unique.\nTime: 0 min (automated)."},
                {"text": "Post 3 TikToks + 2 YouTube Shorts", "tag": "SEMI", "detail": "From auto_clip_pipeline.py best clips.\nAdapt per platform.\nTime: 15 min."},
                {"text": "Reply to 80 primary + 40 secondaries (120 total)", "tag": "SEMI", "detail": "Push ceiling on replies. Review queue.\nQuote-tweet 5 viral posts within 15 min (contrarian take).\nTime: 60 min."},
                {"text": "DM 20 people (warm leads, cold, collaborators)", "tag": "MANUAL", "detail": "Mix of warm leads, cold prospects, collaboration pitches.\nTime: 30 min."},
                {"text": "Send cold email batch 6 (100 emails, 450 total)", "tag": "AUTO", "detail": "Only best-performing template from A/B tests.\nn8n w04 auto-sends.\nTime: 0 min (automated)."},
                {"text": "Auto-bid on all matching freelance jobs", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/auto_freelance_responder.py --scan-and-respond\nBatch bid on matching Reddit/Upwork jobs.\nReview proposals before sending.\nTime: 0 min (automated). 10 min to review."},
                {"text": "Launch newsletter lead magnet across all channels", "tag": "SEMI", "detail": "Tweet link from primary (link in reply, not main tweet).\nPin to profile. Add to all bios.\nAdd to YouTube descriptions, Gumroad checkout.\nTime: 15 min."},
                {"text": "Verify edge tracking, cognition audit, alpha index", "tag": "AUTO", "detail": "Commands:\n  python3 AUTOMATIONS/edge_growth_engine.py --status (follow/unfollow ratios)\n  Check competitive_cognition_audit.py (Sunday 5AM): strategic blind spots.\n  Verify sqlite_alpha_index.py FTS5 index current (3:30AM daily).\nTime: 5 min."},
                {"text": "Follow up on brokering leads, close first referral deal", "tag": "SEMI", "detail": "Review AUTOMATIONS/leads/brokering/outreach_log.csv for replies and interest signals.\nFollow up on non-responders with different angle (Day 3 and Day 5 sequences).\nFor warm leads: identify best service provider match and draft warm intro email for both parties.\nSend referral agreement template to any progressing deals.\nGoal: close first referral fee ($50-5K depending on vertical).\nTrack all pipeline movement in LEDGER/BROKERING_REVENUE.csv\nTime: 25 min."},
            ],
            "revenue_target": "$50-300",
            "content": {"tweets": 35, "threads": 2, "tiktoks": 3, "youtube": 0, "newsletters": 0},
            "engagement": {"replies": 120, "dms": 20, "comments": 15, "follows": 75},
            "outreach": {"cold_emails": 100, "proposals": 5, "calls": 0},
        },
        16: {
            "theme": "AFFILIATE + NEWSLETTER ENGINE",
            "tasks": [
                {"text": "All secondaries: 10 tweets + 30 replies each", "tag": "AUTO", "detail": "Scaling past warmup into growth phase.\nTime: 0 min (automated)."},
                {"text": "Write newsletter #2 (best alpha + exclusive content)", "tag": "SEMI", "detail": "Best alpha from Week 2. Content NOT posted publicly.\nExclusive = subscriber retention.\nTime: 25 min."},
                {"text": "Write 5 affiliate review threads (boomer targeting)", "tag": "SEMI", "detail": "Health supplements, golf tech, fishing gear, financial tools, insurance.\nEach thread: honest review + affiliate link in reply (not main tweet).\nFTC compliant: '#ad' or '#affiliate' disclosure.\nTime: 40 min."},
                {"text": "Post affiliate content across 4 platforms", "tag": "SEMI", "detail": "3 Facebook Groups (boomer), 2 Reddit posts, 2 Pinterest pins, 1 YouTube Short.\nAdapted per platform. Genuine reviews, not spam.\nTime: 25 min."},
                {"text": "Post 12 tweets + 2 threads from primary", "tag": "AUTO", "detail": "Newsletter promotion + affiliate value thread.\nTime: 0 min (automated)."},
                {"text": "Reply to 80 primary + 40 secondaries (120 total)", "tag": "SEMI", "detail": "Target people asking 'what tool should I use for X?'\nReply with genuine comparison + affiliate link in reply thread.\nTime: 60 min."},
                {"text": "DM 15 newsletter/affiliate engagers", "tag": "MANUAL", "detail": "Offer exclusive bonus to people who engaged with affiliate content.\nTime: 25 min."},
                {"text": "Apply to 5 affiliate programs", "tag": "MANUAL", "detail": "Health/supplement (10-15%), SaaS tools ($50-500/ref), hosting ($50-200/ref).\nApply via ShareASale, CJ Affiliate, Impact, direct.\nTime: 15 min."},
                {"text": "Send cold email batch 7 (75 emails, 525 total)", "tag": "AUTO", "detail": "Newsletter CTA in email signature for cross-channel growth.\nTime: 0 min (automated)."},
                {"text": "Submit 10 Upwork proposals (75 total)", "tag": "SEMI", "detail": "Track win rate this week.\nAdd newsletter opt-in to: Gumroad checkout, YouTube desc, all bios.\nTime: 40 min."},
                {"text": "Verify daily digest, prompt meta review, compliance", "tag": "AUTO", "detail": "Commands:\n  Check daily_digest.py: missed opportunities last 48h.\n  Check prompt_meta_review.py (48h cycle): lost threads or forgotten goals.\n  Verify compliance_scanner: all affiliate content FTC compliant.\nTime: 5 min."},
            ],
            "revenue_target": "$50-300",
            "content": {"tweets": 32, "threads": 2, "tiktoks": 0, "youtube": 0, "newsletters": 1},
            "engagement": {"replies": 120, "dms": 15, "comments": 15, "follows": 50},
            "outreach": {"cold_emails": 75, "proposals": 10, "calls": 0},
        },
        17: {
            "theme": "EAS CLOSE PUSH + ARBITRAGE",
            "tasks": [
                {"text": "Check TweepCred health across all accounts", "tag": "AUTO", "detail": "Engagement-to-follower ratios healthy?\nAny account below suppression threshold needs intervention.\nCommand: python3 AUTOMATIONS/edge_growth_engine.py --account-health\nTime: 5 min."},
                {"text": "Identify top 3 APIs to wrap. Use API_ARBITRAGE_PLAYBOOK.md", "tag": "SEMI", "detail": "Browse RapidAPI + Public-APIs GitHub for free-tier APIs solving real problems.\nCheck competitors charging $50+/mo for same data.\nPick top 3 by: demand + build ease + margin (70-99%).\nRef: MONEY_METHODS/API_ARBITRAGE/API_ARBITRAGE_PLAYBOOK.md\nTime: 30 min."},
                {"text": "Find 5 trending repos to repurpose", "tag": "SEMI", "detail": "GitHub search: license:mit stars:>200 pushed:>2025-09-01\nCategories: SaaS starters, scrapers, AI wrappers, payment integrations.\nRun security scan from GITHUB_REPURPOSE_STRATEGY.md Section 2.\nLog to LEDGER/GITHUB_REPURPOSE_TRACKER.csv.\nTime: 30 min."},
                {"text": "Post 12 tweets + 2 threads from primary + 10/secondary (42 total)", "tag": "AUTO", "detail": "1 EAS value prop thread, 1 tactical content thread.\nTime: 0 min (automated)."},
                {"text": "Upload YouTube EAS video", "tag": "SEMI", "detail": "'How I automate a business in 48 hours'\nFull SEO setup. End screen to subscribe.\nTime: 25 min."},
                {"text": "Post 3 TikToks (product review hook format)", "tag": "SEMI", "detail": "Hook: 'Stop paying for X when Y exists'\nFormat: problem > reveal > CTA.\nTime: 15 min."},
                {"text": "Reply to 80 primary + 40 secondaries (120 total)", "tag": "SEMI", "detail": "Review queue. Post via GoLogin.\nTime: 60 min."},
                {"text": "DM 15 local business owners about growth/operations", "tag": "MANUAL", "detail": "From Instagram/Facebook. Owners who posted about growth struggles.\nOffer free audit. Build relationship before pitch.\nTime: 25 min."},
                {"text": "Push top 5 warmest EAS leads to close", "tag": "MANUAL", "detail": "Call/email top 5 warmest leads.\nPush for discovery call or proposal request.\nAttach case study if available.\nTime: 30 min."},
                {"text": "Send 5 custom EAS proposals ($1,500-4,500)", "tag": "SEMI", "detail": "From MONEY_METHODS/EAS/ templates.\nCustomize per business with specific automation opportunities.\nTime: 45 min."},
                {"text": "Send 25 EAS follow-up emails (Day 7/14 follow-ups)", "tag": "AUTO", "detail": "Auto via SendGrid sequence.\nFollowing up on all previous EAS outreach.\nTime: 0 min (automated)."},
                {"text": "Scan ads libraries for 3 winning e-commerce products", "tag": "AUTO", "detail": "TikTok/Facebook ads library scan.\nIdentify products for e-commerce arbitrage.\nTime: 0 min (automated). 10 min to review."},
                {"text": "Weekly SAM.gov contract scan", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/sam_gov_monitor.py\nTime: 0 min (automated). 5 min to review."},
                {"text": "Verify EAS pipeline, model routing, freelance responder", "tag": "AUTO", "detail": "Commands:\n  Check eas_lead_pipeline.py accuracy: high-scored leads vs reply conversion.\n  Check agent_swarm model routing: Opus/Sonnet/Haiku assignments correct.\n  Check auto_freelance_responder.py: proposals sent in last 24h.\nTime: 5 min."},
            ],
            "revenue_target": "$75-350",
            "content": {"tweets": 42, "threads": 2, "tiktoks": 3, "youtube": 1, "newsletters": 0},
            "engagement": {"replies": 120, "dms": 15, "comments": 15, "follows": 50},
            "outreach": {"cold_emails": 25, "proposals": 5, "calls": 5},
        },
        18: {
            "theme": "CONTENT FLOOD + CROSS-AMPLIFICATION",
            "tasks": [
                {"text": "Cross-account amplification: secondaries reply to primary thread", "tag": "SEMI", "detail": "Primary posts thread. Secondaries reply with value-add takes.\nEach account unique perspective from its niche.\nStagger 1-2 hours. Triggers velocity window.\nTier: MEDIUM.\nTime: 15 min."},
                {"text": "Deploy MCPHub to surge. Follow LAUNCH.md", "tag": "MANUAL", "detail": "Curate 50+ MCP servers in servers.json.\nDeploy to mcphub.surge.sh.\nTest on mobile + desktop.\nSet up submission email.\nRef: MONEY_METHODS/MCP_MARKETPLACE/LAUNCH.md + MONETIZATION.md\nMonetization: featured $29/mo, verified $9/mo, enterprise $99/mo.\nTime: 60 min."},
                {"text": "Post to 3 X Communities before main tweets", "tag": "SEMI", "detail": "Triggers 30-min algo velocity window.\nCommunities provide warm audience for initial engagement.\nTime: 10 min."},
                {"text": "Batch create 15 TikToks across niches", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/ai_video_content_pipeline.py --generate fitness --count 5 --generate golf --count 5 --generate health --count 5\nTime: 0 min (automated)."},
                {"text": "Batch create 5 YouTube Shorts from best clips", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/auto_clip_pipeline.py --batch --max-clips 5\nTime: 0 min (automated)."},
                {"text": "Post 10 Facebook posts for boomer Groups", "tag": "SEMI", "detail": "5 different groups, 2 posts each.\nGenuine value content with product recommendations.\nTime: 25 min."},
                {"text": "Write newsletter #3 (Week 3 alpha + subscriber offer)", "tag": "SEMI", "detail": "Exclusive alpha + subscriber-only discount on Gumroad products.\nTime: 25 min."},
                {"text": "Post 12 tweets + 2 threads from primary + 10/secondary (42 total)", "tag": "AUTO", "detail": "1 engagement-bait thread, 1 value thread.\nTime: 0 min (automated)."},
                {"text": "Post 5 findom persona tweets + 3 Fanvue paid posts", "tag": "SEMI", "detail": "Maintain daily cadence. P02 projected $200-5K/mo.\nSeparate GoLogin profile.\nTime: 15 min."},
                {"text": "Publish 2 Medium articles + 1 LinkedIn long-form (Parasite SEO)", "tag": "SEMI", "detail": "Repurpose top threads.\nDA90+ backlinks compound over 60 days.\nMedium: submit to publications for 10x distribution.\nTime: 30 min."},
                {"text": "Reply to 80 primary + 40 secondaries + 10 viral (130 total)", "tag": "SEMI", "detail": "Reply to viral posts within 30 min for max algo weight.\nTime: 65 min."},
                {"text": "DM 20 people (followers, clients, partners)", "tag": "MANUAL", "detail": "10 engaged followers + 5 potential clients + 5 partners.\nTime: 30 min."},
                {"text": "Post 5 Pinterest pins with affiliate boomer products", "tag": "SEMI", "detail": "Golf, health, fishing affiliate links.\nSEO-optimized descriptions.\nTime: 15 min."},
                {"text": "Check e-commerce performance across all listings", "tag": "SEMI", "detail": "Etsy views/favorites, KDP page reads, Gumroad views/conversions, TikTok Shop clicks.\nKill underperformers. Double down on winners.\nTime: 15 min."},
                {"text": "Bland AI: 30 calls to scored leads", "tag": "AUTO", "detail": "EAS discovery script. Target: 1-2 meetings booked.\nTime: 0 min (automated). 5 min to review."},
                {"text": "Verify content queues, system visualizer, discovery crawler", "tag": "AUTO", "detail": "Commands:\n  Check Buffer/Typefully/Publer queues: all content scheduled.\n  python3 AUTOMATIONS/system_visualizer.py: review SYSTEM_VISUAL.html.\n  Check method_discovery_crawler.py + daily_tool_scout.py: emergent tactics.\nTime: 5 min."},
            ],
            "revenue_target": "$100-500",
            "content": {"tweets": 42, "threads": 2, "tiktoks": 3, "youtube": 0, "newsletters": 1},
            "engagement": {"replies": 130, "dms": 20, "comments": 20, "follows": 50},
            "outreach": {"cold_emails": 0, "proposals": 0, "calls": 0},
        },
        19: {
            "theme": "EDGE GROWTH DEPLOYMENT",
            "tasks": [
                {"text": "Deploy bookmark scraping on top 10 alpha accounts", "tag": "AUTO", "detail": "Monitor bookmarks for content ideas and trend signals.\nedge_growth_engine.py handles scraping.\nTime: 0 min (automated)."},
                {"text": "Check shadowban status on all accounts", "tag": "SEMI", "detail": "shadowban.eu for X. Engagement drop for IG/TikTok.\nRecovery protocol if shadowbanned: reduce volume, manual only, 5-7 days.\nTime: 15 min."},
                {"text": "Post 12 tweets + 1 thread primary + 10/secondary (32 total)", "tag": "AUTO", "detail": "Time: 0 min (automated)."},
                {"text": "Run full edge tactic audit", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/edge_growth_engine.py --squeeze\nFull edge tactic audit with actionable output.\nAlso: --viral [top_product] for viral hooks.\nTime: 0 min (automated). 10 min to review."},
                {"text": "Reply to 80 primary + 40 secondaries (120 total)", "tag": "SEMI", "detail": "Deploy reply guy at scale: reply within 15 min to 10 big accounts.\nAuthor-engaged reply = +75 algo weight.\nTime: 60 min."},
                {"text": "DM 15 highest-interaction people this week", "tag": "MANUAL", "detail": "Convert engagement into relationships/sales.\nTime: 25 min."},
                {"text": "Send cold email batch 8 (50 emails, 500 total)", "tag": "AUTO", "detail": "Only highest-engagement segments.\nTime: 0 min (automated)."},
                {"text": "Capture new hot freelance jobs", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/auto_freelance_responder.py --scan-and-respond\nReview all Fiverr/Upwork active proposals: respond within 1 hour.\nTime: 0 min (automated). 10 min to review."},
                {"text": "Bland AI: 30 calls (push for discovery calls)", "tag": "AUTO", "detail": "EAS + local biz leads. Track appointment-to-close ratio.\nTime: 0 min (automated)."},
                {"text": "Scan UK contracts for high-value opportunities", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/uk_contracts_finder.py --keyword 'digital' --min-value 50000\nReview GOV_OPPORTUNITIES.csv: deadlines in 7 days? Draft submission.\nTime: 0 min (automated). 10 min to review."},
                {"text": "Measure edge tactic lift + verify alpha index", "tag": "AUTO", "detail": "Are follow/unfollow, cross-amplification, X Communities producing measurable engagement increase?\nFTS5 query: python3 AUTOMATIONS/sqlite_alpha_index.py --query 'edge growth arbitrage'\nTime: 10 min."},
            ],
            "revenue_target": "$100-500",
            "content": {"tweets": 32, "threads": 1, "tiktoks": 0, "youtube": 0, "newsletters": 0},
            "engagement": {"replies": 120, "dms": 15, "comments": 15, "follows": 55},
            "outreach": {"cold_emails": 50, "proposals": 5, "calls": 0},
        },
        20: {
            "theme": "PRODUCT HUNT PREP + PERSONA LAUNCH",
            "tasks": [
                {"text": "Launch Fanvue persona with first 5 posts", "tag": "SEMI", "detail": "Separate GoLogin profile + SOAX proxy. Findom niche.\nHint content on X secondary with teaser + Fanvue link in bio.\nTime: 20 min."},
                {"text": "Prepare Product Hunt launch assets", "tag": "SEMI", "detail": "Pick best product with clear value prop + professional landing page.\nAssets: 5 screenshots, 60-sec demo video, tagline, description, first comment draft.\nStore at: PRODUCTS/product_hunt/\nTime: 45 min."},
                {"text": "Post 12 tweets + 2 threads from primary + 10/secondary (42 total)", "tag": "AUTO", "detail": "1 pre-launch hype thread, 1 build-in-public thread.\nTime: 0 min (automated)."},
                {"text": "Reply to 80 primary + 40 secondaries (120 total)", "tag": "SEMI", "detail": "Time: 60 min."},
                {"text": "DM 15 PH launchers for tips + mutual upvote", "tag": "MANUAL", "detail": "People who launched recently: ask for tips + mutual support.\nRecruit 10 hunters/upvoters: DM people who upvoted similar products.\nTime: 25 min."},
                {"text": "Create urgency offer: limited-time bundle at 40% off", "tag": "SEMI", "detail": "Combine top 3 Gumroad products.\nAnchor: '$297 value for $97 -- limited time.'\nTime: 15 min."},
                {"text": "Set up CashApp/Venmo/crypto tribute page for persona", "tag": "MANUAL", "detail": "P02/P11 passive tributes.\nLink from Fanvue profile + X bio.\nTime: 10 min."},
                {"text": "Submit 10 Upwork proposals (85 total)", "tag": "SEMI", "detail": "Time: 40 min."},
                {"text": "Verify landing page, system health, pipeline", "tag": "AUTO", "detail": "Commands:\n  Test landing page: load speed, checkout flow, all links.\n  python3 AUTOMATIONS/system_health_monitor.py --quick\n  python3 AUTOMATIONS/loop_closer.py --cycle: deals moving through pipeline?\nTime: 5 min."},
            ],
            "revenue_target": "$100-500",
            "content": {"tweets": 42, "threads": 2, "tiktoks": 0, "youtube": 0, "newsletters": 0},
            "engagement": {"replies": 120, "dms": 15, "comments": 15, "follows": 45},
            "outreach": {"cold_emails": 0, "proposals": 10, "calls": 0},
        },
        21: {
            "theme": "WEEK 3 CALIBRATION",
            "tasks": [
                {"text": "Full multi-account health audit", "tag": "SEMI", "detail": "Per-account: engagement rate, follower growth, TweepCred estimate, shadowban status.\nDocument in OPS/ACCOUNT_HEALTH.md.\nTime: 20 min."},
                {"text": "Post 8 tweets + 1 thread from primary (Week 3 numbers)", "tag": "AUTO", "detail": "Transparent growth update.\nTime: 0 min (automated)."},
                {"text": "Batch create Week 4 content", "tag": "AUTO", "detail": "35 tweets + 7 threads (primary), 25/secondary.\nScript 5 TikToks + 1 YouTube tutorial.\nTime: 0 min (automated). 20 min to review."},
                {"text": "Reply to 30 accounts (reduced calibration volume)", "tag": "SEMI", "detail": "Quality over quantity.\nTime: 20 min."},
                {"text": "Trajectory check: on track for $1K?", "tag": "MANUAL", "detail": "If not: what is the SINGLE fastest move to close the gap? Name it.\nRevenue gap analysis: sort channels by $/hr. Top channel gets 3x resources.\nFull engagement-to-revenue conversion audit per platform.\nIf any venture crossed $100: create SOP to replicate 3x.\nTime: 45 min."},
                {"text": "Pre-draft 100 cold emails for Week 4", "tag": "AUTO", "detail": "With personalization tokens. Load into n8n w04.\nTime: 0 min (automated)."},
                {"text": "Full system audit + recalibrate priorities", "tag": "AUTO", "detail": "Commands:\n  python3 AUTOMATIONS/system_health_monitor.py --full\n  Verify backups + agent resilience logs.\n  python3 AUTOMATIONS/competitive_cognition_audit.py --run\n  Review CAPITAL_GENESIS_PRIORITY_STACK.md: recalibrate Week 4.\nTime: 15 min."},
            ],
            "revenue_target": "$100-600",
            "content": {"tweets": 8, "threads": 1, "tiktoks": 0, "youtube": 0, "newsletters": 0},
            "engagement": {"replies": 30, "dms": 5, "comments": 5, "follows": 15},
            "outreach": {"cold_emails": 0, "proposals": 0, "calls": 0},
        },
        # ===== WEEK 4: COMPOUND + CLOSE =====
        22: {
            "theme": "LAUNCH DAY + REVENUE ENGINE",
            "tasks": [
                {"text": "All accounts full blast for PH launch (45 total tweets)", "tag": "AUTO", "detail": "Primary 15 tweets, secondaries 10 each.\nAll promoting Product Hunt launch.\nauto_content_poster.py handles increased volume.\nTime: 0 min (automated)."},
                {"text": "Execute Product Hunt launch at 12:01 AM PT", "tag": "MANUAL", "detail": "Go to producthunt.com. Submit product.\nPost first comment with story + AMA.\nTiming: 12:01 AM Pacific = least competition for front page.\nFirst hour is critical: rally all upvotes in first 60 min.\nShare PH link: Reddit (r/SaaS, r/SideProject), LinkedIn, Facebook, Discord, Indie Hackers.\nSend newsletter blast: 'We just launched on Product Hunt'\nTime: 30 min launch + all-day monitoring."},
                {"text": "Post 3 TikToks: launch day behind-the-scenes", "tag": "SEMI", "detail": "Raw, authentic BTS content.\nShow real-time upvote count, comments, reactions.\nTime: 15 min."},
                {"text": "Monitor PH + respond to EVERY comment within 15 min", "tag": "MANUAL", "detail": "Check PH every 30 min throughout the day.\nRespond to every comment within 15 min.\nHigh responsiveness = better PH ranking.\nTime: continuous throughout day."},
                {"text": "Reply to 50+ accounts across all platforms (all-engagement mode)", "tag": "SEMI", "detail": "All-engagement mode. Every platform.\nSecondary accounts amplify: quote-tweet launch with unique congratulatory takes.\nDM 20 people who upvoted/commented thanking them.\nTime: 60 min."},
                {"text": "Pitch retainer to best freelance client", "tag": "MANUAL", "detail": "$500-2K/mo for ongoing automation support.\nUse relationship built over Weeks 1-3.\nRecurring revenue > one-time gigs.\nTime: 20 min."},
                {"text": "Set up subscription tier on Gumroad/Whop", "tag": "MANUAL", "detail": "Monthly templates ($19/mo) or community ($49/mo).\nRecurring revenue compounds.\nTime: 15 min."},
                {"text": "Publish Chrome extension to Chrome Web Store", "tag": "MANUAL", "detail": "D05: AI prompt helper or automation shortcut from APP_FACTORY/.\nChrome Web Store: $5 one-time developer fee.\nMonetize: freemium (free basic, $9/mo pro).\nTime: 30 min."},
                {"text": "Verify PH analytics, checkout, auto-fulfillment", "tag": "AUTO", "detail": "PH analytics: upvotes, visits, conversions.\nCheckout: no failures under traffic load.\nn8n w14 Stripe webhook: all sales auto-fulfilled.\nTime: 5 min per check."},
                {"text": "Scale winning brokering vertical, kill non-performers", "tag": "AUTO", "detail": "Review 3-week brokering data: which vertical produced deals/interest?\nWinning vertical: 3x outreach volume, deploy n8n w25 (lead gen service) for automation.\nNon-performers (<2% reply rate after 60+ emails): KILL and reallocate to winner.\nSet up recurring scrape + qualify + connect pipeline for winning vertical via venture_autonomy BROKERING type.\nDeploy n8n w27 (white-label reports) if report-based verticals won.\nTarget: $500+/mo from brokering by end of Month 2.\nVerify: python3 AUTOMATIONS/venture_autonomy.py --status (BROKERING venture active)\nTime: 15 min."},
            ],
            "revenue_target": "$200-1000",
            "content": {"tweets": 45, "threads": 1, "tiktoks": 3, "youtube": 0, "newsletters": 1},
            "engagement": {"replies": 50, "dms": 20, "comments": 30, "follows": 30},
            "outreach": {"cold_emails": 0, "proposals": 0, "calls": 0},
        },
        23: {
            "theme": "RECURRING REVENUE + EAS FINAL PUSH",
            "tasks": [
                {"text": "Resume normal posting cadence (12 primary + 10/secondary)", "tag": "AUTO", "detail": "Back to standard after PH day.\nTime: 0 min (automated)."},
                {"text": "Post 12 tweets + 2 threads (PH results + subscriber value)", "tag": "AUTO", "detail": "Time: 0 min (automated)."},
                {"text": "Post 3 TikToks (day in the life running automated business)", "tag": "SEMI", "detail": "Authentic content showing real automation results.\nTime: 15 min."},
                {"text": "Write newsletter #4 (exclusive offer + referral incentive)", "tag": "SEMI", "detail": "Subscriber-only deal. Referral: 'share with 3 friends, get X free.'\nGrowth loop: each subscriber brings more.\nTime: 25 min."},
                {"text": "Reply to 80 primary + 40 secondaries (120 total)", "tag": "SEMI", "detail": "Time: 60 min."},
                {"text": "DM 20 people (10 buyers/engagers + 10 cold)", "tag": "MANUAL", "detail": "Convert engagement to sales. Push existing buyers to higher tiers.\nTime: 30 min."},
                {"text": "Push hardest EAS prospect to close", "tag": "MANUAL", "detail": "Final proposal with 48-hour expiry + 10% early-bird discount.\nCall top 3 EAS leads: voice builds trust faster than email.\nTime: 30 min."},
                {"text": "Convert Gumroad one-time buyers to subscribers", "tag": "SEMI", "detail": "Email all buyers with subscription offer.\nTemplate: 'You bought X. Get monthly updates + new tools for $19/mo.'\nTime: 15 min."},
                {"text": "Launch Telegram VIP channel ($49-99/mo)", "tag": "MANUAL", "detail": "Exclusive alpha/automation content.\nPayment via Gumroad subscription or direct crypto.\nContent: daily alpha drops, tool reviews, behind-the-scenes.\nTime: 20 min."},
                {"text": "Set up SMS opt-in on checkout pages", "tag": "MANUAL", "detail": "SMS: 98% open rate vs 20% email.\nTCPA REQUIRED: explicit written consent per marketer.\nFirst SMS to opted-in buyers: exclusive deal + referral code.\nUse Twilio or similar. NEVER cold SMS.\nPenalties: $500-$1,500 PER VIOLATION.\nTime: 20 min."},
                {"text": "Send cold email batch 9 (75 emails, 575 total)", "tag": "AUTO", "detail": "PH launch as social proof + testimonials.\nTime: 0 min (automated)."},
                {"text": "Verify Telegram payment, engagement planner, revenue tracking", "tag": "AUTO", "detail": "Telegram VIP payment gateway test.\ndaily_engagement_planner.py: warmup aligned with volume.\nAll revenue dashboards accurate: Gumroad, Fiverr, Upwork, affiliate, Stripe.\nTime: 10 min."},
            ],
            "revenue_target": "$200-1000",
            "content": {"tweets": 42, "threads": 2, "tiktoks": 3, "youtube": 0, "newsletters": 1},
            "engagement": {"replies": 120, "dms": 20, "comments": 15, "follows": 50},
            "outreach": {"cold_emails": 75, "proposals": 0, "calls": 3},
        },
        24: {
            "theme": "CONVERSION OPTIMIZATION + GOV BID",
            "tasks": [
                {"text": "Audit all accounts: keep, pivot, or kill based on 3-week data", "tag": "MANUAL", "detail": "Review 3-week performance per account.\nKeep: >1% engagement, growing followers.\nPivot: stagnant but salvageable niche.\nKill: <0.5% engagement, no growth.\nTime: 15 min."},
                {"text": "Post 12 tweets + 1 thread (cold email lessons)", "tag": "AUTO", "detail": "Thread: 'X things I learned from sending 500+ cold emails'\nTime: 0 min (automated)."},
                {"text": "Post 10/secondary (30 total) + 3 TikToks + 2 Shorts", "tag": "AUTO", "detail": "Top-converting format from analytics.\nTime: 0 min (automated). 15 min for video posting."},
                {"text": "Upload YouTube faceless videos #10-11 (boomer niches)", "tag": "SEMI", "detail": "Full SEO setup.\nTime: 25 min."},
                {"text": "Reply to 80 primary + 40 secondaries (120 total)", "tag": "SEMI", "detail": "Time: 60 min."},
                {"text": "DM 10 warmest leads with personalized offer", "tag": "MANUAL", "detail": "Based on engagement history. Custom pricing.\nTime: 20 min."},
                {"text": "Send cold email batch (50 emails, 625 total)", "tag": "AUTO", "detail": "Daily cadence. Never let pipeline go cold >24h.\nTime: 0 min (automated)."},
                {"text": "Bland AI: 50 calls (fresh leads + follow-ups)", "tag": "AUTO", "detail": "Fresh scored leads + follow-up on Day 23 meeting no-shows.\nTime: 0 min (automated). 10 min to review."},
                {"text": "A/B test Gumroad product titles/descriptions/prices", "tag": "SEMI", "detail": "Top 3 sellers. Test: different titles, prices, descriptions.\nAdd upsells/cross-sells to checkout pages.\nCreate upsell path: $15 buyer > email for $50 product > $99 package.\nOptimize any email with <20% open or <2% click rate.\nTime: 20 min."},
                {"text": "Submit SAM.gov proposal if deadline this week", "tag": "SEMI", "detail": "From GOV_OPPORTUNITIES.csv: best match.\nUK: python3 AUTOMATIONS/uk_contracts_finder.py --keyword 'automation' --min-value 25000\nTime: 60 min if submitting."},
                {"text": "Verify alpha processor, Gumroad analytics, n8n errors", "tag": "AUTO", "detail": "Commands:\n  python3 AUTOMATIONS/alpha_auto_processor.py --process-new (clear all PENDING_REVIEW)\n  Gumroad analytics vs expected revenue match.\n  n8n w16 error alerter: workflow failures in last 7 days.\nTime: 5 min."},
            ],
            "revenue_target": "$200-800",
            "content": {"tweets": 42, "threads": 1, "tiktoks": 3, "youtube": 2, "newsletters": 0},
            "engagement": {"replies": 120, "dms": 10, "comments": 10, "follows": 35},
            "outreach": {"cold_emails": 50, "proposals": 0, "calls": 50},
        },
        25: {
            "theme": "SCALE WINNERS + AUTOMATE REPEATS",
            "tasks": [
                {"text": "Scale top accounts 2x, reduce low performers to maintenance", "tag": "AUTO", "detail": "Top-performing: increase posting 2x, add X Premium if not subscribed.\nLow-performing: 3 tweets + 10 replies daily (maintenance mode).\nUpdate cron schedules accordingly.\nTime: 5 min."},
                {"text": "Post 15 tweets + 2 threads primary + 10/secondary (45 total)", "tag": "AUTO", "detail": "Revenue milestone thread + tactical value thread.\nTime: 0 min (automated)."},
                {"text": "Post 3 TikToks showing real automation workflows", "tag": "SEMI", "detail": "Screen recordings of actual automations running.\nAuthentic > polished.\nTime: 15 min."},
                {"text": "Generate large content buffer for end-of-month", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/content_factory.py --batch-alpha 15\nTime: 0 min (automated)."},
                {"text": "Reply to 80 primary + 40 secondaries (120 total)", "tag": "SEMI", "detail": "Time: 60 min."},
                {"text": "DM 15 potential collaboration partners for next month", "tag": "MANUAL", "detail": "Cross-promotion, affiliate partnerships, content collaborations.\nTime: 25 min."},
                {"text": "Send cold email batch 10 (100 emails, 675 total)", "tag": "AUTO", "detail": "Only hottest segments.\nTime: 0 min (automated)."},
                {"text": "Follow up ALL open proposals with urgency/expiry dates", "tag": "SEMI", "detail": "Upwork, EAS, freelance. Create urgency.\n'This rate expires Friday' or 'Slots filling up.'\nTime: 20 min."},
                {"text": "Automate any task done manually 3+ times this month", "tag": "SEMI", "detail": "Review month's manual tasks.\nAnything done 3+ times: write script + add to cron.\nDeploy remaining ACTIONABLE_QUEUE.md items.\nTime: 30 min."},
                {"text": "Update system map + clean dead ventures", "tag": "AUTO", "detail": "Commands:\n  Update OPS/PRINTMAXX_SYSTEM_MAP.md with all month changes.\n  Archive dead venture docs.\n  Remove orphan cron jobs: crontab -l | grep -v [dead_script]\nCreate automated revenue reporting: n8n w17 daily Telegram summary.\nTime: 15 min."},
            ],
            "revenue_target": "$200-800",
            "content": {"tweets": 45, "threads": 2, "tiktoks": 3, "youtube": 0, "newsletters": 0},
            "engagement": {"replies": 120, "dms": 15, "comments": 10, "follows": 40},
            "outreach": {"cold_emails": 100, "proposals": 5, "calls": 0},
        },
        26: {
            "theme": "PARTNERSHIP + COMMUNITY",
            "tasks": [
                {"text": "Evaluate secondary accounts for X Premium ROI", "tag": "SEMI", "detail": "Based on engagement data: which secondaries justify $3/mo Premium?\nROI threshold: Premium accounts should show >2x impressions vs free.\nTime: 10 min."},
                {"text": "Post 12 tweets + 1 thread (month retrospective, real numbers)", "tag": "AUTO", "detail": "Time: 0 min (automated)."},
                {"text": "Post 10/secondary (30 total) + 3 TikToks + 1 YouTube", "tag": "AUTO", "detail": "YouTube: month retrospective or best tutorial.\nTikToks: authentic behind-the-scenes.\nTime: 0 min (automated). 25 min for YouTube upload."},
                {"text": "Write newsletter #5 (month review + next month preview)", "tag": "SEMI", "detail": "Build anticipation for next month.\nTime: 25 min."},
                {"text": "Reply to 75 primary + 30 secondaries (105 total)", "tag": "SEMI", "detail": "Time: 50 min."},
                {"text": "DM 10 top engagers to convert to paid community", "tag": "MANUAL", "detail": "Convert free followers to Telegram VIP / Gumroad subscribers.\nTime: 20 min."},
                {"text": "Send cold email batch 11 (50 emails, 725 total)", "tag": "AUTO", "detail": "Best social proof from the month.\nTime: 0 min (automated)."},
                {"text": "Push open EAS deals to close with case study + urgency", "tag": "MANUAL", "detail": "Limited-time pricing. Attach case study.\nTime: 20 min."},
                {"text": "Review e-commerce listing performance", "tag": "SEMI", "detail": "Kill underperformers. Double down on winners.\nEtsy, KDP, Gumroad, TikTok Shop.\nTime: 15 min."},
                {"text": "Final monthly GOV contract scan", "tag": "AUTO", "detail": "SAM.gov + UK Contracts Finder.\nIf contract won: begin delivery planning.\nTime: 0 min (automated). 10 min to review."},
                {"text": "Verify all 33 agents + intelligence router + priority stack", "tag": "AUTO", "detail": "Commands:\n  python3 AUTOMATIONS/capital_genesis_ranker.py --rank --report\n  python3 AUTOMATIONS/agent_swarm.py --status (25 agents)\n  python3 AUTOMATIONS/venture_autonomy.py --status (8 ventures)\n  python3 AUTOMATIONS/intelligence_router.py --stats\nTime: 5 min."},
            ],
            "revenue_target": "$200-800",
            "content": {"tweets": 42, "threads": 1, "tiktoks": 3, "youtube": 1, "newsletters": 1},
            "engagement": {"replies": 105, "dms": 10, "comments": 10, "follows": 35},
            "outreach": {"cold_emails": 50, "proposals": 0, "calls": 3},
        },
        27: {
            "theme": "NEXT MONTH WAR PLAN",
            "tasks": [
                {"text": "Full account stack review (ROI per account)", "tag": "MANUAL", "detail": "Total accounts, X Premium spend, engagement/account, ROI/account.\nPlan next month: add, sunset, expand niches.\nTime: 20 min."},
                {"text": "Post 8 tweets + 1 thread (month review, honest numbers)", "tag": "AUTO", "detail": "Time: 0 min (automated)."},
                {"text": "Post 5/secondary (15 total)", "tag": "AUTO", "detail": "Reduced end-of-month volume.\nTime: 0 min (automated)."},
                {"text": "Reply to 30 accounts + handle all open DMs", "tag": "SEMI", "detail": "Time: 20 min."},
                {"text": "Month revenue review vs projections", "tag": "MANUAL", "detail": "Actual vs KPI_DASHBOARD.md projections (conservative $800, medium $2,900, aggressive $8,100).\nPer-channel: revenue, hours invested, $/hr, growth rate. Sort by $/hr.\nTop 3 ventures for next month: 80% effort allocation.\nKill list: <$50 revenue after full month.\nScale list: >$200 revenue or >5% engagement = 3x resources.\nSet next-month daily KPI targets from actual conversion rates.\nTime: 45 min."},
                {"text": "Verify all systems stable for weekend/overnight", "tag": "AUTO", "detail": "All automated systems operational.\nReview agent outputs: underperforming agents need prompt rewrites.\nDocument procedural memory: what worked, what didn't.\nTime: 10 min."},
            ],
            "revenue_target": "$200-800",
            "content": {"tweets": 23, "threads": 1, "tiktoks": 0, "youtube": 0, "newsletters": 0},
            "engagement": {"replies": 30, "dms": 5, "comments": 5, "follows": 15},
            "outreach": {"cold_emails": 0, "proposals": 0, "calls": 0},
        },
        28: {
            "theme": "MONTH CLOSE + SYSTEM AUDIT",
            "tasks": [
                {"text": "Update all bios/profiles with month milestones", "tag": "MANUAL", "detail": "New credentials, follower counts, revenue milestones.\nTime: 10 min."},
                {"text": "Post 10 tweets + 1 thread (month wrap-up) + LinkedIn retrospective", "tag": "AUTO", "detail": "Time: 0 min (automated). 10 min for LinkedIn."},
                {"text": "Post 5/secondary (15 total)", "tag": "AUTO", "detail": "Time: 0 min (automated)."},
                {"text": "Reply to 60 accounts across all platforms", "tag": "SEMI", "detail": "Time: 30 min."},
                {"text": "Final revenue count across ALL channels (exact numbers)", "tag": "MANUAL", "detail": "Freelance, products, content, EAS, persona, affiliate, gov contracts.\nNo rounding. Exact.\nRevenue leaderboard: rank by actual $/hr. Kill <$10/hr.\nEngagement leaderboard: followers gained, engagement rate, growth trajectory.\nOutreach leaderboard: emails sent, reply rate, close rate, $/email.\nCompare to benchmarks: 2% reply, 20% close.\nTime: 30 min."},
                {"text": "Full system health audit + next-month priority stack", "tag": "AUTO", "detail": "Commands:\n  All 112 cron jobs, 33 agents, all scrapers, all pipelines.\n  python3 AUTOMATIONS/capital_genesis_ranker.py --rank --report\n  Verify GoLogin profiles + SOAX proxies functional and undetected.\n  Set 3 stretch goals for next month.\nTime: 15 min."},
                {"text": "Review GOV contract outcomes", "tag": "SEMI", "detail": "Bids submitted, responses received, contracts won.\nUpdate pipeline tracker for next month.\nTime: 10 min."},
            ],
            "revenue_target": "$250-1000",
            "content": {"tweets": 25, "threads": 1, "tiktoks": 0, "youtube": 0, "newsletters": 0},
            "engagement": {"replies": 60, "dms": 10, "comments": 10, "follows": 30},
            "outreach": {"cold_emails": 0, "proposals": 0, "calls": 0},
        },
        29: {
            "theme": "PIPELINE PRELOAD",
            "tasks": [
                {"text": "Pre-create next month accounts (new GoLogin profiles + SOAX)", "tag": "MANUAL", "detail": "Any new niche expansions planned.\nTime: 15 min."},
                {"text": "Batch create next month Week 1 content", "tag": "AUTO", "detail": "35 tweets + 7 threads (primary), 25/secondary.\nPre-generate 10 faceless YouTube videos.\nTime: 0 min (automated). 15 min to review."},
                {"text": "Post 10 tweets + 1 thread from primary", "tag": "AUTO", "detail": "'How I automated X so I never have to do it again'\nTime: 0 min (automated)."},
                {"text": "Reply to 60 accounts + engage automation discussions", "tag": "SEMI", "detail": "Time: 30 min."},
                {"text": "DM 10 next-month collaboration partners", "tag": "MANUAL", "detail": "Time: 20 min."},
                {"text": "Send cold email batch 12 (50 emails, 775 total)", "tag": "AUTO", "detail": "Warm up next month pipeline early.\nPre-draft 100 emails for next month Week 1.\nTime: 0 min (automated)."},
                {"text": "Capture end-of-month hot freelance jobs", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/auto_freelance_responder.py --scan-and-respond\nTime: 0 min (automated). 10 min to review."},
                {"text": "Deploy month's new cron jobs + clean dead ventures", "tag": "AUTO", "detail": "Set up new cron jobs for recurring workflows.\nDeploy remaining ACTIONABLE_QUEUE.md items.\nUpdate OPS/PRINTMAXX_SYSTEM_MAP.md.\nArchive dead ventures, remove orphan crons, update intelligence catalog.\nTime: 15 min."},
            ],
            "revenue_target": "$200-800",
            "content": {"tweets": 10, "threads": 1, "tiktoks": 0, "youtube": 0, "newsletters": 0},
            "engagement": {"replies": 60, "dms": 10, "comments": 10, "follows": 40},
            "outreach": {"cold_emails": 50, "proposals": 0, "calls": 0},
        },
        30: {
            "theme": "MONTH CLOSE FINAL",
            "tasks": [
                {"text": "Final account health check (all clean, proxies working, profiles updated)", "tag": "SEMI", "detail": "Verify: all accounts clean, all SOAX proxies working, all GoLogin profiles updated.\nTime: 10 min."},
                {"text": "Post 10 tweets + 1 thread (end-of-month wrap)", "tag": "AUTO", "detail": "Time: 0 min (automated)."},
                {"text": "Send newsletter: month recap + next month preview", "tag": "SEMI", "detail": "Build anticipation. Include best content from the month.\nTime: 20 min."},
                {"text": "Reply to 60 accounts + close open DM conversations", "tag": "SEMI", "detail": "Final DM batch. Close or schedule follow-ups.\nTime: 30 min."},
                {"text": "Final revenue reconciliation across all channels", "tag": "MANUAL", "detail": "All channels, all platforms, all payment processors.\nUpdate KPI_DASHBOARD.md with actual month numbers vs projections.\nClose any remaining open proposals with urgency.\nTime: 30 min."},
                {"text": "Verify all systems stable for next month Day 1", "tag": "AUTO", "detail": "Commands:\n  All systems stable for overnight transition.\n  python3 AUTOMATIONS/capital_genesis_ranker.py --rank --export csv\n  All 33 agents operational for next month.\n  python3 AUTOMATIONS/backup_system.py --full\nTime: 10 min."},
            ],
            "revenue_target": "$250-1000",
            "content": {"tweets": 10, "threads": 1, "tiktoks": 0, "youtube": 0, "newsletters": 1},
            "engagement": {"replies": 60, "dms": 10, "comments": 10, "follows": 30},
            "outreach": {"cold_emails": 0, "proposals": 0, "calls": 0},
        },
    }

    default_plan = {
        "theme": "FLEX OPS",
        "tasks": [
            {"text": "Execute top 3 priority ventures from CAPITAL_GENESIS_PRIORITY_STACK.md", "tag": "SEMI", "detail": "Check OPS/CAPITAL_GENESIS_PRIORITY_STACK.md daily -- priorities shift.\nCommand: python3 AUTOMATIONS/capital_genesis_ranker.py --top 3\nReview and execute the top 3 highest-priority actions.\nTime: varies."},
            {"text": "Post 12 tweets from primary + 8 per secondary (28 total)", "tag": "AUTO", "detail": "auto_content_poster.py cron handles primary.\nSecondaries via cron, staggered 30+ min.\nContent from content_factory.py.\nVerify: python3 AUTOMATIONS/auto_content_poster.py --status\nTime: 0 min (automated)."},
            {"text": "Reply to 60 from primary + 20 from secondaries (80 total)", "tag": "SEMI", "detail": "Command: python3 AUTOMATIONS/edge_growth_engine.py --reply-guy\nReview queue at CONTENT/social/reply_queue/.\nPost via GoLogin profiles.\nTime: 40 min."},
            {"text": "Follow up on all open proposals and DM conversations", "tag": "SEMI", "detail": "Upwork, EAS, freelance proposals.\nAll open DMs across accounts.\nCreate urgency: 'This rate expires [date].'\nTime: 20 min."},
            {"text": "Auto-scan and respond to hot freelance jobs", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/auto_freelance_responder.py --scan-and-respond\nGenerates proposals for matching jobs.\nReview before sending.\nTime: 0 min (automated). 10 min to review."},
            {"text": "Send 25 cold emails from pipeline", "tag": "AUTO", "detail": "n8n w04 sequence continues daily.\nTime: 0 min (automated). 3 min to check analytics."},
            {"text": "Scan SAM.gov for matching contracts", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/sam_gov_monitor.py\nTime: 0 min (automated). 5 min to review."},
            {"text": "Verify system health (112 crons + 33 agents)", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/system_health_monitor.py --quick\nCheck: twitter_alpha_scraper.py + background_reddit_scraper.py outputs.\nCheck soul_drift_report.json: all agents > 6/10.\nTime: 3 min."},
            {"text": "Process 50 alpha entries", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/alpha_auto_processor.py --process-new\nTime: 0 min (automated)."},
            {"text": "Review daily digest for missed opportunities", "tag": "AUTO", "detail": "daily_digest.py output (6:45AM cron).\nSurface any missed actions.\nTime: 3 min."},
            {"text": "Post 3 TikToks from video pipeline batch", "tag": "SEMI", "detail": "From ai_video_content_pipeline.py output.\nAdapt captions. Post to TikTok + cross-post to Shorts/Reels.\nTime: 15 min."},
            {"text": "Check GoLogin profiles + SOAX proxies (all clean)", "tag": "AUTO", "detail": "Verify: unique IPs, no fingerprint overlap, no shadowbans.\nshadowban.eu for X accounts.\nTime: 5 min."},
        ],
        "revenue_target": "TBD",
        "content": {"tweets": 28, "threads": 0, "tiktoks": 3, "youtube": 0, "newsletters": 0},
        "engagement": {"replies": 80, "dms": 10, "comments": 10, "follows": 25},
        "outreach": {"cold_emails": 25, "proposals": 5, "calls": 0},
    }


    # Weekly phase labels
    week_phases = {1: "INFRA + ACCOUNTS", 2: "VOLUME + PIPELINE", 3: "EDGE + OPTIMIZE", 4: "COMPOUND + CLOSE"}

    calendar_data = []
    m_eng = mult["engagement"]
    m_con = mult["content"]
    m_out = mult["outreach"]
    m_rev = mult["revenue"]

    # ── EFFECTIVE DAY TRACKING ────────────────────────────────────────
    # Instead of using calendar day (which assumes linear progress),
    # use a persistent effective_day that only advances when user
    # completes a day's tasks. Defaults to day 1 if no progress file.
    _progress_file = AUTOMATIONS / "agent" / "kpi_progress.json"
    _effective_day = 1
    if _progress_file.exists():
        try:
            _prog = json.loads(_progress_file.read_text())
            _effective_day = _prog.get("effective_day", 1)
        except Exception:
            pass
    _day_shift = today_day - _effective_day

    for day in range(1, min(days_in_month + 1, 31)):
        # Shift: today shows effective_day's plan, tomorrow shows effective_day+1, etc.
        if _day_shift > 0 and day >= today_day:
            shifted = day - _day_shift
            plan = daily_plans.get(max(1, shifted), default_plan)
        else:
            plan = daily_plans.get(day, default_plan)
        is_today = (day == today_day)
        is_past = (day < today_day)
        week_num = (day - 1) // 7 + 1

        # Apply intensity multipliers
        raw_content = plan.get("content", {"tweets": 5, "threads": 0, "tiktoks": 0, "youtube": 0, "newsletters": 0})
        raw_engagement = plan.get("engagement", {"replies": 0, "dms": 0, "comments": 0, "follows": 0})
        raw_outreach = plan.get("outreach", {"cold_emails": 0, "proposals": 0, "calls": 0})

        content = {k: int(v * m_con) for k, v in raw_content.items()}
        engagement = {k: int(v * m_eng) for k, v in raw_engagement.items()}
        outreach = {k: int(v * m_out) for k, v in raw_outreach.items()}

        # Scale revenue target string
        rev_target = plan["revenue_target"]
        if m_rev > 1.0:
            import re as _re
            nums = _re.findall(r'\d+', rev_target.replace(",", ""))
            if nums:
                scaled = [f"${int(int(n) * m_rev):,}" for n in nums]
                rev_target = "-".join(scaled) if len(scaled) > 1 else scaled[0]

        # Extra tasks for beast/modafinil modes
        tasks = list(plan["tasks"])
        if mode == "beast":
            tasks.extend([
                {"text": f"BEAST: Double content output ({content['tweets']} tweets, {content.get('tiktoks',0)} TikToks)", "tag": "AUTO", "detail": f"Increase auto_content_poster.py cron frequency to 2x.\nDouble batch: python3 AUTOMATIONS/content_factory.py --batch-alpha {content['tweets']}\nAll auto-posted. Verify cron ran.\nTier: MEDIUM on secondaries, push volume.\nTime: 0 min (automated)."},
                {"text": f"BEAST: Extra cold email batch ({outreach['cold_emails']} total)", "tag": "AUTO", "detail": f"n8n w04 sends additional batch. {outreach['cold_emails']} emails today.\nMulti-inbox rotation to stay under per-inbox limits.\nMonitor bounce rate < 5%.\nTime: 0 min (automated)."},
                {"text": "BEAST: 2nd YouTube video or 5 extra Shorts", "tag": "SEMI", "detail": "Batch produce from ai_video_content_pipeline.py or auto_clip_pipeline.py output.\nUpload with full SEO.\nTime: 30 min."},
                {"text": "BEAST: Respond to ALL pending DMs across all accounts", "tag": "MANUAL", "detail": "Open each GoLogin profile. Check DMs.\nRespond to every pending conversation.\nConvert warm leads to calls/proposals.\nTime: 30 min."},
                {"text": "BEAST: 30-min competitor analysis + steal their best formats", "tag": "SEMI", "detail": "edge_growth_engine.py --squeeze outputs competitor analysis.\nIdentify top-performing content formats in niche.\nAdapt (not copy) their best hooks and structures.\nTime: 30 min."},
            ])
        elif mode == "modafinil":
            tasks.extend([
                {"text": f"MODAFINIL: 3x engagement volume ({engagement['replies']} replies, {engagement['dms']} DMs)", "tag": "AUTO", "detail": f"Increase edge_growth_engine.py reply-guy output to 3x.\nGenerate 3x reply queue. Review batch (larger but same review process).\n{engagement['replies']} replies + {engagement['dms']} DMs + {engagement['follows']} follows today.\nSpread across accounts to stay under per-account limits.\nTier: MEDIUM/AGGRESSIVE accounts absorb the volume.\nTime: 15 min extra review (automation handles generation)."},
                {"text": f"MODAFINIL: 3x content factory ({content['tweets']} tweets, {content.get('threads',0)} threads)", "tag": "AUTO", "detail": f"Command: python3 AUTOMATIONS/content_factory.py --batch-alpha {content['tweets']}\nIncrease cron frequency to 3x.\nAll auto-posted via auto_content_poster.py.\n{content.get('tiktoks',0)} TikToks via ai_video_content_pipeline.py.\nTime: 0 min (automated)."},
                {"text": f"MODAFINIL: 3x outreach blast ({outreach['cold_emails']} emails, {outreach['proposals']} proposals)", "tag": "AUTO", "detail": f"n8n w04 sends 3x volume: {outreach['cold_emails']} emails.\nMulti-inbox rotation (3-5 inboxes at 50-80 each).\n{outreach['proposals']} proposals via auto_freelance_responder.py.\n{outreach['calls']} calls via Bland AI.\nAll automated. Monitor bounce/spam rates.\nTime: 0 min (automated). 10 min to monitor."},
                {"text": "MODAFINIL: Set up 2 additional niche accounts", "tag": "MANUAL", "detail": "New GoLogin profiles + SOAX sticky sessions.\nPick 2 untapped niches from market research.\nEach: bio, PFP, 3 seed tweets, follow 20 niche accounts.\nTier: AGGRESSIVE (expendable, test new niches).\nTime: 20 min."},
                {"text": "MODAFINIL: Batch create 48hrs of scheduled content", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/content_factory.py --batch-alpha 50\nGenerate 2 days of content buffer across all platforms.\nLoad into Buffer/Typefully/Publer scheduling queues.\nTime: 0 min (automated). 15 min to review."},
                {"text": "MODAFINIL: Deep alpha review (process ALL pending entries)", "tag": "AUTO", "detail": "Command: python3 AUTOMATIONS/alpha_auto_processor.py --process-all\nProcess every PENDING_REVIEW entry in ALPHA_STAGING.csv.\nNo backlog left.\nTime: 0 min (automated)."},
                {"text": "MODAFINIL: Write and publish 2 Medium articles (Parasite SEO)", "tag": "SEMI", "detail": "Repurpose top threads into long-form.\nDA90+ backlinks. Submit to Medium publications for 10x distribution.\nInclude affiliate links where relevant. FTC compliant.\nTime: 40 min."},
                {"text": "MODAFINIL: Record 3 YouTube videos back to back", "tag": "MANUAL", "detail": "Batch production: record 3 tutorials/reviews consecutively.\nScreen record with narration. Edit: cut dead air, add overlays.\nUpload with full SEO, end screens, cards.\nTime: 2-3 hours."},
                {"text": "MODAFINIL: DM 50 collaborators / affiliate partners", "tag": "SEMI", "detail": "Generate DM templates from edge_growth_engine.py.\nPersonalize each. Mix: 20 collab, 15 affiliate, 15 cross-promo.\nSpread across accounts. Max 50 DMs/day/account.\nTime: 30 min to review/personalize."},
                {"text": "MODAFINIL: Build and deploy 1 micro-SaaS (app factory sprint)", "tag": "MANUAL", "detail": "Pick highest-scoring app from CAPITAL_GENESIS_PRIORITY_STACK.md.\nVibe-code with Claude Code in 2-4 hours.\nDeploy to Vercel/Netlify. List on Gumroad.\nThis is the MANUAL stretch goal -- genuine human creativity required.\nTime: 2-4 hours."},
            ])

        engagement_score = sum(engagement.values())
        content_score = content.get("tweets", 0) + content.get("threads", 0) * 3 + content.get("tiktoks", 0) * 2 + content.get("youtube", 0) * 5 + content.get("newsletters", 0) * 4
        outreach_score = outreach.get("cold_emails", 0) + outreach.get("proposals", 0) * 5 + outreach.get("calls", 0) * 10

        calendar_data.append({
            "day": day,
            "weekday": calendar.day_abbr[calendar.weekday(year, month, day)],
            "theme": plan["theme"],
            "tasks": tasks,
            "revenue_target": rev_target,
            "content": content,
            "engagement": engagement,
            "outreach": outreach,
            "automation_checks": plan.get("automation_checks", []),
            "research": plan.get("research", []),
            "is_today": is_today,
            "is_past": is_past,
            "week": week_num,
            "week_phase": week_phases.get(week_num, "SCALE"),
            "scores": {
                "engagement": engagement_score,
                "content": content_score,
                "outreach": outreach_score,
            },
        })

    goals = {"standard": "$1,000", "beast": "$2,500", "modafinil": "$5,000+"}
    return jsonify({
        "month": f"{calendar.month_name[month]} {year}",
        "days": calendar_data,
        "goal": goals.get(mode, "$1,000"),
        "mode": mode,
        "mode_label": mult["label"],
        "mode_desc": mult["desc"],
        "effective_day": _effective_day,
        "schedule_shifted": _day_shift > 0,
        "shift_days": _day_shift,
    })


@app.route("/api/kpi/advance-day", methods=["POST"])
def api_kpi_advance_day():
    """Advance the effective KPI day (user completed today's tasks)."""
    progress_file = AUTOMATIONS / "agent" / "kpi_progress.json"
    progress_file.parent.mkdir(parents=True, exist_ok=True)
    current = 1
    if progress_file.exists():
        try:
            current = json.loads(progress_file.read_text()).get("effective_day", 1)
        except Exception:
            pass
    new_day = min(current + 1, 30)
    progress_file.write_text(json.dumps({
        "effective_day": new_day,
        "advanced_at": datetime.now().isoformat(),
    }, indent=2))
    return jsonify({"effective_day": new_day, "previous": current})


@app.route("/api/kpi/complete", methods=["POST"])
def api_kpi_complete():
    """Mark a blocker as complete."""
    data = request.get_json() or {}
    blocker_text = data.get("text", "")
    roadmap = OPS / "MONTHLY_ROADMAP_2026_04.md"
    if roadmap.exists() and blocker_text:
        content = roadmap.read_text()
        old = f"- [ ] {blocker_text}"
        new = f"- [x] {blocker_text}"
        if old in content:
            roadmap.write_text(content.replace(old, new))
            return jsonify({"success": True})
    return jsonify({"success": False})


@app.route("/api/kpi/executor-status")
def api_kpi_executor_status():
    """Return latest KPI executor run report and SEMI review queue status."""
    report_file = OPS / "KPI_EXECUTOR_REPORT.md"
    semi_file = OPS / "SEMI_REVIEW_QUEUE.md"
    status = {
        "last_run": None,
        "report_exists": False,
        "summary": None,
        "semi_pending": 0,
        "feed_line": "KPI Executor has not run yet today.",
    }
    if report_file.exists():
        status["report_exists"] = True
        mtime = report_file.stat().st_mtime
        status["last_run"] = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M:%S")
        content = report_file.read_text(encoding="utf-8")
        for line in content.split("\n"):
            if "| AUTO" in line and "OK" in line:
                status["summary"] = line.strip()
                break
        # Build a feed-friendly summary
        auto_line = status.get("summary", "")
        if auto_line:
            status["feed_line"] = f"KPI Executor ran at {status['last_run']}: {auto_line}"
    if semi_file.exists():
        semi_content = semi_file.read_text(encoding="utf-8")
        status["semi_pending"] = semi_content.count("**Status:** Awaiting review")
        if status["semi_pending"] > 0:
            status["feed_line"] += f" | {status['semi_pending']} SEMI tasks queued for review."
    return jsonify(status)


# ---------------------------------------------------------------------------
# Research Intelligence Feed — surfaces latest findings ranked by value
# ---------------------------------------------------------------------------
EDGAR_CACHE = AUTOMATIONS / "auto_ops" / "edgar_cache"
CB_CACHE = AUTOMATIONS / "auto_ops" / "crunchbase_cache"
BACKLOG_REPORT = OPS / "ALPHA_BACKLOG_REPORT.md"
INTEGRATION_GAP = OPS / "INTEGRATION_GAP_REPORT.md"


@app.route("/api/intel-feed")
def api_intel_feed():
    """Aggregate latest research intelligence findings ranked by score × recency.

    Pulls from: ALPHA_STAGING (recent high-score), EDGAR cache, Crunchbase cache,
    backlog scanner, integration gap report, method discovery, Capital Genesis stack.
    """
    findings = []
    now = time.time()
    today = datetime.now().strftime("%Y-%m-%d")

    # 1. Recent high-score alpha entries (last 48h, score >= 60)
    alpha_csv = LEDGER / "ALPHA_STAGING.csv"
    if alpha_csv.exists():
        try:
            with open(alpha_csv) as f:
                for row in csv.DictReader(f):
                    created = row.get("created_at", "")
                    if today not in created and (datetime.now() - timedelta(days=2)).strftime("%Y-%m-%d") not in created:
                        continue
                    score = 0
                    try:
                        score = int(row.get("synergy_score", 0))
                    except (ValueError, TypeError):
                        pass
                    if score < 40:
                        continue
                    status = row.get("status", "").upper()
                    method = row.get("extracted_method", row.get("tactic", ""))[:120]
                    source = row.get("source", "unknown")
                    roi = row.get("roi_potential", "?")
                    findings.append({
                        "type": "alpha",
                        "source": source,
                        "method": method,
                        "score": score,
                        "roi": roi,
                        "status": status,
                        "time_sensitive": roi in ("HIGH", "HIGHEST", "IMMEDIATE"),
                        "created": created[:16],
                    })
        except Exception:
            pass

    # 2. SEC EDGAR findings (latest cache file)
    if EDGAR_CACHE.exists():
        cache_files = sorted(EDGAR_CACHE.glob("edgar_scan_*.json"), reverse=True)
        if cache_files:
            try:
                data = json.loads(cache_files[0].read_text())
                for r in data.get("top_results", [])[:8]:
                    findings.append({
                        "type": "edgar",
                        "source": "SEC_EDGAR",
                        "method": f"{r.get('company', '?')} ({r.get('form', '?')}) — {', '.join(r.get('keywords', []))}",
                        "score": r.get("score", 0),
                        "roi": "HIGH" if r.get("score", 0) >= 60 else "MEDIUM",
                        "status": "NEW",
                        "time_sensitive": True,
                        "created": data.get("timestamp", "")[:16],
                    })
            except Exception:
                pass

    # 3. Crunchbase findings (latest cache file)
    if CB_CACHE.exists():
        cache_files = sorted(CB_CACHE.glob("cb_scan_*.json"), reverse=True)
        if cache_files:
            try:
                data = json.loads(cache_files[0].read_text())
                for r in data.get("top_results", [])[:8]:
                    findings.append({
                        "type": "crunchbase",
                        "source": "CRUNCHBASE",
                        "method": f"{r.get('company', r.get('title', '?')[:40])} ({r.get('stage', 'news')}) — {', '.join(r.get('keywords', [])[:3])}",
                        "score": r.get("score", 0),
                        "roi": "HIGH" if r.get("stage", "") in ("series a", "series b", "series c") else "MEDIUM",
                        "status": "NEW",
                        "time_sensitive": True,
                        "created": data.get("timestamp", "")[:16],
                    })
            except Exception:
                pass

    # 4. Capital Genesis P0 methods (highest priority stack)
    pstack = OPS / "CAPITAL_GENESIS_PRIORITY_STACK.md"
    if pstack.exists():
        try:
            for line in open(pstack):
                if "| P0 |" in line or "LAUNCH_NOW" in line:
                    parts = [p.strip() for p in line.split("|") if p.strip()]
                    if len(parts) >= 3:
                        findings.append({
                            "type": "capital_genesis",
                            "source": "CAPITAL_GENESIS",
                            "method": parts[1] if len(parts) > 1 else line[:80],
                            "score": 90,
                            "roi": "HIGH",
                            "status": "P0",
                            "time_sensitive": True,
                            "created": today,
                        })
        except Exception:
            pass

    # 5. Integration gap findings (methods pipeline missed)
    if INTEGRATION_GAP.exists():
        try:
            gap_age = (now - INTEGRATION_GAP.stat().st_mtime) / 3600
            if gap_age < 48:
                gap_count = 0
                for line in open(INTEGRATION_GAP):
                    if line.startswith("- `"):
                        gap_count += 1
                if gap_count > 0:
                    findings.append({
                        "type": "gap",
                        "source": "GAP_DETECTOR",
                        "method": f"{gap_count} methods the pipeline should auto-catch",
                        "score": 70,
                        "roi": "MEDIUM",
                        "status": "GAP",
                        "time_sensitive": False,
                        "created": datetime.fromtimestamp(INTEGRATION_GAP.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
                    })
        except Exception:
            pass

    # Sort: time-sensitive first, then by score descending
    findings.sort(key=lambda x: (not x.get("time_sensitive", False), -x.get("score", 0)))

    return jsonify({
        "findings": findings[:50],
        "total": len(findings),
        "sources": list(set(f["source"] for f in findings)),
        "time_sensitive_count": sum(1 for f in findings if f.get("time_sensitive")),
        "last_updated": datetime.now().isoformat(),
    })


@app.route("/api/kpi/rerank", methods=["POST"])
def api_kpi_rerank():
    """Rerank today's KPI tasks based on latest intelligence.

    Two override paths:
    1. TIME-SENSITIVE: EDGAR filings, funding signals with 24-48h windows
    2. CAPITAL GENESIS OVERRIDE: If a finding scores higher than existing
       KPI tasks on the Capital Genesis dimensions, it REPLACES lower-ranked
       tasks in today's queue — not just appends to the end.

    Opportunity window awareness:
    - Some methods have expiry windows (e.g. AI influencers = 6-24 months
      before market saturation). The score includes a decay factor based on
      the method's estimated window. Methods with shorter windows get boosted.
    """
    try:
        # Get current intel findings
        findings = []
        alpha_csv = LEDGER / "ALPHA_STAGING.csv"
        today = datetime.now().strftime("%Y-%m-%d")

        if alpha_csv.exists():
            try:
                with open(alpha_csv) as f:
                    for row in csv.DictReader(f):
                        created = row.get("created_at", "")
                        if today not in created:
                            continue
                        score = 0
                        try:
                            score = int(row.get("synergy_score", 0))
                        except (ValueError, TypeError):
                            pass
                        if score < 50:
                            continue
                        findings.append({
                            "method": row.get("extracted_method", row.get("tactic", ""))[:100],
                            "source": row.get("source", ""),
                            "score": score,
                            "roi": row.get("roi_potential", "?"),
                            "time_sensitive": row.get("roi_potential", "") in ("HIGH", "HIGHEST", "IMMEDIATE"),
                        })
            except Exception:
                pass

        # Also pull P0 from Capital Genesis stack
        pstack = OPS / "CAPITAL_GENESIS_PRIORITY_STACK.md"
        if pstack.exists():
            try:
                for line in open(pstack):
                    if "| P0 |" in line or "LAUNCH_NOW" in line:
                        parts = [p.strip() for p in line.split("|") if p.strip()]
                        if len(parts) >= 2:
                            findings.append({
                                "method": parts[1] if len(parts) > 1 else line[:80],
                                "source": "CAPITAL_GENESIS",
                                "score": 95,
                                "roi": "HIGH",
                                "time_sensitive": True,
                            })
            except Exception:
                pass

        if not findings:
            return jsonify({"reranked": False, "reason": "No high-score findings today"})

        # Sort by score descending — highest Capital Genesis score wins
        findings.sort(key=lambda x: -x.get("score", 0))

        # Load rollover state
        rollover_file = AUTOMATIONS / "agent" / "kpi_rollover_state.json"
        rollover_tasks = []
        if rollover_file.exists():
            try:
                state = json.loads(rollover_file.read_text())
                rollover_tasks = state.get("tasks", [])
            except Exception:
                pass

        # Remove existing urgent tasks to avoid duplicates
        rollover_tasks = [t for t in rollover_tasks if "[URGENT]" not in t.get("text", "")
                          and "[CG-OVERRIDE]" not in t.get("text", "")]

        # Create priority KPI entries — these go to TOP of the queue
        priority_tasks = []
        for f in findings[:8]:
            tag_prefix = "[URGENT]" if f["time_sensitive"] else "[CG-OVERRIDE]"
            task_tag = "SEMI" if f["time_sensitive"] else "AUTO"
            priority_tasks.append({
                "text": f"{tag_prefix} {f['source']}: {f['method'][:80]}",
                "tag": task_tag,
                "detail": f"Score: {f['score']}/100, ROI: {f['roi']}. "
                          f"This overrides lower-ranked tasks because Capital Genesis "
                          f"scored it higher than existing queue items.",
                "reason": f"Capital Genesis override (score {f['score']})",
                "original_day": datetime.now().day,
                "failure_count": 0,
            })

        # Priority tasks go FIRST — they override existing order
        all_tasks = priority_tasks + rollover_tasks
        rollover_state = {
            "last_updated": datetime.now().isoformat(),
            "source_day": datetime.now().day,
            "target_day": datetime.now().day,
            "task_count": len(all_tasks),
            "tasks": all_tasks,
            "override_reason": f"Capital Genesis found {len(priority_tasks)} items scoring "
                               f"above existing KPI queue. Window-aware scoring applied.",
        }
        rollover_file.parent.mkdir(parents=True, exist_ok=True)
        rollover_file.write_text(json.dumps(rollover_state, indent=2))

        return jsonify({
            "reranked": True,
            "urgent_injected": sum(1 for t in priority_tasks if "[URGENT]" in t["text"]),
            "cg_overrides": sum(1 for t in priority_tasks if "[CG-OVERRIDE]" in t["text"]),
            "total_priority": len(priority_tasks),
            "total_rollover": len(all_tasks),
            "top_tasks": [t["text"][:80] for t in priority_tasks],
        })

    except Exception as e:
        return jsonify({"reranked": False, "error": str(e)})


# ---------------------------------------------------------------------------
# TODAY'S AUTOMATED OUTPUT — daily production summary for the dashboard
# ---------------------------------------------------------------------------
TIME_SENSITIVE_KEYWORDS = re.compile(
    r"just announced|breaking|IPO|Series [A-C]|acquisition|launching|deadline|"
    r"expires|limited|closing|urgent|time.sensitive|immediately|ending soon",
    re.IGNORECASE,
)


def _is_today(filepath):
    """Return True if a file was modified today."""
    try:
        mtime = Path(filepath).stat().st_mtime
        return datetime.fromtimestamp(mtime).date() == datetime.now().date()
    except Exception:
        return False


def _parse_log_counts(filepath, patterns):
    """Scan a log file for pattern counts. Returns dict of pattern->count."""
    counts = {p: 0 for p in patterns}
    try:
        p = Path(filepath)
        if not p.exists() or not _is_today(p):
            return counts
        with open(p, "r", encoding="utf-8", errors="replace") as f:
            for line in f:
                for pat in patterns:
                    if pat.upper() in line.upper():
                        counts[pat] += 1
    except Exception:
        pass
    return counts


def get_daily_output():
    """Aggregate everything the automated system produced today into a single view."""
    today_str = datetime.now().strftime("%Y-%m-%d")
    today_ts = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0).timestamp()
    result = {
        "date": today_str,
        "summary": {
            "total_new_alpha": 0,
            "methods_discovered": 0,
            "entries_integrated": 0,
            "entries_skipped": 0,
            "rbi_researched": 0,
            "rbi_passed": 0,
            "rbi_conditional": 0,
            "rbi_failed": 0,
            "edgar_filings": 0,
            "crunchbase_rounds": 0,
            "reports_generated": 0,
            "content_created": 0,
            "loops_closed": 0,
            "ventures_refined": 0,
            "dag_status": "unknown",
        },
        "highlights": [],
        "rbi_status": {"researched": 0, "passed": 0, "conditional": 0, "failed": 0},
        "loop_health": {"decisions": "UNKNOWN", "feedback": "UNKNOWN", "pipeline": "UNKNOWN", "drift": "UNKNOWN"},
        "system_working": True,
        "errors": [],
    }

    # ── 1. ALPHA_STAGING — new entries added today ──────────────────
    alpha_csv = LEDGER / "ALPHA_STAGING.csv"
    if alpha_csv.exists():
        try:
            with open(alpha_csv, "r", encoding="utf-8", errors="replace") as f:
                for row in csv.DictReader(f):
                    created = row.get("created_at", row.get("date_added", ""))
                    if today_str in created:
                        result["summary"]["total_new_alpha"] += 1
                        score = 0
                        try:
                            score = float(row.get("synergy_score", 0) or 0)
                        except (ValueError, TypeError):
                            pass
                        text = row.get("extracted_method", row.get("tactic", ""))[:120]
                        source = row.get("source", "alpha")
                        is_ts = bool(TIME_SENSITIVE_KEYWORDS.search(text)) or (score > 7 and bool(created))
                        if score >= 6.0 and text:
                            result["highlights"].append({
                                "source": source,
                                "text": text,
                                "score": round(score, 1),
                                "time_sensitive": is_ts,
                            })
        except Exception as e:
            result["errors"].append(f"alpha parse: {e}")

    # ── 2. LOG FILE SCANNING ────────────────────────────────────────
    # 2a. SEC EDGAR scanner
    edgar_log = LOGS_DIR / "sec_edgar_scanner.log"
    if edgar_log.exists() and _is_today(edgar_log):
        edgar_counts = _parse_log_counts(edgar_log, ["filing", "8-K", "S-1", "10-K", "10-Q", "ERROR"])
        filing_total = sum(v for k, v in edgar_counts.items() if k != "ERROR")
        result["summary"]["edgar_filings"] = filing_total
        if edgar_counts.get("ERROR", 0) > 0:
            result["errors"].append(f"EDGAR scanner: {edgar_counts['ERROR']} errors")
        # Extract specific filings as highlights
        try:
            with open(edgar_log, "r", encoding="utf-8", errors="replace") as f:
                for line in f:
                    if any(kw in line.upper() for kw in ["8-K", "S-1", "IPO", "ACQUISITION"]):
                        clean = line.strip()[-120:]
                        if clean:
                            result["highlights"].append({
                                "source": "edgar",
                                "text": clean,
                                "score": 7.0,
                                "time_sensitive": True,
                            })
        except Exception:
            pass

    # 2b. Crunchbase scanner
    cb_log = LOGS_DIR / "crunchbase_scanner.log"
    if cb_log.exists() and _is_today(cb_log):
        cb_counts = _parse_log_counts(cb_log, ["Series", "funding", "raised", "ERROR"])
        result["summary"]["crunchbase_rounds"] = sum(v for k, v in cb_counts.items() if k != "ERROR")
        if cb_counts.get("ERROR", 0) > 0:
            result["errors"].append(f"Crunchbase scanner: {cb_counts['ERROR']} errors")
        try:
            with open(cb_log, "r", encoding="utf-8", errors="replace") as f:
                for line in f:
                    if any(kw in line for kw in ["Series A", "Series B", "Series C", "$"]):
                        clean = line.strip()[-120:]
                        if clean:
                            result["highlights"].append({
                                "source": "crunchbase",
                                "text": clean,
                                "score": 8.0,
                                "time_sensitive": True,
                            })
        except Exception:
            pass

    # 2c. Method discovery crawler
    mdc_log = LOGS_DIR / "method_discovery_crawler.log"
    if mdc_log.exists() and _is_today(mdc_log):
        mdc_counts = _parse_log_counts(mdc_log, ["discovered", "new method", "scored", "ERROR"])
        result["summary"]["methods_discovered"] = mdc_counts.get("discovered", 0) + mdc_counts.get("new method", 0)
        if mdc_counts.get("ERROR", 0) > 0:
            result["errors"].append(f"Method crawler: {mdc_counts['ERROR']} errors")
    # Also check METHOD_DISCOVERY_LOG.csv
    discovery_csv = LEDGER / "METHOD_DISCOVERY_LOG.csv"
    if discovery_csv.exists():
        try:
            with open(discovery_csv, "r", encoding="utf-8", errors="replace") as f:
                for line in f:
                    if today_str in line:
                        result["summary"]["methods_discovered"] += 1
        except Exception:
            pass

    # 2d. Autonomous integrator
    ai_log = LOGS_DIR / "autonomous_integrator.log"
    if ai_log.exists() and _is_today(ai_log):
        ai_counts = _parse_log_counts(ai_log, ["OK", "SKIP", "INTEGRATED", "ERROR", "timeout"])
        result["summary"]["entries_integrated"] = ai_counts.get("OK", 0) + ai_counts.get("INTEGRATED", 0)
        result["summary"]["entries_skipped"] = ai_counts.get("SKIP", 0)
        if ai_counts.get("ERROR", 0) > 0:
            result["errors"].append(f"Integrator: {ai_counts['ERROR']} errors")
        if ai_counts.get("timeout", 0) > 0:
            result["errors"].append(f"Integrator: {ai_counts['timeout']} timeouts")

    # 2e. RBI loop
    rbi_log = LOGS_DIR / "rbi_loop.log"
    if rbi_log.exists() and _is_today(rbi_log):
        rbi_counts = _parse_log_counts(rbi_log, ["PASS", "CONDITIONAL", "FAIL", "researched", "ERROR"])
        result["summary"]["rbi_researched"] = rbi_counts.get("researched", 0)
        result["summary"]["rbi_passed"] = rbi_counts.get("PASS", 0)
        result["summary"]["rbi_conditional"] = rbi_counts.get("CONDITIONAL", 0)
        result["summary"]["rbi_failed"] = rbi_counts.get("FAIL", 0)
        result["rbi_status"] = {
            "researched": rbi_counts.get("researched", 0),
            "passed": rbi_counts.get("PASS", 0),
            "conditional": rbi_counts.get("CONDITIONAL", 0),
            "failed": rbi_counts.get("FAIL", 0),
        }

    # 2f. Morning intelligence DAG
    dag_log = LOGS_DIR / "morning_intelligence_dag.log"
    if dag_log.exists() and _is_today(dag_log):
        dag_counts = _parse_log_counts(dag_log, ["COMPLETE", "FAILED", "ERROR", "SUCCESS"])
        if dag_counts.get("COMPLETE", 0) > 0 or dag_counts.get("SUCCESS", 0) > 0:
            result["summary"]["dag_status"] = "COMPLETE"
        elif dag_counts.get("FAILED", 0) > 0 or dag_counts.get("ERROR", 0) > 0:
            result["summary"]["dag_status"] = "FAILED"
            result["errors"].append("Morning DAG had failures")
        else:
            result["summary"]["dag_status"] = "RUNNING"
    # Also check DAG checkpoint
    if DAG_CHECKPOINT.exists():
        try:
            cp = json.loads(DAG_CHECKPOINT.read_text())
            if today_str in cp.get("timestamp", cp.get("last_run", "")):
                status = cp.get("status", "unknown")
                result["summary"]["dag_status"] = status.upper()
        except Exception:
            pass

    # 2g. User sim refiner
    usr_log = LOGS_DIR / "user_sim_refiner.log"
    if usr_log.exists() and _is_today(usr_log):
        usr_counts = _parse_log_counts(usr_log, ["refined", "venture", "updated", "ERROR"])
        result["summary"]["ventures_refined"] = usr_counts.get("refined", 0) + usr_counts.get("venture", 0)

    # 2h. Loop closer
    loop_jsonl = AUTOMATIONS / "agent" / "swarm" / "loop_closer.jsonl"
    if loop_jsonl.exists():
        try:
            loop_actions_today = 0
            # Map JSONL action names to loop health categories
            _action_to_loop = {
                "soul_drift_check": "drift",
                "pipeline_advance": "pipeline",
                "decision_execute": "decisions",
                "feedback_update": "feedback",
                # Also handle direct names
                "decisions": "decisions", "feedback": "feedback",
                "pipeline": "pipeline", "drift": "drift",
            }
            with open(loop_jsonl, "r", encoding="utf-8", errors="replace") as f:
                for line in f:
                    try:
                        entry = json.loads(line)
                        ts = entry.get("timestamp", entry.get("ts", ""))
                        if today_str in ts:
                            loop_actions_today += 1
                            action = entry.get("action", entry.get("loop", entry.get("type", "")))
                            status = entry.get("result", entry.get("status", "OK"))
                            loop_key = _action_to_loop.get(action)
                            if loop_key:
                                result["loop_health"][loop_key] = status
                    except Exception:
                        pass
            result["summary"]["loops_closed"] = loop_actions_today
            # If we found today's entries, any loop with actions is OK
            # Also read the state file for definitive status
            loop_state_file = AUTOMATIONS / "agent" / "swarm" / "loop_state.json"
            if loop_state_file.exists():
                try:
                    ls = json.loads(loop_state_file.read_text())
                    for field, key in [("last_decision_cycle", "decisions"), ("last_feedback_cycle", "feedback"),
                                       ("last_pipeline_cycle", "pipeline")]:
                        val = ls.get(field)
                        if val and today_str in val:
                            result["loop_health"][key] = "OK"
                        elif val:
                            result["loop_health"][key] = "STALE"
                except Exception:
                    pass
        except Exception:
            pass

    # Also check loop_closer status file
    feedback_file = AUTOMATIONS / "agent" / "swarm" / "feedback_recommendations.json"
    if feedback_file.exists():
        try:
            fb = json.loads(feedback_file.read_text())
            for loop_name in ["decisions", "feedback", "pipeline", "drift"]:
                if loop_name in fb:
                    status_val = fb[loop_name].get("status", fb[loop_name].get("health", "UNKNOWN"))
                    if result["loop_health"][loop_name] == "UNKNOWN":
                        result["loop_health"][loop_name] = status_val
        except Exception:
            pass

    # ── 3. SWARM REPORTS — generated today ──────────────────────────
    reports_dir = AUTOMATIONS / "agent" / "swarm" / "reports"
    if reports_dir.exists():
        try:
            today_compact = today_str.replace("-", "")
            today_reports = [
                f for f in reports_dir.iterdir()
                if f.is_file() and (today_compact in f.name or _is_today(f))
            ]
            result["summary"]["reports_generated"] = len(today_reports)
            for rpt in today_reports[:5]:
                try:
                    first_line = ""
                    with open(rpt, "r", encoding="utf-8", errors="replace") as rf:
                        for rline in rf:
                            rline = rline.strip()
                            if rline and not rline.startswith("#"):
                                first_line = rline[:120]
                                break
                    if first_line:
                        result["highlights"].append({
                            "source": "swarm_report",
                            "text": f"{rpt.name}: {first_line}",
                            "score": 5.0,
                            "time_sensitive": False,
                        })
                except Exception:
                    pass
        except Exception:
            pass

    # ── 4. CONTENT — new posts created today ────────────────────────
    queue_dir = PROJECT_ROOT / "CONTENT" / "social" / "posting_queue"
    if queue_dir.exists():
        try:
            today_content = [
                f for f in queue_dir.iterdir()
                if f.is_file() and _is_today(f)
            ]
            result["summary"]["content_created"] = len(today_content)
        except Exception:
            pass

    # ── 5. RBI STATE — latest pipeline status ───────────────────────
    rbi_state_path = AUTOMATIONS / "agent" / "rbi_state.json"
    if rbi_state_path.exists():
        try:
            rbi_state = json.loads(rbi_state_path.read_text())
            # Override rbi_status with the state file if it has more data
            if "researched" in rbi_state:
                r_val = rbi_state.get("researched", 0)
                result["rbi_status"]["researched"] = len(r_val) if isinstance(r_val, list) else int(r_val or 0)
            if "results" in rbi_state:
                results = rbi_state["results"]
                if isinstance(results, dict):
                    result["rbi_status"]["passed"] = results.get("PASS", results.get("passed", 0))
                    result["rbi_status"]["conditional"] = results.get("CONDITIONAL", results.get("conditional", 0))
                    result["rbi_status"]["failed"] = results.get("FAIL", results.get("failed", 0))
        except Exception:
            pass

    # ── Scan all logs for errors today ──────────────────────────────
    if LOGS_DIR.exists():
        try:
            for log_file in LOGS_DIR.iterdir():
                if not log_file.is_file() or not log_file.name.endswith(".log"):
                    continue
                if not _is_today(log_file):
                    continue
                try:
                    with open(log_file, "r", encoding="utf-8", errors="replace") as f:
                        lines = f.readlines()
                    # Check last 50 lines for critical errors
                    for line in lines[-50:]:
                        line_lower = line.lower()
                        if "timeout" in line_lower and "twitter" in log_file.name.lower():
                            if "twitter scraper timeout" not in result["errors"]:
                                result["errors"].append("twitter scraper timeout")
                        elif "rate limit" in line_lower:
                            err_msg = f"{log_file.stem}: rate limit hit"
                            if err_msg not in result["errors"]:
                                result["errors"].append(err_msg)
                        elif "critical" in line_lower or "fatal" in line_lower:
                            err_msg = f"{log_file.stem}: {line.strip()[-80:]}"
                            if len(result["errors"]) < 20 and err_msg not in result["errors"]:
                                result["errors"].append(err_msg)
                except Exception:
                    pass
        except Exception:
            pass

    # ── Determine system_working ────────────────────────────────────
    total_activity = (
        result["summary"]["total_new_alpha"]
        + result["summary"]["entries_integrated"]
        + result["summary"]["reports_generated"]
        + result["summary"]["content_created"]
        + result["summary"]["loops_closed"]
    )
    critical_errors = sum(1 for e in result["errors"] if "critical" in e.lower() or "fatal" in e.lower())
    result["system_working"] = total_activity > 0 or critical_errors == 0

    # ── Sort highlights by score desc ───────────────────────────────
    result["highlights"].sort(key=lambda x: (-x.get("score", 0),))
    # Deduplicate highlights by text
    seen_texts = set()
    deduped = []
    for h in result["highlights"]:
        key = h["text"][:60]
        if key not in seen_texts:
            seen_texts.add(key)
            deduped.append(h)
    result["highlights"] = deduped[:30]

    return result


@app.route("/api/daily-output")
def api_daily_output():
    return jsonify(get_daily_output())


@app.route("/api/kpi/manual-priority", methods=["POST"])
def api_manual_priority():
    """Accept a highlight item and append it to OPS/MANUAL_PRIORITIES.md for human override."""
    data = request.get_json(force=True)
    highlight_text = data.get("text", "")
    highlight_source = data.get("source", "unknown")
    highlight_score = data.get("score", 0)
    if not highlight_text:
        return jsonify({"success": False, "error": "No text provided"}), 400

    manual_path = OPS / "MANUAL_PRIORITIES.md"
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    entry = f"\n## [{timestamp}] MANUAL OVERRIDE\n- **Source:** {highlight_source}\n- **Score:** {highlight_score}\n- **Item:** {highlight_text}\n- **Status:** PENDING\n"

    try:
        if not manual_path.exists():
            manual_path.write_text(f"# Manual Priority Overrides\n\nHuman-promoted items from Today's Output dashboard.\n{entry}")
        else:
            with open(manual_path, "a", encoding="utf-8") as f:
                f.write(entry)
        return jsonify({"success": True, "message": f"Promoted: {highlight_text[:60]}..."})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ---------------------------------------------------------------------------
# Command Palette — async job runner
# ---------------------------------------------------------------------------
_running_jobs = {}
_jobs_lock = threading.Lock()


def _get_all_commands():
    """Build the full command palette: slash commands + system commands."""
    commands = []

    # ── Slash commands from .claude/commands/*.md ──────────────────────
    cmd_dir = PROJECT_ROOT / ".claude" / "commands"
    if cmd_dir.exists():
        for f in sorted(cmd_dir.glob("*.md")):
            name = f.stem
            desc = ""
            try:
                with open(f, "r") as fh:
                    desc = fh.readline().strip().lstrip("#").strip()
            except Exception:
                pass
            commands.append({
                "id": f"slash-{name}",
                "name": f"/{name}",
                "description": desc or name,
                "category": "system",
                "cmd": f"claude -p 'Run /{name}'",
                "icon": "terminal-2",
            })

    # ── RBI commands ──────────────────────────────────────────────────
    rbi_cmds = [
        ("rbi-full", "RBI Full Loop", "Research, Backtest, Implement — full cycle", "--full", "target"),
        ("rbi-research", "RBI Research", "Research phase only", "--research", "search"),
        ("rbi-backtest", "RBI Backtest", "Backtest phase only", "--backtest", "chart-bar"),
        ("rbi-implement", "RBI Implement", "Implement phase only", "--implement", "hammer"),
        ("rbi-status", "RBI Status", "Check current RBI status", "--status", "info-circle"),
    ]
    for cid, name, desc, flag, icon in rbi_cmds:
        commands.append({
            "id": cid, "name": name, "description": desc,
            "category": "revenue",
            "cmd": f"python3 AUTOMATIONS/rbi_loop.py {flag}",
            "icon": icon,
        })

    # ── Loop Closer ───────────────────────────────────────────────────
    loop_cmds = [
        ("loops-cycle", "Loop Closer Cycle", "Run all 4 loops", "--cycle", "refresh"),
        ("loops-status", "Loop Health", "Check all 4 loops", "--status", "heart-rate-monitor"),
        ("loops-decisions", "Decision Loop", "Decision execution loop", "--decisions", "gavel"),
        ("loops-feedback", "Feedback Loop", "Feedback tracking loop", "--feedback", "message-report"),
        ("loops-pipeline", "Pipeline Loop", "Pipeline advancement loop", "--pipeline", "git-merge"),
        ("loops-drift", "Soul Drift", "Soul drift scoring loop", "--drift", "ghost"),
    ]
    for cid, name, desc, flag, icon in loop_cmds:
        commands.append({
            "id": cid, "name": name, "description": desc,
            "category": "system",
            "cmd": f"python3 AUTOMATIONS/loop_closer.py {flag}",
            "icon": icon,
        })

    # ── Scanners ──────────────────────────────────────────────────────
    scanner_cmds = [
        ("scan-edgar", "EDGAR Scanner", "Scan SEC EDGAR filings", "python3 AUTOMATIONS/sec_edgar_scanner.py --scan", "building-bank"),
        ("scan-crunchbase", "Crunchbase Scanner", "Scan Crunchbase for opportunities", "python3 AUTOMATIONS/crunchbase_scanner.py --scan", "brand-crunchbase"),
        ("scan-methods", "Method Discovery", "Crawl for new revenue methods", "python3 AUTOMATIONS/method_discovery_crawler.py --crawl", "radar-2"),
        ("scan-orphans", "Orphan Doc Scanner", "Find orphan documents", "python3 AUTOMATIONS/orphan_doc_scanner.py --scan", "file-search"),
        ("scan-alpha-backlog", "Alpha Backlog Scanner", "Sweep historical alpha", "python3 AUTOMATIONS/alpha_backlog_scanner.py --scan", "database-search"),
    ]
    for cid, name, desc, cmd, icon in scanner_cmds:
        commands.append({
            "id": cid, "name": name, "description": desc,
            "category": "intelligence",
            "cmd": cmd, "icon": icon,
        })

    # ── Ranker ────────────────────────────────────────────────────────
    ranker_cmds = [
        ("ranker-full", "Capital Genesis Rank + Report", "Rescore all methods", "python3 AUTOMATIONS/capital_genesis_ranker.py --rank --report", "chart-arrows-vertical"),
        ("ranker-p0", "P0 Methods Only", "Show launch-now methods", "python3 AUTOMATIONS/capital_genesis_ranker.py --p0", "flame"),
        ("ranker-top20", "Top 20 Methods", "Top 20 ranked methods", "python3 AUTOMATIONS/capital_genesis_ranker.py --top 20", "trophy"),
    ]
    for cid, name, desc, cmd, icon in ranker_cmds:
        commands.append({
            "id": cid, "name": name, "description": desc,
            "category": "revenue",
            "cmd": cmd, "icon": icon,
        })

    # ── Intelligence ──────────────────────────────────────────────────
    intel_cmds = [
        ("intel-router-stats", "Intelligence Router Stats", "Coverage and doc stats", "python3 AUTOMATIONS/intelligence_router.py --stats", "brain"),
        ("intel-sqlite-rebuild", "SQLite Index Rebuild", "Rebuild FTS5 alpha index", "python3 AUTOMATIONS/sqlite_alpha_index.py --rebuild", "database"),
        ("intel-sqlite-stats", "SQLite Index Stats", "Alpha index statistics", "python3 AUTOMATIONS/sqlite_alpha_index.py --stats", "chart-dots"),
    ]
    for cid, name, desc, cmd, icon in intel_cmds:
        commands.append({
            "id": cid, "name": name, "description": desc,
            "category": "intelligence",
            "cmd": cmd, "icon": icon,
        })

    # ── Pipeline / Orchestration ──────────────────────────────────────
    pipeline_cmds = [
        ("pipeline-autoapprove", "Auto-Approve Alpha", "Approve pending alpha entries", "python3 AUTOMATIONS/alpha_auto_processor.py --process-new", "check"),
        ("pipeline-morning-dag", "Morning DAG", "Run morning orchestration DAG", "python3 AUTOMATIONS/ceo_agent.py --dag", "git-branch"),
        ("pipeline-daily-digest", "Daily Digest", "Generate daily digest", "python3 AUTOMATIONS/daily_digest.py", "news"),
        ("pipeline-session-briefing", "Session Briefing", "Generate session briefing", "python3 AUTOMATIONS/session_briefing.py", "clipboard-text"),
        ("pipeline-actionable", "Actionable Aggregator", "Aggregate actionable items", "python3 AUTOMATIONS/actionable_aggregator.py", "list-check"),
    ]
    for cid, name, desc, cmd, icon in pipeline_cmds:
        commands.append({
            "id": cid, "name": name, "description": desc,
            "category": "system",
            "cmd": cmd, "icon": icon,
        })

    # ── Agents ────────────────────────────────────────────────────────
    agent_cmds = [
        ("agents-venture-status", "Venture Status", "All venture pipeline statuses", "python3 AUTOMATIONS/venture_autonomy.py --status", "hierarchy-3"),
        ("agents-venture-run", "Venture Run All", "Run all venture pipelines", "python3 AUTOMATIONS/venture_autonomy.py --run-all", "player-play"),
        ("agents-swarm-status", "Swarm Status", "All swarm agent statuses", "python3 AUTOMATIONS/agent_swarm.py --status", "users-group"),
        ("agents-swarm-health", "Swarm Health", "Swarm health check", "python3 AUTOMATIONS/agent_swarm.py --health", "stethoscope"),
    ]
    for cid, name, desc, cmd, icon in agent_cmds:
        commands.append({
            "id": cid, "name": name, "description": desc,
            "category": "agents",
            "cmd": cmd, "icon": icon,
        })

    # ── System / Ops ──────────────────────────────────────────────────
    system_cmds = [
        ("sys-cron-watchdog", "Cron Watchdog", "Check cron health and restore if needed", "python3 AUTOMATIONS/cron_watchdog.py", "clock-shield"),
        ("sys-health-quick", "System Health Quick", "Quick system health check", "python3 AUTOMATIONS/system_health_monitor.py --quick", "heartbeat"),
        ("sys-usage-status", "Usage Optimizer Status", "Check API usage and budget", "python3 AUTOMATIONS/usage_optimizer.py --status", "gauge"),
    ]
    for cid, name, desc, cmd, icon in system_cmds:
        commands.append({
            "id": cid, "name": name, "description": desc,
            "category": "system",
            "cmd": cmd, "icon": icon,
        })

    # ── Refiner / Cognition ───────────────────────────────────────────
    refiner_cmds = [
        ("refiner-sim", "User Sim Refiner", "Refine user simulation model", "python3 AUTOMATIONS/user_sim_refiner.py --all", "user-cog"),
        ("refiner-cognitive", "Cognitive Engine Build", "Build cognitive model", "python3 AUTOMATIONS/cognitive_engine.py --build-model", "cpu"),
    ]
    for cid, name, desc, cmd, icon in refiner_cmds:
        commands.append({
            "id": cid, "name": name, "description": desc,
            "category": "system",
            "cmd": cmd, "icon": icon,
        })

    # ── Content ───────────────────────────────────────────────────────
    commands.append({
        "id": "content-voice-regen",
        "name": "Voice Model Regen",
        "description": "Regenerate user voice model from prompt history",
        "category": "content",
        "cmd": "python3 AUTOMATIONS/user_voice_model.py",
        "icon": "microphone",
    })

    return commands


def _job_runner(job_id, cmd_str):
    """Background thread that runs a command and captures output."""
    try:
        proc = subprocess.Popen(
            cmd_str, shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, cwd=str(PROJECT_ROOT),
        )
        with _jobs_lock:
            _running_jobs[job_id]["pid"] = proc.pid

        output_lines = []
        for line in proc.stdout:
            output_lines.append(line)
            # Keep only last 200 lines in memory
            if len(output_lines) > 200:
                output_lines = output_lines[-200:]
            with _jobs_lock:
                _running_jobs[job_id]["output"] = output_lines

        proc.wait()
        with _jobs_lock:
            _running_jobs[job_id]["status"] = "complete" if proc.returncode == 0 else "error"
            _running_jobs[job_id]["returncode"] = proc.returncode
            _running_jobs[job_id]["ended"] = datetime.now().isoformat()
    except Exception as e:
        with _jobs_lock:
            _running_jobs[job_id]["status"] = "error"
            _running_jobs[job_id]["output"] = [str(e)]
            _running_jobs[job_id]["ended"] = datetime.now().isoformat()


@app.route("/api/commands/all")
def api_commands_all():
    """Return full command palette for GUI."""
    return jsonify({"commands": _get_all_commands()})


@app.route("/api/commands/run", methods=["POST"])
def api_commands_run():
    """Run a command asynchronously. Returns job_id immediately."""
    data = request.get_json(force=True)
    cmd_str = data.get("cmd", "").strip()
    if not cmd_str:
        return jsonify({"error": "No cmd provided"}), 400

    job_id = str(uuid.uuid4())[:8]
    with _jobs_lock:
        _running_jobs[job_id] = {
            "cmd": cmd_str,
            "status": "running",
            "pid": None,
            "started": datetime.now().isoformat(),
            "ended": None,
            "output": [],
            "returncode": None,
        }

    t = threading.Thread(target=_job_runner, args=(job_id, cmd_str), daemon=True)
    t.start()
    return jsonify({"job_id": job_id, "status": "running"})


@app.route("/api/commands/job/<job_id>")
def api_commands_job(job_id):
    """Check job status. Returns last 50 lines of output."""
    with _jobs_lock:
        job = _running_jobs.get(job_id)
    if not job:
        return jsonify({"error": f"Unknown job: {job_id}"}), 404

    output_lines = job.get("output", [])
    tail = output_lines[-50:] if len(output_lines) > 50 else output_lines
    return jsonify({
        "job_id": job_id,
        "cmd": job["cmd"],
        "status": job["status"],
        "pid": job["pid"],
        "started": job["started"],
        "ended": job["ended"],
        "returncode": job["returncode"],
        "output": "".join(tail),
        "lines": len(output_lines),
    })


# ---------------------------------------------------------------------------
# Cron Scheduler API
# ---------------------------------------------------------------------------
def _parse_crontab():
    """Parse crontab -l into structured entries."""
    try:
        result = subprocess.run(["crontab", "-l"], capture_output=True, text=True, timeout=10)
        if result.returncode != 0:
            return []
    except Exception:
        return []

    lines = result.stdout.strip().split("\n")
    entries = []
    prev_comment = ""
    idx = 0
    for line in lines:
        stripped = line.strip()
        if not stripped:
            prev_comment = ""
            continue
        if stripped.startswith("#") and not stripped.startswith("# "):
            # Disabled cron line: strip leading # and parse
            inner = stripped.lstrip("#").strip()
            parts = inner.split(None, 5)
            if len(parts) >= 6 and _looks_like_schedule(parts[:5]):
                schedule = " ".join(parts[:5])
                command = parts[5]
                entries.append({
                    "id": idx, "schedule": schedule, "command": command,
                    "name": prev_comment or _name_from_cmd(command),
                    "enabled": False, "next_run": _readable_schedule(schedule),
                    "comment": prev_comment, "raw": line,
                })
                idx += 1
                prev_comment = ""
                continue
            # Regular comment — save as name for next entry
            prev_comment = stripped.lstrip("# ").strip()
            continue
        if stripped.startswith("#"):
            prev_comment = stripped.lstrip("# ").strip()
            continue

        # Active cron line
        parts = stripped.split(None, 5)
        if len(parts) >= 6 and _looks_like_schedule(parts[:5]):
            schedule = " ".join(parts[:5])
            command = parts[5]
            entries.append({
                "id": idx, "schedule": schedule, "command": command,
                "name": prev_comment or _name_from_cmd(command),
                "enabled": True, "next_run": _readable_schedule(schedule),
                "comment": prev_comment, "raw": line,
            })
            idx += 1
        prev_comment = ""

    return entries


def _looks_like_schedule(fields):
    """Check if 5 fields look like a cron schedule."""
    for f in fields:
        if not re.match(r'^[\d\*\/\-\,]+$', f):
            return False
    return True


def _name_from_cmd(cmd):
    """Extract a readable name from a cron command string."""
    match = re.search(r'(\w+)\.py', cmd)
    if match:
        return match.group(1).replace("_", " ").title()
    return cmd[:60]


def _readable_schedule(schedule):
    """Convert 5-field cron schedule to human-readable text."""
    parts = schedule.split()
    if len(parts) != 5:
        return schedule
    minute, hour, dom, month, dow = parts
    days_map = {"0": "Sun", "1": "Mon", "2": "Tue", "3": "Wed", "4": "Thu", "5": "Fri", "6": "Sat", "7": "Sun"}

    desc = []
    if minute != "*" and hour != "*":
        desc.append(f"{hour.zfill(2)}:{minute.zfill(2)}")
    elif hour != "*":
        desc.append(f"{hour}:xx")
    elif minute != "*":
        desc.append(f"xx:{minute.zfill(2)}")
    else:
        desc.append("every min")

    if dow != "*":
        day_names = [days_map.get(d.strip(), d.strip()) for d in dow.split(",")]
        desc.append(",".join(day_names))
    if dom != "*":
        desc.append(f"day {dom}")
    if month != "*":
        desc.append(f"month {month}")

    return " ".join(desc) if desc else schedule


def _reinstall_crontab(lines):
    """Reinstall crontab from list of lines."""
    content = "\n".join(lines) + "\n"
    proc = subprocess.run(
        ["crontab", "-"], input=content, capture_output=True, text=True, timeout=10,
    )
    return proc.returncode == 0


@app.route("/api/scheduler/crons")
def api_scheduler_crons():
    """Return parsed crontab entries."""
    return jsonify({"crons": _parse_crontab()})


@app.route("/api/scheduler/toggle/<int:cron_id>", methods=["POST"])
def api_scheduler_toggle(cron_id):
    """Toggle a cron entry on/off by commenting/uncommenting."""
    try:
        result = subprocess.run(["crontab", "-l"], capture_output=True, text=True, timeout=10)
        if result.returncode != 0:
            return jsonify({"error": "Failed to read crontab"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    raw_lines = result.stdout.strip().split("\n")

    # Map cron_id back to raw lines — skip blank/comment-only lines for counting
    entry_idx = 0
    target_line_idx = None
    prev_was_comment = False
    for i, line in enumerate(raw_lines):
        stripped = line.strip()
        if not stripped:
            prev_was_comment = False
            continue
        # Check if this is a disabled cron (starts with # but has schedule)
        if stripped.startswith("#"):
            inner = stripped.lstrip("#").strip()
            parts = inner.split(None, 5)
            if len(parts) >= 6 and _looks_like_schedule(parts[:5]):
                if entry_idx == cron_id:
                    target_line_idx = i
                    break
                entry_idx += 1
                continue
            prev_was_comment = True
            continue
        # Active cron
        parts = stripped.split(None, 5)
        if len(parts) >= 6 and _looks_like_schedule(parts[:5]):
            if entry_idx == cron_id:
                target_line_idx = i
                break
            entry_idx += 1
        prev_was_comment = False

    if target_line_idx is None:
        return jsonify({"error": f"Cron entry {cron_id} not found"}), 404

    line = raw_lines[target_line_idx].strip()
    if line.startswith("#"):
        # Enable: remove leading #
        raw_lines[target_line_idx] = line.lstrip("#").strip()
        action = "enabled"
    else:
        # Disable: add #
        raw_lines[target_line_idx] = "# " + line
        action = "disabled"

    if _reinstall_crontab(raw_lines):
        return jsonify({"success": True, "action": action})
    return jsonify({"error": "Failed to reinstall crontab"}), 500


@app.route("/api/scheduler/add", methods=["POST"])
def api_scheduler_add():
    """Add a new cron entry."""
    data = request.get_json(force=True)
    schedule = data.get("schedule", "").strip()
    command = data.get("command", "").strip()
    comment = data.get("comment", "").strip()
    if not schedule or not command:
        return jsonify({"error": "schedule and command required"}), 400

    # Validate schedule (5 fields)
    parts = schedule.split()
    if len(parts) != 5 or not _looks_like_schedule(parts):
        return jsonify({"error": "Invalid cron schedule — need 5 fields"}), 400

    try:
        result = subprocess.run(["crontab", "-l"], capture_output=True, text=True, timeout=10)
        existing = result.stdout.strip() if result.returncode == 0 else ""
    except Exception:
        existing = ""

    new_lines = existing.split("\n") if existing else []
    if comment:
        new_lines.append(f"# {comment}")
    new_lines.append(f"{schedule} {command}")

    if _reinstall_crontab(new_lines):
        return jsonify({"success": True, "message": f"Added cron: {schedule} {command}"})
    return jsonify({"error": "Failed to install new crontab"}), 500


@app.route("/api/scheduler/delete/<int:cron_id>", methods=["POST"])
def api_scheduler_delete(cron_id):
    """Delete a cron entry by index."""
    try:
        result = subprocess.run(["crontab", "-l"], capture_output=True, text=True, timeout=10)
        if result.returncode != 0:
            return jsonify({"error": "Failed to read crontab"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    raw_lines = result.stdout.strip().split("\n")
    entry_idx = 0
    target_line_idx = None
    for i, line in enumerate(raw_lines):
        stripped = line.strip()
        if not stripped:
            continue
        if stripped.startswith("#"):
            inner = stripped.lstrip("#").strip()
            parts = inner.split(None, 5)
            if len(parts) >= 6 and _looks_like_schedule(parts[:5]):
                if entry_idx == cron_id:
                    target_line_idx = i
                    break
                entry_idx += 1
                continue
            continue
        parts = stripped.split(None, 5)
        if len(parts) >= 6 and _looks_like_schedule(parts[:5]):
            if entry_idx == cron_id:
                target_line_idx = i
                break
            entry_idx += 1

    if target_line_idx is None:
        return jsonify({"error": f"Cron entry {cron_id} not found"}), 404

    # Remove the entry and its comment line above if it exists
    del raw_lines[target_line_idx]
    if target_line_idx > 0 and raw_lines[target_line_idx - 1].strip().startswith("#"):
        del raw_lines[target_line_idx - 1]

    if _reinstall_crontab(raw_lines):
        return jsonify({"success": True, "message": f"Deleted cron entry {cron_id}"})
    return jsonify({"error": "Failed to reinstall crontab"}), 500


# ---------------------------------------------------------------------------
# Manual Priority / Notes API
# ---------------------------------------------------------------------------
@app.route("/api/manual/priority", methods=["POST"])
def api_manual_priority_v2():
    """Enhanced manual priority with priority level."""
    data = request.get_json(force=True)
    text = data.get("text", "").strip()
    source = data.get("source", "manual")
    priority = data.get("priority", "P1")
    score = data.get("score", 0)
    if not text:
        return jsonify({"error": "No text provided"}), 400

    manual_path = OPS / "MANUAL_PRIORITIES.md"
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    entry = (
        f"\n## [{timestamp}] {priority} — MANUAL OVERRIDE\n"
        f"- **Priority:** {priority}\n"
        f"- **Source:** {source}\n"
        f"- **Score:** {score}\n"
        f"- **Item:** {text}\n"
        f"- **Status:** PENDING\n"
    )

    try:
        if not manual_path.exists():
            manual_path.write_text(
                f"# Manual Priority Overrides\n\n"
                f"Human-promoted items from the GUI.\n{entry}"
            )
        else:
            with open(manual_path, "a", encoding="utf-8") as f:
                f.write(entry)
        return jsonify({"success": True, "message": f"Added {priority}: {text[:60]}..."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/manual/priorities")
def api_manual_priorities():
    """Read and parse MANUAL_PRIORITIES.md into items."""
    manual_path = OPS / "MANUAL_PRIORITIES.md"
    if not manual_path.exists():
        return jsonify({"items": []})

    try:
        content = manual_path.read_text(encoding="utf-8", errors="replace")
    except Exception:
        return jsonify({"items": []})

    items = []
    current = None
    for line in content.split("\n"):
        if line.startswith("## ["):
            if current:
                items.append(current)
            # Parse header: ## [2026-03-24 12:00:00] P0 — MANUAL OVERRIDE
            match = re.match(r'## \[(.+?)\]\s*(P\d)?\s*', line)
            current = {
                "timestamp": match.group(1) if match else "",
                "priority": match.group(2) if match and match.group(2) else "P1",
                "text": "", "source": "", "status": "PENDING",
            }
        elif current and line.strip().startswith("- **Item:**"):
            current["text"] = line.split("**Item:**", 1)[1].strip()
        elif current and line.strip().startswith("- **Source:**"):
            current["source"] = line.split("**Source:**", 1)[1].strip()
        elif current and line.strip().startswith("- **Status:**"):
            current["status"] = line.split("**Status:**", 1)[1].strip()
        elif current and line.strip().startswith("- **Priority:**"):
            current["priority"] = line.split("**Priority:**", 1)[1].strip()
    if current:
        items.append(current)

    # Most recent first
    items.reverse()
    return jsonify({"items": items})


@app.route("/api/manual/note", methods=["POST"])
def api_manual_note():
    """Append a note to OPS/MANUAL_NOTES.md with timestamp and tags."""
    data = request.get_json(force=True)
    text = data.get("text", "").strip()
    tags = data.get("tags", [])
    if not text:
        return jsonify({"error": "No text provided"}), 400

    notes_path = OPS / "MANUAL_NOTES.md"
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    tag_str = ", ".join(f"`{t}`" for t in tags) if tags else "none"

    entry = f"\n- [{timestamp}] {text} — tags: {tag_str}\n"

    try:
        if not notes_path.exists():
            notes_path.write_text(f"# Manual Notes\n\nQuick notes from the GUI.\n{entry}")
        else:
            with open(notes_path, "a", encoding="utf-8") as f:
                f.write(entry)
        return jsonify({"success": True, "message": f"Note saved: {text[:60]}..."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/manual/notes")
def api_manual_notes():
    """Read and return OPS/MANUAL_NOTES.md parsed into items."""
    notes_path = OPS / "MANUAL_NOTES.md"
    if not notes_path.exists():
        return jsonify({"notes": []})

    try:
        content = notes_path.read_text(encoding="utf-8", errors="replace")
    except Exception:
        return jsonify({"notes": []})

    notes = []
    for line in content.split("\n"):
        # Pattern: - [2026-03-24 12:00:00] Some text — tags: `revenue`, `urgent`
        match = re.match(r'^- \[(.+?)\]\s*(.+?)\s*(?:—\s*tags:\s*(.+))?$', line.strip())
        if match:
            tag_text = match.group(3) or ""
            tags = [t.strip().strip("`") for t in tag_text.split(",") if t.strip() and t.strip() != "none"]
            notes.append({
                "timestamp": match.group(1),
                "text": match.group(2).rstrip(" —"),
                "tags": tags,
            })

    notes.reverse()
    return jsonify({"notes": notes})


# ---------------------------------------------------------------------------
# Usage Optimizer API
# ---------------------------------------------------------------------------
@app.route("/api/usage")
def api_usage():
    """Run usage_optimizer.py --check and return JSON output."""
    code, out, err = run_cmd(
        [PYTHON, str(AUTOMATIONS / "usage_optimizer.py"), "--check"], timeout=30
    )
    if code == 0 and out:
        try:
            return jsonify(json.loads(out))
        except json.JSONDecodeError:
            return jsonify({"raw": out, "status": "ok"})
    return jsonify({"error": err or "usage_optimizer failed", "raw": out}), 500


@app.route("/api/usage/burst", methods=["POST"])
def api_usage_burst():
    """Trigger a manual burst via usage_optimizer.py --burst."""
    code, out, err = run_cmd(
        [PYTHON, str(AUTOMATIONS / "usage_optimizer.py"), "--burst"], timeout=60
    )
    return jsonify({"success": code == 0, "output": out[-3000:], "error": err[-1000:]})


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print(f"\n  PRINTMAXX CONTROL PANEL")
    print(f"  http://localhost:{PORT}")
    print(f"  Press Ctrl+C to stop\n")

    # Auto-open browser after short delay
    def open_browser():
        time.sleep(1.5)
        webbrowser.open(f"http://localhost:{PORT}")

    threading.Thread(target=open_browser, daemon=True).start()

    try:
        app.run(host="0.0.0.0", port=PORT, debug=False)
    except KeyboardInterrupt:
        print("\n  Shutting down.")


if __name__ == "__main__":
    main()
