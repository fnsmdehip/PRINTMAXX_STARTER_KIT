#!/usr/bin/env python3
"""
PRINTMAXX CEO AGENT — 24/7 Autonomous Multi-Venture Orchestrator
=================================================================

The CEO brain that sits ABOVE the ops manager. Reads the master xlsx,
scores all 182 ops dynamically, makes strategic decisions, delegates
to venture sub-agents, and protects good ops with git-based failsafes.

Architecture:
  CEO AGENT (this file) — FULL ORCHESTRATOR of ALL systems
  ├── GitGuard           — auto-snapshot before changes, rollback on failure
  ├── XlsxIntel          — reads 182 ops from master xlsx dynamically
  ├── VentureScorer      — multi-signal scoring (readiness + automation + synergy + revenue)
  ├── CEOBrain           — strategic decisions (PROMOTE / ENHANCE / CREATE / KILL / DISCOVER)
  ├── VentureRunner      — delegates to existing + dynamic venture agents
  ├── AuditTrail         — regression detection, protected ops enforcement
  ├── Alpha Pipeline     — twitter/reddit scrapers + alpha processor
  ├── Research Pipeline  — daily_research_orchestrator (once/day)
  ├── Decision Engine    — closed-loop decision_engine.py
  ├── Content Generation — printmaxx_agent content + upgrade missions
  ├── System Health      — system_health_monitor + auto-fix
  └── Cron Management    — read/write/track scheduled tasks

Failsafes:
  - Git commit before every decision batch
  - Protected ops list (never auto-kill above performance threshold)
  - Max changes per cycle (prevent runaway)
  - Post-cycle audit: rollback if regression detected
  - Lock file prevents double-runs
  - All file ops locked to PROJECT root (inherited guardrails)

Usage:
  python3 AUTOMATIONS/ceo_agent.py                  # One CEO cycle (full orchestration)
  python3 AUTOMATIONS/ceo_agent.py --status          # Full status dashboard
  python3 AUTOMATIONS/ceo_agent.py --daemon           # Run forever (24/7)
  python3 AUTOMATIONS/ceo_agent.py --score            # Score all ops from xlsx
  python3 AUTOMATIONS/ceo_agent.py --decide           # Show decisions without executing
  python3 AUTOMATIONS/ceo_agent.py --protect OP_ID    # Add op to protected list
  python3 AUTOMATIONS/ceo_agent.py --rollback         # Rollback last CEO change batch
  python3 AUTOMATIONS/ceo_agent.py --alpha            # Run alpha pipeline only
  python3 AUTOMATIONS/ceo_agent.py --research         # Run research pipeline only
  python3 AUTOMATIONS/ceo_agent.py --content          # Run content generation only
  python3 AUTOMATIONS/ceo_agent.py --health           # Run system health check only
  python3 AUTOMATIONS/ceo_agent.py --decision-engine  # Run decision engine only
  python3 AUTOMATIONS/ceo_agent.py --ventures         # Run dynamic ventures only
  python3 AUTOMATIONS/ceo_agent.py --cron-list        # Show managed cron entries
  python3 AUTOMATIONS/ceo_agent.py --cron-add "..."   # Add a cron entry
"""

import argparse
import csv
import json
import os
import subprocess
import sys
import time
import shutil
import fcntl
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

# === GUARDRAILS (inherited from ops_manager) ===
PROJECT = Path("/Users/macbookpro/Documents/p/PRINTMAXX_STARTER_KITttttt")
CEO_DIR = PROJECT / "AUTOMATIONS" / "agent" / "ceo_agent"
STATE_FILE = CEO_DIR / "ceo_state.json"
DECISION_LOG = CEO_DIR / "decisions.jsonl"
AUDIT_LOG = CEO_DIR / "audit.jsonl"
LOCK_FILE = CEO_DIR / "ceo.lock"
PYTHON = sys.executable

BLOCKED_DIRS = [
    Path.home() / d for d in
    ["Desktop", "Downloads", "Pictures", "Music", "Movies", "Library", ".ssh", ".aws", ".gnupg"]
] + [Path(d) for d in ["/System", "/Library", "/usr", "/bin", "/etc", "/var"]]

BLOCKED_COMMANDS = ["rm -rf /", "rm -rf ~", "dd if=", "diskutil erase", "mkfs", ":(){ :|:& };:"]

# CEO config — tune these thresholds
MAX_CHANGES_PER_CYCLE = 5       # prevent runaway mutations
KILL_THRESHOLD = 15             # score below this = eligible for kill
PROTECT_THRESHOLD = 60          # score above this = protected from kill
PROMOTE_THRESHOLD = 70          # score above this = double down
CYCLE_INTERVAL_HOURS = 1        # how often the CEO runs
DISCOVER_INTERVAL_HOURS = 4     # how often to hunt new ventures
ALPHA_INTERVAL_HOURS = 2        # how often to run alpha scrapers
RESEARCH_INTERVAL_HOURS = 24    # daily research (once per day)
CONTENT_INTERVAL_HOURS = 6      # content generation interval
HEALTH_INTERVAL_HOURS = 1       # system health check interval
OPENCLAW_INTERVAL_HOURS = 4     # openclaw local biz pipeline interval
AUTONOMY_INTERVAL_HOURS = 2    # venture autonomy engine interval
GIT_SNAPSHOT_BEFORE_CHANGES = True

# OpenClaw city rotation — cycles through these automatically
OPENCLAW_CITIES = [
    ("Austin TX", "dentist"), ("Austin TX", "plumber"), ("Austin TX", "lawyer"),
    ("Houston TX", "dentist"), ("Houston TX", "plumber"), ("Houston TX", "lawyer"),
    ("Dallas TX", "dentist"), ("Dallas TX", "plumber"), ("Dallas TX", "lawyer"),
    ("Miami FL", "dentist"), ("Miami FL", "plumber"), ("Miami FL", "lawyer"),
    ("Atlanta GA", "dentist"), ("Atlanta GA", "plumber"), ("Atlanta GA", "lawyer"),
    ("Chicago IL", "dentist"), ("Chicago IL", "plumber"), ("Chicago IL", "lawyer"),
    ("Phoenix AZ", "dentist"), ("Phoenix AZ", "plumber"), ("Phoenix AZ", "lawyer"),
    ("Denver CO", "dentist"), ("Denver CO", "plumber"), ("Denver CO", "lawyer"),
    ("Seattle WA", "dentist"), ("Seattle WA", "plumber"), ("Seattle WA", "lawyer"),
    ("New York NY", "dentist"), ("New York NY", "plumber"), ("New York NY", "lawyer"),
]

# Master xlsx path (find the latest one)
XLSX_PATTERNS = [
    "PRINTMAXX_MASTER_OPS_ENHANCED_*.xlsx",
    "PRINTMAXX_MASTER_OPS*.xlsx",
]


def safe_path(p):
    resolved = Path(p).resolve()
    project_resolved = PROJECT.resolve()
    if not str(resolved).startswith(str(project_resolved)):
        raise ValueError(f"GUARDRAIL BLOCKED: {resolved} outside project {project_resolved}")
    for blocked in BLOCKED_DIRS:
        if str(resolved).startswith(str(blocked.resolve())):
            raise ValueError(f"GUARDRAIL BLOCKED: {resolved} in protected dir {blocked}")
    return resolved


def safe_command(cmd_str):
    lower = cmd_str.lower()
    for bad in BLOCKED_COMMANDS:
        if bad in lower:
            raise ValueError(f"GUARDRAIL BLOCKED: {bad}")


def disk_free_gb():
    try:
        st = os.statvfs("/")
        return round(st.f_bavail * st.f_frsize / (1024**3), 1)
    except Exception:
        return -1


def ts():
    return datetime.now().strftime("%H:%M:%S")


def log(msg, level="INFO"):
    line = f"[{ts()}] [CEO] [{level}] {msg}"
    print(line)
    log_path = PROJECT / "AUTOMATIONS" / "logs" / "ceo_agent.log"
    try:
        log_path.parent.mkdir(parents=True, exist_ok=True)
        with open(log_path, "a") as f:
            f.write(line + "\n")
    except Exception:
        pass


def run_script(script_path, args="", timeout_sec=300, label=None):
    """Run a PRINTMAXX script with guardrails. Returns (success, output)."""
    full_path = PROJECT / "AUTOMATIONS" / script_path if "/" not in script_path else PROJECT / script_path
    if not full_path.exists():
        msg = f"Script not found: {full_path}"
        log(msg, "WARN")
        return False, msg

    cmd = f"{PYTHON} {full_path} {args}".strip()
    safe_command(cmd)
    tag = label or script_path
    log(f"Running: {tag} {args}")

    try:
        result = subprocess.run(
            cmd, shell=True, capture_output=True, text=True,
            timeout=timeout_sec, cwd=str(PROJECT)
        )
        output = (result.stdout or "") + (result.stderr or "")
        output = output[-2000:]  # keep last 2000 chars
        if result.returncode == 0:
            log(f"  OK: {tag}")
            return True, output
        else:
            log(f"  FAIL (rc={result.returncode}): {tag}", "WARN")
            return False, output
    except subprocess.TimeoutExpired:
        log(f"  TIMEOUT ({timeout_sec}s): {tag}", "WARN")
        return False, f"Timed out after {timeout_sec}s"
    except Exception as e:
        log(f"  ERROR: {tag}: {e}", "ERROR")
        return False, str(e)[:500]


def _hours_since(iso_ts):
    """Return hours elapsed since an ISO timestamp string. Returns float('inf') if None/invalid."""
    if not iso_ts:
        return float('inf')
    try:
        return (datetime.now() - datetime.fromisoformat(str(iso_ts))).total_seconds() / 3600
    except Exception:
        return float('inf')


# ============================================================================
# GIT GUARD — snapshot before changes, rollback on failure
# ============================================================================

class GitGuard:
    """Git-based failsafe. Snapshots state before CEO makes changes."""

    def __init__(self):
        self.last_snapshot_hash = None
        self.snapshot_tag = None

    def snapshot(self, label="ceo_pre_change"):
        """Commit current state before CEO makes changes."""
        if not GIT_SNAPSHOT_BEFORE_CHANGES:
            return True
        try:
            tag = f"{label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

            # Stage all changes
            subprocess.run(
                ["git", "add", "-A"],
                cwd=str(PROJECT), capture_output=True, timeout=30
            )

            # Commit
            result = subprocess.run(
                ["git", "commit", "-m", f"CEO snapshot: {tag}",
                 "--allow-empty"],
                cwd=str(PROJECT), capture_output=True, text=True, timeout=30
            )

            # Get the hash
            hash_result = subprocess.run(
                ["git", "rev-parse", "HEAD"],
                cwd=str(PROJECT), capture_output=True, text=True, timeout=10
            )
            self.last_snapshot_hash = hash_result.stdout.strip()
            self.snapshot_tag = tag

            log(f"Git snapshot: {self.last_snapshot_hash[:8]} ({tag})")
            return True
        except Exception as e:
            log(f"Git snapshot failed: {e}", "WARN")
            return False

    def rollback(self):
        """Rollback to last snapshot. Only used if post-audit detects regression."""
        if not self.last_snapshot_hash:
            log("No snapshot hash to rollback to", "WARN")
            return False
        try:
            # Soft reset — keeps changes staged so we can inspect
            result = subprocess.run(
                ["git", "reset", "--soft", self.last_snapshot_hash],
                cwd=str(PROJECT), capture_output=True, text=True, timeout=30
            )
            if result.returncode == 0:
                log(f"ROLLBACK to {self.last_snapshot_hash[:8]}", "WARN")
                return True
            log(f"Rollback failed: {result.stderr}", "ERROR")
            return False
        except Exception as e:
            log(f"Rollback exception: {e}", "ERROR")
            return False

    def post_change_commit(self, summary):
        """Commit after CEO changes, with a descriptive message."""
        try:
            subprocess.run(
                ["git", "add", "-A"],
                cwd=str(PROJECT), capture_output=True, timeout=30
            )
            subprocess.run(
                ["git", "commit", "-m", f"CEO changes: {summary}",
                 "--allow-empty"],
                cwd=str(PROJECT), capture_output=True, text=True, timeout=30
            )
            log(f"Post-change commit: {summary}")
            return True
        except Exception as e:
            log(f"Post-change commit failed: {e}", "WARN")
            return False


# ============================================================================
# XLSX INTEL — reads master xlsx for dynamic venture data
# ============================================================================

class XlsxIntel:
    """Reads the PRINTMAXX master xlsx to get all 182 ops and their metadata."""

    def __init__(self):
        self.xlsx_path = self._find_xlsx()
        self._cache = {}
        self._cache_time = None

    def _find_xlsx(self):
        """Find the latest master xlsx."""
        for pattern in XLSX_PATTERNS:
            matches = sorted(PROJECT.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
            if matches:
                return matches[0]
        return None

    def _load(self):
        """Load xlsx data with caching (reload every 5 min)."""
        if self._cache and self._cache_time and (datetime.now() - self._cache_time).seconds < 300:
            return self._cache

        if not self.xlsx_path or not self.xlsx_path.exists():
            log("No master xlsx found", "WARN")
            return {}

        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(self.xlsx_path), read_only=True, data_only=True)
            data = {}
            for name in wb.sheetnames:
                ws = wb[name]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    continue
                headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
                sheet_data = []
                for row in rows[1:]:
                    entry = {}
                    for i, val in enumerate(row):
                        if i < len(headers):
                            entry[headers[i]] = val
                    if any(v for v in entry.values()):
                        sheet_data.append(entry)
                data[name] = sheet_data
            wb.close()
            self._cache = data
            self._cache_time = datetime.now()
            log(f"Loaded xlsx: {len(data)} sheets, {sum(len(v) for v in data.values())} rows")
            return data
        except Exception as e:
            log(f"xlsx load error: {e}", "ERROR")
            return {}

    def get_all_ops(self):
        """Get all ops from ALL OPS MASTER sheet."""
        data = self._load()
        return data.get("ALL OPS MASTER", [])

    def get_auto_status(self):
        """Get automation status for all ops."""
        data = self._load()
        return data.get("AUTO_STATUS_LIVE", [])

    def get_priority_launch(self):
        """Get priority-ranked ops for launch."""
        data = self._load()
        return data.get("PRIORITY LAUNCH", [])

    def get_synergy_stacks(self):
        """Get synergy combos with revenue multipliers."""
        data = self._load()
        return data.get("SYNERGY STACKS", [])

    def get_venture_map(self):
        """Get venture-to-automation mapping."""
        data = self._load()
        return data.get("VENTURE_AUTOMATION_MAP", [])

    def get_expansion_queue(self):
        """Get expansion queue for scaling ops."""
        data = self._load()
        return data.get("ETC_EXPANSION_QUEUE", [])

    def get_op_by_id(self, op_id):
        """Find a specific op by its ID."""
        for op in self.get_all_ops():
            if op.get("OP_ID") == op_id:
                return op
        return None

    def get_ops_by_category(self, category):
        """Get all ops in a category."""
        return [op for op in self.get_all_ops() if op.get("CATEGORY", "").upper() == category.upper()]

    def get_ready_ops(self):
        """Get ops marked as READY in auto status."""
        return [op for op in self.get_auto_status()
                if str(op.get("READINESS", "")).upper() == "READY"]

    def get_blocked_ops(self):
        """Get ops with blockers."""
        return [op for op in self.get_auto_status()
                if op.get("BLOCKER_KEY") and str(op.get("BLOCKER_KEY")).strip()]


# ============================================================================
# CEO STATE — persistent state for CEO decisions
# ============================================================================

class CEOState:
    """Persistent state for the CEO agent."""

    def __init__(self):
        CEO_DIR.mkdir(parents=True, exist_ok=True)
        self.data = self._load()

    def _load(self):
        if STATE_FILE.exists():
            try:
                return json.loads(STATE_FILE.read_text())
            except Exception:
                return self._default()
        return self._default()

    def _default(self):
        return {
            "last_cycle": None,
            "cycles_run": 0,
            "total_decisions": 0,
            "op_scores": {},
            "op_history": {},        # op_id -> list of score snapshots
            "protected_ops": [],     # ops that can never be auto-killed
            "killed_ops": [],        # ops the CEO decided to sunset
            "promoted_ops": [],      # ops getting extra resources
            "created_ventures": [],  # dynamically created ventures
            "enhanced_ops": [],      # ops that got enhanced
            "last_discover": None,
            "last_git_snapshot": None,
            "last_alpha_scrape": None,    # last alpha pipeline run
            "last_research": None,        # last daily research run
            "last_content_gen": None,     # last content generation run
            "last_health_check": None,    # last system health check
            "last_decision_engine": None, # last decision engine cycle
            "managed_crons": [],          # cron entries managed by CEO
            "alpha_pipeline_results": {}, # latest alpha pipeline stats
            "health_issues": [],          # outstanding health issues
            "last_openclaw": None,        # last openclaw pipeline run
            "openclaw_city_index": 0,     # current city rotation index
            "openclaw_stats": {},         # openclaw pipeline stats
            "performance_baseline": {},   # pre-change metrics for regression check
            "config": {
                "max_changes": MAX_CHANGES_PER_CYCLE,
                "kill_threshold": KILL_THRESHOLD,
                "protect_threshold": PROTECT_THRESHOLD,
                "promote_threshold": PROMOTE_THRESHOLD,
            }
        }

    def save(self):
        safe_path(STATE_FILE)
        STATE_FILE.write_text(json.dumps(self.data, indent=2, default=str))

    def log_decision(self, decision):
        """Append a decision to the JSONL log."""
        decision["ts"] = datetime.now().isoformat()
        safe_path(DECISION_LOG)
        with open(DECISION_LOG, "a") as f:
            f.write(json.dumps(decision, default=str) + "\n")
        self.data["total_decisions"] += 1

    def log_audit(self, audit_entry):
        """Append an audit entry."""
        audit_entry["ts"] = datetime.now().isoformat()
        safe_path(AUDIT_LOG)
        with open(AUDIT_LOG, "a") as f:
            f.write(json.dumps(audit_entry, default=str) + "\n")

    def is_protected(self, op_id):
        """Check if an op is protected from auto-kill."""
        return op_id in self.data.get("protected_ops", [])

    def protect_op(self, op_id):
        """Add op to protected list."""
        if op_id not in self.data["protected_ops"]:
            self.data["protected_ops"].append(op_id)
            self.save()
            log(f"Protected op: {op_id}")

    def get_score_trend(self, op_id, periods=5):
        """Get recent score trend for an op."""
        history = self.data.get("op_history", {}).get(op_id, [])
        return history[-periods:]


# ============================================================================
# VENTURE SCORER — multi-signal scoring
# ============================================================================

class VentureScorer:
    """Scores all ops using xlsx data + runtime metrics."""

    def __init__(self, xlsx: XlsxIntel, state: CEOState):
        self.xlsx = xlsx
        self.state = state

    def score_all(self):
        """Score every op in the xlsx. Returns sorted list of scored ops."""
        all_ops = self.xlsx.get_all_ops()
        auto_status = {op.get("OP_ID"): op for op in self.xlsx.get_auto_status()}
        venture_map = {op.get("VENTURE_ID"): op for op in self.xlsx.get_venture_map()}
        synergies = self.xlsx.get_synergy_stacks()

        scored = []
        for op in all_ops:
            op_id = op.get("OP_ID", "")
            if not op_id:
                continue

            score = self._score_op(op, auto_status.get(op_id, {}),
                                   venture_map.get(op_id, {}), synergies)
            scored.append(score)

        # Sort by total score descending
        scored.sort(key=lambda x: x["total_score"], reverse=True)

        # Save to state
        self.state.data["op_scores"] = {s["op_id"]: s for s in scored}

        # Update history (keep last 20 snapshots per op)
        history = self.state.data.setdefault("op_history", {})
        for s in scored:
            op_hist = history.setdefault(s["op_id"], [])
            op_hist.append({
                "ts": datetime.now().isoformat(),
                "score": s["total_score"]
            })
            if len(op_hist) > 20:
                history[s["op_id"]] = op_hist[-20:]

        self.state.save()
        return scored

    def _score_op(self, op, auto_status, venture_info, synergies):
        """Score a single op across multiple dimensions."""
        op_id = op.get("OP_ID", "")

        # Dimension 1: Readiness (0-25)
        readiness_str = str(auto_status.get("READINESS", "")).upper()
        readiness_map = {"READY": 25, "BUILD": 15, "PLAN": 5}
        readiness_score = readiness_map.get(readiness_str, 0)

        # Dimension 2: Automation level (0-25)
        auto_val = auto_status.get("AUTOMATION_SCORE_100")
        try:
            auto_score = min(25, int(float(auto_val or 0)) / 4)
        except (ValueError, TypeError):
            auto_score = 0

        # Dimension 3: Signal count / market validation (0-25)
        signal_val = auto_status.get("SIGNAL_COUNT")
        try:
            signals = int(float(signal_val or 0))
            signal_score = min(25, signals / 40)  # 1000 signals = 25
        except (ValueError, TypeError):
            signal_score = 0

        # Dimension 4: Revenue potential (0-15)
        rev_str = str(op.get("REVENUE_RANGE", "$0"))
        rev_score = self._parse_revenue_score(rev_str)

        # Dimension 5: Synergy bonus (0-10)
        synergy_score = 0
        for syn in synergies:
            methods = str(syn.get("METHODS_COMBINED", ""))
            if op_id in methods:
                try:
                    mult = float(syn.get("REVENUE_MULTIPLIER", "1").replace("x", ""))
                    synergy_score = min(10, mult * 2)
                except (ValueError, TypeError):
                    synergy_score = 2
                break

        total = round(readiness_score + auto_score + signal_score + rev_score + synergy_score, 1)

        return {
            "op_id": op_id,
            "name": op.get("OP_NAME", ""),
            "category": op.get("CATEGORY", ""),
            "total_score": total,
            "readiness": readiness_score,
            "automation": round(auto_score, 1),
            "signals": round(signal_score, 1),
            "revenue_potential": rev_score,
            "synergy": round(synergy_score, 1),
            "blocker": auto_status.get("BLOCKER_KEY", ""),
            "lane": auto_status.get("LANE", ""),
            "revenue_range": op.get("REVENUE_RANGE", ""),
            "is_protected": self.state.is_protected(op_id),
        }

    def _parse_revenue_score(self, rev_str):
        """Parse revenue range string into a 0-15 score."""
        if not rev_str or rev_str == "$0":
            return 0
        # Extract the upper bound number
        import re
        numbers = re.findall(r'[\d,]+', rev_str.replace(",", ""))
        if not numbers:
            return 2
        try:
            upper = max(int(n) for n in numbers)
            if "K" in rev_str.upper() or upper >= 1000:
                if upper >= 20000 or upper >= 20:
                    return 15
                if upper >= 10000 or upper >= 10:
                    return 12
                if upper >= 5000 or upper >= 5:
                    return 8
                return 5
            return 3
        except (ValueError, TypeError):
            return 2


# ============================================================================
# CEO BRAIN — strategic decision engine
# ============================================================================

class CEOBrain:
    """Makes strategic decisions about ventures."""

    def __init__(self, scorer: VentureScorer, xlsx: XlsxIntel, state: CEOState):
        self.scorer = scorer
        self.xlsx = xlsx
        self.state = state

    def analyze_and_decide(self, dry_run=False):
        """
        Run full CEO analysis cycle. Returns list of decisions.

        Decision types:
          PROMOTE  — double down on winners (score > promote_threshold)
          ENHANCE  — improve underperformers that have potential
          CREATE   — spin up new ventures from high-readiness xlsx ops
          KILL     — sunset dead ops (score < kill_threshold, not protected)
          DISCOVER — find entirely new opportunities
        """
        scores = self.scorer.score_all()
        decisions = []
        changes_budget = self.state.data["config"]["max_changes"]

        # 1. PROMOTE winners
        for op in scores:
            if len(decisions) >= changes_budget:
                break
            if op["total_score"] >= self.state.data["config"]["promote_threshold"]:
                if op["op_id"] not in self.state.data.get("promoted_ops", []):
                    decisions.append({
                        "type": "PROMOTE",
                        "op_id": op["op_id"],
                        "name": op["name"],
                        "score": op["total_score"],
                        "reason": f"Score {op['total_score']} above promote threshold {self.state.data['config']['promote_threshold']}",
                        "action": "Increase cycle frequency, allocate more resources",
                    })

        # 2. ENHANCE underperformers with potential
        for op in scores:
            if len(decisions) >= changes_budget:
                break
            score = op["total_score"]
            if (self.state.data["config"]["kill_threshold"] < score <
                    self.state.data["config"]["promote_threshold"]):
                # Check trend — is it improving or declining?
                trend = self.state.get_score_trend(op["op_id"])
                if len(trend) >= 2:
                    recent_scores = [t["score"] for t in trend]
                    declining = recent_scores[-1] < recent_scores[0]
                else:
                    declining = False

                if op["blocker"] and declining:
                    decisions.append({
                        "type": "ENHANCE",
                        "op_id": op["op_id"],
                        "name": op["name"],
                        "score": score,
                        "reason": f"Score {score} declining with blocker: {op['blocker']}",
                        "action": f"Address blocker: {op['blocker']}",
                        "blocker": op["blocker"],
                    })

        # 3. CREATE new ventures from ready xlsx ops not yet running
        existing_venture_ids = set()
        # Check ops_manager venture names
        try:
            from printmaxx_ops_manager import VENTURE_CLASSES
            for VC in VENTURE_CLASSES:
                existing_venture_ids.add(VC.name)
        except ImportError:
            pass
        # Also check created ventures
        for cv in self.state.data.get("created_ventures", []):
            existing_venture_ids.add(cv.get("op_id", ""))

        ready_ops = self.xlsx.get_ready_ops()
        for op in ready_ops:
            if len(decisions) >= changes_budget:
                break
            op_id = op.get("OP_ID", "")
            if op_id and op_id not in existing_venture_ids:
                score_data = self.state.data.get("op_scores", {}).get(op_id, {})
                if score_data.get("total_score", 0) >= 30:
                    decisions.append({
                        "type": "CREATE",
                        "op_id": op_id,
                        "name": op.get("OP_NAME", ""),
                        "score": score_data.get("total_score", 0),
                        "reason": f"READY status, score {score_data.get('total_score', 0)}, not yet running",
                        "action": f"Create venture agent for {op.get('OP_NAME', '')}",
                        "lane": op.get("LANE", ""),
                        "category": op.get("CATEGORY", ""),
                    })

        # 4. KILL dead ops (with protection check)
        for op in reversed(scores):  # lowest scores first
            if len(decisions) >= changes_budget:
                break
            if op["total_score"] < self.state.data["config"]["kill_threshold"]:
                if op["is_protected"]:
                    log(f"PROTECTED from kill: {op['op_id']} ({op['name']})")
                    continue
                if op["op_id"] in [d.get("op_id") for d in self.state.data.get("killed_ops", [])]:
                    continue  # already killed
                decisions.append({
                    "type": "KILL",
                    "op_id": op["op_id"],
                    "name": op["name"],
                    "score": op["total_score"],
                    "reason": f"Score {op['total_score']} below kill threshold {self.state.data['config']['kill_threshold']}",
                    "action": "Deprioritize, remove from active rotation",
                })

        # 5. DISCOVER — if we haven't discovered recently
        last_discover = self.state.data.get("last_discover")
        should_discover = True
        if last_discover:
            try:
                elapsed = (datetime.now() - datetime.fromisoformat(last_discover)).total_seconds()
                should_discover = elapsed > DISCOVER_INTERVAL_HOURS * 3600
            except Exception:
                pass

        if should_discover and len(decisions) < changes_budget:
            # Check expansion queue for untapped ops
            expansion = self.xlsx.get_expansion_queue()
            for eq_op in expansion[:3]:
                op_id = eq_op.get("OP_ID", "")
                if op_id and op_id not in existing_venture_ids:
                    score_data = self.state.data.get("op_scores", {}).get(op_id, {})
                    if score_data.get("total_score", 0) >= 25:
                        decisions.append({
                            "type": "DISCOVER",
                            "op_id": op_id,
                            "name": eq_op.get("OP_NAME", ""),
                            "score": score_data.get("total_score", 0),
                            "reason": f"In expansion queue with score {score_data.get('total_score', 0)}",
                            "action": "Research and evaluate for venture creation",
                            "expansion_logic": eq_op.get("EXPANSION_LOGIC", ""),
                        })
                        break  # one discover per cycle

        if dry_run:
            log(f"DRY RUN: {len(decisions)} decisions generated (not executed)")
        else:
            log(f"CEO DECISIONS: {len(decisions)} actions planned")

        return decisions


# ============================================================================
# VENTURE RUNNER — executes CEO decisions
# ============================================================================

class VentureRunner:
    """Executes CEO decisions by delegating to venture agents."""

    def __init__(self, state: CEOState, xlsx: XlsxIntel):
        self.state = state
        self.xlsx = xlsx

    def execute_decisions(self, decisions):
        """Execute a list of CEO decisions. Returns execution results."""
        results = []
        for decision in decisions:
            dtype = decision["type"]
            try:
                if dtype == "PROMOTE":
                    result = self._execute_promote(decision)
                elif dtype == "ENHANCE":
                    result = self._execute_enhance(decision)
                elif dtype == "CREATE":
                    result = self._execute_create(decision)
                elif dtype == "KILL":
                    result = self._execute_kill(decision)
                elif dtype == "DISCOVER":
                    result = self._execute_discover(decision)
                else:
                    result = {"status": "skipped", "reason": f"Unknown type: {dtype}"}
            except Exception as e:
                result = {"status": "failed", "error": str(e)[:200]}
                log(f"Decision execution failed: {dtype} {decision.get('op_id')}: {e}", "ERROR")

            result["decision"] = decision
            results.append(result)
            self.state.log_decision({**decision, "execution": result.get("status", "unknown")})

        return results

    def _execute_promote(self, decision):
        """Promote a winning op — create a focused run config."""
        op_id = decision["op_id"]
        op_name = decision["name"]

        # Add to promoted list
        promoted = self.state.data.setdefault("promoted_ops", [])
        if op_id not in promoted:
            promoted.append(op_id)

        # Auto-protect promoted ops
        self.state.protect_op(op_id)

        # Write a promotion brief
        brief_dir = CEO_DIR / "promotions"
        brief_dir.mkdir(parents=True, exist_ok=True)
        brief_path = brief_dir / f"promote_{op_id}_{datetime.now().strftime('%Y%m%d')}.md"
        safe_path(brief_path)

        op_data = self.xlsx.get_op_by_id(op_id)
        brief = f"# PROMOTED: {op_name} ({op_id})\n\n"
        brief += f"Score: {decision['score']}\n"
        brief += f"Revenue range: {op_data.get('REVENUE_RANGE', 'N/A') if op_data else 'N/A'}\n"
        brief += f"Promoted: {datetime.now().isoformat()}\n\n"
        brief += "## Action: Double down\n"
        brief += "- Increase cycle frequency\n"
        brief += "- Allocate more resources\n"
        brief += "- Auto-protected from kill\n"
        brief_path.write_text(brief)

        self.state.save()
        log(f"PROMOTED {op_id} ({op_name}) — score {decision['score']}")
        return {"status": "success", "action": "promoted and protected"}

    def _execute_enhance(self, decision):
        """Enhance an underperforming op by addressing its blocker."""
        op_id = decision["op_id"]
        blocker = decision.get("blocker", "unknown")

        enhanced = self.state.data.setdefault("enhanced_ops", [])
        if op_id not in enhanced:
            enhanced.append(op_id)

        # Write enhancement plan
        plan_dir = CEO_DIR / "enhancements"
        plan_dir.mkdir(parents=True, exist_ok=True)
        plan_path = plan_dir / f"enhance_{op_id}_{datetime.now().strftime('%Y%m%d')}.md"
        safe_path(plan_path)

        plan = f"# ENHANCE: {decision['name']} ({op_id})\n\n"
        plan += f"Score: {decision['score']}\n"
        plan += f"Blocker: {blocker}\n"
        plan += f"Trend: declining\n\n"
        plan += "## Blocker Resolution Plan\n"
        plan += f"1. Identify what {blocker} requires\n"
        plan += "2. Check if blocker can be automated\n"
        plan += "3. If human action needed, add to OPS/HUMAN_LOOP_QUEUE.md\n"
        plan += "4. Re-score after blocker resolved\n"
        plan_path.write_text(plan)

        # Try to run relevant scripts if the blocker suggests automation
        if "ACCOUNT" in blocker.upper():
            log(f"ENHANCE {op_id}: blocker is account-related ({blocker}) — needs human action")
        elif "SCRIPT" in blocker.upper() or "AUTO" in blocker.upper():
            # Try running relevant automation
            self._try_auto_fix(op_id, blocker)

        self.state.save()
        log(f"ENHANCED {op_id} — plan written, blocker: {blocker}")
        return {"status": "success", "action": f"enhancement plan for blocker {blocker}"}

    def _execute_create(self, decision):
        """Create a new dynamic venture from xlsx op data."""
        op_id = decision["op_id"]
        op_name = decision["name"]
        lane = decision.get("lane", "general")
        category = decision.get("category", "")

        # Write venture definition
        venture_dir = CEO_DIR / "ventures"
        venture_dir.mkdir(parents=True, exist_ok=True)
        venture_path = venture_dir / f"venture_{op_id}.json"
        safe_path(venture_path)

        op_data = self.xlsx.get_op_by_id(op_id) or {}

        venture_def = {
            "op_id": op_id,
            "name": op_name,
            "category": category,
            "lane": lane,
            "created": datetime.now().isoformat(),
            "score": decision["score"],
            "revenue_range": op_data.get("REVENUE_RANGE", ""),
            "description": str(op_data.get("DESCRIPTION", ""))[:500],
            "automation_level": op_data.get("AUTOMATION_LEVEL", ""),
            "status": "CREATED",
            "cycles_run": 0,
            "last_cycle": None,
        }
        venture_path.write_text(json.dumps(venture_def, indent=2, default=str))

        # Add to created ventures
        created = self.state.data.setdefault("created_ventures", [])
        created.append({"op_id": op_id, "name": op_name, "created": datetime.now().isoformat()})

        self.state.save()
        log(f"CREATED venture {op_id} ({op_name}) — lane: {lane}")
        return {"status": "success", "action": f"venture created: {op_name}"}

    def _execute_kill(self, decision):
        """Sunset a dead op. Does NOT delete anything — just deprioritizes."""
        op_id = decision["op_id"]

        # Safety: double-check protection
        if self.state.is_protected(op_id):
            log(f"KILL BLOCKED — {op_id} is protected", "WARN")
            return {"status": "blocked", "reason": "op is protected"}

        killed = self.state.data.setdefault("killed_ops", [])
        killed.append({
            "op_id": op_id,
            "name": decision["name"],
            "score": decision["score"],
            "killed_at": datetime.now().isoformat(),
            "reason": decision["reason"],
        })

        # Write kill record
        kill_dir = CEO_DIR / "kills"
        kill_dir.mkdir(parents=True, exist_ok=True)
        kill_path = kill_dir / f"kill_{op_id}_{datetime.now().strftime('%Y%m%d')}.md"
        safe_path(kill_path)

        record = f"# KILLED: {decision['name']} ({op_id})\n\n"
        record += f"Score: {decision['score']}\n"
        record += f"Reason: {decision['reason']}\n"
        record += f"Killed: {datetime.now().isoformat()}\n\n"
        record += "## Note\n"
        record += "This op was deprioritized, NOT deleted. All files remain intact.\n"
        record += "To resurrect: python3 AUTOMATIONS/ceo_agent.py --protect {op_id}\n"
        kill_path.write_text(record)

        self.state.save()
        log(f"KILLED {op_id} ({decision['name']}) — score {decision['score']}")
        return {"status": "success", "action": "deprioritized (files preserved)"}

    def _execute_discover(self, decision):
        """Research a discovered opportunity."""
        op_id = decision["op_id"]

        discover_dir = CEO_DIR / "discoveries"
        discover_dir.mkdir(parents=True, exist_ok=True)
        discover_path = discover_dir / f"discover_{op_id}_{datetime.now().strftime('%Y%m%d')}.md"
        safe_path(discover_path)

        record = f"# DISCOVERED: {decision['name']} ({op_id})\n\n"
        record += f"Score: {decision['score']}\n"
        record += f"Reason: {decision['reason']}\n"
        record += f"Expansion logic: {decision.get('expansion_logic', 'N/A')}\n"
        record += f"Discovered: {datetime.now().isoformat()}\n\n"
        record += "## Next Steps\n"
        record += "1. Full research on this opportunity\n"
        record += "2. Validate with real data\n"
        record += "3. If validated, CEO will CREATE in next cycle\n"
        discover_path.write_text(record)

        self.state.data["last_discover"] = datetime.now().isoformat()
        self.state.save()
        log(f"DISCOVERED {op_id} ({decision['name']})")
        return {"status": "success", "action": "discovery logged for evaluation"}

    def run_dynamic_ventures(self):
        """Read CEO-created venture JSON files and run relevant scripts for them."""
        venture_dir = CEO_DIR / "ventures"
        if not venture_dir.exists():
            log("No ventures directory yet — skipping dynamic ventures")
            return []

        results = []
        venture_files = sorted(venture_dir.glob("venture_*.json"))
        if not venture_files:
            log("No dynamic venture definitions found")
            return results

        # Map lane/category to relevant automation scripts
        lane_scripts = {
            "CONTENT": [
                ("printmaxx_agent.py", "--mission content"),
            ],
            "OUTBOUND": [
                ("printmaxx_agent.py", "--mission outreach"),
            ],
            "APP_FACTORY": [
                ("printmaxx_agent.py", "--mission apps"),
            ],
            "RESEARCH": [
                ("daily_research_orchestrator.py", "--full"),
            ],
            "SCRAPING": [
                ("twitter_alpha_scraper.py", "--all"),
                ("background_reddit_scraper.py", "--scrape"),
            ],
            "MONETIZATION": [
                ("printmaxx_agent.py", "--mission monetize"),
            ],
            "SEO": [
                ("printmaxx_agent.py", "--mission seo"),
            ],
            "SOCIAL": [
                ("printmaxx_agent.py", "--mission social"),
            ],
        }

        for vf in venture_files:
            try:
                vdef = json.loads(vf.read_text())
            except Exception as e:
                log(f"Failed to read venture {vf.name}: {e}", "WARN")
                continue

            op_id = vdef.get("op_id", "")
            v_name = vdef.get("name", "")
            v_lane = str(vdef.get("lane", "")).upper().strip()
            v_category = str(vdef.get("category", "")).upper().strip()
            v_status = vdef.get("status", "CREATED")

            # Skip killed ventures
            if v_status in ("KILLED", "PAUSED"):
                continue

            # Find matching scripts by lane, then by category
            scripts_to_run = lane_scripts.get(v_lane, [])
            if not scripts_to_run:
                scripts_to_run = lane_scripts.get(v_category, [])

            if not scripts_to_run:
                log(f"  Venture {op_id} ({v_name}): no matching scripts for lane={v_lane} cat={v_category}")
                results.append({"op_id": op_id, "name": v_name, "status": "no_scripts"})
                continue

            # Run each matched script
            for script_name, script_args in scripts_to_run:
                script_path = PROJECT / "AUTOMATIONS" / script_name
                if not script_path.exists():
                    log(f"  Script not found for venture {op_id}: {script_name}", "WARN")
                    continue

                ok, output = run_script(script_name, script_args, timeout_sec=180,
                                        label=f"venture:{op_id}:{script_name}")
                results.append({
                    "op_id": op_id,
                    "name": v_name,
                    "script": script_name,
                    "status": "success" if ok else "failed",
                })

            # Update venture run stats
            vdef["cycles_run"] = vdef.get("cycles_run", 0) + 1
            vdef["last_cycle"] = datetime.now().isoformat()
            try:
                safe_path(vf)
                vf.write_text(json.dumps(vdef, indent=2, default=str))
            except Exception as e:
                log(f"  Failed to update venture file {vf.name}: {e}", "WARN")

        log(f"Dynamic ventures: ran {len(results)} tasks across {len(venture_files)} ventures")
        return results

    def _try_auto_fix(self, op_id, blocker):
        """Attempt to automatically fix a blocker."""
        # Map common blockers to scripts
        auto_fixes = {
            "SCRIPT_BROKEN": "system_health_monitor.py --quick",
            "CRON_DEAD": "system_health_monitor.py --quick",
            "DATA_STALE": "daily_research_orchestrator.py --full",
        }
        for pattern, script in auto_fixes.items():
            if pattern in blocker.upper():
                script_path = PROJECT / "AUTOMATIONS" / script.split()[0]
                if script_path.exists():
                    try:
                        args = " ".join(script.split()[1:])
                        cmd = f"{PYTHON} {script_path} {args}"
                        safe_command(cmd)
                        subprocess.run(cmd, shell=True, capture_output=True,
                                       timeout=120, cwd=str(PROJECT))
                        log(f"Auto-fix attempted for {op_id}: {script}")
                    except Exception as e:
                        log(f"Auto-fix failed for {op_id}: {e}", "WARN")
                break

    def run_existing_ventures(self):
        """Run existing venture classes from ops_manager."""
        try:
            # Import at runtime to avoid circular deps
            sys.path.insert(0, str(PROJECT / "AUTOMATIONS"))
            from printmaxx_ops_manager import (
                VENTURE_CLASSES, OpsState, run_full_cycle, prioritize_ventures
            )

            ops_state = OpsState()
            prioritize_ventures(ops_state)
            run_full_cycle(ops_state)
            return True, "Existing ventures ran successfully"
        except Exception as e:
            log(f"Existing venture run failed: {e}", "ERROR")
            return False, str(e)[:200]


# ============================================================================
# AUDIT TRAIL — regression detection
# ============================================================================

class AuditTrail:
    """Detects regressions after CEO changes."""

    def __init__(self, state: CEOState):
        self.state = state

    def capture_baseline(self):
        """Capture performance baseline before changes."""
        baseline = {
            "ts": datetime.now().isoformat(),
            "scores": dict(self.state.data.get("op_scores", {})),
            "promoted_count": len(self.state.data.get("promoted_ops", [])),
            "killed_count": len(self.state.data.get("killed_ops", [])),
            "created_count": len(self.state.data.get("created_ventures", [])),
            "disk_gb": disk_free_gb(),
        }
        self.state.data["performance_baseline"] = baseline
        self.state.save()
        return baseline

    def check_regression(self, results):
        """Check if CEO decisions caused any regression."""
        issues = []

        # Check: did any execution fail?
        failed = [r for r in results if r.get("status") == "failed"]
        if failed:
            issues.append(f"{len(failed)} decision(s) failed to execute")

        # Check: disk space didn't drop dramatically
        disk_now = disk_free_gb()
        baseline_disk = self.state.data.get("performance_baseline", {}).get("disk_gb", disk_now)
        if isinstance(baseline_disk, (int, float)) and disk_now < baseline_disk - 5:
            issues.append(f"Disk dropped {baseline_disk - disk_now:.1f}GB during CEO cycle")

        # Check: we didn't kill too many ops in one cycle
        killed_this_cycle = sum(1 for r in results
                                if r.get("decision", {}).get("type") == "KILL"
                                and r.get("status") == "success")
        if killed_this_cycle > 3:
            issues.append(f"Killed {killed_this_cycle} ops in one cycle (aggressive)")

        # Log audit
        audit = {
            "cycle": self.state.data.get("cycles_run", 0),
            "decisions_executed": len(results),
            "failed": len(failed),
            "issues": issues,
            "regression_detected": len(issues) > 0,
            "disk_gb": disk_now,
        }
        self.state.log_audit(audit)

        if issues:
            log(f"AUDIT ISSUES: {'; '.join(issues)}", "WARN")
        else:
            log("AUDIT: clean — no regressions detected")

        return issues


# ============================================================================
# LOCK FILE — prevent double-runs
# ============================================================================

class LockGuard:
    """Prevents concurrent CEO cycles."""

    def __init__(self):
        self.lock_fd = None

    def acquire(self):
        CEO_DIR.mkdir(parents=True, exist_ok=True)
        try:
            self.lock_fd = open(LOCK_FILE, "w")
            fcntl.flock(self.lock_fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
            self.lock_fd.write(f"PID={os.getpid()} TS={datetime.now().isoformat()}\n")
            self.lock_fd.flush()
            return True
        except (IOError, OSError):
            log("CEO cycle already running (lock held)", "WARN")
            if self.lock_fd:
                self.lock_fd.close()
                self.lock_fd = None
            return False

    def release(self):
        if self.lock_fd:
            try:
                fcntl.flock(self.lock_fd, fcntl.LOCK_UN)
                self.lock_fd.close()
                LOCK_FILE.unlink(missing_ok=True)
            except Exception:
                pass
            self.lock_fd = None


# ============================================================================
# MAIN CEO CYCLE
# ============================================================================

def run_ceo_cycle(dry_run=False):
    """Run one full CEO cycle: score → decide → snapshot → execute → audit."""
    lock = LockGuard()
    if not lock.acquire():
        return

    try:
        state = CEOState()
        xlsx = XlsxIntel()
        git = GitGuard()
        scorer = VentureScorer(xlsx, state)
        brain = CEOBrain(scorer, xlsx, state)
        runner = VentureRunner(state, xlsx)
        audit = AuditTrail(state)

        disk = disk_free_gb()
        log("=" * 70)
        log(f"CEO CYCLE #{state.data['cycles_run'] + 1} — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        log(f"Disk: {disk}GB | Decisions so far: {state.data['total_decisions']}")
        log("=" * 70)

        if disk < 2:
            log("DISK CRITICAL < 2GB — CEO cycle paused", "WARN")
            return

        # 1. Score all ops
        log("Phase 1: Scoring all ops from xlsx...")
        scores = scorer.score_all()
        top5 = scores[:5]
        log(f"  Top 5: {', '.join(f'{s['op_id']}={s['total_score']}' for s in top5)}")

        # 2. CEO decisions
        log("Phase 2: CEO making strategic decisions...")
        decisions = brain.analyze_and_decide(dry_run=dry_run)

        issues = []  # track issues across all phases

        if not decisions:
            log("No strategic decisions needed this cycle — all ops stable")
        else:
            # Log decisions
            for d in decisions:
                log(f"  [{d['type']}] {d['op_id']} ({d['name']}): {d['reason']}")

            if dry_run:
                log("DRY RUN — decisions not executed")
                state.data["last_cycle"] = datetime.now().isoformat()
                state.data["cycles_run"] += 1
                state.save()
                return

            # 3. Git snapshot BEFORE changes
            log("Phase 3: Git snapshot before changes...")
            audit.capture_baseline()
            git.snapshot("ceo_pre_decisions")

            # 4. Execute decisions
            log("Phase 4: Executing decisions...")
            results = runner.execute_decisions(decisions)

            for r in results:
                d = r.get("decision", {})
                status = r.get("status", "unknown")
                log(f"  [{status.upper()}] {d.get('type', '?')} {d.get('op_id', '?')}")

            # 5. Post-decision audit
            log("Phase 5: Post-decision audit...")
            issues = audit.check_regression(results)

            if issues and any("failed" in i.lower() for i in issues):
                log("REGRESSION DETECTED — considering rollback", "WARN")
                # Only rollback if multiple failures
                failed_count = sum(1 for r in results if r.get("status") == "failed")
                if failed_count > 2:
                    log("ROLLING BACK — too many failures", "ERROR")
                    git.rollback()
                else:
                    log("Minor failures — proceeding without rollback")

            # 6. Post-change commit
            summary_parts = []
            for dtype in ["PROMOTE", "ENHANCE", "CREATE", "KILL", "DISCOVER"]:
                count = sum(1 for d in decisions if d["type"] == dtype)
                if count:
                    summary_parts.append(f"{count} {dtype.lower()}")
            summary = ", ".join(summary_parts) if summary_parts else "no changes"
            git.post_change_commit(summary)

        # 7. Optionally run existing venture cycles
        # (only every other CEO cycle to avoid overload)
        if state.data["cycles_run"] % 2 == 0:
            log("Phase 6: Running existing venture sub-agents...")
            ok, msg = runner.run_existing_ventures()
            log(f"  Ventures: {'OK' if ok else 'FAILED'} — {msg[:80]}")

        # ======================================================================
        # NEW ORCHESTRATION PHASES — CEO as full system orchestrator
        # ======================================================================

        # Phase 7: Alpha Pipeline Integration
        # Run scrapers + alpha processor (max every ALPHA_INTERVAL_HOURS)
        hours_since_alpha = _hours_since(state.data.get("last_alpha_scrape"))
        if hours_since_alpha >= ALPHA_INTERVAL_HOURS:
            log("Phase 7: Alpha Pipeline — scraping + processing...")
            alpha_results = {}

            # Twitter alpha scraper
            ok, out = run_script("twitter_alpha_scraper.py", "--all",
                                 timeout_sec=300, label="alpha:twitter")
            alpha_results["twitter_scraper"] = "ok" if ok else "failed"

            # Reddit scraper
            ok, out = run_script("background_reddit_scraper.py", "--scrape",
                                 timeout_sec=300, label="alpha:reddit")
            alpha_results["reddit_scraper"] = "ok" if ok else "failed"

            # Alpha auto processor — processes scraped data into actionable items
            ok, out = run_script("alpha_auto_processor.py", "--process-new",
                                 timeout_sec=180, label="alpha:processor")
            alpha_results["alpha_processor"] = "ok" if ok else "failed"

            state.data["last_alpha_scrape"] = datetime.now().isoformat()
            state.data["alpha_pipeline_results"] = alpha_results
            state.save()
            successes = sum(1 for v in alpha_results.values() if v == "ok")
            log(f"  Alpha pipeline: {successes}/{len(alpha_results)} succeeded")
        else:
            log(f"Phase 7: Alpha Pipeline — skipped ({hours_since_alpha:.1f}h < {ALPHA_INTERVAL_HOURS}h interval)")

        # Phase 8: Research Orchestration (once per day)
        hours_since_research = _hours_since(state.data.get("last_research"))
        if hours_since_research >= RESEARCH_INTERVAL_HOURS:
            log("Phase 8: Daily Research Orchestration...")
            ok, out = run_script("daily_research_orchestrator.py", "--full",
                                 timeout_sec=600, label="research:daily")
            state.data["last_research"] = datetime.now().isoformat()
            state.save()
            log(f"  Research: {'OK' if ok else 'FAILED'}")
        else:
            log(f"Phase 8: Research — skipped ({hours_since_research:.1f}h < {RESEARCH_INTERVAL_HOURS}h interval)")

        # Phase 9: Decision Engine — closed-loop processing
        log("Phase 9: Decision Engine — processing pending data into actions...")
        ok, out = run_script("decision_engine.py", "--cycle",
                             timeout_sec=300, label="decision:engine")
        state.data["last_decision_engine"] = datetime.now().isoformat()
        state.save()
        log(f"  Decision engine: {'OK' if ok else 'FAILED'}")

        # Phase 10: Content Generation (every CONTENT_INTERVAL_HOURS)
        hours_since_content = _hours_since(state.data.get("last_content_gen"))
        if hours_since_content >= CONTENT_INTERVAL_HOURS:
            log("Phase 10: Content Generation from intelligence...")
            content_results = {}

            ok, out = run_script("printmaxx_agent.py", "--mission content",
                                 timeout_sec=300, label="content:generate")
            content_results["content_gen"] = "ok" if ok else "failed"

            ok, out = run_script("printmaxx_agent.py", "--mission upgrade",
                                 timeout_sec=300, label="content:upgrade")
            content_results["upgrade"] = "ok" if ok else "failed"

            state.data["last_content_gen"] = datetime.now().isoformat()
            state.save()
            successes = sum(1 for v in content_results.values() if v == "ok")
            log(f"  Content: {successes}/{len(content_results)} succeeded")
        else:
            log(f"Phase 10: Content — skipped ({hours_since_content:.1f}h < {CONTENT_INTERVAL_HOURS}h interval)")

        # Phase 11: System Health Monitor
        hours_since_health = _hours_since(state.data.get("last_health_check"))
        if hours_since_health >= HEALTH_INTERVAL_HOURS:
            log("Phase 11: System Health Check...")
            ok, out = run_script("system_health_monitor.py", "--quick",
                                 timeout_sec=120, label="health:check")
            state.data["last_health_check"] = datetime.now().isoformat()

            # Parse health output for issues and attempt auto-fix
            if ok and out:
                # Look for lines containing FAIL, ERROR, BROKEN, DOWN
                health_issues = []
                for line in out.split("\n"):
                    line_upper = line.upper()
                    if any(kw in line_upper for kw in ["FAIL", "ERROR", "BROKEN", "DOWN", "DEAD"]):
                        health_issues.append(line.strip()[:200])
                state.data["health_issues"] = health_issues[-20:]  # keep last 20

                if health_issues:
                    log(f"  Health issues found: {len(health_issues)}")
                    # Attempt auto-fix by re-running health monitor with fix flag
                    fix_ok, fix_out = run_script("system_health_monitor.py", "--quick --fix",
                                                  timeout_sec=120, label="health:autofix")
                    if fix_ok:
                        log("  Auto-fix attempted")
                    else:
                        log("  Auto-fix failed or not supported", "WARN")
                else:
                    log("  System healthy — no issues detected")
            else:
                log(f"  Health check: {'OK' if ok else 'FAILED'}")

            state.save()
        else:
            log(f"Phase 11: Health — skipped ({hours_since_health:.1f}h < {HEALTH_INTERVAL_HOURS}h interval)")

        # Phase 12: Dynamic Ventures — run CEO-created venture agents
        if state.data["cycles_run"] % 3 == 0:  # every 3rd cycle to avoid overload
            log("Phase 12: Running dynamic venture agents...")
            dv_results = runner.run_dynamic_ventures()
            dv_success = sum(1 for r in dv_results if r.get("status") == "success")
            log(f"  Dynamic ventures: {dv_success}/{len(dv_results)} succeeded")

        # Phase 13: OpenClaw Local Biz Pipeline — autonomous lead gen + site building
        hours_since_openclaw = _hours_since(state.data.get("last_openclaw"))
        if hours_since_openclaw >= OPENCLAW_INTERVAL_HOURS:
            log("Phase 13: OpenClaw Local Biz Pipeline...")
            openclaw_script = PROJECT / "AUTOMATIONS" / "openclaw_local_biz.py"
            if openclaw_script.exists():
                # Rotate through cities
                city_idx = state.data.get("openclaw_city_index", 0) % len(OPENCLAW_CITIES)
                city, niche = OPENCLAW_CITIES[city_idx]
                log(f"  OpenClaw: discovering {niche} in {city} (rotation #{city_idx + 1}/{len(OPENCLAW_CITIES)})")

                # Step 1: Discover leads
                ok, out = run_script("openclaw_local_biz.py", f'--discover "{city}" {niche}',
                                     timeout_sec=120, label=f"openclaw:discover:{city}:{niche}")
                openclaw_results = {"discover": "ok" if ok else "failed", "city": city, "niche": niche}

                # Step 2: Build preview sites for F-grade leads
                if ok:
                    ok2, out2 = run_script("openclaw_local_biz.py", "--build",
                                           timeout_sec=180, label="openclaw:build")
                    openclaw_results["build"] = "ok" if ok2 else "failed"

                    # Step 3: Generate outreach
                    ok3, out3 = run_script("openclaw_local_biz.py", "--outreach",
                                           timeout_sec=60, label="openclaw:outreach")
                    openclaw_results["outreach"] = "ok" if ok3 else "failed"

                state.data["last_openclaw"] = datetime.now().isoformat()
                state.data["openclaw_city_index"] = (city_idx + 1) % len(OPENCLAW_CITIES)
                state.data["openclaw_stats"] = openclaw_results
                state.save()
                successes = sum(1 for v in openclaw_results.values() if v == "ok")
                log(f"  OpenClaw: {successes}/{len(openclaw_results)} phases succeeded ({city} {niche})")
            else:
                log("  OpenClaw script not found — skipping", "WARN")
        else:
            log(f"Phase 13: OpenClaw — skipped ({hours_since_openclaw:.1f}h < {OPENCLAW_INTERVAL_HOURS}h interval)")

        # Phase 14: Scheduled Task Management — sync cron state
        log("Phase 14: Cron/scheduled task sync...")
        try:
            cron_result = subprocess.run(
                ["crontab", "-l"], capture_output=True, text=True, timeout=10
            )
            if cron_result.returncode == 0:
                cron_lines = [
                    line.strip() for line in cron_result.stdout.split("\n")
                    if line.strip() and not line.strip().startswith("#")
                ]
                state.data["managed_crons"] = cron_lines[-50:]  # track last 50 entries
                log(f"  Cron: {len(cron_lines)} active entries tracked")
            else:
                log("  Cron: no crontab found (empty or not set)")
                state.data["managed_crons"] = []
        except Exception as e:
            log(f"  Cron sync failed: {e}", "WARN")

        state.save()

        # Phase 15: Venture Autonomy Engine — run all autonomous venture pipelines
        hours_since_autonomy = _hours_since(state.data.get("last_autonomy_run"))
        if hours_since_autonomy >= AUTONOMY_INTERVAL_HOURS:
            log("Phase 15: Venture Autonomy Engine — running all autonomous ventures...")
            autonomy_script = PROJECT / "AUTOMATIONS" / "venture_autonomy.py"
            if autonomy_script.exists():
                ok, out = run_script("venture_autonomy.py", "--run-all",
                                     timeout_sec=600, label="autonomy:run-all")
                state.data["last_autonomy_run"] = datetime.now().isoformat()

                # Parse output for stats
                autonomy_stats = {"status": "ok" if ok else "failed"}
                if ok and out:
                    for line in out.split("\n"):
                        if "Ran" in line and "ventures" in line:
                            autonomy_stats["summary"] = line.strip()[:200]
                state.data["autonomy_stats"] = autonomy_stats
                state.save()
                log(f"  Autonomy engine: {'OK' if ok else 'FAILED'}")

                # Self-management: auto-install missing, fix broken, adjust intervals, prune dead
                log("Phase 15b: Self-management — auto-install/fix/adjust/prune...")
                ok2, out2 = run_script("venture_autonomy.py", "--self-manage",
                                       timeout_sec=300, label="autonomy:self-manage")
                sm_actions = 0
                if ok2 and out2:
                    for line in out2.split("\n"):
                        if "actions taken" in line:
                            try:
                                sm_actions = int(line.split(":")[1].strip().split()[0])
                            except Exception:
                                pass
                state.data["last_self_manage"] = datetime.now().isoformat()
                state.data["self_manage_actions"] = sm_actions
                state.save()
                log(f"  Self-management: {sm_actions} actions")
            else:
                log("  Autonomy engine script not found — skipping", "WARN")
        else:
            log(f"Phase 15: Autonomy — skipped ({hours_since_autonomy:.1f}h < {AUTONOMY_INTERVAL_HOURS}h interval)")

        # Update state
        state.data["last_cycle"] = datetime.now().isoformat()
        state.data["cycles_run"] += 1
        state.save()

        log("=" * 70)
        log(f"CEO CYCLE COMPLETE — {len(decisions)} decisions, {len(issues)} issues")
        log("=" * 70)

    except Exception as e:
        log(f"CEO CYCLE FATAL ERROR: {e}", "ERROR")
        import traceback
        log(traceback.format_exc(), "ERROR")
    finally:
        lock.release()


# ============================================================================
# STATUS DASHBOARD
# ============================================================================

def show_status():
    """Show full CEO agent status dashboard."""
    state = CEOState()
    xlsx = XlsxIntel()
    scorer = VentureScorer(xlsx, state)

    print("=" * 70)
    print("PRINTMAXX CEO AGENT — STATUS DASHBOARD")
    print("=" * 70)
    print(f"Cycles run:       {state.data['cycles_run']}")
    print(f"Total decisions:  {state.data['total_decisions']}")
    print(f"Last cycle:       {state.data.get('last_cycle', 'never')}")
    print(f"Last discovery:   {state.data.get('last_discover', 'never')}")
    print(f"Last alpha:       {state.data.get('last_alpha_scrape', 'never')}")
    print(f"Last research:    {state.data.get('last_research', 'never')}")
    print(f"Last content:     {state.data.get('last_content_gen', 'never')}")
    print(f"Last health:      {state.data.get('last_health_check', 'never')}")
    print(f"Last decision:    {state.data.get('last_decision_engine', 'never')}")
    print(f"Last openclaw:    {state.data.get('last_openclaw', 'never')}")
    print(f"Last autonomy:    {state.data.get('last_autonomy_run', 'never')}")
    print(f"Last self-manage: {state.data.get('last_self_manage', 'never')} ({state.data.get('self_manage_actions', 0)} actions)")
    oc_idx = state.data.get("openclaw_city_index", 0)
    if oc_idx < len(OPENCLAW_CITIES):
        next_city, next_niche = OPENCLAW_CITIES[oc_idx]
        print(f"Next openclaw:    {next_niche} in {next_city} (#{oc_idx + 1}/{len(OPENCLAW_CITIES)})")
    print(f"Managed crons:    {len(state.data.get('managed_crons', []))}")
    print(f"Disk free:        {disk_free_gb()}GB")

    # Score all ops
    scores = scorer.score_all()

    print(f"\nALL OPS SCORED ({len(scores)} total):")
    print(f"{'RANK':>4} {'OP_ID':<6} {'NAME':<40} {'SCORE':>6} {'STATUS':<10}")
    print("-" * 70)

    for i, s in enumerate(scores[:25]):
        status = ""
        if s["op_id"] in state.data.get("promoted_ops", []):
            status = "PROMOTED"
        elif s["op_id"] in [k.get("op_id") for k in state.data.get("killed_ops", [])]:
            status = "KILLED"
        elif s["op_id"] in [c.get("op_id") for c in state.data.get("created_ventures", [])]:
            status = "CREATED"
        elif s["is_protected"]:
            status = "PROTECTED"
        elif s["blocker"]:
            status = "BLOCKED"

        name = s["name"][:38] if s["name"] else ""
        print(f"#{i+1:>3} {s['op_id']:<6} {name:<40} {s['total_score']:>5.1f} {status:<10}")

    # Bottom 5
    if len(scores) > 5:
        print(f"\nBOTTOM 5 (kill candidates):")
        for s in scores[-5:]:
            protected = " [PROTECTED]" if s["is_protected"] else ""
            name = s["name"][:38] if s["name"] else ""
            print(f"  {s['op_id']:<6} {name:<40} {s['total_score']:>5.1f}{protected}")

    # Protected ops
    protected = state.data.get("protected_ops", [])
    if protected:
        print(f"\nPROTECTED OPS ({len(protected)}):")
        for p in protected:
            print(f"  {p}")

    # Recent decisions
    if DECISION_LOG.exists():
        print(f"\nRECENT DECISIONS (last 10):")
        lines = DECISION_LOG.read_text().strip().split("\n")
        for line in lines[-10:]:
            try:
                d = json.loads(line)
                print(f"  [{d.get('type', '?')}] {d.get('ts', '')[:16]} "
                      f"{d.get('op_id', '?')} — {d.get('execution', '?')}")
            except Exception:
                pass

    # Audit issues
    if AUDIT_LOG.exists():
        print(f"\nRECENT AUDITS (last 5):")
        lines = AUDIT_LOG.read_text().strip().split("\n")
        for line in lines[-5:]:
            try:
                a = json.loads(line)
                status = "CLEAN" if not a.get("regression_detected") else "ISSUES"
                print(f"  [{status}] cycle {a.get('cycle', '?')}: "
                      f"{a.get('decisions_executed', 0)} decisions, "
                      f"{a.get('failed', 0)} failed")
            except Exception:
                pass

    # Alpha pipeline results
    alpha_results = state.data.get("alpha_pipeline_results", {})
    if alpha_results:
        print(f"\nALPHA PIPELINE (last run):")
        for k, v in alpha_results.items():
            print(f"  {k}: {v}")

    # Health issues
    health_issues = state.data.get("health_issues", [])
    if health_issues:
        print(f"\nHEALTH ISSUES ({len(health_issues)}):")
        for issue in health_issues[-5:]:
            print(f"  {issue[:80]}")

    # Managed crons
    managed_crons = state.data.get("managed_crons", [])
    if managed_crons:
        print(f"\nMANAGED CRONS ({len(managed_crons)}):")
        for cron in managed_crons[-10:]:
            print(f"  {cron[:80]}")

    # Dynamic ventures
    venture_dir = CEO_DIR / "ventures"
    if venture_dir.exists():
        venture_files = list(venture_dir.glob("venture_*.json"))
        if venture_files:
            print(f"\nDYNAMIC VENTURES ({len(venture_files)}):")
            for vf in venture_files[:10]:
                try:
                    vdef = json.loads(vf.read_text())
                    v_name = vdef.get("name", "")[:35]
                    v_status = vdef.get("status", "?")
                    v_cycles = vdef.get("cycles_run", 0)
                    v_lane = vdef.get("lane", "?")
                    print(f"  {vdef.get('op_id', '?'):<6} {v_name:<35} {v_status:<10} lane={v_lane} runs={v_cycles}")
                except Exception:
                    pass

    # Config
    config = state.data.get("config", {})
    print(f"\nCONFIG:")
    print(f"  Max changes/cycle: {config.get('max_changes', MAX_CHANGES_PER_CYCLE)}")
    print(f"  Kill threshold:    < {config.get('kill_threshold', KILL_THRESHOLD)}")
    print(f"  Protect threshold: > {config.get('protect_threshold', PROTECT_THRESHOLD)}")
    print(f"  Promote threshold: > {config.get('promote_threshold', PROMOTE_THRESHOLD)}")
    print(f"  Cycle interval:    {CYCLE_INTERVAL_HOURS}h")
    print(f"  Discover interval: {DISCOVER_INTERVAL_HOURS}h")
    print(f"  Alpha interval:    {ALPHA_INTERVAL_HOURS}h")
    print(f"  Research interval: {RESEARCH_INTERVAL_HOURS}h")
    print(f"  Content interval:  {CONTENT_INTERVAL_HOURS}h")
    print(f"  Health interval:   {HEALTH_INTERVAL_HOURS}h")
    print(f"  OpenClaw interval: {OPENCLAW_INTERVAL_HOURS}h")

    # OpenClaw stats
    oc_stats = state.data.get("openclaw_stats", {})
    if oc_stats:
        print(f"\nOPENCLAW PIPELINE (last run):")
        for k, v in oc_stats.items():
            print(f"  {k}: {v}")
        # Also show live stats if the script exists
        oc_script = PROJECT / "AUTOMATIONS" / "openclaw_local_biz.py"
        if oc_script.exists():
            try:
                oc_state_file = PROJECT / "AUTOMATIONS" / "leads" / "openclaw" / ".openclaw_state.json"
                if oc_state_file.exists():
                    oc_live = json.loads(oc_state_file.read_text())
                    print(f"  total_leads: {oc_live.get('total_leads', 0)}")
                    print(f"  sites_built: {oc_live.get('total_sites_built', 0)}")
            except Exception:
                pass

    print("=" * 70)


# ============================================================================
# DAEMON MODE — 24/7 operation
# ============================================================================

def run_daemon():
    """Run the CEO agent 24/7."""
    log("CEO AGENT DAEMON STARTING — 24/7 autonomous mode")
    log(f"Cycle every {CYCLE_INTERVAL_HOURS}h | Discover every {DISCOVER_INTERVAL_HOURS}h")
    log(f"Max {MAX_CHANGES_PER_CYCLE} changes/cycle | Kill < {KILL_THRESHOLD} | Promote > {PROMOTE_THRESHOLD}")

    while True:
        try:
            run_ceo_cycle()

            next_run = datetime.now() + timedelta(hours=CYCLE_INTERVAL_HOURS)
            log(f"Next cycle at {next_run.strftime('%H:%M')}")
            time.sleep(CYCLE_INTERVAL_HOURS * 3600)

        except KeyboardInterrupt:
            log("CEO DAEMON STOPPED by user")
            break
        except Exception as e:
            log(f"DAEMON ERROR: {e} — retrying in 10 min", "ERROR")
            time.sleep(600)


# ============================================================================
# CLI
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description="PRINTMAXX CEO Agent")
    parser.add_argument("--status", action="store_true", help="Show status dashboard")
    parser.add_argument("--daemon", action="store_true", help="Run 24/7")
    parser.add_argument("--score", action="store_true", help="Score all ops from xlsx")
    parser.add_argument("--decide", action="store_true", help="Show decisions (dry run)")
    parser.add_argument("--run", action="store_true", help="Run one CEO cycle")
    parser.add_argument("--protect", type=str, help="Add op to protected list")
    parser.add_argument("--rollback", action="store_true", help="Rollback last CEO changes")
    parser.add_argument("--alpha", action="store_true", help="Run alpha pipeline only")
    parser.add_argument("--research", action="store_true", help="Run research pipeline only")
    parser.add_argument("--content", action="store_true", help="Run content generation only")
    parser.add_argument("--health", action="store_true", help="Run system health check only")
    parser.add_argument("--decision-engine", action="store_true", help="Run decision engine only")
    parser.add_argument("--ventures", action="store_true", help="Run dynamic ventures only")
    parser.add_argument("--cron-list", action="store_true", help="Show managed cron entries")
    parser.add_argument("--cron-add", type=str, help="Add a cron entry (e.g., '0 */2 * * * python3 ...')")
    args = parser.parse_args()

    if args.status:
        show_status()
    elif args.daemon:
        run_daemon()
    elif args.score:
        state = CEOState()
        xlsx = XlsxIntel()
        scorer = VentureScorer(xlsx, state)
        scores = scorer.score_all()
        print(f"\nScored {len(scores)} ops. Top 10:")
        for i, s in enumerate(scores[:10]):
            name = s["name"][:35] if s["name"] else ""
            print(f"  #{i+1} {s['op_id']:<6} {name:<35} score={s['total_score']:.1f} "
                  f"(R={s['readiness']:.0f} A={s['automation']:.0f} "
                  f"S={s['signals']:.0f} $={s['revenue_potential']:.0f} X={s['synergy']:.0f})")
    elif args.decide:
        run_ceo_cycle(dry_run=True)
    elif args.protect:
        state = CEOState()
        state.protect_op(args.protect)
        print(f"Protected: {args.protect}")
    elif args.rollback:
        git = GitGuard()
        # Find last CEO snapshot
        try:
            result = subprocess.run(
                ["git", "log", "--oneline", "--grep=CEO snapshot", "-1"],
                cwd=str(PROJECT), capture_output=True, text=True, timeout=10
            )
            if result.stdout.strip():
                hash_val = result.stdout.strip().split()[0]
                print(f"Rolling back to: {result.stdout.strip()}")
                git.last_snapshot_hash = hash_val
                git.rollback()
                print("Rollback complete (changes are staged, review with git diff --cached)")
            else:
                print("No CEO snapshot found to rollback to")
        except Exception as e:
            print(f"Rollback failed: {e}")
    elif args.alpha:
        log("Running alpha pipeline (standalone)...")
        ok1, _ = run_script("twitter_alpha_scraper.py", "--all", label="alpha:twitter")
        ok2, _ = run_script("background_reddit_scraper.py", "--scrape", label="alpha:reddit")
        ok3, _ = run_script("alpha_auto_processor.py", "--process-new", label="alpha:processor")
        print(f"Alpha pipeline: twitter={'OK' if ok1 else 'FAIL'} reddit={'OK' if ok2 else 'FAIL'} processor={'OK' if ok3 else 'FAIL'}")
    elif args.research:
        log("Running research pipeline (standalone)...")
        ok, out = run_script("daily_research_orchestrator.py", "--full", timeout_sec=600, label="research:daily")
        print(f"Research: {'OK' if ok else 'FAILED'}")
        if out:
            print(out[-500:])
    elif args.content:
        log("Running content generation (standalone)...")
        ok1, _ = run_script("printmaxx_agent.py", "--mission content", label="content:generate")
        ok2, _ = run_script("printmaxx_agent.py", "--mission upgrade", label="content:upgrade")
        print(f"Content: generate={'OK' if ok1 else 'FAIL'} upgrade={'OK' if ok2 else 'FAIL'}")
    elif args.health:
        log("Running system health check (standalone)...")
        ok, out = run_script("system_health_monitor.py", "--quick", label="health:check")
        print(f"Health: {'OK' if ok else 'FAILED'}")
        if out:
            print(out[-1000:])
    elif args.decision_engine:
        log("Running decision engine (standalone)...")
        ok, out = run_script("decision_engine.py", "--cycle", label="decision:engine")
        print(f"Decision engine: {'OK' if ok else 'FAILED'}")
        if out:
            print(out[-500:])
    elif args.ventures:
        log("Running dynamic ventures (standalone)...")
        state = CEOState()
        xlsx = XlsxIntel()
        runner = VentureRunner(state, xlsx)
        results = runner.run_dynamic_ventures()
        for r in results:
            print(f"  {r.get('op_id', '?')}: {r.get('status', '?')} ({r.get('script', '')})")
        print(f"Total: {len(results)} venture tasks")
    elif args.cron_list:
        state = CEOState()
        managed = state.data.get("managed_crons", [])
        print(f"Managed cron entries ({len(managed)}):")
        for c in managed:
            print(f"  {c}")
        # Also show live crontab
        try:
            result = subprocess.run(["crontab", "-l"], capture_output=True, text=True, timeout=10)
            if result.returncode == 0:
                live_lines = [l.strip() for l in result.stdout.split("\n") if l.strip() and not l.strip().startswith("#")]
                print(f"\nLive crontab ({len(live_lines)} entries):")
                for l in live_lines:
                    print(f"  {l}")
        except Exception as e:
            print(f"Could not read crontab: {e}")
    elif args.cron_add:
        cron_entry = args.cron_add.strip()
        if not cron_entry:
            print("Empty cron entry — nothing to add")
        else:
            try:
                # Read current crontab
                result = subprocess.run(["crontab", "-l"], capture_output=True, text=True, timeout=10)
                current = result.stdout if result.returncode == 0 else ""

                # Check for duplicate
                if cron_entry in current:
                    print(f"Cron entry already exists: {cron_entry}")
                else:
                    # Append new entry
                    new_crontab = current.rstrip("\n") + "\n" + cron_entry + "\n"
                    install = subprocess.run(
                        ["crontab", "-"], input=new_crontab, capture_output=True,
                        text=True, timeout=10
                    )
                    if install.returncode == 0:
                        print(f"Added cron entry: {cron_entry}")
                        # Track in state
                        state = CEOState()
                        managed = state.data.setdefault("managed_crons", [])
                        managed.append(cron_entry)
                        state.save()
                    else:
                        print(f"Failed to install cron: {install.stderr}")
            except Exception as e:
                print(f"Cron add failed: {e}")
    elif args.run:
        run_ceo_cycle()
    else:
        # Default: one cycle
        run_ceo_cycle()


if __name__ == "__main__":
    main()
