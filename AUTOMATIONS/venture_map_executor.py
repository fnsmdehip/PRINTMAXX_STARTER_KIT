#!/usr/bin/env python3
"""Plan/execute all Venture Automation Map rows with cooldown + dedupe controls.

Reads the `VENTURE_AUTOMATION_MAP` sheet from the latest enhanced Master Ops
workbook, prioritizes ventures by readiness/score/signal, and executes
command templates with:
  - blocker-aware gating
  - per-lane caps
  - per-command dedupe
  - persistent cooldown state between runs
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import subprocess
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Tuple
from zipfile import BadZipFile

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent.parent
OPS_DIR = BASE_DIR / "OPS"
LEDGER_DIR = BASE_DIR / "LEDGER"
OUT_DIR = BASE_DIR / "output" / "venture_map_exec"
STATE_JSON = LEDGER_DIR / "VENTURE_MAP_EXEC_STATE.json"
RUNS_CSV = LEDGER_DIR / "VENTURE_MAP_EXEC_RUNS.csv"

QUEUE_MD = OPS_DIR / "HUMAN_LOOP_QUEUE.md"
APPROVALS_CSV = OPS_DIR / "HUMAN_APPROVALS.csv"

READINESS_RANK = {
    "READY": 0,
    "BLOCKED": 1,
    "BUILD": 2,
    "EXPLORE": 3,
}

SOURCE_RANK = {
    "PRIORITY_LAUNCH": 0,
    "LLM_ALPHA_THESIS": 1,
}

BLOCKER_ALIASES = {
    "GUMROAD_ACCOUNT": ["ACCOUNT_GUMROAD", "GUMROAD_API_TOKEN"],
    "FIVERR_UPWORK_ACCOUNT": ["ACCOUNT_FIVERR", "ACCOUNT_UPWORK", "ACCOUNT_FIVERR_UPWORK"],
    "X_MULTI_ACCOUNT_STACK": ["ACCOUNT_X", "ACCOUNT_X_MULTI", "ACCOUNT_X_STACK"],
    "STORE_ACCOUNT_AND_PAYMENT": ["PAYMENT_RISK_APPROVED", "DEPLOY_APPS", "STORE_ACCOUNT"],
    "ACCOUNT_EBAY/ETSY/AMAZON": ["ACCOUNT_EBAY", "ACCOUNT_ETSY", "ACCOUNT_AMAZON"],
}


def now_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def parse_iso(ts: str) -> datetime | None:
    s = str(ts or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def clean(v: Any) -> str:
    return str(v or "").strip()


def safe_int(v: Any, default: int = 0) -> int:
    try:
        return int(float(str(v).strip()))
    except Exception:
        return default


def sha12(text: str) -> str:
    return hashlib.sha1(text.encode("utf-8", errors="replace")).hexdigest()[:12]


def normalize_key(key: str) -> str:
    return clean(key).upper().replace("`", "")


def find_latest_workbook(explicit: str) -> Path:
    if explicit:
        p = Path(explicit).expanduser().resolve()
        if p.exists():
            return p
        raise FileNotFoundError(f"Workbook not found: {p}")

    candidates = sorted(BASE_DIR.glob("PRINTMAXX_MASTER_OPS_ENHANCED_*.xlsx"))
    if not candidates:
        raise FileNotFoundError("No PRINTMAXX_MASTER_OPS_ENHANCED_*.xlsx found")

    last_error = ""
    for candidate in reversed(candidates):
        try:
            wb = load_workbook(candidate, read_only=True, data_only=True)
            wb.close()
            return candidate
        except Exception as exc:
            last_error = str(exc)
            continue
    raise FileNotFoundError(f"No readable enhanced workbook found. last_error={last_error}")


def parse_approved_keys(path: Path) -> set[str]:
    out: set[str] = set()
    if not path.exists():
        return out
    with open(path, "r", encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            key = normalize_key(row.get("key", ""))
            status = normalize_key(row.get("status", ""))
            if key and status in {"APPROVED", "ACTIVE", "TRUE", "YES"}:
                out.add(key)
    return out


def parse_pending_keys(path: Path) -> set[str]:
    out: set[str] = set()
    if not path.exists():
        return out
    for raw in path.read_text(encoding="utf-8", errors="replace").splitlines():
        s = raw.strip()
        if not s.startswith("- [ ]"):
            continue
        if "`" in s:
            bits = s.split("`")
            if len(bits) >= 3:
                out.add(normalize_key(bits[1]))
                continue
        out.add(normalize_key(s))
    return out


def blocker_is_unapproved(blocker_key: str, approved: set[str], pending: set[str]) -> bool:
    key = normalize_key(blocker_key)
    if key in {"", "NONE", "N/A"}:
        return False
    aliases = [key] + BLOCKER_ALIASES.get(key, [])
    normalized_aliases = [normalize_key(a) for a in aliases]
    if any(a in approved for a in normalized_aliases):
        return False
    if any(a in pending for a in normalized_aliases):
        return True
    # Conservative default: if blocker exists and not explicitly approved, treat as blocked.
    return True


def read_venture_rows(path: Path) -> List[Dict[str, Any]]:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except BadZipFile as exc:
        raise FileNotFoundError(f"Workbook is invalid/corrupt: {path} ({exc})") from exc

    if "VENTURE_AUTOMATION_MAP" not in wb.sheetnames:
        wb.close()
        return []

    ws = wb["VENTURE_AUTOMATION_MAP"]
    headers = [clean(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(headers) if h}

    required = {"VENTURE_ID", "VENTURE_NAME", "LANE", "READINESS", "BLOCKER_KEY", "COMMAND_TEMPLATE"}
    if not required.issubset(set(idx.keys())):
        wb.close()
        return []

    rows: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        venture_id = clean(ws.cell(r, idx["VENTURE_ID"]).value)
        if not venture_id:
            continue
        cmd = clean(ws.cell(r, idx["COMMAND_TEMPLATE"]).value)
        rows.append(
            {
                "SOURCE": clean(ws.cell(r, idx.get("SOURCE", 0)).value) if idx.get("SOURCE") else "",
                "VENTURE_ID": venture_id,
                "VENTURE_NAME": clean(ws.cell(r, idx["VENTURE_NAME"]).value),
                "LANE": clean(ws.cell(r, idx["LANE"]).value).lower() or "other",
                "READINESS": clean(ws.cell(r, idx["READINESS"]).value).upper() or "EXPLORE",
                "BLOCKER_KEY": clean(ws.cell(r, idx["BLOCKER_KEY"]).value).upper() or "NONE",
                "AUTOMATION_SCORE_100": safe_int(ws.cell(r, idx.get("AUTOMATION_SCORE_100", 0)).value if idx.get("AUTOMATION_SCORE_100") else 0, 0),
                "SIGNAL_COUNT": safe_int(ws.cell(r, idx.get("SIGNAL_COUNT", 0)).value if idx.get("SIGNAL_COUNT") else 0, 0),
                "COMMAND_TEMPLATE": cmd,
                "NEXT_ACTION": clean(ws.cell(r, idx.get("NEXT_ACTION", 0)).value) if idx.get("NEXT_ACTION") else "",
            }
        )
    wb.close()
    return rows


def row_sort_key(row: Dict[str, Any]) -> Tuple[int, int, int, int, str]:
    readiness = clean(row.get("READINESS")).upper()
    source = clean(row.get("SOURCE")).upper()
    score = safe_int(row.get("AUTOMATION_SCORE_100"), 0)
    signal = safe_int(row.get("SIGNAL_COUNT"), 0)
    return (
        READINESS_RANK.get(readiness, 9),
        SOURCE_RANK.get(source, 9),
        -score,
        -signal,
        clean(row.get("VENTURE_ID")),
    )


def cooldown_minutes_for_row(row: Dict[str, Any], args: argparse.Namespace) -> int:
    readiness = clean(row.get("READINESS")).upper()
    if readiness == "READY":
        return max(0, args.cooldown_ready_min)
    if readiness == "BLOCKED":
        return max(0, args.cooldown_blocked_min)
    if readiness == "BUILD":
        return max(0, args.cooldown_build_min)
    return max(0, args.cooldown_explore_min)


def load_state() -> Dict[str, Any]:
    if not STATE_JSON.exists():
        return {
            "version": "2026-02-19",
            "updated_at": "",
            "command_last_run": {},
            "venture_last_run": {},
        }
    try:
        payload = json.loads(STATE_JSON.read_text(encoding="utf-8"))
        if not isinstance(payload, dict):
            raise ValueError("state must be object")
        payload.setdefault("command_last_run", {})
        payload.setdefault("venture_last_run", {})
        return payload
    except Exception:
        return {
            "version": "2026-02-19",
            "updated_at": "",
            "command_last_run": {},
            "venture_last_run": {},
        }


def save_state(state: Dict[str, Any]) -> None:
    LEDGER_DIR.mkdir(parents=True, exist_ok=True)
    state["updated_at"] = now_iso()
    STATE_JSON.write_text(json.dumps(state, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def ensure_runs_csv() -> None:
    LEDGER_DIR.mkdir(parents=True, exist_ok=True)
    if RUNS_CSV.exists():
        return
    with open(RUNS_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "timestamp",
                "workbook",
                "command_hash",
                "status",
                "return_code",
                "duration_sec",
                "row_refs",
                "venture_ids",
                "command",
            ]
        )


def command_or_venture_on_cooldown(
    row: Dict[str, Any],
    command_hash: str,
    state: Dict[str, Any],
    cooldown_min: int,
    ignore_cooldown: bool,
) -> Tuple[bool, int]:
    if ignore_cooldown or cooldown_min <= 0:
        return False, 0

    now = datetime.now()
    threshold = timedelta(minutes=cooldown_min)
    venture_id = clean(row.get("VENTURE_ID"))
    command_last = state.get("command_last_run", {}).get(command_hash, {})
    venture_last = state.get("venture_last_run", {}).get(venture_id, {})

    candidates = []
    cmd_ts = parse_iso(clean(command_last.get("last_run_at")))
    if cmd_ts:
        candidates.append(cmd_ts)
    venture_ts = parse_iso(clean(venture_last.get("last_run_at")))
    if venture_ts:
        candidates.append(venture_ts)
    if not candidates:
        return False, 0

    latest = max(candidates)
    elapsed = now - latest
    if elapsed >= threshold:
        return False, 0
    remaining = threshold - elapsed
    remaining_min = max(1, int(remaining.total_seconds() // 60))
    return True, remaining_min


def plan_rows(
    rows: List[Dict[str, Any]],
    state: Dict[str, Any],
    approved: set[str],
    pending: set[str],
    args: argparse.Namespace,
) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
    rows_sorted = sorted(rows, key=row_sort_key)
    planned: List[Dict[str, Any]] = []
    skip = Counter()

    lane_counts: Dict[str, int] = defaultdict(int)
    seen_venture: set[str] = set()
    planned_command_hashes: set[str] = set()
    unique_run_commands = 0

    for row in rows_sorted:
        if len(planned) >= max(1, args.max_rows):
            skip["max_rows_reached"] += 1
            break

        venture_id = clean(row.get("VENTURE_ID"))
        lane = clean(row.get("LANE")).lower() or "other"
        readiness = clean(row.get("READINESS")).upper() or "EXPLORE"
        score = safe_int(row.get("AUTOMATION_SCORE_100"), 0)
        command = clean(row.get("COMMAND_TEMPLATE"))
        blocker = clean(row.get("BLOCKER_KEY")).upper() or "NONE"

        if not venture_id:
            skip["missing_venture_id"] += 1
            continue
        if venture_id in seen_venture:
            skip["duplicate_venture_id"] += 1
            continue
        seen_venture.add(venture_id)

        if not args.include_explore and readiness == "EXPLORE":
            skip["explore_excluded"] += 1
            continue
        if score < args.min_score:
            skip["below_min_score"] += 1
            continue
        if lane_counts[lane] >= max(1, args.max_per_lane):
            skip["lane_cap"] += 1
            continue
        if not command:
            skip["missing_command"] += 1
            continue

        blocked = blocker_is_unapproved(blocker, approved=approved, pending=pending)
        if blocked and not args.allow_blocked_prework:
            skip["blocked_skipped"] += 1
            continue

        command_hash = sha12(command)
        cooldown_min = cooldown_minutes_for_row(row, args)
        cooldown_hit, remaining_min = command_or_venture_on_cooldown(
            row=row,
            command_hash=command_hash,
            state=state,
            cooldown_min=cooldown_min,
            ignore_cooldown=bool(args.ignore_cooldown),
        )
        if cooldown_hit:
            skip["cooldown"] += 1
            continue

        run_mode = "run"
        deduped_to = ""
        if args.dedupe_commands and command_hash in planned_command_hashes:
            run_mode = "dedupe"
            deduped_to = command_hash
        else:
            if unique_run_commands >= max(1, args.max_commands):
                skip["max_commands_reached"] += 1
                continue
            planned_command_hashes.add(command_hash)
            unique_run_commands += 1

        lane_counts[lane] += 1
        planned.append(
            {
                **row,
                "COMMAND_HASH": command_hash,
                "RUN_MODE": run_mode,
                "DEDUPED_TO": deduped_to,
                "BLOCKED_BY_APPROVAL": blocked,
                "COOLDOWN_MIN": cooldown_min,
                "COOLDOWN_REMAINING_MIN": remaining_min,
            }
        )

    return planned, {k: int(v) for k, v in skip.items()}


def run_commands(planned: List[Dict[str, Any]], timeout_sec: int, apply: bool) -> List[Dict[str, Any]]:
    command_map: Dict[str, str] = {}
    refs: Dict[str, List[str]] = defaultdict(list)
    for row in planned:
        h = clean(row.get("COMMAND_HASH"))
        venture_id = clean(row.get("VENTURE_ID"))
        refs[h].append(venture_id)
        if clean(row.get("RUN_MODE")) == "run":
            command_map[h] = clean(row.get("COMMAND_TEMPLATE"))

    if not apply:
        out = []
        for h, command in command_map.items():
            out.append(
                {
                    "command_hash": h,
                    "command": command,
                    "status": "PLANNED",
                    "return_code": 0,
                    "duration_sec": 0.0,
                    "output_tail": "",
                    "row_refs": len(refs.get(h, [])),
                    "venture_ids": refs.get(h, []),
                }
            )
        return out

    out: List[Dict[str, Any]] = []
    for h, command in command_map.items():
        started = datetime.now()
        status = "FAILED"
        return_code = 1
        output_tail = ""
        try:
            proc = subprocess.run(
                ["bash", "-lc", command],
                cwd=BASE_DIR,
                capture_output=True,
                text=True,
                timeout=max(60, timeout_sec),
                check=False,
            )
            merged = ((proc.stdout or "") + ("\n" + proc.stderr if proc.stderr else "")).strip()
            return_code = int(proc.returncode)
            status = "OK" if proc.returncode == 0 else "FAILED"
            output_tail = merged[-600:]
        except subprocess.TimeoutExpired:
            status = "TIMEOUT"
            return_code = 124
            output_tail = f"Timed out after {timeout_sec}s"
        duration = round((datetime.now() - started).total_seconds(), 3)
        out.append(
            {
                "command_hash": h,
                "command": command,
                "status": status,
                "return_code": return_code,
                "duration_sec": duration,
                "output_tail": output_tail,
                "row_refs": len(refs.get(h, [])),
                "venture_ids": refs.get(h, []),
            }
        )
    return out


def update_state(state: Dict[str, Any], planned: List[Dict[str, Any]], command_results: List[Dict[str, Any]], apply: bool) -> None:
    if not apply:
        return
    result_map = {clean(r.get("command_hash")): r for r in command_results}
    timestamp = now_iso()

    command_last = state.setdefault("command_last_run", {})
    venture_last = state.setdefault("venture_last_run", {})

    for row in planned:
        venture_id = clean(row.get("VENTURE_ID"))
        command_hash = clean(row.get("COMMAND_HASH"))
        result = result_map.get(command_hash, {})
        status = clean(result.get("status")) or ("PLANNED" if not apply else "UNKNOWN")

        venture_last[venture_id] = {
            "last_run_at": timestamp,
            "status": status,
            "readiness": clean(row.get("READINESS")),
            "command_hash": command_hash,
        }

        if clean(row.get("RUN_MODE")) == "run":
            command_last[command_hash] = {
                "last_run_at": timestamp,
                "status": status,
                "command": clean(row.get("COMMAND_TEMPLATE")),
            }


def append_runs_csv(workbook: Path, command_results: List[Dict[str, Any]], apply: bool) -> None:
    ensure_runs_csv()
    mode_status = "PLANNED" if not apply else ""
    with open(RUNS_CSV, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for r in command_results:
            writer.writerow(
                [
                    now_iso(),
                    str(workbook),
                    clean(r.get("command_hash")),
                    mode_status or clean(r.get("status")),
                    safe_int(r.get("return_code"), 0),
                    float(r.get("duration_sec") or 0.0),
                    safe_int(r.get("row_refs"), 0),
                    "|".join([clean(v) for v in r.get("venture_ids", []) if clean(v)]),
                    clean(r.get("command")),
                ]
            )


def summarize_statuses(command_results: List[Dict[str, Any]]) -> Dict[str, int]:
    c = Counter()
    for r in command_results:
        c[clean(r.get("status")) or "UNKNOWN"] += 1
    return {k: int(v) for k, v in sorted(c.items())}


def write_outputs(payload: Dict[str, Any]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    manifest = OUT_DIR / "manifest.json"
    latest = OUT_DIR / "latest.md"
    manifest.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")

    lines: List[str] = []
    lines.append("# Venture Map Executor")
    lines.append("")
    lines.append(f"Generated: {clean(payload.get('generated_at'))}")
    lines.append(f"Workbook: {clean(payload.get('workbook'))}")
    lines.append(f"Mode: {clean(payload.get('mode'))}")
    lines.append("")
    lines.append("## Summary")
    lines.append("")
    lines.append(f"- rows_loaded={safe_int(payload.get('rows_loaded'), 0)}")
    lines.append(f"- rows_planned={safe_int(payload.get('rows_planned'), 0)}")
    lines.append(f"- unique_commands={safe_int(payload.get('unique_commands'), 0)}")
    lines.append(f"- status_counts={json.dumps(payload.get('status_counts', {}), sort_keys=True)}")
    lines.append(f"- skip_counts={json.dumps(payload.get('skip_counts', {}), sort_keys=True)}")
    lines.append("")
    lines.append("## Planned Rows")
    lines.append("")
    planned = payload.get("planned_rows", [])
    if not planned:
        lines.append("- none")
    else:
        for row in planned:
            lines.append(
                "- "
                f"{clean(row.get('VENTURE_ID'))} [{clean(row.get('SOURCE'))}] "
                f"lane={clean(row.get('LANE'))} readiness={clean(row.get('READINESS'))} "
                f"score={safe_int(row.get('AUTOMATION_SCORE_100'), 0)} signal={safe_int(row.get('SIGNAL_COUNT'), 0)} "
                f"blocked={bool(row.get('BLOCKED_BY_APPROVAL'))} run_mode={clean(row.get('RUN_MODE'))}"
            )
            lines.append(f"  - {clean(row.get('VENTURE_NAME'))}")
            lines.append(f"  - cmd_hash={clean(row.get('COMMAND_HASH'))}")
    lines.append("")
    lines.append("## Command Results")
    lines.append("")
    results = payload.get("command_results", [])
    if not results:
        lines.append("- none")
    else:
        for r in results:
            lines.append(
                "- "
                f"`{clean(r.get('command'))}` -> {clean(r.get('status'))} "
                f"(rc={safe_int(r.get('return_code'), 0)}, duration_sec={float(r.get('duration_sec') or 0.0):.2f}, "
                f"rows={safe_int(r.get('row_refs'), 0)})"
            )
    latest.write_text("\n".join(lines) + "\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Plan/execute all Venture Automation Map rows with cooldown and dedupe")
    ap.add_argument("--workbook", default="", help="Enhanced workbook path (default: latest)")
    ap.add_argument("--max-rows", type=int, default=120, help="Max venture rows to process")
    ap.add_argument("--max-per-lane", type=int, default=25, help="Max venture rows per lane")
    ap.add_argument("--max-commands", type=int, default=20, help="Max unique commands to run")
    ap.add_argument("--min-score", type=int, default=0, help="Skip rows below this automation score")
    ap.add_argument("--include-explore", action="store_true", default=True, help="Include EXPLORE readiness rows (default: enabled)")
    ap.add_argument("--exclude-explore", dest="include_explore", action="store_false", help="Exclude EXPLORE rows")
    ap.add_argument("--allow-blocked-prework", action="store_true", default=True, help="Allow safe prework for blocked rows (default: enabled)")
    ap.add_argument("--no-allow-blocked-prework", dest="allow_blocked_prework", action="store_false", help="Skip rows with unresolved blockers")
    ap.add_argument("--dedupe-commands", action="store_true", default=True, help="Run each command template once per pass (default: enabled)")
    ap.add_argument("--no-dedupe-commands", dest="dedupe_commands", action="store_false", help="Allow repeated commands in same pass")
    ap.add_argument("--cooldown-ready-min", type=int, default=180, help="Cooldown for READY rows")
    ap.add_argument("--cooldown-blocked-min", type=int, default=360, help="Cooldown for BLOCKED rows")
    ap.add_argument("--cooldown-build-min", type=int, default=720, help="Cooldown for BUILD rows")
    ap.add_argument("--cooldown-explore-min", type=int, default=1440, help="Cooldown for EXPLORE rows")
    ap.add_argument("--ignore-cooldown", action="store_true", help="Ignore persisted cooldown state")
    ap.add_argument("--reset-state", action="store_true", help="Delete persisted cooldown state before planning")
    ap.add_argument("--apply", action="store_true", help="Execute commands (default: dry run)")
    ap.add_argument("--timeout-sec", type=int, default=1200, help="Per-command timeout")
    return ap.parse_args()


def main() -> int:
    args = parse_args()
    workbook = find_latest_workbook(args.workbook)

    if args.reset_state and STATE_JSON.exists():
        STATE_JSON.unlink()

    rows = read_venture_rows(workbook)
    state = load_state()
    approved = parse_approved_keys(APPROVALS_CSV)
    pending = parse_pending_keys(QUEUE_MD)

    planned, skip_counts = plan_rows(
        rows=rows,
        state=state,
        approved=approved,
        pending=pending,
        args=args,
    )
    command_results = run_commands(
        planned=planned,
        timeout_sec=max(60, args.timeout_sec),
        apply=bool(args.apply),
    )

    update_state(state, planned, command_results, apply=bool(args.apply))
    save_state(state)
    append_runs_csv(workbook, command_results, apply=bool(args.apply))

    unique_commands = len({clean(r.get("command_hash")) for r in command_results if clean(r.get("command_hash"))})
    status_counts = summarize_statuses(command_results)
    payload = {
        "generated_at": now_iso(),
        "workbook": str(workbook),
        "mode": "apply" if args.apply else "dry_run",
        "rows_loaded": len(rows),
        "rows_planned": len(planned),
        "unique_commands": unique_commands,
        "skip_counts": skip_counts,
        "status_counts": status_counts,
        "planned_rows": planned,
        "command_results": command_results,
        "state_path": str(STATE_JSON),
        "runs_csv": str(RUNS_CSV),
    }
    write_outputs(payload)

    print("venture_map_executor: wrote")
    print(f"- {OUT_DIR / 'manifest.json'}")
    print(f"- {OUT_DIR / 'latest.md'}")
    print(f"- rows_loaded={len(rows)} rows_planned={len(planned)} unique_commands={unique_commands}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
