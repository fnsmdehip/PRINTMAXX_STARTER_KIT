from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Palette
D = PatternFill('solid', fgColor='0D1117')
H = PatternFill('solid', fgColor='0A1628')
S = PatternFill('solid', fgColor='112240')
R1 = PatternFill('solid', fgColor='0D1117')
R2 = PatternFill('solid', fgColor='161B22')
G = PatternFill('solid', fgColor='0B3D0B')
W = PatternFill('solid', fgColor='2D2500')
RD = PatternFill('solid', fgColor='2D0B0B')
A = PatternFill('solid', fgColor='0B2D3D')
P1 = PatternFill('solid', fgColor='0B3D2B')
ALPHA = PatternFill('solid', fgColor='1A0B3D')

FT = Font(name='Arial', size=20, bold=True, color='00D4FF')
FS = Font(name='Arial', size=12, bold=True, color='8B949E')
FH = Font(name='Arial', size=11, bold=True, color='00D4FF')
FSC = Font(name='Arial', size=12, bold=True, color='58A6FF')
FB = Font(name='Arial', size=10, color='C9D1D9')
FW = Font(name='Arial', size=10, bold=True, color='FFFFFF')
FC = Font(name='Arial', size=10, bold=True, color='00D4FF')
FG = Font(name='Arial', size=10, bold=True, color='3FB950')
FGO = Font(name='Arial', size=10, bold=True, color='F0C000')
FO = Font(name='Arial', size=10, color='F0883E')
FR = Font(name='Arial', size=10, color='F85149')
FP = Font(name='Arial', size=10, bold=True, color='A371F7')
FHL = Font(name='Arial', size=11, bold=True, color='FFFFFF')
FSM = Font(name='Arial', size=9, color='8B949E')
FALPHA = Font(name='Arial', size=10, bold=True, color='E040FB')

BD = Border(bottom=Side(style='thin', color='21262D'))
BS = Border(bottom=Side(style='medium', color='00D4FF'))
AC = Alignment(horizontal='center', vertical='center')
AL = Alignment(horizontal='left', vertical='center', wrap_text=True)
AW = Alignment(horizontal='left', vertical='top', wrap_text=True)

def sw(ws, w):
    for i, v in enumerate(w, 1): ws.column_dimensions[get_column_letter(i)].width = v
def fb(ws, r, c, f):
    for i in range(1, c+1): ws.cell(row=r, column=i).fill = f
def sh(ws, r, mc, t, font=FSC):
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=mc-1)
    ws.cell(row=r, column=2, value=t).font = font
    fb(ws, r, mc, S)
    for c in range(1, mc+1): ws.cell(row=r, column=c).border = BS
    ws.row_dimensions[r].height = 28
    return r + 1
def hd(ws, r, mc, h):
    for i, v in enumerate(h):
        c = ws.cell(row=r, column=i+1, value=v); c.font = FH; c.fill = H; c.alignment = AC
    ws.row_dimensions[r].height = 24
    return r + 1

MC = 9

# ============================================================
# EVERY PARALLEL OP WITH LLM-IN-LOOP AUTOMATION
# Format: (op_id, op_name, category, revenue, automation_level,
#          alpha_reason, infra_stack, setup_steps, algo_guide,
#          shadowban_avoid, llm_loop_desc, manual_first, auto_after)
# ============================================================

OPS = [
    # ── ECOMM ARB / DROPSHIP (LLM-automated) ──
    {
        'id': 'OP01', 'name': 'ECOMM TRENDING PRODUCT ARB',
        'cat': 'ECOMM', 'rev': '$50-500/mo → $2K+ at scale',
        'auto': 'FULL AUTO (after validation)',
        'alpha': 'LLM scrapes trending products from TikTok Shop, Etsy trending, Amazon Movers & Shakers, AliExpress hot products. Cross-references demand vs supply. Lists on Etsy/eBay/Mercari with AI-generated descriptions + SEO titles. Dropships from Temu/AliExpress/1688. Most sellers still manually browse — LLM scans 10,000 products/hour vs human scanning 50.',
        'infra': 'Playwright (scraping) + Claude API (analysis) + Etsy/eBay API (listing) + Temu/AliExpress (fulfillment) + Oracle VPS (24/7 cron)',
        'setup': '1. Create Etsy shop ($0 to list, $0.20/listing fee). Create eBay account ($0).\n2. Set up Playwright scraper on Oracle VPS: scrape TikTok Shop trending, Etsy trending searches, Amazon Movers.\n3. Claude API prompt: "Analyze this product. Estimate demand (search volume proxy from listings count). Find cheapest source on Temu/AliExpress. Calculate margin after fees (Etsy 6.5% + $0.20, eBay 12.9%). Only return products with >40% margin."\n4. Auto-generate listings: Claude writes title (SEO-optimized, 140 chars), 13 tags, 5 bullet description, pricing.\n5. Upload via Etsy API or manual CSV bulk upload.\n6. When order comes in: auto-purchase from Temu/AliExpress with customer shipping address.\n7. Cron job: re-scan trending every 6 hours. Delist products that drop off trending. List new ones.',
        'algo': 'Etsy Search Algorithm:\n- Recency boost: new listings get 24-48hr visibility boost. Relist every 2-3 days.\n- Relevancy: first 3 words of title are most weighted. Put exact search term first.\n- Quality score: conversion rate + favorites + shop reviews matter.\n- Tags: use all 13. Long-tail > generic ("minimalist gold necklace" > "jewelry").\n- Free shipping: Etsy penalizes listings without free shipping. Bake into price.\n\neBay Best Match:\n- Recent sales velocity matters most. Price competitively on first 5 sales.\n- Item specifics: fill EVERY field. eBay penalizes incomplete listings.\n- Promoted listings: 2-5% ad rate to boost visibility initially.',
        'shadow': 'Etsy: Never copy-paste descriptions from other listings (duplicate content flag). Never use same photos as another seller. Never list >50 items in 24hr (new shop). Ramp: 5/day week 1, 10/day week 2, 20/day week 3.\neBay: Below Standard = death. Ship within stated handling time. Never cancel orders. Keep defect rate <2%.',
        'llm_loop': 'Claude runs on 6hr cron: scrape → analyze → score → list → monitor sales → delist losers → find replacements. Fully autonomous after initial shop setup.',
        'manual': 'Week 1-2: Manually list 10 products. Validate demand. Test margins. Learn platform rules.',
        'auto_after': 'Week 3+: Playwright + Claude API + Etsy API. Full loop. Human only handles customer service escalations.',
    },
    {
        'id': 'OP02', 'name': 'PRINT ON DEMAND AUTO-DESIGN',
        'cat': 'ECOMM', 'rev': '$0-300/mo → $1K+ at scale',
        'auto': 'FULL AUTO',
        'alpha': 'LLM monitors trending memes, quotes, cultural moments. Generates designs via prompt → image API. Lists on Redbubble/TeePublic/Merch by Amazon. While humans manually design 5-10/day, LLM generates 200+ designs/day with trending keywords baked into titles.',
        'infra': 'Claude API (trend detection + copywriting) + Leonardo.ai/DALL-E (design gen) + Redbubble/TeePublic accounts + Oracle VPS',
        'setup': '1. Create Redbubble account ($0), TeePublic ($0), apply for Merch by Amazon (waitlist).\n2. Claude monitors: Reddit trending, Twitter/X trending topics, Google Trends, meme subreddits.\n3. Prompt: "Generate 5 t-shirt design concepts based on this trending topic. Each should be: text-based OR simple graphic, high contrast, printable on black and white shirts. Include SEO title and 15 tags."\n4. Generate designs via Leonardo.ai (150 free tokens/day) or Canva (free).\n5. Upload to Redbubble (auto-applies to 70+ products per design). TeePublic for additional reach.\n6. Cron: daily trend scan → generate → upload → track sales → double down on winners.',
        'algo': 'Redbubble Search:\n- Tags are EVERYTHING. Use all 15. Mix broad ("funny shirt") and specific ("software developer gift").\n- Title: exact search phrase first, then descriptors.\n- Seasonal: Christmas designs in October, Valentine in January. Lead time matters.\n- Niche down: "Dog Mom Labrador" beats "Dog Lover".\n\nMerch by Amazon:\n- Tier system: start at 10 designs, scale to 500+.\n- BSR (Best Seller Rank): track and iterate.\n- Keyword research: use Merch Informer or manual Amazon autocomplete.',
        'shadow': 'Redbubble: No trademark violations (Marvel, Disney, NFL = instant ban). No "thin content" (generic shapes with no creativity). Dont upload 100+ in one day.\nMerch by Amazon: Rejection = reset. Upload slowly. No copyrighted phrases. Research trademarks on USPTO before every design.',
        'llm_loop': 'Daily: trend scan → concept generation → design creation → listing → sales tracking → winner amplification. Weekly: retire 0-sale designs after 30 days. Monthly: seasonal prep.',
        'manual': 'Week 1: Manually create 20 designs. Learn what sells. Study top sellers in your niche.',
        'auto_after': 'Week 3+: Claude + Leonardo.ai + Redbubble API. Aim for 50 designs/week automated.',
    },
    {
        'id': 'OP03', 'name': 'AMAZON KDP BOOK FACTORY',
        'cat': 'ECOMM', 'rev': '$0-200/mo → $1K+',
        'auto': 'SEMI-AUTO',
        'alpha': 'LLM generates low-content books (journals, planners, coloring books, puzzle books) at 100x human speed. Keyword research + interior generation + cover design + listing optimization all LLM-driven. Most KDP sellers still manually create 1-2 books/week. LLM does 5-10/day.',
        'infra': 'Claude API (content gen + keyword research) + Canva (covers) + KDP account ($0) + Oracle VPS',
        'setup': '1. Create KDP account ($0, linked to Amazon account).\n2. Keyword research: use Amazon autocomplete + Publisher Rocket ($0 alt: manual autocomplete scraping). Claude prompt: "Find 20 low-competition journal niches on Amazon KDP. Target: <10K competing results, >1000 monthly searches."\n3. Interior generation: Claude generates journal prompts, planner layouts, puzzle content. Export as PDF.\n4. Cover design: Canva templates. Claude writes title + subtitle for SEO.\n5. Upload to KDP. Set price ($6.99-$14.99 for journals). Enable expanded distribution.\n6. Track BSR. Double down on sub-100K BSR titles.',
        'algo': 'Amazon A9 Algorithm:\n- Title: primary keyword first. "Gratitude Journal for Women" not "Beautiful Journal".\n- Backend keywords: 7 keyword fields × 50 chars. No commas, no repetition.\n- Categories: choose 2 most specific. Avoid ultra-competitive.\n- Price sweet spot: $6.99-$9.99 for journals, $12.99-$19.99 for planners.\n- Reviews: first 10 reviews = massive BSR boost. Use launch list.',
        'shadow': 'KDP: No AI-generated cover art (must disclose). No copyrighted content. No duplicate interiors across books. No keyword stuffing in description. Account termination for quality complaints — ensure PDFs render correctly.',
        'llm_loop': 'Weekly: niche research → interior gen → cover design → upload → BSR tracking. Monthly: retire underperformers, scale winners to series.',
        'manual': 'Week 1-2: Create 3 books manually. Learn KDP formatting rules. Get first sales.',
        'auto_after': 'Month 2+: Claude generates interiors + keywords. Human reviews covers + quality checks.',
    },
    # ── CONTENT AUTOMATION ──
    {
        'id': 'OP04', 'name': 'FACELESS YOUTUBE LONGFORM',
        'cat': 'CONTENT', 'rev': '$100-2K/mo → $10K+',
        'auto': 'FULL AUTO',
        'alpha': 'LLM writes scripts, ElevenLabs voices them, Kling/CapCut creates visuals, auto-uploads to YouTube. Faceless channels in finance/true crime/history do $5-50K/mo. Most still manually script + edit. LLM pipeline produces 1 video/day vs manual 1/week.',
        'infra': 'Claude API (scripts) + ElevenLabs (voice, 10K free chars/mo) + Kling (video, 66 free credits/day) + CapCut (editing) + YouTube API (upload) + Oracle VPS',
        'setup': '1. Create YouTube channel in Dolphin Anty profile (dedicated fingerprint).\n2. Niche selection: Finance explainers, AI news, true crime, history — all high RPM ($5-30 CPM).\n3. Script pipeline: Claude writes 1500-2000 word scripts with hook (first 30 sec), conflict, resolution structure.\n4. Voiceover: ElevenLabs API. Clone a custom voice or use stock. 10K chars = ~7 min audio.\n5. Visual assembly: Kling generates B-roll clips. Stock footage from Pexels (free). CapCut auto-edits with captions.\n6. Thumbnail: Canva template + Leonardo.ai for eye-catching imagery.\n7. Upload via YouTube API with Claude-generated title, description (2000 chars, keyword-rich), tags.\n8. Schedule: 3-5 videos/week for first 100 videos. Algorithm rewards consistency.',
        'algo': 'YouTube Algorithm Deep Dive:\n- CTR (Click-Through Rate): #1 factor. Must be >5%. Thumbnail + title = 80% of success.\n- AVD (Average View Duration): must be >50% of video length. Front-load value. Open loops.\n- First 48 hours: YouTube tests with small audience. High CTR + AVD = wider distribution.\n- Session time: videos that keep people on YouTube get boosted. End screens + playlists.\n- Upload consistency: same time, same days. YouTube learns your schedule.\n- Shorts vs Longform: separate algorithms. Shorts viewers rarely convert to longform. Run separately.\n- Community tab: post polls/updates. Engagement signals activity.\n- Revenue niches by CPM: Finance ($15-30), Tech ($12-25), Business ($10-20), Health ($8-15), Gaming ($2-5).',
        'shadow': 'YouTube:\n- Community Guidelines strikes: 3 strikes in 90 days = termination. Avoid controversial topics.\n- Reused content: YouTube detects recycled footage. Use >50% original or heavily edited content.\n- AI disclosure: YouTube requires AI content labeling since March 2024. Always check the "altered/synthetic" box.\n- Music copyright: use royalty-free only (Epidemic Sound, Artlist, or YouTube Audio Library).\n- Title/thumbnail mismatch: clickbait without delivery = low AVD = algorithm punishment.\n- Comment engagement: reply to first 10 comments within 1 hour. Signals active creator.',
        'llm_loop': 'Daily: Claude writes 1 script → ElevenLabs voices → Kling generates visuals → CapCut assembles → YouTube API uploads → Claude analyzes last 7 days performance → adjusts topics/hooks.',
        'manual': 'Week 1-2: Make 3 videos manually. Learn what CTR/AVD looks like. Study top channels in niche.',
        'auto_after': 'Week 3+: Full pipeline. Human does final quality check on thumbnail + first 30 seconds only.',
    },
    {
        'id': 'OP05', 'name': 'YOUTUBE SHORTS FARM (13 NICHES)',
        'cat': 'CONTENT', 'rev': '$200-5K/mo across all channels',
        'auto': 'FULL AUTO',
        'alpha': 'One LLM pipeline feeds 13 different content farm channels simultaneously. Each channel targets different niche (sleep, memes, finance, sports, etc). Cross-post to TikTok + IG Reels + FB Reels. FB Reels pays 40-200x more than YouTube Shorts. Nobody is running 13 channels from one automated pipeline.',
        'infra': '13 Dolphin profiles (need Incogniton for overflow) + Kling + CapCut + Claude API + Buffer + Tweetlio + Oracle VPS',
        'setup': '1. Create 13 YouTube channels, one per CF niche (CF001-CF013).\n2. Each channel gets unique Dolphin profile + proxy.\n3. Content pipeline per channel:\n   - CF001 (Relax): 8-10hr ambient loops. Generate with Suno (music) + stock nature footage.\n   - CF004 (Memes): scrape viral memes from Reddit/Twitter, compile into 60-sec montages.\n   - CF008 (Finance): Claude writes 60-sec market updates, ElevenLabs voices.\n   - CF010 (Sports): clip highlights from free sources, add commentary.\n4. Cross-post EVERY short to: TikTok, Instagram Reels, Facebook Reels, Pinterest Idea Pins.\n5. FB Reels arbitrage: $4.40 CPM vs YouTube Shorts $0.05 CPM = 88x difference.\n6. Scale winning channels, sunset losers after 30 days of no traction.',
        'algo': 'YouTube Shorts Algorithm:\n- Hook in first 1 second. Text overlay immediately.\n- Loop-worthy: ending connects to beginning. Increases watch time.\n- Trending audio: use trending sounds (YouTube audio library).\n- Hashtags: #Shorts + 2-3 niche tags. Dont over-tag.\n- Posting time: 6-9 AM, 12-2 PM, 7-10 PM local time.\n- Frequency: 1-3/day per channel. More = more lottery tickets.\n\nFB Reels:\n- Overlay text on EVERY frame. FB algorithm prefers text-heavy reels.\n- 30-90 seconds optimal. Under 15 = too short for monetization.\n- Original audio pays more than licensed music.\n- Engagement bait: "Wait for the ending" / "Comment your answer" works on FB.',
        'shadow': 'YouTube Shorts: Reused content = demonetization. Must add commentary, editing, or transformation.\nTikTok: Watermarked content from other platforms = suppressed. Always upload native.\nIG Reels: Remove TikTok watermark before cross-posting (use SnapTik or re-export from CapCut).\nFB Reels: Original content bonus program. No watermarks from other platforms.',
        'llm_loop': 'Hourly: scrape trending content → Claude selects + writes hooks → CapCut auto-edit → distribute across 4 platforms × 13 channels. Daily: performance review → adjust.',
        'manual': 'Week 1: Start 3 channels. Post 2/day each. Learn what works per platform.',
        'auto_after': 'Week 3+: Full pipeline. Scale to 13 channels as Dolphin profiles allow.',
    },
    # ── AFFILIATE & SEO ──
    {
        'id': 'OP06', 'name': 'AFFILIATE SEO CONTENT SITES',
        'cat': 'AFFILIATE', 'rev': '$500-5K/mo per site',
        'auto': 'SEMI-AUTO',
        'alpha': 'LLM generates 10-20 SEO articles/day targeting buyer-intent keywords. "Best X for Y" articles with affiliate links. Most affiliate sites publish 2-4 articles/week manually. LLM does 10-20/day with proper E-E-A-T signals, internal linking, schema markup.',
        'infra': 'Claude API (content) + Vercel/Cloudflare (hosting) + WordPress or Next.js + Ahrefs free alt (Ubersuggest free) + Oracle VPS',
        'setup': '1. Domain: buy niche .com on Porkbun ($9.73). Example: bestaitools.com, topfaithapps.com.\n2. Hosting: Cloudflare Pages (free, unlimited) or Vercel.\n3. CMS: Next.js static site (fastest) or WordPress on Oracle VPS.\n4. Keyword research: Claude + free tools. Prompt: "Find 50 buyer-intent keywords for [niche] with: commercial intent, low DR competitors in top 10, question-format preferred."\n5. Content pipeline: Claude writes 2000-3000 word articles with:\n   - H2/H3 structure matching search intent\n   - Comparison tables\n   - Pros/cons for each product\n   - FAQ schema markup\n   - Internal links to related articles\n6. Affiliate programs: Amazon Associates (1-10%), ShareASale, Impact, CJ Affiliate.\n7. Publish 3-5 articles/day. Interlink everything. Build topical authority.\n8. Track rankings: Google Search Console (free) + manual SERP checks.',
        'algo': 'Google SEO Algorithm (2026):\n- E-E-A-T: Experience, Expertise, Authoritativeness, Trust. Add author bios, cite sources, show real experience.\n- Helpful Content Update: content must satisfy search intent completely. No thin pages.\n- Core Web Vitals: LCP <2.5s, FID <100ms, CLS <0.1. Static sites on Cloudflare = auto pass.\n- Topical authority: publish 30+ articles on same topic before expecting rankings. Depth > breadth.\n- Internal linking: every article links to 3-5 related articles. Builds PageRank internally.\n- Title tag: exact keyword + modifier ("2026", "Best", "Review"). Keep under 60 chars.\n- Featured snippets: answer the question in first paragraph, then elaborate. Use tables + lists.',
        'shadow': 'Google: AI content is fine IF it adds genuine value. Thin/unhelpful AI content gets demoted.\nAmazon Associates: No cloaking affiliate links. No incentivizing clicks. No using in email.\nGeneral: Disclose affiliate relationships (FTC requirement). Add "As an Amazon Associate I earn from qualifying purchases" on every page.',
        'llm_loop': 'Daily: Claude researches 5 keywords → writes 5 articles → auto-publishes → tracks rankings weekly → updates underperforming articles with more depth → builds internal links.',
        'manual': 'Month 1: Write 20 articles manually. Learn what ranks. Study competitors.',
        'auto_after': 'Month 2+: Claude writes + publishes. Human reviews for quality + adds personal experience sections.',
    },
    # ── NEWSLETTER + MEDIA ──
    {
        'id': 'OP07', 'name': 'NEWSLETTER EMPIRE (BEEHIIV + SUBSTACK)',
        'cat': 'NEWSLETTER', 'rev': '$500-68K/mo combined',
        'auto': 'SEMI-AUTO',
        'alpha': 'LLM writes 3+ newsletters per week across multiple publications. Curates from signal sources, adds analysis, generates subject lines A/B tested. Most newsletter writers spend 4-8 hrs per issue. LLM draft in 15 min, human polish in 30 min.',
        'infra': 'Claude API (writing) + Beehiiv ×3 (free) + Substack (free) + Buffer (social promo) + TrulyInbox (warmup)',
        'setup': '1. Beehiiv: 3 publications (2,500 subs each free). STACKPRINT (tech), MORNING MANNA (faith), SAUCE LETTER (money).\n2. Substack: unlimited subs, 10% of paid revenue. Use for additional niches.\n3. Claude pipeline: scrape 145+ signal sources daily → extract top 5-7 insights → write newsletter draft → generate 5 subject line variants → schedule.\n4. Growth: cross-promote between newsletters. Share signup links in all social content. Beehiiv referral program (readers refer readers for rewards).\n5. Monetization ladder:\n   - 0-1K subs: grow only, no monetization\n   - 1K-2.5K: Beehiiv Ad Network ($200-500/mo)\n   - 2.5K-5K: Sponsorships ($200-500/issue)\n   - 5K-10K: Premium tier ($5-10/mo)\n   - 10K+: Sponsorships at $20-200 CPM\n6. Welcome sequence: 5 emails over 7 days. Each delivers value + soft-sells paid product.',
        'algo': 'Email Deliverability:\n- SPF, DKIM, DMARC: configure ALL three. Non-negotiable.\n- Warmup: TrulyInbox for 14-21 days before sending to cold list.\n- List hygiene: remove bounces immediately. Remove inactive (no opens in 90 days).\n- Subject lines: 6-10 words. Personalization tokens. Curiosity gaps. Numbers.\n- Send time: Tue/Wed/Thu at 7-9 AM local time. Avoid Mon (inbox overload) and Fri (checked out).\n- Engagement: high open rate (>40%) + click rate (>3%) = good sender reputation.\n- Avoid spam triggers: no ALL CAPS, no excessive exclamation marks, no "free money" phrases.',
        'shadow': 'Email: CAN-SPAM compliance required. Always include unsubscribe link (Beehiiv handles automatically). Never buy email lists. Never add people without consent. Double opt-in preferred.\nBeehiiv: respect their ToS on AI content — you can use AI to draft but should review/edit before sending.',
        'llm_loop': 'Daily: Claude scrapes signal sources → generates newsletter draft → human reviews (15 min) → schedules. Weekly: analyze open rates, click rates → adjust topics/format.',
        'manual': 'Week 1-2: Write first 4 issues manually. Learn your voice. Get first 100 subs from warm outreach.',
        'auto_after': 'Week 3+: Claude drafts, human polishes. Growth automated via cross-promo + referral programs.',
    },
    # ── COLD OUTBOUND ──
    {
        'id': 'OP08', 'name': 'COLD EMAIL B2B OUTBOUND',
        'cat': 'OUTBOUND', 'rev': '$1K-10K/mo',
        'auto': 'SEMI-AUTO',
        'alpha': 'LLM personalizes every cold email at scale. Scrapes prospect website/LinkedIn → generates hyper-personalized first line → writes value prop specific to their pain points. Most cold emailers use generic templates. LLM-personalized emails get 3-5x reply rates.',
        'infra': 'Claude API (personalization) + TrulyInbox (warmup) + Gmail/GWorkspace (sending) + Apollo.io or LinkedIn (lead data) + Playwright (scraping) + Oracle VPS',
        'setup': '1. Domain setup: buy 3 cold email domains ($30/yr total). NEVER use primary brand domain.\n2. Email setup: 2 inboxes per domain via Gmail ($0 free tier) or Google Workspace ($6/user/mo).\n3. Warmup: TrulyInbox all 6 inboxes for 21 days (free forever).\n4. Lead sourcing: Apollo.io free tier (limited) or Playwright scrape LinkedIn/Google Maps.\n5. Personalization pipeline: Claude reads prospect website → extracts 3 key facts → writes personalized opener + pain point + CTA.\n6. Sequence: 4-email sequence over 14 days.\n   - Email 1: Personalized observation + question\n   - Email 2: Case study / social proof (day 3)\n   - Email 3: Different angle / objection handling (day 7)\n   - Email 4: Breakup email (day 14)\n7. Send limits: 30-50/day per inbox. 6 inboxes = 180-300 emails/day.\n8. Track replies in HubSpot (free CRM).',
        'algo': 'Cold Email Deliverability:\n- Warmup 21 days minimum before sending. Non-negotiable.\n- Send limits: max 50/day per inbox. 30 is safer for new domains.\n- Reply rate target: 3-8% = good. <1% = something broken.\n- Bounce rate: keep <2%. Clean list before sending.\n- Domain age: older domains = better deliverability. Buy aged domains if possible.\n- Send time: Tue-Thu, 8-10 AM recipient local time.\n- Plain text > HTML. No images, no tracking pixels initially.\n- One link max. No attachments on first email.\n- Subject line: 3-5 words, lowercase, no special chars. "quick question" outperforms branded subjects.',
        'shadow': 'CAN-SPAM: must include physical address + unsubscribe. Use PostalMate for virtual address ($10/mo).\nGDPR: if emailing EU, need legitimate interest basis. B2B is generally ok, B2C is not.\nGoogle: sending limits enforced. Exceeding = temporary lock → permanent ban.\nBlacklists: check MXToolbox weekly. If blacklisted, switch domain immediately.',
        'llm_loop': 'Daily: scrape 50 prospect websites → Claude personalizes 50 emails → send via Gmail → track replies → Claude drafts follow-ups for non-responders.',
        'manual': 'Week 1-3: Warmup only. Week 4: Send first 50 manually. Test messaging. Learn objections.',
        'auto_after': 'Month 2+: Playwright + Claude + Gmail API. Human only handles live replies.',
    },
    # ── APP FACTORY ──
    {
        'id': 'OP09', 'name': 'APP FACTORY (2-3 APPS/MONTH)',
        'cat': 'APPS', 'rev': '$1K-50K/mo portfolio',
        'auto': 'SEMI-AUTO',
        'alpha': 'Claude Code builds entire apps in 48-hour sprints. Most indie devs ship 1 app every 2-3 months. This ships 2-3/month. Portfolio of 30 apps by month 12. Each app targets different niche from N001-N038. RevenueCat handles IAP/subs across all apps from one dashboard.',
        'infra': 'Claude Code (development) + Xcode/Android Studio (build) + RevenueCat (monetization) + TestFlight/Play Console (distribution) + Vercel (landing pages)',
        'setup': '1. Apple Developer ($99/yr) + Google Play Console ($25 one-time). REQUIRED.\n2. RevenueCat account (free to $2.5K MRR).\n3. App pipeline: each niche (N001-N038) gets a tailored app.\n   - N002 (Faith): PrayerLock (daily prayer streak tracker)\n   - N003 (Fitness): StepUnlock (gamified walking)\n   - N006 (ADHD): FocusBurn (dopamine-friendly pomodoro)\n   - N011 (Sleep): SleepCodex (sleep tracking + ambient sounds)\n4. Dev sprint: Claude Code builds MVP in 24-48 hrs. Human reviews + submits.\n5. ASO (App Store Optimization): Claude writes title, subtitle, keywords, description.\n6. Launch sequence: social accounts promote → newsletter features → first 50 downloads → ratings.\n7. Revenue: freemium model. Free tier hooks. $4.99/mo or $29.99/yr premium.\n8. Portfolio effect: 30 apps × $100/mo average = $3K/mo. 30 apps × $500/mo = $15K/mo.',
        'algo': 'App Store Algorithm (ASO):\n- Title: primary keyword. 30 chars max (iOS). Use it.\n- Subtitle (iOS): secondary keyword. 30 chars.\n- Keyword field (iOS): 100 chars. No spaces, comma-separated. No duplicates with title.\n- Description: first 3 lines visible without "read more". Put value prop + keywords.\n- Downloads velocity: first 48 hours matter most. Coordinate launch with social + newsletter.\n- Ratings: 4.5+ stars = significant boost. Prompt for review after positive action (completed streak, hit goal).\n- Updates: publish update every 2-4 weeks. Algorithm rewards active apps.\n- Category: choose least competitive. Better to rank #50 in small category than #5000 in big one.\n\nGoogle Play:\n- Title: 50 chars. Keyword-rich.\n- Short description: 80 chars. Keyword-rich.\n- Long description: 4000 chars. Keywords naturally woven in.\n- Install velocity + uninstall rate = key ranking factors.',
        'shadow': 'Apple: rejection reasons → fix and resubmit. Common: missing privacy policy, undisclosed data collection, broken links.\nGoogle: faster approval but harsher on policy violations. No misleading claims.\nBoth: AI-generated content must add genuine value. No clone apps. No spam apps (same app, different skin).\nRevenueCat: handles Apple/Google compliance for subscriptions automatically.',
        'llm_loop': 'Bi-weekly sprint: Claude Code builds app → human tests → submit to stores → Claude writes ASO listing → social channels promote → track downloads + revenue.',
        'manual': 'First 2 apps: build and submit manually. Learn the review process. Get first revenue.',
        'auto_after': 'App 3+: Claude Code handles 90% of development. Human handles submission + quality testing.',
    },
    # ── AI INFLUENCER ──
    {
        'id': 'OP10', 'name': 'AI INFLUENCER PERSONA NETWORK',
        'cat': 'AI PERSONA', 'rev': '$500-20K/mo per persona',
        'auto': 'SEMI-AUTO',
        'alpha': 'LLM generates consistent persona content (posts, replies, stories) across platforms. Leonardo.ai generates consistent face shots. ElevenLabs creates consistent voice. While humans can run 1-2 influencer accounts, LLM runs 8 simultaneous personas with consistent voice/aesthetic.',
        'infra': 'Claude API (content) + Leonardo.ai (face gen) + ElevenLabs (voice) + Kling (video) + Dolphin Anty (8 profiles needed) + Decodo proxies',
        'setup': '1. Persona creation: Claude creates detailed persona doc (backstory, voice, aesthetic, posting style).\n2. Face consistency: Leonardo.ai with fixed seed + reference images. Generate 50 photos in different outfits/settings.\n3. Voice: ElevenLabs voice clone from generated audio sample. Consistent across all content.\n4. Platforms per persona: X + TikTok + Instagram minimum. Add YouTube for high-value niches.\n5. Content pipeline per persona:\n   - 3-5 tweets/day (Tweetlio)\n   - 1 TikTok/day (manual or Buffer)\n   - 1 IG post + 3-5 stories/day\n6. Monetization per persona type:\n   - AI001 (Expert): courses, consulting, newsletter → $500-10K/mo\n   - AI005 (Fitness): workout programs, supplement affiliate → $500-10K/mo\n   - AI006 (Lifestyle): brand deals, affiliate → $300-5K/mo\n7. FTC compliance: ALWAYS disclose AI-generated persona in bio.',
        'algo': 'IG Algorithm for Personas:\n- Reels: hook in 0.5 sec. Text overlay. Trending audio. 7-15 sec optimal.\n- Stories: 5-7 per day. Polls + questions = 2x engagement. Swipe-up after 10K followers.\n- Posts: carousel > single image. 10 slides for maximum engagement.\n- Hashtags: 3-5 niche-specific. No banned hashtags.\n- Engagement window: first 30 min after posting. Reply to every comment.\n\nTikTok for Personas:\n- First 3 seconds = hook or die. Text overlay + movement.\n- Trending sounds: boost distribution 3-5x.\n- Post time: 7-9 AM, 12-3 PM, 7-11 PM.\n- Duet/Stitch trending videos in your niche.\n- Bio link: Linktree or direct to highest-converting page.',
        'shadow': 'AI Personas CRITICAL:\n- ALWAYS disclose "AI-generated persona" in bio. FTC requires this.\n- Never claim to be a real person in DMs or comments.\n- Instagram: no buying followers/likes. Shadowban from fake engagement.\n- TikTok: no spam following. No duplicate content across accounts.\n- X: no bulk following/unfollowing same day (1000 follow limit).\n- Consistency: same face, same voice, same aesthetic. Inconsistency = detected = reported.\n- Engagement: respond to comments in persona voice. Claude API with persona system prompt.',
        'llm_loop': 'Daily: Claude (with persona prompt) generates content → Leonardo.ai adds consistent visuals → schedule across platforms. Replies automated via Claude with persona context.',
        'manual': 'Week 1-2: Build persona docs. Generate first 50 photos. Post manually to test audience response.',
        'auto_after': 'Week 3+: Content pipeline + scheduling automated. Human monitors DMs and brand deal inquiries.',
    },
    # ── TIKTOK SHOP ──
    {
        'id': 'OP11', 'name': 'TIKTOK SHOP AFFILIATE + DROPSHIP',
        'cat': 'ECOMM', 'rev': '$500-20K/mo',
        'auto': 'SEMI-AUTO',
        'alpha': 'LLM identifies trending TikTok Shop products → generates UGC-style review scripts → Kling/HeyGen creates AI UGC videos → posts to TikTok. 10-30% affiliate commission on $23.4B US market. Small creators get 3.2x higher CTR than brands. LLM can test 50 products/week vs human testing 5.',
        'infra': 'Claude API (product research + scripts) + Kling (AI video) + TikTok Shop affiliate account ($0) + Dolphin Anty + Decodo proxy',
        'setup': '1. TikTok Shop affiliate: apply at affiliate.tiktok.com. $0, instant approval.\n2. Product research: Claude analyzes TikTok Shop trending → filters for: >1000 units sold, >10% commission, <$30 price point, viral video potential.\n3. Script: Claude writes 30-60 sec UGC-style review. Problem → solution → demo → CTA.\n4. Video: Kling generates AI UGC or record yourself. CapCut edits.\n5. Post with TikTok Shop product tag. Earn commission on every sale.\n6. Scale winners: product getting sales → make 5 more videos with different hooks.\n7. LIVE selling: highest conversion (3-5x video). Schedule 1-2 LIVEs/week for top products.\n8. Dropship angle: find products on Temu/1688 for cheaper → list on TikTok Shop as seller (not just affiliate).',
        'algo': 'TikTok Shop Algorithm:\n- Product videos with shop tags get BOOSTED distribution vs regular videos.\n- LIVE sessions: TikTok pushes LIVE aggressively. 30-60 min sessions.\n- Commission tiers: higher commission = TikTok promotes more (they earn more too).\n- Sample requests: request free samples from sellers for authentic reviews.\n- Trending products tab: TikTok Shop has internal trending. Check daily.\n- Hashtags: #TikTokShop #TikTokMadeMeBuyIt #Unboxing + niche tags.\n- Sound: trending sounds + voiceover hybrid = best performance.',
        'shadow': 'TikTok Shop: No misleading product claims. No health/medical claims without evidence. No showing competitor products negatively. Fake reviews = permanent ban.\nContent: must be original. No stealing other creators review videos. AI UGC must be disclosed if required by platform.',
        'llm_loop': 'Daily: Claude scrapes trending products → scores by commission×demand×viral potential → writes 5 scripts → Kling creates videos → post → track conversion → scale winners.',
        'manual': 'Week 1: Manually review 5 products. Post videos. Learn what converts.',
        'auto_after': 'Week 3+: Product research + script writing automated. Human handles LIVE sessions.',
    },
    # ── MUSIC / SUNO ──
    {
        'id': 'OP12', 'name': 'AI MUSIC FACTORY (SUNO + DISTROKID)',
        'cat': 'MUSIC', 'rev': '$0-500/mo → $2K+ with viral hit',
        'auto': 'FULL AUTO',
        'alpha': 'Suno generates full songs from text prompts. DistroKid distributes to Spotify/Apple/YouTube Music ($22.99/yr). Most musicians spend weeks on one song. LLM + Suno produces 10 songs/day across genres. Lofi/ambient/study music channels generate $500-5K/mo on Spotify alone. Nobody is running 20 Spotify artist profiles from one pipeline.',
        'infra': 'Suno (50 free credits/day) + DistroKid ($22.99/yr) + Claude API (lyrics + prompts) + Canva (album art) + Spotify for Artists ($0)',
        'setup': '1. Create Suno account (50 free credits/day = ~10 songs).\n2. DistroKid account ($22.99/yr, unlimited uploads to all platforms).\n3. Genre strategy: lofi hip hop, ambient, study music, sleep sounds, worship music — all high playlist placement potential.\n4. Claude writes Suno prompts: "[genre], [mood], [tempo BPM], [instruments]. Lyrics about [theme]."\n5. Generate 5-10 tracks/day. Select best 2-3.\n6. Cover art: Canva or Leonardo.ai. Consistent artist branding.\n7. Upload to DistroKid → distributes to Spotify, Apple Music, YouTube Music, Amazon, Tidal, etc.\n8. Playlist pitching: Spotify for Artists has free playlist submission. Submit every release 7 days before launch.\n9. Revenue: $0.003-0.005 per stream on Spotify. 1M streams = $3-5K. Ambient/lofi playlists get millions of passive streams.',
        'algo': 'Spotify Algorithm:\n- Release Radar: new songs appear in followers Release Radar. Build follower base.\n- Discover Weekly: algorithmic. Saves + full listens = strongest signals.\n- Playlist placement: submit via Spotify for Artists 7 days before release. Pick specific playlists.\n- Consistent releases: weekly > monthly. Algorithm rewards active artists.\n- Song length: >30 seconds counts as a stream. But 2-4 min optimal for playlists.\n- Pre-save campaigns: Distrokid generates pre-save links. Share on social.',
        'shadow': 'Spotify: No artificial streaming (botting = instant ban + revenue clawback). No uploading same song on multiple artist profiles. No excessive release volume (>1 song/day per artist raises flags).\nDistroKid: keeps 0% royalties but charges $22.99/yr. Cancel = music stays up.\nAI Music: Currently legal to distribute. Spotify allows AI music but policies evolving. Dont claim to be a human artist.',
        'llm_loop': 'Daily: Claude generates 10 Suno prompts → Suno creates tracks → human selects top 3 → DistroKid uploads → social channels promote → Spotify for Artists submits to playlists.',
        'manual': 'Week 1: Generate 20 songs. Learn Suno prompt engineering. Upload first EP.',
        'auto_after': 'Week 3+: Claude prompts → Suno generates → auto-upload pipeline. Human curates quality.',
    },
    # ── NOTION TEMPLATES ──
    {
        'id': 'OP13', 'name': 'NOTION TEMPLATE STORE',
        'cat': 'DIGITAL PRODUCTS', 'rev': '$500-10K/mo',
        'auto': 'SEMI-AUTO',
        'alpha': 'LLM creates Notion templates + duplicates at 50x human speed. Templates sell for $5-97 on Gumroad/Etsy. Top Notion creators do $10-50K/mo. LLM generates template structure + marketing copy + social content promoting it. Most creators make 1-2 templates/month. LLM makes 5-10/week.',
        'infra': 'Claude API (template design + marketing) + Notion (free) + Gumroad ($0 until sale) + Canva (mockups) + Twitter/TikTok (marketing)',
        'setup': '1. Create Notion account (free).\n2. Research: Claude analyzes top-selling Notion templates on Gumroad/Etsy. Identify gaps.\n3. Template creation: Claude designs template structure (databases, views, formulas, automations). Human builds in Notion.\n4. Pricing: $9-27 for basic, $27-67 for premium, $67-97 for bundles.\n5. Distribution: Gumroad (free until sale, 10% fee) + Etsy ($0.20/listing).\n6. Marketing: create "template reveal" TikToks/Reels showing the template in action. These go viral.\n7. SEO: Gumroad product pages rank on Google. Optimize titles + descriptions.\n8. Bundle strategy: sell 5 templates individually, then bundle at 60% discount.',
        'algo': 'Gumroad Discovery:\n- Product title: exact search phrase. "ADHD Daily Planner Notion Template".\n- Tags: use all available. Gumroad search is tag-based.\n- Cover image: show the template in use. Not just a logo.\n- Price psychology: $9 = impulse buy. $27 = considered. $67+ = need strong value prop.\n- Reviews: first 10 reviews = massive visibility boost. Give free copies for reviews.\n\nEtsy for Digital:\n- Same SEO rules as physical products. All 13 tags. Long-tail titles.\n- Digital download = no shipping. Higher margin.\n- "Instant download" in title = buyer confidence.',
        'shadow': 'Notion: Templates must be original. Dont copy other creators layouts exactly.\nGumroad: No refund manipulation. Honor refund requests.\nEtsy: Digital products must deliver instantly. Test download flow.',
        'llm_loop': 'Weekly: Claude identifies trending template niches → designs structure → human builds in Notion → Claude writes marketing copy → Canva creates mockups → list on Gumroad + Etsy → social promotion.',
        'manual': 'Week 1-2: Build 3 templates manually. Learn what sells. Study top creators.',
        'auto_after': 'Week 3+: Claude designs + writes copy. Human builds in Notion (cant fully automate Notion UI yet).',
    },
    # ── DOMAIN FLIPPING ──
    {
        'id': 'OP14', 'name': 'DOMAIN FLIPPING (LLM-SCOUTED)',
        'cat': 'DOMAINS', 'rev': '$500-20K/mo',
        'auto': 'SEMI-AUTO',
        'alpha': 'LLM scans expiring domain lists (10,000+ daily) and scores by: brandability, keyword value, extension, length, existing backlinks, similar domain sale history. Humans manually check maybe 50 domains/day. LLM scores 10,000 in minutes. Buy at $10, sell at $100-10,000.',
        'infra': 'Claude API (scoring) + Porkbun/Namecheap (buying) + Dan.com/Afternic/Sedo (selling) + ExpiredDomains.net (sourcing)',
        'setup': '1. Source: ExpiredDomains.net (free). Filter: .com only, <15 chars, no hyphens, no numbers.\n2. Claude scoring prompt: "Score this domain 1-100 on: brandability (does it sound like a company?), keyword value (is the keyword searched?), memorability (can you remember it after hearing once?), extension (.com = 100, .io = 60, .net = 40). Also check if similar domains sold for >$500 on NameBio.com."\n3. Buy: domains scoring >70 at auction price ($10-50 typically).\n4. List on: Dan.com (free listing, 9% commission), Afternic (free, 20% commission), Sedo.\n5. Add "buy this domain" lander page using Dan.com or simple Cloudflare page.\n6. Outbound: if domain matches a startup/company name, email them offering the domain.\n7. Portfolio management: hold max 50 domains. Renewal cost = $10/yr each. Drop non-sellers after 1 year.',
        'algo': 'Domain Valuation Factors:\n- Length: 1-word .com = $10K-1M+. 2-word = $500-50K. 3+ word = usually <$500.\n- Extension: .com is 10x .io, 20x .net, 50x .xyz.\n- Keyword: "AI" domains premium 2026. "Crypto" declining. Check Google Trends.\n- Brandability: made-up words that sound good (Spotify, Canva) = premium.\n- Backlinks: domain with existing backlinks = instant SEO value. Check Ahrefs.\n- Comparable sales: NameBio.com has 500K+ historical sales. Always check comps.',
        'shadow': 'Domain squatting on trademarks = UDRP (domain seizure). NEVER buy domains matching existing company trademarks. Check USPTO trademark database before buying.',
        'llm_loop': 'Daily: scrape 1000 expiring domains → Claude scores all → human reviews top 20 → buy top 5 → list on marketplaces. Monthly: review portfolio, drop non-sellers.',
        'manual': 'Week 1: Study NameBio.com sales. Learn what sells. Buy 5 test domains.',
        'auto_after': 'Week 3+: scraping + scoring automated. Human does final buy/no-buy decision.',
    },
    # ── ALGO TRADING SIGNALS ──
    {
        'id': 'OP15', 'name': 'ALGO TRADING SIGNALS + EDUCATION',
        'cat': 'FINANCE', 'rev': '$1K-50K/mo',
        'auto': 'SEMI-AUTO',
        'alpha': 'LLM analyzes on-chain data, options flow, earnings transcripts, Fed speeches in real-time. Generates signal reports + educational content. Most signal services are manual analysis. LLM processes 100x more data points and publishes faster.',
        'infra': 'Claude API (analysis) + GlassNode/Arkham data (free tiers) + Newsletter (Beehiiv/Substack) + X account (@bidaskspread)',
        'setup': '1. Data sources: GlassNode free tier (on-chain), Yahoo Finance API (free), Federal Reserve FRED (free).\n2. Claude analysis: "Analyze todays options flow data. Identify unusual volume. Cross-reference with insider trading filings. Generate a signal report with: ticker, direction (bullish/bearish), confidence (1-10), catalyst, risk."\n3. Distribution: X thread daily + newsletter weekly digest.\n4. Monetization: free signal report → paid premium ($29-99/mo) → course ($297-997).\n5. Compliance: ALWAYS include disclaimer. Not financial advice. Past performance ≠ future results.\n6. Paper trade first: 30 days of signals tracked publicly before charging.',
        'algo': 'FinTwit Algorithm:\n- X algo boosts long-form threads with data + charts.\n- Post market analysis 15 min after market close (4:15 PM ET).\n- Morning pre-market posts (6-7 AM ET) get high engagement.\n- Use $TICKER cashtags for discovery.\n- Quote-tweet breaking financial news with your analysis.\n- Consistency: same time, same format, every trading day.',
        'shadow': 'SEC: Never guarantee returns. Never provide personalized investment advice (requires registration). Always disclaim "Not financial advice. For educational purposes only."\nPlatforms: No pump-and-dump coordination. No encouraging market manipulation.',
        'llm_loop': 'Daily: Claude scrapes market data → generates morning brief + after-hours analysis → posts to X → weekly newsletter compilation. Human reviews compliance.',
        'manual': 'Month 1: Paper trade. Post analysis daily. Build track record.',
        'auto_after': 'Month 2+: Data scraping + analysis + posting automated. Human reviews for compliance + adds commentary.',
    },
    # ── COLD OUTREACH WEB REDESIGN ──
    {
        'id': 'OP16', 'name': 'WEB REDESIGN COLD OUTREACH',
        'cat': 'AGENCY', 'rev': '$3K-50K/mo',
        'auto': 'SEMI-AUTO',
        'alpha': 'LLM scrapes local businesses with bad websites → auto-generates redesign mockup in 5 min → sends cold email with before/after. 25% close rate proven. Most web design agencies manually prospect. LLM finds 100 prospects/hour and creates personalized mockups.',
        'infra': 'Playwright (scraping) + Claude API (analysis + email) + Figma/Canva (mockups) + Cold email stack (TrulyInbox + Gmail)',
        'setup': '1. Target: local businesses (dentists, plumbers, restaurants, lawyers) with outdated websites.\n2. Scraping: Playwright hits Google Maps → extracts business name + website + email.\n3. Analysis: Claude visits website → scores on: mobile responsiveness, load speed, design quality, SEO basics.\n4. Mockup: for businesses scoring <40/100, Claude generates a redesign brief → Canva/Figma creates quick mockup.\n5. Cold email: "Hi [Name], I noticed [specific issue with their site]. I put together a quick redesign concept — would you like to see it? [Screenshot attached]"\n6. Pricing: $500-2000 for simple redesign. $2000-5000 for full rebuild. $200-500/mo maintenance.\n7. Fulfillment: Claude Code builds the site in 24-48 hrs. Vercel/Cloudflare hosts. You manage.',
        'algo': 'Google Maps Scraping Best Practices:\n- Search "[industry] in [city]" → scrape first 60 results.\n- Filter: has website + has email. Skip chains/franchises.\n- Rotate user agents. Use delays between requests (2-5 sec).\n- Target cities with 50K-500K population (enough businesses, less competition from agencies).',
        'shadow': 'Google Maps: dont scrape too aggressively (IP ban). Use residential proxies + delays.\nCold email: follow all CAN-SPAM rules. Include opt-out. Physical address.\nWeb design: never use copyrighted images/content from competitors.',
        'llm_loop': 'Daily: Playwright scrapes 100 businesses → Claude scores websites → generates redesign mockups for bottom 20 → sends personalized cold emails → tracks replies → Claude drafts proposals for interested prospects.',
        'manual': 'Week 1-2: Manually find 20 businesses. Send cold emails. Close first client.',
        'auto_after': 'Week 3+: Scraping + scoring + email automated. Human handles sales calls + fulfillment.',
    },
]

# ============================================================
# SHEET 1: ALL OPS MASTER (overview)
# ============================================================
ws = wb.active
ws.title = 'ALL OPS'
ws.sheet_properties.tabColor = '00D4FF'
sw(ws, [3, 8, 30, 14, 20, 18, 48, 3])
MC = 8

for r in range(1, 250):
    fb(ws, r, MC, D)

ws.merge_cells('B2:G2')
ws.cell(row=2, column=2, value='PRINTMAXX PARALLEL OPS — LLM-IN-THE-LOOP AUTOMATION').font = FT
ws.row_dimensions[2].height = 35
ws.merge_cells('B3:G3')
ws.cell(row=3, column=2, value='16 ops running 24/7. Alpha exists because most people arent automating with LLMs yet.').font = FS

r = 5
r = hd(ws, r, MC, ['', 'OP ID', 'OPERATION', 'CATEGORY', 'REVENUE', 'AUTOMATION LVL', 'LLM ALPHA (WHY THIS PRINTS)'])

for idx, op in enumerate(OPS):
    bg = ALPHA if 'FULL AUTO' in op['auto'] else (G if 'SEMI' in op['auto'] else R1)
    fb(ws, r, MC, bg)
    ws.cell(row=r, column=2, value=op['id']).font = FC
    ws.cell(row=r, column=2).alignment = AC
    ws.cell(row=r, column=3, value=op['name']).font = FHL
    ws.cell(row=r, column=4, value=op['cat']).font = FGO
    ws.cell(row=r, column=4).alignment = AC
    ws.cell(row=r, column=5, value=op['rev']).font = FG
    ws.cell(row=r, column=5).alignment = AC
    auto_font = FALPHA if 'FULL' in op['auto'] else FG
    ws.cell(row=r, column=6, value=op['auto']).font = auto_font
    ws.cell(row=r, column=6).alignment = AC
    ws.cell(row=r, column=7, value=op['alpha'][:200] + '...').font = FB
    ws.cell(row=r, column=7).alignment = AL
    for c in range(1, MC+1): ws.cell(row=r, column=c).border = BD
    ws.row_dimensions[r].height = 65
    r += 1


# ============================================================
# SHEET 2: DEEP PLAYBOOK (each op with full instructions)
# ============================================================
ws2 = wb.create_sheet('DEEP PLAYBOOK')
ws2.sheet_properties.tabColor = 'E040FB'
sw(ws2, [3, 18, 80, 3])
MC2 = 4

for r in range(1, 2000):
    fb(ws2, r, MC2, D)

ws2.merge_cells('B2:C2')
ws2.cell(row=2, column=2, value='DEEP OPS PLAYBOOK — FULL INSTRUCTIONS').font = FT
ws2.row_dimensions[2].height = 35
ws2.merge_cells('B3:C3')
ws2.cell(row=3, column=2, value='Setup → Algorithm Guide → Shadowban Avoidance → LLM Loop → Manual First → Auto After').font = FS

r = 5
for op in OPS:
    # Op header
    ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws2.cell(row=r, column=2, value=f"{op['id']}: {op['name']} — {op['rev']}").font = Font(name='Arial', size=14, bold=True, color='00D4FF')
    fb(ws2, r, MC2, S)
    for c in range(1, MC2+1): ws2.cell(row=r, column=c).border = BS
    ws2.row_dimensions[r].height = 32
    r += 1

    sections = [
        ('LLM ALPHA', op['alpha'], FALPHA),
        ('INFRA STACK', op['infra'], FG),
        ('SETUP INSTRUCTIONS', op['setup'], FB),
        ('ALGORITHM GUIDE', op['algo'], FGO),
        ('SHADOWBAN / BAN AVOIDANCE', op['shadow'], FR),
        ('LLM-IN-THE-LOOP (24/7)', op['llm_loop'], FALPHA),
        ('MANUAL FIRST (VALIDATE)', op['manual'], FO),
        ('AUTOMATE AFTER (SCALE)', op['auto_after'], FG),
    ]

    for label, content, font in sections:
        # Label row
        fb(ws2, r, MC2, H)
        ws2.cell(row=r, column=2, value=label).font = FH
        ws2.cell(row=r, column=3, value='').font = FB
        ws2.row_dimensions[r].height = 22
        r += 1

        # Content - split by newlines for readability
        lines = content.split('\n')
        for line in lines:
            fb(ws2, r, MC2, R2)
            ws2.cell(row=r, column=2, value='').font = FB
            ws2.cell(row=r, column=3, value=line).font = font
            ws2.cell(row=r, column=3).alignment = AW
            for c in range(1, MC2+1): ws2.cell(row=r, column=c).border = BD
            h = max(22, min(60, len(line) // 2 + 10))
            ws2.row_dimensions[r].height = h
            r += 1
    r += 2  # gap between ops


# ============================================================
# SHEET 3: LLM ALPHA THESIS
# ============================================================
ws3 = wb.create_sheet('LLM ALPHA THESIS')
ws3.sheet_properties.tabColor = 'F0C000'
sw(ws3, [3, 30, 36, 28, 3])
MC3 = 5

for r in range(1, 60):
    fb(ws3, r, MC3, D)

ws3.merge_cells('B2:D2')
ws3.cell(row=2, column=2, value='LLM AUTOMATION ALPHA — WHY THIS EDGE EXISTS').font = FT
ws3.row_dimensions[2].height = 32
ws3.merge_cells('B3:D3')
ws3.cell(row=3, column=2, value='Alpha = advantage from LLM automation that disappears once mass adoption occurs.').font = FS

r = 5
r = sh(ws3, r, MC3, 'ALPHA OPPORTUNITIES RANKED BY EDGE DURATION')
r = hd(ws3, r, MC3, ['', 'OPPORTUNITY', 'WHY LLM CREATES EDGE', 'EDGE DURATION ESTIMATE'])

alphas = [
    ('Ecomm product scouting + auto-listing', 'Human scans 50 products/day. LLM scans 10,000. First-mover on trending products = 40% margins before competition arrives.', '12-24 months'),
    ('Hyper-personalized cold email at scale', 'Manual personalization = 20 emails/hour. LLM = 500 emails/hour with website-scraped personalization. 3-5x reply rate vs templates.', '18-36 months'),
    ('SEO content production (10-20 articles/day)', 'Human writes 1-2 articles/day. LLM writes 10-20 with proper E-E-A-T, schema, internal linking. Topical authority in weeks vs months.', '6-18 months (narrowing)'),
    ('App factory (2-3 apps/month)', 'Manual dev = 1 app/3 months. Claude Code = 2-3/month. Portfolio of 30 apps in 1 year = massive surface area.', '24-36 months'),
    ('AI UGC video production', 'Manual UGC = hire creators at $50-500/video. AI UGC via Kling/HeyGen = $0. 100 videos/week vs 5.', '12-18 months'),
    ('Multi-channel content repurposing', 'Human repurposes 1 piece to 3 platforms. LLM repurposes to 13 channels × 5 platforms = 65 posts from 1 piece.', '6-12 months (narrowing fast)'),
    ('Real-time market signal analysis', 'Human reads 10 earnings calls/quarter. LLM reads ALL earnings calls + on-chain data + options flow simultaneously.', '24-36 months'),
    ('Notion/digital product factory', 'Human creates 1 template/week. LLM generates structure + copy + SEO for 5-10/week.', '12-24 months'),
    ('Domain scoring at scale', 'Human evaluates 50 domains/day. LLM scores 10,000 by brandability, keyword, comps, backlinks.', '24-36 months'),
    ('AI music production + distribution', 'Musician creates 1 song/week. Suno + LLM prompts = 10 songs/day across genres. Spotify passive income.', '12-24 months'),
    ('Newsletter curation + writing', 'Human curator: 4-8 hrs per issue. LLM: scrape 145 sources + draft in 15 min. Human polish in 30 min.', '12-18 months'),
    ('Web redesign cold outreach with mockups', 'Agency manually prospects 10 businesses/day. LLM scrapes 100 + auto-generates redesign mockups + sends personalized emails.', '24-36 months'),
    ('YouTube script + video pipeline', 'Human creates 1 video/week. LLM pipeline: script + voice + visuals + upload = 1 video/day per channel.', '12-24 months'),
    ('Print on demand trend-reactive design', 'Human designer: 5-10 designs/day. LLM monitors trends + generates 200 designs/day with SEO titles.', '12-18 months'),
    ('Multi-persona social management', 'Human manages 1-2 accounts. LLM manages 8 personas × 4 platforms = 32 accounts with consistent voice.', '18-24 months'),
    ('Automated A/B testing at scale', 'Human tests 2-3 variants. LLM generates 50 variants + auto-measures + doubles down on winners.', '24-36 months'),
]

for idx, (opp, edge, duration) in enumerate(alphas):
    bg = ALPHA if '24-36' in duration else (G if '18' in duration else W)
    fb(ws3, r, MC3, bg)
    ws3.cell(row=r, column=2, value=opp).font = FW
    ws3.cell(row=r, column=2).alignment = AL
    ws3.cell(row=r, column=3, value=edge).font = FB
    ws3.cell(row=r, column=3).alignment = AL
    dur_font = FALPHA if '24-36' in duration else (FG if '18' in duration else FO)
    ws3.cell(row=r, column=4, value=duration).font = dur_font
    ws3.cell(row=r, column=4).alignment = AC
    for c in range(1, MC3+1): ws3.cell(row=r, column=c).border = BD
    ws3.row_dimensions[r].height = 55
    r += 1

# Thesis section
r += 2
r = sh(ws3, r, MC3, 'THE PRINTMAXX THESIS')
thesis = [
    'We are in a 12-36 month window where LLM-in-the-loop automation creates structural alpha across ecommerce, content, outbound, apps, and finance.',
    'The alpha exists because: (1) most people dont know how to build LLM automation pipelines, (2) those who do are focused on building tools, not using them to print, (3) the tools are free or near-free.',
    'The edge narrows as: (1) no-code LLM tools mature (make.com, n8n templates), (2) platform detection improves, (3) competition increases in each vertical.',
    'Strategy: run ALL 16 ops simultaneously. Each op is a lottery ticket. The portfolio approach guarantees at least 3-5 winners. Winners get doubled down.',
    'Manual first → validate unit economics → automate → scale. Never automate before validating. Never scale before automating.',
    'The compounding effect: social accounts promote newsletters which promote apps which fund ads which amplify everything. Every op feeds every other op.',
]
for t in thesis:
    fb(ws3, r, MC3, R2)
    ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws3.cell(row=r, column=2, value=t).font = FB
    ws3.cell(row=r, column=2).alignment = AL
    ws3.row_dimensions[r].height = 40
    r += 1


OUTPUT = '/sessions/awesome-nice-brown/mnt/PRINTMAXX_STARTER_KITttttt/PRINTMAXX_OPS_PLAYBOOK.xlsx'
wb.save(OUTPUT)
print(f'Saved to {OUTPUT}')
print(f'Total ops: {len(OPS)}')
