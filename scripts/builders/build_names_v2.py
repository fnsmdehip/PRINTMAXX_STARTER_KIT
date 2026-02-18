from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Colors
BG_DARK = PatternFill('solid', fgColor='0D1117')
BG_HEADER = PatternFill('solid', fgColor='0A1628')
BG_SECTION = PatternFill('solid', fgColor='112240')
BG_ROW1 = PatternFill('solid', fgColor='0D1117')
BG_ROW2 = PatternFill('solid', fgColor='161B22')
BG_PICK = PatternFill('solid', fgColor='0B3D0B')
BG_ALT = PatternFill('solid', fgColor='1A2332')
BG_GOLD = PatternFill('solid', fgColor='2D2500')
BG_P1 = PatternFill('solid', fgColor='0B3D0B')  # Phase 1 green
BG_P2 = PatternFill('solid', fgColor='1A2D4A')  # Phase 2 blue
BG_P3 = PatternFill('solid', fgColor='2D1B4E')  # Phase 3 purple

FONT_TITLE = Font(name='Arial', size=20, bold=True, color='00D4FF')
FONT_SUB = Font(name='Arial', size=12, bold=True, color='8B949E')
FONT_HEAD = Font(name='Arial', size=11, bold=True, color='00D4FF')
FONT_SEC = Font(name='Arial', size=12, bold=True, color='58A6FF')
FONT_BODY = Font(name='Arial', size=10, color='C9D1D9')
FONT_WHITE = Font(name='Arial', size=10, bold=True, color='FFFFFF')
FONT_CYAN = Font(name='Arial', size=10, bold=True, color='00D4FF')
FONT_GREEN = Font(name='Arial', size=10, bold=True, color='3FB950')
FONT_GOLD = Font(name='Arial', size=10, bold=True, color='F0C000')
FONT_ORANGE = Font(name='Arial', size=10, color='F0883E')
FONT_PURPLE = Font(name='Arial', size=10, bold=True, color='A371F7')
FONT_RED = Font(name='Arial', size=10, color='F85149')
FONT_HANDLE = Font(name='Arial', size=11, bold=True, color='FFFFFF')
FONT_SMALL = Font(name='Arial', size=9, color='8B949E')
FONT_PHASE = Font(name='Arial', size=9, bold=True, color='3FB950')

BORDER = Border(bottom=Side(style='thin', color='21262D'))
BORDER_SEC = Border(bottom=Side(style='medium', color='00D4FF'))
ALIGN_C = Alignment(horizontal='center', vertical='center')
ALIGN_L = Alignment(horizontal='left', vertical='center', wrap_text=True)

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def fill_bg(ws, row, cols, fill):
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).fill = fill

def section_header(ws, row, cols, text, font=FONT_SEC):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=cols-1)
    ws.cell(row=row, column=2, value=text).font = font
    fill_bg(ws, row, cols, BG_SECTION)
    for c in range(1, cols+1):
        ws.cell(row=row, column=c).border = BORDER_SEC
    ws.row_dimensions[row].height = 28
    return row + 1

def col_headers(ws, row, cols, headers):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=i+1, value=h)
        c.font = FONT_HEAD
        c.fill = BG_HEADER
        c.alignment = ALIGN_C
    ws.row_dimensions[row].height = 24
    return row + 1

# ============================================================
# MASTER DATA: ALL 38 NICHES + CONTENT FARM + AI PERSONAS
# Each entry: (niche_id, niche_name, phase, handle, alt_handle, tiktok, youtube, reason, alt_reason)
# ============================================================

# Format: list of tuples
# (id, name, phase, target, top_handle, top_reason, alt_handle, alt_reason, extra1, extra1_reason, extra2, extra2_reason)

ALL_NICHES = [
    # ── PHASE 1 NICHES (12) ──
    ('N001', 'AI Utilities', 'P1', 'Tech workers 25-40',
     '@patchnotes_', '"Patch notes" = software update logs. Every dev checks these. Underscore = clean tech energy.',
     '@inferenceloop', '"Inference loop" = ML model running predictions. Pure insider ML terminology.',
     '@hotreload', '"Hot reload" = dev term for live code refresh. Short, punchy, universally dev-recognized.',
     '@promptdrift', '"Prompt drift" = real LLM concept where outputs shift. Insider, technical, memorable.'),

    ('N002', 'Faith Streak', 'P1', 'Christians 18-35',
     '@morningwatch_', '"Morning watch" = old Christian term for early prayer/devotion. Historical, authentic, not mega-church-cringe.',
     '@selahmoments', '"Selah" = untranslated Hebrew word in Psalms meaning pause/reflect. Deep biblical insider. Beautiful.',
     '@narrowgate_', '"Narrow gate" = Matthew 7:13-14 reference. Commitment signal without being preachy.',
     '@vesperbell', '"Vespers" = evening prayer service. "Bell" = church bell. Atmospheric, poetic, ancient.'),

    ('N003', 'Fitness Men', 'P1', 'Men 25-45',
     '@repscheme_', '"Rep scheme" = specific set/rep structure. Every gym rat knows this. Underscore = clean.',
     '@deloadweek', '"Deload week" = planned recovery period. Only serious lifters know this. Insider signal.',
     '@compoundonly', '"Compound lifts" = squat/deadlift/bench. "Only" = purist philosophy. No machines.',
     '@failureset', '"Training to failure" = going until you cant. "Set" = one round. Raw gym floor language.'),

    ('N006', 'Neurodivergent ADHD', 'P1', 'ADHD adults 20-45',
     '@executivedysf', '"Executive dysfunction" shortened. THE defining ADHD term. Every ADHDer immediately recognizes it.',
     '@dopaminemenu', '"Dopamine menu" = ADHD coping technique (list of dopamine-friendly activities). Insider self-care term.',
     '@walloftasks', '"Wall of awful" concept + task paralysis. ADHD community immediately gets it.',
     '@hyperfixated_', '"Hyperfixation" = ADHD deep-focus state. Common community term. Underscore = personal brand.'),

    ('N011', 'Sleep Optimization', 'P1', 'Sleep-deprived 25-55',
     '@sleepcodex', '"Codex" = ancient manuscript. "Sleep codex" = collected knowledge of sleep. Elevated, not generic.',
     '@circadianrift', '"Circadian" = body clock. "Rift" = disruption. The problem IS the name.',
     '@remscore_', '"REM" = sleep stage. "Score" = quantified metric. Sleep tracker energy.',
     '@melatoninwindow', '"Melatonin window" = optimal time when melatonin kicks in. Insider sleep science term.'),

    ('N015', 'Motivation / Self-Improvement', 'P1', 'Self-improvers 20-40',
     '@quotesupply', '"Supply" = industrial feel. Not "daily quotes" cringe — supply chain of quotes. Functional, not inspirational.',
     '@atomicloop_', '"Atomic Habits" reference + "loop" = feedback cycle. Self-improvement insider book reference.',
     '@compounddaily', '"Compound" = compound interest metaphor for daily improvement. 1% better framing.',
     '@marginalgains_', '"Marginal gains" = British cycling strategy (1% improvements). Sports science → self-improvement crossover.'),

    ('N021', 'Autism / Neurodivergent', 'P1', 'Autistic adults 18-45',
     '@stimboard_', '"Stim" = stimming (self-stimulatory behavior). "Board" = mood/stim board. Autistic community insider term.',
     '@maskingtax', '"Masking" = autistic masking (hiding traits). "Tax" = the energy cost. Community-specific concept.',
     '@infodumpzone', '"Info dump" = autistic trait of deep-sharing about special interests. "Zone" = safe space.',
     '@spooncount_', '"Spoon theory" = disability/chronic illness energy management. "Count" = tracking spoons left.'),

    ('N031', 'Voice AI Service Businesses', 'P1', 'Local service biz 30-60',
     '@callrouted', '"Call routed" = telephony term. The call was handled. Clean, functional, B2B.',
     '@voicefunnel_', '"Voice" + "funnel" = phone calls as a sales funnel. B2B terminology mashup.',
     '@afterhourspick', '"After hours" = when businesses miss calls. "Pick up" = answering. The problem is the name.',
     '@dialtonedead', '"Dial tone dead" = missed calls killing revenue. Attention-grabbing, problem-focused.'),

    ('N032', 'Faceless TikTok Shop', 'P1', 'Creators 18-35',
     '@nofacesales', '"No face" = faceless content. "Sales" = what it does. Direct, no pretense.',
     '@ghostedshop', '"Ghosted" = invisible/faceless. "Shop" = TikTok Shop. Also plays on "getting ghosted" culture.',
     '@clipcommission', '"Clip" = short video. "Commission" = affiliate earnings. Format + revenue model as name.',
     '@facelessflip', '"Faceless" = no-face content. "Flip" = product flipping. Action-oriented.'),

    ('N034', 'Gen Z Workplace Burnout', 'P1', 'Gen Z workers 18-24',
     '@quietquitter_', '"Quiet quitting" = the Gen Z work philosophy that went viral. Immediate generational recognition.',
     '@sundayscaries', '"Sunday scaries" = anxiety before Monday. EVERY office worker knows this feeling.',
     '@clockedout_', '"Clocked out" = mentally checked out from work. Double meaning: done for the day + done with this BS.',
     '@corpcore', '"Corp core" = corporate aesthetic. Short, TikTok-native term. Ironic but relatable.'),

    ('N035', 'Freelancer Late Payment', 'P1', 'Freelancers worldwide',
     '@net30never', '"Net 30" = payment terms (30 days). "Never" = they never actually pay on time. Every freelancer feels this.',
     '@invoicepurge', '"Invoice purge" = clearing unpaid invoices. Aggressive, action-oriented. Takes no BS.',
     '@pastdueposts', '"Past due" = overdue payment. "Posts" = content about it. The name IS the content format.',
     '@chasecheck_', '"Chase" = chasing payments. "Check" = payment check. Double meaning: chase the check / check on payment.'),

    ('N036', 'Senior Smart Home', 'P1', 'Seniors 60+',
     '@okhomesetup', '"OK Home" = like "OK Boomer" but warm. "Setup" = installation. Friendly, non-patronizing.',
     '@simplecircuit', '"Simple circuit" = easy tech. "Circuit" = technology/electronics. Clean, not condescending.',
     '@helphometech', '"Help" + "home tech" = exactly what it does. Clear for the target demo.',
     '@easewired', '"Ease" = make easy. "Wired" = connected/tech. Calming + technological.'),

    # ── PHASE 2 NICHES (12) ──
    ('N004', 'Pet Wellness', 'P2', 'Pet owners 25-55',
     '@vetvisitlog', '"Vet visit log" = tracking pet health visits. Functional, clear, useful name.',
     '@zoomieswatch', '"Zoomies" = when dogs run around crazy. Every pet owner knows this word. "Watch" = tracking.',
     '@kibbleaudit', '"Kibble" = pet food. "Audit" = reviewing quality. Playful + useful.',
     '@tailwagdata', '"Tail wag" = happy pet signal. "Data" = quantified. Cute but data-driven.'),

    ('N005', 'Senior Tech', 'P2', 'Seniors 60+',
     '@bigbuttonapp', '"Big button" = large UI elements for seniors. Literally what they need. Honest name.',
     '@techforfolks', '"Folks" = warm, inclusive term. Not "seniors" or "elderly." Respectful positioning.',
     '@slowscrollhelp', '"Slow scroll" = navigating at comfortable pace. "Help" = assistance. Non-judgmental.',
     '@largefontlife', '"Large font" = accessibility need. "Life" = lifestyle brand. Functional + positive.'),

    ('N007', 'Gen Z Finance', 'P2', 'Gen Z 18-28',
     '@splitthecheck', '"Split the check" = Gen Z money moment. Social dining culture + finance. Relatable entry point.',
     '@vibecheck_money', '"Vibe check" = Gen Z phrase. Applied to money = checking your financial vibe. Native terminology.',
     '@brokeish_', '"Broke-ish" = not fully broke but not rich. Gen Z financial reality. Self-deprecating, honest.',
     '@sideincomeera', '"Side income era" = TikTok "era" culture. Income + cultural framing.'),

    ('N009', 'Women Wellness', 'P2', 'Women 25-45',
     '@cyclesync_', '"Cycle syncing" = aligning habits with menstrual cycle. Women\'s health insider term. Growing trend.',
     '@cortisollevel', '"Cortisol" = stress hormone. "Level" = measurement. Science-based wellness, not woo-woo.',
     '@glowprotocol', '"Glow" = skin/health. "Protocol" = structured routine. Medical feel + beauty aspiration.',
     '@lutealphase_', '"Luteal phase" = menstrual cycle phase. Deep women\'s health terminology. Signals real knowledge.'),

    ('N010', 'Students / Education', 'P2', 'Students 16-25',
     '@cramwindow', '"Cram" = study intensely. "Window" = time window before exam. Every student knows this desperation.',
     '@flashcardgrind', '"Flashcard" = study tool. "Grind" = hard work. Study culture + action.',
     '@studydebt_', '"Study debt" = double meaning (student loans + sleep debt from studying). Clever wordplay.',
     '@lastminuterev', '"Last minute revision" = universal student behavior. Honest, relatable, not aspirational.'),

    ('N012', 'Gaming / Esports', 'P2', 'Gamers 16-30',
     '@framedata_', '"Frame data" = fighting game term (how many frames each move takes). Hardcore gaming insider term.',
     '@hitboxleak', '"Hitbox" = collision detection area in games. "Leak" = revealing hidden info. Gaming + drama.',
     '@respawnloop', '"Respawn" = coming back after death. "Loop" = doing it repeatedly. Dark humor, universally gaming.',
     '@patchmeta_', '"Patch meta" = the dominant strategy after a game update. Competitive gaming terminology.'),

    ('N013', 'Crypto Trading', 'P2', 'Crypto traders 25-45',
     '@mempooldive', '"Mempool" = where pending txs wait. "Dive" = deep analysis. Insider crypto infra terminology.',
     '@noncetrouble', '"Nonce" = cryptographic number used once. Sounds playful but deeply technical.',
     '@gasfeesonly', '"Gas fees" = ETH transaction costs. Self-deprecating: all profits go to gas. Relatable.',
     '@epochwatcher', '"Epoch" = time period in blockchain consensus. "Watcher" = monitoring.'),

    ('N016', 'Tech Explainers', 'P2', 'Tech enthusiasts 25-45',
     '@latentspace_', '"Latent space" = ML concept (compressed representation). Signals deep knowledge without explaining it.',
     '@stacktrace_', '"Stack trace" = error debugging output. Every developer has stared at one. Universal dev recognition.',
     '@bytesized_', '"Byte sized" = small + computing unit. Perfect for explainer content: small, digestible tech breakdowns.',
     '@kernelpanic_', '"Kernel panic" = critical system error. Dramatic but technical. Commands attention.'),

    ('N018', 'News / Current Events', 'P2', 'News consumers 25-55',
     '@breakwire_', '"Breaking wire" = news wire (AP, Reuters) breaking news alert. Old-school journalism terminology.',
     '@headlinedrift', '"Headline drift" = when stories shift from original facts. Media criticism angle.',
     '@primefold', '"Above the fold" journalism term + "prime" = most important. Prime fold = top story.',
     '@tickertape_', '"Ticker tape" = old news/stock ticker. Retro journalism aesthetic + urgency.'),

    ('N022', 'Menopause Women', 'P2', 'Women 40-60',
     '@hotflashlog', '"Hot flash" = primary menopause symptom. "Log" = tracking. Functional, direct, no euphemism.',
     '@secondspring_', '"Second spring" = Chinese/Japanese term for menopause (positive reframing). Cultural, empowering.',
     '@hormoneshift', '"Hormone shift" = what menopause literally is. Medical accuracy, not cutesy.',
     '@periwatch_', '"Perimenopause" shortened + "watch" = monitoring. Stage-specific, insider medical term.'),

    ('N023', 'Parent Burnout', 'P2', 'Parents 30-50',
     '@mentalloadmap', '"Mental load" = the invisible task management parents (esp. mothers) carry. Viral concept.',
     '@naptimewindow', '"Nap time" = the only free time parents get. "Window" = precious time slot.',
     '@touchedout_', '"Touched out" = parent term for sensory overload from kids constantly touching you. Deep insider.',
     '@snackshift', '"Snack" = constant child demand. "Shift" = work shift. Parenting as a shift job. Dark humor.'),

    ('N025', 'Caregiver Support', 'P2', 'Family caregivers 35-65',
     '@respitecheck', '"Respite" = temporary relief from caregiving. "Check" = checking in. Clinical + warm.',
     '@burnshield_', '"Burnout" + "shield" = protection from caregiver burnout. Protective positioning.',
     '@carerotation', '"Care rotation" = scheduling caregiving shifts among family. Practical, functional term.',
     '@compassionfatigue_', '"Compassion fatigue" = real clinical term for caregiver emotional exhaustion. Insider medical term.'),

    ('N029', 'Career Pivoters', 'P2', 'Professionals 28-45',
     '@lateralshift_', '"Lateral" = sideways career move. "Shift" = change. Not a step down, a pivot. Positive framing.',
     '@skillgapreport', '"Skill gap" = the gap between current skills and new career requirements. HR/career insider term.',
     '@pivotstory_', '"Pivot" = career change. "Story" = narrative you tell employers. Personal branding term.',
     '@transferable_', '"Transferable skills" = skills that cross industries. THE career pivot concept. Shortened = clean.'),

    ('N030', 'Microentrepreneurs', 'P2', 'Side hustlers 25-40',
     '@cashflowfirst', '"Cash flow" = fundamental business metric. "First" = priority. Business fundamentals, not hustle culture.',
     '@marginwatch_', '"Margin" = profit margin. "Watch" = monitoring. Functional business tracking.',
     '@revenuerun_', '"Revenue run" = revenue run rate. Financial modeling term applied to micro-business. Professional framing.',
     '@uniteconomics_', '"Unit economics" = cost/revenue per unit sold. The real metric that matters. MBA terminology for side hustles.'),

    ('N033', 'Vertical SaaS Clinics', 'P2', 'Clinic owners 35-60',
     '@chartpulse_', '"Chart" = patient chart/EMR. "Pulse" = vital sign + real-time. Medical + tech fusion.',
     '@clinicstack', '"Clinic" = target. "Stack" = tech stack. B2B SaaS naming pattern.',
     '@intakeflow_', '"Intake" = patient intake form. "Flow" = workflow automation. Exact pain point as name.',
     '@waitroomfix', '"Wait room" = the problem (long waits). "Fix" = the solution. Pain point naming.'),

    ('N037', 'Field Service Route Optimization', 'P2', 'Field service SMBs',
     '@routedrift_', '"Route drift" = when technicians deviate from optimal route. The problem IS the name.',
     '@dispatchloop', '"Dispatch" = sending technicians. "Loop" = feedback/optimization loop. Operations terminology.',
     '@drivetimekill', '"Drive time" = non-billable transit. "Kill" = eliminate/reduce. Every field service owner hates drive time.',
     '@jobprofitmap', '"Job profit" = per-job profitability. "Map" = visual route mapping. Two metrics in one name.'),

    ('N038', 'SMB Compliance', 'P2', 'SMBs 10-500 employees',
     '@auditready_', '"Audit ready" = the state every business wants to be in. Aspirational + functional.',
     '@compliancegap', '"Compliance gap" = distance between current state and requirements. Risk/audit terminology.',
     '@policystack_', '"Policy" = compliance documents. "Stack" = collection/tech stack. GRC terminology.',
     '@riskledger', '"Risk" = compliance risk. "Ledger" = record-keeping. Financial + compliance mashup.'),

    # ── PHASE 3 NICHES (8) ──
    ('N008', 'Couples / Relationship', 'P3', 'Couples 25-45',
     '@bidsandturns', '"Bids" and "turns" = Gottman Institute research terms for relationship interactions. Therapy insider.',
     '@loveledger_', '"Love" + "ledger" = tracking relationship investments. Business metaphor for relationships.',
     '@attachmentstyle_', '"Attachment style" = psychology term (anxious, avoidant, secure). Therapy/self-help insider.',
     '@repairconvo', '"Repair conversation" = Gottman term for resolving conflict. Couples therapy insider.'),

    ('N014', 'Stock Investing', 'P3', 'Investors 30-55',
     '@bidaskspread', '"Bid-ask spread" = fundamental market mechanics. Every trader knows it. Clean, technical.',
     '@darkpoolflow', '"Dark pool" = off-exchange trading. "Flow" = order flow. Institutional insider terminology.',
     '@maxpainlevel', '"Max pain" = strike price where most options expire worthless. Technical, insider.',
     '@thetadecay_', '"Theta decay" = options Greek measuring time decay. Underscore = technical handle energy.'),

    ('N017', 'Sports Betting', 'P3', 'Sports fans 21-45',
     '@sharpaction_', '"Sharp action" = bets from professional/smart bettors. Sportsbook insider terminology.',
     '@closingline_', '"Closing line" = the final odds before a game. Getting closing line value = peak betting.',
     '@steampick', '"Steam" = sudden line movement from sharp money. "Pick" = selection. Insider sportsbook term.',
     '@coverthenut', '"Cover the nut" = bet enough to cover costs. Gambling terminology, raw, insider.'),

    ('N019', 'ASMR / Relaxation', 'P3', 'Relaxation seekers 18-45',
     '@tingletrace', '"Tingle" = ASMR sensation. "Trace" = gentle movement/tracing triggers. ASMR-specific vocabulary.',
     '@whisperlayer', '"Whisper" = most common ASMR trigger. "Layer" = layered audio. Production technique term.',
     '@triggermap_', '"Trigger" = ASMR trigger type. "Map" = mapping what works for you. Community insider.',
     '@softclickloop', '"Soft click" = keyboard/tapping ASMR. "Loop" = continuous relaxation. Specific trigger naming.'),

    ('N020', 'Lifestyle / Aesthetic', 'P3', 'Aspirational 20-40',
     '@moodboardleak', '"Mood board" = design/aesthetic planning. "Leak" = revealing/sharing. Curated aesthetic energy.',
     '@cleanfeedonly', '"Clean feed" = curated Instagram aesthetic. "Only" = purist. Visual culture insider.',
     '@editedreality_', '"Edited reality" = self-aware about curation. Meta-commentary on lifestyle content.',
     '@tonaldrift', '"Tonal" = color tone. "Drift" = gradual aesthetic shift. Photography/design terminology.'),

    ('N026', 'Chronic Pain', 'P3', 'Chronic pain 30-60',
     '@flarelog_', '"Flare" = pain flare-up. "Log" = tracking. Clinical + functional. The community uses this term daily.',
     '@painbaseline', '"Baseline" = normal pain level. "Pain baseline" = the level you live with. Clinical term, community insider.',
     '@triggerstacked', '"Trigger stacked" = multiple pain triggers happening at once. Community term for bad days.',
     '@pacecheck_', '"Pacing" = chronic pain management strategy (not overdoing it). "Check" = self-monitoring.'),

    ('N027', 'Introverts Socializing', 'P3', 'Introverts 20-40',
     '@socialbattery_', '"Social battery" = energy for social interaction. THE introvert metaphor. Universal recognition.',
     '@rechargezone', '"Recharge" = what introverts need after socializing. "Zone" = safe space for recovery.',
     '@smalltalkdebt', '"Small talk" = what introverts dread. "Debt" = the energy cost. Relatable, self-deprecating.',
     '@exitstrategy_', '"Exit strategy" = plan to leave social events early. Startup term repurposed for introverts. Clever.'),

    ('N028', 'Grief / Bereavement', 'P3', 'Bereaved adults 25-65',
     '@griefledger_', '"Grief" + "ledger" = recording grief journey. Quiet, dignified, not performative.',
     '@wavestill_', '"Waves" = grief comes in waves (universal grief metaphor). "Still" = finding stillness in grief.',
     '@emptychair_', '"Empty chair" = Gestalt therapy technique + literal absence at holidays. Deeply resonant.',
     '@afterlosslog', '"After loss" = life after loss. "Log" = journaling/tracking. Clear, compassionate, not clinical.'),
]

# ── CONTENT FARM SUB-METHODS (CF001-CF013) ──
CONTENT_FARMS = [
    ('CF001', 'Relax Channels', '@ambientvault', '"Ambient" = background sound genre. "Vault" = large collection. Sleep/relax YouTube channel.'),
    ('CF002', 'Sleep Timer / Alarm', '@sleepinterval', '"Sleep interval" = time between sleep cycles. Functional, science-based.'),
    ('CF003', 'News Socials', '@breakwire_', '"Breaking wire" = news wire alert. Old-school journalism urgency.'),
    ('CF004', 'Meme Channels', '@foldersdump', '"Folders dump" = dumping saved memes. Internet slang, casual energy.'),
    ('CF005', 'Women Appreciation', '@softeraclub', '"Soft era" = TikTok trend of embracing femininity. "Club" = community.'),
    ('CF006', 'Clip Channels', '@rawclipvault', '"Raw" = unedited. "Clip vault" = stored clips. Archive positioning.'),
    ('CF007', 'Motivation Quotes', '@marginalgains_', '"Marginal gains" = 1% improvement philosophy. Sports science crossover.'),
    ('CF008', 'Finance News', '@morningdelta', '"Morning" = AM delivery. "Delta" = change (options Greek). Finance insider.'),
    ('CF009', 'Tech Explainers', '@bytesized_', '"Byte sized" = computing unit + small/digestible. Tech explainer perfect name.'),
    ('CF010', 'Sports Highlights', '@buzzerfile', '"Buzzer" = buzzer beater. "File" = archived highlights. Sports + archive.'),
    ('CF011', 'Crypto News', '@blockheight_', '"Block height" = blockchain measurement. Pure crypto infra terminology.'),
    ('CF012', 'Stock Research', '@earningsdrift', '"Earnings drift" = post-earnings announcement drift (finance concept). Insider terminology.'),
    ('CF013', 'YouTube Shorts', '@shortcutfeed', '"Shortcut" = abbreviated + YouTube Shorts. "Feed" = content stream.'),
]

# ── AI INFLUENCER PERSONAS (AI001-AI008) ──
AI_PERSONAS = [
    ('AI001', 'Niche Experts', '@kiraellis_', 'Sounds like a real thought leader. "Kira" = memorable. "Ellis" = professional surname.', '@marcusvane_', '"Marcus Vane" = authoritative male expert persona. Sounds like a real consultant.'),
    ('AI002', 'FinDom', '@jadevelvet', '"Jade" = precious stone name. "Velvet" = luxury texture. Elegant, not explicit.', '@ivorycommand', '"Ivory" = luxury. "Command" = authority. Power dynamic naming.'),
    ('AI003', 'OnlyFans/Fansly', '@novawinters', '"Nova" = star. "Winters" = seasonal surname. Atmospheric, sounds real.', '@sierrablake_', '"Sierra" = nature. "Blake" = gender-neutral surname. Influencer-coded.'),
    ('AI004', 'ASMR', '@whispervelvet', '"Whisper" = ASMR core trigger. "Velvet" = soft texture. Sensory naming pattern.', '@softclickemiko', '"Soft click" = ASMR trigger. "Emiko" = Japanese name. International ASMR aesthetic.'),
    ('AI005', 'Fitness Coaches', '@ivorsteel', '"Ivor" = strong Anglo name. "Steel" = strength. Gym influencer without "fit" or "gains".', '@tannerforged', '"Tanner" = rugged name. "Forged" = built through effort. Blacksmith metaphor.'),
    ('AI006', 'Lifestyle Curators', '@siennacurates', '"Sienna" = warm name. "Curates" = taste/selection. Lifestyle influencer pattern.', '@lunavoss', '"Luna" = moon/celestial. "Voss" = clean European surname. Aspirational.'),
    ('AI007', 'Gaming Personas', '@glitchframe', '"Glitch" = gaming error/aesthetic. "Frame" = frame rate. Gaming-native naming.', '@lagspike_', '"Lag spike" = sudden latency. Every gamer knows this rage. Relatable.'),
    ('AI008', 'Relationship Advice', '@lunarcassidy', '"Lunar" = celestial/emotional. "Cassidy" = approachable Western name. Spiritually-coded.', '@reyforbes', '"Rey" = gender-neutral cool. "Forbes" = wealth/success signal. Finance-adjacent.'),
]

# ── KEY MONEY METHODS (top names for business brands) ──
MONEY_METHODS = [
    ('MM001', 'App Factory', 'streakpilot', '"Streak" = habit streaks. "Pilot" = navigator. Habit tracking app name.', 'focusburn', '"Focus" + "burn" = mental energy. Study/focus app.'),
    ('MM006', 'Content Farm', '@rawclipvault', '"Raw clip vault" = unedited clip archive. Content farm name.', '@deepcutsfeed', '"Deep cuts" = music term for non-obvious tracks. Applied to content = hidden gems.'),
    ('MM007', 'Cold Outbound', '@pipelinepoisn', '"Pipeline poison" = bad leads. Shortened. Inside sales term.', '@coldsequence', '"Cold sequence" = automated email chain. SDR terminology.'),
    ('MM009', 'AI Influencer', '@emikohaze', '"Emiko" = Japanese name. "Haze" = mystique. International, not obviously AI.', '@reyforbes', '"Rey" = cool. "Forbes" = wealth. Finance influencer energy.'),
    ('MM012', 'Algo Trading', '@darkpoolflow', '"Dark pool" = off-exchange. "Flow" = order flow. Institutional insider.', '@vwapdeviator', '"VWAP deviation" = trading strategy. Day trader insider.'),
    ('MM015', 'Newsletter', '@signalnoise_', '"Signal vs noise" = information theory. Filter noise, deliver signal.', '@thebrief_', '"Brief" = short report. "The" = authority. Newsletter naming pattern.'),
    ('MM016', 'TikTok Shop', '@clipcommission', '"Clip" = video. "Commission" = affiliate earnings. Format + revenue as name.', '@ghostedshop', '"Ghosted" = faceless. "Shop" = TikTok Shop. Plays on culture.'),
    ('MM019', 'Portfolio App Builder', 'routinedrift', '"Routine drift" = when habits slip. Problem-as-name. App catches the drift.', 'pulsestack', '"Pulse" = real-time. "Stack" = tech stack. Analytics app name.'),
    ('MM020', 'X Launch Viral', '@shipclock', '"Ship" = launch. "Clock" = speed/time. Ship fast or die.', '@coldstart_', '"Cold start problem" = classic growth term. Getting first users.'),
    ('MM021', 'Personal Brand SEO', '@printmaxxer', '"Maxxer" = optimizer (internet culture). Self-aware, build-in-public.', '@exitvelocity_', '"Exit velocity" = physics/startup term. Escaping competition.'),
    ('MM041', 'Directory Listing', 'listrank_', '"List" + "rank" = directory positioning. SEO-native naming.', 'indexgrid', '"Index" = search index. "Grid" = structured layout. Directory tool.'),
    ('MM044', 'Rapid Build Monetization', '@pushtomaster', '"Push to master" = git command. Ship to production. Bold dev humor.', '@ramenclub', '"Ramen profitable" = startup survival term. Community framing.'),
    ('MM045', 'Domain Flipping', '@flipregistry', '"Flip" = buy/sell. "Registry" = domain registry. Domain trading name.', '@parkedgold', '"Parked" = parked domain. "Gold" = valuable finds.'),
    ('MM046', 'Notion Templates', '@templatedrift', '"Template" = what it sells. "Drift" = aesthetic movement. Clean, professional.', '@notionstack_', '"Notion" = platform. "Stack" = collection. Direct but branded.'),
    ('MM070', 'Web Redesign Outreach', '@siteauditcold', '"Site audit" = web analysis. "Cold" = cold outreach. The whole method as name.', '@redesignpitch', '"Redesign pitch" = exactly the service. Transparent, B2B.'),
]


# ============================================================
# SHEET 1: MASTER — ALL 38 NICHES
# ============================================================
ws = wb.active
ws.title = 'ALL 38 NICHES'
ws.sheet_properties.tabColor = '00D4FF'
set_widths(ws, [3, 8, 22, 6, 16, 20, 44, 20, 44, 3])
max_col = 10

for r in range(1, 400):
    fill_bg(ws, r, max_col, BG_DARK)

ws.merge_cells('B2:I2')
ws.cell(row=2, column=2, value='ALL 38 NICHES — SOCIAL ACCOUNT NAMES').font = FONT_TITLE
ws.row_dimensions[2].height = 35
ws.merge_cells('B3:I3')
ws.cell(row=3, column=2, value='Inside-baseball handles for every niche. 4 options each. 152 names total.').font = FONT_SUB

r = 5
headers = ['', 'ID', 'NICHE', 'PH', 'TARGET', 'TOP PICK', 'WHY IT WORKS', 'ALT PICK', 'ALT REASONING']
r = col_headers(ws, r, max_col, headers)

current_phase = None
for (nid, name, phase, target, top, top_r, alt, alt_r, ex1, ex1_r, ex2, ex2_r) in ALL_NICHES:
    # Phase separator
    phase_labels = {'P1': 'PHASE 1 — LAUNCH IMMEDIATELY (12 NICHES)', 'P2': 'PHASE 2 — GROWTH STAGE (14 NICHES)', 'P3': 'PHASE 3 — EXPANSION (8 NICHES)'}
    if phase != current_phase:
        current_phase = phase
        r = section_header(ws, r, max_col, phase_labels.get(phase, phase))

    # Top pick row
    fill_bg(ws, r, max_col, BG_PICK)
    ws.cell(row=r, column=2, value=nid).font = FONT_CYAN
    ws.cell(row=r, column=2).alignment = ALIGN_C
    ws.cell(row=r, column=3, value=name).font = FONT_WHITE
    ws.cell(row=r, column=4, value=phase).font = FONT_PHASE
    ws.cell(row=r, column=4).alignment = ALIGN_C
    ws.cell(row=r, column=5, value=target).font = FONT_BODY
    ws.cell(row=r, column=5).alignment = ALIGN_C
    ws.cell(row=r, column=6, value=top).font = FONT_HANDLE
    ws.cell(row=r, column=7, value=top_r).font = FONT_BODY
    ws.cell(row=r, column=7).alignment = ALIGN_L
    ws.cell(row=r, column=8, value=alt).font = FONT_GOLD
    ws.cell(row=r, column=9, value=alt_r).font = FONT_BODY
    ws.cell(row=r, column=9).alignment = ALIGN_L
    for c in range(1, max_col+1):
        ws.cell(row=r, column=c).border = BORDER
    ws.row_dimensions[r].height = 42
    r += 1

    # Extra options row
    fill_bg(ws, r, max_col, BG_ROW2)
    ws.cell(row=r, column=6, value=ex1).font = FONT_BODY
    ws.cell(row=r, column=7, value=ex1_r).font = FONT_SMALL
    ws.cell(row=r, column=7).alignment = ALIGN_L
    ws.cell(row=r, column=8, value=ex2).font = FONT_BODY
    ws.cell(row=r, column=9, value=ex2_r).font = FONT_SMALL
    ws.cell(row=r, column=9).alignment = ALIGN_L
    for c in range(1, max_col+1):
        ws.cell(row=r, column=c).border = BORDER
    ws.row_dimensions[r].height = 32
    r += 1


# ============================================================
# SHEET 2: CONTENT FARM NAMES (CF001-CF013)
# ============================================================
ws2 = wb.create_sheet('CONTENT FARMS')
ws2.sheet_properties.tabColor = 'A371F7'
set_widths(ws2, [3, 8, 22, 22, 48, 3])
max_col2 = 6

for r in range(1, 30):
    fill_bg(ws2, r, max_col2, BG_DARK)

ws2.merge_cells('B2:E2')
ws2.cell(row=2, column=2, value='CONTENT FARM ACCOUNTS — CF001-CF013').font = FONT_TITLE
ws2.row_dimensions[2].height = 32
ws2.merge_cells('B3:E3')
ws2.cell(row=3, column=2, value='Faceless niche page names across 13 content verticals.').font = FONT_SUB

r = 5
headers2 = ['', 'ID', 'CONTENT TYPE', 'HANDLE', 'WHY IT WORKS']
r = col_headers(ws2, r, max_col2, headers2)

for idx, (cid, ctype, handle, reason) in enumerate(CONTENT_FARMS):
    bg = BG_ROW1 if idx % 2 == 0 else BG_ROW2
    fill_bg(ws2, r, max_col2, bg)
    ws2.cell(row=r, column=2, value=cid).font = FONT_CYAN
    ws2.cell(row=r, column=2).alignment = ALIGN_C
    ws2.cell(row=r, column=3, value=ctype).font = FONT_WHITE
    ws2.cell(row=r, column=4, value=handle).font = FONT_HANDLE
    ws2.cell(row=r, column=5, value=reason).font = FONT_BODY
    ws2.cell(row=r, column=5).alignment = ALIGN_L
    for c in range(1, max_col2+1):
        ws2.cell(row=r, column=c).border = BORDER
    ws2.row_dimensions[r].height = 32
    r += 1


# ============================================================
# SHEET 3: AI PERSONA NAMES (AI001-AI008)
# ============================================================
ws3 = wb.create_sheet('AI PERSONAS')
ws3.sheet_properties.tabColor = 'FF69B4'
set_widths(ws3, [3, 8, 22, 22, 48, 22, 48, 3])
max_col3 = 8

for r in range(1, 25):
    fill_bg(ws3, r, max_col3, BG_DARK)

ws3.merge_cells('B2:G2')
ws3.cell(row=2, column=2, value='AI INFLUENCER PERSONAS — AI001-AI008').font = FONT_TITLE
ws3.row_dimensions[2].height = 32
ws3.merge_cells('B3:G3')
ws3.cell(row=3, column=2, value='Names that pass as real people. Not obviously AI.').font = FONT_SUB

r = 5
headers3 = ['', 'ID', 'PERSONA TYPE', 'TOP NAME', 'REASONING', 'ALT NAME', 'ALT REASONING']
r = col_headers(ws3, r, max_col3, headers3)

for idx, (aid, atype, top, top_r, alt, alt_r) in enumerate(AI_PERSONAS):
    bg = BG_ROW1 if idx % 2 == 0 else BG_ROW2
    fill_bg(ws3, r, max_col3, bg)
    ws3.cell(row=r, column=2, value=aid).font = FONT_CYAN
    ws3.cell(row=r, column=2).alignment = ALIGN_C
    ws3.cell(row=r, column=3, value=atype).font = FONT_WHITE
    ws3.cell(row=r, column=4, value=top).font = FONT_HANDLE
    ws3.cell(row=r, column=5, value=top_r).font = FONT_BODY
    ws3.cell(row=r, column=5).alignment = ALIGN_L
    ws3.cell(row=r, column=6, value=alt).font = FONT_GOLD
    ws3.cell(row=r, column=7, value=alt_r).font = FONT_BODY
    ws3.cell(row=r, column=7).alignment = ALIGN_L
    for c in range(1, max_col3+1):
        ws3.cell(row=r, column=c).border = BORDER
    ws3.row_dimensions[r].height = 42
    r += 1


# ============================================================
# SHEET 4: MONEY METHOD BRANDS (MM names)
# ============================================================
ws4 = wb.create_sheet('MONEY METHODS')
ws4.sheet_properties.tabColor = '3FB950'
set_widths(ws4, [3, 8, 24, 22, 48, 22, 48, 3])
max_col4 = 8

for r in range(1, 35):
    fill_bg(ws4, r, max_col4, BG_DARK)

ws4.merge_cells('B2:G2')
ws4.cell(row=2, column=2, value='MONEY METHOD BRANDS — MM001-MM070').font = FONT_TITLE
ws4.row_dimensions[2].height = 32
ws4.merge_cells('B3:G3')
ws4.cell(row=3, column=2, value='Business/product names for each monetization method.').font = FONT_SUB

r = 5
headers4 = ['', 'ID', 'METHOD', 'TOP NAME', 'REASONING', 'ALT NAME', 'ALT REASONING']
r = col_headers(ws4, r, max_col4, headers4)

for idx, (mid, mtype, top, top_r, alt, alt_r) in enumerate(MONEY_METHODS):
    bg = BG_ROW1 if idx % 2 == 0 else BG_ROW2
    fill_bg(ws4, r, max_col4, bg)
    ws4.cell(row=r, column=2, value=mid).font = FONT_CYAN
    ws4.cell(row=r, column=2).alignment = ALIGN_C
    ws4.cell(row=r, column=3, value=mtype).font = FONT_WHITE
    ws4.cell(row=r, column=4, value=top).font = FONT_HANDLE
    ws4.cell(row=r, column=5, value=top_r).font = FONT_BODY
    ws4.cell(row=r, column=5).alignment = ALIGN_L
    ws4.cell(row=r, column=6, value=alt).font = FONT_GOLD
    ws4.cell(row=r, column=7, value=alt_r).font = FONT_BODY
    ws4.cell(row=r, column=7).alignment = ALIGN_L
    for c in range(1, max_col4+1):
        ws4.cell(row=r, column=c).border = BORDER
    ws4.row_dimensions[r].height = 38
    r += 1


# ============================================================
# SHEET 5: CROSS-PLATFORM MAP (all top picks mapped)
# ============================================================
ws5 = wb.create_sheet('CROSS-PLATFORM')
ws5.sheet_properties.tabColor = 'F0C000'
set_widths(ws5, [3, 8, 18, 22, 22, 22, 22, 28, 3])
max_col5 = 9

for r in range(1, 80):
    fill_bg(ws5, r, max_col5, BG_DARK)

ws5.merge_cells('B2:H2')
ws5.cell(row=2, column=2, value='CROSS-PLATFORM HANDLE MAPPING').font = FONT_TITLE
ws5.row_dimensions[2].height = 32
ws5.merge_cells('B3:H3')
ws5.cell(row=3, column=2, value='Every top pick mapped across X, TikTok, YouTube, Instagram, Email.').font = FONT_SUB

r = 5
headers5 = ['', 'ID', 'NICHE', 'X / TWITTER', 'TIKTOK', 'YOUTUBE', 'INSTAGRAM', 'EMAIL']
r = col_headers(ws5, r, max_col5, headers5)

for idx, (nid, name, phase, target, top, top_r, alt, alt_r, ex1, ex1_r, ex2, ex2_r) in enumerate(ALL_NICHES):
    bg = BG_ROW1 if idx % 2 == 0 else BG_ROW2
    fill_bg(ws5, r, max_col5, bg)

    # Generate cross-platform names from X handle
    x_handle = top
    # Strip @ and _
    clean = x_handle.replace('@', '').rstrip('_')
    tiktok = f'@{clean}'
    youtube = clean.replace('_', ' ').title()
    instagram = f'@{clean}'
    email = f'{clean}@protonmail.com'

    ws5.cell(row=r, column=2, value=nid).font = FONT_CYAN
    ws5.cell(row=r, column=2).alignment = ALIGN_C
    ws5.cell(row=r, column=3, value=name).font = FONT_WHITE
    ws5.cell(row=r, column=4, value=x_handle).font = FONT_HANDLE
    ws5.cell(row=r, column=5, value=tiktok).font = FONT_BODY
    ws5.cell(row=r, column=6, value=youtube).font = FONT_BODY
    ws5.cell(row=r, column=7, value=instagram).font = FONT_BODY
    ws5.cell(row=r, column=8, value=email).font = FONT_SMALL
    for c in range(1, max_col5+1):
        ws5.cell(row=r, column=c).border = BORDER
        ws5.cell(row=r, column=c).alignment = ALIGN_C
    ws5.row_dimensions[r].height = 26
    r += 1


# ============================================================
# SHEET 6: NAMING RULES (carried from v1)
# ============================================================
ws6 = wb.create_sheet('NAMING RULES')
ws6.sheet_properties.tabColor = 'F0883E'
set_widths(ws6, [3, 22, 42, 42, 3])
max_col6 = 5

for r in range(1, 50):
    fill_bg(ws6, r, max_col6, BG_DARK)

ws6.merge_cells('B2:D2')
ws6.cell(row=2, column=2, value='ANTI-CRINGE NAMING RULES').font = FONT_TITLE
ws6.row_dimensions[2].height = 32
ws6.merge_cells('B3:D3')
ws6.cell(row=3, column=2, value='Patterns from 145+ scraped accounts that actually work').font = FONT_SUB

r = 5
rules = [
    ('RULE', 'DO THIS', 'NOT THIS'),
    ('Use insider terminology', '@unusual_whales (trading lingo), @deloadweek (gym), @executivedysf (ADHD)', '@TradingTips, @FitnessGuru, @ADHDHelper'),
    ('Lowercase casual energy', '@videosinfolder, @levelsio, @cursaboringdays, @clockedout_', '@EPIC_CONTENT, @BEST_TRADES, @BURNOUT_HELP'),
    ('Contradiction/paradox', '@oddlyterrifying, @lowresgold, @factcoldcuts, @smalltalkdebt', '@AmazingFacts, @CoolContent, @IntrovertLife'),
    ('Specific > generic', '@NoContextBrits (region), @net30never (exact pain), @hotflashlog (exact symptom)', '@FunnyMemes, @FreelancerProbs, @WomenHealth'),
    ('Underscore = dev/tech signal', '@patchnotes_, @latentspace_, @flarelog_, @socialbattery_', 'No underscore = fine too. Just adds technical/personal brand energy.'),
    ('Name + concept (personas)', '@kiraellis_, @ivorsteel, @novawinters, @siennacurates', '@AI_Coach, @Digital_Lisa, @VirtualJane, @FitBot'),
    ('Problem-as-name', '@promptdrift (LLM issue), @routedrift_ (field service), @mentalloadmap (parenting)', '@EmailSolutions, @PainFixer, @RouteOptimizer'),
    ('Format-as-name', '@foldersdump (content dump), @rawclipvault (clip archive), @flarelog_ (pain diary)', '@ContentCollection, @VideoArchive, @PainTracker'),
    ('Numbers add mystique', '@Rainmaker1973, @tom777kruise, @KCodes7777', 'Random numbers: @user38492, @john123'),
    ('Self-deprecating humor', '@spamfolder_, @gasfeesonly, @brokeish_, @quietquitter_', '@SuccessfulTrader, @WealthBuilder, @ProductivityKing'),
    ('Community insider language', '@dopaminemenu (ADHD), @cyclesync_ (women), @bidsandturns (couples therapy)', '@ADHDSupport, @WomenWellness, @RelationshipGoals'),
]

for i, (rule, do, dont) in enumerate(rules):
    bg = BG_HEADER if i == 0 else (BG_ROW1 if i % 2 == 0 else BG_ROW2)
    font = FONT_HEAD if i == 0 else FONT_GOLD
    fill_bg(ws6, r, max_col6, bg)
    ws6.cell(row=r, column=2, value=rule).font = font
    ws6.cell(row=r, column=2).alignment = ALIGN_L
    ws6.cell(row=r, column=3, value=do).font = FONT_GREEN if i > 0 else FONT_HEAD
    ws6.cell(row=r, column=3).alignment = ALIGN_L
    ws6.cell(row=r, column=4, value=dont).font = FONT_RED if i > 0 else FONT_HEAD
    ws6.cell(row=r, column=4).alignment = ALIGN_L
    for c in range(1, max_col6):
        ws6.cell(row=r, column=c).border = BORDER
    ws6.row_dimensions[r].height = 40 if i > 0 else 26
    r += 1

r += 2
ws6.merge_cells(f'B{r}:D{r}')
ws6.cell(row=r, column=2, value='BANNED WORDS — NEVER USE IN HANDLES').font = Font(name='Arial', size=14, bold=True, color='F85149')
fill_bg(ws6, r, max_col6, BG_SECTION)
r += 1

banned = [
    'guru, expert, master, coach, pro, tips, hacks, mindset, grind, hustle',
    'unlock, secrets, revealed, proven, guaranteed, ultimate, best, top',
    'alpha (as "alpha male"), sigma, based (when forced), king, queen',
    'AI_ prefix (screams bot), Digital_, Virtual_, Smart_, Automated_',
    'Daily[Niche], [Niche]Tips, [Niche]Guru, [Niche]Master, [Niche]Pro',
    'journey, warrior, tribe, empower, transform, manifest, vibe (as brand)',
]
for ban in banned:
    fill_bg(ws6, r, max_col6, BG_DARK)
    ws6.merge_cells(f'B{r}:D{r}')
    ws6.cell(row=r, column=2, value=f'  {ban}').font = FONT_RED
    ws6.cell(row=r, column=2).alignment = ALIGN_L
    ws6.row_dimensions[r].height = 22
    r += 1


# ============================================================
# SAVE
# ============================================================
OUTPUT = '/sessions/awesome-nice-brown/mnt/PRINTMAXX_STARTER_KITttttt/PRINTMAXX_BRAND_NAMES.xlsx'
wb.save(OUTPUT)
print(f'Saved to {OUTPUT}')
print(f'Niches: {len(ALL_NICHES)}')
print(f'Content Farms: {len(CONTENT_FARMS)}')
print(f'AI Personas: {len(AI_PERSONAS)}')
print(f'Money Methods: {len(MONEY_METHODS)}')
total = len(ALL_NICHES)*4 + len(CONTENT_FARMS) + len(AI_PERSONAS)*2 + len(MONEY_METHODS)*2
print(f'Total names: {total}')
