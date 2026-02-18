from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Colors
BG_DARK = PatternFill('solid', fgColor='0D1117')
BG_HEADER = PatternFill('solid', fgColor='0A1628')
BG_SECTION = PatternFill('solid', fgColor='112240')
BG_ROW1 = PatternFill('solid', fgColor='0D1117')
BG_ROW2 = PatternFill('solid', fgColor='161B22')
BG_FREE = PatternFill('solid', fgColor='0B3D0B')
BG_WARN = PatternFill('solid', fgColor='2D2500')
BG_LIMIT = PatternFill('solid', fgColor='4A1A1A')
BG_CARD = PatternFill('solid', fgColor='161B22')

FONT_TITLE = Font(name='Arial', size=20, bold=True, color='00D4FF')
FONT_SUB = Font(name='Arial', size=12, bold=True, color='8B949E')
FONT_HEAD = Font(name='Arial', size=11, bold=True, color='00D4FF')
FONT_SEC = Font(name='Arial', size=12, bold=True, color='58A6FF')
FONT_BODY = Font(name='Arial', size=10, color='C9D1D9')
FONT_WHITE = Font(name='Arial', size=10, bold=True, color='FFFFFF')
FONT_CYAN = Font(name='Arial', size=10, bold=True, color='00D4FF')
FONT_GREEN = Font(name='Arial', size=10, bold=True, color='3FB950')
FONT_GOLD = Font(name='Arial', size=10, bold=True, color='F0C000')
FONT_ORANGE = Font(name='Arial', size=10, color='F0883E')
FONT_RED = Font(name='Arial', size=10, color='F85149')
FONT_STAT = Font(name='Arial', size=28, bold=True, color='00D4FF')
FONT_STAT_L = Font(name='Arial', size=10, color='8B949E')
FONT_HANDLE = Font(name='Arial', size=11, bold=True, color='FFFFFF')
FONT_SMALL = Font(name='Arial', size=9, color='8B949E')

BORDER = Border(bottom=Side(style='thin', color='21262D'))
BORDER_SEC = Border(bottom=Side(style='medium', color='00D4FF'))
ALIGN_C = Alignment(horizontal='center', vertical='center')
ALIGN_L = Alignment(horizontal='left', vertical='center', wrap_text=True)

def sw(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def fb(ws, row, cols, fill):
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).fill = fill

def sh(ws, row, cols, text, font=FONT_SEC):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=cols-1)
    ws.cell(row=row, column=2, value=text).font = font
    fb(ws, row, cols, BG_SECTION)
    for c in range(1, cols+1):
        ws.cell(row=row, column=c).border = BORDER_SEC
    ws.row_dimensions[row].height = 28
    return row + 1

def ch(ws, row, cols, headers):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=i+1, value=h)
        c.font = FONT_HEAD; c.fill = BG_HEADER; c.alignment = ALIGN_C
    ws.row_dimensions[row].height = 24
    return row + 1

MC = 9  # max columns

# ============================================================
# SHEET 1: $0 DEPLOYMENT DASHBOARD
# ============================================================
ws = wb.active
ws.title = '$0 DEPLOYMENT'
ws.sheet_properties.tabColor = '00D4FF'
sw(ws, [3, 24, 18, 20, 18, 18, 30, 18, 3])

for r in range(1, 100):
    fb(ws, r, MC, BG_DARK)

ws.merge_cells('B2:H2')
ws.cell(row=2, column=2, value='$0 PRINTMAXX DEPLOYMENT — MAXXED FREE TIERS').font = FONT_TITLE
ws.row_dimensions[2].height = 35
ws.merge_cells('B3:H3')
ws.cell(row=3, column=2, value='Every free tier squeezed to the limit. Zero monthly recurring cost.').font = FONT_SUB

# Stats
r = 5
stats = [
    ('$0/mo', 'RECURRING'),
    ('47', 'FREE TOOLS'),
    ('12', 'SOCIAL ACCTS'),
    ('3', 'NEWSLETTERS'),
    ('UNLIMITED', 'LANDING PAGES'),
    ('7,500', 'FREE SUBS'),
    ('3,000', 'EMAILS/DAY'),
]
for i, (num, lab) in enumerate(stats):
    col = 2 + i
    ws.cell(row=5, column=col, value=num).font = FONT_STAT if i == 0 else Font(name='Arial', size=18, bold=True, color='3FB950' if i < 3 else '00D4FF')
    ws.cell(row=5, column=col).fill = BG_CARD; ws.cell(row=5, column=col).alignment = ALIGN_C
    ws.cell(row=6, column=col, value=lab).font = FONT_STAT_L
    ws.cell(row=6, column=col).fill = BG_CARD; ws.cell(row=6, column=col).alignment = ALIGN_C
ws.row_dimensions[5].height = 40
ws.row_dimensions[6].height = 20

# Free tier tool table
r = 8
r = sh(ws, r, MC, 'CORE FREE STACK — WHAT YOU GET AT $0')
headers = ['', 'TOOL', 'CATEGORY', 'FREE TIER DETAIL', 'LIMIT', 'BOTTLENECK AT', 'UPGRADE COST', 'NOTES']
r = ch(ws, r, MC, headers)

free_stack = [
    ('Dolphin Anty', 'Anti-Detect', '10 browser profiles forever', '10 profiles', '5+ personas', '$89/mo', 'TIER 0. 3.3x more free profiles than GoLogin.'),
    ('Buffer', 'Social Scheduling', '3 channels free', '3 platforms', '4+ platforms', '$6/mo', 'TIER 0. Covers TikTok + IG + one more.'),
    ('Tweetlio', 'Social (X/Twitter)', 'Free plan, unlimited posts', 'Unlimited X posts', 'Never', '$12/mo', 'Adds X to Buffer stack = 4 platforms free.'),
    ('Canva', 'Design', '2.1M templates free', 'Unlimited designs', 'Never', '$13/mo', 'Everything design. Thumbnails, carousels, PDFs.'),
    ('CapCut', 'Video Editing', 'Full features free', 'Full suite', 'Never', 'N/A', 'Complete video editing at $0. TikTok-native.'),
    ('Leonardo.ai', 'AI Images', '150 tokens/day free', '~5-10 images/day', '20+ images/day', '$12/mo', 'Best free AI image gen. Persona photos, thumbnails.'),
    ('ElevenLabs', 'AI Voice', '10K chars/mo free', '~10 min audio/mo', 'Heavy audio', '$5/mo', 'AI voiceover for faceless content.'),
    ('Kling', 'AI Video', '66 credits/day free', '~3-5 clips/day', 'Mass production', '$5/mo', 'Best free AI video. TikTok Shop UGC.'),
    ('Beehiiv', 'Newsletter', '2,500 subs free', '2,500 per publication', '2,500 subs', '$49/mo', 'TIER 0. Run 3 publications = 7,500 free subs.'),
    ('Kit (ConvertKit)', 'Newsletter Alt', '10,000 subs + unlimited landing pages', '10,000 subs', '10K subs', '$29/mo', '4x Beehiiv free tier. Best for landing pages.'),
    ('Substack', 'Newsletter Alt', 'Unlimited subs', 'Unlimited', 'Never (10% take)', '10% of paid', 'Zero friction. Discovery network built in.'),
    ('TrulyInbox', 'Email Warmup', 'Free forever, unlimited', 'Unlimited warmup', 'Never', 'N/A', 'Only free-forever warmup tool. Critical for deliverability.'),
    ('Carrd', 'Landing Pages', '3 sites free', '3 landing pages', '4+ brands', '$19/year', 'Simple one-pagers. Perfect for newsletter capture.'),
    ('Systeme.io', 'All-in-One Funnels', '2K contacts + funnels + courses + email', '2,000 contacts', '2K contacts', '$27/mo', 'Replaces 3-4 tools. Non-tech founder dream.'),
    ('Oracle Cloud', 'VPS Hosting', '4 cores, 24GB RAM, always free', 'Unlimited', 'Never', 'N/A', 'Best free VPS on earth. Run n8n, PostHog, anything.'),
    ('Vercel', 'Frontend Hosting', '100GB/mo bandwidth free', '100GB/mo', 'Viral traffic', '$20/mo', 'Next.js default. 5-10 landing pages easy.'),
    ('Cloudflare Pages', 'Static Hosting', 'Unlimited free + free DNS', 'Unlimited', 'Never', '$20/mo Pro', 'Static sites + DNS for all domains. $0 forever.'),
    ('GA4', 'Web Analytics', 'Unlimited free', 'Unlimited events', 'Never', 'N/A', 'Standard analytics. Required for every site.'),
    ('PostHog', 'Product Analytics', '1M events/mo free', '1M events', '1M+ events', '$0 self-host', 'Self-host on Oracle VPS for unlimited.'),
    ('HubSpot', 'CRM', 'Unlimited contacts free', 'Unlimited', 'Never', '$50/mo', 'Free CRM forever. Contact management.'),
    ('Gumroad', 'Digital Products', 'Free until first sale (10% fee)', 'Unlimited products', 'Never', '10% fee', 'Launch digital products with zero upfront.'),
    ('Stripe', 'Payments', 'Free until sale (2.9% + $0.30)', 'Unlimited', 'Never', '2.9% + $0.30', 'Payment processing. Required for revenue.'),
    ('Notion', 'Docs/Wiki', 'Free personal plan', 'Unlimited pages', 'Team use', '$10/mo', 'Knowledge base, templates, digital products.'),
    ('Porkbun', 'Domains', '$9.73/yr per .com', '~$0.81/mo per domain', 'Many domains', '$9.73/yr', 'Cheapest registrar. Annual cost, not monthly.'),
    ('n8n', 'Workflow Automation', 'Self-host free on Oracle', 'Unlimited', 'Never (self-host)', '$20/mo cloud', 'Zapier alternative at $0. Run on Oracle VPS.'),
    ('Figma', 'UI Design', 'Free Starter plan', '3 projects', '4+ projects', '$15/mo', 'Industry standard. App mockups, landing page design.'),
    ('Photopea', 'Photo Editing', 'Completely free', 'Unlimited', 'Never', 'N/A', 'Browser Photoshop. $0 forever.'),
    ('Fiverr', 'Freelance Hiring', 'Free to browse', 'Unlimited browsing', 'Never', 'Per gig', 'UGC sourcing, VA hiring, dev work.'),
]

for idx, (tool, cat, free_detail, limit, bottleneck, upgrade, notes) in enumerate(free_stack):
    bg = BG_FREE if bottleneck == 'Never' else (BG_ROW1 if idx % 2 == 0 else BG_ROW2)
    fb(ws, r, MC, bg)
    ws.cell(row=r, column=2, value=tool).font = FONT_WHITE
    ws.cell(row=r, column=3, value=cat).font = FONT_BODY
    ws.cell(row=r, column=4, value=free_detail).font = FONT_GREEN
    ws.cell(row=r, column=4).alignment = ALIGN_L
    ws.cell(row=r, column=5, value=limit).font = FONT_BODY
    ws.cell(row=r, column=5).alignment = ALIGN_C
    bn_font = FONT_GREEN if bottleneck == 'Never' else (FONT_GOLD if 'Never' not in bottleneck else FONT_GREEN)
    ws.cell(row=r, column=6, value=bottleneck).font = bn_font
    ws.cell(row=r, column=6).alignment = ALIGN_C
    ws.cell(row=r, column=7, value=upgrade).font = FONT_ORANGE
    ws.cell(row=r, column=7).alignment = ALIGN_C
    ws.cell(row=r, column=8, value=notes).font = FONT_SMALL
    ws.cell(row=r, column=8).alignment = ALIGN_L
    for c in range(1, MC+1):
        ws.cell(row=r, column=c).border = BORDER
    ws.row_dimensions[r].height = 28
    r += 1


# ============================================================
# SHEET 2: BOTTLENECK MAP
# ============================================================
ws2 = wb.create_sheet('BOTTLENECK MAP')
ws2.sheet_properties.tabColor = 'F0C000'
sw(ws2, [3, 24, 22, 20, 24, 20, 30, 3])
MC2 = 8

for r in range(1, 50):
    fb(ws2, r, MC2, BG_DARK)

ws2.merge_cells('B2:G2')
ws2.cell(row=2, column=2, value='BOTTLENECK MAP — WHERE FREE TIERS HIT LIMITS').font = FONT_TITLE
ws2.row_dimensions[2].height = 32
ws2.merge_cells('B3:G3')
ws2.cell(row=3, column=2, value='Green = unlimited. Gold = generous but limited. Red = hard limit.').font = FONT_SUB

r = 5
r = sh(ws2, r, MC2, 'TIER 1: UNLIMITED FOREVER ($0)')
headers2 = ['', 'RESOURCE', 'FREE LIMIT', 'BOTTLENECK', 'SOLUTION', 'COST', 'STATUS']
r = ch(ws2, r, MC2, headers2)

unlimited = [
    ('HubSpot Contacts', 'Unlimited', 'Never', 'N/A', '$0', 'UNLIMITED'),
    ('GA4 Analytics', 'Unlimited events', 'Never', 'N/A', '$0', 'UNLIMITED'),
    ('Oracle Cloud VPS', '4 cores, 24GB RAM', 'Never', 'N/A', '$0', 'UNLIMITED'),
    ('Cloudflare Pages', 'Unlimited sites', 'Never', 'N/A', '$0', 'UNLIMITED'),
    ('Substack Subs', 'Unlimited', 'Never (10% take)', 'N/A', '$0', 'UNLIMITED'),
    ('Newsletter Sends', 'Unlimited per publication', 'Never', 'N/A', '$0', 'UNLIMITED'),
    ('Stripe/Gumroad', 'Unlimited products', 'Never (% fee)', 'N/A', '$0', 'UNLIMITED'),
    ('Photopea Editing', 'Unlimited', 'Never', 'N/A', '$0', 'UNLIMITED'),
    ('n8n (self-hosted)', 'Unlimited workflows', 'Never', 'N/A', '$0', 'UNLIMITED'),
]

for idx, (res, limit, bn, sol, cost, status) in enumerate(unlimited):
    fb(ws2, r, MC2, BG_FREE)
    ws2.cell(row=r, column=2, value=res).font = FONT_WHITE
    ws2.cell(row=r, column=3, value=limit).font = FONT_GREEN
    ws2.cell(row=r, column=4, value=bn).font = FONT_GREEN
    ws2.cell(row=r, column=4).alignment = ALIGN_C
    ws2.cell(row=r, column=5, value=sol).font = FONT_BODY
    ws2.cell(row=r, column=6, value=cost).font = FONT_GREEN
    ws2.cell(row=r, column=6).alignment = ALIGN_C
    ws2.cell(row=r, column=7, value=status).font = FONT_GREEN
    ws2.cell(row=r, column=7).alignment = ALIGN_C
    for c in range(1, MC2+1):
        ws2.cell(row=r, column=c).border = BORDER
    ws2.row_dimensions[r].height = 24
    r += 1

r += 1
r = sh(ws2, r, MC2, 'TIER 2: GENEROUS BUT LIMITED (HIT AT SCALE)')
r = ch(ws2, r, MC2, headers2)

generous = [
    ('Dolphin Anty Profiles', '10 profiles', '5+ personas', 'Add Incogniton (10 more free)', '$0', 'GENEROUS'),
    ('Beehiiv Subscribers', '2,500 per pub', '3 pubs scaling', 'Switch to Kit (10K free) or upgrade', '$49/mo', 'GENEROUS'),
    ('Kit Subscribers', '10,000 subs', 'Beyond 10K subs', 'Upgrade to Creator tier', '$29/mo', 'GENEROUS'),
    ('Email Volume', '~3,000/day (6 inboxes)', 'Heavy outbound', 'Add more inboxes via DeliverOn', '$49/mo', 'GENEROUS'),
    ('Vercel Bandwidth', '100GB/mo', 'Viral traffic spikes', 'Use Cloudflare Pages (unlimited)', '$0', 'GENEROUS'),
    ('Leonardo.ai Tokens', '150/day (~10 images)', '20+ images/day', 'Upgrade or use DALL-E free tier', '$12/mo', 'GENEROUS'),
    ('ElevenLabs Chars', '10K/mo (~10 min)', 'Heavy audio content', 'Upgrade to Starter', '$5/mo', 'GENEROUS'),
    ('Buffer Channels', '3 channels', '4+ platforms', 'Add Tweetlio (X) + manual YT', '$0', 'GENEROUS'),
    ('Carrd Sites', '3 free sites', '4+ brands/niches', 'Switch to Kit (unlimited LP) or Systeme.io', '$0', 'GENEROUS'),
]

for idx, (res, limit, bn, sol, cost, status) in enumerate(generous):
    fb(ws2, r, MC2, BG_WARN)
    ws2.cell(row=r, column=2, value=res).font = FONT_WHITE
    ws2.cell(row=r, column=3, value=limit).font = FONT_GOLD
    ws2.cell(row=r, column=4, value=bn).font = FONT_GOLD
    ws2.cell(row=r, column=4).alignment = ALIGN_C
    ws2.cell(row=r, column=5, value=sol).font = FONT_BODY
    ws2.cell(row=r, column=5).alignment = ALIGN_L
    ws2.cell(row=r, column=6, value=cost).font = FONT_ORANGE
    ws2.cell(row=r, column=6).alignment = ALIGN_C
    ws2.cell(row=r, column=7, value=status).font = FONT_GOLD
    ws2.cell(row=r, column=7).alignment = ALIGN_C
    for c in range(1, MC2+1):
        ws2.cell(row=r, column=c).border = BORDER
    ws2.row_dimensions[r].height = 28
    r += 1

r += 1
r = sh(ws2, r, MC2, 'TIER 3: HARD LIMITS (REQUIRES PAID UPGRADE)')
r = ch(ws2, r, MC2, headers2)

hard = [
    ('Google Workspace Email', 'No free tier', 'Immediately', 'Use free Gmail (500/day limit) or buy domains only', '$6/user/mo', 'HARD LIMIT'),
    ('Cold Email Tool', 'Trial only (14 days)', 'After trial ends', 'Manual sending via Gmail + TrulyInbox warmup', '$37-49/mo', 'HARD LIMIT'),
    ('Residential Proxies', 'Free datacenter only', 'Social media accounts', 'Soax Mobile ($50/mo) when farming accounts', '$50/mo', 'HARD LIMIT'),
    ('Account Marketplace', 'No free tier', 'Need pre-warmed accounts', 'Create accounts manually (slower but $0)', '$0.20-80/acct', 'HARD LIMIT'),
    ('Paid Ads', 'No free tier', 'Want paid acquisition', 'Organic only until $2K+ MRR', '$100+/mo', 'HARD LIMIT'),
    ('Premium Accounts (SWAPD)', 'No free tier', 'Want verified/aged accounts', 'Build organically first', '$500+', 'HARD LIMIT'),
]

for idx, (res, limit, bn, sol, cost, status) in enumerate(hard):
    fb(ws2, r, MC2, BG_LIMIT)
    ws2.cell(row=r, column=2, value=res).font = FONT_WHITE
    ws2.cell(row=r, column=3, value=limit).font = FONT_RED
    ws2.cell(row=r, column=4, value=bn).font = FONT_RED
    ws2.cell(row=r, column=4).alignment = ALIGN_C
    ws2.cell(row=r, column=5, value=sol).font = FONT_BODY
    ws2.cell(row=r, column=5).alignment = ALIGN_L
    ws2.cell(row=r, column=6, value=cost).font = FONT_RED
    ws2.cell(row=r, column=6).alignment = ALIGN_C
    ws2.cell(row=r, column=7, value=status).font = FONT_RED
    ws2.cell(row=r, column=7).alignment = ALIGN_C
    for c in range(1, MC2+1):
        ws2.cell(row=r, column=c).border = BORDER
    ws2.row_dimensions[r].height = 28
    r += 1


# ============================================================
# SHEET 3: 30-DAY LAUNCH SEQUENCE
# ============================================================
ws3 = wb.create_sheet('30-DAY LAUNCH')
ws3.sheet_properties.tabColor = '3FB950'
sw(ws3, [3, 10, 16, 38, 18, 20, 3])
MC3 = 7

for r in range(1, 80):
    fb(ws3, r, MC3, BG_DARK)

ws3.merge_cells('B2:F2')
ws3.cell(row=2, column=2, value='30-DAY $0 LAUNCH SEQUENCE').font = FONT_TITLE
ws3.row_dimensions[2].height = 32
ws3.merge_cells('B3:F3')
ws3.cell(row=3, column=2, value='Day-by-day deployment. Everything bootstrapped. Zero cost.').font = FONT_SUB

r = 5
headers3 = ['', 'DAY', 'CATEGORY', 'TASK', 'TOOL', 'COST']
r = ch(ws3, r, MC3, headers3)

launch = [
    # WEEK 1: Foundation
    ('WEEK 1: INFRASTRUCTURE', '', '', '', ''),
    ('Day 1', 'Anti-Detect', 'Create Dolphin Anty account. Set up 6 browser profiles (3 personas × 2 each).', 'Dolphin Anty', '$0'),
    ('Day 1', 'Email', 'Create 6 Gmail accounts (2 per niche). Start TrulyInbox warmup on all 6.', 'TrulyInbox + Gmail', '$0'),
    ('Day 2', 'Domains', 'Buy 6 .com domains on Porkbun ($9.73 each). Point all DNS to Cloudflare.', 'Porkbun + Cloudflare', '$58/yr'),
    ('Day 2', 'DNS', 'Set up SPF, DKIM, DMARC records for all 6 domains in Cloudflare.', 'Cloudflare', '$0'),
    ('Day 3', 'Newsletter', 'Create 3 Beehiiv publications. Configure welcome sequences. Add lead magnets.', 'Beehiiv', '$0'),
    ('Day 3', 'Landing Pages', 'Build 3 Carrd landing pages (1 per newsletter). Connect custom domains.', 'Carrd', '$0'),
    ('Day 4', 'Design', 'Create 3 lead magnet PDFs in Canva (one per niche). Brand kit setup.', 'Canva', '$0'),
    ('Day 4', 'Hosting', 'Claim Oracle Cloud free tier. Set up VPS. Install n8n + PostHog.', 'Oracle Cloud', '$0'),
    ('Day 5', 'Analytics', 'Set up GA4 for all 3 landing pages. Configure PostHog on Oracle VPS.', 'GA4 + PostHog', '$0'),
    ('Day 5', 'CRM', 'Create HubSpot free account. Import niche contact lists. Set up pipelines.', 'HubSpot', '$0'),
    ('Day 6-7', 'Social Accounts', 'Create 12 social accounts (4 per persona: X, TikTok, IG, YT). Complete all profiles 100%.', 'Dolphin Anty', '$0'),
    # WEEK 2: Content Foundation
    ('WEEK 2: CONTENT FOUNDATION', '', '', '', ''),
    ('Day 8-9', 'Content', 'Batch create 30 posts per niche in Canva (90 total). Schedule Buffer for week 3.', 'Canva + Buffer', '$0'),
    ('Day 8-9', 'Video', 'Record/edit 10 short videos per niche in CapCut (30 total). AI voiceover via ElevenLabs.', 'CapCut + ElevenLabs', '$0'),
    ('Day 10', 'AI Images', 'Generate persona photos + thumbnail templates in Leonardo.ai. Save brand presets.', 'Leonardo.ai', '$0'),
    ('Day 10', 'Newsletter', 'Write first 3 newsletter issues. Schedule for week 3. Set up Beehiiv automations.', 'Beehiiv', '$0'),
    ('Day 11-12', 'Email Warmup', 'Continue warmup (now Day 11). Aim for 10-15 warmup emails/day per inbox.', 'TrulyInbox', '$0'),
    ('Day 13-14', 'Monetization', 'Set up Gumroad (3 digital products). Connect Stripe. Create product pages.', 'Gumroad + Stripe', '$0'),
    # WEEK 3: Soft Launch
    ('WEEK 3: SOFT LAUNCH', '', '', '', ''),
    ('Day 15', 'Social Launch', 'First posts go live on all 12 accounts. Text-only, no links. Build algorithm trust.', 'Buffer + Tweetlio', '$0'),
    ('Day 16-18', 'Engagement', '50-100 manual engagements/day per account (likes, comments, follows). Use Dolphin profiles.', 'Dolphin Anty', '$0'),
    ('Day 19', 'Newsletter', 'Send first newsletter issues. Share signup links with warm contacts (20 per niche via DM).', 'Beehiiv', '$0'),
    ('Day 20-21', 'Content Ramp', 'Increase to 2 posts/day per account. Introduce video content (CapCut + Kling AI clips).', 'Buffer + Kling', '$0'),
    # WEEK 4: Growth
    ('WEEK 4: GROWTH MODE', '', '', '', ''),
    ('Day 22-25', 'Links + CTAs', 'Add bio links to all social profiles. Start linking to newsletter + lead magnets in posts.', 'All platforms', '$0'),
    ('Day 26-27', 'Cold Outreach', 'Email warmup complete (Day 26). Start cold outreach: 50 emails/day per inbox.', 'Gmail + TrulyInbox', '$0'),
    ('Day 28', 'Automation', 'Set up n8n workflows: auto-post → cross-platform. Newsletter → social promotion.', 'n8n (Oracle VPS)', '$0'),
    ('Day 29-30', 'Review', 'Audit analytics. Track subs, engagement, email open rates. Adjust content mix.', 'GA4 + PostHog', '$0'),
    ('Day 30', 'Revenue', 'Launch first paid product. Announce via newsletter + social. Track conversions.', 'Gumroad + Stripe', '$0'),
]

for idx, (day, cat, task, tool, cost) in enumerate(launch):
    if not cat and not task:  # Section header
        r = sh(ws3, r, MC3, day)
        continue
    bg = BG_ROW1 if idx % 2 == 0 else BG_ROW2
    if 'Launch' in task or 'Revenue' in task:
        bg = BG_FREE
    fb(ws3, r, MC3, bg)
    ws3.cell(row=r, column=2, value=day).font = FONT_CYAN
    ws3.cell(row=r, column=2).alignment = ALIGN_C
    ws3.cell(row=r, column=3, value=cat).font = FONT_GOLD
    ws3.cell(row=r, column=4, value=task).font = FONT_BODY
    ws3.cell(row=r, column=4).alignment = ALIGN_L
    ws3.cell(row=r, column=5, value=tool).font = FONT_WHITE
    ws3.cell(row=r, column=5).alignment = ALIGN_C
    ws3.cell(row=r, column=6, value=cost).font = FONT_GREEN if cost == '$0' else FONT_ORANGE
    ws3.cell(row=r, column=6).alignment = ALIGN_C
    for c in range(1, MC3+1):
        ws3.cell(row=r, column=c).border = BORDER
    ws3.row_dimensions[r].height = 32
    r += 1


# ============================================================
# SHEET 4: PLATFORM LIMITS CHEAT SHEET
# ============================================================
ws4 = wb.create_sheet('PLATFORM LIMITS')
ws4.sheet_properties.tabColor = 'F85149'
sw(ws4, [3, 18, 20, 20, 22, 24, 3])
MC4 = 7

for r in range(1, 50):
    fb(ws4, r, MC4, BG_DARK)

ws4.merge_cells('B2:F2')
ws4.cell(row=2, column=2, value='SOCIAL PLATFORM RATE LIMITS').font = FONT_TITLE
ws4.row_dimensions[2].height = 32
ws4.merge_cells('B3:F3')
ws4.cell(row=3, column=2, value='Stay under these limits to avoid bans. Per-account daily maximums.').font = FONT_SUB

r = 5
headers4 = ['', 'PLATFORM', 'SAFE DAILY ACTIONS', 'POSTING LIMIT', 'FOLLOW/LIKE LIMIT', 'BAN TRIGGERS']
r = ch(ws4, r, MC4, headers4)

limits = [
    ('X / Twitter', '300-500 total actions', 'Unlimited posts', '400 follows, 1000 likes', 'Aggressive follow/unfollow same day. Duplicate content. API spam.'),
    ('Instagram', '100-150 total actions', 'Unlimited posts + stories', '60 follows, 100 likes, 20 DMs', 'Action blocks from rapid liking. Same-IP multi-account. Mass DMs.'),
    ('TikTok', '50-100 total actions', '1-3 posts/day optimal', '30 follows, 50 likes', 'Too many accounts per device. Bot-like patterns. Duplicate content.'),
    ('YouTube', 'No strict action limits', '1-2 videos/day', 'Unlimited subs/likes', 'Copyright strikes. Spam comments. Misleading metadata.'),
    ('LinkedIn', '50-80 total actions', '1-2 posts/day', '100 connections/week', 'Automated messaging. Too many connections. Sales Navigator abuse.'),
    ('Reddit', '10-20 total actions', '5-10 comments/day', 'Karma-gated participation', 'Self-promotion ratio (10:1 rule). New account restrictions.'),
]

for idx, (plat, safe, post, follow, triggers) in enumerate(limits):
    bg = BG_ROW1 if idx % 2 == 0 else BG_ROW2
    fb(ws4, r, MC4, bg)
    ws4.cell(row=r, column=2, value=plat).font = FONT_WHITE
    ws4.cell(row=r, column=3, value=safe).font = FONT_GOLD
    ws4.cell(row=r, column=3).alignment = ALIGN_C
    ws4.cell(row=r, column=4, value=post).font = FONT_GREEN
    ws4.cell(row=r, column=4).alignment = ALIGN_C
    ws4.cell(row=r, column=5, value=follow).font = FONT_ORANGE
    ws4.cell(row=r, column=5).alignment = ALIGN_C
    ws4.cell(row=r, column=6, value=triggers).font = FONT_RED
    ws4.cell(row=r, column=6).alignment = ALIGN_L
    for c in range(1, MC4+1):
        ws4.cell(row=r, column=c).border = BORDER
    ws4.row_dimensions[r].height = 42
    r += 1

# Anti-detect protocol
r += 2
r = sh(ws4, r, MC4, 'ANTI-DETECT PROTOCOL — ACCOUNT SAFETY')
safety_rules = [
    ('1 Dolphin profile per persona', 'Never share profiles between personas. Fingerprint isolation prevents cross-contamination.'),
    ('Different proxy per profile', 'Residential > datacenter for social. Use Webshare free datacenter for non-social only.'),
    ('Warm up 7 days before posting', 'Browse, like, follow naturally for 7 days. Algorithm trusts aged + warmed accounts.'),
    ('Never login same account on 2 profiles', 'One account = one profile = one proxy = one fingerprint. No exceptions.'),
    ('Stagger post times by 2+ hours', 'Dont post to all accounts at same time. Looks automated. Buffer handles this.'),
    ('Use mobile user agents for IG/TikTok', 'These platforms flag desktop-only usage. Set Dolphin to mobile UA.'),
]

for idx, (rule, reason) in enumerate(safety_rules):
    bg = BG_ROW1 if idx % 2 == 0 else BG_ROW2
    fb(ws4, r, MC4, bg)
    ws4.merge_cells(f'B{r}:C{r}')
    ws4.cell(row=r, column=2, value=rule).font = FONT_GOLD
    ws4.cell(row=r, column=2).alignment = ALIGN_L
    ws4.merge_cells(f'D{r}:F{r}')
    ws4.cell(row=r, column=4, value=reason).font = FONT_BODY
    ws4.cell(row=r, column=4).alignment = ALIGN_L
    for c in range(1, MC4+1):
        ws4.cell(row=r, column=c).border = BORDER
    ws4.row_dimensions[r].height = 32
    r += 1


# ============================================================
# SHEET 5: REVENUE MILESTONES
# ============================================================
ws5 = wb.create_sheet('REVENUE MILESTONES')
ws5.sheet_properties.tabColor = 'F0C000'
sw(ws5, [3, 16, 20, 30, 24, 20, 3])
MC5 = 7

for r in range(1, 30):
    fb(ws5, r, MC5, BG_DARK)

ws5.merge_cells('B2:F2')
ws5.cell(row=2, column=2, value='REVENUE MILESTONES — WHEN TO UPGRADE').font = FONT_TITLE
ws5.row_dimensions[2].height = 32
ws5.merge_cells('B3:F3')
ws5.cell(row=3, column=2, value='Only spend money when revenue justifies it. Never upgrade speculatively.').font = FONT_SUB

r = 5
headers5 = ['', 'MILESTONE', 'TRIGGERED BY', 'WHAT TO UPGRADE', 'COST', 'EXPECTED ROI']
r = ch(ws5, r, MC5, headers5)

milestones = [
    ('500 newsletter subs', 'Organic growth from social', 'Nothing — Beehiiv Boost program unlocks ($1-3/sub earned)', '$0', 'Earn $500-1,500'),
    ('1,000 newsletter subs', 'Continued growth', 'Beehiiv Ad Network unlocks — sell ad inventory', '$0', '$200-500/mo ad revenue'),
    ('2,500 newsletter subs', 'Hit Beehiiv free cap', 'Upgrade Beehiiv to Growth ($49/mo) for custom domain', '$49/mo', 'Custom domain = 2x open rate'),
    ('$100-300 MRR', 'First paying customers', 'Buy Soax Mobile proxies ($50/mo) for IG/TikTok safety', '$50/mo', 'Prevent $500+ in account bans'),
    ('$500 MRR', 'Growing revenue', 'Add Publer ($12/mo) for 5 social channels. Add DistroKid ($23/yr) if music.', '$12/mo', '2x social output'),
    ('$1,000 MRR', 'Real business', 'Add Instantly.ai ($37/mo) for cold email. Add Google Workspace ($6/user/mo).', '$43/mo', '10-20x ROI on cold email'),
    ('$2,000 MRR', 'Scaling', 'Add Skool ($99/mo) for community. Consider Hetzner backup VPS ($4/mo).', '$103/mo', 'Recurring community revenue'),
    ('$5,000 MRR', 'Full operation', 'Meta/TikTok Ads ($100/mo test). Apollo leads ($79/mo). Expandi LinkedIn ($99/mo).', '$278/mo', '3-10x ROAS on paid ads'),
]

for idx, (milestone, trigger, upgrade, cost, roi) in enumerate(milestones):
    bg = BG_FREE if idx < 2 else (BG_WARN if idx < 5 else BG_ROW1 if idx % 2 == 0 else BG_ROW2)
    fb(ws5, r, MC5, bg)
    ws5.cell(row=r, column=2, value=milestone).font = FONT_WHITE
    ws5.cell(row=r, column=2).alignment = ALIGN_C
    ws5.cell(row=r, column=3, value=trigger).font = FONT_BODY
    ws5.cell(row=r, column=3).alignment = ALIGN_L
    ws5.cell(row=r, column=4, value=upgrade).font = FONT_BODY
    ws5.cell(row=r, column=4).alignment = ALIGN_L
    ws5.cell(row=r, column=5, value=cost).font = FONT_GREEN if cost == '$0' else FONT_ORANGE
    ws5.cell(row=r, column=5).alignment = ALIGN_C
    ws5.cell(row=r, column=6, value=roi).font = FONT_GOLD
    ws5.cell(row=r, column=6).alignment = ALIGN_C
    for c in range(1, MC5+1):
        ws5.cell(row=r, column=c).border = BORDER
    ws5.row_dimensions[r].height = 42
    r += 1


OUTPUT = '/sessions/awesome-nice-brown/mnt/PRINTMAXX_STARTER_KITttttt/PRINTMAXX_ZERO_COST_DEPLOYMENT.xlsx'
wb.save(OUTPUT)
print(f'Saved to {OUTPUT}')
