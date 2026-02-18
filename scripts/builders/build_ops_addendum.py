import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load existing workbook
wb = openpyxl.load_workbook('/sessions/awesome-nice-brown/mnt/PRINTMAXX_STARTER_KITttttt/PRINTMAXX_OPS_PLAYBOOK.xlsx')

# === THEME (match existing) ===
D = PatternFill('solid', fgColor='0D1117')
H = PatternFill('solid', fgColor='0A1628')
S = PatternFill('solid', fgColor='112240')
R1 = PatternFill('solid', fgColor='0D1117')
R2 = PatternFill('solid', fgColor='161B22')

FT = Font(name='Arial', size=20, bold=True, color='00D4FF')
FS = Font(name='Arial', size=12, bold=True, color='8B949E')
FH = Font(name='Arial', size=11, bold=True, color='00D4FF')
FSC = Font(name='Arial', size=12, bold=True, color='58A6FF')
FB = Font(name='Arial', size=10, color='C9D1D9')
FW = Font(name='Arial', size=10, bold=True, color='FFFFFF')
FC = Font(name='Arial', size=10, bold=True, color='00D4FF')
FG = Font(name='Arial', size=10, bold=True, color='3FB950')
FGO = Font(name='Arial', size=10, bold=True, color='F0C000')

THIN = Border(bottom=Side(style='thin', color='30363D'))

# ============================================================
# ADD NEW OPS TO "ALL OPS" SHEET
# ============================================================
ws1 = wb['ALL OPS']

# Find the last data row
last_row = ws1.max_row
# Skip to after the last OP entry (find first empty row after OP16)
insert_row = 22  # OP16 is at row 21, so OP17 starts at 22

new_ops = [
    ('OP17', 'FREELANCE SERVICE ARBITRAGE (CLAUDE CODE)', 'FREELANCE', '$1K-20K/mo', 'FULL AUTO (after intake)',
     'Claude Code Max = unlimited usage. List services on 10+ platforms (Fiverr, Upwork, Contra, LinkedIn, Reddit). Each gig takes 5-15 min to build, clients pay $50-500 for 2-3 day delivery. 95%+ margins. Volume play: 10-20 gigs/day = $500-5K/day. Pet project founders + small biz don\'t know Claude Code exists.',
     'Fiverr + Upwork + Contra + LinkedIn Services + Reddit r/forhire + PeoplePerHour + Freelancer.com + Guru + Twitter/X inbound',
     'FULL PLAYBOOK: See PRINTMAXX_FREELANCE_ARB.xlsx'),
    ('OP18', 'AI UGC VIDEO FACTORY', 'CONTENT', '$500-10K/mo', 'FULL AUTO',
     'HeyGen/Kling generate AI UGC videos at $0.50-2/video vs $50-500 human UGC creators. Sell to DTC brands, list on Billo/Insense as AI UGC creator. Batch produce 50-100 UGC ads/day. Sell packages: 10 UGC ads for $200-500.',
     'HeyGen ($29/mo) + Kling (free tier) + Claude (scripts) + Canva (overlays)',
     'Full pipeline: trending product → Claude writes script → HeyGen generates video → edit in CapCut → deliver to client or post to TikTok Shop'),
    ('OP19', 'MICRO-SAAS PORTFOLIO', 'SAAS', '$500-50K/mo portfolio', 'SEMI-AUTO',
     'Claude Code builds micro-SaaS tools in 24-48 hours. Target: $29-99/mo single-feature tools. Portfolio approach: build 10-20 tools, 5% hit rate = 1-2 winners. Each winner grows to $5K-50K MRR. Build → launch on ProductHunt → iterate.',
     'Claude Code + Vercel (free) + Supabase (free) + Stripe + ProductHunt',
     'Every micro-SaaS is an experiment. Build fast, launch fast, kill fast if no traction.'),
    ('OP20', 'COURSE/INFO PRODUCT FACTORY', 'INFO PRODUCTS', '$1K-50K/mo', 'SEMI-AUTO',
     'Claude writes full courses (50-100 lessons) in hours. Package as: Gumroad digital products ($19-97), Udemy courses ($12-50), Skool communities ($99/mo), Notion course vaults. Zero content creation cost.',
     'Claude (content) + Gumroad (sales, free) + Canva (visuals) + ElevenLabs (narration) + Udemy/Skillshare (distribution)',
     'Test with free lead magnet → upgrade to paid mini-course ($19) → full course ($97) → community ($99/mo). Each niche gets one course.'),
    ('OP21', 'LOCAL BIZ LEAD GEN MACHINE', 'AGENCY', '$2K-20K/mo', 'SEMI-AUTO',
     'Scrape local businesses with bad websites (Google Maps API). Claude generates redesign mockup. Cold email with before/after. Close at $500-2K/project. At scale: 100 emails/day → 2-5% response → 1-2 clients/week.',
     'Google Maps API (free tier) + Playwright (scraping) + Claude (mockups + email) + Instantly (cold email)',
     'Full pipeline already built: AUTOMATIONS/local_biz_pipeline.py. Just needs cold email warmup and execution.'),
    ('OP22', 'PINTEREST AFFILIATE AUTOMATION', 'AFFILIATE', '$200-5K/mo', 'FULL AUTO',
     'Pinterest has zero competition for AI-generated content. Create 50-100 pins/day with Claude-written descriptions + Canva templates. Link to Amazon affiliate, blog posts, or direct product pages. Pinterest traffic compounds over months.',
     'Canva (free) + Claude (descriptions + SEO) + Tailwind ($15/mo, auto-schedule) + Amazon Associates (free)',
     'Most underexploited platform for LLM automation. Pins have 6-month+ shelf life unlike social posts. Compounding traffic.'),
]

# Add to ALL OPS sheet
for i, op in enumerate(new_ops):
    row = insert_row + i
    # Insert row
    ws1.insert_rows(row)
    for j, val in enumerate(op, 1):
        c = ws1.cell(row=row, column=j, value=val)
        c.fill = R1 if i % 2 == 0 else R2
        c.border = THIN
        c.alignment = Alignment(wrap_text=True, vertical='top')
        if j == 1:
            c.font = FC
        elif j == 5:
            c.font = FG if 'FULL' in str(val) else FGO
        elif j == 6:
            c.font = FB
        else:
            c.font = FW

# ============================================================
# ADD TO DEEP PLAYBOOK
# ============================================================
ws2 = wb['DEEP PLAYBOOK']
last_row2 = ws2.max_row + 3

deep_ops = [
    {
        'id': 'OP17',
        'title': 'FREELANCE SERVICE ARBITRAGE (CLAUDE CODE MAX)',
        'revenue': '$1K-20K/mo',
        'sections': {
            'LLM ALPHA': 'Claude Code Max = near-unlimited usage for practical purposes. Services that take freelancers 2-5 days take Claude Code 5-60 minutes. List on 10+ platforms simultaneously. Every gig is 95%+ margin because the only cost is your flat monthly subscription. This is the purest form of LLM arbitrage available right now.',
            'INFRA STACK': 'Claude Code Max ($200/mo) + Fiverr (free) + Upwork (free) + Contra (free, 0% commission) + LinkedIn Services (free) + PeoplePerHour (free) + Reddit r/forhire (free) + Freelancer.com (free) + Guru.com (free). Total cost: $200/mo (Claude Max). Everything else is free.',
            'SETUP INSTRUCTIONS': [
                '1. Build 5 portfolio pieces with Claude Code: landing page, Chrome extension, Discord bot, web scraper, data dashboard. Takes 30min total.',
                '2. Create accounts on ALL platforms: Fiverr (5 gigs), Upwork (profile + 10 proposals), Contra (3 services), LinkedIn Services page, Reddit [For Hire] post.',
                '3. Set pricing at 25th-40th percentile. NOT cheapest (suspicious). NOT expensive (no reviews yet). Example: landing page $100, Chrome extension $200, script $75.',
                '4. Create 3-tier packages on Fiverr: Basic ($X) / Standard ($2X) / Premium ($3X). Most buy Standard. Premium is anchor.',
                '5. Offer unlimited revisions (costs you nothing with Claude Code). This eliminates buyer hesitation.',
                '6. Apply to 10 Upwork jobs/day for first 2 weeks. Use LLM to personalize each proposal from job description.',
                '7. First 5 gigs: offer 50% discount for 5-star reviews. Investment in algorithmic boost + social proof.',
                '8. After 10 reviews: raise prices 25%. Apply for Fiverr Pro. Add premium services.',
                '9. Build case study from each gig. Post on Twitter/LinkedIn: "Built a Chrome extension in 15 minutes for a client."',
                '10. Week 3+: Set up semi-automated pipeline. LLM reads order → Claude Code builds → you review (2-3 min) → deliver.'
            ],
            'ALGORITHM GUIDE': [
                'Fiverr Search Algorithm:',
                '- Response time matters: reply to all messages <1 hour. Fiverr tracks this.',
                '- Order completion rate: cancel = death. Never cancel, always deliver something.',
                '- Seller level: New → Level 1 (10 orders, 60 days) → Level 2 (50 orders) → Top Rated (100 orders). Each level = algorithmic boost.',
                '- Gig SEO: title keywords match buyer search. First 3 words most weighted. "I will" is mandatory on Fiverr.',
                '- Tags: use all 5. Mix broad and specific. "python script" + "web scraping" + "data extraction" + "automation" + "csv".',
                '- Video intros: 2-3x higher conversion. Use ElevenLabs or record yourself.',
                '',
                'Upwork Algorithm:',
                '- Job Success Score (JSS): most important metric. 90%+ JSS = Top Rated. Every bad review destroys months of work.',
                '- Rising Talent: new accounts get 30-day boost. Use this window to get first 5 reviews.',
                '- Connects: 16 free/month. Buy more ($0.15 each). Apply to jobs with <5 proposals first.',
                '- Profile completeness: 100% complete profile ranks higher. Add portfolio, skills test, video intro.',
                '- Specialized profiles: create one per category (Web Dev, Automation, Data). Separate proposal templates.',
            ],
            'SHADOWBAN / BAN AVOIDANCE': [
                'Fiverr: Never communicate off-platform (instant ban). Never ask for reviews (against TOS, ask naturally: "If satisfied, I\'d appreciate your feedback").',
                'Upwork: Never include contact info in proposals. Never suggest off-platform payment. Both = permanent ban.',
                'Delivery timing: NEVER deliver complex project in <4 hours. Wait minimum 4-6 hours even if done in 15 min. Instant delivery looks automated.',
                'Code variation: vary variable naming, file structure, comment style between deliveries. Same template every time = flaggable.',
                'Communication cadence: send progress updates even though done. "Working on it now" → 2hrs → "Testing" → deliver. Builds perceived value.',
                'Multiple platforms: different email per platform. Don\'t cross-link. Each platform thinks you\'re exclusive to them.',
            ],
            'LLM-IN-THE-LOOP (24/7)': 'Pipeline: order notification → LLM parses requirements (10 sec) → LLM drafts acknowledgment (10 sec) → Claude Code builds project (5-60 min) → auto-run QA checks (1 min) → human spot-check (2-3 min) → LLM drafts delivery message → human sends. Total human time per gig: 3-5 minutes. Claude runs follow-up cron: 48hr review request, 7-day maintenance upsell, case study auto-capture.',
            'MANUAL FIRST (VALIDATE)': 'Week 1-2: Manually deliver first 10 gigs. Learn platform quirks. Test pricing. Validate demand for each service type. Identify highest-converting services.',
            'AUTOMATE AFTER (SCALE)': 'Week 3+: LLM proposal writer for Upwork. Auto-ack messages. Claude Code pipeline for builds. Auto-follow-up cron. Case study auto-post. Target: 10-20 gigs/day at 5 min human time each = 1-2 hours/day for $500-5K/day revenue.',
        }
    },
    {
        'id': 'OP18',
        'title': 'AI UGC VIDEO FACTORY',
        'revenue': '$500-10K/mo',
        'sections': {
            'LLM ALPHA': 'Human UGC creators charge $50-500/video. AI UGC via HeyGen/Kling costs $0.50-2/video. You can produce 50-100 per day. DTC brands need 10-50 UGC ads per product for testing. This is pure cost arbitrage. List on Billo/Insense as a creator, or sell direct to brands via cold email.',
            'INFRA STACK': 'HeyGen ($29/mo, 15 min/mo video) + Kling (free tier) + Claude API (scripts) + CapCut (free, editing) + Canva (free, overlays) + Billo/Insense (marketplace listing)',
            'SETUP INSTRUCTIONS': [
                '1. Create HeyGen account ($29/mo). Choose 3-5 AI avatar personas that look authentic (avoid obviously AI faces).',
                '2. Claude writes UGC scripts: hook (3 sec) → problem (5 sec) → solution (10 sec) → CTA (3 sec). 15-30 seconds total.',
                '3. Batch produce: write 20 scripts at once with Claude, generate all 20 videos in HeyGen, edit batch in CapCut.',
                '4. List on Billo ($0 to join) and Insense as UGC creator. Set rate: $25-50/video (undercuts human creators by 5-10x).',
                '5. Cold email DTC brands: "I produce high-converting UGC ads. 10 videos for $200. 48-hour turnaround." Include 3 sample videos.',
                '6. For your own TikTok Shop: use AI UGC to promote affiliate products. Commission 10-30% per sale.',
            ],
            'ALGORITHM GUIDE': [
                'TikTok UGC performance factors:',
                '- Hook in first 1 second (text hook overlay + face). Average watch time >3 sec = algorithmic push.',
                '- Raw/authentic style outperforms polished. Add slight camera shake, natural lighting, casual tone.',
                '- Product in hand by second 3. Don\'t wait. Attention span is 1.5 seconds average.',
                '- CTA: "link in bio" or "TikTok Shop" orange button. Don\'t say "buy" say "I got mine here".',
                '- Post 3-5x/day. Volume matters more than quality for testing. Winners get boosted.',
            ],
            'SHADOWBAN / BAN AVOIDANCE': [
                'Disclose AI: FTC requires disclosure. Add "#AIgenerated" or similar. Most platforms don\'t penalize AI UGC for ads.',
                'Don\'t claim human: never say "this is my honest review" with AI avatar. Say "here\'s what makes this product great."',
                'Billo/Insense: some marketplaces don\'t allow AI UGC yet. Check TOS. Use on platforms that explicitly allow it.',
                'Vary avatars: don\'t use same face for competing brands. Rotate 3-5 personas.',
            ],
            'LLM-IN-THE-LOOP (24/7)': 'Claude API monitors trending products (TikTok Shop, Amazon Movers). Auto-generates 20 scripts per trending product. HeyGen API batch produces videos. CapCut templates auto-apply branding. Human reviews final batch (2 min per 20 videos).',
            'MANUAL FIRST (VALIDATE)': 'Week 1: Create 20 UGC videos manually. Test on TikTok + send to 5 brands. Validate: conversion rates, brand response, cost per video.',
            'AUTOMATE AFTER (SCALE)': 'Week 3+: Full pipeline. Claude → HeyGen → CapCut → post/deliver. 50-100 videos/day. Revenue from: direct sales to brands + TikTok Shop commissions + marketplace gigs.',
        }
    },
    {
        'id': 'OP19',
        'title': 'MICRO-SAAS PORTFOLIO BUILDER',
        'revenue': '$500-50K/mo portfolio',
        'sections': {
            'LLM ALPHA': 'Traditional SaaS takes 3-6 months to build. Claude Code builds MVPs in 24-48 hours. Portfolio approach: build 10-20 micro-tools, launch each on ProductHunt. 5-10% hit rate = 1-2 winners. Each winner can reach $5K-50K MRR. The edge: volume of experiments at near-zero marginal cost.',
            'INFRA STACK': 'Claude Code Max ($200/mo) + Vercel (free tier, 100 deploys/day) + Supabase (free tier, 500MB) + Stripe ($0 until revenue) + ProductHunt (free) + LemonSqueezy ($0 until revenue)',
            'SETUP INSTRUCTIONS': [
                '1. Identify micro-SaaS ideas: one-feature tools that solve one specific pain point. Examples: PDF merger, image compressor, invoice generator, QR code maker, email signature builder.',
                '2. Use Claude Code to build MVP in 24-48 hours. Next.js + Supabase + Stripe. Single page app.',
                '3. Deploy to Vercel (free). Custom domain via Cloudflare ($8-12/yr).',
                '4. Launch on ProductHunt (Tuesday-Thursday, 12:01 AM PT). Have 10 friends upvote in first hour.',
                '5. Add analytics (Plausible, free tier). Monitor usage. If tool gets 100+ daily users in first week → invest more.',
                '6. Monetize: freemium with $9-29/mo pro tier. Or one-time $29-97 payment via LemonSqueezy.',
                '7. If no traction after 30 days → archive, move to next idea. Don\'t fall in love with duds.',
                '8. Repeat: build 2-3 per month. By month 6: 12-18 tools launched, 1-3 generating revenue.',
            ],
            'ALGORITHM GUIDE': [
                'ProductHunt Algorithm:',
                '- Launch day: Tuesday-Thursday best. Avoid Monday (competition) and Friday (low traffic).',
                '- Time: 12:01 AM PT. First 4 hours are critical for ranking.',
                '- Hunter: get a known hunter to post your product. Their followers see it immediately.',
                '- First comment: post detailed maker comment explaining the tool, your story, ask for feedback.',
                '- Engage: reply to EVERY comment within 1 hour. Engagement signals rank your product higher.',
                '',
                'SEO for micro-SaaS:',
                '- Target "[thing] tool" and "[thing] generator" keywords. Example: "pdf merger" gets 165K searches/mo.',
                '- Simple landing page: hero → demo → pricing → FAQ. That\'s it.',
                '- Add blog with 5-10 SEO articles about the problem your tool solves.',
            ],
            'SHADOWBAN / BAN AVOIDANCE': [
                'ProductHunt: don\'t ask for upvotes explicitly (against TOS). Say "check out my launch, would love feedback."',
                'Don\'t launch same idea twice on PH. If first launch flops, rebrand and pivot the angle.',
                'Vercel: stay within free tier limits (100GB bandwidth). Add caching headers. Upgrade if needed ($20/mo).',
            ],
            'LLM-IN-THE-LOOP (24/7)': 'Claude Code builds entire MVPs. Claude API writes landing page copy + blog posts + email sequences. Claude monitors competitor pricing + feature gaps. Automated: build → deploy → launch → monitor → iterate.',
            'MANUAL FIRST (VALIDATE)': 'Build first 3 tools manually (with Claude Code). Learn the deploy → launch → market cycle. Validate which niches respond.',
            'AUTOMATE AFTER (SCALE)': 'Month 2+: Templated build pipeline. Claude Code starts from boilerplate → customizes for each idea. Semi-automated PH launches. Aim: 2-3 launches/month.',
        }
    },
    {
        'id': 'OP20',
        'title': 'COURSE / INFO PRODUCT FACTORY',
        'revenue': '$1K-50K/mo',
        'sections': {
            'LLM ALPHA': 'Human course creators spend 100-500 hours per course. Claude writes full 50-100 lesson courses in 2-4 hours of prompting. Package as: Gumroad digital downloads ($19-97), Udemy courses ($12-50, their traffic), Skool communities ($99/mo recurring). The arbitrage: content production cost approaches $0 while value to student remains high.',
            'INFRA STACK': 'Claude (content) + Gumroad ($0 until sale, 5% fee) + Udemy (free, 63% rev share on their traffic) + Skool ($99/mo) + Canva (visuals, free) + ElevenLabs ($5/mo for narration) + Descript ($24/mo for video editing)',
            'SETUP INSTRUCTIONS': [
                '1. Pick your first 3 course topics from high-demand, low-competition areas. Check Udemy: courses with 1,000+ students and <20 competitors.',
                '2. Claude writes full course outline: 8-12 modules, 5-10 lessons each. Include exercises, quizzes, action items.',
                '3. Claude generates all lesson content. For text courses: PDF/Notion delivery. For video courses: Claude writes scripts.',
                '4. For video: ElevenLabs narrates scripts. Use Canva for slides. Descript for editing. Or use Loom with voiceover.',
                '5. Upload to Gumroad ($0 cost): price at $19 for mini-courses, $47-97 for full courses.',
                '6. Upload to Udemy: they bring traffic (marketplace of 40M+ learners). Lower price but free distribution.',
                '7. Create free lead magnet (first 3 lessons free) → email funnel → paid course upsell.',
                '8. Launch Skool community ($99/mo) for premium access: all courses + Q&A + group calls.',
            ],
            'ALGORITHM GUIDE': [
                'Udemy Algorithm: enrollment velocity in first 7 days matters most. Price low ($12.99) initially for volume. Udemy promotes courses with 4.0+ rating and 50+ reviews.',
                'Gumroad: SEO your product page title. "Python for Data Science: Complete 2026 Guide" beats "Learn Python". Price anchoring works: show $297 crossed out, $47 actual.',
                'Skool: community engagement = retention. Post daily challenge or prompt. Members who post in first 48 hours retain 5x longer.',
            ],
            'SHADOWBAN / BAN AVOIDANCE': [
                'Udemy: don\'t say "AI-generated content." Students expect human expertise. Add personal insights and real-world examples.',
                'Quality gate: always review and edit AI-generated content. Add your perspective. Minimum 20% human value-add per lesson.',
                'Avoid: making false income claims in marketing. FTC requires disclaimers for any earnings representations.',
            ],
            'LLM-IN-THE-LOOP (24/7)': 'Claude generates course outlines from trending topic detection. Writes all lesson content. ElevenLabs narrates. Auto-generate Gumroad product pages. Auto-write email sequences for funnel. Human reviews final product before listing.',
            'MANUAL FIRST (VALIDATE)': 'Create first course manually (with Claude writing assist). Validate: sales, completion rate, reviews. Identify winning topic + format.',
            'AUTOMATE AFTER (SCALE)': 'Month 2+: 1-2 new courses/month across different niches. Automated pipeline: trend detection → outline → content → narration → listing → email funnel. Each course = perpetual passive income.',
        }
    },
    {
        'id': 'OP21',
        'title': 'LOCAL BIZ LEAD GEN MACHINE',
        'revenue': '$2K-20K/mo',
        'sections': {
            'LLM ALPHA': 'Millions of local businesses have terrible websites. They don\'t know a redesign can be done in 30 minutes. Scrape Google Maps → find businesses with outdated sites → Claude generates beautiful mockup → cold email with before/after comparison → close at $500-2K. At 100 emails/day, 2-5% response, 10-20% close = 1-3 clients/week.',
            'INFRA STACK': 'Google Maps API (free 28K calls/mo) + Playwright (free, scraping) + Claude (mockup + email) + Instantly ($30/mo, cold email) + Cloudflare Pages (free, deploy mockups)',
            'SETUP INSTRUCTIONS': [
                '1. Pipeline already built: AUTOMATIONS/local_biz_pipeline.py. It scrapes → analyzes → generates page → cold emails.',
                '2. Set up Instantly ($30/mo): warm 3 email accounts for 2 weeks before sending.',
                '3. Target: restaurants, dentists, plumbers, law offices, salons in cities with 50K-200K population.',
                '4. Claude generates one-page redesign deployed to Cloudflare Pages. Include their actual business info (phone, address, hours).',
                '5. Cold email: "I noticed your website doesn\'t match the quality of your business. I built a concept for you [link]. Free consultation?"',
                '6. Close on call: $500 for the page you already built + $200/mo maintenance. Upsell: SEO ($300/mo), Google Ads management ($300/mo).',
                '7. Deliver the page you already built (took 5 minutes). Offer additional pages at $100-200/page.',
                '8. Scale: VA handles initial outreach + scheduling. You handle closes + Claude Code delivers.',
            ],
            'ALGORITHM GUIDE': [
                'Cold email deliverability: warm emails 2 weeks before first campaign. Send 20/day per inbox, max 50. Ramp slowly.',
                'Subject lines: personalization = open rate. "{Business Name} website upgrade concept" converts 15-25%.',
                'Google Maps scraping: filter by rating (3.5-4.5 stars = active biz, care about reputation, but not perfect). Avoid 5-star (probably already invested in marketing).',
                'Best niches by close rate: dentists > plumbers > restaurants > lawyers > salons. Medical/dental pay most ($2K-5K).',
            ],
            'SHADOWBAN / BAN AVOIDANCE': [
                'CAN-SPAM compliance: include unsubscribe link, physical address, don\'t mislead subject line.',
                'Don\'t scrape Google Maps too aggressively. Rate limit to 1 req/sec. Rotate IPs if doing 10K+ lookups.',
                'Never claim existing relationship. Say "I noticed" not "as discussed."',
            ],
            'LLM-IN-THE-LOOP (24/7)': 'Daily cron: scrape 100 local businesses → Claude scores website quality (1-10) → filter sites scoring <5 → Claude generates redesign mockup → deploy to Cloudflare → Claude writes personalized cold email → queue in Instantly. Human: review emails before send (5 min/day), take sales calls.',
            'MANUAL FIRST (VALIDATE)': 'Week 1: Manually email 50 local businesses. Take calls. Close 1-2 deals. Validate pricing + close rate.',
            'AUTOMATE AFTER (SCALE)': 'Week 3+: Full pipeline automated. 100 emails/day. VA handles scheduling. You close deals and Claude Code delivers. Target: 2-5 new clients/week = $4K-40K/month.',
        }
    },
    {
        'id': 'OP22',
        'title': 'PINTEREST AFFILIATE AUTOMATION',
        'revenue': '$200-5K/mo',
        'sections': {
            'LLM ALPHA': 'Pinterest is the most underexploited platform for LLM automation. Pins have 6-12 month shelf life (vs 24 hours on Twitter/TikTok). Create 50-100 pins/day with Claude-written descriptions + Canva templates. Link to Amazon affiliate products, blog posts, or direct product pages. Traffic compounds over months. Most affiliate marketers ignore Pinterest = low competition.',
            'INFRA STACK': 'Canva (free) + Claude (descriptions + SEO) + Tailwind ($15/mo, auto-schedule) + Amazon Associates (free, 1-10% commission) + ShareASale (free) + Cloudflare Pages (free, for blog)',
            'SETUP INSTRUCTIONS': [
                '1. Create Pinterest business account ($0). Set up 10 niche boards (fitness, tech, home, cooking, fashion, etc).',
                '2. Create 5 Canva pin templates per niche. Standard sizes: 1000x1500px. Clean, text overlay, product image.',
                '3. Claude writes pin descriptions: 100-500 chars, include 3-5 relevant keywords, include CTA, include hashtags.',
                '4. Link pins to: Amazon affiliate links (direct), blog posts with affiliate links embedded, Gumroad product pages.',
                '5. Use Tailwind to schedule: 10-15 pins/day. Best times: 8-11 PM, 2-4 PM. Spread across boards.',
                '6. Create "idea pins" (multi-slide): Claude writes the content, Canva creates slides. Higher engagement than standard pins.',
                '7. After 30 days: analyze top-performing pins. Double down on winning niches + product categories.',
                '8. Scale: 50-100 pins/day across multiple niche accounts. Use different emails per account.',
            ],
            'ALGORITHM GUIDE': [
                'Pinterest Algorithm (2026):',
                '- Fresh pins (new images) rank higher than repins. Create new visuals, don\'t just repin.',
                '- Keyword SEO: Pinterest is a search engine. Pin title + description + board name all matter.',
                '- Idea pins get 5-10x more distribution than standard pins. Multi-slide format preferred.',
                '- Consistency > volume: 10-15 pins/day beats 100 pins on Monday and 0 rest of week.',
                '- Rich pins (pull metadata from your site) rank higher. Set up on your blog.',
                '- Click-through rate matters. Pins that get clicks → get more distribution.',
            ],
            'SHADOWBAN / BAN AVOIDANCE': [
                'Don\'t use shortened URLs (bit.ly etc). Pinterest penalizes them. Use direct URLs.',
                'Don\'t pin same URL to too many boards. Max 3-5 boards per URL.',
                'Affiliate link disclosure: add #affiliate or #ad to pin description. FTC requirement.',
                'Don\'t spam: if every pin goes to same site, Pinterest limits distribution. Mix in other content.',
                'New account warmup: 5-10 pins/day for first 2 weeks. Ramp to 15-20. Never jump to 100.',
            ],
            'LLM-IN-THE-LOOP (24/7)': 'Daily cron: Claude identifies trending products on Amazon (Movers & Shakers) → writes pin descriptions with SEO keywords → Canva API generates pin images from template → Tailwind API schedules → track affiliate conversions. Fully automated after template creation.',
            'MANUAL FIRST (VALIDATE)': 'Week 1-2: Create 100 pins manually across 5 niches. Track: impressions, clicks, affiliate clicks. Identify winning niches.',
            'AUTOMATE AFTER (SCALE)': 'Week 3+: Full automation. Canva templates + Claude descriptions + Tailwind scheduling. 50-100 pins/day. Revenue compounds as old pins continue driving traffic. Month 6: $500-2K/mo passive.',
        }
    },
]

# Write deep playbook entries
for op in deep_ops:
    r = last_row2
    # Title row (merged B:C)
    ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    c = ws2.cell(row=r, column=2, value=f"{op['id']}: {op['title']} — {op['revenue']}")
    c.font = Font(name='Arial', size=12, bold=True, color='00D4FF')
    c.fill = S
    c.alignment = Alignment(wrap_text=True, vertical='top')
    r += 1

    for section_name, section_content in op['sections'].items():
        # Section header
        c = ws2.cell(row=r, column=2, value=section_name)
        c.font = Font(name='Arial', size=10, bold=True, color='F0C000')
        c.fill = H
        r += 1

        # Section content
        if isinstance(section_content, list):
            for line in section_content:
                c = ws2.cell(row=r, column=2, value=line)
                c.font = FB
                c.fill = R1
                c.alignment = Alignment(wrap_text=True, vertical='top')
                r += 1
        else:
            c = ws2.cell(row=r, column=2, value=section_content)
            c.font = FB
            c.fill = R1
            c.alignment = Alignment(wrap_text=True, vertical='top')
            r += 1

    # Spacing
    r += 2
    last_row2 = r

# ============================================================
# ADD TO LLM ALPHA THESIS
# ============================================================
ws3 = wb['LLM ALPHA THESIS']
last_row3 = ws3.max_row

new_alpha = [
    ('Freelance service arbitrage (Claude Code)', 'Freelancers spend 2-5 days on $50-500 projects. Claude Code does it in 5-60 min. 95%+ margin on every gig. Volume play: 10-20 gigs/day.', '24-48 months', 'Pet project founders + small biz don\'t know Claude Code exists. As AI awareness grows, some shift to DIY. But custom work + support + reliability = durable moat.'),
    ('AI UGC video production at scale', 'Human UGC: $50-500/video. AI UGC: $0.50-2/video. 100x cost reduction. Sell to DTC brands at 5-10x markup. Volume: 50-100 videos/day.', '12-18 months', 'HeyGen/Kling quality improving fast. Window to establish brand + client relationships before market saturates.'),
    ('Micro-SaaS portfolio via Claude Code', 'Traditional dev: 3-6 months per SaaS. Claude Code: 24-48 hours per MVP. Portfolio approach: build 20 tools, 2-3 become winners.', '24-36 months', 'Building software remains hard even with AI. The edge: speed of iteration + volume of experiments. Coding AI improves but so does your ability to use it.'),
    ('Course/info product factory', 'Human course creation: 100-500 hours. Claude: 2-4 hours. Production cost approaches $0. Distribution via Udemy (40M learners) + Gumroad.', '12-24 months', 'AI-generated course quality is visibly improving. Early movers build brand + review count. Eventually: AI courses become commodity. Win now.'),
    ('Local biz website redesign pipeline', 'Manual prospecting: 10 businesses/day. Automated: 100/day with mockup already built. Claude generates redesign before first contact.', '24-36 months', 'Local businesses are slowest to adopt AI. This edge lasts longer than consumer/tech markets. Most local biz still don\'t know what a landing page builder is.'),
    ('Pinterest affiliate compound machine', 'Most affiliate marketers ignore Pinterest. Pins have 6-12 month shelf life. LLM generates 50-100 optimized pins/day. Traffic compounds.', '18-24 months', 'Pinterest is a search engine with less competition than Google. AI-generated pins at scale = compounding traffic asset.'),
]

insert_row3 = last_row3 - 10  # Insert before the thesis summary section
for i, (opp, why, duration, notes) in enumerate(new_alpha):
    row = 23 + i  # After existing entries
    ws3.insert_rows(row)
    for j, val in enumerate([opp, why, duration, notes], 1):
        c = ws3.cell(row=row, column=j, value=val)
        c.fill = R1 if i % 2 == 0 else R2
        c.font = FB
        c.alignment = Alignment(wrap_text=True, vertical='top')
        c.border = THIN

# Update op count in title
for row in ws1.iter_rows(min_row=2, max_row=3, max_col=8):
    for cell in row:
        if cell.value and '16 ops' in str(cell.value):
            cell.value = str(cell.value).replace('16 ops', '22 ops')

# SAVE
out = '/sessions/awesome-nice-brown/mnt/PRINTMAXX_STARTER_KITttttt/PRINTMAXX_OPS_PLAYBOOK.xlsx'
wb.save(out)
print(f'Saved updated playbook to {out}')
print(f'Added ops: OP17-OP22')
print(f'Added alpha thesis entries: {len(new_alpha)}')
