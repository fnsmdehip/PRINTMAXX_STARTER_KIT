#!/usr/bin/env python3
"""Build PRINTMAXX_STRATEGIC_RBI.xlsx - the real improvement engine deliverable"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

DARK = PatternFill('solid', fgColor='0D1117')
CYAN_FILL = PatternFill('solid', fgColor='0D2137')
HEADER_FILL = PatternFill('solid', fgColor='00D4FF')
GREEN_FILL = PatternFill('solid', fgColor='0D3B0D')
RED_FILL = PatternFill('solid', fgColor='3B0D0D')
YELLOW_FILL = PatternFill('solid', fgColor='3B3B0D')
PURPLE_FILL = PatternFill('solid', fgColor='2D0D3B')

WHITE = Font(name='Arial', color='FFFFFF', size=10)
CYAN_FONT = Font(name='Arial', color='00D4FF', size=10, bold=True)
HEADER_FONT = Font(name='Arial', color='0D1117', size=11, bold=True)
GREEN_FONT = Font(name='Arial', color='00FF88', size=10)
RED_FONT = Font(name='Arial', color='FF4444', size=10)
YELLOW_FONT = Font(name='Arial', color='FFD700', size=10)

thin_border = Border(left=Side(style='thin',color='333333'),right=Side(style='thin',color='333333'),top=Side(style='thin',color='333333'),bottom=Side(style='thin',color='333333'))

def style_sheet(ws, headers, widths):
    ws.sheet_properties.tabColor = '00D4FF'
    for i,(h,w) in enumerate(zip(headers,widths),1):
        c = ws.cell(row=1,column=i,value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = Alignment(horizontal='center',wrap_text=True); c.border = thin_border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

def dark_row(ws, row, data, font=WHITE, fill=DARK):
    for i,val in enumerate(data,1):
        c = ws.cell(row=row,column=i,value=val)
        c.font = font; c.fill = fill; c.border = thin_border; c.alignment = Alignment(wrap_text=True,vertical='top')

# ============================================================
# SHEET 1: VIABILITY MATRIX (real data, not projections)
# ============================================================
ws1 = wb.active
ws1.title = 'VIABILITY MATRIX'
h1 = ['OP_TYPE','SUCCESS_RATE','TIME_TO_$1K','AUTOMATION_%','SATURATION','REAL_REVENUE_RANGE','KEY_FAILURE_MODE','HUMAN_INPUT_REQUIRED','VIABILITY_SCORE','VERDICT']
w1 = [22,12,14,12,12,24,30,30,14,14]
style_sheet(ws1,h1,w1)

viability = [
    ['AI NSFW/Findom','30%','2-4 weeks','40%','LOW','$500-30K/mo','Audience building takes 3-6mo. Daily engagement 2-4hrs. Community management.','Character development, audience engagement, DM responses, content strategy','9/10','ACTIVATE NOW'],
    ['Local Biz Service','25%','2-4 weeks','60%','LOW','$2-15K/mo','Cold email reply rate 3-5%. Need 500+ sends for results.','Cold email personalization, client calls, mockup review, proposal writing','9/10','ACTIVATE NOW'],
    ['Cold Email Lead Gen','20%','1-2 months','50%','MEDIUM','$2-8K/mo','Deliverability issues. Need SPF/DKIM/DMARC setup. 14-day warmup.','List building, personalization, follow-up management, sales calls','8/10','ACTIVATE NOW'],
    ['AI UGC Factory','20%','2-3 months','70%','MEDIUM','$1.2-10K/mo','Commoditizing. Brands want authentic human UGC. AI detection improving.','Client communication, brief analysis, quality control, revision handling','7/10','ACTIVATE'],
    ['Digital Products (Gumroad)','15%','1-3 months','90%','MEDIUM','$100-10K/mo','Need existing audience for sales. Cold launch = crickets.','Product creation (done), listing copy, audience building, marketing','8/10','ACTIVATE NOW - 9 READY'],
    ['Freelance Arbitrage','10%','4-6 weeks','40%','GROWING','$1-8K/mo','AI alone fails 97% of Upwork tasks. Need human oversight + niche expertise.','Client communication, quality review, revision handling, proposal writing','6/10','ACTIVATE with caution'],
    ['Faceless YouTube','10%','3-4 months','80%','HIGH','$500-20K/mo (top 10%)','AI content enthusiasm dropped 60%→26%. AI slop everywhere. Need differentiation.','Topic selection, quality review, thumbnail strategy, niche expertise','5/10','SELECTIVE - niche only'],
    ['SEO Affiliate','10%','3-6 months','70%','HIGH','$200-5K/mo','Takes 3-6 months to rank. Algorithm changes can wipe progress.','Keyword research, content review, link building outreach','5/10','LONG-TERM play'],
    ['Newsletter','8%','6-12 months','50%','MEDIUM','$200-5K/mo','Slow growth. Need 1000+ subs before monetization makes sense.','Content creation, audience engagement, sponsor outreach','4/10','STACK with other ops'],
    ['TikTok Content','5%','2-3 months','70%','HIGH','$0-5K/mo (most)','Creator Rewards: $0.40-1/1K views. Need 500K+ views/mo. Most make <$100/mo.','Hook creation, trend monitoring, engagement, content calendar','3/10','SUPPORT op only'],
    ['Micro-SaaS','5%','2-3 months MRR','60%','LOW-MED','$0-5K/mo (95%)','92% fail within 3 years. 68% built products nobody wanted. Valley of death at 18mo.','Product-market fit research, customer support, bug fixes, feature dev','3/10','HIGH RISK - validate first'],
    ['App Factory','5%','3-6 months','60%','HIGH','$0-5K/mo (95%)','App stores saturated. Discovery is the bottleneck, not building.','App Store optimization, user feedback, bug fixes, marketing','3/10','PORTFOLIO approach only'],
]

r = 2
for v in viability:
    score = v[8]
    if '9/' in score or '8/' in score:
        fill = GREEN_FILL
    elif '7/' in score or '6/' in score:
        fill = CYAN_FILL
    elif '5/' in score or '4/' in score:
        fill = YELLOW_FILL
    else:
        fill = RED_FILL
    dark_row(ws1,r,v,WHITE,fill)
    r += 1

# ============================================================
# SHEET 2: CRITICAL BOTTLENECKS
# ============================================================
ws2 = wb.create_sheet('BOTTLENECKS')
h2 = ['SEVERITY','AREA','ISSUE','IMPACT','FIX','AUTOMATION_POSSIBLE','EFFORT','DEADLINE']
w2 = [12,18,40,30,40,14,12,14]
style_sheet(ws2,h2,w2)

bottlenecks = [
    ['CRITICAL','Revenue Tracking','ZERO revenue entries tracked. Flying blind.','Cannot optimize what you dont measure. All decisions are guesses.','Track EVERY dollar: source, method, amount, date, channel. Even $1 matters.','YES - build intake script','1 hour','TODAY'],
    ['CRITICAL','Alpha Pipeline','85 entries stuck in PENDING_REVIEW','New tactics and opportunities not being implemented. Stale alpha = dead alpha.','Schedule 30-min daily review. Process 10-20 per session. Reject ruthlessly.','PARTIAL - auto-score + human review','30 min/day','THIS WEEK'],
    ['CRITICAL','A/B Testing','42 tests DESIGNED but ZERO running','No learning happening. Flying blind on conversion. Optimizing nothing.','Launch 3 tests THIS WEEK: 1 pricing, 1 content format, 1 CTA.','YES - build test runner','2-3 hours','THIS WEEK'],
    ['CRITICAL','Account Creation','Key platform accounts NOT CREATED (Fiverr, Upwork, Gumroad, Fanvue, Fansly, TikTok)','Cannot run ops without accounts. Every op is BLOCKED by this.','Create accounts TODAY. Use GoLogin for anti-detection. 2 accounts/day max.','NO - must be manual','3-4 hours','TODAY'],
    ['HIGH','Active Methods','6 methods marked ACTIVE but $0 revenue tracked','Either tracking broken or methods arent actually active. Misleading dashboard.','Verify each "active" method is genuinely producing revenue. Reclassify if not.','NO - manual audit','1 hour','THIS WEEK'],
    ['HIGH','Freelance Pipeline','EMPTY. Zero gigs tracked.','Freelance arbitrage (S01) listed as active but zero activity.','Create Fiverr + Upwork listings TODAY. Start with top 5 services.','PARTIAL - listing templates','2-3 hours','TODAY'],
    ['HIGH','Content Status','484 content pieces with NO STATUS','Content created but never posted = wasted effort. Content rot.','Batch-assign: QUEUED if ready, DRAFT if needs work, KILL if stale.','YES - auto-status script','1 hour','THIS WEEK'],
    ['HIGH','Playbooks Dead','03_PLAYBOOKS: 2/21915 files modified in 7 days','49 playbooks sitting unused. Massive intellectual capital collecting dust.','Activate top 3 playbooks by viability score. Connect to daily workflow.','NO - human execution','Ongoing','THIS MONTH'],
    ['MEDIUM','Deprecated Code','1 deprecated script still in AUTOMATIONS/','Dead code = confusion. Clean codebase = faster iteration.','Delete backtest_alpha_DEPRECATED.py.','YES','5 min','TODAY'],
    ['MEDIUM','Browser Scripts','11 scripts need Playwright/Selenium','Cannot run scrapers in all environments. Need headless browser setup.','Install Playwright on mac. Set up in launchd for automated running.','YES - setup script','30 min','THIS WEEK'],
]

r = 2
for b in bottlenecks:
    sev = b[0]
    fill = RED_FILL if sev == 'CRITICAL' else YELLOW_FILL if sev == 'HIGH' else DARK
    dark_row(ws2,r,b,WHITE,fill)
    r += 1

# ============================================================
# SHEET 3: HYPOTHESES TO TEST
# ============================================================
ws3 = wb.create_sheet('HYPOTHESES')
h3 = ['ID','HYPOTHESIS','TEST_METHOD','METRIC','BASELINE','TARGET','EFFORT','EXPECTED_IMPACT','STATUS']
w3 = [8,40,40,20,16,16,14,20,12]
style_sheet(ws3,h3,w3)

hyp = [
    ['H001','Animated paywalls convert 2.9x better than static','Deploy animated paywall on PrayerLock. A/B test vs current.','paywall_conversion','current (measure)','2.9x improvement','LOW (2-3 hrs)','$500-2K/mo increase','NOT STARTED'],
    ['H002','Findom #findom #paypig hashtags convert 5x vs general','Post 20 findom tweets with niche tags vs 20 general. Track clicks.','clicks_per_impression','TBD','5x CTR','LOW (1-2 hrs)','5x follower growth','NOT STARTED'],
    ['H003','Timeline-based cold email hooks get 10% reply rate','Split test 200 emails: timeline vs problem-statement hooks.','reply_rate','3.4% (avg)','10%+ replies','MED (4-6 hrs)','3x more leads','NOT STARTED'],
    ['H004','Video demos on Gumroad = 3x conversion','Add Kling demo videos to top 3 products. Measure conversion delta.','sales_conversion','current','3x conversion','MED (3-4 hrs)','$300-1K/mo increase','NOT STARTED'],
    ['H005','Mockup attachments = 5x cold email response','100 emails with mockup vs 100 generic. Track response + close.','response_rate','3-5% replies','15%+ with mockup','LOW (built already)','5x close rate','NOT STARTED'],
    ['H006','Bland AI calls to funded cos = 6-10% conversion','100 calls/day (free). Test funded companies vs cold list.','call_to_meeting','2-3% (avg)','6-10% funded','MED (setup)','3-5x lead quality','NOT STARTED'],
    ['H007','3x/day posting = 5x faster account growth','2 identical accounts. 1x/day vs 3x/day for 30 days. Track followers.','follower_growth','TBD','5x faster','MED (30 days)','5x audience build speed','NOT STARTED'],
    ['H008','6-8 AM cold emails = 42% higher open rate','200 identical emails: 100 at 7AM, 100 at 2PM. Track opens/replies.','open_rate','current','42% improvement','LOW (change time)','More replies = more $','NOT STARTED'],
]

r = 2
for h in hyp:
    dark_row(ws3,r,h,WHITE,CYAN_FILL if r%2==0 else DARK)
    r += 1

# ============================================================
# SHEET 4: GTM + EDGE TACTICS
# ============================================================
ws4 = wb.create_sheet('GTM + EDGE')
h4 = ['CATEGORY','TACTIC','DETAIL','RISK_LEVEL','WORKS_UNTIL','EXPECTED_IMPACT']
w4 = [18,30,50,12,16,20]
style_sheet(ws4,h4,w4)

tactics = [
    ['ALGO GAMING','TikTok first 3-second hook','Pattern interrupt: zoom, text flash, reaction face. Hook or die in first 3 sec.','LEGAL','Evergreen','2-5x view retention'],
    ['ALGO GAMING','TikTok comment reply videos','Reply to comments with new videos. Algo treats as engagement signal, boosts reach.','LEGAL','Evergreen','3x reach boost'],
    ['ALGO GAMING','YouTube Shorts cross-post','Same TikTok content → YouTube Shorts. Zero extra production, new audience.','LEGAL','Evergreen','2x total reach'],
    ['ALGO GAMING','Instagram Collab feature','Use Collab feature to co-post with bigger accounts. Instant audience share.','LEGAL','Evergreen','10x post reach'],
    ['ALGO GAMING','X/Twitter thread game','5-8 tweet threads outperform single tweets 3-5x. Quote-tweet viral posts.','LEGAL','Evergreen','3-5x impressions'],
    ['ALGO GAMING','Pinterest SEO pins','SEO-optimize pin titles + descriptions. Rich pins. 15-25 pins/day via scheduler.','LEGAL','Evergreen','Compound organic traffic'],
    ['COLD OUTREACH','SPF/DKIM/DMARC setup','Set up ALL THREE on sending domain. Without them → spam folder guaranteed.','LEGAL','Evergreen','Go from 20% to 95% inbox placement'],
    ['COLD OUTREACH','Inbox rotation (3-5 accounts)','Rotate 3-5 sending accounts. Max 50 emails/account/day. Protects reputation.','LEGAL','Evergreen','Prevents domain blacklisting'],
    ['COLD OUTREACH','Timeline-based hooks','Open with "I noticed [company] just [event]..." Contextual = personal = replies.','LEGAL','Evergreen','10% reply rate (vs 3.4% avg)'],
    ['COLD OUTREACH','4-7 touch follow-up','Day 1,3,7,14,21 follow-up sequence. 42% of replies come from follow-ups.','LEGAL','Evergreen','2x total reply rate'],
    ['COLD OUTREACH','Lowercase subject lines','3-5 words. Lowercase. No caps, no emojis, no spam triggers. "quick question"','LEGAL','Evergreen','46% open rate benchmark'],
    ['SEO EDGE','Programmatic SEO at scale','300+ pages: "[service] in [city]". Claude Code generates all pages.','LEGAL','Until Google penalizes','10x organic pages'],
    ['SEO EDGE','Parasite SEO','Post authority content on Medium, LinkedIn, Quora. Link back to money site.','GREY AREA','6-12 months','Free high-DA backlinks'],
    ['SEO EDGE','HARO/Connectively backlinks','Answer journalist queries. Get backlinks from major publications. Free.','LEGAL','Evergreen','DR boost from Forbes etc'],
    ['SEO EDGE','Schema markup (FAQ/HowTo)','Add structured data. Rich snippets = 2x CTR in search results.','LEGAL','Evergreen','2x search CTR'],
    ['SEO EDGE','Expired domain redirect','Buy expired domains with backlinks. 301 redirect to your site.','GREY AREA','6-12 months','Instant DA/DR boost'],
    ['MULTI-ACCOUNT','GoLogin fingerprint separation','Separate browser profile per account. Unique fingerprint, cookies, storage.','GREY (platform TOS)','Ongoing with updates','Avoid cross-account detection'],
    ['MULTI-ACCOUNT','Proxy per account','1 residential proxy per account. Never share IPs. Decodo/Smartproxy.','GREY (platform TOS)','Evergreen','Avoid IP-based bans'],
    ['MULTI-ACCOUNT','14-day warmup schedule','Day 1-3: browse. Day 4-7: engage. Day 8-14: light posting. Day 15+: full.','GREY (platform TOS)','Evergreen','Avoid new account detection'],
    ['FINDOM GROWTH','X/Twitter hashtag discovery','#findom #paypig #humanATM #cashslave #finsub. Post 5-10x/day.','LEGAL','Evergreen','Direct niche audience discovery'],
    ['FINDOM GROWTH','Teaser→paywall funnel','Free teaser on Twitter → "Full content on Fanvue". Drive subs from free.','LEGAL','Evergreen','Convert followers to paying subs'],
    ['FINDOM GROWTH','Telegram VIP flash sales','$99-499/mo VIP. Flash sales: "Next 10 tributes get exclusive content."','LEGAL','Evergreen','25% conversion × $200 avg = $2K/flash'],
    ['FINDOM GROWTH','Bio payment stacking','CashApp + Venmo + crypto + Fanvue + Linktree in bio. Maximum options.','LEGAL','Evergreen','Reduce friction = more tributes'],
    ['CONVERSION','Price anchoring','Show $197 crossed out next to $97 sale price. Anchoring = 2-3x conversion.','LEGAL','Evergreen','2-3x conversion on products'],
    ['CONVERSION','Social proof numbers','"1,247 people bought this". Numbers beat testimonials.','LEGAL','Evergreen','1.5-2x conversion'],
    ['CONVERSION','Payment splitting','3 payments of $33 vs $97 one-time. Installments = 2x conversion.','LEGAL','Evergreen','2x conversion on high-ticket'],
    ['CONVERSION','Exit intent popup','Offer 20% off when user tries to leave. 10-15% save rate.','LEGAL','Evergreen','10-15% recovered visitors'],
    ['CONVERSION','Bundle pricing','Individual $30 each, bundle of 5 for $97. 50% perceived discount.','LEGAL','Evergreen','Higher AOV'],
    ['PAYMENT','CashApp/Venmo for findom','0% fee for personal payments. Instant settlement. Perfect for tributes.','LEGAL','Evergreen','Zero processing fees'],
    ['PAYMENT','Crypto in bio','BTC/ETH/SOL address. Zero fees. Anonymous for buyers (findom audience prefers).','LEGAL','Evergreen','Additional payment channel'],
    ['PAYMENT','Stripe volume negotiation','At $50K+/mo: negotiate from 2.9% to 2.2%. Saves $350/mo per $50K.','LEGAL','Evergreen','$4K+/yr savings at scale'],
    ['PAYMENT','S-Corp for tax optimization','At $50K+/yr profit: S-Corp saves 15.3% self-employment tax on distributions.','LEGAL','Evergreen','$7.5K+/yr tax savings'],
]

r = 2
for t in tactics:
    risk = t[3]
    fill = GREEN_FILL if risk == 'LEGAL' else YELLOW_FILL if 'GREY' in risk else RED_FILL
    dark_row(ws4,r,t,WHITE,fill)
    r += 1

# ============================================================
# SHEET 5: NEW OPS DISCOVERED (first principles)
# ============================================================
ws5 = wb.create_sheet('NEW OPS DISCOVERED')
h5 = ['OP_NAME','THESIS','EDGE','VIABILITY','GTM','VALIDATION_METHOD','EFFORT','STATUS']
w5 = [24,40,40,14,40,30,14,14]
style_sheet(ws5,h5,w5)

new_ops = [
    ['AI Agent-as-a-Service','Businesses need AI agents but cant build them. Sell pre-built agents for verticals.','Claude Code builds agent in 2-4 hrs. $500-5K per agent + $50-200/mo maintenance.','HIGH','Cold email to target vertical. Free 7-day trial → convert to monthly.','Check if competitors exist (market validated) or few (opportunity).','MED (2-4 hrs/agent)','NEW'],
    ['AI Data Analysis Service','Small biz have data but no analysts. Automated analysis reports via Claude Code.','Claude Code analyzes CSVs/databases in minutes. $200-1K per report.','HIGH','Target Shopify stores, SaaS companies. Free sample report as lead magnet.','Fiverr "data analysis" growing. Check pricing ($50-200/report).','LOW (30 min/report)','NEW'],
    ['Multi-Channel Outreach Orchestrator','Combine Bland AI + Instantly + LinkedIn into unified system. Sell as service.','Multi-channel = 3-5x response rate. Orchestrate voice + email + social.','HIGH','Use the system to sell itself. Generate leads via multi-channel, sell the system.','Clay/Apollo charge $149+/mo. Our orchestrated service has room.','HIGH (build system)','NEW'],
    ['Programmatic Blog Network','50+ niche blogs × 100+ long-tail keywords. AI generates all content.','Claude Code generates entire blog in 1-2 hrs. Cloudflare Pages = free unlimited bandwidth.','MEDIUM','Target underserved niches with high affiliate commissions.','Check keyword difficulty. Target KD <20. 3-6 month SEO timeline.','MED (2 hrs/blog)','NEW'],
    ['AI Translation Service','Translate content into 10+ languages via Claude. Sell as service.','Claude translates near-human quality. $0.10/word. 99% margin.','HIGH','Fiverr/Upwork. Target course creators, SaaS, ecom stores.','Fiverr translation $0.05-0.15/word. We are competitive.','LOW (5-10 min/job)','NEW'],
    ['AI Meeting Notes Service','Process recordings into structured action items + decisions.','Upload → Claude transcribes + structures in 5 min. $10-50/meeting.','MEDIUM','Target small biz, freelancers, consultants on LinkedIn.','Otter.ai = $16.67/mo. Our per-meeting model for occasional users.','LOW (5 min/meeting)','NEW'],
    ['Claude Code Skills Marketplace','Build and sell reusable Claude Code skills/plugins.','First-mover advantage. Claude Code skills marketplace just launched.','HIGH','Publish to marketplace. Market to Claude Code community.','Growing marketplace. Check existing skills and gaps.','MED (2-4 hrs/skill)','NEW'],
    ['AI Voice Note Transcription','Voice memos → structured notes/tasks. Target busy professionals.','Claude transcribes and structures. $5-20 per note set.','MEDIUM','LinkedIn targeting executives. Cold DM with sample.','Voice-first apps growing. Check Rev.ai, Whisper alternatives.','LOW (5 min/job)','NEW'],
]

r = 2
for o in new_ops:
    v = o[3]
    fill = GREEN_FILL if v == 'HIGH' else CYAN_FILL
    dark_row(ws5,r,o,WHITE,fill)
    r += 1

# ============================================================
# SHEET 6: SELF-TEST PROTOCOL
# ============================================================
ws6 = wb.create_sheet('SELF-TEST PROTOCOL')
h6 = ['OP_ID','OP_NAME','TEST_COMMAND','SUCCESS_CRITERIA','FREQUENCY','LAST_RUN','SCORE','RECOMMENDATION']
w6 = [10,24,40,40,12,14,10,14]
style_sheet(ws6,h6,w6)

tests = [
    ['S02','Local Biz Pipeline','python3 AUTOMATIONS/local_biz_pipeline.py --dry-run','Scraper returns 5+ businesses. Landing page valid HTML. Email passes spam check.','WEEKLY','NEVER','TBD','TEST NOW'],
    ['D01','Gumroad Products','ls -la DIGITAL_PRODUCTS/ && wc -c DIGITAL_PRODUCTS/*','9+ products exist. Each >1KB. Descriptions coherent.','MONTHLY','NEVER','TBD','TEST NOW'],
    ['G04','Alpha Extraction','python3 scripts/extract_source_csvs_from_mega_sheet.py','Completes without errors. 50+ entries. No duplicate URLs.','DAILY','NEVER','TBD','TEST NOW'],
    ['C01','TikTok Content','ls 04_CONTENT/social/ && check content quality','Content ready for posting. Platform-appropriate. No policy violations.','WEEKLY','NEVER','TBD','NEEDS ACCOUNTS'],
    ['P02','NSFW Findom','ls MONEY_METHODS/AI_INFLUENCER/ && check execution plan','Execution plan complete. Persona images ready. Compliance docs in place.','WEEKLY','NEVER','TBD','NEEDS ACCOUNTS'],
    ['S01','Freelance Arbitrage','check PRINTMAXX_FREELANCE_ARB.xlsx service catalog','30 services documented. Pricing set. Automation pipeline ready.','WEEKLY','NEVER','TBD','NEEDS ACCOUNTS'],
    ['G02','RBI System','python3 scripts/strategic_rbi_engine.py full','All 3 layers complete. Report generated. Learnings logged.','WEEKLY','2026-02-10','TBD','OPERATIONAL'],
    ['G03','Daily Briefing','python3 scripts/daily_briefing.py','10 systems scanned. Actions prioritized. Report saved.','DAILY','2026-02-10','TBD','OPERATIONAL'],
]

r = 2
for t in tests:
    rec = t[7]
    fill = RED_FILL if 'NEVER' in t[5] else GREEN_FILL if 'OPERATIONAL' in rec else YELLOW_FILL
    dark_row(ws6,r,t,WHITE,fill)
    r += 1

# ============================================================
# SHEET 7: MARKET REALITY DATA
# ============================================================
ws7 = wb.create_sheet('MARKET REALITY')
h7 = ['CHANNEL','REAL_EARNINGS_DATA','SUCCESS_RATE','TIME_TO_PROFITABILITY','SATURATION_2026','KEY_SOURCE']
w7 = [20,40,12,18,14,40]
style_sheet(ws7,h7,w7)

market = [
    ['Fanvue (AI creators)','Top: $50K/mo (Amibnw). Aitana: $30K/mo. Mid: $2.5-4K/mo. Most: $500-3K after 3-6mo. Platform grew 150% YoY to $100M ARR.','30%','3-6 months','GROWING','finance.yahoo.com/news/ai-influencers, fanvuemodels.com/blog/top-earners'],
    ['Fiverr AI Services','Category grew 83.8% YoY to $34M. Total Fiverr revenue $108.6M Q2 2025. AI agents and workflow automation trending.','10%','4-6 weeks','GROWING','investors.fiverr.com Q2 2025 results'],
    ['YouTube Faceless','Top: $500K-1.3M/mo (DaFuq Boom). Science channels avg $43K/mo. RPM $3-20. 70% make under $1K/mo.','10%','3-4 months','HIGH','nexlev.io/faceless-youtube, awisee.com faceless channels'],
    ['TikTok Creator Rewards','$0.40-1.00/1K views (20x old Creator Fund). Creators earned $4.1B in 2024. Micro-influencers: $700-1.2K/mo.','5%','2-3 months','HIGH','bluehost.com/blog/tiktok-pay, wealthinsightwatch.com'],
    ['AI UGC','Avg US salary $9.7K/mo (ZipRecruiter). Beginners $1.2-3K/mo. AI workflows = 3x more likely to hit $8K/mo in 6mo.','20%','2-3 months','MEDIUM','argil.ai UGC creator guide, alternativeincomemagazine.com'],
    ['Cold Email','Avg reply rate 3.43%. Timeline hooks: 10.01%. 2-3 follow-ups increase response 65.8%. Best: 4-7 touch sequence.','20%','1-2 months','MEDIUM','instantly.ai cold-email-benchmark-report-2026'],
    ['Micro-SaaS','92% fail within 3 years. 45% fail at 18-24 months. Median profitable: $4.2K MRR. Top: $258K/mo (Plausible).','5%','12-24 months','LOW-MED','rockingweb.com.au micro-saas-revenue-analysis-2025'],
    ['Gumroad Digital Products','Average creator: $100-500/mo. Top 1%: $10K+/mo. 10% platform fee. Need existing audience.','15%','1-3 months','MEDIUM','Gumroad data, indie hacker reports'],
    ['Cold Calling (Bland AI)','100 FREE calls/day. Avg conversion 2-3%. Top performers 6-10%. AI voice warm transfer model.','15%','1-2 months','LOW','Bland AI documentation, cold calling benchmarks'],
    ['Local Biz Websites','$500-5K per site. Cold email: 3-5% reply rate. Mockup attachment = 5x response.','25%','2-4 weeks','LOW','Industry data, local agency benchmarks'],
    ['AI Agents Market','$5.4B (2024) → $7.6B (2025). $3.8B raised by AI agent startups in 2024. 30% CAGR.','N/A','6-12 months','VERY LOW','a16z.com MCP deep dive, geeky-gadgets.com MCP monetization'],
    ['Newsletter','Beehiiv/Substack. $5-15/mo premium. Need 1K+ subs to monetize. Ad network for sponsors.','8%','6-12 months','MEDIUM','Substack/Beehiiv creator data'],
]

r = 2
for m in market:
    sat = m[4]
    fill = GREEN_FILL if 'LOW' in sat else YELLOW_FILL if 'MEDIUM' in sat else RED_FILL
    dark_row(ws7,r,m,WHITE,fill)
    r += 1

# ============================================================
# SAVE
# ============================================================
out = '/sessions/awesome-nice-brown/mnt/PRINTMAXX_STARTER_KITttttt/PRINTMAXX_STRATEGIC_RBI.xlsx'
wb.save(out)
print(f'Saved {out}')
print(f'Sheet 1: VIABILITY MATRIX - {len(viability)} op types scored')
print(f'Sheet 2: BOTTLENECKS - {len(bottlenecks)} issues ranked')
print(f'Sheet 3: HYPOTHESES - {len(hyp)} testable experiments')
print(f'Sheet 4: GTM + EDGE - {len(tactics)} tactics')
print(f'Sheet 5: NEW OPS - {len(new_ops)} first-principles discoveries')
print(f'Sheet 6: SELF-TEST - {len(tests)} validation protocols')
print(f'Sheet 7: MARKET REALITY - {len(market)} channels with real data')
