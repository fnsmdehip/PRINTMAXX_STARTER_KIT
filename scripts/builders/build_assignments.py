from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

BG_DARK = PatternFill('solid', fgColor='0D1117')
BG_HEADER = PatternFill('solid', fgColor='0A1628')
BG_SECTION = PatternFill('solid', fgColor='112240')
BG_R1 = PatternFill('solid', fgColor='0D1117')
BG_R2 = PatternFill('solid', fgColor='161B22')
BG_FREE = PatternFill('solid', fgColor='0B3D0B')
BG_ACTIVE = PatternFill('solid', fgColor='0B3D2B')
BG_CARD = PatternFill('solid', fgColor='161B22')
BG_WARN = PatternFill('solid', fgColor='2D2500')
BG_V1 = PatternFill('solid', fgColor='0D2818')  # venture color 1
BG_V2 = PatternFill('solid', fgColor='1A1A3D')
BG_V3 = PatternFill('solid', fgColor='2D1B1B')
BG_V4 = PatternFill('solid', fgColor='1A2D3D')
BG_V5 = PatternFill('solid', fgColor='2D2D1A')

FT = Font(name='Arial', size=20, bold=True, color='00D4FF')
FS = Font(name='Arial', size=12, bold=True, color='8B949E')
FH = Font(name='Arial', size=11, bold=True, color='00D4FF')
FSC = Font(name='Arial', size=12, bold=True, color='58A6FF')
FB = Font(name='Arial', size=10, color='C9D1D9')
FW = Font(name='Arial', size=10, bold=True, color='FFFFFF')
FC = Font(name='Arial', size=10, bold=True, color='00D4FF')
FG = Font(name='Arial', size=10, bold=True, color='3FB950')
FGO = Font(name='Arial', size=10, bold=True, color='F0C000')
FO = Font(name='Arial', size=10, color='F0883E')
FR = Font(name='Arial', size=10, color='F85149')
FP = Font(name='Arial', size=10, bold=True, color='A371F7')
FHL = Font(name='Arial', size=11, bold=True, color='FFFFFF')
FSM = Font(name='Arial', size=9, color='8B949E')
FSTAT = Font(name='Arial', size=24, bold=True, color='00D4FF')

BD = Border(bottom=Side(style='thin', color='21262D'))
BS = Border(bottom=Side(style='medium', color='00D4FF'))
AC = Alignment(horizontal='center', vertical='center')
AL = Alignment(horizontal='left', vertical='center', wrap_text=True)

def sw(ws, w):
    for i, v in enumerate(w, 1): ws.column_dimensions[get_column_letter(i)].width = v
def fb(ws, r, c, f):
    for i in range(1, c+1): ws.cell(row=r, column=i).fill = f
def sh(ws, r, mc, t):
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=mc-1)
    ws.cell(row=r, column=2, value=t).font = FSC
    fb(ws, r, mc, BG_SECTION)
    for c in range(1, mc+1): ws.cell(row=r, column=c).border = BS
    ws.row_dimensions[r].height = 28
    return r + 1
def hd(ws, r, mc, h):
    for i, v in enumerate(h):
        c = ws.cell(row=r, column=i+1, value=v); c.font = FH; c.fill = BG_HEADER; c.alignment = AC
    ws.row_dimensions[r].height = 24
    return r + 1

# ============================================================
# THE MASTER DATA: 5 VENTURES × EXACT INFRA ASSIGNMENTS
# ============================================================

# Dolphin Anty has 10 free profiles. We assign 2 per venture (main + backup).
# Buffer has 3 free channels. Tweetlio covers X. So Buffer = TikTok + IG + LinkedIn (rotate).
# Decodo/Smartproxy proxies: 1 residential IP per venture.
# Each venture gets: 1 Dolphin profile (main) + 1 backup, 1 proxy, 4 social accounts, 1 newsletter, 1 landing page, 1 email.

VENTURES = [
    {
        'id': 'V1', 'name': 'PRINTMAXX META BRAND', 'niche': 'Build in Public / Solopreneur',
        'color': BG_V1, 'phase': 'P1 ACTIVE',
        'methods': 'MM021 (Personal Brand SEO), MM044 (Rapid Build), MM015 (Newsletter)',
        'revenue': '$2K-20K/mo',
        'dolphin_profile': 'DOLPHIN-01 (Main)', 'dolphin_backup': 'DOLPHIN-02 (Backup)',
        'proxy': 'Decodo-1 (US Residential)', 'proxy_type': 'Smartproxy Residential',
        'x_handle': '@printmaxxer', 'x_email': 'printmaxxer@protonmail.com',
        'tiktok': '@printmaxxer', 'instagram': '@printmaxxer',
        'youtube': 'PrintMaxxer', 'linkedin': 'Shane Donnelly (personal)',
        'newsletter': 'STACKPRINT (Beehiiv #1)', 'newsletter_limit': '2,500 subs free',
        'landing': 'Carrd Site #1 → stackprint.com', 'domain': 'stackprint.com ($9.73/yr)',
        'buffer_channels': 'TikTok + Instagram (2 of 3 free slots)',
        'x_scheduler': 'Tweetlio (free, unlimited X posts)',
        'content_tools': 'Canva (design), CapCut (video), Leonardo.ai (AI thumbnails)',
        'ai_tools': 'ElevenLabs (voiceover), Kling (AI video clips)',
        'analytics': 'GA4 + PostHog (self-hosted Oracle)',
        'monetization': 'Gumroad (digital products) + Stripe (payments) + Beehiiv Ads',
        'crm': 'HubSpot (unlimited contacts free)',
        'email_warmup': 'TrulyInbox (free forever)',
        'hosting': 'Vercel (landing) + Oracle Cloud (backend)',
        'dns': 'Cloudflare (free DNS + SSL)',
        'accounts': [
            ('X/Twitter', '@printmaxxer', 'DOLPHIN-01', 'Decodo-1', 'Tweetlio', 'printmaxxer@protonmail.com'),
            ('TikTok', '@printmaxxer', 'DOLPHIN-01', 'Decodo-1', 'Buffer Ch.1', 'printmaxxer@protonmail.com'),
            ('Instagram', '@printmaxxer', 'DOLPHIN-01', 'Decodo-1', 'Buffer Ch.2', 'printmaxxer@protonmail.com'),
            ('YouTube', 'PrintMaxxer', 'DOLPHIN-01', 'Decodo-1', 'Manual', 'printmaxxer@protonmail.com'),
            ('LinkedIn', 'Shane Donnelly', 'DOLPHIN-01', 'Decodo-1', 'Manual', 'personal'),
        ],
    },
    {
        'id': 'V2', 'name': 'AI UTILITIES / TECH', 'niche': 'AI Tools, Automation, Workflows',
        'color': BG_V2, 'phase': 'P1 ACTIVE',
        'methods': 'MM001 (App Factory), CF009 (Tech Explainers), AI001 (Niche Expert)',
        'revenue': '$1K-50K/mo',
        'dolphin_profile': 'DOLPHIN-03 (Main)', 'dolphin_backup': 'DOLPHIN-04 (Backup)',
        'proxy': 'Decodo-2 (US Residential)', 'proxy_type': 'Smartproxy Residential',
        'x_handle': '@patchnotes_', 'x_email': 'ai.workflows.tips@protonmail.com',
        'tiktok': '@patchnotes', 'instagram': '@patchnotes.ai',
        'youtube': 'Patch Notes', 'linkedin': 'N/A (use V1 LinkedIn)',
        'newsletter': 'STACKPRINT shared (tech section) OR separate Substack',
        'newsletter_limit': 'Unlimited (Substack) or shared Beehiiv',
        'landing': 'Vercel landing page → patchnotes.dev', 'domain': 'patchnotes.dev ($12/yr est.)',
        'buffer_channels': 'TikTok (shared Buffer Ch.1 rotation)',
        'x_scheduler': 'Tweetlio (free, separate login)',
        'content_tools': 'Canva (design), CapCut (video), Leonardo.ai (AI screenshots)',
        'ai_tools': 'ElevenLabs (voiceover), Kling (demo videos)',
        'analytics': 'GA4 + PostHog',
        'monetization': 'App Factory (RevenueCat IAP) + Gumroad (templates) + Affiliate links',
        'crm': 'HubSpot (shared)',
        'email_warmup': 'TrulyInbox (free)',
        'hosting': 'Vercel (app landing) + Oracle (API backend)',
        'dns': 'Cloudflare',
        'accounts': [
            ('X/Twitter', '@patchnotes_', 'DOLPHIN-03', 'Decodo-2', 'Tweetlio', 'ai.workflows.tips@protonmail.com'),
            ('TikTok', '@patchnotes', 'DOLPHIN-03', 'Decodo-2', 'Buffer Ch.1 (rotate)', 'ai.workflows.tips@protonmail.com'),
            ('Instagram', '@patchnotes.ai', 'DOLPHIN-03', 'Decodo-2', 'Manual / Publer later', 'ai.workflows.tips@protonmail.com'),
            ('YouTube', 'Patch Notes', 'DOLPHIN-03', 'Decodo-2', 'Manual', 'ai.workflows.tips@protonmail.com'),
        ],
    },
    {
        'id': 'V3', 'name': 'FAITH / DEVOTIONAL', 'niche': 'Daily Devotionals, Prayer, Scripture',
        'color': BG_V3, 'phase': 'P1 ACTIVE',
        'methods': 'MM001 (PrayerLock App), CF007 (Motivation Quotes), MM002 (Info Products)',
        'revenue': '$500-10K/mo',
        'dolphin_profile': 'DOLPHIN-05 (Main)', 'dolphin_backup': 'DOLPHIN-06 (Backup)',
        'proxy': 'Decodo-3 (US Residential)', 'proxy_type': 'Smartproxy Residential',
        'x_handle': '@morningwatch_', 'x_email': 'daily.anchor.faith@protonmail.com',
        'tiktok': '@morningwatch', 'instagram': '@morningwatch_',
        'youtube': 'Morning Watch', 'linkedin': 'N/A',
        'newsletter': 'MORNING MANNA (Beehiiv #2)', 'newsletter_limit': '2,500 subs free',
        'landing': 'Carrd Site #2 → morningmanna.com', 'domain': 'morningmanna.com ($9.73/yr)',
        'buffer_channels': 'Instagram (shared Buffer Ch.2 rotation)',
        'x_scheduler': 'Tweetlio (free)',
        'content_tools': 'Canva (scripture graphics), CapCut (devotional reels)',
        'ai_tools': 'ElevenLabs (prayer voiceover), Leonardo.ai (faith imagery)',
        'analytics': 'GA4',
        'monetization': 'PrayerLock App (RevenueCat) + Gumroad (devotional guides) + Beehiiv Ads',
        'crm': 'HubSpot (shared)',
        'email_warmup': 'TrulyInbox (free)',
        'hosting': 'Vercel (PrayerLock landing) + Oracle (app backend)',
        'dns': 'Cloudflare',
        'accounts': [
            ('X/Twitter', '@morningwatch_', 'DOLPHIN-05', 'Decodo-3', 'Tweetlio', 'daily.anchor.faith@protonmail.com'),
            ('TikTok', '@morningwatch', 'DOLPHIN-05', 'Decodo-3', 'Buffer Ch.2 (rotate)', 'daily.anchor.faith@protonmail.com'),
            ('Instagram', '@morningwatch_', 'DOLPHIN-05', 'Decodo-3', 'Buffer Ch.2 (rotate)', 'daily.anchor.faith@protonmail.com'),
            ('YouTube', 'Morning Watch', 'DOLPHIN-05', 'Decodo-3', 'Manual', 'daily.anchor.faith@protonmail.com'),
        ],
    },
    {
        'id': 'V4', 'name': 'FITNESS / HEALTH', 'niche': 'Workouts, Nutrition, Progressive Overload',
        'color': BG_V4, 'phase': 'P1 ACTIVE',
        'methods': 'MM001 (StepUnlock/WalkToUnlock App), AI005 (Fitness Coach), CF007 (Motivation)',
        'revenue': '$500-10K/mo',
        'dolphin_profile': 'DOLPHIN-07 (Main)', 'dolphin_backup': 'DOLPHIN-08 (Backup)',
        'proxy': 'Decodo-4 (US Residential)', 'proxy_type': 'Smartproxy Residential',
        'x_handle': '@repscheme_', 'x_email': 'three.hour.physique@protonmail.com',
        'tiktok': '@repscheme', 'instagram': '@repscheme',
        'youtube': 'Rep Scheme', 'linkedin': 'N/A',
        'newsletter': 'Substack (free, unlimited)', 'newsletter_limit': 'Unlimited (Substack)',
        'landing': 'Carrd Site #3 → repscheme.com', 'domain': 'repscheme.com ($9.73/yr)',
        'buffer_channels': 'Buffer Ch.3 (TikTok or IG rotation)',
        'x_scheduler': 'Tweetlio (free)',
        'content_tools': 'Canva (workout graphics), CapCut (exercise demos)',
        'ai_tools': 'Leonardo.ai (fitness imagery), ElevenLabs (coaching voice)',
        'analytics': 'GA4',
        'monetization': 'StepUnlock App (RevenueCat) + Gumroad (workout PDFs) + Affiliate (supplements)',
        'crm': 'HubSpot (shared)',
        'email_warmup': 'TrulyInbox (free)',
        'hosting': 'Vercel (app landing) + Oracle (backend)',
        'dns': 'Cloudflare',
        'accounts': [
            ('X/Twitter', '@repscheme_', 'DOLPHIN-07', 'Decodo-4', 'Tweetlio', 'three.hour.physique@protonmail.com'),
            ('TikTok', '@repscheme', 'DOLPHIN-07', 'Decodo-4', 'Buffer Ch.3', 'three.hour.physique@protonmail.com'),
            ('Instagram', '@repscheme', 'DOLPHIN-07', 'Decodo-4', 'Buffer Ch.3 (rotate)', 'three.hour.physique@protonmail.com'),
            ('YouTube', 'Rep Scheme', 'DOLPHIN-07', 'Decodo-4', 'Manual', 'three.hour.physique@protonmail.com'),
        ],
    },
    {
        'id': 'V5', 'name': 'MEME / CONTENT FARM', 'niche': 'Viral Clips, Memes, No-Context, Compilations',
        'color': BG_V5, 'phase': 'P1 ACTIVE',
        'methods': 'MM006 (Content Farm), CF004 (Memes), CF006 (Clips), SWARM001 (Swarm Promo)',
        'revenue': '$500-10K/mo',
        'dolphin_profile': 'DOLPHIN-09 (Main)', 'dolphin_backup': 'DOLPHIN-10 (Backup)',
        'proxy': 'Decodo-5 (US Residential)', 'proxy_type': 'Smartproxy Residential',
        'x_handle': '@foldersdump', 'x_email': 'foldersdump@protonmail.com',
        'tiktok': '@foldersdump', 'instagram': '@foldersdump',
        'youtube': 'Folders Dump', 'linkedin': 'N/A',
        'newsletter': 'N/A (traffic → other ventures via SWARM)', 'newsletter_limit': 'N/A',
        'landing': 'N/A (social-only venture)', 'domain': 'N/A',
        'buffer_channels': 'Tweetlio only (X-focused). IG/TikTok manual or Publer later.',
        'x_scheduler': 'Tweetlio (free)',
        'content_tools': 'CapCut (clip editing), Canva (meme templates)',
        'ai_tools': 'Kling (AI-enhanced clips)',
        'analytics': 'GA4 (link tracking only)',
        'monetization': 'Creator Fund (TikTok/YT) + SWARM traffic to V1-V4 ventures',
        'crm': 'N/A',
        'email_warmup': 'N/A',
        'hosting': 'N/A (social only)',
        'dns': 'N/A',
        'accounts': [
            ('X/Twitter', '@foldersdump', 'DOLPHIN-09', 'Decodo-5', 'Tweetlio', 'foldersdump@protonmail.com'),
            ('TikTok', '@foldersdump', 'DOLPHIN-09', 'Decodo-5', 'Manual', 'foldersdump@protonmail.com'),
            ('Instagram', '@foldersdump', 'DOLPHIN-09', 'Decodo-5', 'Manual', 'foldersdump@protonmail.com'),
            ('YouTube', 'Folders Dump', 'DOLPHIN-09', 'Decodo-5', 'Manual', 'foldersdump@protonmail.com'),
        ],
    },
]

MC = 10

# ============================================================
# SHEET 1: VENTURE → INFRA ASSIGNMENT MAP
# ============================================================
ws = wb.active
ws.title = 'VENTURE ASSIGNMENTS'
ws.sheet_properties.tabColor = '00D4FF'
sw(ws, [3, 20, 22, 22, 22, 22, 22, 22, 22, 3])

for r in range(1, 250):
    fb(ws, r, MC, BG_DARK)

ws.merge_cells('B2:I2')
ws.cell(row=2, column=2, value='VENTURE → INFRASTRUCTURE ASSIGNMENTS').font = FT
ws.row_dimensions[2].height = 35
ws.merge_cells('B3:I3')
ws.cell(row=3, column=2, value='5 ventures × 10 Dolphin profiles × 5 proxies × 21 social accounts × exact tool pairings').font = FS

r = 5
for v in VENTURES:
    # Venture header
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=MC-1)
    ws.cell(row=r, column=2, value=f"{v['id']}: {v['name']}").font = Font(name='Arial', size=14, bold=True, color='00D4FF')
    fb(ws, r, MC, BG_SECTION)
    for c in range(1, MC+1): ws.cell(row=r, column=c).border = BS
    ws.row_dimensions[r].height = 32
    r += 1

    # Venture info grid (2 columns of key-value pairs)
    info_pairs = [
        ('Niche', v['niche'], 'Phase', v['phase']),
        ('Methods', v['methods'], 'Revenue Target', v['revenue']),
        ('Dolphin Profile', v['dolphin_profile'], 'Backup Profile', v['dolphin_backup']),
        ('Proxy', v['proxy'], 'Proxy Type', v['proxy_type']),
        ('X Handle', v['x_handle'], 'X Scheduler', v['x_scheduler']),
        ('TikTok', v['tiktok'], 'Instagram', v['instagram']),
        ('YouTube', v['youtube'], 'LinkedIn', v['linkedin']),
        ('Newsletter', v['newsletter'], 'Sub Limit', v['newsletter_limit']),
        ('Landing Page', v['landing'], 'Domain', v['domain']),
        ('Buffer', v['buffer_channels'], 'Email', v['x_email']),
        ('Content Tools', v['content_tools'], 'AI Tools', v['ai_tools']),
        ('Monetization', v['monetization'], 'CRM', v['crm']),
        ('Hosting', v['hosting'], 'DNS', v['dns']),
        ('Email Warmup', v['email_warmup'], 'Analytics', v['analytics']),
    ]

    for idx, (k1, v1, k2, v2) in enumerate(info_pairs):
        bg = v['color'] if idx < 4 else (BG_R1 if idx % 2 == 0 else BG_R2)
        fb(ws, r, MC, bg)
        ws.cell(row=r, column=2, value=k1).font = FGO
        ws.cell(row=r, column=3, value=v1).font = FW if idx < 4 else FB
        ws.cell(row=r, column=3).alignment = AL
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        ws.cell(row=r, column=6, value=k2).font = FGO
        ws.cell(row=r, column=7, value=v2).font = FW if idx < 4 else FB
        ws.cell(row=r, column=7).alignment = AL
        ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=9)
        for c in range(1, MC+1): ws.cell(row=r, column=c).border = BD
        ws.row_dimensions[r].height = 24
        r += 1

    # Account table for this venture
    r += 1
    acct_headers = ['', 'PLATFORM', 'HANDLE', 'DOLPHIN PROFILE', 'PROXY', 'SCHEDULER', 'EMAIL', '', '']
    r = hd(ws, r, MC, acct_headers)

    for aidx, (plat, handle, dolphin, proxy, sched, email) in enumerate(v['accounts']):
        bg = BG_ACTIVE
        fb(ws, r, MC, bg)
        ws.cell(row=r, column=2, value=plat).font = FC
        ws.cell(row=r, column=2).alignment = AC
        ws.cell(row=r, column=3, value=handle).font = FHL
        ws.cell(row=r, column=4, value=dolphin).font = FP
        ws.cell(row=r, column=4).alignment = AC
        ws.cell(row=r, column=5, value=proxy).font = FO
        ws.cell(row=r, column=5).alignment = AC
        ws.cell(row=r, column=6, value=sched).font = FG
        ws.cell(row=r, column=6).alignment = AC
        ws.cell(row=r, column=7, value=email).font = FSM
        ws.cell(row=r, column=7).alignment = AL
        ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=9)
        for c in range(1, MC+1): ws.cell(row=r, column=c).border = BD
        ws.row_dimensions[r].height = 26
        r += 1
    r += 2


# ============================================================
# SHEET 2: DOLPHIN PROFILE ALLOCATION
# ============================================================
ws2 = wb.create_sheet('DOLPHIN PROFILES')
ws2.sheet_properties.tabColor = 'A371F7'
sw(ws2, [3, 16, 24, 20, 20, 22, 22, 3])
MC2 = 8

for r in range(1, 30):
    fb(ws2, r, MC2, BG_DARK)

ws2.merge_cells('B2:G2')
ws2.cell(row=2, column=2, value='DOLPHIN ANTY — 10 PROFILE ALLOCATION').font = FT
ws2.row_dimensions[2].height = 32
ws2.merge_cells('B3:G3')
ws2.cell(row=3, column=2, value='10 free profiles. 5 ventures × 2 each (main + backup). Zero profiles wasted.').font = FS

r = 5
r = hd(ws2, r, MC2, ['', 'PROFILE #', 'ASSIGNED TO', 'ROLE', 'PROXY PAIRED', 'PLATFORMS', 'STATUS'])

profiles = [
    ('DOLPHIN-01', 'V1: PRINTMAXX META', 'MAIN', 'Decodo-1 (US)', 'X + TT + IG + YT + LI', 'ACTIVE'),
    ('DOLPHIN-02', 'V1: PRINTMAXX META', 'BACKUP', 'Decodo-1 (US)', 'Backup accounts', 'STANDBY'),
    ('DOLPHIN-03', 'V2: AI / TECH', 'MAIN', 'Decodo-2 (US)', 'X + TT + IG + YT', 'ACTIVE'),
    ('DOLPHIN-04', 'V2: AI / TECH', 'BACKUP', 'Decodo-2 (US)', 'Backup accounts', 'STANDBY'),
    ('DOLPHIN-05', 'V3: FAITH', 'MAIN', 'Decodo-3 (US)', 'X + TT + IG + YT', 'ACTIVE'),
    ('DOLPHIN-06', 'V3: FAITH', 'BACKUP', 'Decodo-3 (US)', 'Backup accounts', 'STANDBY'),
    ('DOLPHIN-07', 'V4: FITNESS', 'MAIN', 'Decodo-4 (US)', 'X + TT + IG + YT', 'ACTIVE'),
    ('DOLPHIN-08', 'V4: FITNESS', 'BACKUP', 'Decodo-4 (US)', 'Backup accounts', 'STANDBY'),
    ('DOLPHIN-09', 'V5: MEME/CONTENT', 'MAIN', 'Decodo-5 (US)', 'X + TT + IG + YT', 'ACTIVE'),
    ('DOLPHIN-10', 'V5: MEME/CONTENT', 'BACKUP', 'Decodo-5 (US)', 'Backup accounts', 'STANDBY'),
]

for idx, (prof, assigned, role, proxy, plats, status) in enumerate(profiles):
    bg = BG_ACTIVE if role == 'MAIN' else BG_R2
    fb(ws2, r, MC2, bg)
    ws2.cell(row=r, column=2, value=prof).font = FP
    ws2.cell(row=r, column=2).alignment = AC
    ws2.cell(row=r, column=3, value=assigned).font = FW
    ws2.cell(row=r, column=4, value=role).font = FG if role == 'MAIN' else FSM
    ws2.cell(row=r, column=4).alignment = AC
    ws2.cell(row=r, column=5, value=proxy).font = FO
    ws2.cell(row=r, column=5).alignment = AC
    ws2.cell(row=r, column=6, value=plats).font = FB
    ws2.cell(row=r, column=6).alignment = AC
    ws2.cell(row=r, column=7, value=status).font = FG if status == 'ACTIVE' else FSM
    ws2.cell(row=r, column=7).alignment = AC
    for c in range(1, MC2+1): ws2.cell(row=r, column=c).border = BD
    ws2.row_dimensions[r].height = 28
    r += 1

# Rules
r += 2
r = sh(ws2, r, MC2, 'PROFILE ISOLATION RULES')
rules = [
    'ONE venture per Dolphin profile. Never cross-contaminate fingerprints.',
    'ONE proxy per profile. Decodo-N always pairs with DOLPHIN-(2N-1) and DOLPHIN-(2N).',
    'BACKUP profiles use same proxy as main. Only activate if main gets flagged.',
    'NEVER log into a V1 account from a V3 profile. Fingerprint leak = instant ban.',
    'Set User-Agent to MOBILE for IG/TikTok profiles. Desktop UA gets flagged.',
    'Rotate cookies every 7 days. Clear cache monthly. Never "restore" from backup profile.',
]
for rule in rules:
    fb(ws2, r, MC2, BG_R2)
    ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    ws2.cell(row=r, column=2, value=f'  {rule}').font = FB
    ws2.cell(row=r, column=2).alignment = AL
    ws2.row_dimensions[r].height = 24
    r += 1


# ============================================================
# SHEET 3: PROXY ASSIGNMENT
# ============================================================
ws3 = wb.create_sheet('PROXY MAP')
ws3.sheet_properties.tabColor = 'F0883E'
sw(ws3, [3, 16, 24, 20, 20, 24, 24, 3])
MC3 = 8

for r in range(1, 25):
    fb(ws3, r, MC3, BG_DARK)

ws3.merge_cells('B2:G2')
ws3.cell(row=2, column=2, value='PROXY → VENTURE ASSIGNMENT MAP').font = FT
ws3.row_dimensions[2].height = 32
ws3.merge_cells('B3:G3')
ws3.cell(row=3, column=2, value='5 residential proxies. 1 per venture. Strict isolation.').font = FS

r = 5
r = hd(ws3, r, MC3, ['', 'PROXY ID', 'VENTURE', 'TYPE', 'REGION', 'DOLPHIN PROFILES', 'SOCIAL ACCOUNTS'])

proxies = [
    ('Decodo-1', 'V1: PRINTMAXX', 'Residential', 'US East', 'DOLPHIN-01, 02', '@printmaxxer (X, TT, IG, YT, LI)'),
    ('Decodo-2', 'V2: AI/TECH', 'Residential', 'US West', 'DOLPHIN-03, 04', '@patchnotes_ (X, TT, IG, YT)'),
    ('Decodo-3', 'V3: FAITH', 'Residential', 'US South', 'DOLPHIN-05, 06', '@morningwatch_ (X, TT, IG, YT)'),
    ('Decodo-4', 'V4: FITNESS', 'Residential', 'US Central', 'DOLPHIN-07, 08', '@repscheme_ (X, TT, IG, YT)'),
    ('Decodo-5', 'V5: MEME/CONTENT', 'Residential', 'US Random', 'DOLPHIN-09, 10', '@foldersdump (X, TT, IG, YT)'),
]

for idx, (pid, vent, ptype, region, dps, accts) in enumerate(proxies):
    bg = [BG_V1, BG_V2, BG_V3, BG_V4, BG_V5][idx]
    fb(ws3, r, MC3, bg)
    ws3.cell(row=r, column=2, value=pid).font = FO
    ws3.cell(row=r, column=2).alignment = AC
    ws3.cell(row=r, column=3, value=vent).font = FW
    ws3.cell(row=r, column=4, value=ptype).font = FG
    ws3.cell(row=r, column=4).alignment = AC
    ws3.cell(row=r, column=5, value=region).font = FB
    ws3.cell(row=r, column=5).alignment = AC
    ws3.cell(row=r, column=6, value=dps).font = FP
    ws3.cell(row=r, column=6).alignment = AC
    ws3.cell(row=r, column=7, value=accts).font = FB
    ws3.cell(row=r, column=7).alignment = AL
    for c in range(1, MC3+1): ws3.cell(row=r, column=c).border = BD
    ws3.row_dimensions[r].height = 32
    r += 1

# Free proxy options
r += 2
r = sh(ws3, r, MC3, 'FREE PROXY OPTIONS (NON-SOCIAL ONLY)')
free_proxies = [
    ('Webshare', '10 datacenter IPs', '$0 forever', 'Scraping, research, non-social ONLY. Gets banned on social platforms.'),
    ('Bright Data', '15 IPs + 2GB/mo', '$0 trial', 'Testing residential quality. Good for initial account creation.'),
    ('Geonode', 'Public proxy list', '$0 forever', 'Lowest quality. Research/non-critical use only.'),
]
r = hd(ws3, r, MC3, ['', 'PROVIDER', 'FREE TIER', 'COST', 'USE CASE', '', ''])
for idx, (prov, free, cost, use) in enumerate(free_proxies):
    fb(ws3, r, MC3, BG_R1 if idx % 2 == 0 else BG_R2)
    ws3.cell(row=r, column=2, value=prov).font = FW
    ws3.cell(row=r, column=3, value=free).font = FG
    ws3.cell(row=r, column=4, value=cost).font = FG
    ws3.cell(row=r, column=4).alignment = AC
    ws3.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
    ws3.cell(row=r, column=5, value=use).font = FB
    ws3.cell(row=r, column=5).alignment = AL
    for c in range(1, MC3+1): ws3.cell(row=r, column=c).border = BD
    ws3.row_dimensions[r].height = 26
    r += 1


# ============================================================
# SHEET 4: BUFFER + SCHEDULER ALLOCATION
# ============================================================
ws4 = wb.create_sheet('SCHEDULER MAP')
ws4.sheet_properties.tabColor = '3FB950'
sw(ws4, [3, 16, 22, 22, 22, 28, 3])
MC4 = 7

for r in range(1, 40):
    fb(ws4, r, MC4, BG_DARK)

ws4.merge_cells('B2:F2')
ws4.cell(row=2, column=2, value='BUFFER + TWEETLIO CHANNEL ALLOCATION').font = FT
ws4.row_dimensions[2].height = 32
ws4.merge_cells('B3:F3')
ws4.cell(row=3, column=2, value='Buffer: 3 free channels. Tweetlio: unlimited X. Manual: YouTube. Total: 21 accounts covered.').font = FS

r = 5
r = sh(ws4, r, MC4, 'BUFFER (3 FREE CHANNELS) — HIGH-VALUE SLOTS')
r = hd(ws4, r, MC4, ['', 'CHANNEL #', 'PLATFORM', 'VENTURE', 'HANDLE', 'SCHEDULE'])

buffer_channels = [
    ('Buffer Ch.1', 'TikTok', 'V1: PRINTMAXX + V2: AI (rotation)', '@printmaxxer / @patchnotes', 'M/W/F = V1, T/Th/Sa = V2. 1 post/day.'),
    ('Buffer Ch.2', 'Instagram', 'V1: PRINTMAXX + V3: FAITH (rotation)', '@printmaxxer / @morningwatch_', 'M/W/F = V1, T/Th/Sa = V3. 1 post/day.'),
    ('Buffer Ch.3', 'TikTok', 'V4: FITNESS', '@repscheme', 'Daily. Workout demos + tips. 1-2 posts/day.'),
]

for idx, (ch, plat, vent, handle, sched) in enumerate(buffer_channels):
    fb(ws4, r, MC4, BG_FREE)
    ws4.cell(row=r, column=2, value=ch).font = FG
    ws4.cell(row=r, column=2).alignment = AC
    ws4.cell(row=r, column=3, value=plat).font = FC
    ws4.cell(row=r, column=3).alignment = AC
    ws4.cell(row=r, column=4, value=vent).font = FW
    ws4.cell(row=r, column=5, value=handle).font = FHL
    ws4.cell(row=r, column=6, value=sched).font = FB
    ws4.cell(row=r, column=6).alignment = AL
    for c in range(1, MC4+1): ws4.cell(row=r, column=c).border = BD
    ws4.row_dimensions[r].height = 36
    r += 1

r += 1
r = sh(ws4, r, MC4, 'TWEETLIO (FREE, UNLIMITED X/TWITTER POSTS)')
r = hd(ws4, r, MC4, ['', 'VENTURE', 'X HANDLE', 'DOLPHIN', 'PROXY', 'SCHEDULE'])

tweetlio = [
    ('V1: PRINTMAXX', '@printmaxxer', 'DOLPHIN-01', 'Decodo-1', '3-5 tweets/day. Threads M/W. Revenue screenshots F.'),
    ('V2: AI/TECH', '@patchnotes_', 'DOLPHIN-03', 'Decodo-2', '3-5 tweets/day. Tool reviews T/Th. AI news daily.'),
    ('V3: FAITH', '@morningwatch_', 'DOLPHIN-05', 'Decodo-3', '2-3 tweets/day. Scripture AM. Devotional PM.'),
    ('V4: FITNESS', '@repscheme_', 'DOLPHIN-07', 'Decodo-4', '2-3 tweets/day. Workout AM. Nutrition PM.'),
    ('V5: MEME', '@foldersdump', 'DOLPHIN-09', 'Decodo-5', '5-10 tweets/day. Viral clips. Meme threads.'),
]

for idx, (vent, handle, dolphin, proxy, sched) in enumerate(tweetlio):
    bg = BG_R1 if idx % 2 == 0 else BG_R2
    fb(ws4, r, MC4, bg)
    ws4.cell(row=r, column=2, value=vent).font = FW
    ws4.cell(row=r, column=3, value=handle).font = FHL
    ws4.cell(row=r, column=4, value=dolphin).font = FP
    ws4.cell(row=r, column=4).alignment = AC
    ws4.cell(row=r, column=5, value=proxy).font = FO
    ws4.cell(row=r, column=5).alignment = AC
    ws4.cell(row=r, column=6, value=sched).font = FB
    ws4.cell(row=r, column=6).alignment = AL
    for c in range(1, MC4+1): ws4.cell(row=r, column=c).border = BD
    ws4.row_dimensions[r].height = 30
    r += 1

r += 1
r = sh(ws4, r, MC4, 'MANUAL UPLOAD (YOUTUBE — ALL 5 VENTURES)')
r = hd(ws4, r, MC4, ['', 'VENTURE', 'YT CHANNEL', 'DOLPHIN', 'PROXY', 'CADENCE'])
yt = [
    ('V1: PRINTMAXX', 'PrintMaxxer', 'DOLPHIN-01', 'Decodo-1', '1-2 long-form/week. "How I Built X" format.'),
    ('V2: AI/TECH', 'Patch Notes', 'DOLPHIN-03', 'Decodo-2', '1 video/week. Tool reviews, workflow walkthroughs.'),
    ('V3: FAITH', 'Morning Watch', 'DOLPHIN-05', 'Decodo-3', '2-3 Shorts/week + 1 long devotional/month.'),
    ('V4: FITNESS', 'Rep Scheme', 'DOLPHIN-07', 'Decodo-4', '2-3 Shorts/week. Full workout demos monthly.'),
    ('V5: MEME', 'Folders Dump', 'DOLPHIN-09', 'Decodo-5', '3-5 Shorts/week. Viral compilations.'),
]
for idx, (vent, chan, dolphin, proxy, cad) in enumerate(yt):
    bg = BG_R1 if idx % 2 == 0 else BG_R2
    fb(ws4, r, MC4, bg)
    ws4.cell(row=r, column=2, value=vent).font = FW
    ws4.cell(row=r, column=3, value=chan).font = FHL
    ws4.cell(row=r, column=4, value=dolphin).font = FP; ws4.cell(row=r, column=4).alignment = AC
    ws4.cell(row=r, column=5, value=proxy).font = FO; ws4.cell(row=r, column=5).alignment = AC
    ws4.cell(row=r, column=6, value=cad).font = FB; ws4.cell(row=r, column=6).alignment = AL
    for c in range(1, MC4+1): ws4.cell(row=r, column=c).border = BD
    ws4.row_dimensions[r].height = 30
    r += 1


# ============================================================
# SHEET 5: TOOL → VENTURE MATRIX
# ============================================================
ws5 = wb.create_sheet('TOOL MATRIX')
ws5.sheet_properties.tabColor = 'F0C000'
sw(ws5, [3, 24, 14, 14, 14, 14, 14, 3])
MC5 = 8

for r in range(1, 80):
    fb(ws5, r, MC5, BG_DARK)

ws5.merge_cells('B2:G2')
ws5.cell(row=2, column=2, value='TOOL → VENTURE USAGE MATRIX').font = FT
ws5.row_dimensions[2].height = 32
ws5.merge_cells('B3:G3')
ws5.cell(row=3, column=2, value='Which tools each venture uses. Checkmarks = active assignment.').font = FS

r = 5
r = hd(ws5, r, MC5, ['', 'TOOL', 'V1 META', 'V2 AI', 'V3 FAITH', 'V4 FIT', 'V5 MEME'])

tools_matrix = [
    ('ANTI-DETECT & PROXIES', '', '', '', '', ''),
    ('Dolphin Anty (profiles)', 'D01-02', 'D03-04', 'D05-06', 'D07-08', 'D09-10'),
    ('Decodo Residential Proxy', 'Decodo-1', 'Decodo-2', 'Decodo-3', 'Decodo-4', 'Decodo-5'),
    ('SOCIAL SCHEDULING', '', '', '', '', ''),
    ('Tweetlio (X/Twitter)', '3-5/day', '3-5/day', '2-3/day', '2-3/day', '5-10/day'),
    ('Buffer Ch.1 (TikTok)', 'M/W/F', 'T/Th/Sa', '—', '—', '—'),
    ('Buffer Ch.2 (Instagram)', 'M/W/F', '—', 'T/Th/Sa', '—', '—'),
    ('Buffer Ch.3 (TikTok)', '—', '—', '—', 'DAILY', '—'),
    ('YouTube (manual)', '1-2/wk', '1/wk', '2-3 Shorts/wk', '2-3 Shorts/wk', '3-5 Shorts/wk'),
    ('CONTENT CREATION', '', '', '', '', ''),
    ('Canva (design)', 'Thumbnails', 'Screenshots', 'Scripture art', 'Workout cards', 'Meme templates'),
    ('CapCut (video)', 'Demos', 'Walkthroughs', 'Reels', 'Exercise demos', 'Clip edits'),
    ('Leonardo.ai (images)', 'Thumbnails', 'AI screenshots', 'Faith imagery', 'Fitness art', 'Meme enhance'),
    ('ElevenLabs (voice)', 'Voiceover', 'Voiceover', 'Prayer audio', 'Coaching', '—'),
    ('Kling (AI video)', 'Demo clips', 'Demo clips', '—', '—', 'Enhanced clips'),
    ('NEWSLETTER & LANDING', '', '', '', '', ''),
    ('Beehiiv (#1)', 'STACKPRINT', '(shared)', '—', '—', '—'),
    ('Beehiiv (#2)', '—', '—', 'MANNA', '—', '—'),
    ('Substack', '—', '(optional)', '—', 'Rep Scheme', '—'),
    ('Carrd (#1)', 'stackprint.com', '—', '—', '—', '—'),
    ('Carrd (#2)', '—', '—', 'morningmanna.com', '—', '—'),
    ('Carrd (#3)', '—', '—', '—', 'repscheme.com', '—'),
    ('MONETIZATION', '', '', '', '', ''),
    ('Gumroad', 'Templates', 'AI workflows', 'Devotionals', 'Workout PDFs', '—'),
    ('Stripe', 'Payments', 'Payments', 'Payments', 'Payments', '—'),
    ('RevenueCat', '—', 'App IAP', 'PrayerLock', 'StepUnlock', '—'),
    ('Beehiiv Ads', 'STACKPRINT', '—', 'MANNA', '—', '—'),
    ('Affiliate', 'Tool links', 'SaaS affiliate', 'Faith products', 'Supplements', '—'),
    ('SWARM traffic', '—', '—', '—', '—', '→ V1-V4'),
    ('INFRASTRUCTURE', '', '', '', '', ''),
    ('Oracle Cloud VPS', 'Backend', 'API', 'App backend', 'App backend', '—'),
    ('Vercel', 'Landing', 'Landing', 'Landing', 'Landing', '—'),
    ('Cloudflare DNS', 'All domains', 'patchnotes.dev', 'morningmanna.com', 'repscheme.com', '—'),
    ('GA4 Analytics', 'All sites', 'All sites', 'All sites', 'All sites', 'Link tracking'),
    ('PostHog', 'Product', 'Product', '—', '—', '—'),
    ('HubSpot CRM', 'Contacts', 'Shared', 'Shared', 'Shared', '—'),
    ('TrulyInbox', 'Warmup', 'Warmup', 'Warmup', 'Warmup', '—'),
    ('n8n Automation', 'Workflows', 'Workflows', 'Workflows', 'Workflows', '—'),
]

for idx, (tool, v1, v2, v3, v4, v5) in enumerate(tools_matrix):
    if not v1 and not v2 and not v3 and not v4 and not v5:
        r = sh(ws5, r, MC5, tool)
        continue
    bg = BG_R1 if idx % 2 == 0 else BG_R2
    fb(ws5, r, MC5, bg)
    ws5.cell(row=r, column=2, value=tool).font = FW
    for ci, val in enumerate([v1, v2, v3, v4, v5], 3):
        cell = ws5.cell(row=r, column=ci, value=val)
        cell.alignment = AC
        if val == '—':
            cell.font = FSM
        elif val:
            cell.font = FG
            ws5.cell(row=r, column=ci).fill = BG_ACTIVE
        else:
            cell.font = FSM
    for c in range(1, MC5+1): ws5.cell(row=r, column=c).border = BD
    ws5.row_dimensions[r].height = 24
    r += 1


OUTPUT = '/sessions/awesome-nice-brown/mnt/PRINTMAXX_STARTER_KITttttt/PRINTMAXX_INFRA_ASSIGNMENTS.xlsx'
wb.save(OUTPUT)
print(f'Saved to {OUTPUT}')
