import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load tools data
tools = []
with open('/sessions/awesome-nice-brown/mnt/PRINTMAXX_STARTER_KITttttt/LEDGER/TOOLS_SERVICES_MASTER.csv') as f:
    reader = csv.DictReader(f)
    for row in reader:
        tools.append(row)

wb = Workbook()

# ── Color palette (dark/cyan PrintMaxx aesthetic) ──
BG_DARK = PatternFill('solid', fgColor='0D1117')
BG_HEADER = PatternFill('solid', fgColor='0A1628')
BG_SECTION = PatternFill('solid', fgColor='112240')
BG_ROW_ALT1 = PatternFill('solid', fgColor='0D1117')
BG_ROW_ALT2 = PatternFill('solid', fgColor='161B22')
BG_FREE = PatternFill('solid', fgColor='0B3D0B')
BG_TIER50 = PatternFill('solid', fgColor='1A3A4A')
BG_TIER150 = PatternFill('solid', fgColor='2D1B4E')
BG_TIER500 = PatternFill('solid', fgColor='4A1A1A')
BG_ACTIVE = PatternFill('solid', fgColor='0B3D0B')
BG_CYAN_ACCENT = PatternFill('solid', fgColor='00D4FF')
BG_DASHBOARD_CARD = PatternFill('solid', fgColor='161B22')
BG_GOLD = PatternFill('solid', fgColor='3D3500')

FONT_TITLE = Font(name='Arial', size=20, bold=True, color='00D4FF')
FONT_SUBTITLE = Font(name='Arial', size=13, bold=True, color='8B949E')
FONT_HEADER = Font(name='Arial', size=11, bold=True, color='00D4FF')
FONT_SECTION = Font(name='Arial', size=12, bold=True, color='58A6FF')
FONT_BODY = Font(name='Arial', size=10, color='C9D1D9')
FONT_BODY_BRIGHT = Font(name='Arial', size=10, color='FFFFFF')
FONT_FREE = Font(name='Arial', size=10, bold=True, color='3FB950')
FONT_ACTIVE = Font(name='Arial', size=10, bold=True, color='3FB950')
FONT_COST = Font(name='Arial', size=10, color='F0883E')
FONT_URL = Font(name='Arial', size=10, color='58A6FF', underline='single')
FONT_STAT_NUM = Font(name='Arial', size=28, bold=True, color='00D4FF')
FONT_STAT_LABEL = Font(name='Arial', size=10, color='8B949E')
FONT_TIER_TITLE = Font(name='Arial', size=16, bold=True, color='00D4FF')
FONT_GOLD = Font(name='Arial', size=11, bold=True, color='F0C000')
FONT_WHITE_BOLD = Font(name='Arial', size=11, bold=True, color='FFFFFF')
FONT_UPGRADE = Font(name='Arial', size=10, italic=True, color='F0883E')

BORDER_THIN = Border(
    bottom=Side(style='thin', color='21262D')
)
BORDER_SECTION = Border(
    bottom=Side(style='medium', color='00D4FF')
)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')

def style_sheet_tab(ws, color):
    ws.sheet_properties.tabColor = color

def fill_row_bg(ws, row, max_col, fill):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── Categorize tools by tier ──
tiers = {'$0': [], '$50': [], '$150': [], '$500': []}
for t in tools:
    tier = t['budget_tier']
    if tier in tiers:
        tiers[tier].append(t)

# Category display order
CAT_ORDER = [
    'anti_detect', 'proxies', 'accounts', 'email', 'social', 'content',
    'newsletter', 'funnels', 'hosting', 'analytics', 'monetization',
    'automation', 'crm', 'leads', 'linkedin', 'ads', 'growth',
    'hiring', 'community', 'productivity'
]
CAT_LABELS = {
    'anti_detect': 'Anti-Detect Browsers', 'proxies': 'Proxies & IPs',
    'accounts': 'Account Sourcing', 'email': 'Email Infrastructure',
    'social': 'Social Scheduling', 'content': 'Content Creation',
    'newsletter': 'Newsletter Platforms', 'funnels': 'Funnels & Landing Pages',
    'hosting': 'Hosting & Domains', 'analytics': 'Analytics',
    'monetization': 'Monetization', 'automation': 'Automation',
    'crm': 'CRM', 'leads': 'Lead Generation',
    'linkedin': 'LinkedIn Automation', 'ads': 'Paid Ads',
    'growth': 'Growth Services', 'hiring': 'Hiring & VAs',
    'community': 'Community', 'productivity': 'Productivity'
}

def get_cat_tools(tool_list, category):
    return [t for t in tool_list if t['category'] == category]

# ============================================================
# SHEET 1: DASHBOARD OVERVIEW
# ============================================================
ws = wb.active
ws.title = 'DASHBOARD'
style_sheet_tab(ws, '00D4FF')
set_col_widths(ws, [3, 22, 18, 18, 18, 18, 22, 3])

# Dark background
for r in range(1, 50):
    fill_row_bg(ws, r, 8, BG_DARK)

r = 2
ws.merge_cells('B2:G2')
c = ws.cell(row=2, column=2, value='PRINTMAXX INFRASTRUCTURE STACKS')
c.font = FONT_TITLE
c.alignment = ALIGN_LEFT
ws.row_dimensions[2].height = 35

r = 3
ws.merge_cells('B3:G3')
c = ws.cell(row=3, column=2, value='From $0 to Scale — Every Tool You Need at Every Stage')
c.font = FONT_SUBTITLE

# Stats row
r = 5
stats = [
    (f'{len(tools)}', 'Total Tools'),
    (f'{len(tiers["$0"])}', 'Free ($0)'),
    (f'{len(tiers["$50"])}', 'Growth ($50)'),
    (f'{len(tiers["$150"])}', 'Scale ($150)'),
    (f'{len(tiers["$500"])}', 'Enterprise ($500)')
]
for i, (num, label) in enumerate(stats):
    col = 2 + i
    ws.cell(row=5, column=col, value=int(num) if num.isdigit() else num).font = FONT_STAT_NUM
    ws.cell(row=5, column=col).fill = BG_DASHBOARD_CARD
    ws.cell(row=5, column=col).alignment = ALIGN_CENTER
    ws.cell(row=6, column=col, value=label).font = FONT_STAT_LABEL
    ws.cell(row=6, column=col).fill = BG_DASHBOARD_CARD
    ws.cell(row=6, column=col).alignment = ALIGN_CENTER
ws.row_dimensions[5].height = 45
ws.row_dimensions[6].height = 22

# Tier breakdown table
r = 8
headers = ['', 'TIER', 'MONTHLY COST', 'TOOLS INCLUDED', 'CUMULATIVE TOOLS', 'KEY ADDITIONS', 'BEST FOR']
for i, h in enumerate(headers):
    c = ws.cell(row=r, column=i+1, value=h)
    c.font = FONT_HEADER
    c.fill = BG_HEADER
    c.border = BORDER_SECTION
    c.alignment = ALIGN_CENTER
fill_row_bg(ws, r, 8, BG_HEADER)
ws.row_dimensions[r].height = 28

tier_data = [
    ('FREE STACK', '$0/mo', len(tiers['$0']), len(tiers['$0']),
     'Anti-detect, proxies, email warmup, Buffer, Canva, CapCut, newsletters, hosting, analytics, payments',
     'Everyone — absolute zero cost startup'),
    ('GROWTH STACK', '~$50/mo', len(tiers['$50']), len(tiers['$0']) + len(tiers['$50']),
     'Soax mobile proxies, Publer, AccsMarket accounts, SMSPool, Hetzner VPS, dev accounts',
     'First revenue ($100-500/mo) — unlock account farming & mobile proxies'),
    ('SCALE STACK', '~$150/mo', len(tiers['$150']), len(tiers['$0']) + len(tiers['$50']) + len(tiers['$150']),
     'Cold email (Instantly/Smartlead), Google Workspace, residential proxies, AI avatars, Skool',
     'Scaling ($500-2K/mo) — outbound sales & premium content'),
    ('ENTERPRISE', '~$500/mo', len(tiers['$500']), len(tools),
     'Multilogin, Meta/TikTok Ads, Expandi, Apollo, Clay, SWAPD premium accounts',
     'Full operation ($5K+/mo) — paid acquisition & enterprise tooling'),
]
tier_fills = [BG_FREE, BG_TIER50, BG_TIER150, BG_TIER500]
tier_fonts = [
    Font(name='Arial', size=11, bold=True, color='3FB950'),
    Font(name='Arial', size=11, bold=True, color='58D5DB'),
    Font(name='Arial', size=11, bold=True, color='A371F7'),
    Font(name='Arial', size=11, bold=True, color='F85149'),
]

for idx, (name, cost, count, cumul, additions, best) in enumerate(tier_data):
    r = 9 + idx
    ws.cell(row=r, column=2, value=name).font = tier_fonts[idx]
    ws.cell(row=r, column=3, value=cost).font = FONT_COST
    ws.cell(row=r, column=3).alignment = ALIGN_CENTER
    ws.cell(row=r, column=4, value=count).font = FONT_BODY_BRIGHT
    ws.cell(row=r, column=4).alignment = ALIGN_CENTER
    ws.cell(row=r, column=5, value=cumul).font = FONT_BODY_BRIGHT
    ws.cell(row=r, column=5).alignment = ALIGN_CENTER
    ws.cell(row=r, column=6, value=additions).font = FONT_BODY
    ws.cell(row=r, column=6).alignment = ALIGN_LEFT
    ws.cell(row=r, column=7, value=best).font = FONT_BODY
    ws.cell(row=r, column=7).alignment = ALIGN_LEFT
    for c in range(1, 9):
        ws.cell(row=r, column=c).fill = tier_fills[idx]
        ws.cell(row=r, column=c).border = BORDER_THIN
    ws.row_dimensions[r].height = 42

# Active tools section
r = 14
fill_row_bg(ws, r, 8, BG_DARK)
r = 15
ws.merge_cells('B15:G15')
ws.cell(row=15, column=2, value='CURRENTLY ACTIVE TOOLS').font = FONT_SECTION
fill_row_bg(ws, r, 8, BG_DARK)
ws.row_dimensions[r].height = 28

r = 16
active_tools = [t for t in tools if t['status'] == 'ACTIVE']
headers2 = ['', 'TOOL', 'CATEGORY', 'FREE TIER', 'PRIORITY', 'URL', 'NOTES']
for i, h in enumerate(headers2):
    c = ws.cell(row=r, column=i+1, value=h)
    c.font = FONT_HEADER
    c.fill = BG_HEADER
    c.border = BORDER_SECTION

for idx, t in enumerate(active_tools):
    r = 17 + idx
    bg = BG_ROW_ALT1 if idx % 2 == 0 else BG_ROW_ALT2
    fill_row_bg(ws, r, 8, bg)
    ws.cell(row=r, column=2, value=t['tool_name']).font = FONT_BODY_BRIGHT
    ws.cell(row=r, column=3, value=CAT_LABELS.get(t['category'], t['category'])).font = FONT_BODY
    ws.cell(row=r, column=4, value=t['free_tier_detail']).font = FONT_FREE
    ws.cell(row=r, column=5, value=t['priority']).font = FONT_GOLD if t['priority'] == 'HIGHEST' else FONT_BODY
    ws.cell(row=r, column=6, value=t['url']).font = FONT_URL
    ws.cell(row=r, column=7, value=t['notes']).font = FONT_BODY
    ws.cell(row=r, column=7).alignment = ALIGN_LEFT
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = BORDER_THIN
    ws.row_dimensions[r].height = 24


# ============================================================
# TIER SHEETS: $0, $50, $150, $500
# ============================================================
def build_tier_sheet(wb, title, tab_color, tier_key, tier_label, monthly_cost,
                     bg_accent, font_accent, include_previous=None, description=''):
    ws = wb.create_sheet(title)
    style_sheet_tab(ws, tab_color)
    set_col_widths(ws, [3, 24, 16, 14, 24, 20, 30, 3])

    for r in range(1, 120):
        fill_row_bg(ws, r, 8, BG_DARK)

    # Title
    ws.merge_cells('B2:G2')
    ws.cell(row=2, column=2, value=f'{tier_label}').font = FONT_TIER_TITLE
    ws.row_dimensions[2].height = 32
    ws.merge_cells('B3:G3')
    ws.cell(row=3, column=2, value=description).font = FONT_SUBTITLE

    # Gather all tools for this tier (cumulative)
    all_tier_tools = []
    tier_order = ['$0', '$50', '$150', '$500']
    current_idx = tier_order.index(tier_key)
    for i in range(current_idx + 1):
        all_tier_tools.extend(tiers[tier_order[i]])

    # NEW at this tier
    new_at_tier = tiers[tier_key]

    # Stats
    ws.cell(row=5, column=2, value=f'{len(all_tier_tools)} tools').font = FONT_STAT_NUM
    ws.cell(row=5, column=2).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=5, column=3, value=f'at {monthly_cost}').font = Font(name='Arial', size=16, color='F0883E')
    ws.row_dimensions[5].height = 40

    if tier_key != '$0':
        ws.cell(row=5, column=5, value=f'+{len(new_at_tier)} new tools at this tier').font = font_accent
        ws.cell(row=5, column=5).alignment = Alignment(horizontal='left', vertical='center')

    # Build by category
    r = 7
    for cat in CAT_ORDER:
        cat_tools_all = [t for t in all_tier_tools if t['category'] == cat]
        if not cat_tools_all:
            continue

        # Category header
        ws.merge_cells(f'B{r}:G{r}')
        ws.cell(row=r, column=2, value=CAT_LABELS.get(cat, cat).upper()).font = FONT_SECTION
        fill_row_bg(ws, r, 8, BG_SECTION)
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = BORDER_SECTION
        ws.row_dimensions[r].height = 26
        r += 1

        # Column headers
        col_headers = ['', 'TOOL', 'SUB-CATEGORY', 'FREE TIER', 'WHAT YOU GET', 'PAID UPGRADE', 'NOTES']
        for i, h in enumerate(col_headers):
            c = ws.cell(row=r, column=i+1, value=h)
            c.font = FONT_HEADER
            c.fill = BG_HEADER
        ws.row_dimensions[r].height = 22
        r += 1

        for idx, t in enumerate(cat_tools_all):
            bg = BG_ROW_ALT1 if idx % 2 == 0 else BG_ROW_ALT2
            is_new = t in new_at_tier and tier_key != '$0'

            # Mark new-at-tier tools
            tool_name = t['tool_name']
            if is_new:
                tool_name = f"+ {tool_name}"
                bg = bg_accent

            fill_row_bg(ws, r, 8, bg)
            ws.cell(row=r, column=2, value=tool_name).font = font_accent if is_new else FONT_BODY_BRIGHT
            ws.cell(row=r, column=3, value=t['sub_category']).font = FONT_BODY
            ws.cell(row=r, column=4, value='YES' if t['free_tier'] == 'YES' else 'NO').font = FONT_FREE if t['free_tier'] == 'YES' else Font(name='Arial', size=10, color='F85149')
            ws.cell(row=r, column=4).alignment = ALIGN_CENTER
            ws.cell(row=r, column=5, value=t['free_tier_detail']).font = FONT_BODY
            ws.cell(row=r, column=5).alignment = ALIGN_LEFT
            ws.cell(row=r, column=6, value=t['paid_from']).font = FONT_COST
            ws.cell(row=r, column=7, value=t['notes']).font = FONT_BODY
            ws.cell(row=r, column=7).alignment = ALIGN_LEFT
            for c in range(1, 8):
                ws.cell(row=r, column=c).border = BORDER_THIN
            ws.row_dimensions[r].height = 24
            r += 1
        r += 1  # gap between categories

    return ws

build_tier_sheet(wb, 'FREE ($0)', '3FB950', '$0', 'FREE STACK — $0/mo',
    '$0/mo', BG_FREE,
    Font(name='Arial', size=10, bold=True, color='3FB950'),
    description='The ultimate zero-cost setup. 47 tools, every category covered. Launch with nothing.')

build_tier_sheet(wb, 'GROWTH ($50)', '58D5DB', '$50', 'GROWTH STACK — ~$50/mo',
    '~$50/mo', BG_TIER50,
    Font(name='Arial', size=10, bold=True, color='58D5DB'),
    description='First revenue unlocked. Mobile proxies, account farming, dev accounts. +11 tools.')

build_tier_sheet(wb, 'SCALE ($150)', 'A371F7', '$150', 'SCALE STACK — ~$150/mo',
    '~$150/mo', BG_TIER150,
    Font(name='Arial', size=10, bold=True, color='A371F7'),
    description='Outbound sales engine + premium content. Cold email, Google Workspace, AI avatars. +18 tools.')

build_tier_sheet(wb, 'ENTERPRISE ($500)', 'F85149', '$500', 'ENTERPRISE STACK — ~$500/mo',
    '~$500/mo', BG_TIER500,
    Font(name='Arial', size=10, bold=True, color='F85149'),
    description='Full operation. Paid ads, enterprise anti-detect, LinkedIn automation, lead enrichment. +8 tools.')


# ============================================================
# SHEET: VENTURE STACKS (which tools per venture type)
# ============================================================
ws_v = wb.create_sheet('VENTURE STACKS')
style_sheet_tab(ws_v, 'F0C000')
set_col_widths(ws_v, [3, 24, 16, 14, 20, 18, 30, 3])

for r in range(1, 100):
    fill_row_bg(ws_v, r, 8, BG_DARK)

ws_v.merge_cells('B2:G2')
ws_v.cell(row=2, column=2, value='VENTURE-SPECIFIC STACKS').font = FONT_TIER_TITLE
ws_v.row_dimensions[2].height = 32
ws_v.merge_cells('B3:G3')
ws_v.cell(row=3, column=2, value='Curated tool stacks for each venture type. Tools marked ALL appear in every stack.').font = FONT_SUBTITLE

venture_types = [
    ('COLD_OUTBOUND', 'Cold Outbound / B2B Sales', '58D5DB',
     'Lead gen, cold email campaigns, LinkedIn automation, CRM'),
    ('CONTENT_MEDIA', 'Content & Media Empire', 'A371F7',
     'Social scheduling, AI content creation, music, video, growth hacking'),
    ('AI_PERSONAS', 'AI Personas / UGC Factory', 'F0883E',
     'Anti-detect browsers, AI images, AI avatars, AI voice, mobile fingerprints'),
    ('APPS_SOFTWARE', 'Apps & Software', '3FB950',
     'Hosting, analytics, app stores, UI design, subscription management'),
    ('INFO_EDUCATION', 'Info Products & Education', 'F0C000',
     'All-in-one funnels, communities, courses, newsletters, landing pages'),
]

r = 5
for vtype, vlabel, vcolor, vdesc in venture_types:
    # Section header
    ws_v.merge_cells(f'B{r}:G{r}')
    ws_v.cell(row=r, column=2, value=vlabel.upper()).font = Font(name='Arial', size=13, bold=True, color=vcolor)
    fill_row_bg(ws_v, r, 8, BG_SECTION)
    for c in range(1, 9):
        ws_v.cell(row=r, column=c).border = BORDER_SECTION
    ws_v.row_dimensions[r].height = 28
    r += 1

    ws_v.merge_cells(f'B{r}:G{r}')
    ws_v.cell(row=r, column=2, value=vdesc).font = FONT_SUBTITLE
    r += 1

    # Column headers
    for i, h in enumerate(['', 'TOOL', 'CATEGORY', 'TIER', 'FREE DETAIL', 'PRIORITY', 'NOTES']):
        ws_v.cell(row=r, column=i+1, value=h).font = FONT_HEADER
        ws_v.cell(row=r, column=i+1).fill = BG_HEADER
    r += 1

    # Tools for this venture (ALL + specific)
    venture_tools = [t for t in tools if t['venture_fit'] in (vtype, 'ALL')]
    venture_tools.sort(key=lambda x: (
        ['HIGHEST', 'HIGH', 'MEDIUM', 'LOW'].index(x['priority']) if x['priority'] in ['HIGHEST', 'HIGH', 'MEDIUM', 'LOW'] else 4,
        x['budget_tier']
    ))

    for idx, t in enumerate(venture_tools):
        bg = BG_ROW_ALT1 if idx % 2 == 0 else BG_ROW_ALT2
        fill_row_bg(ws_v, r, 8, bg)
        ws_v.cell(row=r, column=2, value=t['tool_name']).font = FONT_BODY_BRIGHT
        ws_v.cell(row=r, column=3, value=CAT_LABELS.get(t['category'], t['category'])).font = FONT_BODY
        tier_val = t['budget_tier']
        tier_font_map = {'$0': FONT_FREE, '$50': Font(name='Arial', size=10, color='58D5DB'),
                         '$150': Font(name='Arial', size=10, color='A371F7'), '$500': Font(name='Arial', size=10, color='F85149')}
        ws_v.cell(row=r, column=4, value=tier_val).font = tier_font_map.get(tier_val, FONT_BODY)
        ws_v.cell(row=r, column=4).alignment = ALIGN_CENTER
        ws_v.cell(row=r, column=5, value=t['free_tier_detail']).font = FONT_BODY
        ws_v.cell(row=r, column=6, value=t['priority']).font = FONT_GOLD if t['priority'] == 'HIGHEST' else FONT_BODY
        ws_v.cell(row=r, column=6).alignment = ALIGN_CENTER
        ws_v.cell(row=r, column=7, value=t['notes']).font = FONT_BODY
        ws_v.cell(row=r, column=7).alignment = ALIGN_LEFT
        for c in range(1, 8):
            ws_v.cell(row=r, column=c).border = BORDER_THIN
        ws_v.row_dimensions[r].height = 22
        r += 1

    r += 2  # gap between ventures


# ============================================================
# SHEET: UPGRADE PATH (decision matrix)
# ============================================================
ws_u = wb.create_sheet('UPGRADE PATH')
style_sheet_tab(ws_u, 'F0883E')
set_col_widths(ws_u, [3, 20, 16, 30, 30, 20, 3])

for r in range(1, 30):
    fill_row_bg(ws_u, r, 7, BG_DARK)

ws_u.merge_cells('B2:F2')
ws_u.cell(row=2, column=2, value='WHEN TO UPGRADE — DECISION MATRIX').font = FONT_TIER_TITLE
ws_u.row_dimensions[2].height = 32
ws_u.merge_cells('B3:F3')
ws_u.cell(row=3, column=2, value='Revenue milestones that trigger each tier upgrade').font = FONT_SUBTITLE

r = 5
for i, h in enumerate(['', 'MILESTONE', 'UPGRADE TO', 'WHAT YOU UNLOCK', 'WHY IT MATTERS', 'ROI TRIGGER']):
    ws_u.cell(row=r, column=i+1, value=h).font = FONT_HEADER
    ws_u.cell(row=r, column=i+1).fill = BG_HEADER
    ws_u.cell(row=r, column=i+1).border = BORDER_SECTION
ws_u.row_dimensions[r].height = 26

upgrades = [
    ('$0 MRR\n(Day 1)', 'FREE STACK',
     'Anti-detect (10 profiles), free proxies, email warmup, Buffer, Canva, CapCut, Beehiiv, Oracle VPS, Stripe, Gumroad',
     'Launch everything with zero risk. No credit card needed anywhere. Every category covered.',
     'Immediate'),
    ('$100-300 MRR', 'GROWTH ($50)',
     'Soax mobile proxies, AccsMarket accounts, SMSPool verification, Publer scheduling, Hetzner VPS backup',
     'Mobile proxies = account safety on IG/TikTok. Account marketplace = faster scaling. VPS backup = reliability.',
     '3-5x ROI on proxy spend'),
    ('$500-1K MRR', 'SCALE ($150)',
     'Instantly/Smartlead cold email, Google Workspace, residential proxies, D-ID/HeyGen AI avatars, Skool community',
     'Cold email is the highest-ROI B2B channel. AI avatars = UGC at scale. Community = recurring revenue.',
     '10-20x ROI on cold email'),
    ('$2-5K MRR', 'ENTERPRISE ($500)',
     'Meta/TikTok Ads ($100 budgets), Multilogin, Expandi LinkedIn, Apollo leads, Clay enrichment',
     'Paid ads compound organic. Enterprise anti-detect for 100+ profiles. Full outbound stack.',
     '3-10x ROAS on ads'),
]

for idx, (milestone, tier, unlock, why, roi) in enumerate(upgrades):
    r = 6 + idx
    bg = [BG_FREE, BG_TIER50, BG_TIER150, BG_TIER500][idx]
    fill_row_bg(ws_u, r, 7, bg)
    ws_u.cell(row=r, column=2, value=milestone).font = FONT_BODY_BRIGHT
    ws_u.cell(row=r, column=2).alignment = ALIGN_CENTER
    ws_u.cell(row=r, column=3, value=tier).font = [
        Font(name='Arial', size=10, bold=True, color='3FB950'),
        Font(name='Arial', size=10, bold=True, color='58D5DB'),
        Font(name='Arial', size=10, bold=True, color='A371F7'),
        Font(name='Arial', size=10, bold=True, color='F85149'),
    ][idx]
    ws_u.cell(row=r, column=4, value=unlock).font = FONT_BODY
    ws_u.cell(row=r, column=4).alignment = ALIGN_LEFT
    ws_u.cell(row=r, column=5, value=why).font = FONT_BODY
    ws_u.cell(row=r, column=5).alignment = ALIGN_LEFT
    ws_u.cell(row=r, column=6, value=roi).font = FONT_GOLD
    ws_u.cell(row=r, column=6).alignment = ALIGN_CENTER
    for c in range(1, 7):
        ws_u.cell(row=r, column=c).border = BORDER_THIN
    ws_u.row_dimensions[r].height = 65

# Quick rules
r = 12
ws_u.merge_cells('B12:F12')
ws_u.cell(row=12, column=2, value='GOLDEN RULES').font = FONT_SECTION
fill_row_bg(ws_u, 12, 7, BG_SECTION)

rules = [
    'Never upgrade until the current tier is generating revenue that justifies the next tier cost',
    'Free tier tools are not "starter" — many are permanent (Buffer, Canva, GA4, Stripe stay forever)',
    'Proxy spend is the first justified cost — account bans cost more than proxies',
    'Cold email has the highest single-tool ROI — one client pays for 6+ months of tooling',
    'Paid ads are LAST — only after organic channels are proven and profitable',
]
for i, rule in enumerate(rules):
    r = 13 + i
    fill_row_bg(ws_u, r, 7, BG_DARK)
    ws_u.cell(row=r, column=2, value=f'{i+1}.').font = FONT_GOLD
    ws_u.cell(row=r, column=2).alignment = ALIGN_CENTER
    ws_u.merge_cells(f'C{r}:F{r}')
    ws_u.cell(row=r, column=3, value=rule).font = FONT_BODY
    ws_u.cell(row=r, column=3).alignment = ALIGN_LEFT
    ws_u.row_dimensions[r].height = 24


# ============================================================
# SHEET: FULL MASTER LIST (raw data, sortable)
# ============================================================
ws_m = wb.create_sheet('FULL MASTER')
style_sheet_tab(ws_m, '8B949E')
headers_master = ['tool_id', 'tool_name', 'category', 'sub_category', 'free_tier',
                  'free_tier_detail', 'paid_from', 'status', 'priority', 'budget_tier',
                  'venture_fit', 'url', 'notes']
header_labels = ['ID', 'TOOL', 'CATEGORY', 'SUB-CATEGORY', 'FREE?',
                 'FREE DETAIL', 'PAID FROM', 'STATUS', 'PRIORITY', 'TIER',
                 'VENTURE FIT', 'URL', 'NOTES']
set_col_widths(ws_m, [8, 22, 16, 16, 8, 28, 14, 12, 12, 10, 18, 28, 36])

for i, h in enumerate(header_labels):
    c = ws_m.cell(row=1, column=i+1, value=h)
    c.font = FONT_HEADER
    c.fill = BG_HEADER
    c.border = BORDER_SECTION
    c.alignment = ALIGN_CENTER
ws_m.row_dimensions[1].height = 26

for idx, t in enumerate(tools):
    r = idx + 2
    bg = BG_ROW_ALT1 if idx % 2 == 0 else BG_ROW_ALT2
    for i, key in enumerate(headers_master):
        cell = ws_m.cell(row=r, column=i+1, value=t[key])
        cell.font = FONT_BODY
        cell.fill = bg
        cell.border = BORDER_THIN
        if key == 'status' and t[key] == 'ACTIVE':
            cell.font = FONT_ACTIVE
        elif key == 'priority' and t[key] == 'HIGHEST':
            cell.font = FONT_GOLD
        elif key == 'free_tier' and t[key] == 'YES':
            cell.font = FONT_FREE
        elif key == 'url':
            cell.font = FONT_URL
        elif key == 'budget_tier':
            tier_font_map = {'$0': FONT_FREE, '$50': Font(name='Arial', size=10, color='58D5DB'),
                             '$150': Font(name='Arial', size=10, color='A371F7'), '$500': Font(name='Arial', size=10, color='F85149')}
            cell.font = tier_font_map.get(t[key], FONT_BODY)
            cell.alignment = ALIGN_CENTER
    ws_m.row_dimensions[r].height = 22

# Add autofilter
ws_m.auto_filter.ref = f'A1:M{len(tools)+1}'

# Save
OUTPUT = '/sessions/awesome-nice-brown/mnt/PRINTMAXX_STARTER_KITttttt/PRINTMAXX_INFRA_STACKS.xlsx'
wb.save(OUTPUT)
print(f'Saved to {OUTPUT}')
