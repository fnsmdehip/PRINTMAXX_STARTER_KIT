import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# === THEME ===
BG = PatternFill('solid', fgColor='0D1117')
CYAN = Font(name='Arial', color='00D4FF', bold=True, size=12)
CYAN_SM = Font(name='Arial', color='00D4FF', bold=True, size=10)
WHITE = Font(name='Arial', color='FFFFFF', size=10)
WHITE_B = Font(name='Arial', color='FFFFFF', bold=True, size=10)
GOLD = Font(name='Arial', color='FFD700', bold=True, size=10)
GREEN = Font(name='Arial', color='00FF88', size=10)
GREEN_B = Font(name='Arial', color='00FF88', bold=True, size=10)
RED = Font(name='Arial', color='FF4444', size=10)
HEADER_FILL = PatternFill('solid', fgColor='1A2332')
ROW_ALT = PatternFill('solid', fgColor='151B23')
GOLD_FILL = PatternFill('solid', fgColor='2A2000')
GREEN_FILL = PatternFill('solid', fgColor='0A2A15')
THIN = Border(
    bottom=Side(style='thin', color='333333')
)

def style_sheet(ws, cols, widths):
    ws.sheet_properties.tabColor = '00D4FF'
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=cols):
        for cell in row:
            cell.fill = BG
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = THIN

def header_row(ws, row, headers, fonts=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = fonts[i-1] if fonts else CYAN_SM
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# ============================================================
# SHEET 1: FREELANCE SERVICE CATALOG
# ============================================================
ws1 = wb.active
ws1.title = 'SERVICE CATALOG'

ws1.cell(row=2, column=2, value='PRINTMAXX FREELANCE ARBITRAGE — CLAUDE CODE MAX SERVICES').font = Font(name='Arial', color='00D4FF', bold=True, size=16)
ws1.merge_cells('B2:J2')
ws1.cell(row=3, column=2, value='Near-infinite Claude Code usage = every gig is pure margin. 5-min delivery on $50-500 services.').font = WHITE
ws1.merge_cells('B3:J3')

headers = ['SVC ID', 'SERVICE NAME', 'CATEGORY', 'FIVERR PRICE', 'UPWORK RATE', 'DELIVERY TIME', 'CLAUDE CODE TIME', 'MARGIN %', 'DEMAND LEVEL', 'GIG DESCRIPTION (TEMPLATE)']
header_row(ws1, 5, headers)

services = [
    ('SVC01', 'Landing Page (HTML/CSS/JS)', 'WEB DEV', '$75-150', '$40/hr', '24hr', '10-20 min', '95%+', 'VERY HIGH',
     'I will create a modern, responsive landing page optimized for conversions. Mobile-friendly, fast-loading, with SEO basics. Includes hero, features, testimonials, CTA sections. Unlimited revisions.'),
    ('SVC02', 'Full Website (5-10 pages)', 'WEB DEV', '$200-500', '$50/hr', '2-3 days', '30-60 min', '95%+', 'VERY HIGH',
     'Professional multi-page website with responsive design, contact forms, about page, services page, and blog structure. Clean code, fast loading, SEO-optimized.'),
    ('SVC03', 'React/Next.js Web App', 'WEB DEV', '$300-800', '$60/hr', '3-5 days', '1-3 hrs', '90%+', 'HIGH',
     'Custom React or Next.js web application with modern UI, state management, API integration, and responsive design. Full source code delivered.'),
    ('SVC04', 'Chrome Extension', 'DEV', '$150-400', '$50/hr', '2-3 days', '30-60 min', '95%+', 'HIGH',
     'Custom Chrome extension with popup UI, content scripts, background processing. Published to Chrome Web Store or delivered as source. Perfect for productivity tools, scrapers, UI enhancers.'),
    ('SVC05', 'Discord Bot (Python/JS)', 'DEV', '$100-300', '$45/hr', '1-2 days', '20-45 min', '95%+', 'HIGH',
     'Custom Discord bot with slash commands, moderation, auto-roles, welcome messages, polls, or custom integrations. Hosted instructions included.'),
    ('SVC06', 'Telegram Bot', 'DEV', '$100-250', '$45/hr', '1-2 days', '20-40 min', '95%+', 'HIGH',
     'Telegram bot for notifications, group management, payments, or custom workflows. Python-based with inline keyboards and webhook support.'),
    ('SVC07', 'Python Script / Automation', 'AUTOMATION', '$50-200', '$40/hr', '12-24hr', '5-20 min', '97%+', 'VERY HIGH',
     'Custom Python script for any task: web scraping, data processing, API integration, file conversion, automation. Clean, documented code.'),
    ('SVC08', 'Web Scraper (Any Site)', 'AUTOMATION', '$75-250', '$45/hr', '12-24hr', '10-30 min', '95%+', 'VERY HIGH',
     'Custom web scraper using Playwright/BeautifulSoup/Selenium. Handles pagination, login walls, dynamic content. Outputs CSV/JSON/Excel.'),
    ('SVC09', 'API Integration / Wrapper', 'DEV', '$100-300', '$50/hr', '1-2 days', '15-45 min', '95%+', 'HIGH',
     'Connect any two APIs, build REST API wrappers, create webhook handlers, or integrate third-party services. Full documentation included.'),
    ('SVC10', 'Data Analysis + Visualization', 'DATA', '$75-250', '$45/hr', '12-24hr', '10-30 min', '95%+', 'HIGH',
     'Data cleaning, analysis, and visualization using Python (pandas, matplotlib, plotly). Deliver insights report with charts. Excel/CSV input accepted.'),
    ('SVC11', 'Excel/Google Sheets Automation', 'DATA', '$50-150', '$35/hr', '12-24hr', '5-15 min', '97%+', 'VERY HIGH',
     'Complex spreadsheet formulas, macros, VBA scripts, Google Apps Script. Dashboards, automated reports, data validation. Any complexity.'),
    ('SVC12', 'Database Design + Setup', 'DEV', '$100-300', '$50/hr', '1-2 days', '20-45 min', '95%+', 'MEDIUM',
     'PostgreSQL/MySQL/MongoDB database design, schema creation, migrations, seed data, and CRUD API. Optimized queries included.'),
    ('SVC13', 'Mobile App Prototype (React Native)', 'MOBILE', '$300-800', '$60/hr', '3-5 days', '2-4 hrs', '90%+', 'HIGH',
     'Cross-platform mobile app prototype in React Native or Expo. Working screens, navigation, API calls, local storage. iOS + Android from one codebase.'),
    ('SVC14', 'WordPress Customization', 'WEB DEV', '$50-200', '$40/hr', '12-24hr', '15-30 min', '95%+', 'VERY HIGH',
     'WordPress theme customization, plugin development, speed optimization, WooCommerce setup, or custom post types. Fix any WP issue.'),
    ('SVC15', 'SEO Audit + Technical Fix', 'SEO', '$100-300', '$50/hr', '1-2 days', '20-40 min', '95%+', 'HIGH',
     'Comprehensive technical SEO audit with actionable fixes. Site speed, meta tags, schema markup, internal linking, XML sitemap optimization. Detailed report.'),
    ('SVC16', 'Email Template (HTML)', 'MARKETING', '$50-150', '$35/hr', '12-24hr', '5-15 min', '97%+', 'HIGH',
     'Responsive HTML email template for newsletters, drip campaigns, or transactional emails. Works in Gmail, Outlook, Apple Mail. Mailchimp/SendGrid ready.'),
    ('SVC17', 'Shopify Store Setup', 'ECOMM', '$200-500', '$50/hr', '2-3 days', '1-2 hrs', '90%+', 'VERY HIGH',
     'Complete Shopify store setup: theme customization, product pages, collections, checkout optimization, shipping rules, payment gateway. Launch-ready.'),
    ('SVC18', 'Notion Workspace Setup', 'PRODUCTIVITY', '$50-200', '$35/hr', '12-24hr', '10-20 min', '95%+', 'HIGH',
     'Custom Notion workspace with databases, dashboards, project trackers, CRM, content calendar, or operations hub. Templates included.'),
    ('SVC19', 'CLI Tool / Developer Tool', 'DEV', '$100-250', '$50/hr', '1-2 days', '20-45 min', '95%+', 'MEDIUM',
     'Custom CLI tool in Python/Node/Go for any workflow: deployment, file processing, code generation, testing, or DevOps automation.'),
    ('SVC20', 'AI/LLM Integration', 'AI', '$150-400', '$60/hr', '2-3 days', '30-60 min', '95%+', 'VERY HIGH',
     'Integrate OpenAI/Claude/Llama into your app. Chatbots, content generators, document analyzers, RAG systems, or custom AI features. Full prompt engineering.'),
    ('SVC21', 'Stripe/Payment Integration', 'DEV', '$100-300', '$50/hr', '1-2 days', '20-40 min', '95%+', 'HIGH',
     'Stripe integration: subscriptions, one-time payments, checkout sessions, webhooks, customer portal. Or PayPal/LemonSqueezy integration.'),
    ('SVC22', 'Bug Fix / Code Review', 'DEV', '$25-100', '$40/hr', '4-12hr', '5-15 min', '97%+', 'VERY HIGH',
     'Fix any bug in your codebase. JavaScript, Python, React, Node, etc. Includes root cause analysis and prevention recommendations.'),
    ('SVC23', 'Airtable/Zapier Automation', 'AUTOMATION', '$50-200', '$40/hr', '12-24hr', '10-25 min', '95%+', 'HIGH',
     'Build complex Airtable bases with automations, or Zapier/Make.com workflows connecting your tools. Replace manual processes with automated pipelines.'),
    ('SVC24', 'Technical Writing / Documentation', 'CONTENT', '$75-200', '$40/hr', '1-2 days', '15-30 min', '95%+', 'MEDIUM',
     'API documentation, README files, developer guides, architecture docs, or user manuals. Clear, professional technical writing.'),
    ('SVC25', 'PDF Report / Business Document', 'CONTENT', '$50-150', '$35/hr', '12-24hr', '10-20 min', '95%+', 'HIGH',
     'Professional business reports, pitch decks in PDF, data analysis reports, or formatted documents. Clean design with charts and branding.'),
    ('SVC26', 'Supabase/Firebase Backend', 'DEV', '$150-400', '$55/hr', '2-3 days', '30-60 min', '95%+', 'HIGH',
     'Complete backend setup with Supabase or Firebase: auth, database, storage, realtime, edge functions. Connected to your frontend.'),
    ('SVC27', 'Tailwind CSS Website Redesign', 'WEB DEV', '$100-300', '$45/hr', '1-2 days', '20-45 min', '95%+', 'HIGH',
     'Modernize your website with Tailwind CSS. Clean, responsive, accessible design. Convert any old site to modern stack.'),
    ('SVC28', 'N8N/Make.com Workflow', 'AUTOMATION', '$75-250', '$45/hr', '1-2 days', '15-30 min', '95%+', 'HIGH',
     'Build complex automation workflows in n8n or Make.com. Connect APIs, process data, trigger actions. Replace manual tasks with reliable automations.'),
    ('SVC29', 'Google Apps Script', 'AUTOMATION', '$50-150', '$40/hr', '12-24hr', '5-15 min', '97%+', 'VERY HIGH',
     'Custom Google Apps Scripts for Sheets, Forms, Docs, Gmail. Automated reports, email triggers, data processing, custom menus.'),
    ('SVC30', 'TypeScript/Node.js Backend', 'DEV', '$200-500', '$55/hr', '2-4 days', '1-2 hrs', '90%+', 'HIGH',
     'Production-ready backend in TypeScript/Node.js with Express or Fastify. REST or GraphQL API, auth, database, tests, Docker deployment.'),
]

for i, svc in enumerate(services):
    row = 6 + i
    for j, val in enumerate(svc, 1):
        c = ws1.cell(row=row, column=j, value=val)
        c.font = WHITE if j > 1 else GOLD
        if i % 2 == 1:
            c.fill = ROW_ALT
        if j == 8:
            c.font = GREEN_B
        if j == 9:
            c.font = GREEN_B if 'VERY' in str(val) else GOLD

# Summary row
summary_row = 6 + len(services) + 2
ws1.cell(row=summary_row, column=2, value='TOTAL SERVICES: 30').font = CYAN_SM
ws1.cell(row=summary_row+1, column=2, value='AVG MARGIN: 95%+ (Claude does the work in minutes, client pays for days of delivery)').font = GREEN_B
ws1.cell(row=summary_row+2, column=2, value='KEY INSIGHT: Price at 25th percentile (not cheapest = suspicious). Compete on SPEED and QUALITY, not price.').font = GOLD
ws1.cell(row=summary_row+3, column=2, value='VOLUME TARGET: 20-30 active gigs/week = $1,500-5,000/week = $6K-20K/month pure margin').font = GREEN_B

style_sheet(ws1, 10, [8, 30, 12, 12, 12, 12, 14, 10, 12, 55])

# ============================================================
# SHEET 2: PLATFORM STRATEGY
# ============================================================
ws2 = wb.create_sheet('PLATFORM STRATEGY')

ws2.cell(row=2, column=2, value='FREELANCE PLATFORM DEPLOYMENT MAP').font = Font(name='Arial', color='00D4FF', bold=True, size=16)
ws2.merge_cells('B2:H2')
ws2.cell(row=3, column=2, value='List on ALL platforms simultaneously. Each platform = new deal flow channel.').font = WHITE
ws2.merge_cells('B3:H3')

headers2 = ['PLATFORM', 'MONTHLY FEE', 'COMMISSION', 'BEST FOR', 'ACCOUNT STRATEGY', 'LISTING PRIORITY', 'CONVERSION TIPS', 'AUTOMATION POTENTIAL']
header_row(ws2, 5, headers2)

platforms = [
    ('Fiverr', '$0', '20%', 'Volume small gigs ($25-300)',
     'Create 7-10 gigs across categories. Use Fiverr Pro application after 10 reviews. Specialize each gig listing.',
     'TOP PRIORITY',
     'Use video intros (AI-generated). Offer 3 packages (Basic/Standard/Premium). Fast response time (<1hr). Include FAQ. Add gig extras (+$25 rush, +$50 source code).',
     'FULL AUTO: Claude Code receives requirements via API/email → builds → delivers ZIP. Human reviews final output only.'),
    ('Upwork', '$0 (Connects: $0.15/ea)', '10-20%', 'Higher-value projects ($200-2K+)',
     'Create Top Rated profile. Apply to 10-15 jobs/day. Use Connects strategically on <5 proposals jobs. Specialized profile for each category.',
     'TOP PRIORITY',
     'Personalized proposals (LLM writes). Address client pain points specifically. Include relevant portfolio piece. Offer free mini-audit as hook.',
     'SEMI-AUTO: LLM writes personalized proposals from job description. Claude Code builds deliverable. Human submits proposal + final review.'),
    ('Freelancer.com', '$0', '10-13%', 'Contest-based + projects',
     'Bid on fixed-price projects. Enter design/code contests (easy wins with Claude Code speed). Build reputation through contest wins.',
     'MEDIUM',
     'Enter every relevant contest (Claude Code can produce entries in minutes). Low competition on technical contests. Bid slightly below average.',
     'SEMI-AUTO: Monitor contests via RSS. Claude Code generates entries. Human submits.'),
    ('PeoplePerHour', '$0', '15-20%', 'UK/EU market projects',
     'Create Hourlies (fixed-price offers). Target UK small businesses. Different market than Fiverr/Upwork = less competition for US-based sellers.',
     'MEDIUM',
     'Hourlies convert best (pre-packaged services). Focus on web dev and automation. UK businesses pay premium for quality.',
     'SEMI-AUTO: Hourlies are set-and-forget listings. Fulfillment via Claude Code.'),
    ('Contra', '$0', '0% (!)', 'Commission-free freelancing',
     'Zero commission = keep 100%. Newer platform = less competition. Build portfolio with case studies. Target startup founders.',
     'HIGH PRIORITY',
     'Zero commission is your pitch ("I keep 100% so I price fairly"). Portfolio-first platform. Connect with startups directly.',
     'SEMI-AUTO: Less volume but higher margin per gig.'),
    ('Guru.com', '$0-$40/mo', '5-9%', 'Lowest commission marketplace',
     'Lowest fees in the industry. Create WorkRooms for repeat clients. Good for ongoing relationships.',
     'LOW',
     'Low commission = price lower than Fiverr for same work. WorkRooms for repeat clients. Less competition overall.',
     'SEMI-AUTO: Standard fulfillment via Claude Code.'),
    ('Toptal', '$0', '~30-40% (they take margin)', 'Premium clients ($100+/hr)',
     'Apply to Toptal network (screening process). If accepted, access to Fortune 500 clients. Highest rates in industry.',
     'LONG-TERM',
     'Must pass technical screening. Once in, jobs come to you. $100-200/hr rates. Only pursue after establishing track record.',
     'LOW AUTO: Premium clients expect communication. Claude Code handles deliverables. Human manages relationship.'),
    ('LinkedIn Services', '$0', '0%', 'B2B direct relationships',
     'Enable Services section on LinkedIn profile. Post case studies weekly. Target small business owners in feed.',
     'HIGH PRIORITY',
     'Zero fees. Direct relationships. Post "I built this in 30 minutes" case studies. DM prospects who engage with content.',
     'SEMI-AUTO: Content is LLM-generated. Claude Code builds case study deliverables. Human manages DMs.'),
    ('Twitter/X Services', '$0', '0%', 'Direct inbound from content',
     'Tweet about builds ("Just built a Chrome extension in 15 minutes for a client"). Pin services tweet. DM-based intake.',
     'HIGH PRIORITY',
     'Build in public content drives inbound. "I will build your [X] for $[Y]" tweets. Show speed/quality combo. 0% commission.',
     'FULL AUTO: Auto-post builds. Claude Code creates portfolio content. Human reviews DM inquiries.'),
    ('Reddit (r/forhire)', '$0', '0%', 'Direct hiring, tech-savvy clients',
     'Post [For Hire] threads weekly. r/forhire, r/slavelabour, r/programminghelp. Show portfolio + pricing.',
     'MEDIUM',
     'Tech audience respects speed. Show Claude Code outputs (anonymized). Price competitively. Fast turnaround = reviews.',
     'SEMI-AUTO: LLM writes posts. Claude Code fulfills. Human manages inbox.'),
]

for i, p in enumerate(platforms):
    row = 6 + i
    for j, val in enumerate(p, 1):
        c = ws2.cell(row=row, column=j, value=val)
        c.font = WHITE if j > 1 else GOLD
        if i % 2 == 1:
            c.fill = ROW_ALT
        if j == 6:
            if 'TOP' in str(val):
                c.font = GREEN_B
            elif 'HIGH' in str(val):
                c.font = Font(name='Arial', color='00FF88', size=10)
            else:
                c.font = GOLD

style_sheet(ws2, 8, [18, 14, 12, 28, 45, 14, 45, 45])

# ============================================================
# SHEET 3: PRICING STRATEGY + CONVERSION
# ============================================================
ws3 = wb.create_sheet('PRICING + CONVERSION')

ws3.cell(row=2, column=2, value='PRICING STRATEGY — NOT CHEAPEST, NOT EXPENSIVE').font = Font(name='Arial', color='00D4FF', bold=True, size=16)
ws3.merge_cells('B2:G2')
ws3.cell(row=3, column=2, value='Sweet spot: 25th-40th percentile pricing. Compete on SPEED + QUALITY, not price. Too cheap = suspicious.').font = GOLD
ws3.merge_cells('B3:G3')

headers3 = ['STRATEGY', 'DETAILS']
header_row(ws3, 5, headers3)

pricing_data = [
    ('PRICING RULES', ''),
    ('Rule 1: Never be cheapest', 'Bottom 10% pricing = suspicious + attracts nightmare clients. Price at 25th-40th percentile. Shows confidence in quality.'),
    ('Rule 2: 3-tier packages', 'Basic ($X) / Standard ($2X) / Premium ($3X). Most buy Standard. Premium is anchor price. Always include "extras" (+rush delivery, +source code, +revisions).'),
    ('Rule 3: Rush delivery premium', 'Standard delivery: 2-3 days. Express: 24hr (+50%). Ultra-rush: 6hr (+100%). Claude Code does it in 15min regardless — the premium is pure profit.'),
    ('Rule 4: Revision insurance', 'Offer "unlimited revisions" — it costs you nothing (Claude Code revises in seconds). This eliminates buyer hesitation and justifies higher pricing.'),
    ('Rule 5: Source code upsell', 'Base price = deployed/built product. +$25-50 for full source code + documentation. Many clients want this. Zero extra effort.'),
    ('Rule 6: Maintenance packages', 'After delivery, offer $50-200/month maintenance packages. Bug fixes, updates, small changes. Claude Code handles in minutes. Recurring revenue.'),
    ('', ''),
    ('CONVERSION OPTIMIZATION', ''),
    ('Portfolio first 48hrs', 'Before launching ANY gig, build 3-5 portfolio pieces. Claude Code can create showcase projects in 30 minutes total. Screenshots + live demos.'),
    ('Video introductions', 'Use ElevenLabs or record 30-sec intro video per gig. Gigs with videos convert 2-3x better. Script: "Hi, I specialize in X, here\'s what I deliver..."'),
    ('First 10 reviews strategy', 'Price first 10 gigs at 50% discount for 5-star reviews. Total cost: $500-1000 in discount. Value: unlocks algorithmic boost + social proof forever.'),
    ('Response time', 'Reply to ALL inquiries within 1 hour (set up notifications). Fast response = 30% higher conversion. LLM can draft responses, human sends.'),
    ('Custom samples', 'For $200+ gigs, offer free 10-minute custom sample. Claude Code creates it while you\'re chatting with client. Jaw-dropping close rate.'),
    ('Case study content', 'After every gig, screenshot the result + write 2-sentence case study. Post on Twitter/LinkedIn. "Just built a Chrome extension in 15min" = inbound leads.'),
    ('', ''),
    ('ANTI-DETECTION / SUSTAINABILITY', ''),
    ('Delivery timing', 'NEVER deliver a complex project in 15 minutes. Always wait minimum 4-6 hours even if done instantly. Builds perceived value + avoids AI suspicion.'),
    ('Personalized touches', 'Add a README.md with client\'s name. Include deployment instructions specific to their stack. Comment code naturally. Shows "human touch."'),
    ('Communication cadence', 'Send "progress update" messages even though it\'s done. "Working on your project now, looking great!" → 2hrs later → "Almost done, testing a few things" → deliver.'),
    ('Code style variation', 'Don\'t deliver identical patterns every time. Vary: variable naming, file structure, comment style. Each delivery should feel hand-crafted.'),
    ('Revision handling', 'Welcome revisions enthusiastically. "Great feedback, let me tweak that!" Claude Code revises in 30 seconds. Client feels heard + valued.'),
]

for i, (strat, details) in enumerate(pricing_data):
    row = 6 + i
    c1 = ws3.cell(row=row, column=1, value=strat)
    c2 = ws3.cell(row=row, column=2, value=details)
    if details == '':
        c1.font = CYAN_SM
        c1.fill = HEADER_FILL
        c2.fill = HEADER_FILL
    else:
        c1.font = GOLD
        c2.font = WHITE
        if i % 2 == 1:
            c1.fill = ROW_ALT
            c2.fill = ROW_ALT

style_sheet(ws3, 2, [30, 100])

# ============================================================
# SHEET 4: AUTOMATION PIPELINE
# ============================================================
ws4 = wb.create_sheet('AUTOMATION PIPELINE')

ws4.cell(row=2, column=2, value='FREELANCE FULFILLMENT AUTOMATION — LLM-IN-THE-LOOP').font = Font(name='Arial', color='00D4FF', bold=True, size=16)
ws4.merge_cells('B2:F2')
ws4.cell(row=3, column=2, value='Goal: human touches each gig for <5 minutes. Claude Code does 95%+ of the work.').font = GREEN_B
ws4.merge_cells('B3:F3')

headers4 = ['PHASE', 'STEP', 'WHO DOES IT', 'TIME', 'DETAILS']
header_row(ws4, 5, headers4)

pipeline_data = [
    ('PHASE 1: INTAKE', '', '', '', ''),
    ('', '1. Client submits order/message', 'CLIENT', '—', 'Order comes in via Fiverr/Upwork/etc. Notification pushed to phone + Slack.'),
    ('', '2. LLM parses requirements', 'CLAUDE API', '10 sec', 'Claude API reads order description, extracts: project type, tech stack, deliverables, special requests. Outputs structured JSON.'),
    ('', '3. LLM drafts acknowledgment', 'CLAUDE API', '10 sec', 'Auto-draft: "Thanks for your order! I\'m reviewing your requirements now and will begin shortly. Expected delivery: [X]."'),
    ('', '4. Human reviews + sends ack', 'HUMAN', '30 sec', 'Quick review of parsed requirements. Tweak ack message if needed. Send. Total human time: 30 seconds.'),
    ('PHASE 2: BUILD', '', '', '', ''),
    ('', '5. Claude Code builds project', 'CLAUDE CODE', '5-60 min', 'Full project build: code, structure, readme, tests. Claude Code does this autonomously. Human can do other things.'),
    ('', '6. Auto-run QA checks', 'SCRIPT', '1 min', 'Automated: lint, build, test, screenshot. Script validates project compiles, renders, meets basic quality bar.'),
    ('', '7. Human spot-check', 'HUMAN', '2-3 min', 'Quick visual review: does it look good? Does it match requirements? If yes → deliver. If issues → tell Claude Code to fix.'),
    ('PHASE 3: DELIVER', '', '', '', ''),
    ('', '8. Package deliverables', 'SCRIPT', '30 sec', 'Auto-zip code, generate deployment instructions, create README with client name, add portfolio screenshot.'),
    ('', '9. LLM drafts delivery message', 'CLAUDE API', '10 sec', 'Personalized delivery message: "Hi [Name], your [project] is ready! Here\'s what I built: [features]. Let me know if you need any adjustments!"'),
    ('', '10. Human sends delivery', 'HUMAN', '30 sec', 'Review delivery message, attach files, send. Total human time: 30 seconds.'),
    ('PHASE 4: POST-DELIVERY', '', '', '', ''),
    ('', '11. Follow-up for review', 'LLM CRON', '0', '48hr auto-follow-up: "Hope you\'re enjoying [project]! If satisfied, a 5-star review would mean the world."'),
    ('', '12. Revision handling', 'CLAUDE CODE', '1-5 min', 'If revision requested: Claude Code fixes in 30 seconds. Human reviews, sends. Total: 2 minutes per revision.'),
    ('', '13. Case study capture', 'SCRIPT', '0', 'Auto-screenshot result, auto-generate 2-sentence case study, auto-queue for Twitter/LinkedIn posting.'),
    ('', '14. Maintenance upsell', 'LLM CRON', '0', '7 days post-delivery: "Want me to maintain your [project]? $99/month for unlimited bug fixes + small changes."'),
    ('', '', '', '', ''),
    ('TOTAL HUMAN TIME PER GIG', '', 'HUMAN', '3-5 MIN', 'Review requirements (30s) + spot-check build (2-3min) + send delivery (30s) = 3-5 minutes of actual human work per $50-500 gig.'),
    ('THEORETICAL THROUGHPUT', '', '—', '—', 'At 5 min/gig, 8 active hours = 96 gigs/day theoretical max. Realistic with communication overhead: 10-20 gigs/day = $500-5,000/day.'),
]

for i, (phase, step, who, time, details) in enumerate(pipeline_data):
    row = 6 + i
    c1 = ws4.cell(row=row, column=1, value=phase)
    c2 = ws4.cell(row=row, column=2, value=step)
    c3 = ws4.cell(row=row, column=3, value=who)
    c4 = ws4.cell(row=row, column=4, value=time)
    c5 = ws4.cell(row=row, column=5, value=details)
    if step == '' and phase != '':
        c1.font = CYAN_SM
        for c in [c1,c2,c3,c4,c5]:
            c.fill = HEADER_FILL
    elif 'TOTAL' in phase or 'THEORETICAL' in phase:
        c1.font = GREEN_B
        c2.font = GREEN_B
        c3.font = GREEN_B
        c4.font = GREEN_B
        c5.font = GREEN_B
        for c in [c1,c2,c3,c4,c5]:
            c.fill = GREEN_FILL
    else:
        c1.font = WHITE
        c2.font = WHITE_B
        c3.font = GOLD if who == 'HUMAN' else (GREEN if who in ('CLAUDE CODE', 'CLAUDE API') else WHITE)
        c4.font = WHITE
        c5.font = WHITE
        if i % 2 == 1:
            for c in [c1,c2,c3,c4,c5]:
                c.fill = ROW_ALT

style_sheet(ws4, 5, [22, 30, 14, 10, 65])

# ============================================================
# SHEET 5: DEMAND HEATMAP — WHAT PET PROJECTS NEED
# ============================================================
ws5 = wb.create_sheet('DEMAND HEATMAP')

ws5.cell(row=2, column=2, value='DEMAND HEATMAP — PET PROJECT / BOOTSTRAP BUSINESSES BUYING').font = Font(name='Arial', color='00D4FF', bold=True, size=16)
ws5.merge_cells('B2:G2')
ws5.cell(row=3, column=2, value='These people don\'t know Claude Code exists. They\'re paying $200-500 for what takes us 15 minutes.').font = GOLD
ws5.merge_cells('B3:G3')

headers5 = ['BUYER PERSONA', 'WHAT THEY NEED', 'WHAT THEY PAY', 'OUR COST', 'SEARCH VOLUME', 'NOTES']
header_row(ws5, 5, headers5)

demand_data = [
    ('Indie maker / solopreneur', 'Landing page + waitlist + Stripe integration', '$150-400', '$0 (Claude Code)', 'MASSIVE', 'Every ProductHunt launch needs this. Recurring demand. They launch monthly.'),
    ('Small business owner', 'Website redesign / modernization', '$200-800', '$0', 'MASSIVE', 'Millions of small biz with 2015-era websites. They search Fiverr first. Easy close.'),
    ('Startup founder (non-technical)', 'MVP / prototype app', '$500-2000', '$0', 'HIGH', 'They have ideas but can\'t code. Will pay premium for working prototype. Upsell to full build.'),
    ('Content creator', 'Custom tools (link-in-bio, analytics, bots)', '$50-300', '$0', 'HIGH', 'TikTokers, YouTubers need custom tools. Discord bots, dashboards, link pages.'),
    ('E-commerce seller', 'Shopify customization, product scrapers', '$100-500', '$0', 'VERY HIGH', 'Always need: custom themes, inventory scripts, competitor scrapers, CSV imports.'),
    ('Agency owner', 'White-label dev work', '$200-1000', '$0', 'HIGH', 'Agencies sub out work. They charge client $2K, pay you $500. Win-win. Recurring.'),
    ('SaaS founder', 'Feature builds, bug fixes, integrations', '$100-500', '$0', 'HIGH', 'Need specific features added, API integrations, bug fixes. Ongoing relationships.'),
    ('Marketing team', 'Automation scripts, data analysis, dashboards', '$75-300', '$0', 'HIGH', 'Can\'t code but need: scrapers, report generators, email automation, CRM scripts.'),
    ('Real estate agent', 'Property listing sites, CRM, lead capture', '$200-600', '$0', 'MEDIUM', 'Want custom IDX sites, lead magnets, auto-email sequences. Premium pricing.'),
    ('Restaurant/local biz', 'Online ordering, menu site, booking system', '$150-500', '$0', 'HIGH', 'Post-COVID every restaurant needs digital. Simple builds = high value to them.'),
    ('Educator / course creator', 'Course platform, quiz apps, student portals', '$150-400', '$0', 'MEDIUM', 'Teachable is expensive. Custom simple course site = value proposition.'),
    ('Non-profit', 'Donation pages, volunteer portals, event sites', '$100-300', '$0', 'MEDIUM', 'Budget-conscious but willing to pay for quality. Good for portfolio building.'),
    ('Crypto/DeFi project', 'Token dashboards, wallet connectors, dApps', '$300-1000', '$0', 'MEDIUM', 'High budgets, fast timelines, willing to pay premium. Claude Code handles Solidity/React.'),
    ('Pet project hobbyist', 'Personal tools, hobby trackers, game mods', '$25-150', '$0', 'HIGH', 'r/slavelabour, r/forhire goldmine. Small gigs but VOLUME. 5-min builds.'),
    ('HR / recruiting', 'ATS customization, job board scrapers, reports', '$100-400', '$0', 'MEDIUM', 'Need: resume parsers, LinkedIn scrapers, interview schedulers. Repeat clients.'),
]

for i, d in enumerate(demand_data):
    row = 6 + i
    for j, val in enumerate(d, 1):
        c = ws5.cell(row=row, column=j, value=val)
        c.font = WHITE if j > 1 else GOLD
        if i % 2 == 1:
            c.fill = ROW_ALT
        if j == 4:
            c.font = GREEN_B
        if j == 5 and ('MASSIVE' in str(val) or 'VERY' in str(val)):
            c.font = GREEN_B

style_sheet(ws5, 6, [25, 38, 14, 16, 14, 50])

# ============================================================
# SHEET 6: 30-DAY LAUNCH PLAN
# ============================================================
ws6 = wb.create_sheet('30-DAY LAUNCH')

ws6.cell(row=2, column=2, value='30-DAY FREELANCE ARBITRAGE LAUNCH SEQUENCE').font = Font(name='Arial', color='00D4FF', bold=True, size=16)
ws6.merge_cells('B2:F2')

headers6 = ['DAY', 'TASK', 'PLATFORM', 'TIME REQ', 'OUTCOME']
header_row(ws6, 4, headers6)

launch_data = [
    ('WEEK 1: FOUNDATION', '', '', '', ''),
    ('Day 1', 'Build 5 portfolio projects with Claude Code (landing page, Chrome ext, Discord bot, scraper, dashboard)', 'ALL', '2 hrs', '5 showcase pieces ready'),
    ('Day 2', 'Create Fiverr account. Set up 5 gigs: landing page, web scraper, Python script, Chrome extension, bug fix', 'Fiverr', '1 hr', '5 live gigs on Fiverr'),
    ('Day 3', 'Create Upwork profile. Write 10 proposals for active jobs using LLM personalization', 'Upwork', '1 hr', 'Upwork profile live + 10 proposals sent'),
    ('Day 4', 'Create Contra + PeoplePerHour accounts. List top 3 services on each', 'Contra/PPH', '45 min', '6 more listings live'),
    ('Day 5', 'Create LinkedIn Services page. Post first "I built this in 15min" case study', 'LinkedIn', '30 min', 'LinkedIn inbound channel open'),
    ('Day 6', 'Set up Reddit threads: r/forhire, r/slavelabour [For Hire] posts', 'Reddit', '30 min', 'Reddit deal flow open'),
    ('Day 7', 'Twitter/X: Pin services tweet, post first build showcase thread', 'Twitter/X', '30 min', 'Twitter inbound channel open'),
    ('WEEK 2: FIRST REVENUE', '', '', '', ''),
    ('Day 8-10', 'Apply to 10 Upwork jobs/day. Offer 50% discount on first 3 Fiverr gigs for reviews', 'Upwork/Fiverr', '1 hr/day', 'First orders coming in'),
    ('Day 11-12', 'Deliver first gigs. Request 5-star reviews. Post case studies on Twitter/LinkedIn', 'ALL', '2 hrs/day', 'First reviews earned'),
    ('Day 13-14', 'Add 3 more Fiverr gigs (Shopify, Notion, API integration). Create gig video intros', 'Fiverr', '1 hr', '8 total Fiverr gigs'),
    ('WEEK 3: SCALE', '', '', '', ''),
    ('Day 15-17', 'Set up automation pipeline: LLM proposal writer, delivery templater, follow-up cron', 'INTERNAL', '3 hrs total', 'Semi-automated fulfillment'),
    ('Day 18-19', 'Raise Fiverr prices by 25% (you have reviews now). Add premium tiers', 'Fiverr', '30 min', 'Higher margins per gig'),
    ('Day 20-21', 'Apply for Fiverr Pro. Add Guru.com and Freelancer.com listings', 'Fiverr/Guru/FL', '1 hr', '3 more platforms live'),
    ('WEEK 4: OPTIMIZE', '', '', '', ''),
    ('Day 22-24', 'Analyze which services sell most. Double down on top 3. Kill bottom performers', 'ALL', '1 hr', 'Optimized service catalog'),
    ('Day 25-26', 'Set up maintenance package offerings for past clients. LLM sends upsell messages', 'ALL', '1 hr', 'Recurring revenue started'),
    ('Day 27-28', 'Build white-label agency offering. Reach out to 5 agencies for subcontract work', 'LinkedIn/Email', '2 hrs', 'Agency pipeline opened'),
    ('Day 29-30', 'Review month 1 metrics. Adjust pricing. Plan month 2 expansion services', 'ALL', '1 hr', 'Month 1 complete, $1K-3K earned'),
    ('', '', '', '', ''),
    ('MONTH 1 TARGET', '$1,000-3,000 revenue', 'ALL', '~2 hrs/day', 'Foundation built, reviews earned, pipeline running'),
    ('MONTH 3 TARGET', '$3,000-8,000/month', 'ALL', '~3 hrs/day', 'Repeat clients, agency subs, optimized listings'),
    ('MONTH 6 TARGET', '$8,000-20,000/month', 'ALL', '~4 hrs/day', 'Pro status, premium pricing, maintenance MRR, multiple platforms'),
]

for i, d in enumerate(launch_data):
    row = 5 + i
    for j, val in enumerate(d, 1):
        c = ws6.cell(row=row, column=j, value=val)
        if d[1] == '' and d[0] != '':
            c.font = CYAN_SM
            c.fill = HEADER_FILL
        elif 'TARGET' in str(d[0]):
            c.font = GREEN_B
            c.fill = GREEN_FILL
        else:
            c.font = WHITE if j > 1 else GOLD
            if i % 2 == 1:
                c.fill = ROW_ALT

style_sheet(ws6, 5, [16, 65, 16, 12, 35])

# ============================================================
# SHEET 7: LLM ALPHA THESIS — WHY THIS PRINTS
# ============================================================
ws7 = wb.create_sheet('WHY THIS PRINTS')

ws7.cell(row=2, column=2, value='THE FREELANCE ARB ALPHA THESIS').font = Font(name='Arial', color='00D4FF', bold=True, size=16)
ws7.merge_cells('B2:C2')

thesis_lines = [
    ('THE EDGE', ''),
    ('Claude Code Max = ~unlimited usage', 'Flat monthly cost, no per-token billing. Every gig is pure margin. The more gigs you do, the lower your effective cost approaches $0.'),
    ('5-minute delivery on 3-day projects', 'Clients expect 2-5 day delivery. You build in 15 minutes, wait 6+ hours, deliver "early." They think you\'re a rockstar.'),
    ('Quality exceeds most freelancers', 'Claude Code produces cleaner, better-documented, more tested code than 80% of Fiverr freelancers. Real competitive advantage.'),
    ('Unlimited revisions at zero cost', 'Most freelancers dread revisions (more unpaid work). You WELCOME them — Claude Code revises in 30 seconds. This becomes your differentiator.'),
    ('', ''),
    ('WHY CLIENTS DON\'T JUST USE AI THEMSELVES', ''),
    ('They don\'t know how', 'Most small biz owners, founders, and content creators have zero technical ability. They don\'t know Claude Code exists.'),
    ('They don\'t want to learn', 'Even if they know about AI coding, they\'d rather pay $100 than spend 3 hours figuring it out. Time > money for them.'),
    ('They need it to work', 'AI output requires debugging, testing, deployment knowledge. They need a "developer" to make it production-ready.'),
    ('They need support', 'They want someone to message when it breaks. You provide that (Claude Code fixes it in 30 seconds).'),
    ('', ''),
    ('MARKET SIZE', ''),
    ('Fiverr: 4M+ active buyers', 'Programming & Tech category growing 25%+ YoY. Massive demand for web dev, automation, scripts.'),
    ('Upwork: $3.8B annual freelancer earnings', 'Tech is highest-paying category. Average project: $500-5K. Your margin: 90%+.'),
    ('Total addressable: $50B+/yr', 'Global freelance dev market is massive and growing. You\'re capturing margin that used to require real labor.'),
    ('', ''),
    ('EDGE DURATION: 24-48 MONTHS', ''),
    ('Why it lasts', 'Even as AI awareness grows, most people still can\'t prompt effectively, debug output, or deploy. The "AI-assisted developer" role is a new profession.'),
    ('When it narrows', 'Eventually no-code AI builders (v0, Bolt, Replit Agent) will make simple projects truly self-serve. But custom work, integrations, and complex projects will always need a human-in-loop.'),
    ('How to extend the edge', 'Build relationships, earn Pro status, accumulate reviews. Once you\'re Top Rated on Upwork or Fiverr Pro, the platform promotes you. That moat compounds.'),
]

for i, (key, val) in enumerate(thesis_lines):
    row = 4 + i
    c1 = ws7.cell(row=row, column=1, value=key)
    c2 = ws7.cell(row=row, column=2, value=val)
    if val == '' and key != '':
        c1.font = CYAN_SM
        c1.fill = HEADER_FILL
        c2.fill = HEADER_FILL
    elif key == '':
        pass
    else:
        c1.font = GOLD
        c2.font = WHITE
        if i % 2 == 1:
            c1.fill = ROW_ALT
            c2.fill = ROW_ALT

style_sheet(ws7, 2, [35, 100])

# ============================================================
# SHEET 8: DAILY MORNING BRIEFING SPEC
# ============================================================
ws8 = wb.create_sheet('DAILY BRIEFING SPEC')

ws8.cell(row=2, column=2, value='DAILY MORNING BRIEFING — HUMAN-IN-LOOP REQUIRED ACTIONS').font = Font(name='Arial', color='00D4FF', bold=True, size=16)
ws8.merge_cells('B2:D2')
ws8.cell(row=3, column=2, value='Auto-generated every morning at 5:00 AM. Consults all ledgers, ops, logs, and monitoring.').font = WHITE
ws8.merge_cells('B3:D3')

headers8 = ['COMPONENT', 'DATA SOURCE', 'WHAT IT CHECKS', 'OUTPUT']
header_row(ws8, 5, headers8)

briefing_spec = [
    ('AUTOMATION STATUS', 'LEDGER/AUTOMATION_RESULTS.csv + logs/', 'Did all cron jobs run? Any failures? Alpha extracted count. Content generated count. Opus audit verdicts.', 'GREEN/YELLOW/RED status per automation + error details'),
    ('REVENUE CHECK', 'FINANCIALS/ + LEDGER/ACTIVE_INVESTMENTS.csv', 'Any new revenue? Pending payments? Investment positions changed? Expense anomalies?', 'Revenue delta since yesterday + alerts'),
    ('CONTENT QUEUE', 'LEDGER/CONTENT_CALENDAR_30DAY.csv', 'What content is scheduled for today? Any gaps? Platforms needing manual posting?', 'Today\'s posting schedule + manual action items'),
    ('ALPHA DIGEST', 'LEDGER/ALPHA_STAGING.csv + ALPHA_BY_CATEGORY/', 'New alpha entries pending review. High-priority items. Expiring opportunities.', 'Top 5 alpha to review today + action needed'),
    ('ACCOUNT HEALTH', 'LEDGER/ACCOUNT_HEALTH_DAILY.csv', 'Any accounts flagged? Shadowban detection? Engagement drops? Warmup status per platform.', 'Account health dashboard + warnings'),
    ('FREELANCE PIPELINE', 'NEW: freelance_orders.csv (to be created)', 'New orders received overnight. Pending deliveries. Review requests. Revision requests.', 'Active gig queue + deadlines'),
    ('OPS STATUS', 'OPS/ directory scan', 'Which ops are active? Which need attention? Blocked ops? New opportunities?', 'Op-by-op status summary'),
    ('EXPERIMENT RESULTS', 'LEDGER/MEGA_SHEET/TAB9_EXPERIMENTS_METRICS.csv', 'Any A/B tests concluded? Winners to implement? New tests to start?', 'Experiment scorecard'),
    ('FILE SYSTEM CHANGES', 'git diff / file watchers', 'New files created? Large files? Orphan files? Config changes?', 'File system health check'),
    ('TOOL STATUS', 'LEDGER/TOOLS_SERVICES_MASTER.csv', 'Free tier limits approaching? Tools needing renewal? New tool opportunities?', 'Tool health + upgrade alerts'),
    ('', '', '', ''),
    ('OUTPUT FORMAT', '', '', ''),
    ('Morning email/report', 'ALL ABOVE', 'Aggregated into single morning report at 5:00 AM', 'DAILY_BRIEFING_YYYY-MM-DD.md in LEDGER/'),
    ('Terminal dashboard', 'PRINTMAXX Terminal.app', 'Live dashboard showing today\'s priorities', 'TUI tab: "TODAY" with action items'),
    ('Slack/notification', 'Webhook', 'Push notification with top 3 action items', 'Phone notification at 6:00 AM'),
]

for i, d in enumerate(briefing_spec):
    row = 6 + i
    for j, val in enumerate(d, 1):
        c = ws8.cell(row=row, column=j, value=val)
        if d[0] == '':
            pass
        elif d[0] == 'OUTPUT FORMAT':
            c.font = CYAN_SM
            c.fill = HEADER_FILL
        else:
            c.font = WHITE if j > 1 else GOLD
            if i % 2 == 1:
                c.fill = ROW_ALT

style_sheet(ws8, 4, [22, 35, 50, 40])

# SAVE
out = '/sessions/awesome-nice-brown/mnt/PRINTMAXX_STARTER_KITttttt/PRINTMAXX_FREELANCE_ARB.xlsx'
wb.save(out)
print(f'Saved to {out}')
print(f'Sheets: {wb.sheetnames}')
print(f'Services: {len(services)}')
print(f'Platforms: {len(platforms)}')
