from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Colors
BG_DARK = PatternFill('solid', fgColor='0D1117')
BG_HEADER = PatternFill('solid', fgColor='0A1628')
BG_SECTION = PatternFill('solid', fgColor='112240')
BG_ROW1 = PatternFill('solid', fgColor='0D1117')
BG_ROW2 = PatternFill('solid', fgColor='161B22')
BG_PICK = PatternFill('solid', fgColor='0B3D0B')
BG_ALT = PatternFill('solid', fgColor='1A2332')
BG_GOLD = PatternFill('solid', fgColor='2D2500')

FONT_TITLE = Font(name='Arial', size=20, bold=True, color='00D4FF')
FONT_SUB = Font(name='Arial', size=12, bold=True, color='8B949E')
FONT_HEAD = Font(name='Arial', size=11, bold=True, color='00D4FF')
FONT_SEC = Font(name='Arial', size=12, bold=True, color='58A6FF')
FONT_BODY = Font(name='Arial', size=10, color='C9D1D9')
FONT_WHITE = Font(name='Arial', size=10, bold=True, color='FFFFFF')
FONT_CYAN = Font(name='Arial', size=10, bold=True, color='00D4FF')
FONT_GREEN = Font(name='Arial', size=10, bold=True, color='3FB950')
FONT_GOLD = Font(name='Arial', size=10, bold=True, color='F0C000')
FONT_ORANGE = Font(name='Arial', size=10, color='F0883E')
FONT_PURPLE = Font(name='Arial', size=10, bold=True, color='A371F7')
FONT_HANDLE = Font(name='Arial', size=11, bold=True, color='FFFFFF')
FONT_SMALL = Font(name='Arial', size=9, color='8B949E')

BORDER = Border(bottom=Side(style='thin', color='21262D'))
BORDER_SEC = Border(bottom=Side(style='medium', color='00D4FF'))
ALIGN_C = Alignment(horizontal='center', vertical='center')
ALIGN_L = Alignment(horizontal='left', vertical='center', wrap_text=True)

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def fill_bg(ws, row, cols, fill):
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).fill = fill

def section_header(ws, row, cols, text, font=FONT_SEC):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=cols-1)
    ws.cell(row=row, column=2, value=text).font = font
    fill_bg(ws, row, cols, BG_SECTION)
    for c in range(1, cols+1):
        ws.cell(row=row, column=c).border = BORDER_SEC
    ws.row_dimensions[row].height = 28
    return row + 1

def col_headers(ws, row, cols, headers):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=i+1, value=h)
        c.font = FONT_HEAD
        c.fill = BG_HEADER
        c.alignment = ALIGN_C
    ws.row_dimensions[row].height = 24
    return row + 1

# ============================================================
# NAMING PHILOSOPHY:
# 1. Inside baseball — uses niche terminology that insiders recognize
# 2. Not cringe — no "guru", "mindset", "unlock", "hustle"
# 3. Good SEO — keywords embedded naturally
# 4. Sounds like a real person/brand — not AI-generated
# 5. Modeled after successful scraped accounts
# ============================================================

# All account data organized by category
ACCOUNTS = {
    'BUILD IN PUBLIC / META BRAND': {
        'color': '00D4FF',
        'desc': 'The main PrintMaxx build-in-public account. Revenue screenshots, automation demos, behind-the-scenes.',
        'names': [
            ('@printmaxxer', 'X/Twitter', 'Build in Public', 'Main brand. "Maxxer" = someone who optimizes everything (looksmaxxing/gymmaxxing culture). Not cringe — it\'s self-aware internet culture.', 'TOP PICK'),
            ('@printmaxxer', 'TikTok', 'Build in Public', 'Same handle, short-form content demos', ''),
            ('@printmaxxer', 'YouTube', 'Build in Public', 'Long-form: "How I Built 30 Apps in 90 Days"', ''),
            ('@printmaxxer', 'Instagram', 'Build in Public', 'Revenue screenshots, tool reviews, aesthetic', ''),
        ]
    },
    'AI TOOLS / AUTOMATION': {
        'color': '58A6FF',
        'desc': 'AI tool reviews, workflow automation, Claude/GPT tips. Like @levelsio but focused on AI tooling.',
        'names': [
            ('@patchnotes_', 'X/Twitter', 'AI Tools', '"Patch notes" = software update terminology. Underscore adds dev energy. Implies you\'re tracking every AI update.', 'TOP PICK'),
            ('@toolcache', 'X/Twitter', 'AI Tools', '"Cache" = stored data. You\'re the cached library of every tool worth knowing.', ''),
            ('@latentspace_', 'X/Twitter', 'AI Deep', '"Latent space" = ML concept (compressed representation). Signals deep technical knowledge without explaining it.', 'ALT PICK'),
            ('@inferenceloop', 'X/Twitter', 'AI Technical', '"Inference loop" = model running predictions. Pure ML insider terminology.', ''),
            ('@hotreload', 'X/Twitter', 'Dev/AI', '"Hot reload" = dev term for live code refresh. Short, punchy, universally dev-recognized.', ''),
            ('@promptdrift', 'X/Twitter', 'AI/LLM', '"Prompt drift" = real LLM concept where outputs shift over time. Insider, technical, memorable.', ''),
        ]
    },
    'INDIE HACKING / SOLOPRENEUR': {
        'color': '3FB950',
        'desc': 'Revenue reports, app launches, solopreneur tactics. Like @levelsio / @tdinh_me energy.',
        'names': [
            ('@shipclock', 'X/Twitter', 'Indie', '"Ship" = launch products. "Clock" = speed/time pressure. Ship fast or die.', 'TOP PICK'),
            ('@coldstart_', 'X/Twitter', 'Indie/Growth', '"Cold start problem" = classic growth term. How do you get users when you have none? Insider terminology.', 'ALT PICK'),
            ('@ramenclub', 'X/Twitter', 'Indie', '"Ramen profitable" = startup term for making just enough to survive. Club = community.', ''),
            ('@exitvelocity_', 'X/Twitter', 'Startup', '"Exit velocity" = physics term borrowed by startups. Speed at which you escape gravity (competition).', ''),
            ('@zerotoone_', 'X/Twitter', 'Indie', 'Peter Thiel reference. Zero to One = creating something new. Iconic startup book.', ''),
            ('@pushtomaster', 'X/Twitter', 'Dev/Indie', '"Push to master" = git command. Shipping directly to production. Bold, reckless, dev humor.', ''),
        ]
    },
    'COLD OUTBOUND / B2B SALES': {
        'color': '58D5DB',
        'desc': 'Cold email, LinkedIn outreach, lead gen. Like @pipelineabuser / @codyschneiderxx.',
        'names': [
            ('@pipelinepoisn', 'X/Twitter', 'Cold Email', '"Pipeline poison" = what bad leads do to your funnel. Shortened "poison" for handle. Inside sales terminology.', 'TOP PICK'),
            ('@coldsequence', 'X/Twitter', 'Cold Email', '"Cold sequence" = the automated email chain. Every SDR knows this term.', ''),
            ('@bounceratekid', 'X/Twitter', 'Email', '"Bounce rate" = email deliverability metric. "Kid" adds personality (like @FearBuck pattern).', 'ALT PICK'),
            ('@spamfolder_', 'X/Twitter', 'Email', 'Self-deprecating humor. "I live in the spam folder." Funny, relatable for anyone doing cold email.', ''),
            ('@firsttouchonly', 'X/Twitter', 'Sales', '"First touch" = initial contact in sales funnel. "Only" = we only need one shot.', ''),
            ('@domainwarmup', 'X/Twitter', 'Deliverability', '"Domain warmup" = email deliverability process. Technical, insider, SEO-friendly.', ''),
        ]
    },
    'FINANCE / TRADING': {
        'color': 'F0C000',
        'desc': 'Market analysis, options flow, crypto signals. Like @unusual_whales / @DeItaone.',
        'names': [
            ('@bidaskspread', 'X/Twitter', 'Trading', '"Bid-ask spread" = fundamental market mechanics concept. Every trader knows it. Clean, technical.', 'TOP PICK'),
            ('@darkpoolflow', 'X/Twitter', 'Options', '"Dark pool" = off-exchange trading venue. "Flow" = order flow. Institutional insider terminology.', 'ALT PICK'),
            ('@thetadecay_', 'X/Twitter', 'Options', '"Theta decay" = options Greek measuring time decay. Underscore = technical handle energy.', ''),
            ('@maxpainlevel', 'X/Twitter', 'Options', '"Max pain" = the strike price at which most options expire worthless. Technical, insider.', ''),
            ('@vwapdeviator', 'X/Twitter', 'Day Trading', '"VWAP" = Volume Weighted Average Price. "Deviator" = someone who trades the deviation.', ''),
            ('@openinterest_', 'X/Twitter', 'Futures/Options', '"Open interest" = number of outstanding contracts. Pure market data terminology.', ''),
        ]
    },
    'CRYPTO / ONCHAIN': {
        'color': 'F0883E',
        'desc': 'Crypto whale tracking, DeFi, on-chain analysis. Like @lookonchain / @GlassNode.',
        'names': [
            ('@mempooldive', 'X/Twitter', 'Crypto', '"Mempool" = where pending transactions wait. "Dive" = deep analysis. Insider crypto infra terminology.', 'TOP PICK'),
            ('@gasfeesonly', 'X/Twitter', 'Crypto', '"Gas fees" = transaction costs on ETH. Self-deprecating humor — all profits go to gas.', ''),
            ('@noncetrouble', 'X/Twitter', 'Crypto', '"Nonce" = cryptographic number used once. Sounds playful but is deeply technical.', 'ALT PICK'),
            ('@frontrunbot', 'X/Twitter', 'DeFi', '"Front-running" = MEV strategy. "Bot" = automated. Controversial but insider.', ''),
            ('@rugpullradar', 'X/Twitter', 'Crypto Safety', '"Rug pull" = scam exit. "Radar" = detection. Protective positioning.', ''),
            ('@epochwatcher', 'X/Twitter', 'Crypto', '"Epoch" = time period in blockchain consensus. "Watcher" = monitoring.', ''),
        ]
    },
    'MEME / VIRAL REPURPOSE': {
        'color': 'A371F7',
        'desc': 'Viral content curation, meme pages, no-context clips. Like @LocalBateman / @NoContextBrits.',
        'names': [
            ('@foldersdump', 'X/Twitter', 'Memes', '"Folders dump" = internet slang for dumping saved content. Like @videosinfolder energy. Insider, casual.', 'TOP PICK'),
            ('@nocontextmidwest', 'X/Twitter', 'Regional Memes', '"No context" format (proven by @NoContextBrits) + US midwest. Regional specificity = engagement.', ''),
            ('@postironyclub', 'X/Twitter', 'Memes', '"Post-irony" = meme culture term for sincerity wrapped in irony layers. Gen Z/internet culture insider.', 'ALT PICK'),
            ('@maincharmoment', 'X/Twitter', 'Viral Clips', '"Main character moment" = viral format where someone acts like the main character. TikTok-native term.', ''),
            ('@unhingedfiles', 'X/Twitter', 'Memes', '"Unhinged" = internet slang for wild behavior. "Files" = collected evidence. Meme curator positioning.', ''),
            ('@npcencounters', 'X/Twitter', 'Viral Clips', '"NPC" = Non-Player Character (gaming term for people who act robotically). Viral format.', ''),
            ('@lowresgold', 'X/Twitter', 'Memes', '"Low res" = poor quality. "Gold" = valuable content. Contradiction pattern like @oddlyterrifying.', ''),
            ('@cursedtimeline', 'X/Twitter', 'Memes', '"Cursed timeline" = internet culture for when reality feels wrong. Doom-scrolling energy.', ''),
        ]
    },
    'FITNESS / HEALTH': {
        'color': 'FF6B6B',
        'desc': 'Workout splits, nutrition, progressive overload. Like gym Twitter but not cringe "fitness coach" energy.',
        'names': [
            ('@repscheme_', 'X/Twitter', 'Fitness', '"Rep scheme" = the specific set/rep structure of a program. Every gym rat knows this. Underscore = clean.', 'TOP PICK'),
            ('@workingreps', 'X/Twitter', 'Fitness', '"Working reps" = the actual hard reps (not warmup). Pure gym floor terminology.', ''),
            ('@deloadweek', 'X/Twitter', 'Fitness', '"Deload week" = planned recovery period. Only people who train seriously know this. Insider signal.', 'ALT PICK'),
            ('@failureset', 'X/Twitter', 'Fitness', '"Training to failure" = going until you literally can\'t. "Set" = one round. Raw gym language.', ''),
            ('@compoundonly', 'X/Twitter', 'Fitness', '"Compound lifts" = multi-joint exercises (squat, deadlift, bench). "Only" = purist philosophy.', ''),
            ('@traininglog_', 'X/Twitter', 'Fitness', '"Training log" = workout journal. Simple, functional, no BS. Underscore = dev-adjacent.', ''),
        ]
    },
    'FAITH / DEVOTIONAL': {
        'color': 'FFD700',
        'desc': 'Daily devotionals, scripture, faith community. Authentic, not mega-church-cringe.',
        'names': [
            ('@morningwatch_', 'X/Twitter', 'Faith', '"Morning watch" = old Christian term for early morning prayer/devotion time. Historical, not modern-cringe.', 'TOP PICK'),
            ('@selahmoments', 'X/Twitter', 'Faith', '"Selah" = untranslated Hebrew word in Psalms, meaning pause/reflect. Deep biblical insider term. Beautiful.', 'ALT PICK'),
            ('@dailybread_', 'X/Twitter', 'Faith', '"Daily bread" = Lord\'s Prayer reference. Universal across denominations. Timeless.', ''),
            ('@narrowgate_', 'X/Twitter', 'Faith', '"Narrow gate" = Matthew 7:13-14 reference. Commitment signal without being preachy.', ''),
            ('@vesperbell', 'X/Twitter', 'Faith', '"Vespers" = evening prayer service. "Bell" = church bell calling to worship. Atmospheric, poetic.', ''),
            ('@quiververse', 'X/Twitter', 'Faith', '"Quiver" = Psalm 127:5 ("quiver full of arrows"). "Verse" = scripture. Family + faith signal.', ''),
        ]
    },
    'AI INFLUENCER PERSONAS': {
        'color': 'FF69B4',
        'desc': 'AI-generated personas for niche content. Needs names that pass as real people — not obvious AI.',
        'names': [
            ('@kiraellis_', 'X/Twitter', 'Lifestyle/Fashion', 'Sounds like a real person. "Kira" = common but not generic. "Ellis" = last name energy. Underscore = creator.', 'TOP PICK'),
            ('@jadenarcotics', 'X/Twitter', 'Aesthetic/Alt', '"Jade" = name + color. "Narcotics" = edgy without being illegal. Alt-girl aesthetic naming pattern.', ''),
            ('@siennacurates', 'X/Twitter', 'Lifestyle', '"Sienna" = warm-toned name. "Curates" = implies taste and selection. Lifestyle influencer pattern.', 'ALT PICK'),
            ('@novawinters', 'X/Twitter', 'ASMR/Aesthetic', '"Nova" = star. "Winters" = seasonal surname. Atmospheric, memorable, sounds real.', ''),
            ('@emikohaze', 'X/Twitter', 'Tech/AI', '"Emiko" = Japanese name. "Haze" = mystique. International feel, not obviously American.', ''),
            ('@lunarcassidy', 'X/Twitter', 'Spiritual/Wellness', '"Lunar" = moon. "Cassidy" = Western Americana. Celestial + grounded. Spiritually-coded.', ''),
            ('@reyforbes', 'X/Twitter', 'Finance/Lifestyle', '"Rey" = cool, gender-neutral. "Forbes" = wealth signal without trying. Sounds like a real finance influencer.', ''),
            ('@ivorsteel', 'X/Twitter', 'Fitness/Masc', '"Ivor" = strong Anglo name. "Steel" = literal strength. Gym influencer naming without "fit" or "gains".', ''),
        ]
    },
    'CONTENT FARM / NICHE PAGES': {
        'color': 'C9D1D9',
        'desc': 'Faceless niche pages across verticals. Quote pages, fact pages, compilation channels.',
        'names': [
            ('@rawclipvault', 'X/Twitter', 'Viral Clips', '"Raw" = unedited. "Clip vault" = stored collection. Sounds like an archive, not a content farm.', 'TOP PICK'),
            ('@factcoldcuts', 'X/Twitter', 'Facts/Trivia', '"Cold cuts" = deli meat, but also "cold cuts of information" — quick, digestible facts. Unexpected, memorable.', ''),
            ('@quotesupply', 'X/Twitter', 'Quotes/Motivation', '"Supply" = industrial feel. Not "daily quotes" cringe — it\'s a supply chain of quotes.', 'ALT PICK'),
            ('@nightshiftclips', 'X/Twitter', 'Late Night Content', '"Night shift" = late-night workers/insomniacs. Targeting a vibe, not a topic.', ''),
            ('@bingeclips_', 'X/Twitter', 'Compilations', '"Binge" = watch continuously. "Clips" = short content. The name IS the format.', ''),
            ('@deepcutsfeed', 'X/Twitter', 'Niche Culture', '"Deep cuts" = music term for album tracks beyond singles. Applied to content = non-obvious finds.', ''),
            ('@retrowavefeed', 'X/Twitter', 'Nostalgia', '"Retrowave" = aesthetic/music genre. "Feed" = content stream. Specific subculture targeting.', ''),
            ('@sleepcodex', 'X/Twitter', 'Sleep/ASMR', '"Codex" = ancient manuscript. "Sleep codex" = the collected knowledge of sleep. Elevated, not generic.', ''),
        ]
    },
    'NEWSLETTER / MEDIA BRAND': {
        'color': '00D4FF',
        'desc': 'Email newsletter brands. Like Morning Brew / The Hustle but for specific niches.',
        'names': [
            ('@signalnoise_', 'Newsletter', 'Tech/AI', '"Signal vs noise" = information theory concept. You filter the noise and deliver the signal. Perfect newsletter name.', 'TOP PICK'),
            ('@morningdelta', 'Newsletter', 'Finance', '"Morning" = AM delivery. "Delta" = change/Greek (options). The morning change report.', 'ALT PICK'),
            ('@thecompound_', 'Newsletter', 'Investing', '"Compound" = compound interest. "The" = authority. Investing newsletter with elegant simplicity.', ''),
            ('@weeklyrebase', 'Newsletter', 'Crypto', '"Rebase" = crypto tokenomics term + git term. Weekly cadence. Double insider reference.', ''),
            ('@thebrief_', 'Newsletter', 'Business', '"Brief" = short report + legal brief. Implies concise, authoritative summary. "The" = definitive.', ''),
            ('@patchlog', 'Newsletter', 'Tech/Dev', '"Patch log" = software changelog. Weekly tech updates framed as a patch to your knowledge.', ''),
        ]
    },
    'APPS & SOFTWARE BRANDS': {
        'color': '3FB950',
        'desc': 'Product names for apps, SaaS tools, micro-products. Clean, memorable, domain-available patterns.',
        'names': [
            ('streakpilot', 'Product', 'Habit Tracking', '"Streak" = consecutive-day habit. "Pilot" = navigator/controller. You\'re piloting your streaks.', 'TOP PICK'),
            ('routinedrift', 'Product', 'Productivity', '"Routine drift" = when habits slip. The app catches the drift. Problem-as-name pattern.', ''),
            ('focusburn', 'Product', 'Focus/Study', '"Focus" = attention. "Burn" = calories metaphor for mental energy. Active, energetic.', 'ALT PICK'),
            ('pulsestack', 'Product', 'Analytics', '"Pulse" = heartbeat/real-time check. "Stack" = tech stack. Real-time analytics tool name.', ''),
            ('anchorapp', 'Product', 'Faith/Wellness', '"Anchor" = stability. Already a known word (Spotify\'s podcast tool). Clean, cross-niche.', ''),
            ('vaultmode', 'Product', 'Privacy/Security', '"Vault" = secure storage. "Mode" = activation state. Privacy tool positioning.', ''),
            ('nightowlkit', 'Product', 'Late Night Workers', '"Night owl" = late worker. "Kit" = toolkit. Niche targeting (insomniacs, night shift).', ''),
            ('driftcheck', 'Product', 'Monitoring', '"Drift" = configuration/data drift. "Check" = validation. DevOps-adjacent naming.', ''),
        ]
    },
}

# ============================================================
# SHEET 1: OVERVIEW (all top picks)
# ============================================================
ws = wb.active
ws.title = 'TOP PICKS'
ws.sheet_properties.tabColor = '00D4FF'
set_widths(ws, [3, 22, 14, 16, 44, 12, 3])

for r in range(1, 80):
    fill_bg(ws, r, 7, BG_DARK)

ws.merge_cells('B2:F2')
ws.cell(row=2, column=2, value='PRINTMAXX BRAND & ACCOUNT NAMES').font = FONT_TITLE
ws.row_dimensions[2].height = 35
ws.merge_cells('B3:F3')
ws.cell(row=3, column=2, value='Inside-baseball nomenclature. No cringe. No AI vibes. Sounds like real accounts.').font = FONT_SUB

r = 5
headers = ['', 'HANDLE / NAME', 'PLATFORM', 'NICHE', 'WHY IT WORKS', 'STATUS']
r = col_headers(ws, r, 7, headers)

for cat_name, cat_data in ACCOUNTS.items():
    r = section_header(ws, r, 7, cat_name)
    for name, platform, niche, reason, pick in cat_data['names']:
        if pick != 'TOP PICK':
            continue
        bg = BG_PICK
        fill_bg(ws, r, 7, bg)
        ws.cell(row=r, column=2, value=name).font = FONT_HANDLE
        ws.cell(row=r, column=3, value=platform).font = FONT_BODY
        ws.cell(row=r, column=3).alignment = ALIGN_C
        ws.cell(row=r, column=4, value=niche).font = FONT_CYAN
        ws.cell(row=r, column=4).alignment = ALIGN_C
        ws.cell(row=r, column=5, value=reason).font = FONT_BODY
        ws.cell(row=r, column=5).alignment = ALIGN_L
        ws.cell(row=r, column=6, value='TOP PICK').font = FONT_GREEN
        ws.cell(row=r, column=6).alignment = ALIGN_C
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = BORDER
        ws.row_dimensions[r].height = 38
        r += 1

# ============================================================
# SHEET 2: ALL NAMES (full list with alternates)
# ============================================================
ws2 = wb.create_sheet('ALL NAMES')
ws2.sheet_properties.tabColor = '58A6FF'
set_widths(ws2, [3, 22, 14, 16, 48, 12, 3])

for r in range(1, 200):
    fill_bg(ws2, r, 7, BG_DARK)

ws2.merge_cells('B2:F2')
ws2.cell(row=2, column=2, value='COMPLETE NAME LIBRARY').font = FONT_TITLE
ws2.row_dimensions[2].height = 32
ws2.merge_cells('B3:F3')
ws2.cell(row=3, column=2, value='Every name option across all niches. Top picks highlighted green. Alt picks highlighted gold.').font = FONT_SUB

r = 5
headers = ['', 'HANDLE / NAME', 'PLATFORM', 'NICHE', 'WHY IT WORKS (INSIDER ETYMOLOGY)', 'PICK']
r = col_headers(ws2, r, 7, headers)

for cat_name, cat_data in ACCOUNTS.items():
    r = section_header(ws2, r, 7, f"{cat_name} — {cat_data['desc']}")
    for idx, (name, platform, niche, reason, pick) in enumerate(cat_data['names']):
        if pick == 'TOP PICK':
            bg = BG_PICK
            pick_font = FONT_GREEN
        elif pick == 'ALT PICK':
            bg = BG_GOLD
            pick_font = FONT_GOLD
        else:
            bg = BG_ROW1 if idx % 2 == 0 else BG_ROW2
            pick_font = FONT_SMALL

        fill_bg(ws2, r, 7, bg)
        ws2.cell(row=r, column=2, value=name).font = FONT_HANDLE if pick else FONT_WHITE
        ws2.cell(row=r, column=3, value=platform).font = FONT_BODY
        ws2.cell(row=r, column=3).alignment = ALIGN_C
        ws2.cell(row=r, column=4, value=niche).font = FONT_CYAN if pick == 'TOP PICK' else FONT_BODY
        ws2.cell(row=r, column=4).alignment = ALIGN_C
        ws2.cell(row=r, column=5, value=reason).font = FONT_BODY
        ws2.cell(row=r, column=5).alignment = ALIGN_L
        ws2.cell(row=r, column=6, value=pick if pick else '').font = pick_font
        ws2.cell(row=r, column=6).alignment = ALIGN_C
        for c in range(1, 7):
            ws2.cell(row=r, column=c).border = BORDER
        ws2.row_dimensions[r].height = 38
        r += 1
    r += 1

# ============================================================
# SHEET 3: CROSS-PLATFORM MAPPING
# ============================================================
ws3 = wb.create_sheet('CROSS-PLATFORM')
ws3.sheet_properties.tabColor = '3FB950'
set_widths(ws3, [3, 18, 22, 22, 22, 22, 28, 3])

for r in range(1, 60):
    fill_bg(ws3, r, 8, BG_DARK)

ws3.merge_cells('B2:G2')
ws3.cell(row=2, column=2, value='CROSS-PLATFORM HANDLE MAPPING').font = FONT_TITLE
ws3.row_dimensions[2].height = 32
ws3.merge_cells('B3:G3')
ws3.cell(row=3, column=2, value='Unified naming across X, TikTok, YouTube, Instagram, Email. Handle availability TBD.').font = FONT_SUB

r = 5
headers3 = ['', 'NICHE', 'X / TWITTER', 'TIKTOK', 'YOUTUBE', 'INSTAGRAM', 'EMAIL']
r = col_headers(ws3, r, 8, headers3)

cross_platform = [
    ('Build in Public', '@printmaxxer', '@printmaxxer', 'PrintMaxxer', '@printmaxxer', 'printmaxxer@protonmail.com'),
    ('AI Tools', '@patchnotes_', '@patchnotesai', 'Patch Notes', '@patchnotes.ai', 'patchnotes.ai@protonmail.com'),
    ('Indie Hacking', '@shipclock', '@shipclock', 'ShipClock', '@shipclock', 'shipclock@protonmail.com'),
    ('Cold Email', '@pipelinepoisn', '@pipelinepoisn', 'Pipeline Poison', '@pipelinepoisn', 'pipelinepoisn@protonmail.com'),
    ('Trading', '@bidaskspread', '@bidaskspread', 'Bid Ask Spread', '@bidaskspread', 'bidaskspread@protonmail.com'),
    ('Crypto', '@mempooldive', '@mempooldive', 'Mempool Dive', '@mempooldive', 'mempooldive@protonmail.com'),
    ('Memes', '@foldersdump', '@foldersdump', 'Folders Dump', '@foldersdump', 'foldersdump@protonmail.com'),
    ('Fitness', '@repscheme_', '@repscheme', 'Rep Scheme', '@repscheme', 'repscheme@protonmail.com'),
    ('Faith', '@morningwatch_', '@morningwatch', 'Morning Watch', '@morningwatch_', 'morningwatch@protonmail.com'),
    ('AI Persona 1', '@kiraellis_', '@kiraellis', 'Kira Ellis', '@kiraellis', '—'),
    ('AI Persona 2', '@siennacurates', '@siennacurates', 'Sienna Curates', '@siennacurates', '—'),
    ('AI Persona 3', '@ivorsteel', '@ivorsteel', 'Ivor Steel', '@ivorsteel', '—'),
    ('Viral Clips', '@rawclipvault', '@rawclipvault', 'Raw Clip Vault', '@rawclipvault', 'rawclipvault@protonmail.com'),
    ('Newsletter', '@signalnoise_', '@signalnoise', 'Signal/Noise', '@signalnoise_', 'signalnoise@protonmail.com'),
    ('Sleep/ASMR', '@sleepcodex', '@sleepcodex', 'Sleep Codex', '@sleepcodex', 'sleepcodex@protonmail.com'),
    ('Nostalgia', '@retrowavefeed', '@retrowavefeed', 'Retrowave Feed', '@retrowavefeed', 'retrowavefeed@protonmail.com'),
    ('Quotes', '@quotesupply', '@quotesupply', 'Quote Supply', '@quotesupply', 'quotesupply@protonmail.com'),
    ('Late Night', '@nightshiftclips', '@nightshiftclips', 'Night Shift Clips', '@nightshiftclips', '—'),
]

for idx, (niche, x, tt, yt, ig, email) in enumerate(cross_platform):
    bg = BG_ROW1 if idx % 2 == 0 else BG_ROW2
    fill_bg(ws3, r, 8, bg)
    ws3.cell(row=r, column=2, value=niche).font = FONT_CYAN
    ws3.cell(row=r, column=3, value=x).font = FONT_WHITE
    ws3.cell(row=r, column=4, value=tt).font = FONT_BODY
    ws3.cell(row=r, column=5, value=yt).font = FONT_BODY
    ws3.cell(row=r, column=6, value=ig).font = FONT_BODY
    ws3.cell(row=r, column=7, value=email).font = FONT_SMALL
    for c in range(1, 8):
        ws3.cell(row=r, column=c).border = BORDER
        ws3.cell(row=r, column=c).alignment = ALIGN_C
    ws3.row_dimensions[r].height = 26
    r += 1

# ============================================================
# SHEET 4: NAMING RULES (anti-cringe guide)
# ============================================================
ws4 = wb.create_sheet('NAMING RULES')
ws4.sheet_properties.tabColor = 'F0883E'
set_widths(ws4, [3, 20, 40, 40, 3])

for r in range(1, 45):
    fill_bg(ws4, r, 5, BG_DARK)

ws4.merge_cells('B2:D2')
ws4.cell(row=2, column=2, value='ANTI-CRINGE NAMING RULES').font = FONT_TITLE
ws4.row_dimensions[2].height = 32
ws4.merge_cells('B3:D3')
ws4.cell(row=3, column=2, value='Patterns extracted from 145+ scraped accounts that actually work').font = FONT_SUB

r = 5
rules = [
    ('RULE', 'DO THIS', 'NOT THIS'),
    ('Use insider terminology', '@unusual_whales (trading lingo), @thetadecay_ (options Greek), @deloadweek (gym term)', '@TradingTipsDaily, @FitnessGuru123, @CryptoExpert'),
    ('Lowercase casual energy', '@videosinfolder, @levelsio, @cursaboringdays', '@EPIC_CONTENT, @BEST_TRADES, @TOP_QUOTES'),
    ('Contradiction/paradox', '@oddlyterrifying, @lowresgold, @factcoldcuts', '@AmazingFacts, @CoolContent, @GreatVideos'),
    ('Specific > generic', '@NoContextBrits (region), @OldSchoolCool80 (era)', '@FunnyMemes, @ViralClips, @DailyContent'),
    ('Underscore = dev/tech energy', '@patchnotes_, @latentspace_, @morningwatch_', 'No underscore = fine too, just signals technical'),
    ('Name + concept (for personas)', '@kiraellis_ (real name feel), @ivorsteel (strong name)', '@AI_Fitness_Coach, @Digital_Lisa, @VirtualJane'),
    ('Problem-as-name', '@promptdrift (LLM problem), @bounceratekid (email problem)', '@EmailSolutions, @AIFixer, @GrowthHacker'),
    ('Format-as-name', '@foldersdump (content dump), @rawclipvault (clip archive)', '@ContentCollection, @VideoArchive, @MemeDatabase'),
    ('Numbers add mystique', '@Rainmaker1973, @tom777kruise, @KCodes7777', 'Random numbers with no meaning: @user38492'),
    ('Self-deprecating humor', '@spamfolder_ (lives in spam), @gasfeesonly (all money = gas)', '@SuccessfulTrader, @TopPerformer, @Winner'),
]

for i, (rule, do, dont) in enumerate(rules):
    bg = BG_HEADER if i == 0 else (BG_ROW1 if i % 2 == 0 else BG_ROW2)
    font = FONT_HEAD if i == 0 else FONT_GOLD
    fill_bg(ws4, r, 5, bg)
    ws4.cell(row=r, column=2, value=rule).font = font
    ws4.cell(row=r, column=2).alignment = ALIGN_L
    ws4.cell(row=r, column=3, value=do).font = FONT_GREEN if i > 0 else FONT_HEAD
    ws4.cell(row=r, column=3).alignment = ALIGN_L
    ws4.cell(row=r, column=4, value=dont).font = Font(name='Arial', size=10, color='F85149') if i > 0 else FONT_HEAD
    ws4.cell(row=r, column=4).alignment = ALIGN_L
    for c in range(1, 5):
        ws4.cell(row=r, column=c).border = BORDER
    ws4.row_dimensions[r].height = 38 if i > 0 else 26
    r += 1

# Banned words section
r += 2
ws4.merge_cells(f'B{r}:D{r}')
ws4.cell(row=r, column=2, value='BANNED WORDS — NEVER USE IN HANDLES').font = Font(name='Arial', size=14, bold=True, color='F85149')
fill_bg(ws4, r, 5, BG_SECTION)
r += 1

banned = [
    'guru, expert, master, coach, pro, tips, hacks, mindset, grind, hustle',
    'unlock, secrets, revealed, proven, guaranteed, ultimate, best, top',
    'alpha (as "alpha male"), sigma, based (when forced), king, queen',
    'AI_ prefix (screams bot), Digital_, Virtual_, Smart_, Automated_',
    'Daily[Niche], [Niche]Tips, [Niche]Guru, [Niche]Master, [Niche]Pro',
]
for ban in banned:
    fill_bg(ws4, r, 5, BG_DARK)
    ws4.merge_cells(f'B{r}:D{r}')
    ws4.cell(row=r, column=2, value=f'  {ban}').font = Font(name='Arial', size=10, color='F85149')
    ws4.cell(row=r, column=2).alignment = ALIGN_L
    ws4.row_dimensions[r].height = 22
    r += 1


OUTPUT = '/sessions/awesome-nice-brown/mnt/PRINTMAXX_STARTER_KITttttt/PRINTMAXX_BRAND_NAMES.xlsx'
wb.save(OUTPUT)
print(f'Saved {OUTPUT}')
