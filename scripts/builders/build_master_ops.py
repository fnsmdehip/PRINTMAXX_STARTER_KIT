#!/usr/bin/env python3
"""
PRINTMAXX MASTER OPS — THE REAL ONE
Every possible op from: 71 money methods, 836 alpha entries,
50+ automation scripts, 38 niches, 16 synergy packages,
first principles, and new tool research.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# THEME
BG = PatternFill('solid', fgColor='0D1117')
HDR = PatternFill('solid', fgColor='1A2332')
R1 = PatternFill('solid', fgColor='0D1117')
R2 = PatternFill('solid', fgColor='161B22')
SEC = PatternFill('solid', fgColor='0A1628')
GF = PatternFill('solid', fgColor='0A2A15')
WF = PatternFill('solid', fgColor='2A2000')

T = Font(name='Arial', color='00D4FF', bold=True, size=16)
H = Font(name='Arial', color='00D4FF', bold=True, size=11)
H2 = Font(name='Arial', color='58A6FF', bold=True, size=10)
W = Font(name='Arial', color='FFFFFF', size=10)
WB = Font(name='Arial', color='FFFFFF', bold=True, size=10)
G = Font(name='Arial', color='00FF88', bold=True, size=10)
GO = Font(name='Arial', color='FFD700', bold=True, size=10)
R = Font(name='Arial', color='FF4444', size=10)
B = Font(name='Arial', color='C9D1D9', size=10)
THIN = Border(bottom=Side(style='thin', color='333333'))

def style(ws, ncols, widths):
    ws.sheet_properties.tabColor = '00D4FF'
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ncols):
        for c in row:
            c.fill = BG
            c.alignment = Alignment(wrap_text=True, vertical='top')
            c.border = THIN

def hdr(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = HDR
        c.font = H
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# ============================================================
# SHEET 1: ALL OPS MASTER (THE REAL LIST)
# ============================================================
ws1 = wb.active
ws1.title = 'ALL OPS MASTER'

ws1.cell(row=2, column=2, value='PRINTMAXX MASTER OPS — EVERY POSSIBLE OPERATION').font = Font(name='Arial', color='00D4FF', bold=True, size=18)
ws1.merge_cells('B2:L2')
ws1.cell(row=3, column=2, value='71 money methods + 836 alpha + 38 niches + first principles + tool research. If it can print, it\'s here.').font = B
ws1.merge_cells('B3:L3')

headers = ['OP', 'OPERATION', 'CATEGORY', 'REVENUE', 'AUTO LVL', 'VIDEO STACK', 'HOSTING', 'COST', 'EXISTS IN SYSTEM', 'LLM ALPHA', 'PRIORITY', 'METHOD IDS']
hdr(ws1, 5, headers)

# THE MASTER LIST — organized by category
ops = [
    # === CONTENT OPS ===
    ('', 'CONTENT CREATION OPS', '', '', '', '', '', '', '', '', '', ''),
    ('C01', 'FACELESS YOUTUBE LONGFORM', 'CONTENT', '$100-10K/mo/ch', 'FULL AUTO', 'Kling (66 free/day) + Remotion + ElevenLabs ($5/mo)', 'YouTube', '$5/mo', 'YES: CONTENT_FARM/SLEEP_YOUTUBE/', 'Claude scripts + ElevenLabs voice + Kling visuals. 1 video/day per channel. Human: review script (2 min)', 'P1', 'MM014,MM039,CF001-CF013'),
    ('C02', 'YOUTUBE SHORTS FARM (13 NICHES)', 'CONTENT', '$200-5K/mo', 'FULL AUTO', 'Kling + Remotion + stock footage', 'YouTube', '$0', 'YES: content farm infrastructure', 'Claude writes scripts, Remotion renders, auto-upload. 5-10 shorts/day/niche. 13 niches = 65-130 shorts/day', 'P1', 'CF001-CF013'),
    ('C03', 'TIKTOK CONTENT FARM (38 NICHES)', 'CONTENT', '$500-20K/mo', 'FULL AUTO', 'Kling + CapCut + Remotion', 'TikTok', '$0', 'YES: accounts + content ready', 'Same pipeline as Shorts but vertical. TikTok Creativity Program RPM $0.50-1/1K. 38 niches x 3 posts/day = 114 posts/day', 'P1', 'MM037,CF001-CF013,N001-N038'),
    ('C04', 'INSTAGRAM REELS FARM', 'CONTENT', '$200-5K/mo', 'FULL AUTO', 'Kling + CapCut', 'Instagram', '$0', 'YES: accounts planned', 'Repurpose TikTok/Shorts content to Reels. DM shares = new ranking signal 2026. Cross-post engine', 'P2', 'CF001-CF013'),
    ('C05', 'PINTEREST AFFILIATE ENGINE', 'AFFILIATE', '$200-5K/mo', 'FULL AUTO', 'Canva templates + Nano Banana (2 free/day)', 'Pinterest', '$15/mo Tailwind', 'PARTIAL: MM038 planned', '50-100 pins/day. 6-12 month shelf life per pin. Claude writes SEO descriptions. Compounding traffic', 'P2', 'MM038'),
    ('C06', 'MEDIUM ARTICLE FARM', 'CONTENT', '$100-2K/mo', 'FULL AUTO', 'N/A', 'Medium', '$0', 'YES: medium_articles/ ready', 'Claude writes 2-5 articles/day. Medium Partner Program pays per read. SEO + Medium algo. $0 cost', 'P2', 'MM021'),
    ('C07', 'SUBSTACK NEWSLETTER + NOTES', 'NEWSLETTER', '$500-10K/mo', 'SEMI-AUTO', 'N/A', 'Substack', '$0', 'YES: welcome sequences ready', 'Claude writes newsletter issues. Substack Notes = Twitter-like algo. Paid subs $5-15/mo. 3 publications', 'P1', 'MM015,MM032'),
    ('C08', 'BEEHIIV NEWSLETTER EMPIRE x3', 'NEWSLETTER', '$500-68K/mo', 'SEMI-AUTO', 'N/A', 'Beehiiv', '$0 (free tier x3)', 'YES: 3 pubs planned', 'Claude writes 3+ newsletters/week. Beehiiv free = 2,500 subs x 3 = 7,500 free. Sponsorships + affiliate', 'P1', 'MM015'),
    ('C09', 'LINKEDIN CONTENT ENGINE', 'CONTENT', '$0 + lead gen', 'FULL AUTO', 'N/A', 'LinkedIn', '$0', 'YES: linkedin posts ready', 'Claude writes daily LinkedIn posts. Build authority → inbound leads for services. DM outreach for B2B', 'P2', 'MM021'),
    ('C10', 'X/TWITTER MONETIZATION', 'CONTENT', '$0-5K/mo', 'SEMI-AUTO', 'N/A', 'X/Twitter', '$0', 'YES: 130 tweets ready', 'Build in public. @PRINTMAXXER voice. Subscriptions + tips + affiliate. Tweetlio unlimited free posting', 'P1', 'MM040'),
    ('C11', 'REDDIT CONTENT + AFFILIATE', 'CONTENT', '$200-2K/mo', 'SEMI-AUTO', 'N/A', 'Reddit', '$0', 'YES: reddit posts ready', 'Claude writes posts for 41 subreddits. Affiliate links in bio/comments. GEO-optimized for AI citations', 'P2', 'MM003'),
    ('C12', 'AI PODCAST FACTORY', 'CONTENT', '$200-5K/mo', 'FULL AUTO', 'ElevenLabs ($5/mo) + Remotion', 'Spotify/Apple', '$5/mo', 'NO: new op', 'Claude writes scripts, ElevenLabs voices 2 hosts, Remotion makes video podcast, DistroKid for audio. 3-5 episodes/week', 'P3', 'NEW'),
    ('C13', 'AI MUSIC FACTORY (SUNO)', 'MUSIC', '$0-2K+/mo', 'FULL AUTO', 'Suno (10 free/day) + DistroKid ($23/yr)', 'Spotify/Apple/TikTok', '$23/yr', 'PARTIAL: method documented', '10 songs/day free via Suno. DistroKid distributes to 150+ platforms. Focus: lo-fi, ambient, meditation, workout playlists', 'P3', 'MM012'),
    ('C14', 'REMOTION BRANDED VIDEO PRODUCTION', 'CONTENT', '$0 + force multiplier', 'FULL AUTO', 'Remotion (FREE individual) + Claude Code', 'N/A', '$0', 'YES: Remotion project exists', 'Claude Code writes React video compositions. Auto-render app demos, product videos, social clips. 10 app videos queued', 'P1', 'NEW'),
    ('C15', 'FACEBOOK REELS CROSSPOST', 'CONTENT', '$0-1K/mo', 'FULL AUTO', 'repurpose from C03/C04', 'Facebook', '$0', 'YES: FB_REELS_GTM.md exists', 'Crosspost TikTok/Reels to FB. Lower RPM ($0.02-0.60) but free traffic. Zero marginal effort', 'P3', 'CF001-CF013'),

    # === ECOMM OPS ===
    ('', 'E-COMMERCE OPS', '', '', '', '', '', '', '', '', '', ''),
    ('E01', 'ECOMM TRENDING PRODUCT ARB', 'ECOMM', '$50-2K+/mo', 'FULL AUTO', 'Kling (product demo videos)', 'Etsy/eBay', '$0 to list', 'YES: ecom_arb_scanner.py built', 'Claude scrapes TikTok Shop + Etsy trending. Scores demand. Auto-generates listings. Dropship from Temu/AliExpress. 6hr cron cycle', 'P1', 'MM022,MM023'),
    ('E02', 'PRINT ON DEMAND AUTO-DESIGN', 'ECOMM', '$0-1K+/mo', 'FULL AUTO', 'Leonardo.ai (150 free/day) + Nano Banana', 'Redbubble/TeePublic', '$0', 'YES: POD/ method exists', 'Claude monitors trends → generates design concepts → Leonardo/Nano Banana creates → upload to POD platforms. 20-50 designs/day', 'P2', 'MM024'),
    ('E03', 'AMAZON KDP BOOK FACTORY', 'ECOMM', '$0-1K+/mo', 'FULL AUTO', 'Canva (interiors)', 'Amazon KDP', '$0', 'PARTIAL: method documented', 'Claude generates low-content books: journals, planners, coloring books, logbooks. 2-3 per day. Canva for covers/interiors', 'P2', 'MM026'),
    ('E04', 'TIKTOK SHOP AFFILIATE + DROPSHIP', 'ECOMM', '$500-20K/mo', 'SEMI-AUTO', 'Kling + HeyGen for UGC', 'TikTok Shop', '$0', 'YES: TIKTOK_SHOP/ playbook', 'Find trending products → AI UGC video → post → earn 10-30% commission. Claude identifies winners from trending data', 'P1', 'MM016'),
    ('E05', 'ETSY DIGITAL PRODUCTS', 'ECOMM', '$0-5K/mo', 'FULL AUTO', 'Canva', 'Etsy', '$0.20/listing', 'YES: MM036 documented', 'Claude generates: SVG bundles, invitation templates, resume templates, wall art, planners. $3-15/download. Volume play', 'P2', 'MM036'),
    ('E06', 'SHOPIFY STORE (NICHE)', 'ECOMM', '$500-10K/mo', 'SEMI-AUTO', 'Kling for product videos', 'Shopify', '$39/mo', 'NO: new op', 'Claude Code builds entire Shopify store in 2-4 hours. Niche: faith merch, fitness gear, pet products. Dropship or POD backend', 'P3', 'MM022'),
    ('E07', 'AI STOCK FOOTAGE LIBRARY', 'ECOMM', '$100-2K/mo', 'FULL AUTO', 'Kling + Veo (Google Veo)', 'Shutterstock/Adobe Stock', '$0', 'NO: new op', 'Generate AI video clips via Kling/Veo. Upload to stock footage sites. $0.25-15 per download. Volume: 50-100 clips/day', 'P3', 'NEW'),

    # === DIGITAL PRODUCT OPS ===
    ('', 'DIGITAL PRODUCT OPS', '', '', '', '', '', '', '', '', '', ''),
    ('D01', 'NOTION TEMPLATE STORE', 'DIGITAL', '$500-10K/mo', 'FULL AUTO', 'N/A', 'Gumroad/Notion', '$0', 'YES: MM046 + templates ready', 'Claude creates Notion templates at 50x human speed. Gumroad $0 until sale. Target: productivity, business, students, creators', 'P1', 'MM046'),
    ('D02', 'GUMROAD DIGITAL PRODUCTS', 'DIGITAL', '$500-10K/mo', 'SEMI-AUTO', 'N/A', 'Gumroad', '$0 (5% fee)', 'YES: 4 products spec\'d', 'Ebooks, checklists, frameworks, swipe files, prompt libraries. Claude generates in hours. 4 products already spec\'d and ready', 'P1', 'MM025'),
    ('D03', 'COURSE / INFO PRODUCT FACTORY', 'DIGITAL', '$1K-50K/mo', 'SEMI-AUTO', 'Remotion + ElevenLabs for video courses', 'Udemy/Gumroad/Skool', '$0-99/mo', 'PARTIAL: method documented', 'Claude writes 50-100 lesson courses in hours. Udemy=distribution. Gumroad=higher margin. Skool=$99/mo community', 'P1', 'MM002,MM030'),
    ('D04', 'CANVA TEMPLATE STORE', 'DIGITAL', '$200-5K/mo', 'FULL AUTO', 'Canva (free)', 'Etsy/Gumroad/Creative Market', '$0', 'NO: new op', 'Social media templates, presentation templates, resume templates. Claude designs layouts, Canva implements. 10-20/day', 'P3', 'NEW'),
    ('D05', 'CHROME EXTENSION PORTFOLIO', 'DIGITAL', '$0-5K/mo', 'SEMI-AUTO', 'N/A', 'Chrome Web Store', '$5 one-time', 'NO: new op from first principles', 'Claude Code builds Chrome extensions in 30-60 min each. Freemium model. 10-20 extensions = portfolio. $2-5/mo per user', 'P2', 'MM027,MM028'),
    ('D06', 'AI PROMPT LIBRARY / MARKETPLACE', 'DIGITAL', '$200-5K/mo', 'FULL AUTO', 'N/A', 'Gumroad/PromptBase', '$0', 'NO: new op', 'Curate + create premium prompt packs. Claude writes prompts for: business, coding, writing, marketing. $9-29 per pack', 'P3', 'NEW'),

    # === SERVICE / FREELANCE OPS ===
    ('', 'SERVICE & FREELANCE OPS', '', '', '', '', '', '', '', '', '', ''),
    ('S01', 'FREELANCE SERVICE ARB (CLAUDE CODE MAX)', 'FREELANCE', '$1K-20K/mo', 'FULL AUTO', 'N/A', 'Fiverr/Upwork/Contra/etc', '$0', 'YES: FREELANCE_ARB.xlsx built', '30 services on 10 platforms. 95%+ margin. Claude Code builds in 5-60 min, client pays $50-500 for 2-3 day delivery', 'P1', 'NEW-OP17'),
    ('S02', 'AI UGC VIDEO FACTORY', 'SERVICE', '$500-10K/mo', 'FULL AUTO', 'HeyGen ($29/mo) + Kling (free) + Veo', 'N/A', '$29/mo', 'NO: new op', 'AI UGC at $0.50-2/video vs human $50-500. Sell to DTC brands. List on Billo/Insense. 50-100 videos/day', 'P1', 'MM008,NEW-OP18'),
    ('S03', 'LOCAL BIZ LEAD GEN + WEB REDESIGN', 'SERVICE', '$2K-20K/mo', 'SEMI-AUTO', 'N/A', 'Netlify (free, commercial OK) + Cloudflare', '$30/mo Instantly', 'YES: local_biz_pipeline.py BUILT', 'Pipeline ALREADY BUILT. Scrape → score → Claude generates mockup → deploy Netlify → cold email via Instantly. 100 emails/day', 'P1', 'MM070,MM029'),
    ('S04', 'COLD EMAIL B2B OUTBOUND', 'OUTBOUND', '$1K-10K/mo', 'SEMI-AUTO', 'N/A', 'N/A', '$30/mo Instantly', 'YES: cold outbound playbooks ready', 'Claude personalizes every email. Clay enrichment. Instantly sends. 2-5% reply rate. Lead gen for ANY other op', 'P1', 'MM007'),
    ('S05', 'AI VOICE OUTREACH (BLAND AI)', 'OUTBOUND', '$1K-10K/mo', 'FULL AUTO', 'N/A', 'N/A', '$0 (100 free calls/day)', 'NO: new op from research', 'Bland AI: 100 FREE calls/day. Claude writes call scripts. AI voices make outbound calls. Book meetings. Qualify leads. $0 cost on free tier', 'P1', 'NEW'),
    ('S06', 'CRUNCHBASE FUNDED CO. SCRAPING', 'OUTBOUND', '$0 + lead gen', 'FULL AUTO', 'N/A', 'N/A', '$0-49/mo', 'NO: new op from user suggestion', 'Scrape recently funded companies → identify needs (websites, apps, marketing) → cold email offering services. Funded = budget available', 'P1', 'NEW'),
    ('S07', 'BUILTWITH OUTDATED WEBSITE DETECTOR', 'OUTBOUND', '$0 + lead gen', 'FULL AUTO', 'N/A', 'N/A', '$0 Wappalyzer ext', 'NO: new op from user suggestion', 'Wappalyzer (free browser ext) detects tech stacks. Find sites on WordPress 4.x, jQuery 1.x, no SSL, no mobile. Score outdatedness. Feed to S03 pipeline', 'P1', 'NEW'),
    ('S08', 'SEO/GEO AUDIT SERVICE', 'SERVICE', '$500-5K/project', 'SEMI-AUTO', 'N/A', 'N/A', '$0', 'YES: GEO playbooks exist', 'Claude audits websites for SEO/GEO compliance. Generates report. Offer fix-it service. $500-2K per audit+fix. Claude Code implements fixes', 'P2', 'MM005'),
    ('S09', 'SOCIAL MEDIA MANAGEMENT (AI)', 'SERVICE', '$500-3K/client/mo', 'FULL AUTO', 'Kling + Canva', 'Buffer (free 3 ch)', '$0', 'YES: content infra exists', 'Claude generates all content. Buffer schedules. Report monthly. Manage 5-10 clients at $500-1K/mo each. 90% automated', 'P2', 'MM005'),
    ('S10', 'WHITE LABEL DEV (AGENCY SUB)', 'SERVICE', '$2K-20K/mo', 'SEMI-AUTO', 'N/A', 'Vercel/Netlify', '$0', 'NO: new op', 'Approach agencies, offer white-label dev at 50% of client rate. They charge $5K, pay you $2.5K. Claude Code delivers. Volume', 'P2', 'MM033'),
    ('S11', 'AI CHATBOT INSTALLATION SERVICE', 'SERVICE', '$500-2K/project', 'FULL AUTO', 'N/A', 'Client site', '$0', 'NO: new op from first principles', 'Build custom AI chatbots for businesses using Claude/OpenAI API. Charge $500-2K per install + $100/mo maintenance. Claude Code builds in 30 min', 'P2', 'NEW'),
    ('S12', 'RESUME/CV WRITING SERVICE', 'SERVICE', '$50-200/resume', 'FULL AUTO', 'N/A', 'Fiverr/Upwork', '$0', 'NO: new op', 'List on Fiverr. Claude writes + formats professional resumes in 5 min. $50-200 per resume. Volume: 5-10/day', 'P2', 'NEW'),
    ('S13', 'AI SUBCONTRACTING NETWORK', 'SERVICE', '$5K-50K/mo', 'SEMI-AUTO', 'N/A', 'N/A', '$0', 'NO: new op from first principles', 'Find subcontract work from agencies/founders who need dev work done. Claude Code fulfills. Build reputation → referral network', 'P2', 'MM005,MM033'),

    # === APP / SAAS OPS ===
    ('', 'APP & SAAS OPS', '', '', '', '', '', '', '', '', '', ''),
    ('A01', 'APP FACTORY (2-3/MONTH)', 'APPS', '$1K-50K/mo portfolio', 'SEMI-AUTO', 'Remotion for demos', 'App Store/Play Store', '$99/yr Apple + $25 Google', 'YES: 4 apps built, 13 in dev', 'Claude Code builds apps in 24-48hr sprints. Portfolio approach: 30 apps → 5% hit rate = 1-2 winners. PrayerLock, WalkToUnlock ready', 'P1', 'MM001,MM019'),
    ('A02', 'MICRO-SAAS PORTFOLIO', 'SAAS', '$500-50K/mo', 'SEMI-AUTO', 'N/A', 'Vercel (free) + Netlify', '$0', 'YES: method documented', 'Single-feature tools: PDF merger, invoice gen, QR maker. Claude Code builds MVP in 24hrs. Launch on ProductHunt. Freemium $9-29/mo', 'P1', 'MM028,MM044'),
    ('A03', 'AI WRAPPER PRODUCTS', 'SAAS', '$200-10K/mo', 'SEMI-AUTO', 'N/A', 'Vercel/Netlify', '$0 + API costs', 'YES: MM027 documented', 'Wrap Claude/OpenAI API for niche use cases. AI resume writer, AI blog writer, AI email writer, AI code reviewer. $10-50/mo', 'P2', 'MM027'),
    ('A04', 'CHROME EXTENSION PORTFOLIO', 'APPS', '$0-5K/mo', 'SEMI-AUTO', 'N/A', 'Chrome Web Store', '$5', 'NO: new op', 'Claude Code builds extensions in 30-60 min. Productivity tools, SEO tools, social tools. 10-20 extensions. Freemium', 'P2', 'NEW'),
    ('A05', 'ROBLOX GAME FACTORY', 'APPS', '$100-10K/mo', 'SEMI-AUTO', 'N/A', 'Roblox', '$0', 'YES: robloxmaxx game built', 'Claude Code writes Luau. Tycoon games, obbies, simulators. Robux + premium payouts. Kids demographic = retention', 'P3', 'MM011'),
    ('A06', 'DISCORD BOT MARKETPLACE', 'APPS', '$100-3K/mo', 'SEMI-AUTO', 'N/A', 'Discord', '$0', 'NO: new op', 'Build premium Discord bots: moderation, ticketing, analytics, music. List on top.gg. Freemium $5-15/mo per server', 'P3', 'NEW'),
    ('A07', 'DIRECTORY LISTING SITES', 'SAAS', '$5K-50K/mo', 'SEMI-AUTO', 'N/A', 'Vercel/Netlify', '$0', 'YES: MM041 documented', 'Niche directories: AI tools, SaaS products, agencies. Featured listings $50-500/mo. Claude Code builds entire directory in 48hrs', 'P2', 'MM041'),
    ('A08', 'MCP SERVER MARKETPLACE', 'SAAS', '$2K-20K/mo', 'SEMI-AUTO', 'N/A', 'Vercel/Netlify', '$0', 'YES: MCP build plan exists', 'Build Model Context Protocol servers for niche use cases. Sell on marketplace or as SaaS. First-mover advantage', 'P2', 'TOOL_ALPHA'),

    # === PERSONA / INFLUENCER OPS ===
    ('', 'AI PERSONA & INFLUENCER OPS', '', '', '', '', '', '', '', '', '', ''),
    ('P01', 'AI NICHE EXPERT PERSONAS (8 SIMULTANEOUS)', 'PERSONA', '$500-10K/mo/persona', 'SEMI-AUTO', 'Leonardo.ai (150 free/day) + Nano Banana + Kling', 'Multi-platform', '$0', 'YES: AI001 execution plan', '8 personas across niches. Each: consistent face via Leonardo, Claude writes posts, Kling for video. Revenue: sponsorships, courses, affiliate', 'P1', 'AI001-AI008'),
    ('P02', 'AI FITNESS COACHES', 'PERSONA', '$500-10K/mo', 'SEMI-AUTO', 'Kling for workout demos', 'IG/TikTok/YouTube', '$0', 'YES: AI005 documented', 'AI fitness persona: workout plans, nutrition tips, transformation content. Monetize: programs, affiliate, sponsorships', 'P2', 'AI005'),
    ('P03', 'AI ASMR CHANNELS', 'PERSONA', '$500-5K/mo', 'FULL AUTO', 'ElevenLabs + Kling', 'YouTube/TikTok', '$5/mo', 'YES: AI004 documented', 'AI-generated ASMR content. ElevenLabs for whisper voices. Kling for visuals. YouTube ads + Patreon', 'P3', 'AI004'),
    ('P04', 'AI LIFESTYLE CURATORS', 'PERSONA', '$300-5K/mo', 'FULL AUTO', 'Leonardo.ai + Nano Banana + Kling', 'IG/Pinterest/TikTok', '$0', 'YES: AI006 documented', 'Aesthetic travel/lifestyle content. All AI-generated images+video. Sponsorships, affiliate links, brand deals', 'P3', 'AI006'),
    ('P05', 'AI GAMING PERSONAS', 'PERSONA', '$300-5K/mo', 'SEMI-AUTO', 'Kling + screen recording', 'Twitch/YouTube/TikTok', '$0', 'YES: AI007 documented', 'Gaming commentary, tips, highlights. Claude writes scripts. Merch + sponsorships + subs', 'P3', 'AI007'),
    ('P06', 'AI RELATIONSHIP/DATING ADVICE', 'PERSONA', '$500-10K/mo', 'SEMI-AUTO', 'Kling', 'TikTok/YouTube/IG', '$0', 'YES: AI008 documented', 'Dating tips, relationship advice. Claude writes content. Courses + coaching upsell. High engagement niche', 'P2', 'AI008'),

    # === INVESTMENT / TRADING OPS ===
    ('', 'INVESTMENT & TRADING OPS', '', '', '', '', '', '', '', '', '', ''),
    ('I01', 'DOMAIN FLIPPING (LLM-SCOUTED)', 'TRADING', '$500-20K/mo', 'SEMI-AUTO', 'N/A', 'Namecheap/Cloudflare', '$8-12/domain', 'YES: MM045 documented', 'Claude scores 10,000+ expiring domains daily. Brandability, keyword value, TLD quality. Buy at $8-12, sell at $50-5000', 'P2', 'MM045'),
    ('I02', 'ALGO TRADING SIGNALS + EDUCATION', 'TRADING', '$1K-50K/mo', 'SEMI-AUTO', 'N/A', 'Custom dashboard', '$0', 'YES: algo_trading playbook', 'Claude analyzes on-chain data, options flow, earnings. Sell signals via Telegram/Discord. Courses on trading strategy', 'P3', 'MM012'),
    ('I03', 'MEMECOIN SIGNAL TRACKING', 'TRADING', '$100-10K/mo', 'SEMI-AUTO', 'N/A', 'Telegram/Discord', '$0', 'YES: meme_coin_signal_tracker.py built', '<5% allocation. Claude monitors Reddit/Twitter for early signals. $5-20/bet. High risk, high reward', 'P3', 'MM034'),

    # === COMMUNITY OPS ===
    ('', 'COMMUNITY & MEMBERSHIP OPS', '', '', '', '', '', '', '', '', '', ''),
    ('M01', 'SKOOL COMMUNITY ($99/MO)', 'COMMUNITY', '$5K-50K/mo', 'SEMI-AUTO', 'N/A', 'Skool', '$99/mo', 'YES: method documented', 'Premium community: all courses, Q&A, group calls. Claude generates daily prompts + content. Members who post in 48hrs retain 5x', 'P2', 'MM031'),
    ('M02', 'DISCORD PREMIUM SERVER', 'COMMUNITY', '$500-5K/mo', 'SEMI-AUTO', 'N/A', 'Discord', '$0', 'NO: new op', 'Paid Discord for niche community. Trading signals, alpha, networking. Claude moderates + generates content. $9-29/mo', 'P3', 'MM031'),
    ('M03', 'TELEGRAM VIP CHANNEL', 'COMMUNITY', '$200-5K/mo', 'FULL AUTO', 'N/A', 'Telegram', '$0', 'NO: new op', 'Premium Telegram with daily alpha, signals, content. Claude auto-generates daily posts. $5-15/mo. Low maintenance', 'P3', 'NEW'),
    ('M04', 'PATREON MEMBERSHIP', 'COMMUNITY', '$500-10K/mo', 'SEMI-AUTO', 'N/A', 'Patreon', '$0 (5-12% fee)', 'NO: new op', 'Tiered membership: $5 basic, $15 pro, $50 VIP. Claude generates exclusive content per tier. Works for any niche persona', 'P3', 'NEW'),

    # === AFFILIATE OPS ===
    ('', 'AFFILIATE OPS', '', '', '', '', '', '', '', '', '', ''),
    ('F01', 'AFFILIATE SEO CONTENT SITES', 'AFFILIATE', '$500-5K/mo/site', 'FULL AUTO', 'N/A', 'Netlify (free, commercial OK)', '$0', 'YES: MM003 documented', 'Claude generates 10-20 SEO articles/day targeting buyer intent keywords. Netlify hosting. Amazon/ShareASale affiliate. Programmatic SEO', 'P1', 'MM003'),
    ('F02', 'AMAZON ASSOCIATES PROGRAM', 'AFFILIATE', '$100-5K/mo', 'FULL AUTO', 'N/A', 'Blog + Pinterest', '$0', 'PARTIAL: documented', 'Product review sites + Pinterest pins. Claude writes reviews. 1-10% commission. Volume: 50-100 product pages', 'P2', 'MM003'),
    ('F03', 'SOFTWARE AFFILIATE (HIGH TICKET)', 'AFFILIATE', '$500-10K/mo', 'SEMI-AUTO', 'N/A', 'Blog/Newsletter/YouTube', '$0', 'PARTIAL: alpha references', 'Promote SaaS tools with $50-200/referral commissions. Comparison sites, review videos, tutorial content. Claude writes all content', 'P2', 'MM003'),

    # === GROWTH / INFRASTRUCTURE OPS ===
    ('', 'GROWTH HACKS & INFRASTRUCTURE OPS', '', '', '', '', '', '', '', '', '', ''),
    ('G01', '500-CLIPPER NETWORK MODEL', 'GROWTH', '$2K-20K/mo', 'SEMI-AUTO', 'N/A', 'N/A', '$0 (pay per clip)', 'YES: ALPHA008 documented', '$1/1K views model. 500 clippers create short clips from long content. Massive distribution network. Scale virality', 'P2', 'SWARM001'),
    ('G02', 'CONTENT REPURPOSING ENGINE (1→13)', 'GROWTH', '$0 + force multiplier', 'FULL AUTO', 'Remotion + Kling + Canva', 'All 13 platforms', '$0', 'YES: content_repurposer.py built', '1 piece of content → X thread, TikTok, YT Short, Reel, Pin, LinkedIn, Medium, Substack, Newsletter, Reddit, FB, Podcast, Email. Claude + Remotion', 'P1', 'NEW'),
    ('G03', 'DAILY ALPHA RESEARCH PIPELINE', 'RESEARCH', '$0 + force multiplier', 'FULL AUTO', 'N/A', 'N/A', '$0', 'YES: twitter+reddit scrapers built', '89 Twitter accounts + 41 subreddits + Product Hunt + GitHub trending. Auto-extract alpha → ALPHA_STAGING.csv → review → action', 'P1', 'EXISTING'),
    ('G04', 'PLATFORM ALGORITHM MONITORING', 'RESEARCH', '$0 + force multiplier', 'FULL AUTO', 'N/A', 'N/A', '$0', 'YES: platform_meta_monitor.py built', 'Detect TikTok/IG/YouTube/X algorithm changes in real-time. Adjust content strategy immediately. Competitive edge', 'P1', 'EXISTING'),
    ('G05', 'X LAUNCH VIRAL HACK', 'GROWTH', '$3K-30K/mo', 'SEMI-AUTO', 'N/A', 'X/Twitter', '$0', 'YES: MM020 documented', 'Comment-driven X launches. Build product → tweet thread → comments drive organic virality → MRR. $0 marketing cost', 'P2', 'MM020'),
    ('G06', 'WAITLIST/PRESALE SERVICE', 'GROWTH', '$3K-30K/mo', 'SEMI-AUTO', 'N/A', 'Custom landing page', '$0', 'YES: MM042 documented', '25-85% conversion vs 2-4% traditional. Build hype → presale → validate → deliver. Works for any product launch', 'P2', 'MM042'),
    ('G07', 'SWARM PROMOTION (MULTI-ACCOUNT)', 'GROWTH', '$500-10K/mo', 'FULL AUTO', 'N/A', 'Multi-platform', '$0', 'YES: SWARM001 documented', 'Coordinated multi-account virality. 5-10 accounts amplify each other. Anti-detect browser + proxies. Algorithm gaming', 'P2', 'SWARM001'),
    ('G08', 'CROSS-POLLINATION MATRIX', 'GROWTH', '$0 + force multiplier', 'FULL AUTO', 'N/A', 'N/A', '$0', 'YES: 371 stacks verified', 'Automated synergy detection between methods. 1.5-2.8x revenue multiplier when methods stack. Auto-suggest best combos', 'P1', 'EXISTING'),
    ('G09', 'RBI PERPETUAL IMPROVEMENT SYSTEM', 'RESEARCH', '$0 + force multiplier', 'FULL AUTO', 'N/A', 'N/A', '$0', 'BUILDING NOW', 'Daily automated audit of ALL ops. Score performance. Identify improvements. Suggest new ops. Self-improving system. Jane Street model', 'P1', 'NEW'),
    ('G10', 'PROGRAMMATIC SEO (1000+ PAGES)', 'GROWTH', '$500-10K/mo', 'FULL AUTO', 'N/A', 'Netlify/Vercel', '$0', 'YES: 300 longtail slugs, 200 GEO prompts', 'Claude generates 1000+ landing pages targeting long-tail keywords. Each page = SEO entry point. Programmatic approach', 'P1', 'EXISTING'),
    ('G11', 'DAILY MORNING BRIEFING', 'OPS', '$0 + force multiplier', 'FULL AUTO', 'N/A', 'N/A', '$0', 'YES: daily_briefing.py BUILT', 'Scans 10 systems every morning. Generates prioritized human-action-required report. Wired into printmaxx_cron.sh', 'LIVE', 'EXISTING'),
]

for i, op in enumerate(ops):
    row = 6 + i
    for j, val in enumerate(op, 1):
        c = ws1.cell(row=row, column=j, value=val)
        # Category header rows
        if op[0] == '' and op[1] != '':
            c.font = H
            c.fill = SEC
        elif op[0] == '':
            pass
        else:
            c.font = W if j > 2 else GO
            if i % 2 == 1:
                c.fill = R2
            # Highlight columns
            if j == 5:  # AUTO LVL
                c.font = G if 'FULL' in str(val) else (GO if 'SEMI' in str(val) else W)
            if j == 8:  # COST
                c.font = G if val == '$0' else GO
            if j == 9:  # EXISTS
                if str(val).startswith('YES'):
                    c.font = G
                elif str(val).startswith('PARTIAL'):
                    c.font = GO
                elif str(val).startswith('NO'):
                    c.font = Font(name='Arial', color='FF8844', size=10)
            if j == 11:  # PRIORITY
                c.font = G if val == 'P1' else (GO if val == 'P2' else W)

# Summary
sr = 6 + len(ops) + 2
# Count actual ops (non-header rows)
op_count = len([o for o in ops if o[0] != '' and o[1] != ''])
p1_count = len([o for o in ops if o[10] == 'P1'])
exists_count = len([o for o in ops if str(o[8]).startswith('YES')])
free_count = len([o for o in ops if o[7] == '$0'])

ws1.cell(row=sr, column=2, value=f'TOTAL OPS: {op_count}').font = Font(name='Arial', color='00D4FF', bold=True, size=14)
ws1.cell(row=sr+1, column=2, value=f'P1 PRIORITY: {p1_count} | ALREADY IN SYSTEM: {exists_count} | $0 COST: {free_count}').font = G
ws1.cell(row=sr+2, column=2, value='VIDEO: Kling (66 free/day) + Veo (limited) + Remotion (FREE) + ElevenLabs ($5/mo) + Nano Banana (2 free/day) + Leonardo.ai (150 free/day)').font = GO
ws1.cell(row=sr+3, column=2, value='HOSTING: Netlify (free, commercial OK) + Vercel (free, hobby only) + Cloudflare Pages (free) + GitHub Pages (free)').font = GO
ws1.cell(row=sr+4, column=2, value='LEAD GEN: Bland AI (100 FREE calls/day) + Instantly ($30/mo cold email) + Wappalyzer (free outdated site detection) + Crunchbase (free basic search)').font = GO

style(ws1, 12, [6, 35, 12, 18, 12, 35, 20, 16, 28, 55, 8, 25])

# ============================================================
# SHEET 2: VIDEO & MEDIA STACK
# ============================================================
ws2 = wb.create_sheet('VIDEO & MEDIA STACK')

ws2.cell(row=2, column=2, value='COMPLETE VIDEO / IMAGE / AUDIO PRODUCTION STACK').font = T
ws2.merge_cells('B2:F2')
ws2.cell(row=3, column=2, value='All free/cheap tools for content production. Use in combination.').font = B
ws2.merge_cells('B3:F3')

hdr(ws2, 5, ['TOOL', 'TYPE', 'FREE TIER', 'PAID', 'BEST FOR', 'API?'])

media_tools = [
    ('Kling AI', 'VIDEO GEN', '66 credits/day (720p, 5-10 sec clips). No rollover.', '$7-180/mo (1080p, longer, no watermark)', 'Product demos, faceless content, UGC-style video. Best free video gen available.', 'YES (paid tier)'),
    ('Google Veo', 'VIDEO GEN', 'Very limited free. ~3 videos/day on Pro ($20/mo)', '$20/mo (Google AI Pro)', 'High-quality video gen. Use sparingly for hero content. Complements Kling.', 'YES'),
    ('Remotion', 'VIDEO RENDER', 'FREE for individuals/teams <3. Unlimited videos.', '$25/dev/mo for companies', 'Programmatic video from React. Claude Code writes compositions. Auto-render app demos, social clips, branded content.', 'Self-hosted'),
    ('Nano Banana', 'IMAGE GEN', '2 images/day at 1K res. Watermarked.', '$20/mo (Google AI Pro, 100/day, no watermark)', 'Quick product images, social graphics. Use with Leonardo for volume.', 'YES ($0.13/image)'),
    ('Leonardo.ai', 'IMAGE GEN', '150 tokens/day (20-30 images). Commercial OK.', '$12-28/mo', 'Highest free volume for images. App icons, social graphics, POD designs, personas. 150 tokens = 20-30 images/day FREE.', 'YES'),
    ('ElevenLabs', 'VOICE/TTS', '20 min audio/month. Non-commercial.', '$5/mo (voice cloning, commercial)', 'Voiceovers for faceless YouTube, podcasts, UGC scripts. $5/mo unlocks commercial + voice cloning.', 'YES'),
    ('Suno', 'MUSIC GEN', '10 songs/day. Personal use only. 2 min max.', '$10/mo (500 songs, 8 min, commercial)', 'Lo-fi beats, ambient music, workout tracks. DistroKid distributes. AI music at scale.', 'NO (web only)'),
    ('CapCut', 'VIDEO EDIT', 'FREE. Full editor. Templates. Auto-captions.', '$8/mo (Pro features)', 'Quick edits, auto-captions, effects. Best free video editor for social content.', 'NO'),
    ('Canva', 'DESIGN', 'FREE. Templates, basic features.', '$13/mo (Pro)', 'Pin images, social graphics, presentations, Etsy templates. Best free design tool.', 'YES'),
    ('HeyGen', 'AI AVATAR', '3 min/video free. Limited avatars.', '$29/mo (unlimited standard)', 'AI UGC videos with talking avatars. Sell to brands. $29/mo = unlimited standard videos.', 'YES'),
    ('DistroKid', 'MUSIC DIST', 'N/A', '$23/yr (100% earnings kept)', 'Distribute AI music to Spotify, Apple, TikTok, 150+ platforms. Cheapest distributor.', 'NO'),
    ('Descript', 'PODCAST EDIT', 'Limited free tier', '$24/mo', 'Podcast editing, transcription, filler removal. Text-based editing.', 'NO'),
]

for i, tool in enumerate(media_tools):
    row = 6 + i
    for j, val in enumerate(tool, 1):
        c = ws2.cell(row=row, column=j, value=val)
        c.font = W if j > 1 else GO
        if i % 2 == 1:
            c.fill = R2

style(ws2, 6, [14, 12, 40, 30, 50, 14])

# ============================================================
# SHEET 3: HOSTING & DEPLOYMENT STACK
# ============================================================
ws3 = wb.create_sheet('HOSTING & DEPLOY')

ws3.cell(row=2, column=2, value='HOSTING & DEPLOYMENT — FREE TIER COMPARISON').font = T
ws3.merge_cells('B2:F2')

hdr(ws3, 5, ['PLATFORM', 'FREE TIER', 'COMMERCIAL USE', 'BEST FOR', 'LIMITS', 'RECOMMENDATION'])

hosting = [
    ('Netlify', '100GB bandwidth, 300 build min, unlimited sites', 'YES - commercial allowed on free tier', 'BULK LANDING PAGES. Best for deploying 50+ client mockups for local biz outreach. Commercial use OK.', '100GB/mo bandwidth. 300 build min/mo.', 'USE FOR: client mockup pages, affiliate sites, static landing pages. Commercial = OK.'),
    ('Vercel', '100GB bandwidth, edge functions, optimized for Next.js', 'NO - hobby only on free tier', 'YOUR OWN PROJECTS. Best for Next.js apps, micro-SaaS, portfolio. Not for commercial client work on free tier.', '100GB/mo. Hobby = non-commercial only.', 'USE FOR: your own apps, micro-SaaS, printmaxx-site. Need Pro ($20/mo) for commercial.'),
    ('Cloudflare Pages', '500 builds/mo, unlimited bandwidth, 20K files', 'YES - commercial allowed', 'STATIC SITES. Unlimited bandwidth = best for high-traffic pages. Great for local biz mockups.', '500 builds/mo. 20K files. 25MB file size.', 'USE FOR: high-traffic affiliate sites, mockup pages that get shared. Unlimited bandwidth.'),
    ('GitHub Pages', 'Unlimited repos, 100GB/mo bandwidth', 'YES - commercial allowed', 'SIMPLE SITES. Good for portfolio, docs, project pages. Free custom domain SSL.', '1GB repo size. 100GB/mo bandwidth.', 'USE FOR: portfolio, documentation, simple project sites.'),
    ('Railway', '$5 free credit/mo. Container hosting.', 'YES', 'BACKEND SERVICES. Good for APIs, bots, cron jobs. $5/mo free credit.', '$5/mo credit. Sleep after 30 min inactive.', 'USE FOR: Discord bots, API backends, cron jobs that need a server.'),
    ('Render', '750 hrs free compute, auto-sleep', 'YES', 'WEB SERVICES. Free PostgreSQL (90 days). Auto-deploy from GitHub.', 'Auto-sleep after 15 min. 750 hrs/mo.', 'USE FOR: small backends, API services, free PostgreSQL.'),
    ('Oracle Cloud', '2 VMs (1GB RAM each) FREE FOREVER', 'YES', 'PERSISTENT SERVERS. Free VMs for scrapers, cron jobs, bots. Always-on.', '2 VMs, 1GB RAM each, 1/8 OCPU. 200GB block.', 'USE FOR: persistent scrapers, cron jobs, Always Free tier is real.'),
    ('Supabase', '500MB DB, 1GB storage, 2GB bandwidth', 'YES', 'DATABASE + AUTH. PostgreSQL + auth + storage + realtime. Perfect for micro-SaaS backend.', '500MB DB. 2 projects max. 50K monthly active users.', 'USE FOR: every micro-SaaS and app backend.'),
]

for i, h in enumerate(hosting):
    row = 6 + i
    for j, val in enumerate(h, 1):
        c = ws3.cell(row=row, column=j, value=val)
        c.font = W if j > 1 else GO
        if j == 3:
            c.font = G if 'YES' in str(val) else R
        if i % 2 == 1:
            c.fill = R2

style(ws3, 6, [16, 40, 22, 45, 30, 45])

# ============================================================
# SHEET 4: LEAD GEN & OUTREACH STACK
# ============================================================
ws4 = wb.create_sheet('LEAD GEN STACK')

ws4.cell(row=2, column=2, value='LEAD GENERATION & OUTREACH — TOOLS & METHODS').font = T
ws4.merge_cells('B2:F2')

hdr(ws4, 5, ['TOOL/METHOD', 'TYPE', 'FREE TIER', 'USE CASE', 'INTEGRATION WITH PRINTMAXX', 'PRIORITY'])

lead_tools = [
    ('Bland AI', 'AI VOICE CALLING', '100 calls/day FREE. $0.09/min after.', 'Automated AI voice outreach. Call local businesses, qualify leads, book meetings. Claude writes scripts.', 'NEW: Add to S05 outreach pipeline. 100 free calls/day → qualify → warm email follow-up → close', 'P1 — FREE 100 calls/day is insane'),
    ('Instantly', 'COLD EMAIL', '$30/mo. Unlimited email accounts. Warmup included.', 'Send 100-500 personalized cold emails/day. Automated warmup. Reply tracking. A/B testing.', 'EXISTING: integrated in local_biz_pipeline.py. Claude personalizes every email. Instantly sends.', 'P1'),
    ('Clay', 'DATA ENRICHMENT', '$149/mo (or free trial)', 'Enrich leads with email, phone, company data. Waterfall enrichment across 75+ providers.', 'PLANNED: enrich leads from Crunchbase/Google Maps scraping before email outreach', 'P2 — expensive but powerful'),
    ('Apollo.io', 'B2B DATABASE', 'Free: 10K records/mo. Email sending.', '275M+ contacts database. Find decision-makers. Built-in sequences.', 'ALTERNATIVE to Instantly. Free tier is generous for starting. Good for B2B lead sourcing.', 'P2'),
    ('Wappalyzer', 'TECH DETECTION', 'FREE browser extension', 'Detect website tech stacks. Find sites on outdated WordPress, jQuery, no SSL. Score outdatedness.', 'NEW: Feed outdated site data to S03 (local biz) and S07 pipelines. Identify prospects who NEED upgrades.', 'P1 — FREE and high signal'),
    ('Crunchbase', 'FUNDING DATA', 'Free: basic company search', 'Find recently funded startups → they have budget → they need: websites, apps, marketing, dev work.', 'NEW: Scrape funded companies → identify service needs → cold email offering Claude Code-powered services', 'P1'),
    ('Google Maps API', 'LOCAL BIZ SCRAPING', '28K requests/mo FREE', 'Find local businesses by category + location. Extract: name, address, phone, website, reviews, rating.', 'EXISTING: Used in local_biz_pipeline.py. Scrape dentists/plumbers/restaurants → score website → cold email.', 'P1 — already integrated'),
    ('BuiltWith', 'TECH DETECTION', '$295/mo (paid) or free trial', '673M website database. Historical tech tracking. Find sites using specific outdated tech at scale.', 'FOR SCALE: When Wappalyzer isn\'t enough. Bulk detect 10K+ outdated sites. Feed to outreach pipeline.', 'P3 — expensive, use Wappalyzer first'),
    ('LinkedIn Sales Navigator', 'B2B PROSPECTING', '$79/mo', 'Advanced LinkedIn search. 50 InMails/mo. Lead lists. CRM integration.', 'FOR B2B: When cold email isn\'t enough. Find decision-makers. InMail for warm outreach.', 'P3 — cost/benefit after revenue'),
    ('Hunter.io', 'EMAIL FINDER', '25 searches/mo free', 'Find email addresses from domains. Verify emails before sending.', 'COMPLEMENTARY: Verify emails from Google Maps scraping before sending via Instantly.', 'P2'),
]

for i, tool in enumerate(lead_tools):
    row = 6 + i
    for j, val in enumerate(tool, 1):
        c = ws4.cell(row=row, column=j, value=val)
        c.font = W if j > 1 else GO
        if j == 6:
            c.font = G if 'P1' in str(val) else (GO if 'P2' in str(val) else W)
        if i % 2 == 1:
            c.fill = R2

style(ws4, 6, [18, 16, 30, 45, 50, 28])

# ============================================================
# SHEET 5: RBI PERPETUAL IMPROVEMENT SYSTEM
# ============================================================
ws5 = wb.create_sheet('RBI SYSTEM')

ws5.cell(row=2, column=2, value='RBI — RESEARCH-BASED IMPROVEMENT (PERPETUAL)').font = T
ws5.merge_cells('B2:D2')
ws5.cell(row=3, column=2, value='Jane Street / RenTech model: RESEARCH → SCREEN → PAPER TRADE → DEPLOY → MONITOR → REBALANCE → RESEARCH').font = B
ws5.merge_cells('B3:D3')

hdr(ws5, 5, ['COMPONENT', 'FREQUENCY', 'AUTOMATION', 'WHAT IT DOES', 'OUTPUT'])

rbi_data = [
    ('DAILY ALPHA EXTRACTION', '', '', '', ''),
    ('Twitter 89-account scrape', 'Daily 4 AM', 'FULL AUTO', 'Scrape all 89 high-signal Twitter accounts. Extract alpha: tactics, revenue numbers, tool discoveries', 'ALPHA_STAGING.csv (PENDING_REVIEW entries)'),
    ('Reddit 41-subreddit scrape', 'Daily 4 AM', 'FULL AUTO', 'Scrape r/SideProject, r/juststart, r/Affiliatemarketing + 38 more. Extract: what\'s working, revenue proof', 'ALPHA_STAGING.csv'),
    ('Product Hunt trending', 'Daily 6 AM', 'FULL AUTO', 'Monitor top launches. Identify clone opportunities. Track competitor features.', 'APP_CLONE_OPPORTUNITIES.csv'),
    ('GitHub trending repos', 'Daily 6 AM', 'FULL AUTO', 'New tools, frameworks, MCP servers. Identify automation opportunities.', 'TOOL_ALPHA entries in ALPHA_STAGING.csv'),
    ('Platform algorithm changes', 'Daily 6 AM', 'FULL AUTO', 'TikTok/IG/YouTube/X algorithm shifts. New ranking signals. Content strategy adjustments.', 'PLATFORM_ALGORITHM_RESEARCH.md'),
    ('Competitor monitoring', 'Daily 6 AM', 'FULL AUTO', 'Track competitor pricing, features, launches. Identify gaps and opportunities.', 'COMPETITIVE_LANDSCAPE_MAP.md'),
    ('', '', '', '', ''),
    ('DAILY OPS AUDIT', '', '', '', ''),
    ('Morning briefing generation', 'Daily 5 AM', 'FULL AUTO', 'Scan ALL ledgers, logs, ops, financials. Generate prioritized action-required report.', 'DAILY_BRIEFINGS/DAILY_BRIEFING_YYYY-MM-DD.md'),
    ('Content performance check', 'Daily 7 AM', 'SEMI-AUTO', 'Which posts performed? Which flopped? Why? Adjust strategy.', 'Content performance insights in briefing'),
    ('Account health monitoring', 'Daily 7 AM', 'FULL AUTO', 'Shadowban detection, engagement drops, warmup status per platform.', 'ACCOUNT_HEALTH_DAILY.csv'),
    ('Revenue tracking', 'Daily', 'SEMI-AUTO', 'Log any new revenue. Track against projections. Alert if falling behind.', 'REVENUE_TRACKER.csv + P_AND_L'),
    ('Freelance pipeline review', 'Daily 8 AM', 'SEMI-AUTO', 'Check new orders, pending deliveries, review requests. Prioritize day\'s work.', 'FREELANCE_PIPELINE.csv'),
    ('', '', '', '', ''),
    ('WEEKLY DEEP ANALYSIS', '', '', '', ''),
    ('Method performance ranking', 'Monday 6 AM', 'FULL AUTO', 'Rank all active ops by: revenue/hour, growth rate, margin, scalability. Identify winners and losers.', 'WEEKLY_METHOD_AUDIT.md'),
    ('Cross-pollination scan', 'Monday 6 AM', 'FULL AUTO', 'Check for new synergy opportunities between methods. Update CROSS_POLLINATION_MATRIX.csv.', 'Updated synergy scores + suggestions'),
    ('Alpha batch review', 'Monday', 'SEMI-AUTO', 'Review all PENDING_REVIEW alpha. Approve/reject. Integrate approved into ops.', 'Approved alpha → method updates'),
    ('Tool audit', 'Monday', 'SEMI-AUTO', 'Check free tier limits. Any approaching caps? New tools discovered? Stack optimization.', 'Updated TOOLS_SERVICES_MASTER.csv'),
    ('Experiment results', 'Monday', 'SEMI-AUTO', 'Check A/B test results. Implement winners. Design new experiments.', 'EXPERIMENTS_METRICS.csv'),
    ('', '', '', '', ''),
    ('MONTHLY STRATEGIC', '', '', '', ''),
    ('Full ops portfolio rebalancing', 'Monthly 1st', 'SEMI-AUTO', 'Which ops to double down? Which to kill? Where to allocate more time/resources?', 'Updated priorities in ALL OPS'),
    ('Revenue projection update', 'Monthly 1st', 'FULL AUTO', 'Monte Carlo simulation on all active ops. Kelly Criterion for allocation. Updated projections.', 'REVENUE_PROJECTIONS.md'),
    ('New op identification', 'Monthly 1st', 'SEMI-AUTO', 'First principles scan: what new ops should we add? What\'s changed in the market? What alpha has matured?', 'New ops added to master list'),
    ('Financial deep review', 'Monthly 1st', 'SEMI-AUTO', 'P&L review. Tax deductions. Investment positions. Cash flow planning.', 'P_AND_L_MONTHLY.csv + dashboard'),
    ('Tool stack optimization', 'Monthly 1st', 'SEMI-AUTO', 'Reduce costs. Upgrade where needed. Consolidate tools. Test new tools.', 'TOOLS optimization actions'),
]

for i, d in enumerate(rbi_data):
    row = 6 + i
    for j, val in enumerate(d, 1):
        c = ws5.cell(row=row, column=j, value=val)
        if d[1] == '' and d[0] != '':
            c.font = H
            c.fill = SEC
        elif d[0] == '':
            pass
        else:
            c.font = W if j > 1 else GO
            if j == 3:
                c.font = G if 'FULL' in str(val) else GO
            if i % 2 == 1:
                c.fill = R2

style(ws5, 5, [30, 14, 12, 55, 40])

# SAVE
out = '/sessions/awesome-nice-brown/mnt/PRINTMAXX_STARTER_KITttttt/PRINTMAXX_MASTER_OPS.xlsx'
wb.save(out)
print(f'Saved to {out}')
print(f'Sheets: {wb.sheetnames}')
op_count = len([o for o in ops if o[0] != '' and o[1] != ''])
print(f'Total ops: {op_count}')
