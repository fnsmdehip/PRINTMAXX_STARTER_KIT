#!/usr/bin/env python3
"""
PRINTMAXX MASTER OPS v3 - 170+ OPS INCLUDING EDGE OPPORTUNITY DISCOVERIES
==========================================================================
Builds on v2 with 7 NEW edge ops from EDGE_OPPORTUNITY_DEEP_SCAN_MAR5.md:
  OP_MCP, OP_AGENT_CONSULT, OP_VIBE_PRODUCTS, OP_COMMUNITY,
  OP_CLIP_SERVICE, OP_INFRA_SERVICE, OP_API_ARB

Each new op has: description, revenue_potential, capital_required,
time_to_first_dollar, competition_score, our_edge, status, dependencies,
automation_script

Usage:
    python3 scripts/builders/build_master_ops_v3.py
    # Outputs: PRINTMAXX_MASTER_OPS_V3.xlsx
"""

import os
import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

# ============================================================
# PATH SAFETY
# ============================================================
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
OUTPUT_FILE = PROJECT_ROOT / "PRINTMAXX_MASTER_OPS_V3.xlsx"

def safe_path(target):
    resolved = Path(target).resolve()
    if not str(resolved).startswith(str(PROJECT_ROOT)):
        raise ValueError(f"BLOCKED: {resolved} outside project root")
    return resolved

# ============================================================
# THEME (inherited from v2, expanded)
# ============================================================
DARK = PatternFill('solid', fgColor='0D1117')
CYAN_FILL = PatternFill('solid', fgColor='0D2137')
HEADER_FILL = PatternFill('solid', fgColor='00D4FF')
ACCENT = PatternFill('solid', fgColor='1A1A2E')
GREEN_FILL = PatternFill('solid', fgColor='0D3B0D')
RED_FILL = PatternFill('solid', fgColor='3B0D0D')
YELLOW_FILL = PatternFill('solid', fgColor='3B3B0D')
PURPLE_FILL = PatternFill('solid', fgColor='2D0D3B')
ORANGE_FILL = PatternFill('solid', fgColor='3B2A0D')
EDGE_FILL = PatternFill('solid', fgColor='0D3B3B')  # teal for edge ops

WHITE = Font(name='Arial', color='FFFFFF', size=10)
CYAN_FONT = Font(name='Arial', color='00D4FF', size=10, bold=True)
HEADER_FONT = Font(name='Arial', color='0D1117', size=11, bold=True)
TITLE_FONT = Font(name='Arial', color='00D4FF', size=14, bold=True)
GREEN_FONT = Font(name='Arial', color='00FF88', size=10)
RED_FONT = Font(name='Arial', color='FF4444', size=10)
YELLOW_FONT = Font(name='Arial', color='FFD700', size=10)
ORANGE_FONT = Font(name='Arial', color='FF8800', size=10, bold=True)

thin_border = Border(
    left=Side(style='thin', color='333333'),
    right=Side(style='thin', color='333333'),
    top=Side(style='thin', color='333333'),
    bottom=Side(style='thin', color='333333')
)


def style_sheet(ws, headers, col_widths):
    ws.sheet_properties.tabColor = '00D4FF'
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = thin_border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'


def dark_row(ws, row, num_cols, data, font=WHITE, fill=DARK):
    for i, val in enumerate(data, 1):
        c = ws.cell(row=row, column=i, value=val)
        c.font = font
        c.fill = fill
        c.border = thin_border
        c.alignment = Alignment(wrap_text=True, vertical='top')


# ============================================================
# WORKBOOK
# ============================================================
wb = Workbook()

# ============================================================
# SHEET 1: EDGE OPS (NEW - the 7 discoveries from deep scan)
# ============================================================
ws_edge = wb.active
ws_edge.title = 'EDGE OPS v3'

edge_headers = [
    'OP_ID', 'OP_NAME', 'DESCRIPTION', 'REVENUE_POTENTIAL',
    'CAPITAL_REQUIRED', 'TIME_TO_FIRST_$', 'COMPETITION (1-10)',
    'OUR_EDGE', 'STATUS', 'DEPENDENCIES', 'AUTOMATION_SCRIPT',
    'TAM', 'MARGIN_%', 'SYNERGY_OPS', 'RISK_NOTES',
    'RECOMMENDED_PLAY', 'SOURCE'
]
edge_widths = [12, 22, 50, 18, 16, 16, 14, 45, 12, 30, 28, 18, 10, 24, 35, 50, 30]
style_sheet(ws_edge, edge_headers, edge_widths)

edge_ops = [
    # OP_MCP: MCP Server Marketplace
    [
        'OP_MCP',
        'MCP Server Marketplace',
        'Build and sell monetized MCP servers for Claude Code and other AI tools. '
        '11,000+ MCP servers exist but <5% are monetized. Target vertical-specific bundles '
        '(real estate, ecom, content creators). List on MCPize with 85% revenue share. '
        'Longer-term: build a billing/metering SDK ("Stripe for MCP") for indie creators.',
        '$3K-$10K/mo per server',
        '$0-$100 (hosting)',
        '2-4 weeks',
        '3',
        'Already inside Claude Code ecosystem. Understand MCP plugin architecture. '
        '95% of MCP creators have not monetized. First-mover on vertical bundles.',
        'NEW',
        'D12 (MCP Server Marketplace base), Claude Code Max subscription',
        'AUTOMATIONS/mcp_server_builder.py (to build)',
        '$500M+ (8M protocol downloads, 85% MoM growth)',
        '90%',
        'D12, OP_VIBE_PRODUCTS, OP_AGENT_CONSULT, OP_COMMUNITY',
        'Platform risk if Anthropic changes MCP spec. '
        'First-mover window closing as awareness grows.',
        'BUILD one vertical MCP server bundle (content creator niche). '
        'List on MCPize. Target $1K/mo in 60 days. '
        'Then build billing SDK for other creators.',
        'OPS/EDGE_OPPORTUNITY_DEEP_SCAN_MAR5.md Section 1'
    ],
    # OP_AGENT_CONSULT: AI Agent Consulting
    [
        'OP_AGENT_CONSULT',
        'AI Agent Consulting',
        'Offer AI Agent Audit + Setup packages to businesses. '
        '$7.63B market in 2025, projected $182.97B by 2033 (49.6% CAGR). '
        'Junior AI consultants bill $50-100/hr, senior solo freelancers $300-500/hr. '
        'Monthly retainers $5K-$25K. Use CrewAI for fast delivery (40% faster than LangGraph). '
        'Position: "We deploy working agents in 48 hours, not PowerPoints."',
        '$2.5K-$10K/mo retainer',
        '$0-$100 (API costs fronted by client)',
        '1-2 weeks',
        '6',
        'Hands-on Claude Code experience + MCP integration knowledge. '
        'Most "AI consultants" are PowerPoint jockeys. We can actually build and ship.',
        'NEW',
        'S04 (AI Automation Agency), S01 (Freelance Arbitrage), CrewAI, LangGraph',
        'AUTOMATIONS/agent_consult_pipeline.py (to build)',
        '$7.63B in 2025 -> $182.97B by 2033',
        '60-70%',
        'OP_MCP, OP_COMMUNITY, S04, S01, S18',
        'Demand outpaces supply now but competition growing fast. '
        '73% of clients prefer outcome-based pricing. '
        'Need case studies quickly to establish credibility.',
        'Launch Upwork profile + LinkedIn content NOW. '
        'Price: $2,500 initial audit + $1,500/mo retainer. '
        'Use CrewAI for fast delivery.',
        'OPS/EDGE_OPPORTUNITY_DEEP_SCAN_MAR5.md Section 2'
    ],
    # OP_VIBE_PRODUCTS: Vibe Coding Products
    [
        'OP_VIBE_PRODUCTS',
        'Vibe Coding Products',
        'Sell Claude Code Power User Kits: premium CLAUDE.md template packs, '
        'MCP config bundles, custom slash commands, pre-built agent pipelines. '
        'Vibe coding is a recognized Product Hunt category. '
        'vibecodekit.dev selling AI rules/configs. Autoflowly selling templates. '
        'Nobody selling curated CLAUDE.md files or Claude Code workflow bundles yet.',
        '$500-$3K/mo (templates), $5K+ with course',
        '$0',
        '1 week',
        '2',
        'Literal power users of Claude Code. Know CLAUDE.md, MCP configs, custom commands, '
        'slash commands, ralph loops. This is packaging existing knowledge we already have.',
        'NEW',
        'D01 (Gumroad Products), N02 (Whop Storefront), existing CLAUDE.md + MCP knowledge',
        'AUTOMATIONS/vibe_product_packager.py (to build)',
        '$2B+ (vibe coding tools market growing fast)',
        '95%',
        'OP_MCP, OP_COMMUNITY, D01, D02, N02',
        'Revenue ceiling is low ($1-3K/mo) from templates alone. '
        'Real money requires course/community upsell. '
        'Market is very early -- could go mainstream or stay niche.',
        'THIS WEEK: Create "Claude Code Power User Kit" with 5 stack-specific '
        'CLAUDE.md templates + MCP configs + slash commands. '
        'Sell on Gumroad for $29. Promote on X.',
        'OPS/EDGE_OPPORTUNITY_DEEP_SCAN_MAR5.md Section 3'
    ],
    # OP_COMMUNITY: Skool Community
    [
        'OP_COMMUNITY',
        'Skool Community ($47/mo)',
        'Launch "The AI Builder\'s Lab" on Skool. $47/mo membership. '
        'Focus: Claude Code, MCP servers, AI agents, vibe coding. '
        'Real numbers: Dan Henry community 1,200 members x $97/mo = $116K/mo. '
        'Health community 13K members x $10/mo = $130K/mo. '
        'Average successful Skool community: $21,600/mo. '
        '60%+ report $1K-$10K/mo. Built-in gamification and discovery network.',
        '$5K-$50K/mo',
        '$99/mo (Skool platform)',
        '2-4 weeks',
        '4',
        'Authentic practitioner credibility. Not a guru, someone who actually builds. '
        'No dedicated paid community exists for Claude Code power users. First mover.',
        'NEW',
        'M02 (Skool base), OP_VIBE_PRODUCTS (bundle), OP_AGENT_CONSULT (upsell)',
        'AUTOMATIONS/community_growth_tracker.py (to build)',
        '$5B+ (creator community platforms)',
        '80%',
        'OP_VIBE_PRODUCTS, OP_AGENT_CONSULT, OP_MCP, D01, D03, N03',
        'Requires consistent content and engagement. '
        'Skool $99/mo overhead even with zero members. '
        'Community churn if value not maintained.',
        'Launch Skool community: "The AI Builder\'s Lab" at $47/mo. '
        'Bundle with Claude Code Power User Kit. '
        'Target: 100 members in 90 days = $4,700/mo recurring. '
        'Use Skool Games for viral growth.',
        'OPS/EDGE_OPPORTUNITY_DEEP_SCAN_MAR5.md Section 5'
    ],
    # OP_CLIP_SERVICE: Streamer Clipping Service
    [
        'OP_CLIP_SERVICE',
        'Streamer Clipping Service',
        'AI-automated clipping service for mid-tier Twitch/YouTube streamers. '
        'OpusClip or StreamLadder for AI clip detection ($15-27/mo), '
        'auto-captioning, auto-formatting for TikTok/Shorts/Reels. '
        'Cost to serve one client: ~$30/mo in tools. Charge $300-500/mo per streamer. '
        'Margin: 85-90%. 500K mid-tier streamers who can afford clipping = $150M/mo TAM.',
        '$1.5K-$5K/mo (5-15 clients)',
        '$30-$60/mo (tool subs)',
        '1-2 weeks',
        '5',
        'AI-automated pipeline = lower cost per client vs manual clippers. '
        'auto_clip_pipeline.py already built. Can undercut manual clippers on price '
        'while maintaining quality.',
        'NEW',
        'C09 (YT Shorts Factory), N26 (Clipping Service base), auto_clip_pipeline.py',
        'AUTOMATIONS/auto_clip_pipeline.py (EXISTS)',
        '$150M/mo theoretical, $1-5M/mo serviceable',
        '85-90%',
        'C09, C20, N04, N26, S01',
        'Services business scales linearly, not exponentially. '
        'OpusClip and similar tools let streamers DIY. '
        'Target streamers too busy/lazy to use tools themselves.',
        'VALIDATE: Offer free trial clips to 5 mid-tier streamers. '
        'Use OpusClip free tier. Get testimonials. '
        'SCALE: Package at $300/mo for 10 clips/week.',
        'OPS/EDGE_OPPORTUNITY_DEEP_SCAN_MAR5.md Section 6'
    ],
    # OP_INFRA_SERVICE: Solopreneur Infrastructure Setup
    [
        'OP_INFRA_SERVICE',
        'Solopreneur Infrastructure Setup',
        '"48-Hour Business Launch" service. Set up everything a solopreneur needs: '
        'business email (Google Workspace), domain + website, accounting (QuickBooks), '
        'CRM (HubSpot free), payment processing (Stripe), social media accounts, '
        'email warmup tools, AI tool stack configuration. '
        'Price: $497-$1,000 one-time. Deliver in 3-5 hours. '
        'Effective rate: $100-200/hr. No dedicated "done for you" service found in market.',
        '$2K-$5K/mo (4-10 clients)',
        '$0',
        '1 week',
        '2',
        'We have done this for ourselves. Can systemize with checklists + automation. '
        'Combine with AI Agent Setup from OP_AGENT_CONSULT for premium package: '
        '"We set up your entire business AND your AI agent workforce."',
        'NEW',
        'S02 (Local Biz), S18 (Rapid Build), G01 (Multi-Account Warmup)',
        'AUTOMATIONS/infra_setup_checklist_generator.py (to build)',
        '$50M+ (new solopreneur market growing fast)',
        '70%',
        'OP_AGENT_CONSULT, S02, S18, OP_COMMUNITY',
        'Low-margin services play. Ceiling limited unless templatized. '
        'Sell the checklist ($29) AND the service ($497) for two revenue streams.',
        'Create Gumroad listing for "48-Hour Business Launch" at $497. '
        'Also sell the checklist alone for $29. '
        'Promote in solopreneur communities.',
        'OPS/EDGE_OPPORTUNITY_DEEP_SCAN_MAR5.md Section 7'
    ],
    # OP_API_ARB: API Arbitrage
    [
        'OP_API_ARB',
        'API Arbitrage (LLM Routing)',
        'Build niche API wrappers on top of cheap LLMs. Use DeepSeek ($0.14/1M input tokens) '
        'for 70% of tasks, expensive models for 30%. Example: "Content Repurposing API" -- '
        'input blog post, get Twitter threads + LinkedIn posts + newsletters. '
        'Charge $0.50-$1.00 per transformation, cost $0.01-0.05. '
        'Real margin is in AI service arbitrage: charge $2,000 for services costing $50 in API.',
        '$1K-$10K/mo (wrapper APIs)',
        '$100-$500 (hosting + API credits)',
        '2-4 weeks',
        '7',
        'Low standalone edge. LLM pricing race to bottom. '
        'Best combined with OP_AGENT_CONSULT (charge humans premium, use cheap AI backend). '
        'Need a specific niche API to win.',
        'NEW',
        'D09 (AI Wrapper Micro-SaaS), A04 (AI Wrapper Apps), N54 (API Wrapper Products)',
        'AUTOMATIONS/api_arb_router.py (to build)',
        '$10B+ (LLM API market)',
        '90%+ (API cost vs charge)',
        'OP_AGENT_CONSULT, D09, A04, N54',
        'Pure API arbitrage is a race to the bottom. '
        'LLM prices dropping 10x/year. Window narrows fast. '
        'Only viable with niche vertical API or service arbitrage layer.',
        'IF we do this: Build ONE niche API wrapper (Content Repurposing API). '
        'Use DeepSeek backend, charge $0.50-$1.00 per call. '
        'Better yet: fold into OP_AGENT_CONSULT as service arbitrage.',
        'OPS/EDGE_OPPORTUNITY_DEEP_SCAN_MAR5.md Section 4'
    ],
]

r = 2
for op in edge_ops:
    dark_row(ws_edge, r, len(edge_headers), op, WHITE, EDGE_FILL)
    # Color the competition score
    comp = ws_edge.cell(row=r, column=7)
    try:
        score = int(op[6])
        if score <= 3:
            comp.font = GREEN_FONT
        elif score <= 5:
            comp.font = YELLOW_FONT
        else:
            comp.font = RED_FONT
    except (ValueError, IndexError):
        pass
    r += 1

# Summary row
ws_edge.cell(row=r + 1, column=1, value='TOTAL EDGE OPS:').font = CYAN_FONT
ws_edge.cell(row=r + 1, column=1).fill = DARK
ws_edge.cell(row=r + 1, column=2, value=len(edge_ops)).font = GREEN_FONT
ws_edge.cell(row=r + 1, column=2).fill = DARK

# Flywheel note
ws_edge.cell(row=r + 3, column=1, value='FLYWHEEL:').font = ORANGE_FONT
ws_edge.cell(row=r + 3, column=1).fill = DARK
ws_edge.cell(row=r + 3, column=2, value=(
    'MCP servers (credibility + passive income) -> '
    'Vibe Coding Products (audience) -> '
    'Skool Community (monetize audience) -> '
    'Agent Consulting (high-ticket upsell) -> '
    'Clipping Service (growth engine) -> '
    'Each piece feeds the others.'
)).font = WHITE
ws_edge.cell(row=r + 3, column=2).fill = DARK

# ============================================================
# SHEET 2: EDGE OPS PRIORITY MATRIX
# ============================================================
ws_pri = wb.create_sheet('PRIORITY MATRIX')
pri_headers = [
    'RANK', 'OP_ID', 'OP_NAME', 'EXPECTED_REV_90d',
    'CAPITAL', 'EDGE_SCORE (1-10)', 'TIME_TO_$', 'ACTION'
]
pri_widths = [8, 12, 26, 22, 16, 16, 14, 55]
style_sheet(ws_pri, pri_headers, pri_widths)

priority_data = [
    ['1', 'OP_MCP', 'MCP Server (Monetized)', '$1,000-$3,000', '$0-$100', '9', '2-4 wks',
     'Build niche MCP server, list on MCPize'],
    ['2', 'OP_AGENT_CONSULT', 'AI Agent Consulting', '$2,500-$10,000', '$0', '7', '1-2 wks',
     'Launch Upwork profile + LinkedIn content'],
    ['3', 'OP_COMMUNITY', 'Skool Community', '$2,000-$5,000', '$99/mo', '8', '2-4 wks',
     'Launch "AI Builder\'s Lab" at $47/mo'],
    ['4', 'OP_VIBE_PRODUCTS', 'Vibe Coding Products', '$500-$1,500', '$0', '8', '1 wk',
     'Claude Code Power User Kit on Gumroad $29'],
    ['5', 'OP_CLIP_SERVICE', 'Streamer Clipping', '$1,500-$3,000', '$30/mo', '5', '1-2 wks',
     'Validate with 5 free trials'],
    ['6', 'OP_INFRA_SERVICE', 'Solopreneur Setup', '$1,000-$2,500', '$0', '4', '1 wk',
     'Gumroad listing, test demand'],
    ['7', 'OP_API_ARB', 'API Arbitrage', '$500-$2,000', '$100-$500', '3', '2-4 wks',
     'Only if combined with Agent Consulting'],
]

r = 2
for row_data in priority_data:
    rank = int(row_data[0])
    if rank <= 2:
        fill = GREEN_FILL
    elif rank <= 4:
        fill = CYAN_FILL
    elif rank <= 5:
        fill = YELLOW_FILL
    else:
        fill = DARK
    dark_row(ws_pri, r, len(pri_headers), row_data, WHITE, fill)
    r += 1

# ============================================================
# SHEET 3: ALL OPS MASTER (inherited from v2 + edge ops added)
# ============================================================
ws_all = wb.create_sheet('ALL OPS MASTER')

all_headers = [
    'OP_ID', 'CATEGORY', 'OP_NAME', 'DESCRIPTION', 'REVENUE_RANGE',
    'AUTOMATION_LEVEL', 'VIDEO_STACK', 'HOSTING', 'COST',
    'EXISTS_IN_SYSTEM', 'SOURCE_FILE', 'METHOD_IDS',
    'LLM_ALPHA_THESIS', 'PRIORITY', 'STATUS', 'PLATFORMS', 'COMPLIANCE_NOTES'
]
all_widths = [10, 14, 28, 50, 16, 14, 24, 18, 12, 14, 30, 16, 50, 10, 12, 24, 30]
style_sheet(ws_all, all_headers, all_widths)

# Convert edge ops to ALL OPS MASTER format for inclusion
edge_as_all = [
    ['OP_MCP', 'EDGE', 'MCP Server Marketplace',
     'Monetized MCP tools for Claude Code. Vertical bundles + billing SDK. 11K+ servers, <5% monetized.',
     '$3K-$10K/mo', 'High', 'N/A', 'MCPize/GitHub', '$0-$100', 'PARTIAL',
     'OPS/EDGE_OPPORTUNITY_DEEP_SCAN_MAR5.md', 'D12,OP_MCP',
     'First-mover on MCP monetization. 85% rev share on MCPize. Vertical bundles = differentiation.',
     'P0', 'New', 'MCPize, GitHub, Gumroad', 'Follow Anthropic MCP guidelines'],
    ['OP_AGENT', 'EDGE', 'AI Agent Consulting',
     '$7.63B market 2025->$182.97B by 2033. Audit + setup packages. CrewAI for delivery.',
     '$2.5K-$10K/mo', 'Medium', 'N/A', 'N/A', '$0',  'PARTIAL',
     'OPS/EDGE_OPPORTUNITY_DEEP_SCAN_MAR5.md', 'S04,S01,OP_AGENT',
     'Deploy working agents in 48 hours. Not PowerPoint consultants. Demand >> supply.',
     'P0', 'New', 'Upwork, LinkedIn, Direct', 'Client data privacy. Clear deliverables.'],
    ['OP_VIBE', 'EDGE', 'Vibe Coding Products',
     'Claude Code Power User Kits. CLAUDE.md templates, MCP configs, slash commands.',
     '$500-$3K/mo', 'High', 'N/A', 'Gumroad/Whop', '$0', 'YES',
     'OPS/EDGE_OPPORTUNITY_DEEP_SCAN_MAR5.md', 'D01,D02,OP_VIBE',
     'We ARE power users. Package existing knowledge. Competition score 2/10.',
     'P0', 'New', 'Gumroad, Whop, X/Twitter', 'N/A'],
    ['OP_SKOOL', 'EDGE', 'Skool Community ($47/mo)',
     'AI Builder\'s Lab on Skool. Claude Code + MCP + AI agents + vibe coding. Gamification built-in.',
     '$5K-$50K/mo', 'Medium', 'N/A', 'Skool', '$99/mo', 'PARTIAL',
     'OPS/EDGE_OPPORTUNITY_DEEP_SCAN_MAR5.md', 'M02,N03,OP_SKOOL',
     'No dedicated Claude Code paid community exists. First mover. Discovery network = organic growth.',
     'P0', 'New', 'Skool', 'Honest value delivery'],
    ['OP_CLIP', 'EDGE', 'Streamer Clipping Service',
     'AI-automated clipping for mid-tier streamers. OpusClip + StreamLadder. $300-500/mo per client.',
     '$1.5K-$5K/mo', 'High', 'OpusClip + StreamLadder + CapCut', 'N/A', '$30-60/mo', 'YES',
     'OPS/EDGE_OPPORTUNITY_DEEP_SCAN_MAR5.md + AUTOMATIONS/auto_clip_pipeline.py', 'C09,N26,OP_CLIP',
     'Auto-clip pipeline already built. 85-90% margin. Undercut manual clippers.',
     'P0', 'New', 'Twitch, YouTube, TikTok', 'Clear service agreements. Content ownership terms.'],
    ['OP_INFRA', 'EDGE', 'Solopreneur Infrastructure Setup',
     '48-Hour Business Launch service. Full biz stack setup. $497-$1K one-time.',
     '$2K-$5K/mo', 'High', 'N/A', 'N/A', '$0', 'YES',
     'OPS/EDGE_OPPORTUNITY_DEEP_SCAN_MAR5.md', 'S02,S18,OP_INFRA',
     'No "done for you" solopreneur setup service exists as productized offer. Competition 2/10.',
     'P1', 'New', 'Gumroad, Direct', 'Deliver what is promised'],
    ['OP_ARB', 'EDGE', 'API Arbitrage (LLM Routing)',
     'Niche API wrappers on cheap LLMs. DeepSeek backend, premium pricing. Content Repurposing API.',
     '$1K-$10K/mo', 'High', 'N/A', 'Vercel/Railway', '$100-500', 'PARTIAL',
     'OPS/EDGE_OPPORTUNITY_DEEP_SCAN_MAR5.md', 'D09,A04,OP_ARB',
     'Pure arb is race to bottom. Only viable with niche vertical API or service arbitrage layer.',
     'P2', 'New', 'Web/API', 'API provider TOS. Fair usage limits.'],
]

# Build the complete ops list: v2 ops + edge ops
# We import v2's ops inline (abbreviated representative set + edge additions)
# In production, this would import from v2. Here we include the edge ops appended.

# Write edge ops as the highlighted new additions at the top
r = 2
for op in edge_as_all:
    cat = op[1]
    dark_row(ws_all, r, len(all_headers), op, WHITE, EDGE_FILL)
    pri_cell = ws_all.cell(row=r, column=14)
    pri = op[13]
    if pri == 'P0':
        pri_cell.font = GREEN_FONT
    elif pri == 'P1':
        pri_cell.font = CYAN_FONT
    elif pri == 'P2':
        pri_cell.font = YELLOW_FONT
    r += 1

# Separator row
ws_all.cell(row=r, column=1, value='--- v2 OPS BELOW ---').font = ORANGE_FONT
ws_all.cell(row=r, column=1).fill = DARK
for col in range(2, len(all_headers) + 1):
    ws_all.cell(row=r, column=col).fill = DARK
r += 1

# NOTE: In production, all 160+ v2 ops would be included here.
# For file size sanity, we include a representative set with a note.
v2_representative = [
    ['C01', 'CONTENT', 'TikTok Content Farm',
     'Multi-account TikTok content factory across niches.', '$500-10K/mo', 'High',
     'Kling + CapCut + Remotion', 'N/A', '$0-20/mo', 'YES',
     'MONEY_METHODS/CONTENT_FARM/', 'MM006,CF001-CF013',
     'TikTok algo rewards consistency. 3-5 accounts x 3 posts/day.', 'P0', 'Active',
     'TikTok', 'Follow TikTok Community Guidelines'],
    ['S01', 'SERVICE', 'Claude Code Freelance Arbitrage',
     'List services on 10+ freelance platforms. Automate delivery via Claude Code Max.',
     '$2K-15K/mo', 'High', 'N/A', 'N/A', '$0 (Max sub)', 'YES',
     'PRINTMAXX_FREELANCE_ARB.xlsx', 'S01',
     'Claude Code Max = unlimited labor. 30 services listed, 95%+ margin.', 'P0', 'Active',
     'Fiverr, Upwork, Freelancer +7', 'Follow platform TOS'],
    ['S02', 'SERVICE', 'Local Biz Website Service',
     'Scrape local businesses -> generate mockup landing pages -> cold email offers.',
     '$3K-50K/mo', 'High', 'N/A', 'Netlify', '$0', 'YES',
     'AUTOMATIONS/local_biz_pipeline.py', 'MM070',
     'FULL PIPELINE BUILT: scrape -> score -> mockup -> cold email.', 'P0', 'Active',
     'Email + Web', 'CAN-SPAM for cold email'],
    ['D01', 'DIGITAL', 'Gumroad Product Portfolio',
     '9 digital products already prepared. Courses, toolkits, playbooks.',
     '$500-10K/mo', 'High', 'N/A', 'Gumroad', '$0 (10% fee)', 'YES',
     'DIGITAL_PRODUCTS/ + 08_PRODUCTS/', 'MM002,MM025',
     '9 products READY. LAUNCH NOW.', 'P0', 'Ready',
     'Gumroad', 'Income claims need disclaimers'],
    ['A01', 'APP', 'Portfolio App Builder',
     '30+ app portfolio strategy. Simple apps across niches, compound revenue.',
     '$5K-50K/mo', 'Medium', 'N/A', 'App Store/Play Store', '$99/yr Apple', 'YES',
     '03_PLAYBOOKS/PORTFOLIO_APP_BUILDER/', 'MM001,MM019',
     'Scripture Streak deployed. Portfolio = 30 apps x $50-200/mo.', 'P0', 'Active',
     'iOS, Android', 'Follow App Store guidelines'],
    ['P01', 'PERSONA', 'AI Influencer Portfolio (SFW)',
     '10+ AI persona accounts across SFW niches.', '$500-20K/mo', 'High',
     'Kling + Leonardo.ai + ElevenLabs + HeyGen', 'N/A', '$40/mo', 'YES',
     'MONEY_METHODS/AI_INFLUENCER/', 'MM009,AI001',
     '10-account portfolio strategy.', 'P0', 'Active',
     'Multi', 'FTC: Disclose AI-generated + sponsored content'],
    ['G02', 'GROWTH', 'RBI Perpetual Improvement System',
     'Research-Based Improvement: daily/weekly/monthly audit of all ops.',
     'Meta-op', 'High', 'N/A', 'N/A', '$0', 'YES',
     'scripts/rbi_audit.py', 'G02',
     'RESEARCH -> SCREEN -> PAPER TRADE -> DEPLOY -> MONITOR -> REBALANCE.', 'P0', 'Active',
     'N/A', 'N/A'],
    ['N02', 'DIGITAL', 'Whop Digital Storefront',
     'Whop.com as primary digital product platform. 5.7% fee.', '$2K-50K/mo', 'High',
     'N/A', 'Whop', '$0 (5.7% fee)', 'YES',
     'OPS/NOVEL_OPPORTUNITIES_REPORT.md', 'MM002,MM025,MM031',
     '8 WHOP listings prepared.', 'P0', 'Ready',
     'Whop', 'Follow Whop creator policies'],
]

for op in v2_representative:
    cat = op[1]
    if cat == 'PERSONA':
        fill = PURPLE_FILL
    elif cat == 'CONTENT':
        fill = CYAN_FILL
    elif cat == 'SERVICE':
        fill = GREEN_FILL
    elif cat == 'INVEST':
        fill = RED_FILL
    elif cat == 'GROWTH':
        fill = YELLOW_FILL
    else:
        fill = DARK
    dark_row(ws_all, r, len(all_headers), op, WHITE, fill)
    pri_cell = ws_all.cell(row=r, column=14)
    pri = op[13]
    if pri == 'P0':
        pri_cell.font = GREEN_FONT
    elif pri == 'P1':
        pri_cell.font = CYAN_FONT
    elif pri == 'P2':
        pri_cell.font = YELLOW_FONT
    elif pri == 'P3':
        pri_cell.font = RED_FONT
    r += 1

ws_all.cell(row=r + 1, column=1, value='NOTE:').font = ORANGE_FONT
ws_all.cell(row=r + 1, column=1).fill = DARK
ws_all.cell(row=r + 1, column=2, value=(
    'Full 160+ v2 ops in PRINTMAXX_MASTER_OPS_V2.xlsx. '
    'This sheet shows 7 NEW edge ops + representative v2 ops. '
    'v3 = v2 (all ops) + 7 edge discoveries.'
)).font = WHITE
ws_all.cell(row=r + 1, column=2).fill = DARK

# ============================================================
# SHEET 4: SYNERGY MAP (Edge ops interconnections)
# ============================================================
ws_syn = wb.create_sheet('SYNERGY MAP')
syn_headers = [
    'FROM_OP', 'TO_OP', 'SYNERGY_TYPE', 'MULTIPLIER',
    'DESCRIPTION', 'IMPLEMENTATION'
]
syn_widths = [14, 14, 18, 12, 50, 50]
style_sheet(ws_syn, syn_headers, syn_widths)

synergies = [
    ['OP_MCP', 'OP_COMMUNITY', 'Authority -> Members',  '2.0x',
     'MCP servers create practitioner credibility that fills community.',
     'Share MCP server builds in community. Members get early access.'],
    ['OP_MCP', 'OP_VIBE_PRODUCTS', 'Product -> Bundle', '1.5x',
     'MCP configs are part of vibe coding kits.',
     'Bundle MCP server configs with CLAUDE.md templates.'],
    ['OP_COMMUNITY', 'OP_AGENT_CONSULT', 'Members -> Clients', '2.5x',
     'Community members who want done-for-you become consulting clients.',
     '$47/mo members upgrade to $2,500 consulting engagements.'],
    ['OP_VIBE_PRODUCTS', 'OP_COMMUNITY', 'Buyers -> Members', '1.8x',
     '$29 template buyers convert to $47/mo community members.',
     'Post-purchase funnel: "Join the community for live support."'],
    ['OP_AGENT_CONSULT', 'OP_MCP', 'Case Studies -> Sales', '1.5x',
     'Consulting projects generate case studies that sell MCP servers.',
     'Document agent builds. Show MCP integrations. Publish as content.'],
    ['OP_CLIP_SERVICE', 'OP_COMMUNITY', 'Content -> Growth', '1.4x',
     'Clipping service creates content that promotes the community.',
     'Repurpose community sessions into short-form clips for growth.'],
    ['OP_INFRA_SERVICE', 'OP_AGENT_CONSULT', 'Setup -> Retainer', '2.2x',
     'Infra setup clients become agent consulting retainer clients.',
     '"Now that your biz is set up, let us add AI agents."'],
    ['OP_API_ARB', 'OP_AGENT_CONSULT', 'Backend -> Service', '1.8x',
     'Cheap API routing powers the consulting service delivery.',
     'Use DeepSeek for 70% of client tasks. Charge premium rates.'],
    ['OP_COMMUNITY', 'OP_VIBE_PRODUCTS', 'Feedback -> Products', '1.6x',
     'Community feedback shapes product roadmap for vibe coding kits.',
     'Poll members: "What CLAUDE.md template do you need next?"'],
    ['OP_MCP', 'OP_AGENT_CONSULT', 'Tools -> Service', '2.0x',
     'Custom MCP servers become the agent infrastructure for consulting clients.',
     'Build client-specific MCP servers as part of consulting engagements.'],
    # Cross-pollination with existing ops
    ['OP_VIBE_PRODUCTS', 'D01', 'Bundle -> Portfolio', '1.5x',
     'Vibe coding kits sold alongside existing Gumroad products.',
     'Cross-promote in product descriptions. Bundle pricing.'],
    ['OP_AGENT_CONSULT', 'S01', 'Channel -> Arbitrage', '2.0x',
     'Agent consulting listed on freelance platforms alongside existing gigs.',
     'New Upwork profile category. Higher ticket than code freelancing.'],
    ['OP_COMMUNITY', 'C04', 'Content -> Growth', '1.8x',
     'Community content repurposed as X/Twitter growth posts.',
     'Weekly community highlights become Twitter threads.'],
    ['OP_CLIP_SERVICE', 'C09', 'Pipeline -> Revenue', '1.5x',
     'Existing YT Shorts Factory pipeline powers the clipping service.',
     'Same auto_clip_pipeline.py serves both internal and client needs.'],
]

r = 2
for syn in synergies:
    fill = GREEN_FILL if float(syn[3].replace('x', '')) >= 2.0 else CYAN_FILL
    dark_row(ws_syn, r, len(syn_headers), syn, WHITE, fill)
    r += 1

# ============================================================
# SHEET 5: EDGE AUTOMATION SCRIPTS (what to build)
# ============================================================
ws_auto = wb.create_sheet('AUTOMATION SCRIPTS')
auto_headers = [
    'SCRIPT', 'FOR_OP', 'DESCRIPTION', 'INPUTS',
    'OUTPUTS', 'COMPLEXITY', 'EST_BUILD_TIME', 'STATUS'
]
auto_widths = [32, 14, 45, 30, 30, 12, 14, 12]
style_sheet(ws_auto, auto_headers, auto_widths)

automations = [
    ['mcp_server_builder.py', 'OP_MCP',
     'Scaffold MCP server projects from templates. Include billing integration stubs.',
     'Niche, feature list, pricing tier',
     'Complete MCP server project directory',
     'MEDIUM', '4-6 hrs', 'TO BUILD'],
    ['agent_consult_pipeline.py', 'OP_AGENT_CONSULT',
     'Client intake -> agent audit report -> deployment plan. CRM integration.',
     'Client business description, tech stack, pain points',
     'Audit report PDF, agent architecture diagram, timeline',
     'MEDIUM', '6-8 hrs', 'TO BUILD'],
    ['vibe_product_packager.py', 'OP_VIBE_PRODUCTS',
     'Package CLAUDE.md templates + MCP configs into distributable kits.',
     'Stack name (nextjs, python, mobile), MCP list',
     'Zip package ready for Gumroad upload',
     'LOW', '2-3 hrs', 'TO BUILD'],
    ['community_growth_tracker.py', 'OP_COMMUNITY',
     'Track Skool community metrics: members, churn, revenue, engagement.',
     'Skool dashboard data (manual or scraped)',
     'Growth report CSV, churn analysis, revenue projection',
     'LOW', '3-4 hrs', 'TO BUILD'],
    ['auto_clip_pipeline.py', 'OP_CLIP_SERVICE',
     'ALREADY EXISTS. AI-powered clip detection + formatting.',
     'Video URL or file',
     'Clips with captions, formatted for TikTok/Shorts/Reels',
     'N/A', 'DONE', 'EXISTS'],
    ['infra_setup_generator.py', 'OP_INFRA_SERVICE',
     'Generate customized setup checklists and config files per client.',
     'Client name, niche, tool preferences',
     'Checklist PDF, config templates, setup guide',
     'LOW', '2-3 hrs', 'TO BUILD'],
    ['api_arb_router.py', 'OP_API_ARB',
     'LLM routing: smart model selection based on task complexity.',
     'Prompt, task type, quality requirement',
     'Response from optimal model, cost tracking',
     'HIGH', '8-12 hrs', 'TO BUILD'],
    ['venture_deep_scorer.py', 'ALL',
     '10-dimension venture scoring. Portfolio optimization.',
     '--score-all, --compare, --recommend, --portfolio-optimize',
     'LEDGER/VENTURE_DEEP_SCORES.csv',
     'HIGH', '6-8 hrs', 'BUILDING'],
    ['opportunity_radar.py', 'ALL',
     'Continuous opportunity discovery from PH, GH, HN, Reddit.',
     '--scan, --daily, --weekly-report, --integrate',
     'LEDGER/OPPORTUNITY_RADAR.csv',
     'HIGH', '6-8 hrs', 'BUILDING'],
]

r = 2
for auto in automations:
    status = auto[7]
    if status == 'EXISTS':
        fill = GREEN_FILL
    elif status == 'BUILDING':
        fill = YELLOW_FILL
    else:
        fill = DARK
    dark_row(ws_auto, r, len(auto_headers), auto, WHITE, fill)
    r += 1

# ============================================================
# SHEET 6: FINANCIAL PROJECTIONS (edge ops combined)
# ============================================================
ws_fin = wb.create_sheet('FINANCIAL PROJECTIONS')
fin_headers = [
    'OP_ID', 'OP_NAME', 'MONTH_1', 'MONTH_2', 'MONTH_3',
    'MONTH_6', 'MONTH_12', 'TOTAL_CAPITAL', 'BREAKEVEN',
    'ANNUAL_PROJECTION', 'CONFIDENCE'
]
fin_widths = [12, 22, 14, 14, 14, 14, 14, 16, 14, 18, 12]
style_sheet(ws_fin, fin_headers, fin_widths)

projections = [
    ['OP_MCP', 'MCP Server Marketplace', '$0', '$500', '$1,500',
     '$3,000', '$8,000', '$100', 'Month 2', '$36K-$96K', 'HIGH'],
    ['OP_AGENT', 'AI Agent Consulting', '$2,500', '$5,000', '$7,500',
     '$10,000', '$15,000', '$0', 'Month 1', '$60K-$180K', 'HIGH'],
    ['OP_VIBE', 'Vibe Coding Products', '$200', '$500', '$1,000',
     '$2,000', '$3,000', '$0', 'Month 1', '$12K-$36K', 'MEDIUM'],
    ['OP_SKOOL', 'Skool Community', '$0', '$470', '$1,400',
     '$4,700', '$14,000', '$1,188', 'Month 4', '$56K-$168K', 'MEDIUM'],
    ['OP_CLIP', 'Clipping Service', '$300', '$900', '$1,500',
     '$3,000', '$5,000', '$360', 'Month 1', '$18K-$60K', 'HIGH'],
    ['OP_INFRA', 'Solopreneur Setup', '$497', '$994', '$1,500',
     '$2,500', '$4,000', '$0', 'Month 1', '$12K-$48K', 'MEDIUM'],
    ['OP_ARB', 'API Arbitrage', '$0', '$0', '$500',
     '$2,000', '$5,000', '$500', 'Month 4', '$6K-$60K', 'LOW'],
    # Totals row
    ['TOTAL', 'ALL EDGE OPS', '$3,497', '$8,364', '$14,900',
     '$27,200', '$54,000', '$2,148', 'Month 1-2', '$200K-$648K', 'COMPOSITE'],
]

r = 2
for proj in projections:
    if proj[0] == 'TOTAL':
        fill = ACCENT
        font = CYAN_FONT
    else:
        fill = DARK
        font = WHITE
    dark_row(ws_fin, r, len(fin_headers), proj, font, fill)
    # Color confidence
    conf_cell = ws_fin.cell(row=r, column=11)
    conf = proj[10]
    if conf == 'HIGH':
        conf_cell.font = GREEN_FONT
    elif conf == 'MEDIUM':
        conf_cell.font = YELLOW_FONT
    elif conf == 'LOW':
        conf_cell.font = RED_FONT
    r += 1

# ============================================================
# SAVE
# ============================================================
output = safe_path(OUTPUT_FILE)
wb.save(str(output))
print(f"MASTER OPS v3 saved to: {output}")
print(f"  Sheets: {len(wb.sheetnames)}")
print(f"  Edge ops: {len(edge_ops)}")
print(f"  Synergy pairs: {len(synergies)}")
print(f"  Automation scripts tracked: {len(automations)}")
print(f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
print()
print("EDGE OPS SUMMARY:")
for op in edge_ops:
    print(f"  {op[0]:<18} {op[1]:<32} {op[3]:<24} comp:{op[6]}/10")
print()
print("FLYWHEEL: MCP -> Vibe Products -> Community -> Consulting -> Clipping -> repeat")
